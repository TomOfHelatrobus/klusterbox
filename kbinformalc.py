"""
a klusterbox module 
This module runs the Informal C, a program which allows users to record and track grievance settlements. It keeps 
track of grievance numbers, dates of the violation, at what level the settlement was signed, date of the signing, etc.
It also tracks which carrier was award what amount and if that settlement was paid. Report are available by carrier, 
by grievance as well as summaries. 
"""

# custom modules
import projvar  # defines project variables used in all modules.
from kbtoolbox import commit, dir_path, inquire, \
    isint, titlebar_icon, ProgressBarDe, ReportName, Handler, NameChecker, \
    GrievanceChecker, BackSlashDateChecker, Convert, DateTimeChecker
# standard libraries
from tkinter import messagebox, ttk, Tk, filedialog, Label
from datetime import datetime, timedelta
import os
import sys
import subprocess
import time
from threading import Thread  # run load workbook while progress bar runs
# non standard libraries
from openpyxl import Workbook, load_workbook
from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment
# define globals
global root  # used to hold the Tk() root for the new window used by all Informal C windows.
global pb_flag  #

""" this module has its own MakeWindow() class since it uses a different root. 
So it is not imported from kbtoolbar. """


def informalc_gen_clist(start, end, station):
    """ generates carrier list for informal c. """
    rec = None
    end += timedelta(weeks=52)
    sql = "SELECT * FROM carriers WHERE effective_date<='%s'and station='%s' " \
          "ORDER BY carrier_name, effective_date DESC" % (end, station)
    result = inquire(sql)
    unique_carriers = []  # create non repeating list of otdl carriers
    for name in result:
        if name[1] not in unique_carriers:
            unique_carriers.append(name[1])
    carrier_list = []
    for name in unique_carriers:
        sql = "SELECT effective_date,carrier_name,station FROM carriers WHERE carrier_name='%s' " \
              "ORDER BY effective_date DESC" % name
        after_start = []  # array for records after start date
        before_start = []  # array for records before start date
        added = False
        result = inquire(sql)
        for rec in result:
            if rec[0] >= str(start):
                after_start.append(rec)
            if rec[0] < str(start):
                before_start.append(rec)
        for rec in after_start:
            if not added and rec[2] == station:
                carrier_list.append(rec[1])
                added = True
        if not added and len(before_start) > 0:
            if before_start[0][2] == station:
                carrier_list.append(rec[1])
    return carrier_list


def informalc_date_converter(date):
    """ be sure to run informalc date checker before using this """
    sd = date.get().split("/")
    return datetime(int(sd[2]), int(sd[0]), int(sd[1]))


class InfcSpeedSheetGen:
    """ this generates and reads a speedsheet for the informal c grievance tracker """

    def __init__(self, frame, station, selection_range):
        self.frame = frame
        self.selection_range = selection_range
        self.station = station
        self.titles = []
        self.filename = ""
        # self.ws_titles = ["grievances", "settlements", "non compliance", "batch settlements", "remanded"]
        self.ws_titles = ["grievances", "settlements", "non compliance", "remanded", "batch settlements", "batch gats"]
        # get sql results from the tables.
        self.grievance_onrecs = []
        self.settlement_onrecs = []
        self.index_onrecs = []  # combine nonc, batch and remand onrecs
        self.file_result = []
        self.ws_list = []
        self.wb = Workbook()  # define the workbook
        self.ws = None  # the worksheet of the workbook
        self.ws_header = None  # styles for workbook
        self.list_header = None  # styles for workbook
        self.date_dov = None  # styles for workbook
        self.date_dov_title = None  # styles for workbook
        self.col_header = None  # styles for workbook
        self.input_s = None  # styles for workbook
        self.input_ns = None  # styles for workbook
        self.index_columns = [
            ["settlement", "follow up"],  # non compliance index
            ["remanded", "follow up"],  # remanded index
            ["main", "sub"],  # batch settlement index
            ["main", "sub"]  # batch gats index
        ]

    def run(self):
        pass

    def new(self):
        """ this generates a blank speedsheet for new greivances"""
        self.name_styles()
        self.get_titles()  # generate the title and filename
        self.make_workbook_object()  # make the workbook object
        self.create_ws_headers()
        self.create_grievance_headers()
        self.create_settlement_headers()
        self.create_index_headers()
        self.column_formatting_grievances()  # format sheet column widths, fonts, numbers
        self.column_formatting_settlements()  # format sheet column widths, fonts, numbers
        self.column_formatting_indexes()  # format sheet column widths, fonts, numbers
        self.stopsaveopen()

    def selected(self):
        """ this generates a blank speedsheet for selected range of greivances"""
        self.name_styles()
        self.get_titles()  # generate the title and filename
        self.make_workbook_object()  # make the workbook object
        self.create_ws_headers()
        self.create_grievance_headers()
        self.create_settlement_headers()
        self.create_index_headers()
        self.column_formatting_grievances()  # format sheet column widths, fonts, numbers
        self.column_formatting_settlements()  # format sheet column widths, fonts, numbers
        self.column_formatting_indexes()  # format sheet column widths, fonts, numbers
        self.stopsaveopen()

    def all(self):
        """ this generates a blank speedsheet for all greivances"""
        self.name_styles()
        self.get_titles()  # generate the title and filename
        self.get_onrecs()  # get data from all tables to fill speedsheets
        self.make_workbook_object()  # make the workbook object
        self.create_ws_headers()
        self.create_grievance_headers()
        self.create_settlement_headers()
        self.create_index_headers()
        self.column_formatting_grievances()  # format sheet column widths, fonts, numbers
        self.column_formatting_settlements()  # format sheet column widths, fonts, numbers
        self.column_formatting_indexes()  # format sheet column widths, fonts, numbers
        self.insert_grievance_onrecs()  # fills the grievance speedsheet with data from informalc grievances table
        self.insert_settlement_onrecs()  # fills the settlement speedsheet with data from the informalc settlements
        self.insert_index_onrecs()
        self.stopsaveopen()

    def name_styles(self):
        """ Named styles for workbook """
        bd = Side(style='thin', color="80808080")  # defines borders
        self.ws_header = NamedStyle(name="ws_header", font=Font(bold=True, name='Arial', size=12))
        self.list_header = NamedStyle(name="list_header", font=Font(bold=True, name='Arial', size=9))
        self.date_dov = NamedStyle(name="date_dov", font=Font(name='Arial', size=9))
        self.date_dov_title = NamedStyle(name="date_dov_title", font=Font(bold=True, name='Arial', size=9),
                                         alignment=Alignment(horizontal='right'))
        self.col_header = NamedStyle(name="col_header", font=Font(bold=True, name='Arial', size=9),
                                     border=Border(left=bd, top=bd, right=bd, bottom=bd),
                                     alignment=Alignment(horizontal='left'))

    def get_titles(self):
        """ generate title and filename. The titles and file names vary depending on the selection
        range - new, selected, or all inclusive. This is passed in the command calling SpeedSheetGen. """
        text = "New"
        filetext = "new"
        if self.selection_range == "selected":
            text = "Selected"
            filetext = "selected"
        if self.selection_range == "all":
            text = "All"
            filetext = "all"
        self.titles = [
            "Speedsheet - {} Grievances".format(text),
            "Speedsheet - {} Settlements ".format(text),
            "Speedsheet - {} Non Compliance Index".format(text),
            "Speedsheet - {} Remanded Index".format(text),
            "Speedsheet - {} Batch Settlement Index".format(text),
            "Speedsheet - {} Batch Gats Index".format(text)
        ]
        self.filename = "{}_grievances_speedsheet".format(filetext) + ".xlsx"

    def get_onrecs(self):
        """ get data from tables """
        nonc_onrecs = []
        remand_onrecs = []
        batchset_onrecs = []  # store sql results for batch settlements
        batchgats_onrecs = []  # store sql results for batch gats
        sql = "SELECT * FROM 'informalc_grievances' WHERE station = '%s'" % self.station
        self.grievance_onrecs = inquire(sql)
        grv_list = []  # array to hold all grievance numbers
        for grv in self.grievance_onrecs:
            grv_list.append(grv[2])
        # use arrays and loops to get search results for all the grievances in the grv_list array.
        # search these tables
        tables_array = ("informalc_settlements", "informalc_noncindex", "informalc_remandindex",
                        "informalc_batchindex", "informalc_gatsindex")
        # search these columns in the tables
        search_criteria_array = ("grv_no", "settlement", "remanded", "main", "main")
        for i in range(len(tables_array)):  # loop for each table
            for ii in range(len(grv_list)):  # loop for every grv in the grv list array.
                sql = "SELECT * FROM '%s' WHERE %s = '%s'" % (tables_array[i], search_criteria_array[i], grv_list[ii])
                result = inquire(sql)
                # get the onrecs for informalc settlements
                if tables_array[i] == "informalc_settlements":
                    if result:
                        self.settlement_onrecs.append(result[0])
                if tables_array[i] == "informalc_noncindex":  # get the onrecs for informalc non compliance index
                    if result:  # if there is a result
                        for r in result:  # there can be multiple results
                            nonc_onrecs.append(r)  # add record to the array
                if tables_array[i] == "informalc_remandindex":  # get the onrecs for informalc_remandindex
                    if result:
                        for r in result:
                            remand_onrecs.append(r)
                if tables_array[i] == "informalc_batchindex":  # get the onrecs for informalc_batchindex
                    if result:
                        for r in result:
                            batchset_onrecs.append(r)
                if tables_array[i] == "informalc_batchgats":  # get the onrecs for informalc_batchindex
                    if result:
                        for r in result:
                            batchgats_onrecs.append(r)
        self.index_onrecs = [nonc_onrecs, remand_onrecs, batchset_onrecs, batchgats_onrecs]

    def make_workbook_object(self):
        """ make the workbook object """
        self.ws_list = ["grievances", "settlements", "non compliance", "remanded", "batch settlements", "batch gats"]
        self.ws_list[0] = self.wb.active  # create first worksheet - this will be for grievances
        self.ws_list[0].title = self.ws_titles[0]  # title first worksheet - this is for grievances
        for i in range(1, len(self.ws_list)):  # loop to create all other worksheets
            self.ws_list[i] = self.wb.create_sheet(self.ws_titles[i])

    def create_ws_headers(self):
        """ use a loop to create headers for all the worksheets """
        for i in range(6):  # there are six worksheets
            cell = self.ws_list[i].cell(column=1, row=1)
            cell.value = self.titles[i]
            cell.style = self.ws_header
            self.ws_list[i].merge_cells('A1:G1')
            cell = self.ws_list[i].cell(column=1, row=3)
            cell.value = "Station: "
            cell.style = self.date_dov_title
            cell = self.ws_list[i].cell(column=2, row=3)
            cell.value = self.station
            cell.style = self.date_dov
            self.ws_list[i].merge_cells('B3:C3')

    def create_grievance_headers(self):
        """ create the grievance worksheet. all worksheets must be formatted separately since they all have
        distinct information. """
        cell = self.ws_list[0].cell(column=1, row=5)
        cell.value = "grievant"
        cell.style = self.col_header
        cell = self.ws_list[0].cell(column=2, row=5)
        cell.value = "grievance number"
        cell.style = self.col_header
        cell = self.ws_list[0].cell(column=3, row=5)
        cell.value = "start incident"
        cell.style = self.col_header
        cell = self.ws_list[0].cell(column=4, row=5)
        cell.value = "end incident"
        cell.style = self.col_header
        cell = self.ws_list[0].cell(column=5, row=5)
        cell.value = "meeting date"
        cell.style = self.col_header
        cell = self.ws_list[0].cell(column=6, row=5)
        cell.value = "issue"
        cell.style = self.col_header
        cell = self.ws_list[0].cell(column=7, row=5)
        cell.value = "article"
        cell.style = self.col_header
        cell = self.ws_list[0].cell(column=8, row=5)
        cell.value = "action"
        cell.style = self.col_header
        # freeze panes
        self.ws_list[0].freeze_panes = self.ws_list[0].cell(row=6, column=1)

    def create_settlement_headers(self):
        """ create the grievance worksheet. all worksheets must be formatted separately since they all have
        distinct information. """
        cell = self.ws_list[1].cell(column=1, row=5)
        cell.value = "grievance number"
        cell.style = self.col_header
        cell = self.ws_list[1].cell(column=2, row=5)
        cell.value = "level"
        cell.style = self.col_header
        cell = self.ws_list[1].cell(column=3, row=5)
        cell.value = "date signed"
        cell.style = self.col_header
        cell = self.ws_list[1].cell(column=4, row=5)
        cell.value = "decision"
        cell.style = self.col_header
        cell = self.ws_list[1].cell(column=5, row=5)
        cell.value = "proof due"
        cell.style = self.col_header
        cell = self.ws_list[1].cell(column=6, row=5)
        cell.value = "docs"
        cell.style = self.col_header
        cell = self.ws_list[1].cell(column=7, row=5)
        cell.value = "gats number"
        cell.style = self.col_header
        cell = self.ws_list[1].cell(column=8, row=5)
        cell.value = "action"
        cell.style = self.col_header
        # freeze panes
        self.ws_list[1].freeze_panes = self.ws_list[1].cell(row=6, column=1)

    def create_index_headers(self):
        """ use a loop to fill in the index headers using self.index_columns """
        for i in range(4):
            cell = self.ws_list[i+2].cell(column=1, row=5)
            cell.value = self.index_columns[i][0]
            cell.style = self.col_header
            cell = self.ws_list[i+2].cell(column=2, row=5)
            cell.value = self.index_columns[i][1]
            cell.style = self.col_header
            cell = self.ws_list[i + 2].cell(column=3, row=5)
            cell.value = "action"
            cell.style = self.col_header
            # freeze panes
            self.ws_list[i + 2].freeze_panes = self.ws_list[i + 2].cell(row=6, column=1)

    def column_formatting_grievances(self):
        """ format the columns. this can be overridden by individually formating the cells. """
        self.ws_list[0].oddFooter.center.text = "&A"
        col = self.ws_list[0].column_dimensions["A"]  # grievant
        col.width = 25
        col.font = Font(size=9, name="Arial")
        col = self.ws_list[0].column_dimensions["B"]  # grievance number
        col.width = 20
        col.font = Font(size=9, name="Arial")
        col.number_format = '@'
        col = self.ws_list[0].column_dimensions["C"]  # start incident
        col.width = 12
        col.font = Font(size=9, name="Arial")
        col.number_format = 'MM/DD/YYYY'
        col = self.ws_list[0].column_dimensions["D"]  # end incident
        col.width = 12
        col.font = Font(size=9, name="Arial")
        col.number_format = 'MM/DD/YYYY'
        col = self.ws_list[0].column_dimensions["E"]  # meeting date
        col.width = 12
        col.font = Font(size=9, name="Arial")
        col.number_format = 'MM/DD/YYYY'
        col = self.ws_list[0].column_dimensions["F"]  # issue
        col.width = 25
        col.font = Font(size=9, name="Arial")
        col = self.ws_list[0].column_dimensions["G"]  # article
        col.width = 6
        col.font = Font(size=9, name="Arial")
        col = self.ws_list[0].column_dimensions["H"]  # action
        col.width = 9
        col.font = Font(size=9, name="Arial")
        
    def column_formatting_settlements(self):
        """ format the columns. this can be overridden by individually formating the cells. """
        self.ws_list[1].oddFooter.center.text = "&A"
        col = self.ws_list[1].column_dimensions["A"]  # grievance number
        col.width = 18
        col.font = Font(size=9, name="Arial")
        col.number_format = '@'
        col = self.ws_list[1].column_dimensions["B"]  # level
        col.width = 10
        col.font = Font(size=9, name="Arial")
        col = self.ws_list[1].column_dimensions["C"]  # date signed
        col.width = 10
        col.font = Font(size=9, name="Arial")
        col.number_format = 'MM/DD/YYYY'
        col = self.ws_list[1].column_dimensions["D"]  # decision
        col.width = 20
        col.font = Font(size=9, name="Arial")
        col = self.ws_list[1].column_dimensions["E"]  # proof due
        col.width = 10
        col.font = Font(size=9, name="Arial")
        col.number_format = 'MM/DD/YYYY'
        col = self.ws_list[1].column_dimensions["F"]  # docs
        col.width = 15
        col.font = Font(size=9, name="Arial")
        col = self.ws_list[1].column_dimensions["G"]  # gats_number
        col.width = 12
        col.font = Font(size=9, name="Arial")
        col = self.ws_list[1].column_dimensions["H"]  # action
        col.width = 9
        col.font = Font(size=9, name="Arial")

    def column_formatting_indexes(self):
        """ format the columns of all index worksheets - non compliance, batch settlements and remanded"""
        for i in range(4):
            self.ws_list[i+2].oddFooter.center.text = "&A"
            col = self.ws_list[i+2].column_dimensions["A"]  # settlement/main/remanded
            col.width = 20
            col.font = Font(size=9, name="Arial")
            col.number_format = '@'
            col = self.ws_list[i+2].column_dimensions["B"]  # followup/sub/followup
            col.width = 20
            col.font = Font(size=9, name="Arial")
            col.number_format = '@'
            col = self.ws_list[i + 2].column_dimensions["C"]  # action
            col.width = 9
            col.font = Font(size=9, name="Arial")

    def insert_grievance_onrecs(self):
        """ loop for each grievance on record to fill the grievance speedsheet which is ws.list[0] """
        row = 6  # start on row 6 to make room for headers
        for grv in self.grievance_onrecs:
            grievant = grv[0]
            grievance_number = grv[2]
            start_incident = Convert(grv[3]).dtstr_to_backslashstr()
            end_incident = Convert(grv[4]).dtstr_to_backslashstr()
            meeting_date = Convert(grv[5]).dtstr_to_backslashstr()
            issue = grv[6]
            article = grv[7]
            values_array = [grievant, grievance_number, start_incident, end_incident, meeting_date, 
                            issue, article]
            for i in range(len(values_array)):
                cell = self.ws_list[0].cell(row=row, column=i+1)  # carrier effective date
                cell.value = values_array[i]
                if i in (2, 3, 4):
                    cell.number_format = 'MM/DD/YYYY'
            row += 1

    def insert_settlement_onrecs(self):
        """ loop for each grievance on record to fill the grievance speedsheet which is ws.list[0] """
        row = 6  # start on row 6 to make room for headers
        for sett in self.settlement_onrecs:  # loop for each row
            grievance_number = sett[0]  # define all the fields
            level = sett[1]
            date_signed = Convert(sett[2]).dtstr_to_backslashstr()
            decision = sett[3]
            proofdue = Convert(sett[4]).dtstr_to_backslashstr()
            docs = sett[5]
            gats_number = sett[6]
            values_array = [grievance_number, level, date_signed, decision,
                            proofdue, docs, gats_number]
            for i in range(len(values_array)):  # loop for each column
                cell = self.ws_list[1].cell(row=row, column=i+1)  # define the cell by sheet and cell coordinates
                cell.value = values_array[i]  # insert the appropriate element
                if i in (2, 4):  # for date signed and proof due, format the cell as a date.
                    cell.number_format = 'MM/DD/YYYY'
                    cell.style = self.date_dov
            row += 1

    def insert_index_onrecs(self):
        # """ loop for each table - non compliance, batch and remanded"""
        # sheet_count = 2  # 2 non compliance, 3 batch and 4 remanded
        """ loop for each table - non compliance, remanded, batch settlements and batch gats"""
        sheet_count = 2  # 2 non compliance, 3 remanded 4 batch set and 5 batch gats
        for index in self.index_onrecs:  # there are four indexes...
            row = 6  # start on row 6 to make room for headers
            for rec in index:  # loop for each record in the index
                first = rec[0]
                second = rec[1]
                values_array = [first, second]
                for i in range(len(values_array)):  # loop for each column
                    # define the cell by sheet and cell coordinates
                    cell = self.ws_list[sheet_count].cell(row=row, column=i + 1)
                    cell.value = values_array[i]  # insert the appropriate element
                    cell.number_format = '@'
                    cell.style = self.date_dov
                row += 1
            sheet_count += 1

    def stopsaveopen(self):
        """ save and open the speedsheet. """
        try:
            self.wb.save(dir_path('informalc_speedsheets') + self.filename)
            messagebox.showinfo("Speedsheet Generator",
                                "Your speedsheet was successfully generated. \n"
                                "File is named: {}".format(self.filename),
                                parent=self.frame)
            if sys.platform == "win32":
                os.startfile(dir_path('informalc_speedsheets') + self.filename)
            if sys.platform == "linux":
                subprocess.call(["xdg-open", 'kb_sub/informalc_speedsheets/' + self.filename])
            if sys.platform == "darwin":
                subprocess.call(["open", dir_path('informalc_speedsheets') + self.filename])
        except PermissionError:
            messagebox.showerror("Speedsheet generator",
                                 "The speedsheet was not generated. \n"
                                 "Suggestion: \n"
                                 "Make sure that identically named informalc_speedsheets are closed \n"
                                 "(the file can't be overwritten while open).\n",
                                 parent=self.frame)


class InfcSpeedWorkBookGet:
    """
    this class gets the speedsheet and opens it.
    """

    def __init__(self):
        pass

    @staticmethod
    def get_filepath():
        """ get the file path"""
        if projvar.platform == "macapp" or projvar.platform == "winapp":
            return os.path.join(os.path.sep,
                                os.path.expanduser("~"), 'Documents', 'klusterbox', 'informalc_speedsheets')
        else:
            return 'kb_sub/informalc_speedsheets'

    def get_file(self):
        """ returns the file path if there is one. else no selection/invalid selection. """
        path_ = self.get_filepath()
        file_path = filedialog.askopenfilename(initialdir=path_, filetypes=[("Excel files", "*.xlsx")])
        if file_path[-5:].lower() == ".xlsx":
            return file_path
        elif file_path == "":
            return "no selection"
        else:
            return "invalid selection"

    def open_file(self, frame, interject):
        """ gets the file and calls the speedsheet check and progress bar. """
        global pb_flag
        pb_flag = True
        file_path = self.get_file()
        if file_path == "no selection":
            return
        elif file_path == "invalid selection":
            messagebox.showerror("Report Generator",
                                 "The file you have selected is not an .xlsx file. "
                                 "You must select a file with a .xlsx extension.",
                                 parent=frame)
            return
        else:
            pb = ProgressBarIn(title="Klusterbox", label="SpeedSheeets Loading",
                               text="Loading and reading workbook. This could take a minute")
            wb = SpeedLoadThread(file_path)  # open workbook in separate thread
            wb.start()  # start loading workbook
            pb.start_up()  # start progress bar
            wb.join()  # wait for loading workbook to finish
            pb.stop()  # stop the progress bar and destroy the object
            SpeedSheetCheck(frame, wb.workbook, file_path, interject).check()  # check the speedsheet


class SpeedLoadThread(Thread):
    """ use multithreading to load workbook while progress bar runs """

    def __init__(self, path_):
        Thread.__init__(self)
        self.path_ = path_
        self.workbook = ""

    def run(self):
        """ runs the speedsheet loading. """
        global pb_flag  # this will signal when the thread has ended to end the progress bar
        wb = load_workbook(self.path_)  # load xlsx doc with openpyxl
        self.workbook = wb
        pb_flag = False


class SpeedSheetCheck:
    """ a class for checking the informal c grievance speedsheets. """
    def __init__(self, frame, wb, path_, interject):
        self.frame = frame
        self.station = None
        self.wb = wb
        self.ws = None  # this hold the worksheet
        self.path_ = path_
        self.interject = interject
        self.input_type = None
        self.sheets = None
        self.sheet_count = None
        self.grievance_count = 0  # count of how many grievances have been checked.
        self.fatal_rpt = 0
        self.add_rpt = 0
        self.fyi_rpt = 0
        self.settlement_count = 0  # count of how many settlements have been checked
        self.settlement_fatal_rpt = 0
        self.settlement_add_rpt = 0
        self.settlement_fyi_rpt = 0
        # count of how many index (non compliance, remanded, batch settlements, and batch gats) have been checked
        self.index_count = [0, 0, 0, 0]
        self.index_fatal_rpt = [0, 0, 0, 0]
        self.index_add_rpt = [0, 0, 0, 0]
        self.index_fyi_rpt = [0, 0, 0, 0]
        self.sheet_rowcount = []
        self.row_counter = 0  # get the total amount of rows in the worksheet
        self.start_row = 6  # the row where after the headers
        self.pb = None  # progress bar object
        self.pb_counter = 0
        self.filename = ReportName("speedsheet_precheck").create()  # generate a name for the report
        self.report = open(dir_path('report') + self.filename, "w")  # open the report
        self.grv_mentioned = False  # keeps grievance numbers from being repeated in reports
        self.worksheet = ("grievances", "settlements", "non compliance", "remanded", "batch set", "batch gats")
        self.index_columns = [
            ["settlement", "followup"],  # non compliance index
            ["remanded", "followup"],  # remanded index
            ["main", "sub"],  # batch settlement index
            ["main", "sub"]  # batch gats index
        ]
        self.allowaddrecs = True
        self.fullreport = True
        self.name_mentioned = False
        self.issue_index = []  # get the speedsheet issue index number for issue categories
        self.issue_description = []  # get the issue description for issue categories
        self.issue_article = []  # get the article of the issue for issue catergories
        self.decision_index = []  # get the speedsheet decision index number for decision categories
        self.decision_description = []  # get the decision description for the decision categories
        self.del_settlement = []  # when action is delete - delete these settlements
        self.del_batch = []  # when action is delete - delete these batch index recs
        self.del_gatsbatch = []  # when action is delete - delete these gats batch index recs
        self.del_nonc = []  # when action is delete - delete these non compliance index recs
        self.del_remanded = []  # when action is delete - delete these remanded index recs
        self.del_location = []  # when action is delete - delete these location recs

    def check(self):
        """ master method for running other methods and returns to the mainframe. """
        try:
            self.pb = ProgressBarDe(label="SpeedSheet Checking")
            self.get_issuecats()  # fetch the issue categories from the informalc_issuescategories table
            self.get_decisioncats()  # fetch the decision categories from the informalc_decisioncategories table
            self.set_sheet_facts()
            self.set_station()
            self.start_reporter()
            self.checking()
            self.reporter()
            self.pb.stop()
        except KeyError:  # if wrong type of file is selected, there will be an error
            self.pb.delete()  # stop and destroy progress bar
            self.showerror()

    def get_issuecats(self):
        """ fetch the issue categories from the informalc_issuescategories table of the db and place them in arrays. """
        sql = "SELECT * FROM informalc_issuescategories"
        results = inquire(sql)
        for r in results:
            self.issue_index.append(r[0])
            self.issue_description.append(r[2])
            self.issue_article.append(r[1])

    def get_decisioncats(self):
        """ fetch the decision categories from the informalc_decisioncategories table of the db and place them in
         arrays """
        sql = "SELECT * FROM informalc_decisioncategories"
        results = inquire(sql)
        for r in results:
            self.decision_index.append(r[0])
            self.decision_description.append(r[2])

    def set_sheet_facts(self):
        """ get the worksheet names and number worksheets. """
        # there are three input types: new, selected, or all inclusive
        self.input_type = "new"
        self.sheets = self.wb.sheetnames  # get the names of the worksheets as a list
        self.sheet_count = len(self.sheets)  # get the number of worksheets

    def set_station(self):
        """ gets the station from the speedsheet. """
        self.station = self.wb[self.sheets[0]].cell(row=3, column=2).value  # get the station.

    def start_reporter(self):
        """ starts the report. """
        self.report.write("\nSpeedSheet Pre-Check Report \n")
        self.report.write(">>> {}\n".format(self.path_))

    def row_count(self):
        """ get a count of all rows for all sheets - need for progress bar """
        total_rows = 0
        for i in range(self.sheet_count):
            ws = self.wb[self.sheets[i]]  # assign the worksheet object
            row_count = ws.max_row  # get the total amount of rows in the worksheet
            self.sheet_rowcount.append(row_count)
            total_rows += row_count
        return total_rows

    def showerror(self):
        """ message box for showing errors. """
        messagebox.showerror("Klusterbox SpeedSheets",
                             "SpeedSheets Precheck or Input has failed. \n"
                             "Either you have selected a spreadsheet that is not \n"
                             "a SpeedSheet or your Speedsheet is corrupted. \n"
                             "Suggestion: Verify that the file you are selecting \n "
                             "is a SpeedSheet. \n"
                             "Suggestion: Try re-generating the SpeedSheet.",
                             parent=self.frame)

    def checking(self):
        """ reads rows and send to scan grievances, scan settlements or scan indexes. """
        # self.worksheet = ("grievances", "settlements", "non compliance", "remanded", "batch set", "batch gats")
        count_diff = self.sheet_count * (self.start_row - 1)  # subtract top five/six rows from the row count
        self.pb.max_count(self.row_count() - count_diff)  # get total count of rows for the progress bar
        self.pb.start_up()  # start up the progress bar
        self.pb_counter = 0  # initialize the progress bar counter
        for i in range(self.sheet_count):  # loop once for each worksheet in the workbook
            self.ws = self.wb[self.sheets[i]]  # assign the worksheet object
            self.row_counter = self.ws.max_row  # get the total amount of rows in the worksheet
            if self.worksheet[i] == "grievances":  # execute for grievance speedsheet
                self.scan_grievances(i)
            if self.worksheet[i] == "settlements":  # execute for settlements speedsheet
                self.scan_settlements(i)
            if self.worksheet[i] == "non compliance":
                self.scan_indexes(i)
            if self.worksheet[i] == "remanded":
                self.scan_indexes(i)
            if self.worksheet[i] == "batch set":
                self.scan_indexes(i)
            if self.worksheet[i] == "batch gats":
                self.scan_indexes(i)

    def scan_grievances(self, i):
        """ scan the values of the grievances worksheet, line by line. """
        # loop through all rows, start with row 5 or 6 until the end
        for ii in range(self.start_row, self.row_counter + 1):
            self.pb.move_count(self.pb_counter)
            self.grv_mentioned = False  # keeps names from being repeated in reports
            self.grievance_count += 1  # get a count of the carriers for reports
            grievant = Handler(self.ws.cell(row=ii, column=1).value).nonetype()
            grv_no = Handler(self.ws.cell(row=ii, column=2).value).nonetype()
            startdate = Handler(self.ws.cell(row=ii, column=3).value).nonetype()
            enddate = Handler(self.ws.cell(row=ii, column=4).value).nonetype()
            meetingdate = Handler(self.ws.cell(row=ii, column=5).value).nonetype()
            issue = Handler(self.ws.cell(row=ii, column=6).value).nonetype()
            article = Handler(self.ws.cell(row=ii, column=7).value).nonetype()
            action = Handler(self.ws.cell(row=ii, column=8).value).nonetype()
            self.pb.change_text("Reading Speedcell: {}".format(grv_no))  # update text for progress bar
            SpeedGrvCheck(self, self.sheets[i], ii, grievant, grv_no, startdate, enddate, meetingdate,
                          issue, article, action).check_all()
        self.pb_counter += 1

    def scan_settlements(self, i):
        """ scan the values of the grievances worksheet, line by line. """
        for ii in range(self.start_row,
                        self.row_counter + 1):  # loop through all rows, start with row 5 or 6 until the end
            if self.ws.cell(row=ii, column=1).value is not None:  # if there is a grievance number
                self.pb.move_count(self.pb_counter)
                self.grv_mentioned = False  # keeps names from being repeated in reports
                self.settlement_count += 1  # get a count of the carriers for reports
                grv_no = Handler(self.ws.cell(row=ii, column=1).value).nonetype()
                level = Handler(self.ws.cell(row=ii, column=2).value).nonetype()
                datesigned = Handler(self.ws.cell(row=ii, column=3).value).nonetype()
                decision = Handler(self.ws.cell(row=ii, column=4).value).nonetype()
                proofdue = Handler(self.ws.cell(row=ii, column=5).value).nonetype()
                docs = Handler(self.ws.cell(row=ii, column=6).value).nonetype()
                gatsnumber = Handler(self.ws.cell(row=ii, column=7).value).nonetype()
                action = Handler(self.ws.cell(row=ii, column=8).value).nonetype()
                self.pb.change_text("Reading Speedcell: {}".format(grv_no))  # update text for progress bar
                SpeedSetCheck(self, self.sheets[i], ii, grv_no, level, datesigned, decision, proofdue, docs,
                              gatsnumber, action).check_all()
        self.pb_counter += 1

    def scan_indexes(self, i):
        """ scan the values of the index worksheets, line by line. """
        # loop through all rows, start with row 5 or 6 until the end
        for ii in range(self.start_row, self.row_counter + 1):
            # if there is a grievance number for both columns
            if self.ws.cell(row=ii, column=1).value is not None and self.ws.cell(row=ii, column=2).value is not None:
                self.pb.move_count(self.pb_counter)
                self.grv_mentioned = False  # keeps names from being repeated in reports
                self.index_count[i-2] += 1  # get a count of the carriers for reports
                first = Handler(self.ws.cell(row=ii, column=1).value).nonetype()
                second = Handler(self.ws.cell(row=ii, column=2).value).nonetype()
                action = Handler(self.ws.cell(row=ii, column=3).value).nonetype()
                # update text for progress bar
                self.pb.change_text("Reading Speedcell: {}".format(self.index_columns[i-2][0]))
                SpeedIndexCheck(self, i, self.sheets[i], ii, first, second, action).check_all()
        self.pb_counter += 1

    def reporter(self):
        """ writes the report """
        self.report.write("\n\n----------------------------------")
        # build report summary for grievance checks
        self.report.write("\n\nGrievance SpeedSheet Check Complete.\n\n")
        msg = "grievance{} checked".format(Handler(self.grievance_count).plurals())
        self.report.write('{:>6}  {:<40}\n'.format(self.grievance_count, msg))
        msg = "fatal error{} found".format(Handler(self.fatal_rpt).plurals())
        self.report.write('{:>6}  {:<40}\n'.format(self.fatal_rpt, msg))
        if self.interject:
            msg = "change{} made".format(Handler(self.add_rpt).plurals())
            self.report.write('{:>6}  {:<40}\n'.format(self.add_rpt, msg))
        else:
            msg = "fyi notification{}".format(Handler(self.fyi_rpt).plurals())
            self.report.write('{:>6}  {:<40}\n'.format(self.fyi_rpt, msg))
        # build report summary for settlement checks
        self.report.write("\n\nSettlements SpeedSheet Check Complete.\n\n")
        msg = "settlement{} checked".format(Handler(self.settlement_count).plurals())
        self.report.write('{:>6}  {:<40}\n'.format(self.settlement_count, msg))
        msg = "fatal error{} found".format(Handler(self.settlement_fatal_rpt).plurals())
        self.report.write('{:>6}  {:<40}\n'.format(self.settlement_fatal_rpt, msg))
        if self.interject:
            msg = "change{} made".format(Handler(self.settlement_add_rpt).plurals())
            self.report.write('{:>6}  {:<40}\n'.format(self.settlement_add_rpt, msg))
        else:
            msg = "fyi notification{}".format(Handler(self.settlement_fyi_rpt).plurals())
            self.report.write('{:>6}  {:<40}\n'.format(self.settlement_fyi_rpt, msg))
        # use a loop to write the report for 4 indexes (non compliance, remanded, batch settlements and batch gats).
        index_rpt_subheader = ("\n\nNon Compliance Index Check Complete.\n\n",
                               "\n\nRemanded Index Check Complete.\n\n",
                               "\n\nBatch Settlements Index Check Complete.\n\n",
                               "\n\nBatch Gats Index Check Complete.\n\n")
        for i in range(4):
            self.report.write(index_rpt_subheader[i])
            msg = "record{} checked".format(Handler(self.index_count[i]).plurals())
            self.report.write('{:>6}  {:<40}\n'.format(self.index_count[i], msg))
            msg = "fatal error{} found".format(Handler(self.index_fatal_rpt[i]).plurals())
            self.report.write('{:>6}  {:<40}\n'.format(self.index_fatal_rpt[i], msg))
            if self.interject:
                msg = "change{} made".format(Handler(self.index_add_rpt[i]).plurals())
                self.report.write('{:>6}  {:<40}\n'.format(self.index_add_rpt[i], msg))
            else:
                msg = "fyi notification{}".format(Handler(self.index_fyi_rpt[i]).plurals())
                self.report.write('{:>6}  {:<40}\n'.format(self.index_fyi_rpt[i], msg))
        # close out the report and open in notepad
        self.report.close()
        if sys.platform == "win32":  # open the text document
            os.startfile(dir_path('report') + self.filename)
        if sys.platform == "linux":
            subprocess.call(["xdg-open", 'kb_sub/report/' + self.filename])
        if sys.platform == "darwin":
            subprocess.call(["open", dir_path('report') + self.filename])


class SpeedGrvCheck:
    """ checks one line of the grievance speedsheet when it is called by the SpeedSheetCheck class. """
    def __init__(self, parent, sheet, row, grievant, grv_no, startdate, enddate, meetingdate, issue, article, action):
        self.parent = parent
        self.sheet = sheet  # input here is coming directly from the speedcell
        self.row = str(row)
        self.grievant = grievant
        self.grv_no = grv_no
        self.startdate = startdate
        self.enddate = enddate
        self.meetingdate = meetingdate
        self.input_date = []  # array to hold startdate, enddate and meetingdate - form in check_dates()
        self.issue = issue
        self.article = article
        self.action = action
        # onrec variables - these hold the values of the record currently in the database.
        self.onrec = False  # this value is True if a sql search shows that there is a rec of the grv_no in the db.
        self.onrec_grievant = ""
        # skip station as that is held in self.parent.station
        # skip grievance number as that is self.grv_no
        self.onrec_startdate = ""
        self.onrec_enddate = ""
        self.onrec_meetingdate = ""
        self.onrec_issue = ""
        self.onrec_article = ""
        self.error_array = []  # gives a report of failed checks
        self.attn_array = []  # gives a report of issues to bring to the attention of users
        self.add_array = []  # gives a report of records to add to the database
        self.fyi_array = []  # gives a report of useful information for the user
        self.parent.name_mentioned = False  # reset this so that name is not repeated on reports
        self.parent.allowaddrecs = True  # if False, records will not be added to database
        self.addday = []  # checked input formatted for entry into database
        self.addgrievant = "empty"
        self.addstartdate = "empty"
        self.addenddate = "empty"
        self.addmeetingdate = "empty"
        self.adddate = [self.addstartdate, self.addenddate, self.addmeetingdate]
        self.addissue = "empty"
        self.addarticle = "empty"

    def check_all(self):
        """ master method to run other methods. """
        self.reformat_grv_no()  # reformat the grievance number to all lowercase, no whitespaces, no dashes.
        self.get_onrecs()  # 'on record' - get the record currently in the database if it exist
        if not self.check_delete():
            if self.check_grv_number():  # first check the grievance number. if that is good, then proceed.
                self.check_grievant()
                self.check_dates()
                self.check_issue()
                self.add_recs()  # write changes to the db
        self.generate_report()

    def reformat_grv_no(self):
        """ reformat the grievance number to all lowercase, no whitespaces, no dashes. """
        self.grv_no = self.grv_no.lower()  # convert grievance number to lowercas
        self.grv_no = self.grv_no.strip()  # strip whitespace from start and end of the string.
        self.grv_no = self.grv_no.replace('-', '')  # remove any dashes
        self.grv_no = self.grv_no.replace(' ', '')  # remove any whitespace

    def get_onrecs(self):
        """ check if there is an existing record for the grievance number in the informalc grievances table.
        if so, store the values in the self.onrec variables. if not, the self.onrec variables default to empty. """
        sql = "SELECT * FROM informalc_grievances WHERE grv_no = '%s' and station = '%s'" \
              % (self.grv_no, self.parent.station)
        results = inquire(sql)
        if results:
            self.onrec = True  # this value is True if a sql search shows that there is a rec in the db.
            self.onrec_grievant = results[0][0]
            # skip station as that is held in self.parent.station and is part of the search criteria
            # skip grievance number as that is self.grv_no and is part of the search criteria
            self.onrec_startdate = results[0][3]
            self.onrec_enddate = results[0][4]
            self.onrec_meetingdate = results[0][5]
            self.onrec_issue = results[0][6]
            self.onrec_article = results[0][7]

    def check_delete(self):
        """ check the input for the third column and delete if that action is indicated. """
        delete_array = ("delete", "d/", "erase", "cut")
        if self.action in delete_array:
            if not self.parent.interject:
                if self.onrec:
                    fyi = "     FYI: DELETE Record for grievance number: {}\n"\
                        .format(self.grv_no)
                    self.fyi_array.append(fyi)
                else:
                    error = "     ERROR: CAN NOT DELETE Record for grievance number does not exist in database: {}\n" \
                        .format(self.grv_no)
                    self.error_array.append(error)
            else:
                # execute the delete in the database
                if self.onrec:
                    sql = "DELETE FROM informalc_grievances WHERE grv_no='%s'" % self.grv_no
                    commit(sql)
                    # create a message for the report
                    add = "     DELETE: Grievance record deleted from database >> {}\n" .format(self.grv_no)  # report
                    self.add_array.append(add)
                else:
                    # create a message for the report
                    error = "     CAN NOT DELETE: Grievance record does not exist in database >> {}\n"\
                        .format(self.grv_no)
                    self.error_array.append(error)
            #  when true the grv number will be deleted from settlement table in SpeeedSetCheck()
            #  and index tables in SpeedIndexCheck()
            self.parent.del_settlement.append(self.grv_no)
            self.parent.del_batch.append(self.grv_no)
            self.parent.del_gatsbatch.append(self.grv_no)
            self.parent.del_nonc.append(self.grv_no)
            self.parent.del_remanded.append(self.grv_no)
            self.parent.del_location.append(self.grv_no)
            return True
        return False

    def check_grv_number(self):
        """ check the grievance number input """
        if not GrievanceChecker(self.grv_no).has_value():
            error = "     ERROR: The grievance number must not be blank. \n"
            self.error_array.append(error)
            self.parent.allowaddrecs = False  # do not allow this speedcell be be input into database
            return False
        if not GrievanceChecker(self.grv_no).check_characters():
            error = "     ERROR: The grievance number can only contain numbers and letters. \n"
            self.error_array.append(error)
            self.parent.allowaddrecs = False  # do not allow this speedcell be be input into database
            return False
        if not GrievanceChecker(self.grv_no).min_lenght():
            error = "     ERROR: The grievance number must contain at least 4 characters. \n"
            self.error_array.append(error)
            self.parent.allowaddrecs = False  # do not allow this speedcell be be input into database
            return False
        if not GrievanceChecker(self.grv_no).max_lenght():
            error = "     ERROR: The grievance number can not contain more than 20 characters. \n"
            self.error_array.append(error)
            self.parent.allowaddrecs = False  # do not allow this speedcell be be input into database
            return False
        return True

    def reformat_grievant(self):
        """ reformat the grievant to all lowercase, no whitespaces """
        self.grievant = self.grievant.lower()
        self.grievant = self.grievant.strip()

    def check_grievant(self):
        """ check the grievant input. this is either 'class action' or a carrier name. it can be blank. """
        self.reformat_grievant()  # remove external whitespace and convert to lower case
        not_names = ("class action", "")
        if self.grievant in not_names:  # "class action" is a standard entry
            self.add_grievant()
            return
        if not NameChecker(self.grievant).check_characters():
            error = "     ERROR: Grievant name can not contain numbers or most special characters\n"
            self.error_array.append(error)
            self.parent.allowaddrecs = False  # do not allow this speedcell be be input into database
            return
        if not NameChecker(self.grievant).check_length():
            error = "     ERROR: Grievant name must not exceed 42 characters\n"
            self.error_array.append(error)
            self.parent.allowaddrecs = False  # do not allow this speedcell be be input into database
            return
        if not NameChecker(self.grievant).check_comma():
            error = "     ERROR: Grievant name must contain one comma to separate last name and first initial\n"
            self.error_array.append(error)
            self.parent.allowaddrecs = False  # do not allow this speedcell be be input into database
            return
        if not NameChecker(self.grievant).check_initial():
            attn = "     ATTENTION: Grievant name should must contain one initial ideally, \n" \
                   "                unless more are needed to create a distinct carrier name.\n"
            self.attn_array.append(attn)
        self.add_grievant()

    def add_grievant(self):
        """ add the grievant to add_grivant variable """
        if self.grievant == self.onrec_grievant:
            pass  # retain "empty" value for grievant variable
        else:
            fyi = "     FYI: New or updated grievant: {}\n".format(self.grievant)
            self.fyi_array.append(fyi)
            self.addgrievant = self.grievant  # save to input to dbase

    def check_dates(self):
        """ check the startdate, enddate and meetingdate.
         since these are all dates with similiar criteria, use a loop to check them.
         sometimes, openpyxl sends the dates as strings of datetime objects, instead of the mm/dd/yyyy formated dates,
         the DateTimeChecker() will identify these and skip the checks. """
        self.input_date = [self.startdate, self.enddate, self.meetingdate]
        for i in range(3):
            self.check_date_loop(i)

    def check_date_loop(self, i):
        """ loop from check dates """
        _type = ("start", "end", "meeting")
        if self.input_date[i].strip() == "":  # if the value is blank, skip all the checks
            self.add_date(i)
            return
        # if the value is a valid dt object, skip all the checks
        if DateTimeChecker().check_dtstring(self.input_date[i]):
            self.add_date(i)
            return
        date_object = BackSlashDateChecker(self.input_date[i])  # first create the date_object
        if not date_object.count_backslashes():  # this checks that there are 2 backslashes in the date
            error = "     ERROR: The date for the {} date must have two backslashes. Got instead: {}\n"\
                .format(_type[i], self.input_date[i])
            self.error_array.append(error)
            self.parent.allowaddrecs = False  # do not allow this speedcell be be input into database
            return
        date_object.breaker()  # this breaks the object into month, day and year elements.
        if not date_object.check_numeric():  # check each element in the date to ensure they are numeric
            error = "     ERROR: The month, day and year for the {} date must be numeric\n".format(_type[i])
            self.error_array.append(error)
            self.parent.allowaddrecs = False  # do not allow this speedcell be be input into database
            return
        if not date_object.check_minimums():  # check each element in the date to ensure they are greater than zero
            error = "     ERROR: The month, day and year for the {} date must be greater than zero.\n"\
                .format(_type[i])
            self.error_array.append(error)
            self.parent.allowaddrecs = False  # do not allow this speedcell be be input into database
            return
        if not date_object.check_month():  # returns False if the month is greater than 12.
            error = "     ERROR: The month for the {} date must less than 13.\n".format(_type[i])
            self.error_array.append(error)
            self.parent.allowaddrecs = False  # do not allow this speedcell be be input into database
            return
        if not date_object.check_day():  # return False if the day is greater than 31.
            error = "     ERROR: The day entered for the {} date is must be less than 32.\n".format(_type[i])
            self.error_array.append(error)
            self.parent.allowaddrecs = False  # do not allow this speedcell be be input into database
            return
        if not date_object.check_year():  # returns False if the year does not have 4 digits.
            error = "     ERROR: The year entered for the {} date must have 4 digits.\n".format(_type[i])
            self.error_array.append(error)
            self.parent.allowaddrecs = False  # do not allow this speedcell be be input into database
            return
        if not date_object.valid_date():  # returns False if the date is not a valid date
            error = "     ERROR: The date entered for the {} date is not a valid date.\n".format(_type[i])
            self.error_array.append(error)
            self.parent.allowaddrecs = False  # do not allow this speedcell be be input into database
            return
        # this removes white space from the date and each element of the date.
        self.input_date[i] = self.reformat_date(i)
        # convert the input date into a string of a datetime object.
        self.input_date[i] = Convert(self.input_date[i]).backslashdate_to_dtstring()
        self.add_date(i)  # add the dates to add_date variables

    def reformat_date(self, i):
        """ this removes white space from the date and each element of the date. """
        breakdown = self.input_date[i].strip()
        breakdown = breakdown.split("/")
        month = breakdown[0].strip()
        day = breakdown[1].strip()
        year = breakdown[2].strip()
        return "{}/{}/{}".format(month, day, year)

    def add_date(self, i):
        """ add the dates to add_date variables
         this is self.addstartdate, self.addenddate and self.addmeetingdate
         a counter is passed from the self.check_date method above. """
        onrec_date = [self.onrec_startdate, self.onrec_enddate, self.onrec_meetingdate]
        _type = ("start", "end", "meeting")
        if self.input_date[i] == onrec_date[i]:  # if the new input and the old record are the same - do nothing
            pass  # retain "empty" value for grievant variable
        else:
            fyi = "     FYI: New or updated {} date: {}\n".format(_type[i], self.input_date[i])
            self.fyi_array.append(fyi)
            self.adddate[i] = self.input_date[i]  # save to input to dbase

    def check_issue(self):
        """ check the issue input """
        self.issue = self.issue.strip().lower()  # strip out any whitespace before or after the string
        if self.issue == "":  # accept blank entries
            return
        if isint(self.issue):  # identify issue index entries and execute as valid - this also update the article
            self.check_issue_index()
            return
        self.check_issue_description()

    def check_issue_index(self):
        """ check that the issue index provided by the user is valid.
        use arrays of issue categories and articles collected in the SpeedSheetCheck class"""
        if self.issue in self.parent.issue_index:
            self.addissue = self.parent.issue_description[int(self.issue)-1]
            self.addarticle = self.parent.issue_article[int(self.issue)-1]
            fyi = "     FYI: New or updated issue and article (issue index entry): {} Article: {}\n"\
                .format(self.addissue, self.addarticle)
            self.fyi_array.append(fyi)
            return
        error = "     ERROR: The number for issue is in the index of issues. Got: {}\n".format(self.issue)
        self.error_array.append(error)
        self.parent.allowaddrecs = False  # do not allow this speedcell be be input into database

    def check_issue_description(self):
        """ check if the issue description is already in the list of issues. If so, update article. """
        if self.issue in self.parent.issue_description:
            index = self.parent.issue_description.index(self.issue)
            self.addarticle = self.parent.issue_article[index]
            fyi = "     FYI: New or updated issue and article (issue description entry): {} Article: {}\n" \
                .format(self.addissue, self.addarticle)
            self.add_issue(fyi)
            return
        fyi = "     FYI: New or updated issue: {}\n" \
            .format(self.addissue)
        self.add_issue(fyi)

    def add_issue(self, msg):
        """ add the issue to the add issue var """
        if self.issue == self.onrec_issue:
            pass
        else:
            self.addissue = self.issue
            self.fyi_array.append(msg)

    def check_article(self):
        """ check the article input """
        self.article = self.article.strip()
        if not self.article:
            self.add_article()
            return
        if not isint(self.issue):
            error = "     ERROR: The number the article must be a whole number. Got: {}\n".format(self.issue)
            self.error_array.append(error)
            self.parent.allowaddrecs = False  # do not allow this speedcell be be input into database
            return

    def add_article(self):
        """ add the article to the add_article var """
        if self.article == self.onrec_article:
            return
        else:
            fyi = "     FYI: New or updated article: {}\n".format(self.article)
            self.fyi_array.append(fyi)
            self.addarticle = self.article

    def add_recs(self):
        """ add records using the add___ vars. """
        chg_these = []
        if not self.onrec:  # if there is no record of the grievance number in the db informalc_grievance table
            fyi = "     FYI: New Grievance Number to add to database >>{}\n" \
                .format(self.grv_no)  # report
            self.fyi_array.append(fyi)
            add = "     INPUT: New Grievance Number added to database >>{}\n" \
                .format(self.grv_no)  # report
            self.add_array.append(add)
            chg_these.append('grv_no')
        if not self.parent.allowaddrecs:  # if all checks passed
            return
        # get grievant place
        if self.addgrievant != "empty":
            add = "     INPUT: Grievant added or updated to database >>{}\n" \
                .format(self.addgrievant)  # report
            self.add_array.append(add)
            chg_these.append("grievant")
            grievant_place = self.addgrievant
        else:
            grievant_place = self.onrec_grievant
        # get date places using loop
        onrec_date = [self.onrec_startdate, self.onrec_enddate, self.onrec_meetingdate]
        startdate_place = None
        enddate_place = None
        meetingdate_place = None
        date_place = [startdate_place, enddate_place, meetingdate_place]
        chg_notation = ("startdate", "enddate", "meetingdate")
        _type = ("Start", "End", "Meeting")
        for i in range(3):
            if self.adddate[i] != "empty":
                add = "     INPUT: {} Date added or updated to database >>{}\n".format(_type[i], self.adddate[i])
                self.add_array.append(add)
                chg_these.append(chg_notation[i])
                date_place[i] = self.adddate[i]
            else:
                date_place[i] = onrec_date[i]
        # get issue place
        if self.addissue != "empty":
            add = "     INPUT: Issue added or updated to database >>{}\n".format(self.addissue)  # report
            self.add_array.append(add)
            chg_these.append("issue")
            issue_place = self.addissue
        else:
            issue_place = self.onrec_issue
        # get article place
        # the addarticle might be assigned a value in self.check_issue_description() so check against onrec
        if self.addarticle == self.onrec_article:
            article_place = self.onrec_article
        elif self.addarticle != "empty":
            add = "     INPUT: Article added or updated to database >>{}\n".format(self.addarticle)  # report
            self.add_array.append(add)
            chg_these.append("article")
            article_place = self.addarticle
        else:
            article_place = self.onrec_article
        if not self.parent.interject:  # if 'pre check' is selected
            return  # do no update/insert into the database
        # if any values have changed - form sql statements using _place vars and commit to db.
        if len(chg_these) != 0:  # if change these is empty, then there is no need to insert/update records
            if not self.onrec:  # if there is no rec on file for the grievance, insert the first rec
                sql = "INSERT INTO informalc_grievances(grievant, station, grv_no, startdate, enddate, " \
                      "meetingdate, issue, article) VALUES('%s','%s','%s','%s','%s','%s','%s','%s')" \
                      % (grievant_place, self.parent.station, self.grv_no, date_place[0], date_place[1],
                         date_place[2], issue_place, article_place)
            else:  # update the first rec to replace pre existing record.
                sql = "UPDATE informalc_grievances SET grievant='%s', startdate='%s', enddate ='%s', " \
                      "meetingdate='%s', issue='%s', article='%s' WHERE grv_no='%s' and station='%s'" \
                      % (grievant_place, date_place[0], date_place[1], date_place[2], issue_place, article_place,
                         self.grv_no, self.parent.station)
            commit(sql)

    def generate_report(self):
        """ generate a report """
        self.parent.fatal_rpt += len(self.error_array)
        if not self.parent.interject:  # if 'pre check' is selected
            if len(self.fyi_array):  # if there is anything in the fyi array - increment the add report by 1
                self.parent.fyi_rpt += 1
            master_array = self.error_array + self.attn_array  # use these reports for precheck
            if self.parent.fullreport:  # if the full report option is selected...
                master_array += self.fyi_array   # include the fyi messages.
        else:  # if 'input into database' is selected
            if len(self.add_array):  # if there is anything in the add array - increment the add report by 1
                self.parent.add_rpt += 1
            master_array = self.error_array + self.attn_array  # use these reports for input
            if self.parent.fullreport:  # if the full report option is selected...
                master_array += self.add_array  # include the adds messages.
        if len(master_array) > 0:
            if not self.parent.name_mentioned:
                self.parent.report.write("\nGrievance Number: {}\n".format(self.grv_no))
                self.parent.name_mentioned = True
            self.parent.report.write("   >>> sheet: \"{}\" --> row: \"{}\"  <<<\n".format(self.sheet, self.row))
            if not self.parent.allowaddrecs:
                self.parent.report.write("     GRIEVANCE RECORD ENTRY PROHIBITED: Correct errors!\n")
            for rpt in master_array:  # write all reports that have been keep in arrays.
                self.parent.report.write(rpt)


class SpeedSetCheck:
    """ checks one line of the settlement speedsheet when it is called by the SpeedSheetCheck class. """
    def __init__(self, parent, sheet, row, grv_no, level, datesigned, decision, proofdue, docs, gatsnumber, action):
        self.parent = parent
        self.sheet = sheet  # input here is coming directly from the speedcell
        self.row = str(row)
        self.grv_no = grv_no
        self.level = level
        self.datesigned = datesigned
        self.decision = decision
        self.proofdue = proofdue
        self.input_date = []  # array to hold datesigned and proofdue - form in check_dates()
        self.docs = docs
        self.gatsnumber = gatsnumber
        self.action = action
        self.onrec = False  # this value is True if a sql search shows that there is a rec of the grv_no in the db.
        self.onrec_grv_no = None
        self.onrec_level = None
        self.onrec_datesigned = None
        self.onrec_decision = None
        self.onrec_proofdue = None
        self.onrec_docs = None
        self.onrec_gatsnumber = None
        self.addlevel = "empty"  # post checked values
        self.adddatesigned = "empty"
        self.adddecision = "empty"
        self.addproofdue = "empty"
        self.adddate = [self.adddatesigned, self.addproofdue]  # holds date values for self.add_date() loop
        self.adddocs = "empty"
        self.addgatsnumber = "empty"
        self.error_array = []  # gives a report of failed checks
        self.attn_array = []  # gives a report of issues to bring to the attention of users
        self.add_array = []  # gives a report of records to add to the database
        self.fyi_array = []  # gives a report of useful information for the user
        self.parent.name_mentioned = False  # reset this so that name is not repeated on reports
        self.parent.allowaddrecs = True  # if False, records will not be added to database
        self.levelarray = ("informal a", "formal a", "step b", "pre arb", "arbitration")
        self.docsarray = ("non-applicable", "no", "yes", "unknown", "yes-not paid", "yes-in part",
                          "yes-verified", "no-moot", "no-ignore")

    def check_all(self):
        """ master method to run other methods. """
        self.get_onrecs()  # check to see if a record exist, if so self.onrec == True
        if not self.check_delete():  # if the action is not delete - return False
            if self.check_grv_number():  # check the grievance number input
                self.check_level()
                self.check_dates()
                self.check_decision()
                self.check_docs()
                self.check_gatsnumber()
                self.add_recs()
        self.generate_report()

    def get_onrecs(self):
        """ check if there is an existing record for the grievance number in the informalc grievances table.
        if so, store the values in the self.onrec variables. if not, the self.onrec variables default to empty. """
        sql = "SELECT * FROM informalc_settlements WHERE grv_no = '%s'" % self.grv_no
        results = inquire(sql)
        if results:
            self.onrec = True  # this value is True if a sql search shows that there is a rec in the db.
            # skip grievance number as that is self.grv_no and is part of the search criteria
            self.onrec_level = results[0][1]
            self.onrec_datesigned = results[0][2]
            self.onrec_decision = results[0][3]
            self.onrec_proofdue = results[0][4]
            self.onrec_docs = results[0][5]
            self.onrec_gatsnumber = results[0][6]

    def check_delete(self):
        """ check the input for the third column and delete if that action is indicated. """
        delete_array = ("delete", "d/", "erase", "cut")
        if self.action in delete_array:
            self.delete_rec()
            return True
        if self.grv_no in self.parent.del_settlement:
            self.delete_rec()
            return True
        return False

    def delete_rec(self):
        """ check the input for the third column and delete if that action is indicated. """
        if not self.parent.interject:  # if 'pre check' is selected
            if self.onrec:  # if there is a record in the database
                fyi = "     FYI: DELETE Settlement grievance number: {}\n"\
                    .format(self.grv_no)
                self.fyi_array.append(fyi)
            else:  # if there is no record of the grievance in the settlement table
                error = "     FYI: CAN NOT DELETE Settlement grievance number does not exist in database: {}\n" \
                    .format(self.grv_no)
                self.error_array.append(error)

        else:  # if 'input to database' is selected
            # execute the delete in the database
            if self.onrec:
                sql = "DELETE FROM informalc_settlements WHERE grv_no='%s'" % self.grv_no
                commit(sql)
                # create a message for the report
                add = "     DELETE: Settlement record deleted from database >> {}\n" .format(self.grv_no)  # report
                self.add_array.append(add)
            else:
                error = "     CAN NOT DELETE: Settlement record does not exist in database >> {}\n".format(self.grv_no)
                self.error_array.append(error)
        #  store the grv number so this can be deleted from batch settlement index in SpeedIndexCheck()
        if self.grv_no not in self.parent.del_batch:
            self.parent.del_batch.append(self.grv_no)
        #  store the grv number so this can be deleted from batch gats index in SpeedIndexCheck()
        if self.grv_no not in self.parent.del_gatsbatch:
            self.parent.del_gatsbatch.append(self.grv_no)

    def check_grv_number(self):
        """ check the grievance number input """
        if not GrievanceChecker(self.grv_no).has_value():
            error = "     ERROR: The grievance number must not be blank. \n"
            self.error_array.append(error)
            self.parent.allowaddrecs = False  # do not allow this speedcell be be input into database
            return False
        # check that there is a record of the grievance in informalc_grievances
        sql = "SELECT * FROM informalc_grievances WHERE grv_no = '%s' and station = '%s'" \
              % (self.grv_no, self.parent.station)
        result = inquire(sql)
        if not result:
            error = "     ERROR: There is no record of the grievance. \n"
            self.error_array.append(error)
            self.parent.allowaddrecs = False  # do not allow this speedcell be be input into database
            return False
        if not GrievanceChecker(self.grv_no).check_characters():
            error = "     ERROR: The grievance number can only contain numbers and letters. \n"
            self.error_array.append(error)
            self.parent.allowaddrecs = False  # do not allow this speedcell be be input into database
            return False
        if not GrievanceChecker(self.grv_no).min_lenght():
            error = "     ERROR: The grievance number must contain at least 4 characters. \n"
            self.error_array.append(error)
            self.parent.allowaddrecs = False  # do not allow this speedcell be be input into database
            return False
        if not GrievanceChecker(self.grv_no).max_lenght():
            error = "     ERROR: The grievance number can not contain more than 20 characters. \n"
            self.error_array.append(error)
            self.parent.allowaddrecs = False  # do not allow this speedcell be be input into database
            return False
        return True

    def check_level(self):
        """ check the grievance number input """
        self.level = self.level.strip()
        self.level = self.level.lower()
        if not self.level:  # accept blank entries
            pass
        elif self.level not in self.levelarray:
            error = "     ERROR: The level must be either 'informal a', 'formal a', 'step b', 'pre arb' or \n" \
                    "            'arbitration'. No other values are allowed. \n"
            self.error_array.append(error)
            self.parent.allowaddrecs = False  # do not allow this speedcell be be input into database
            return
        self.add_level()

    def add_level(self):
        """ add level to the self.addlevel var """
        if self.level == self.onrec_level:
            pass
        else:
            fyi = "     FYI: New or updated level: {}\n".format(self.level)
            self.fyi_array.append(fyi)
            self.addlevel = self.level

    def check_dates(self):
        """ check the startdate, enddate and meetingdate.
         since these are all dates with similiar criteria, use a loop to check them.
         sometimes, openpyxl sends the dates as strings of datetime objects, instead of the mm/dd/yyyy formated dates,
         the DateTimeChecker() will identify these and skip the checks. """
        self.input_date = [self.datesigned, self.proofdue]
        for i in range(2):
            self.check_date_loop(i)

    def check_date_loop(self, i):
        """ loop from check dates """
        _type = ("date signed", "proof due")
        if self.input_date[i].strip() == "":  # if the value is blank, skip all the checks
            self.add_date(i)
            return
        # if the value is a valid dt object, skip all the checks
        if DateTimeChecker().check_dtstring(self.input_date[i]):
            self.add_date(i)
            return
        date_object = BackSlashDateChecker(self.input_date[i])  # first create the date_object
        if not date_object.count_backslashes():  # this checks that there are 2 backslashes in the date
            error = "     ERROR: The date for the {} date must have two backslashes. Got instead: {}\n"\
                .format(_type[i], self.input_date[i])
            self.error_array.append(error)
            self.parent.allowaddrecs = False  # do not allow this speedcell be be input into database
            return
        date_object.breaker()  # this breaks the object into month, day and year elements.
        if not date_object.check_numeric():  # check each element in the date to ensure they are numeric
            error = "     ERROR: The month, day and year for the {} date must be numeric\n".format(_type[i])
            self.error_array.append(error)
            self.parent.allowaddrecs = False  # do not allow this speedcell be be input into database
            return
        if not date_object.check_minimums():  # check each element in the date to ensure they are greater than zero
            error = "     ERROR: The month, day and year for the {} date must be greater than zero.\n"\
                .format(_type[i])
            self.error_array.append(error)
            self.parent.allowaddrecs = False  # do not allow this speedcell be be input into database
            return
        if not date_object.check_month():  # returns False if the month is greater than 12.
            error = "     ERROR: The month for the {} date must less than 13.\n".format(_type[i])
            self.error_array.append(error)
            self.parent.allowaddrecs = False  # do not allow this speedcell be be input into database
            return
        if not date_object.check_day():  # return False if the day is greater than 31.
            error = "     ERROR: The day entered for the {} date is must be less than 32.\n".format(_type[i])
            self.error_array.append(error)
            self.parent.allowaddrecs = False  # do not allow this speedcell be be input into database
            return
        if not date_object.check_year():  # returns False if the year does not have 4 digits.
            error = "     ERROR: The year entered for the {} date must have 4 digits.\n".format(_type[i])
            self.error_array.append(error)
            self.parent.allowaddrecs = False  # do not allow this speedcell be be input into database
            return
        if not date_object.valid_date():  # returns False if the date is not a valid date
            error = "     ERROR: The date entered for the {} date is not a valid date.\n".format(_type[i])
            self.error_array.append(error)
            self.parent.allowaddrecs = False  # do not allow this speedcell be be input into database
            return
        # this removes white space from the date and each element of the date.
        self.input_date[i] = self.reformat_date(i)
        # convert the input date into a string of a datetime object.
        self.input_date[i] = Convert(self.input_date[i]).backslashdate_to_dtstring()
        self.add_date(i)  # add the dates to add_date variables

    def reformat_date(self, i):
        """ this removes white space from the date and each element of the date. """
        breakdown = self.input_date[i].strip()
        breakdown = breakdown.split("/")
        month = breakdown[0].strip()
        day = breakdown[1].strip()
        year = breakdown[2].strip()
        return "{}/{}/{}".format(month, day, year)

    def add_date(self, i):
        """ add the dates to add_date variables
         this is self.addstartdate, self.addenddate and self.addmeetingdate
         a counter is passed from the self.check_date method above. """
        onrec_date = [self.onrec_datesigned, self.onrec_proofdue]
        _type = ("date signed", "proof due")
        if self.input_date[i] == onrec_date[i]:  # if the new input and the old record are the same - do nothing
            pass  # retain "empty" value for grievant variable
        else:
            fyi = "     FYI: New or updated {} date: {}\n".format(_type[i], self.input_date[i])
            self.fyi_array.append(fyi)
            self.adddate[i] = self.input_date[i]  # save to input to dbase

    def check_decision(self):
        """ check the decision input """
        self.decision = self.decision.strip()  # strip out any whitespace before or after the string
        if self.decision == "":  # accept blank entries
            msg = ""
            self.add_decision(msg)
        elif isint(self.decision):  # identify decision index entries and execute as valid - this also updates article
            self.check_decision_index()
            return
        self.check_decision_description()

    def check_decision_index(self):
        """ check that the decision index provided by the user is valid.
        use arrays of decision categories and articles collected in the SpeedSheetCheck class"""
        if self.decision in self.parent.decision_index:
            self.adddecision = self.parent.decision_description[int(self.decision)-1]
            fyi = "     FYI: New or updated decision (decision index entry): {}\n"\
                .format(self.adddecision)
            self.fyi_array.append(fyi)
            return
        error = "     ERROR: The number for decision is in the index of decisions. Got: {}\n".format(self.decision)
        self.error_array.append(error)
        self.parent.allowaddrecs = False  # do not allow this speedcell be be input into database

    def check_decision_description(self):
        """ check if the decision description is already in the list of decisions. If so, update article. """
        if self.decision in self.parent.decision_description:
            fyi = "     FYI: New or updated decision and article (decision description entry):\n"\
                .format(self.adddecision)
            self.add_decision(fyi)
            return
        fyi = "     FYI: New or updated decision: {}\n".format(self.adddecision)
        self.add_decision(fyi)

    def add_decision(self, msg):
        """ add the decision to the add decision var """
        if self.decision == self.onrec_decision:
            pass
        else:
            self.adddecision = self.decision
            if msg:
                self.fyi_array.append(msg)

    def check_docs(self):
        """ check the grievance number input """
        self.docs = self.docs.strip()
        self.docs = self.docs.lower()
        if not self.docs:
            pass
        elif self.docs in self.docsarray:
            pass
        else:
            error = "     ERROR: The docs input must be either 'non-applicable', 'no', 'yes', 'unknown', \n" \
                    "            'yes - not paid', 'yes - in part', 'yes - verified', 'no - moot' or \n" \
                    "            'no - ignore'. No other values are allowed. Got: {}\n".format(self.docs)
            self.error_array.append(error)
            self.parent.allowaddrecs = False  # do not allow this speedcell be be input into database
            return False
        self.add_docs()

    def add_docs(self):
        """ add docs to the self.adddocs var """
        if self.docs == self.onrec_docs:
            pass
        else:
            fyi = "     FYI: New or updated docs: {}\n".format(self.docs)
            self.fyi_array.append(fyi)
            self.adddocs = self.docs

    def check_gatsnumber(self):
        """ check the article input - this is an open field that takes almost anything with no limits or indexes. """
        self.gatsnumber = self.gatsnumber.strip()
        self.gatsnumber = self.gatsnumber.lower()
        if not self.gatsnumber:
            pass
        if len(self.gatsnumber) < 30:
            pass
        else:
            error = "     ERROR: The gats number must not be longer than 30 characters.  \n"
            self.error_array.append(error)
            self.parent.allowaddrecs = False  # do not allow this speedcell be be input into database
            return False
        self.add_gatsnumber()

    def add_gatsnumber(self):
        """ add gats number to the self.addgatsnumber var """
        if self.gatsnumber == self.onrec_gatsnumber:
            pass
        else:
            fyi = "     FYI: New or updated gats number: {}\n".format(self.gatsnumber)
            self.fyi_array.append(fyi)
            self.addgatsnumber = self.gatsnumber

    def add_recs(self):
        """ add records using the add___ vars. """
        chg_these = []
        if not self.parent.allowaddrecs:  # if all checks passed
            return  # do not update or insert into the database
        if not self.parent.interject:  # if 'precheck' is selected
            return  # do not update or insert into the database
        if not self.onrec:  # if there is no record of the grievance number in the db informalc_grievance table
            add = "     INPUT: New Grievance Number added to database >>{}\n" \
                .format(self.grv_no)  # report
            self.add_array.append(add)
            chg_these.append('grv_no')
        # get level place
        if self.addlevel != "empty":
            add = "     INPUT: Level added or updated to database >>{}\n" \
                .format(self.addlevel)  # report
            self.add_array.append(add)
            chg_these.append("level")
            level_place = self.addlevel
        else:
            level_place = self.onrec_level
        # get date places using loop
        onrec_date = [self.onrec_datesigned, self.onrec_proofdue]
        datesigned_place = None  # aka date_place[0]
        proofdue_place = None  # aka date_place[1]
        date_place = [datesigned_place, proofdue_place]
        chg_notation = ("datesigned", "proofdue")
        _type = ("Date Signed", "Proof Due Date")
        for i in range(2):
            if self.adddate[i] != "empty":
                add = "     INPUT: {} added or updated to database >>{}\n".format(_type[i], self.adddate[i])
                self.add_array.append(add)
                chg_these.append(chg_notation[i])
                date_place[i] = self.adddate[i]
            else:
                date_place[i] = onrec_date[i]
        # get decision place
        if self.adddecision != "empty":
            add = "     INPUT: Decision added or updated to database >>{}\n".format(self.adddecision)  # report
            self.add_array.append(add)
            chg_these.append("decision")
            decision_place = self.adddecision
        else:
            decision_place = self.onrec_decision
        # get docs place
        if self.adddocs != "empty":
            add = "     INPUT: Docs added or updated to database >>{}\n".format(self.adddocs)  # report
            self.add_array.append(add)
            chg_these.append("docs")
            docs_place = self.adddocs
        else:
            docs_place = self.onrec_docs
        # get gats place
        if self.addgatsnumber != "empty":
            add = "     INPUT: Gats Number added or updated to database >>{}\n".format(self.addgatsnumber)  # report
            self.add_array.append(add)
            chg_these.append("gatsnumber")
            gats_place = self.addgatsnumber
        else:
            gats_place = self.onrec_gatsnumber
        # if any values have changed - form sql statements using _place vars and commit to db.
        if len(chg_these) != 0:  # if change these is empty, then there is no need to insert/update records
            if not self.onrec:  # if there is no rec on file for the grievance, insert the first rec
                sql = "INSERT INTO informalc_settlements(grv_no, level, date_signed, decision, proofdue, " \
                      "docs, gats_number) VALUES('%s','%s','%s','%s','%s','%s','%s')" \
                      % (self.grv_no, level_place, date_place[0], decision_place, date_place[1], docs_place, gats_place)
            else:  # update the first rec to replace pre existing record.
                sql = "UPDATE informalc_settlements SET level='%s', date_signed='%s', decision ='%s', " \
                      "proofdue='%s', docs='%s', gats_number='%s' WHERE grv_no='%s'" \
                      % (level_place, date_place[0], decision_place, date_place[1], docs_place, gats_place,
                         self.grv_no)
            commit(sql)

    def generate_report(self):
        """ generate a report
        """
        self.parent.settlement_fatal_rpt += len(self.error_array)
        if len(self.add_array):  # if there is anything in the add array - increment the add report by 1
            self.parent.settlement_add_rpt += 1
        if len(self.fyi_array):  # if there is anything in the fyi array - increment the add report by 1
            self.parent.settlement_fyi_rpt += 1
        if not self.parent.interject:
            master_array = self.error_array + self.attn_array  # use these reports for precheck
            if self.parent.fullreport:  # if the full report option is selected...
                master_array += self.fyi_array   # include the fyi messages.
        else:
            master_array = self.error_array + self.attn_array  # use these reports for input
            if self.parent.fullreport:  # if the full report option is selected...
                master_array += self.add_array  # include the adds messages.
        if len(master_array) > 0:
            if not self.parent.name_mentioned:
                self.parent.report.write("\nGrievance Number: {}\n".format(self.grv_no))
                self.parent.name_mentioned = True
            self.parent.report.write("   >>> sheet: \"{}\" --> row: \"{}\"  <<<\n".format(self.sheet, self.row))
            if not self.parent.allowaddrecs:
                self.parent.report.write("     SETTLEMENT RECORD ENTRY PROHIBITED: Correct errors!\n")
            for rpt in master_array:  # write all reports that have been keep in arrays.
                self.parent.report.write(rpt)


class SpeedIndexCheck:
    """ checks one line of the index speedsheet when it is called by the SpeedSheetCheck() class.
    SpeedSheetCheck() will call scan each speedsheet, line by line.
    """
    def __init__(self, parent, i, sheet, row, first, second, action):
        self.parent = parent
        self.i = i - 2  # subtract two for the grievance and settlement worksheets
        self.sheet = sheet  # input here is coming directly from the speedcell
        self.row = str(row)
        # convert inputted grievance numbers to lower case and strip whitespace from front and back.
        self.first = first.lower().strip()
        self.second = second.lower().strip()
        self.action = action.lower().strip()
        self.grv_array = [self.first, self.second]  # combine both first and second into an array.
        self.tables = ("informalc_noncindex", "informalc_remandindex", "informalc_batchindex", "informalc_gatsindex")
        self.error_array = []  # gives a report of failed checks
        self.attn_array = []  # gives a report of issues to bring to the attention of users
        self.add_array = []  # gives a report of records to add to the database
        self.fyi_array = []  # gives a report of useful information for the user
        self.parent.name_mentioned = False  # reset this so that name is not repeated on reports
        self.parent.allowaddrecs = True  # if False, records will not be added to database
        self.del_array = [self.parent.del_nonc, self.parent.del_remanded, self.parent.del_batch,
                          self.parent.del_gatsbatch]
        self.onrec = False  # is True if the record already exist in the database

    def check_all(self):
        """ master method to run other methods. """
        self.check_onrecs()
        if not self.check_delete():  # if the action is not delete - return False
            self.check_firstandsecond()
            self.check_exist()
            self.check_same()
            self.add_recs()
        self.generate_report()

    def check_onrecs(self):
        """ check to see if the record is already on record - in the db. """
        sql = "SELECT * FROM %s WHERE %s='%s' and %s='%s'" \
              % (self.tables[self.i], self.parent.index_columns[self.i][0], self.grv_array[0],
                 self.parent.index_columns[self.i][1], self.grv_array[1])
        result = inquire(sql)
        if result:
            self.onrec = True

    def check_delete(self):
        """ check the input for the third column and delete if that action is indicated.
        also check to see if the grievance number is in the delete array. """
        delete_array = ("delete", "del", "d/", "erase", "cut", "remove")
        if self.action in delete_array:
            self.delete_recs()
            return True
        if self.first in self.del_array[self.i] or self.second in self.del_array[self.i]:
            self.delete_recs()
            return True
        return False

    def delete_recs(self):
        """ delete records from indexes or make an fyi report stating that recs are to be deleted. """
        index = ("Non compliance", "Remanded", "Batch Settlement", "Batch Gats")
        if not self.parent.interject:
            if self.onrec:
                fyi = "     FYI: DELETE {} index entry: {} - {}\n"\
                    .format(index[self.i], self.grv_array[0], self.grv_array[1])
                self.fyi_array.append(fyi)
            else:
                error = "     FYI: CAN NOT DELETE {} index entry does not exist in database: {} - {}\n" \
                    .format(index[self.i], self.grv_array[0], self.grv_array[1])
                self.error_array.append(error)
        else:
            # set default message to a failure to add
            if self.onrec:  # if there is a record in the database
                # execute the delete in the database
                sql = "DELETE FROM %s WHERE %s='%s' and %s='%s'" \
                      % (self.tables[self.i], self.parent.index_columns[self.i][0], self.grv_array[0],
                         self.parent.index_columns[self.i][1], self.grv_array[1])
                commit(sql)
                # create a message for the report
                add = "     DELETE: {} record deleted from database >>{} - {}\n" \
                    .format(index[self.i], self.grv_array[0], self.grv_array[1])  # report
                self.add_array.append(add)
            else:
                error = "     CAN NOT DELETE: {} record does not exist in database >>{} - {}\n" \
                    .format(index[self.i], self.grv_array[0], self.grv_array[1])  # report
                self.error_array.append(error)

    def check_firstandsecond(self):
        """ check the grievant input """
        for ii in range(2):
            if not GrievanceChecker(self.grv_array[ii]).has_value():
                error = "     ERROR: The grievance number for {} must not be blank.\n"\
                    .format(self.parent.index_columns[self.i][ii])
                self.error_array.append(error)
                self.parent.allowaddrecs = False  # do not allow this speedcell be be input into database
            if not GrievanceChecker(self.grv_array[ii]).check_characters():
                error = "     ERROR: The grievance number for {} can not contain special characters\n"\
                    .format(self.parent.index_columns[self.i][ii])
                self.error_array.append(error)
                self.parent.allowaddrecs = False  # do not allow this speedcell be be input into database
            if not GrievanceChecker(self.grv_array[ii]).min_lenght():
                error = "     ERROR: The grievance number for {} can not be shorter than 4 characters\n"\
                    .format(self.parent.index_columns[self.i][ii])
                self.error_array.append(error)
                self.parent.allowaddrecs = False  # do not allow this speedcell be be input into database
            if not GrievanceChecker(self.grv_array[ii]).max_lenght():
                error = "     ERROR: The grievance number for {} can not be longer than 20 characters\n"\
                    .format(self.parent.index_columns[self.i][ii])
                self.error_array.append(error)
                self.parent.allowaddrecs = False  # do not allow this speedcell be be input into database

    def check_same(self):
        """ check the grievance number input is the same for both columns. """
        if self.grv_array[0] == self.grv_array[1]:
            error = "     ERROR: The grievance numbers for {} and {} can not be the same\n" \
                .format(self.parent.index_columns[self.i][0], self.parent.index_columns[self.i][1])
            self.error_array.append(error)
            self.parent.allowaddrecs = False  # do not allow this speedcell be be input into database

    def check_exist(self):
        """ check to see if the first and second grievance number is grievance that exist in the database """
        distinct = []
        sql = "SELECT DISTINCT grv_no FROM informalc_grievances"
        result = inquire(sql)
        for r in result:
            distinct.append(r[0])
        for i in range(2):
            if self.grv_array[i] not in distinct:
                error = "     ERROR: Grievance number {} entered for {} does not exist in the database. \n" \
                    .format(self.grv_array[i], self.parent.index_columns[self.i][i])
                self.error_array.append(error)
                self.parent.allowaddrecs = False  # do not allow this speedcell be be input into database

    def add_recs(self):
        """ add records to the database. """
        index = ("Non compliance", "Remanded", "Batch Settlement", "Batch Gats")
        if not self.parent.allowaddrecs:  # do not add records if there is a fatal error
            return
        if self.onrec:  # if there is already a record in the database, do not add another one
            return
        if not self.parent.interject:  # if user selected 'pre check' - only generate message
            fyi = "     FYI: New {} index entry: {} - {}\n".format(index[self.i], self.grv_array[0], self.grv_array[1])
            self.fyi_array.append(fyi)
            return
        sql = "INSERT INTO %s(%s, %s) VALUES('%s', '%s')" \
              % (self.tables[self.i], self.parent.index_columns[self.i][0], self.parent.index_columns[self.i][1],
                 self.grv_array[0], self.grv_array[1])
        commit(sql)
        add = "     INPUT: {} record added to database >>{} - {}\n" \
            .format(index[self.i], self.grv_array[0], self.grv_array[1])  # report
        self.add_array.append(add)

    def generate_report(self):
        """"
        generate the text report
        """
        self.parent.index_fatal_rpt[self.i] += len(self.error_array)
        if len(self.add_array):  # if there is anything in the add array - increment the add report by 1
            self.parent.index_add_rpt[self.i] += 1
        if len(self.fyi_array):  # if there is anything in the fyi array - increment the add report by 1
            self.parent.index_fyi_rpt[self.i] += 1
        if not self.parent.interject:
            master_array = self.error_array + self.attn_array  # use these reports for precheck
            if self.parent.fullreport:  # if the full report option is selected...
                master_array += self.fyi_array  # include the fyi messages.
        else:
            master_array = self.error_array + self.attn_array  # use these reports for input
            if self.parent.fullreport:  # if the full report option is selected...
                master_array += self.add_array  # include the adds messages.
        if len(master_array) > 0:
            if not self.parent.name_mentioned:
                self.parent.report.write("\nGrievance Number: {}\n".format(self.grv_array[0]))
                self.parent.name_mentioned = True
            self.parent.report.write("   >>> sheet: \"{}\" --> row: \"{}\"  <<<\n".format(self.sheet, self.row))
            if not self.parent.allowaddrecs:
                self.parent.report.write("     INDEX RECORD ENTRY PROHIBITED: Correct errors!\n")
            for rpt in master_array:  # write all reports that have been keep in arrays.
                self.parent.report.write(rpt)


class ProgressBarIn:
    """ Indeterminate Progress Bar """

    def __init__(self, title="", label="", text=""):
        self.title = title
        self.label = label
        self.text = text
        self.pb_root = Tk()  # create a window for the progress bar
        self.pb_label = Label(self.pb_root, text=self.label)  # make label for progress bar
        self.pb = ttk.Progressbar(self.pb_root, length=400, mode="indeterminate")  # create progress bar
        self.pb_text = Label(self.pb_root, text=self.text, anchor="w")

    def start_up(self):
        """ starts up the progress bar. """
        titlebar_icon(self.pb_root)  # place icon in titlebar
        self.pb_root.title(self.title)
        self.pb_label.grid(row=0, column=0, sticky="w")
        self.pb.grid(row=1, column=0, sticky="w")
        self.pb_text.grid(row=2, column=0, sticky="w")
        while pb_flag:  # use global as a flag. stop loop when flag is False
            projvar.root.update()
            self.pb['value'] += 1
            time.sleep(.01)

    def stop(self):
        """ stops and destroys the progress bar. """
        self.pb.stop()  # stop and destroy the progress bar
        self.pb_text.destroy()
        self.pb_label.destroy()  # destroy the label for the progress bar
        self.pb.destroy()
        self.pb_root.destroy()


if __name__ == "__main__":
    """ this is where the program starts if not launched from another app. """
    # InformalC().informalc(None)
