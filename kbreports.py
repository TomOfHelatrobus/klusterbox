"""
a klusterbox module: Klusterbox Reports Generator
this module generates text files which provide the user with information, such as the routes of all carriers in the
investigation range, or their ns days, etc. The Messenger class gives the location of the program and also provides
the user information in the form of message boxes.
"""
import projvar
from kbtoolbox import inquire, CarrierList, dt_converter, NsDayDict, dir_path, Convert, check_path, \
    informalc_date_checker, DateTimeChecker, ProgressBarDe, issuedecisionresult_sorter, distinctresult_to_list
from tkinter import messagebox, simpledialog, filedialog
from tkinter.simpledialog import askstring
from shutil import rmtree
import os
import sys
import subprocess
from datetime import datetime, timedelta
from operator import itemgetter
# Spreadsheet Libraries
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment


class Reports:
    """
    generates reports
    """

    def __init__(self, frame):
        self.frame = frame
        self.start_date = projvar.invran_date
        self.end_date = projvar.invran_date
        if projvar.invran_weekly_span:
            self.start_date = projvar.invran_date_week[0]
            self.end_date = projvar.invran_date_week[6]
        self.carrier_list = []
        self.seniority_list = []
        self.positivedate = []
        self.negativedate = []

    def get_carrierlist(self):
        """ gets the carrier list for the investigation range. """
        # get carrier list
        self.carrier_list = CarrierList(self.start_date, self.end_date, projvar.invran_station).get()

    @staticmethod
    def rpt_dt_limiter(date, first_date):
        """ return the first day if it is earlier than the date """
        if date < first_date:
            return first_date
        else:
            return date

    @staticmethod
    def rpt_ns_fixer(nsday_code):
        """ remove the day from the ns_code if fixed """
        if "fixed" in nsday_code:
            fix = nsday_code.split(":")
            return fix[0]
        else:
            return nsday_code

    def rpt_carrier(self):
        """ Generate and display a report of carrier routes and nsday """
        self.get_carrierlist()
        ns_dict = NsDayDict.get_custom_nsday()  # get the ns day names from the dbase
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")  # create a file name
        filename = "report_carrier_route_nsday" + "_" + stamp + ".txt"
        report = open(dir_path('report') + filename, "w")
        report.write("\nCarrier Route and NS Day Report\n\n\n")
        report.write('   Showing results for:\n')
        report.write('      Station: {}\n'.format(projvar.invran_station))
        if not projvar.invran_weekly_span:  # if investigation range is daily
            f_date = projvar.invran_date
            report.write('      Date: {}\n'.format(f_date.strftime("%m/%d/%Y")))
        else:  # if investigation range is weekly
            f_date = projvar.invran_date_week[0]  # use the first day of the service week
            report.write('      Dates: {} through {}\n'
                         .format(projvar.invran_date_week[0].strftime("%m/%d/%Y"),
                                 projvar.invran_date_week[6].strftime("%m/%d/%Y")))
        report.write('      Pay Period: {}\n\n'.format(projvar.pay_period))
        report.write('{:>4} {:<23} {:<13} {:<29} {:<10}\n'.format("", "Carrier Name", "N/S Day", "Route/s",
                                                                  "Start Date"))
        report.write('     ------------------------------------------------------------------- ----------\n')
        i = 1
        for line in self.carrier_list:
            ii = 0
            for rec in reversed(line):
                if not ii:
                    report.write('{:>4} {:<23} {:<4} {:<8} {:<29}\n'
                                 .format(i, rec[1], projvar.ns_code[rec[3]], self.rpt_ns_fixer(ns_dict[rec[3]]),
                                         rec[4]))
                else:
                    report.write('{:>4} {:<23} {:<4} {:<8} {:<29} {:<10}\n'
                                 .format("", rec[1], projvar.ns_code[rec[3]], self.rpt_ns_fixer(ns_dict[rec[3]]),
                                         rec[4], self.rpt_dt_limiter(dt_converter(rec[0]), f_date).strftime("%A")))
                ii += 1
            if i % 3 == 0:
                report.write('     ------------------------------------------------------------------- ----------\n')
            i += 1
        report.close()
        if sys.platform == "win32":  # open the text document
            os.startfile(dir_path('report') + filename)
        if sys.platform == "linux":
            subprocess.call(["xdg-open", 'kb_sub/report/' + filename])
        if sys.platform == "darwin":
            subprocess.call(["open", dir_path('report') + filename])

    def rpt_carrier_route(self):
        """ Generate and display a report of carrier routes """
        self.get_carrierlist()
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = "report_carrier_route" + "_" + stamp + ".txt"
        report = open(dir_path('report') + filename, "w")
        report.write("\nCarrier Route Report\n\n\n")
        report.write('   Showing results for:\n')
        report.write('      Station: {}\n'.format(projvar.invran_station))
        if not projvar.invran_weekly_span:  # if investigation range is daily
            f_date = projvar.invran_date
            report.write('      Date: {}\n'.format(f_date.strftime("%m/%d/%Y")))
        else:
            f_date = projvar.invran_date_week[0]
            report.write('      Date: {} through {}\n'
                         .format(projvar.invran_date_week[0].strftime("%m/%d/%Y"),
                                 projvar.invran_date_week[6].strftime("%m/%d/%Y")))
        report.write('      Pay Period: {}\n\n'.format(projvar.pay_period))
        report.write('{:>4}  {:<22} {:<29}\n'.format("", "Carrier Name", "Route/s"))
        report.write('      ---------------------------------------------------- -------------------\n')
        i = 1
        for line in self.carrier_list:
            ii = 0
            for rec in reversed(line):  # reverse order so earliest one appears first
                if not ii:  # if the first record
                    report.write('{:>4}  {:<22} {:<29}\n'.format(i, rec[1], rec[4]))
                else:  # if not the first record, use alternate format
                    report.write('{:>4}  {:<22} {:<29} effective {:<10}\n'
                                 .format("", rec[1], rec[4],
                                         self.rpt_dt_limiter(dt_converter(rec[0]), f_date).strftime("%A")))
                ii += 1
            if i % 3 == 0:
                report.write('      ---------------------------------------------------- -------------------\n')
            i += 1
        report.close()
        if sys.platform == "win32":  # open the text document
            os.startfile(dir_path('report') + filename)
        if sys.platform == "linux":
            subprocess.call(["xdg-open", 'kb_sub/report/' + filename])
        if sys.platform == "darwin":
            subprocess.call(["open", dir_path('report') + filename])

    def rpt_carrier_nsday(self):
        """ Generate and display a report of carrier ns day """
        self.get_carrierlist()
        ns_dict = NsDayDict.get_custom_nsday()  # get the ns day names from the dbase
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = "report_carrier_nsday" + "_" + stamp + ".txt"
        report = open(dir_path('report') + filename, "w")
        report.write("\nCarrier NS Day\n\n\n")
        report.write('   Showing results for:\n')
        report.write('      Station: {}\n'.format(projvar.invran_station))
        if not projvar.invran_weekly_span:  # if investigation range is daily
            f_date = projvar.invran_date
            report.write('      Date: {}\n'.format(f_date.strftime("%m/%d/%Y")))
        else:
            f_date = projvar.invran_date_week[0]
            report.write('      Date: {} through {}\n'
                         .format(projvar.invran_date_week[0].strftime("%m/%d/%Y"),
                                 projvar.invran_date_week[6].strftime("%m/%d/%Y")))
        report.write('      Pay Period: {}\n\n'.format(projvar.pay_period))
        report.write('{:>4}  {:<22} {:<17}\n'.format("", "Carrier Name", "N/S Day"))
        report.write('      ----------------------------------------  -------------------\n')
        i = 1
        for line in self.carrier_list:
            ii = 0
            for rec in reversed(line):
                if not ii:
                    report.write('{:>4}  {:<22} {:<5}{:<12}\n'
                                 .format(i, rec[1], projvar.ns_code[rec[3]], self.rpt_ns_fixer(ns_dict[rec[3]])))
                else:
                    report.write('{:>4}  {:<22} {:<5}{:<12}  effective {:<10}\n'
                                 .format("", rec[1], projvar.ns_code[rec[3]], self.rpt_ns_fixer(ns_dict[rec[3]]),
                                         self.rpt_dt_limiter(dt_converter(rec[0]), f_date).strftime("%A")))
                ii += 1
            if i % 3 == 0:
                report.write('      ----------------------------------------  -------------------\n')
            i += 1
        report.close()
        if sys.platform == "win32":  # open the text document
            os.startfile(dir_path('report') + filename)
        if sys.platform == "linux":
            subprocess.call(["xdg-open", 'kb_sub/report/' + filename])
        if sys.platform == "darwin":
            subprocess.call(["open", dir_path('report') + filename])

    def rpt_carrier_by_list(self):
        """ generates a report which shows carriers by the list. """
        self.get_carrierlist()
        list_dict = {"nl": "No List", "wal": "Work Assignment List",
                     "otdl": "Overtime Desired List", "ptf": "Part Time Flexible", "aux": "Auxiliary Carrier"}
        # initialize arrays for data sorting
        otdl_array = []
        wal_array = []
        nl_array = []
        ptf_array = []
        aux_array = []
        for line in self.carrier_list:
            for carrier in line:
                if carrier[2] == "otdl":
                    otdl_array.append(carrier)
                if carrier[2] == "wal":
                    wal_array.append(carrier)
                if carrier[2] == "nl":
                    nl_array.append(carrier)
                if carrier[2] == "ptf":
                    ptf_array.append(carrier)
                if carrier[2] == "aux":
                    aux_array.append(carrier)
        array_var = nl_array + wal_array + otdl_array + ptf_array + aux_array  #
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")  # create a file name
        filename = "report_carrier_by_list" + "_" + stamp + ".txt"
        report = open(dir_path('report') + filename, "w")
        report.write("\nCarrier by List\n\n")
        report.write('   Showing results for:\n')
        report.write('      Station: {}\n'.format(projvar.invran_station))
        if not projvar.invran_weekly_span:  # if investigation range is daily
            f_date = projvar.invran_date
            report.write('      Date: {}\n'.format(f_date.strftime("%m/%d/%Y")))
        else:
            f_date = projvar.invran_date_week[0]
            report.write('      Dates: {} through {}\n'
                         .format(projvar.invran_date_week[0].strftime("%m/%d/%Y"),
                                 projvar.invran_date_week[6].strftime("%m/%d/%Y")))
        report.write('      Pay Period: {}\n'.format(projvar.pay_period))
        i = 1
        last_list = ""  # this is a indicator for when a new list is starting
        for line in array_var:
            if last_list != line[2]:  # if the new record is in a different list that the last
                report.write('\n\n      {:<20}\n\n'
                             .format(list_dict[line[2]]))  # write new headers
                report.write('{:>4}  {:<22} {:>4}\n'.format("", "Carrier Name", "List"))
                report.write('      ---------------------------  -------------------\n')
                i = 1
            if dt_converter(line[0]) not in projvar.invran_date_week:
                report.write('{:>4}  {:<22} {:>4}\n'.format(i, line[1], line[2]))
            else:
                report.write('{:>4}  {:<22} {:>4}  effective {:<10}\n'
                             .format(i, line[1], line[2],
                                     self.rpt_dt_limiter(dt_converter(line[0]), f_date).strftime("%A")))
            if i % 3 == 0:
                report.write('      ---------------------------  -------------------\n')
            last_list = line[2]
            i += 1
        report.close()
        if sys.platform == "win32":  # open the text document
            os.startfile(dir_path('report') + filename)
        if sys.platform == "linux":
            subprocess.call(["xdg-open", 'kb_sub/report/' + filename])
        if sys.platform == "darwin":
            subprocess.call(["open", dir_path('report') + filename])

    @staticmethod
    def rpt_carrier_history(carrier):
        """ generates a report showing all records from a specified carrier. """
        sql = "SELECT effective_date, list_status, ns_day, route_s, station" \
              " FROM carriers WHERE carrier_name = '%s' ORDER BY effective_date DESC" % carrier
        results = inquire(sql)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = "report_carrier_history" + "_" + stamp + ".txt"
        report = open(dir_path('report') + filename, "w")
        report.write("\nCarrier Status Change History\n\n")
        report.write('   Showing all status changes in the klusterbox database for {}\n\n'.format(carrier))
        report.write('{:<16}{:<8}{:<10}{:<31}{:<25}\n'
                     .format("Date Effective", "List", "N/S Day", "Route/s", "Station"))
        report.write('----------------------------------------------------------------------------------\n')
        i = 1
        for line in results:
            report.write('{:<16}{:<8}{:<10}{:<31}{:<25}\n'
                         .format(dt_converter(line[0]).strftime("%m/%d/%Y"), line[1], line[2], line[3], line[4]))
            if i % 3 == 0:
                report.write('----------------------------------------------------------------------------------\n')
            i += 1
        report.close()
        if sys.platform == "win32":  # open the text document
            os.startfile(dir_path('report') + filename)
        if sys.platform == "linux":
            subprocess.call(["xdg-open", 'kb_sub/report/' + filename])
        if sys.platform == "darwin":
            subprocess.call(["open", dir_path('report') + filename])

    @staticmethod
    def rpt_all_rings(carrier):
        """ this will generate a report showing all rings for a selected carrier """
        sql = "SELECT rings_date, total, code, bt, rs, et, moves, leave_type, leave_time" \
              " FROM rings3 WHERE carrier_name = '%s' ORDER BY rings_date" % carrier
        results = inquire(sql)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = "report_all_rings" + "_" + stamp + ".txt"
        report = open(dir_path('report') + filename, "w")
        report.write("\nCarrier All Rings History\n\n")
        report.write('   Showing all clock rings in the Klusterbox database for {}\n\n'.format(carrier))
        if not results:  # if there are no rings, show a message
            report.write("   No clock rings for {} were found in the Klusterbox database. ".format(carrier))
        else:  # if there are rings, then write the column headers
            report.write('{:<11}|{:>6} {:>8}|{:>6}|{:>6} {:>6} {:>6}|{:>6}|{:>6}|{:>8} {:>6}\n'
                         .format("Date ", "5200", "Code", "BT", "MV off", "MV on", "Route", "RS", "ET", "Leave", ""))
            report.write('-------------------------------------------------------------------------------------\n')
        i = 1
        for line in results:
            date = dt_converter(line[0]).strftime("%m/%d/%Y")
            total = Convert(line[1]).empty_or_hunredths()
            code = Convert(line[2]).empty_not_none()
            bt = Convert(line[3]).empty_or_hunredths()
            rs = Convert(line[4]).empty_or_hunredths()
            et = Convert(line[5]).empty_or_hunredths()
            lv_type = Convert(line[7]).empty_not_none()
            lv_time = Convert(line[8]).empty_or_hunredths()
            moves = []
            mvoff = ""
            mvon = ""
            rte = ""
            if line[6]:  # if there are moves
                moves = Convert(line[6]).string_to_array()  # change string to an array
                mvoff = Convert(moves[0]).zero_or_hundredths()  # format first move off time
                mvon = Convert(moves[1]).zero_or_hundredths()  # format first move on time
                rte = moves[2]
            report.write('{:<11}|{:>6} {:>8}|{:>6}|{:>6} {:>6} {:>6}|{:>6}|{:>6}|{:>8} {:>6}\n'
                         .format(date, total, code, bt, mvoff, mvon, rte, rs, et, lv_type, lv_time))
            if len(moves) > 3:  # if there is more than one move triad, output move triads on new line
                for ii in range(3, len(moves), 3):
                    mvoff = Convert(moves[ii]).zero_or_hundredths()
                    mvon = Convert(moves[ii + 1]).zero_or_hundredths()
                    rte = moves[ii + 2]
                    report.write('{:<11}|{:>6} {:>8}|{:>6}|{:>6} {:>6} {:>6}|{:>6}|{:>6}|{:>8} {:>6}\n'
                                 .format("", "", "", "", mvoff, mvon, rte, "", "", "", ""))
            if i % 3 == 0:
                report.write('-----------------------------------------------------------------------------------'
                             '--\n')
            i += 1
        report.close()
        if sys.platform == "win32":  # open the text document
            os.startfile(dir_path('report') + filename)
        if sys.platform == "linux":
            subprocess.call(["xdg-open", 'kb_sub/report/' + filename])
        if sys.platform == "darwin":
            subprocess.call(["open", dir_path('report') + filename])

    def rpt_carrier_seniority_id(self):
        """ Generate and display a report of carrier routes """
        self.get_empid_seniority_list()
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = "report_carrier_seniority" + "_" + stamp + ".txt"
        report = open(dir_path('report') + filename, "w")
        report.write("\nCarrier Seniority Report\n\n\n")
        report.write('   Showing results for:\n')
        report.write('      Station: {}\n'.format(projvar.invran_station))
        if not projvar.invran_weekly_span:  # if investigation range is daily
            f_date = projvar.invran_date
            report.write('      Date: {}\n'.format(f_date.strftime("%m/%d/%Y")))
        else:
            report.write('      Date: {} through {}\n'
                         .format(projvar.invran_date_week[0].strftime("%m/%d/%Y"),
                                 projvar.invran_date_week[6].strftime("%m/%d/%Y")))
        report.write('      Pay Period: {}\n\n'.format(projvar.pay_period))

        report.write('{:>4}  {:<25} {:<12} {:<14} {:<4}\n'.
                     format("", "Carrier Name", "Employee ID", "Seniority Date", "Rank"))
        report.write('      ------------------------- -----------  -------------- ---- \n')
        i = 1
        for line in self.seniority_list:
            report.write('{:>4}  {:<25} {:<12} {:<14} {:>4}\n'.format(i, line[0], line[3], line[1], line[2]))
            if i % 3 == 0:
                report.write('      ------------------------- -----------  -------------- ----\n')
            i += 1
        report.close()
        if sys.platform == "win32":  # open the text document
            os.startfile(dir_path('report') + filename)
        if sys.platform == "linux":
            subprocess.call(["xdg-open", 'kb_sub/report/' + filename])
        if sys.platform == "darwin":
            subprocess.call(["open", dir_path('report') + filename])

    def rpt_carrier_seniority(self):
        """ Generate and display a report of carrier routes """
        self.get_seniority_list()
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = "report_carrier_seniority" + "_" + stamp + ".txt"
        report = open(dir_path('report') + filename, "w")
        report.write("\nCarrier Seniority Report\n\n\n")
        report.write('   Showing results for:\n')
        report.write('      Station: {}\n'.format(projvar.invran_station))
        if not projvar.invran_weekly_span:  # if investigation range is daily
            f_date = projvar.invran_date
            report.write('      Date: {}\n'.format(f_date.strftime("%m/%d/%Y")))
        else:
            report.write('      Date: {} through {}\n'
                         .format(projvar.invran_date_week[0].strftime("%m/%d/%Y"),
                                 projvar.invran_date_week[6].strftime("%m/%d/%Y")))
        report.write('      Pay Period: {}\n\n'.format(projvar.pay_period))

        report.write('{:>4}  {:<30} {:<10}\n'.format("", "Carrier Name", "Seniority Date"))
        report.write('      ------------------------------ --------------\n')
        i = 1
        for line in self.seniority_list:
            report.write('{:>4}  {:<30} {:<10}\n'.format(i, line[0], line[1]))
            if i % 3 == 0:
                report.write('      ------------------------------ --------------\n')
            i += 1
        report.close()
        if sys.platform == "win32":  # open the text document
            os.startfile(dir_path('report') + filename)
        if sys.platform == "linux":
            subprocess.call(["xdg-open", 'kb_sub/report/' + filename])
        if sys.platform == "darwin":
            subprocess.call(["open", dir_path('report') + filename])

    def get_seniority_list(self):
        """ returns a carrier list of seniority dates ordered by date """
        self.build_pos_neg()  # creates the positvedate and negative date arrays
        i = 0
        for rec in self.positivedate:
            backslashdate = Convert(rec[1]).dtstring_to_backslashdate()
            self.positivedate[i][1] = backslashdate
            i += 1
        i = 0
        for _ in self.negativedate:
            self.negativedate[i][1] = "no record"
            i += 1
        self.seniority_list = self.positivedate + self.negativedate

    def get_empid_seniority_list(self):
        """ returns a list with employee id, seniority rank and date"""
        self.build_pos_neg()  # creates the positvedate and negative date arrays
        i = 0
        rank = 1
        # rec[0] is the carrier name, rec[1] is the seniority date. append seniority rank and employee id
        for rec in self.positivedate:
            backslashdate = Convert(rec[1]).dtstring_to_backslashdate()
            self.positivedate[i][1] = backslashdate
            self.positivedate[i].append(str(rank))
            empid = self.get_empid(rec[0])
            self.positivedate[i].append(empid)
            rank += 1
            i += 1
        i = 0
        for rec in self.negativedate:
            self.negativedate[i][1] = "no record"
            self.negativedate[i].append("?")
            empid = self.get_empid(rec[0])
            self.negativedate[i].append(empid)
            i += 1
        self.seniority_list = self.positivedate + self.negativedate
        self.seniority_list.sort(key=itemgetter(0))  # sort by name

    @staticmethod
    def get_empid(carrier):
        """ return the employee id for a carrier name """
        sql = "SELECT emp_id FROM name_index WHERE kb_name = '%s'" % carrier
        result = inquire(sql)
        if result:
            if result[0][0] == "":
                return "no record"
            else:
                return result[0][0]
        else:
            return "no record"

    def build_pos_neg(self):
        """ build a positive and negative arrays """
        self.positivedate = []
        self.negativedate = []
        self.get_carrierlist()  # assigns self.carrier_list for investigation range
        for carrier in self.carrier_list:
            sql = "SELECT senior_date FROM seniority WHERE name = '%s'" % carrier[0][1]
            result = inquire(sql)
            if result:  # build an array with seniority dates
                sen_date = result[0][0]
                to_add = [carrier[0][1], sen_date]
                self.positivedate.append(to_add)
            else:  # build an array without seniority dates
                to_add = [carrier[0][1], ""]
                self.negativedate.append(to_add)
        self.positivedate.sort(key=itemgetter(1))  # sort the list with seniority dates by date.

    def pay_period_guide(self):
        """
        creates a txt file which is saved in the archive which list out the pay periods for a year.
        """
        i = 0
        year = simpledialog.askinteger("Pay Period Guide", "Enter the year you want generated.", parent=self.frame,
                                       minvalue=2, maxvalue=9999)
        if year is not None:
            firstday = datetime(1, 12, 22)
            while int(firstday.strftime("%Y")) != year - 1:
                firstday += timedelta(weeks=52)
                if int(firstday.strftime("%m")) <= 12 and int(firstday.strftime("%d")) <= 12:
                    firstday += timedelta(weeks=2)
            filename = "pp_guide" + "_" + str(year) + ".txt"  # create the filename for the text doc
            report = open(dir_path('pp_guide') + filename, "w")  # create the document
            report.write("\nPay Period Guide\n")
            report.write("Year: " + str(year) + "\n")
            report.write("---------------------------------------------\n\n")
            report.write("                 START (Sat):   END (Fri):         \n")
            for i in range(1, 27):
                # calculate dates
                wk1_start = firstday
                wk1_end = firstday + timedelta(days=6)
                wk2_start = firstday + timedelta(days=7)
                wk2_end = firstday + timedelta(days=13)
                report.write("PP: " + str(i).zfill(2) + "\n")
                report.write(
                    "\t week 1: " + wk1_start.strftime("%b %d, %Y") + " - " + wk1_end.strftime("%b %d, %Y") + "\n")
                report.write(
                    "\t week 2: " + wk2_start.strftime("%b %d, %Y") + " - " + wk2_end.strftime("%b %d, %Y") + "\n")
                # increment the first day by two weeks
                firstday += timedelta(days=14)
            # handle cases where there are 27 pay periods
            if int(firstday.strftime("%m")) <= 12 and int(firstday.strftime("%d")) <= 12:
                i += 1
                wk1_start = firstday
                wk1_end = firstday + timedelta(days=6)
                wk2_start = firstday + timedelta(days=7)
                wk2_end = firstday + timedelta(days=13)
                report.write("PP: " + str(i).zfill(2) + "\n")
                report.write(
                    "\t week 1: " + wk1_start.strftime("%b %d, %Y") + " - " + wk1_end.strftime("%b %d, %Y") + "\n")
                report.write(
                    "\t week 2: " + wk2_start.strftime("%b %d, %Y") + " - " + wk2_end.strftime("%b %d, %Y") + "\n")
            report.close()
            if sys.platform == "win32":
                os.startfile(dir_path('pp_guide') + filename)
            if sys.platform == "linux":
                subprocess.call(["xdg-open", 'kb_sub/pp_guide/' + filename])
            if sys.platform == "darwin":
                subprocess.call(["open", dir_path('pp_guide') + filename])

    @staticmethod
    def rpt_dov_history(date_array, history_array):
        """ Generate and display a report of dispatch of value times for a station """
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")  # create a file name
        filename = "report_dov_history" + "_" + stamp + ".txt"
        report = open(dir_path('report') + filename, "w")
        report.write("\nDispatch of Value History\n\n\n")
        report.write('   Showing results for:\n')
        report.write('      Station: {}\n\n'.format(projvar.invran_station))

        report.write('{:>4} {:<16} {:<7} {:<7} {:<7} {:<7} {:<7} {:<7} {:<7} \n'
                     .format("", "Effective Date", "sat", "sun", "mon", "tue", "wed", "thu", "fri"))
        report.write('     -----------------------------------------------------------------------------\n')
        i = 1
        for line in history_array:
            date = Convert(date_array[i - 1]).str_to_dt()
            date = Convert(date).dt_to_backslash_str()
            report.write('{:>4} {:<16} {:<7} {:<7} {:<7} {:<7} {:<7} {:<7} {:<7}\n'
                         .format("", date, line[0], line[1], line[2], line[3], line[4], line[5], line[6]))
            if i % 3 == 0:
                report.write('     -----------------------------------------------------------------------------\n')
            i += 1
        report.write('\n\n\n')
        report.write(' This report shows the settings for dispatch of value (DOV) times. The most recent \n'
                     ' records will be on the top and earlier records lower on the list. \n\n'
                     ' Effective Date is the first day of the service week and will always be a Saturday. \n'
                     ' this is the date on which the record is effective. The record will apply to later \n'
                     ' days of the week until updated/changed.\n\n'
                     ' * Asterisks denote a temporary record. Such records will only apply for one day, \n'
                     ' after which the earlier non-temporary will apply.\n\n'
                     ' The bottom record is the default. It generates automatically and can not/ should \n'
                     ' not be deleted. The time set in the default records is arbitrary and is not \n'
                     ' necessarily the correct time.')
        report.close()
        if sys.platform == "win32":  # open the text document
            os.startfile(dir_path('report') + filename)
        if sys.platform == "linux":
            subprocess.call(["xdg-open", 'kb_sub/report/' + filename])
        if sys.platform == "darwin":
            subprocess.call(["open", dir_path('report') + filename])


class CheatSheet:
    """
    This class generates a cheatsheet which allows the user to view a generated document showing codes they need to
    read TACS reports.
    """

    def __init__(self):
        pass

    @staticmethod
    def tacs_cheatsheet():
        """ generate a tacs cheatsheet for the user. """
        stamp = datetime.now().strftime("%d%H%M%S")  # create a file name
        filename = "tacs_cheatsheet" + "_" + stamp + ".txt"
        report = open(dir_path('report') + filename, "w")
        report.write("TACS Cheat Sheet\n")
        report.write("\nD/A (Designation and Activity) Codes:\n\n")
        report.write("13-4 .......... Full Time Regular (FTR)\n")
        report.write("33-4 .......... Part Time Regular (PTR)\n")
        report.write("43-4 .......... Part Time Flexible (PTF)\n")
        report.write("84-4 .......... City Carrier Assistant (CCA)\n")
        report.write("11-0 .......... Clerk\n")
        report.write("81-3 .......... Clerk\n")
        report.write("16-6 .......... Maintenance\n")
        report.write("09-0 .......... Supervisor/ Manager\n")
        report.write("\nHour Codes:\n\n")
        report.write("5200 .......... Work Hours\n")
        report.write("5300 .......... Overtime Hours\n")
        report.write("4300 .......... Penalty Overtime\n")
        report.write("4800 .......... Holiday Premium Pay\n")
        report.write("49## .......... OWCP Leave Without Pay (LWOP)\n")
        report.write("5400 .......... Night Work Premium\n")
        report.write("55## .......... Annual Leave\n")
        report.write("56## .......... Sick Leave\n")
        report.write("5800 .......... Holiday Leave\n")
        report.write("59## .......... Leave Without Pay (LWOP) - Part Day\n")
        report.write("60## .......... Leave Without Pay (LWOP) - Full Day\n")
        report.write("61## .......... Court Leave (paid)\n")
        report.write("2400 .......... Absent Without Leave (AWOL)\n")
        report.write("\nOperation Codes:\n\n")
        report.write("721 ........... Street Time\n")
        report.write("722 ........... Office Time\n")
        report.write("613 ........... Stewards Time\n")
        report.write("354 ........... Standby Time\n")
        report.write("743 ........... Route Maintenance\n")
        report.write("\nMove Codes:\n\n")
        report.write("BT ............ Begin Tour\n")
        report.write("MV 7210-## .... Move to Street\n")
        report.write("MV 7220-## .... Move to Office\n")
        report.write("073 ........... Out of Schedule Premium\n")
        report.write("093 ........... No Lunch\n")
        report.write("OL ............ Begin Lunch\n")
        report.write("IL ............ End Lunch\n")
        report.write("ET ............ End Tour\n")
        report.close()
        if sys.platform == "win32":  # open the text document
            os.startfile(dir_path('report') + filename)
        if sys.platform == "linux":
            subprocess.call(["xdg-open", dir_path('report') + filename])
        if sys.platform == "darwin":
            subprocess.call(["open", dir_path('report') + filename])

    @staticmethod
    def spdsht_cheatsheet():
        """ generate a speed sheet cheatsheet for the user. """
        stamp = datetime.now().strftime("%d%H%M%S")  # create a file name
        filename = "speedsheet_cheatsheet" + "_" + stamp + ".txt"
        report = open(dir_path('report') + filename, "w")
        report.write("Klusterbox SpeedSheet CheatSheet\n")
        report.write("\n\n")
        report.write("________________________Carrier Information___________________________\n")
        report.write("\n")
        report.write("Days: \"sat\", \"mon\", \"tue\", \"wed\", \"thu\", \"fri\" --> default is \"none\"\n")
        report.write("\n")
        report.write("Carrier Name: full last name, first initial separated by comma\n")
        report.write("\n")
        report.write("List: \"otdl\", \"wal\", \"nl\", \"aux\", \"ptf\" --> default is \"nl\"\n")
        report.write("\n")
        report.write("NS Day: \"sat\", \"mon\", \"tue\", \"wed\", \"thu\", \"fri\" --> default is \"none\"\n")
        report.write("	if rotating:\n")
        report.write("		\"rsat\", \"rmon\", \"rtue\", \"rwed\", \"rthu\", \"rfri\"\n")
        report.write("	if fixed:\n")
        report.write("		\"fsat\", \"fmon\", \"ftue\", \"fwed\", \"fthu\", \"ffri\"\n")
        report.write("\n")
        report.write("Route/s: 4 or 5 digits. 1st and 2nd place are zone\n")
        report.write("			3rd, 4th and 5th are route\n")
        report.write("	e.g. 1024 or 10124 or 0924\n")
        report.write("\n")
        report.write("Emp id: 8 digit employee id number\n")
        report.write("\n\n")
        report.write("_________________________Rings Information____________________________\n")
        report.write("\n")
        report.write("5200: Hours worked: number between 0 and 24\n")
        report.write("\n")
        report.write("MOVES: if \"moves notation route first\" is False:\n")
        report.write("	time move off route + time move on route + route number\n")
        report.write("       move sets separated by \"/\"\n")
        report.write("       e.g. 14.52+15.88+0152/15.89+16.32+0155\n")
        report.write("\n")
        report.write("	if \"moves notation route first\" is True:\n")
        report.write("	route number + time move off route + time move on route\n")
        report.write("       move sets separated by \"/\"\n")
        report.write("       e.g. 0152+14.52+15.88/0155+15.89+16.32\n")
        report.write("\n")
        report.write("RS: Return to station time: number between 0 and 24\n")
        report.write("\n")
        report.write("CODE: if List Status is \"wal\" or \"nl\":\n")
        report.write("           \"none\", \"ns day\"\n")
        report.write("      if List Status is \"otdl\", \"aux\", or \"ptf\":\n")
        report.write("           \"none\", \"ns day\", \"no call\", \"light\", \"sch chg\",")
        report.write("           \"annual\", \"sick\", \"excused\"")
        report.write("\n")
        report.write("LV type: Leave type: \"none\", \"annual\", \"sick\", \"holiday\", \"other\"\n")
        report.write("\n")
        report.write("LV time: Hours of paid leave: number between 0 and 8\n")
        report.close()
        if sys.platform == "win32":  # open the text document
            os.startfile(dir_path('report') + filename)
        if sys.platform == "linux":
            subprocess.call(["xdg-open", dir_path('report') + filename])
        if sys.platform == "darwin":
            subprocess.call(["open", dir_path('report') + filename])


class Messenger:
    """
    The Messenger class gives the location of the program and also provides the user information in the form of
    message boxes.
    """

    def __init__(self, frame):
        self.frame = frame

    def location_klusterbox(self):
        """ provides the location of the program """
        archive = ""
        dbase = None
        path = None
        if sys.platform == "darwin":
            if projvar.platform == "macapp":
                path = "Applications"
                dbase = os.path.expanduser("~") + '/Documents/.klusterbox/' + 'mandates.sqlite'
                archive = os.path.expanduser("~") + '/Documents/klusterbox'
            if projvar.platform == "py":
                path = os.getcwd()
                dbase = os.getcwd() + '/kb_sub/mandates.sqlite'
                archive = os.getcwd() + '/kb_sub'
        else:
            if projvar.platform == "winapp":
                path = os.getcwd()
                dbase = os.path.expanduser("~") + '\\Documents\\.klusterbox\\' + 'mandates.sqlite'
                archive = os.path.expanduser("~") + '\\Documents\\klusterbox'
            else:
                path = os.getcwd()
                dbase = os.getcwd() + '\\kb_sub\\mandates.sqlite'
                archive = os.getcwd() + '\\kb_sub'

        messagebox.showinfo("KLUSTERBOX ",
                            "On this computer Klusterbox is located at:\n"
                            "{}\n\nThe Klusterbox database is located at \n"
                            "{}\n\nThe Klusterbox archive is located at \n"
                            "{}".format(path, dbase, archive),
                            parent=self.frame)

    def tolerance_info(self, switch):
        """ generates a message box giving information on options/buttons. """
        text = ""
        if switch == "OT_own_route":
            text = "Sets the tolerance for no list carrier overtime\n" \
                   "\n" \
                   "Enter a value in clicks between 0 and .99"
        if switch == "OT_off_route":
            text = "Sets the tolerance for no list and work assignment \n" \
                   "list carriers for overtime off their own routes.\n\n" \
                   "Enter a value in clicks between 0 and .99"
        if switch == "availability":
            text = "Sets the tolerance for availability of otdl and " \
                   "aux carriers. Applies to availability to 10, 11.5 \n" \
                   "and 12 hour columns.\n\n" \
                   "Enter a value in clicks between 0 and .99"
        if switch == "min_nl":
            text = "Sets the minimum number of rows for the No List " \
                   "section of the spreadsheet. \n\n" \
                   "Enter a value between 1 and 100"
        if switch == "min_wal":
            text = "Sets the minimum number of rows for the Work Assignment " \
                   "section of the spreadsheet. \n\n" \
                   "Enter a value between 1 and 100"
        if switch == "min_otdl":
            text = "Sets the minimum number of rows for the OT Desired " \
                   "section of the spreadsheet. \n\n" \
                   "Enter a value between 1 and 100"
        if switch == "min_aux":
            text = "Sets the minimum number of rows for the Auxiliary " \
                   "section of the spreadsheet. \n\n" \
                   "Enter a value between 1 and 100"
        if switch == "min_overmax":
            text = "Sets the minimum number of rows for the " \
                   "12 and 60 Hour Violations spreadsheet. \n\n" \
                   "Enter a value between 1 and 100"
        if switch == "pb_nl_wal":
            text = "Creates a page break between No List and " \
                   "Work Assignment List on the spreadsheet \n\n" \
                   "Select ON or OFF"
        if switch == "pb_wal_otdl":
            text = "Creates a page break between Work Assignment List " \
                   "and OT Desired List on the spreadsheet \n\n" \
                   "Select ON or OFF"
        if switch == "pb_otdl_aux":
            text = "Creates a page break between the OT Desired List " \
                   " and the Auxiliary List on the spreadsheet \n\n" \
                   "Select ON or OFF"

        if switch == "pb_wal_aux":
            text = "Creates a page break between the Work Assignment List " \
                   " and the Auxiliary List on the spreadsheet \n\n" \
                   "Select ON or OFF"
        if switch == "pb_aux_otdl":
            text = "Creates a page break between the Auxiliary List \n" \
                   "and the OT Desired List on the spreadsheet \n\n" \
                   "Select ON or OFF"
        if switch == "man4_dis_limit":
            text = "Limits what is displayed in the Improper Mandates No.4 \n" \
                   "Spreadsheet for no-list and work assignment carriers \n" \
                   "(all otdl and auxiliary carriers will be displayed).\n" \
                   "\"show all\" will display all carriers. \n" \
                   "\"only workdays\" will display who worked that day. \n" \
                   "\"only mandates\" will display all carriers who worked \n" \
                   "overtime or off their routes. \n"
        if switch == "min_ot_equit":
            text = "Sets the minimum number of rows for the " \
                   "OTDL Equitability spreadsheet. \n\n" \
                   "Enter a value between 1 and 100"
        if switch == "ot_calc_pref":
            text = "Overtime Calculation Preferences:\n\n" \
                   "all: All overtime over 8 hours as well as ns days \n" \
                   "are calculated as overtime. \n\n" \
                   "off route: Only overtime hours worked off of the \n" \
                   "carrier's assignment are calculated as overtime.\n\n" \
                   "(Overtime for OTDL carriers with no assignment will \n" \
                   "automatically use the \"all\" overtime calculation.)\n"
        if switch == "min_ot_dist":
            text = "Sets the minimum number of rows for the " \
                   "Overtime Distribution spreadsheet. \n\n" \
                   "Enter a value between 1 and 100"
        if switch == "ot_calc_pref_dist":
            text = "Overtime Calculation Preferences:\n\n" \
                   "all: All overtime over 8 hours as well as ns days \n" \
                   "are calculated as overtime. \n\n" \
                   "off route: Only overtime hours worked off of the \n" \
                   "carrier's assignment are calculated as overtime.\n\n" \
                   "(Overtime for OTDL carriers with no assignment will \n" \
                   "automatically use the \"all\" overtime calculation.)\n"
        if switch == "wal_12_hour":
            text = "Work Assignment List 12 Hour Violation: \n\n" \
                   "on: Daily violations for carriers on the Work Assignment \n" \
                   "List will occur after the carrier works 12 hours.  \n" \
                   "(on is the default setting.)\n\n" \
                   "off: Daily violations for carriers on the Work Assignment \n" \
                   "List will occur after the carrier works 11.50 hours"
        if switch == "wal_dec_exempt":
            text = "Work Assignment List December Exemption: \n\n" \
                   "on: Work Assignment List Carriers will be exempted \n" \
                   "from all daily and weekly violations in the month \n" \
                   "of December.\n\n" \
                   "off: Work Assignment List Carriers will not be exempted \n" \
                   "from any violations during the month of December. \n" \
                   "(off is the default setting.)"
        if switch == "offbid_distinctpage":
            text = "Create distinct pages for each carrier. \n\n" \
                   "Selecting 'on' creates a distinct pages for each \n" \
                   "carrier. Selecting 'off' will place all carriers \n" \
                   "on one page. "
        if switch == "offbid_maxpivot":
            text = "Sets the maximum pivot which will count as a \n" \
                   "violation of the carrier's off bid assignment. \n\n" \
                   "Enter a value between 0 and 8"
        messagebox.showinfo("About Tolerances and Settings", text, parent=self.frame)


class Archive:
    """
    This class opens and deletes archives.
    """

    def __init__(self):
        self.frame = None
        # make sure that lenght of path array and label array are the same or else there will be an index error.
        self.path_array = [  # used in clear all
            'spreadsheets',
            'mandates_4',
            'over_max_spreadsheet',
            'speedsheets',
            'over_max',
            'off_bid',
            'ot_equitability',
            'ot_distribution',
            'ee_reader',
            'weekly_availability',
            'pp_guide'
        ]
        self.status_array = []  # used in clear all

    @staticmethod
    def file_dialogue(folder):
        """ opens file folders to access generated kbreports """
        if not os.path.isdir(folder):
            os.makedirs(folder)
        if projvar.platform == "py":
            file_path = filedialog.askopenfilename(initialdir=os.getcwd() + "/" + folder)
        else:
            file_path = filedialog.askopenfilename(initialdir=folder)
        if file_path:
            if sys.platform == "win32":
                os.startfile(file_path)
            if sys.platform == "linux":
                subprocess.call(["xdg-open", file_path])
            if sys.platform == "darwin":
                subprocess.call(["open", file_path])

    @staticmethod
    def remove_file(folder):
        """ removes a file and all contents """
        if os.path.isdir(folder):  # if it exist
            rmtree(folder)  # delete it

    def remove_file_var(self, frame, folder):
        """ removes a file and all contents """
        self.frame = frame
        folder = check_path(folder)
        if sys.platform == "win32":
            folder_name = folder.split("\\")
        else:
            folder_name = folder.split("/")
        folder_name = folder_name[-2]
        if not os.path.isdir(folder):
            messagebox.showwarning("Archive File Management",
                                   "The {} folder is already empty".format(folder_name),
                                   parent=self.frame)
            return
        if not messagebox.askokcancel("Archive File Management",
                                      "This will delete all the files in the {} archive. ".format(folder_name),
                                      parent=self.frame):
            return
        try:
            rmtree(folder)
            if not os.path.isdir(folder):
                messagebox.showinfo("Archive File Management",
                                    "Success! All the files in the {} archive have been deleted."
                                    .format(folder_name),
                                    parent=self.frame)
        except PermissionError:
            messagebox.showerror("Archive File Management",
                                 "Failure! {} can not be deleted because it is being used by another program."
                                 .format(folder_name),
                                 parent=frame)

    def clear_all(self, frame):
        """ this empties and deletes all archive folders."""
        self.frame = frame
        if not messagebox.askokcancel("Archive File Management",
                                      "This will delete all the files in the all archives. \n\n"
                                      "As all data used to generate spreadsheets and reports is "
                                      "kept in the klusterbox database, deleting archives is "
                                      "safe since they can easily be regenerated.",
                                      parent=self.frame):
            return
        for folder in self.path_array:  # for each in the path array
            self.clear_each(check_path(folder))  # delete the folder and record status report.
        status_string = self.build_status_string()
        messagebox.showinfo("Archive File Management",
                            "Delete all archives requested. \n\n"
                            "Report: \n"
                            "{}".format(status_string),
                            parent=self.frame)

    def clear_each(self, folder):
        """ this is called by clear all to delete individual files. """
        if not os.path.isdir(folder):
            self.status_array.append("Already empty - no action taken")
            return
        try:
            rmtree(folder)
            if not os.path.isdir(folder):
                self.status_array.append("Successfully deleted")
        except PermissionError:
            self.status_array.append("Folder in use - action failed.")

    def build_status_string(self):
        """ builds a string for the status report. """
        status_string = ""
        for i in range(len(self.status_array)):
            status_string += "    {}:  {}\n".format(self.path_array[i], self.status_array[i])
        return status_string


class InformalCIndex:
    """
    this class will generate a text file guide for informal c speedsheet indexes. Including guides for Issues,
    Decisions, Level, Docs and Grievants.
    """

    def __init__(self):
        self.station = ""
        self.issue_array = []
        self.level_array = ("informal a", "formal a", "step b", "pre arb", "arbitration")
        self.decision_array = []
        self.docs_array = ("non-applicable", "no", "yes", "unknown", "yes-not paid", "yes-in part",
                           "yes-verified", "no-moot", "no-ignore")
        self.grievant_array = []

    def speedsheet_guide(self):
        """ this method will generate a text file showing acceptable values for issue, level, decision and docs """
        self.get_issue_array()
        self.get_decision_array()
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")  # create a file name
        filename = "informal_c_index" + "_" + stamp + ".txt"
        report = open(dir_path('report') + filename, "w")
        report.write("\nInformal C Index Guide\n\n\n")
        # write the issue index
        report.write("Issue Index\n\n")
        report.write("    (selecting the issue, \n    automatically fills in the article)\n\n")
        report.write('{:>8}  {:<22} {:<4}\n'.format("index", "issue", "article"))
        report.write('----------------------------------------\n')
        for rec in self.issue_array:
            report.write('{:>8}  {:<22} {:<4}\n'.format(rec[0], rec[2], rec[1]))
        report.write("\n\n")
        # write the allowed level entries
        report.write("Level (allowed values)\n\n")
        for elem in self.level_array:
            report.write('      {:<20}\n'.format(elem))
        report.write("\n\n")
        # write decision index
        report.write("Decision Index\n\n")
        report.write('{:>8}  {:<22}\n'.format("index", "decision"))
        report.write('----------------------------\n')
        for rec in self.decision_array:
            report.write('{:>8}  {:<22}\n'.format(rec[0], rec[2]))
        report.write("\n\n")
        # write docs allowed values
        report.write("Docs (allowed values)\n\n")
        for elem in self.docs_array:
            report.write('      {:<20}\n'.format(elem))

        report.close()
        if sys.platform == "win32":  # open the text document
            os.startfile(dir_path('report') + filename)
        if sys.platform == "linux":
            subprocess.call(["xdg-open", 'kb_sub/report/' + filename])
        if sys.platform == "darwin":
            subprocess.call(["open", dir_path('report') + filename])

    def grievant_guide(self, station):
        """ this method will generate a text file showing acceptable values for issue, level, decision and docs """
        self.station = station
        self.get_grievants()
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")  # create a file name
        filename = "informal_c_carriers" + "_" + stamp + ".txt"
        report = open(dir_path('report') + filename, "w")
        report.write("\nInformal C Carrier List\n\n\n")
        # write the list of carriers
        report.write("Grievant List\n\n")
        report.write('----------------------------------------\n')
        for rec in self.grievant_array:
            report.write('    {:<22}\n'.format(rec[0]))
        report.write("\n\n")
        report.write("{} carriers are in the carrier list.\n".format(len(self.grievant_array)))
        report.close()
        if sys.platform == "win32":  # open the text document
            os.startfile(dir_path('report') + filename)
        if sys.platform == "linux":
            subprocess.call(["xdg-open", 'kb_sub/report/' + filename])
        if sys.platform == "darwin":
            subprocess.call(["open", dir_path('report') + filename])
        pass

    def get_issue_array(self):
        """ this method gets the issue array from the informalc_issuecategories table. """
        sql = "SELECT * FROM informalc_issuescategories"
        self.issue_array = inquire(sql)

    def get_decision_array(self):
        """ this method gets the decision array from the informalc_decisioncategories table. """
        sql = "SELECT * FROM informalc_decisioncategories"
        self.decision_array = inquire(sql)

    def get_grievants(self):
        """ this method will get a distinct list of carriers from the station """
        sql = "SELECT DISTINCT carrier_name FROM carriers WHERE station = '%s' " \
              "ORDER BY carrier_name ASC" % self.station
        self.grievant_array = inquire(sql)


class InformalCReports:
    """ generates some of the reports for informal c """

    def __init__(self, parent):
        self.parent = parent

    class AwardReports:
        """ this will create an awards report sorted by either a carrier - showing awards for  all grievances or
        by grievance - showing awards for all carriers.  get the awards from the db for the carrier, format that
        information into rows stored in self.award_stack which can be unpacked for display inside a report. """

        def __init__(self):
            # static variables for carrier award reports
            self.carrier = ""
            self.grv_list = []
            # static variables for grievance award reports
            self.grv_no = ""
            # static variables for both carrier and grievance award reports.
            self.select_grv = False  # false = display carrier awards, true display grievance awards
            self.award_stack = []  # this stores rows of information on carrier awards.
            self.dollar_array = []  # re initialize arrays
            self.hourrate_array = []
            self.gats_dollar_array = []
            self.gats_hourrate_array = []
            self.dollar_total = 0.0
            self.hourrate_total = 0.0
            self.gats_dollar_total = 0.0
            self.gats_hourrate_total = 0.0
            self.substack = []  # index 0 is dollars, index 1 is hourrate
            self.cum_dollar = 0.0  # cumulative dollar awards
            self.cum_hourrate = 0.0  # cumulative hourrate awards
            self.cum_gats_dollar = 0.0  # cumulative dollar gats descrepancies
            self.cum_gats_hourrate = 0.0  # cumulative hourrate gats descrepancies

        def run_grievance(self, grv_no):
            """ a master method for controlling the other methods """
            self.grv_no = grv_no
            self.select_grv = True
            self.build_stack()
            return self.award_stack

        def run_carrier(self, carrier, grv_list):
            """ this is a master method for controlling other methods. """
            self.carrier = carrier
            self.grv_list = grv_list
            self.select_grv = False
            self.build_stack()
            return self.award_stack

        def get_arrays(self, query):
            """ build the dollar, hourrate, gats_dollar and gats_hourrate arrays.
            accept the search results for informalc_awards2. """
            self.dollar_array = []  # re initialize arrays
            self.hourrate_array = []
            self.gats_dollar_array = []
            self.gats_hourrate_array = []
            for rec in query:
                if not rec[2]:  # rec[2] is the award amount
                    pass
                elif "/" in rec[2]:  # if the award is an hour/rate
                    self.hourrate_array.append(rec[2])
                else:  # if the award is a dollar value
                    self.dollar_array.append(rec[2])
                if not rec[3]:  # rec[3] is the gats descrepancy
                    pass
                elif "/" in rec[3]:  # if the gats descrepancy is an hour/rate
                    split_hourrate = rec[3].split(",")  # since the gats descrepancy can contain multiple values
                    for element in split_hourrate:  # add each of those values to the array
                        self.gats_hourrate_array.append(element)
                else:  # if the gats descrepancy is a dollar value
                    split_hourrate = rec[3].split(",")  # since the gats descrepancy can contain multiple values
                    for element in split_hourrate:  # add each of those values to the array
                        self.gats_dollar_array.append(element)

        def get_totals(self):
            """ get the totals from the dollar, hourrate, gats_dollar and gats_hourrate arrays. """

            def hourrate_adjuster(hourrate):
                """ multiple hour by rate and return adjusted hours """
                hourrate_array = hourrate.split("/")
                return float(hourrate_array[0]) * float(hourrate_array[1])

            self.dollar_total = 0.0
            self.hourrate_total = 0.0
            self.gats_dollar_total = 0.0
            self.gats_hourrate_total = 0.0
            if not self.dollar_array:  # get totals for dollar awards
                self.dollar_total = ""
            else:
                for dollar in self.dollar_array:
                    if dollar:
                        self.dollar_total += float(dollar)
                self.cum_dollar += self.dollar_total
            if not self.hourrate_array:  # get totals for hourrate awards
                self.hourrate_total = ""
            else:
                for hour in self.hourrate_array:
                    if hour:
                        self.hourrate_total += hourrate_adjuster(hour)
                self.cum_hourrate += self.hourrate_total
            if not self.gats_dollar_array:  # get totals for dollar gats descrepancies
                self.gats_dollar_total = "---"
            else:
                gats_dollar_total = 0.0
                for g_dollar in self.gats_dollar_array:
                    if g_dollar:
                        gats_dollar_total += float(g_dollar)
                self.gats_dollar_total = self.dollar_total - gats_dollar_total
                if self.gats_dollar_total <= 0:
                    self.gats_dollar_total = "good"
                else:
                    self.cum_gats_dollar += self.gats_dollar_total
            if not self.gats_hourrate_array:  # get totals for hourrate gats descrepancies
                self.gats_hourrate_total = "---"
            else:
                gats_hourrate_total = 0.0
                for g_hour in self.gats_hourrate_array:
                    if g_hour:
                        gats_hourrate_total += hourrate_adjuster(g_hour)
                self.gats_hourrate_total = self.hourrate_total - gats_hourrate_total
                if self.gats_hourrate_total <= 0:
                    self.gats_hourrate_total = "good"
                else:
                    self.cum_gats_hourrate += self.gats_hourrate_total

        def get_substack(self):
            """ get the detailed breakdown to appear under the awards and/or gats descrepancy totals. """
            self.substack = [[], []]  # this list contains a list for dollars and a list for hourrate
            awards = (self.dollar_array, self.hourrate_array)
            gats = (self.gats_dollar_array, self.gats_hourrate_array)
            for i in range(2):  # 0 if for dollar awards, 1 is for hourrate awards.
                greatest_count = max(len(awards[i]), len(gats[i]))
                for ii in range(greatest_count):
                    line = []
                    try:
                        if awards[i][ii]:  # if the index is valid
                            if i == 0 and len(awards[i]) > 1:  # if there is more than one value for dollar awards
                                line.append(awards[i][ii])  # add this to the line
                            elif i == 1 and len(awards[i]) > 0:  # if there is one or more for hourrate awards
                                line.append(awards[i][ii])  # add this to the line
                            else:
                                line.append("---")  # add '---' to the line to indicate no value
                    except IndexError:
                        line.append("---")  # add '---' to the line to indicate no value
                    try:
                        if gats[i][ii]:  # if the index is valid
                            line.append(gats[i][ii])
                    except IndexError:
                        line.append("---")  # add '---' to the line to indicate no value
                    if line != ["---", "---"]:
                        self.substack[i].append(line)

        @staticmethod
        def convert_dollar_hourrate(value, _type):
            """ convert the hourrate or dollars value to a floating value """
            if value == "---":
                return "   ---  "
            if value == "good":
                return "  good  "
            if _type == "dollar":
                return "  ${:,.2f}  ".format(float(value)).lstrip('0')
            if _type == "sub_dollar":
                return "+ ${:,.2f}  ".format(float(value)).lstrip('0')
            if _type == "hourrate":
                return "  {:,.2f}  ".format(float(value)).lstrip('0')
            if _type == "sub_hourrate":
                hourrate = value.split("/")
                hourrate[1] = float(hourrate[1]) * 100
                value_string = str(hourrate[0]) + " @ " + str(int(hourrate[1])) + "%"
                return "+ {}  ".format(value_string)

        def build_stack(self):
            """ this builds the awards stack, each row represents a grievance. """
            noaward_count = 1
            dollar_count = 1
            hourrate_count = 1
            noaward_stack = []
            dollar_stack = []
            hourrate_stack = []
            if self.select_grv:  # find awards for grievances 
                sql = "SELECT DISTINCT(carrier_name) FROM informalc_awards2 WHERE grv_no='%s' ORDER BY carrier_name" \
                      % self.grv_no
                result = inquire(sql)  # get a distinct list of carriers with awards in the settlement
                selection_list = distinctresult_to_list(result)  # convert result from inquiry to a list.
            else:  # find awards for carriers
                sql = "SELECT DISTINCT(grv_no) FROM informalc_awards2 WHERE carrier_name='%s'" \
                      % self.carrier
                result = inquire(sql)  # get a distinct list of carriers with awards in the settlement
                inclusive_list = distinctresult_to_list(result)  # convert result from inquiry to a list.
                selection_list = [x for x in self.grv_list if x in inclusive_list]
            # if self.select_grv is true: selection is grv_no. if false: selection is carrier_name
            for selection in selection_list:
                if self.select_grv:
                    sql = "SELECT * FROM informalc_awards2 WHERE carrier_name='%s' AND grv_no='%s'" \
                          % (selection, self.grv_no)
                    query = inquire(sql)  # get all records of awards for that carrier.
                else:
                    sql = "SELECT * FROM informalc_awards2 WHERE carrier_name='%s' AND grv_no='%s'" \
                          % (self.carrier, selection)
                    query = inquire(sql)  # get all records of awards for that carrier for a specific grievance
                
                self.get_arrays(query)
                self.get_totals()
                self.get_substack()
                # build the award stack, line by line.
                if not self.dollar_array and not self.hourrate_array:  # if there is no award
                    row = '    {:<5}{:<18}{:>15}{:>15}\n' \
                        .format(str(noaward_count), selection, "   ---  ", "   ---  ")
                    noaward_stack.append(row)
                    noaward_count += 1
                if self.dollar_array:  # if there is a dollar award
                    dollar_total_place = self.convert_dollar_hourrate(self.dollar_total, "dollar")
                    gats_dollar_total_place = self.convert_dollar_hourrate(self.gats_dollar_total, "dollar")
                    row = '    {:<5}{:<18}{:>15}{:>15}\n'\
                        .format(str(dollar_count), selection, dollar_total_place, gats_dollar_total_place)
                    dollar_stack.append(row)
                    for element in self.substack[0]:  # for each dollar element in substack
                        awards_place = self.convert_dollar_hourrate(element[0], "sub_dollar")
                        gats_place = self.convert_dollar_hourrate(element[1], "sub_dollar")
                        row = '    {:<5}{:<18}{:>15}{:>15}\n'.format("", "", awards_place, gats_place)
                        dollar_stack.append(row)
                    dollar_count += 1
                if self.hourrate_array:  # if there is an hour rate award
                    hourrate_total_place = self.convert_dollar_hourrate(self.hourrate_total, "hourrate")
                    gats_hourrate_total_place = self.convert_dollar_hourrate(self.gats_hourrate_total, "hourrate")
                    row = '    {:<5}{:<18}{:>15}{:>15}\n' \
                        .format(str(hourrate_count), selection, hourrate_total_place, gats_hourrate_total_place)
                    hourrate_stack.append(row)
                    for element in self.substack[1]:  # for each hourrate element in substack
                        awards_place = self.convert_dollar_hourrate(element[0], "sub_hourrate")
                        gats_place = self.convert_dollar_hourrate(element[1], "sub_hourrate")
                        row = '    {:<5}{:<5}{:>28}{:>15}\n'.format("", "", awards_place, gats_place)
                        hourrate_stack.append(row)
                    hourrate_count += 1
            if dollar_stack or hourrate_stack:  # if there is somthing, write column headers and totals
                totaldollars = "${:,.2f}".format(float(self.cum_dollar))
                totalhours = "{:,.2f}".format(float(self.cum_hourrate))
                totalgatsdollars = "${:,.2f}".format(float(self.cum_gats_dollar))
                totalgatshours = "{:,.2f}".format(float(self.cum_gats_hourrate))
                if self.select_grv:
                    firstrow = ["         Carrier Name             Awards    Gats Descrepancies\n", ]
                else:
                    firstrow = ["         Grievance Number         Awards    Gats Descrepancies\n", ]
                line_row = ["    --------------------------------------------------------------\n", ]
                noaward_label = ["														  no award\n"]
                dollars_label = ["														   dollars\n"]
                hourrate_label = ["													     hour/rate\n"]
                totaldollarsrow = ["    {:<19}{:>17}\n".format("Cumulative dollars:", totaldollars), ]
                totalhoursrow = ["    {:<19}{:>17}\n".format("Cumulative hours:  ", totalhours), ]
                totalgatsdollarsrow = \
                    ["    {:<36}{:>15}\n".format("Cumulative gats dollar descepancies:", totalgatsdollars), ]
                totalgatshoursrow = \
                    ["    {:<36}{:>15}\n".format("Cumulative gats hour descepancies:  ", totalgatshours), ]
                skip_line = ["\n"]
                self.award_stack = firstrow + line_row
                if noaward_stack:
                    self.award_stack += noaward_label
                    self.award_stack += noaward_stack
                    self.award_stack += line_row
                if dollar_stack:
                    self.award_stack += dollars_label
                    self.award_stack += dollar_stack
                    self.award_stack += line_row
                if hourrate_stack:
                    self.award_stack += hourrate_label
                    self.award_stack += hourrate_stack
                    self.award_stack += line_row
                if dollar_stack:
                    self.award_stack += totaldollarsrow
                if hourrate_stack:
                    self.award_stack += totalhoursrow
                if self.cum_gats_dollar or self.cum_gats_hourrate:
                    self.award_stack += skip_line
                if self.cum_gats_dollar:
                    self.award_stack += totalgatsdollarsrow
                if self.cum_gats_hourrate:
                    self.award_stack += totalgatshoursrow
            else:
                self.award_stack = ["    There are no awards entered for this settlement."]

    class IndexReports:
        """ get index data from db, sort the data into 'first array' and 'second array' for each index, 
        then write a report to be displayed for informal c reports. """

        def __init__(self):
            self.index_recs = [[], [], [], []]  # store results in self.index_recs
            self.first_array = [[], [], [], []]  # store first values of index/associations
            self.second_array = [[], [], [], []]  # store second values of index/associations
            self.grv_no = ""  # the grievance number being investigated
            self.reports_array = []  # store all the reports in this array

        def run(self, grv_no):
            """ master method for controlling sequence of methods """
            self.grv_no = grv_no  # the grievance number being investigated
            self.get_index_recs()  # use sql to get data from db
            self.sort_recs()  # sort into 'first array' and 'second array'
            self.gen_index_reports()  # generate reports and place in 'self.reports_array'
            return self.reports_array

        def get_index_recs(self):
            """ calls indexes/associations from tables and puts all in self.index_recs multi array. """
            tables_array = ("informalc_noncindex", "informalc_remandindex",  # search these tables
                            "informalc_batchindex", "informalc_gats")
            # search these columns in the tables
            first_search_criteria = ("followup", "refiling", "main", "grv_no")
            second_search_criteria = ("overdue", "remanded", "sub", "gats_no")
            for i in range(4):  # loop for each table
                sql = "SELECT * FROM %s WHERE %s = '%s' OR %s = '%s'" % \
                      (tables_array[i], first_search_criteria[i], self.grv_no, second_search_criteria[i], self.grv_no)
                # capture all records where the grv no is first or second value
                result = inquire(sql)
                if result:
                    for r in result:
                        self.index_recs[i].append(r)

        def sort_recs(self):
            """ this will sort values into 'first array' and 'second array' sort the records into arrays
            depending on if the grv no is the first or second value in the record."""
            #
            for i in range(4):
                if self.index_recs[i]:  # if there is any rec this iteration...
                    for r in self.index_recs[i]:
                        if r[0] == self.grv_no:  # if the grv no is the first value
                            self.first_array[i].append(r[1])  # capture 2nd: "overdue", "remanded", "sub", "gats_no"
                        if r[1] == self.grv_no:  # if the grv no is the second value
                            self.second_array[i].append(r[0])  # capture 1st: "followup", "refiling", "main", "grv_no"

        def gen_index_reports(self):
            """ generates reports for grievances/ settlements for non compliance, remanded, batch settlement and
                    gats reports indexes. """
            first_message = (
                "    This is a non compliance grievance for: \n",
                "    This is a refiling of a remanded grievance/s. \n",
                "    This settlement is a batch settlement for the following: \n ",
                "    The gats reports for this settlement are: \n"
            )
            second_message = (
                "    This settlement is the subject of non compliance grievance/s: \n",
                "    This grievance was remanded and refiled under grievance/s: \n",
                "    The settlement for this grievance is included in the batch settlement for: \n",
                "    Gats reports (this text generated in error): \n"
            )
            line_text = "    -----------------------------------------------------------------------\n"
            for i in range(4):
                is_gats = False  # detect is the gats report index is being read.
                if i == 3:  # since the gats report is the fourth index
                    is_gats = True
                if self.first_array[i]:  # if there is something in the first array
                    self.reports_array.append(first_message[i])  # add this to text to be displayed
                    self.reports_array.append(line_text)
                    count = 1
                    for r in self.first_array[i]:
                        if not is_gats:  # if the gats report is being read
                            rec = self.get_recforindex(r)  # get the grv/set info for the grievance
                            text = "    {:<4} {:<16}{:<20}{:<15}{:<20}\n".format(count, r, rec[0], rec[1], rec[2])
                        else:  # use alternate format for gats reports
                            text = "    {:<4}{:<20}\n".format(count, r)
                        self.reports_array.append(text)  # add this to text to be displayed
                        count += 1
                if self.second_array[i]:  # if there is something in the second array
                    self.reports_array.append(second_message[i])
                    self.reports_array.append(line_text)
                    count = 1
                    for r in self.second_array[i]:
                        rec = self.get_recforindex(r)  # get the grv/set info for the grievance
                        text = "    {:<4} {:<16}{:<20}{:<15}{:<20}\n".format(count, r, rec[0], rec[1], rec[2])
                        self.reports_array.append(text)  # add this to text to be displayed
                        count += 1
                if self.first_array[i] or self.second_array[i]:
                    self.reports_array.append("\n")  # add blank line for formatting to be displayed

        @staticmethod
        def get_recforindex(sub_grv_no):
            """ will get the records for grievances mentioned in indexes. """
            sql = "SELECT issue, meetingdate FROM informalc_grievances WHERE grv_no = '%s'" % sub_grv_no
            grv_result = inquire(sql)
            issue = "unknown"
            if grv_result[0][0]:
                issue = grv_result[0][0]
            meetingdate = "no date"
            if grv_result[0][1]:
                meetingdate = Convert(grv_result[0][1]).dtstring_to_backslashdate()
            sql = "SELECT decision FROM informalc_settlements WHERE grv_no = '%s'" % sub_grv_no
            set_result = inquire(sql)
            decision = "pending"
            if set_result:
                if set_result[0][0]:
                    decision = set_result[0][0]
            return [issue, meetingdate, decision]

    class EverythingReport:
        """ show a full report with all grievance particulars """

        @staticmethod
        def run(rec, count=None):
            """ show all the particulars of a grievance and the settlement.
            The count is the line number. If the line number is over 99, a new line is created to avoid
            the line number exceeding it's column. """
            everything_stack = []
            # ------------------------------------------------------------------------------------configure first line
            if count:
                num_space = 3 - (len(str(count)))  # number of spaces for number
                space = " "
                space += num_space * " "
                if count > 99:  # create a staggered line number if the line is 100+
                    everything_stack.append(str(count) + "\n" + "    Grievance Number:   " + rec[2] + "\n")
                else:
                    everything_stack.append(str(count) + space + "Grievance Number:   " + rec[2] + "\n")
            else:
                everything_stack.append("    Grievance Number:   " + rec[2] + "\n")
            # ------------------------------------------------------------------------------------- grievance details
            article = ""
            if rec[7]:
                article = "art.{}/ ".format(rec[7])
            everything_stack.append("    Article/Issue:      " + article + rec[6] + "\n")  # display issue
            everything_stack.append("    Grievant:           " + rec[0] + "\n")  # display issue
            start = Convert(rec[3]).dtstr_to_backslashstr()  # format incident start date
            end = Convert(rec[4]).dtstr_to_backslashstr()  # format incident end date
            meet = Convert(rec[5]).dtstr_to_backslashstr()  # format meeting date
            sign = Convert(rec[10]).dtstr_to_backslashstr()  # format date signed
            proof = Convert(rec[12]).dtstr_to_backslashstr()  # format date signed
            everything_stack.append("    Dates of Violation: " + start + " - " + end + "\n")  # display incident dates
            everything_stack.append("    Meeting Date:       " + meet + "\n")
            # ------------------------------------------------------------------------------------- settlement details
            if rec[8]:  # if there is not a grievance number here, there is no settlement
                everything_stack.append("    Decision:           " + rec[11] + "\n")  # display decsion
                everything_stack.append("    Signing Date:       " + sign + "\n")
                everything_stack.append("    Settlement Level    " + rec[9] + "\n")
                # only display proof due and documentation if decision is monetary remedy, back pay or adjustment.
                if rec[11] in ("monetary remedy", "back pay", "adjustment"):
                    everything_stack.append("    Proof Due:          " + proof + "\n")
                    everything_stack.append("    Documentation:      " + rec[13] + "\n\n")
            else:  # if there is no settlement record...
                everything_stack.append("\n")
                everything_stack.append("    There is no settlement entered for this grievance. \n\n")
            return everything_stack

    def everything_all_report(self):
        """ generates a text report for grievance summary.
        this is called the the button 'grievance everything' in the informalc reports screen. """
        if not len(self.parent.parent.search_result):
            return
        result = list(self.parent.parent.search_result)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = "infc_grv_list" + "_" + stamp + ".txt"
        report = open(dir_path('infc_grv') + filename, "w")
        report.write("Grievance Everything Report\n\n")
        report.write("    Station:            " + self.parent.parent.station + "\n\n")
        i = 1
        pb = ProgressBarDe(title="Informal C Reports", label="Generating Grievance Everything Report")
        pb.max_count(len(result))
        pb.start_up()
        for sett in result:
            pb.change_text("Reading settlement: {}".format(sett[2]))  # update the text of the progress bar
            pb.move_count(i)
            # ---------------------------------------------------------------------------------- get everything stack
            everything_stack = self.EverythingReport().run(sett, count=i)
            for row in everything_stack:
                report.write(row)
            # ------------------------------------------------------------------------- get index/associations report
            index_reports = self.IndexReports().run(sett[2])
            for ir in index_reports:
                report.write(ir)  # write index/associations line by line
            if index_reports:
                report.write("\n")
            # ------------------------------------------------------------------------------------- get awards stack
            if sett[11] in ("monetary remedy", "back pay"):  # skip if decision is not either.
                # grv_stack = self.GrvAwardReports().run(sett[2])
                grv_stack = self.AwardReports().run_grievance(sett[2])
                for row in grv_stack:
                    report.write(row)
                report.write("\n")
            report.write("\n\n")
            i += 1
        report.close()
        pb.stop()
        # --------------------------------------------------------------------------------------------- save and open
        try:
            if sys.platform == "win32":
                os.startfile(dir_path('infc_grv') + filename)
            if sys.platform == "linux":
                subprocess.call(["xdg-open", 'kb_sub/infc_grv/' + filename])
            if sys.platform == "darwin":
                subprocess.call(["open", dir_path('infc_grv') + filename])
        except PermissionError:
            messagebox.showerror("Report Generator",
                                 "The report was not generated.", parent=self.parent.win.topframe)

    def monetary_sum(self):
        """ generates text report for settlement list summary showing all grievance settlements.
        if gats_desc = True is passed, the report shows the gats descepancies. """

        def get_gats(grv_no):
            """ get all the gats numbers for the grievance number sent as an argument.
            return a list of gats numbers as an array or return an array with one empty string. """
            sql_ = "SELECT gats_no FROM informalc_gats WHERE grv_no = '%s'" % grv_no
            result = inquire(sql_)
            gats_array = []
            if result:
                for gats in result:
                    gats_array.append(*gats)
                return gats_array
            else:
                return [""]

        if not len(self.parent.parent.search_result):  # if there are no search results
            messagebox.showerror("Report Generator",
                                 "There are no search results to display. The report was not generated.",
                                 parent=self.parent.win.topframe)
            return
        # ------------------------------------------------------------------------------------------ generate file name
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = "infc_grv_list" + "_" + stamp + ".txt"
        report = open(dir_path('infc_grv') + filename, "w")
        # ------------------------------------------------------------------------------------------- generate document
        report.write("   Monetary Remedy Summary\n\n")
        report.write("   Only settlements of \'monetary remedy\' or \'back pay are displayed\'\n\n")
        report.write('  {:<18}{:<12}{:>9}{:>11}{:>12}{:>12}{:>12}\n'
                     .format("    Grievance #", "Date Signed", "GATS #", "Docs?", "Level", "Hours", "Dollars"))
        report.write(
            "      ----------------------------------------------------------------------------------\n")
        # ----------------------------------------------------------------------------------------------- collect data
        total_hour = 0
        total_dollar = 0
        i = 1
        monetary_remedies = []  # store all grievances where there is a monetary remedy settlement
        for sett in self.parent.parent.search_result:
            if sett[11] in ("monetary remedy", "back pay"):  # find decisions of monetary / back pay
                monetary_remedies.append(sett)
        # ----------------------------------------------------------------------------------- loop for each settlement
        for sett in monetary_remedies:  # for each settlement with monetary remedy.
            sql = "SELECT * FROM informalc_awards2 WHERE grv_no='%s'" % sett[2]
            query = inquire(sql)
            award_hour = 0
            award_dollar = 0
            # ------------------------------------------------------------------------------------ loop for each award
            for rec in query:  # for each award in the given settlement
                if not rec[2]:  # if there is no award
                    pass
                elif "/" in rec[2]:  # if the award is an hourrate type
                    hourrate = rec[2].split("/")
                    award_hour += float(hourrate[0]) * float(hourrate[1])  # get the adjusted hours
                else:  # if none of the above apply, the award is a dollar type
                    award_dollar += float(rec[2])
            # --------------------------------------------------- after loop has sorted and added all of the awards...
            total_hour += award_hour  # increment the total that will appear on the bottom of the report.
            total_dollar += award_dollar
            # next, format the awards to the format which will appear on the report.
            award_hour = Convert(award_hour).empty_returns_str("   ----")  # if zero, convert to blank lines
            if award_hour == "   ----":
                pass
            else:  # if award is not blank lines, then format the number
                award_hour = "{:.2f}".format(float(award_hour)).lstrip('0')
            award_dollar = Convert(award_dollar).empty_returns_str("   ----")  # if zero, convert to blank lines
            if award_dollar == "   ----":
                pass
            else:  # if award is not blank lines, then format the number
                award_dollar = "${:.2f}".format(float(award_dollar)).lstrip('0')
            # get the signing date of the settlement
            if not DateTimeChecker().check_dtstring(sett[10]):  # if the date signed can not be made to dt
                sign = ""
            else:
                sign = dt_converter(sett[10]).strftime("%m/%d/%Y")  # convert date time to mm/dd/yyyy format
            # get the level of the settlement
            if sett[9] is None or sett[9] == "unknown":  # format level to '---' or the level as a string.
                lvl = "---"
            else:
                lvl = sett[9]
            s_gats = get_gats(sett[2])  # get all gats information from informalc_gats table
            for gi in range(len(s_gats)):  # for gats_no in s_gats:
                if gi == 0:  # for the first line
                    # line #, Grievance #, Date Signed, GATS #, Docs?, Level, Hours, Dollars
                    report.write('{:>4}  {:<14}{:<12}{:<9}{:>11}{:>12}{:>12}{:>12}\n'
                                 .format(str(i), sett[2], sign, s_gats[gi], sett[13], lvl, award_hour, award_dollar))
                if gi != 0:
                    report.write('{:<32}{:<12}\n'.format("", s_gats[gi]))
            if i % 3 == 0:
                report.write(
                    "      ----------------------------------------------------------------------------------\n")
            i += 1
        report.write("      ----------------------------------------------------------------------------------\n")
        # --------------------------------------------------------------------------------------------- end of report
        report.write("{:<20}{:>56}\n".format("      Total Hours", "{0:.2f}".format(total_hour)))
        report.write("{:<20}{:>68}\n".format("      Total Dollars", "${0:.2f}".format(total_dollar)))
        report.close()
        # --------------------------------------------------------------------------------------------- save and open
        try:
            if sys.platform == "win32":
                os.startfile(dir_path('infc_grv') + filename)
            if sys.platform == "linux":
                subprocess.call(["xdg-open", 'kb_sub/infc_grv/' + filename])
            if sys.platform == "darwin":
                subprocess.call(["open", dir_path('infc_grv') + filename])
        except PermissionError:
            messagebox.showerror("Report Generator",
                                 "The report was not generated.", parent=self.parent.win.topframe)

    def gats_descrepancies(self, fullreport=True):
        """ generates text report for settlement list summary showing all grievance settlements.
        if fullreport = True is passed, show all settlements. If False, show only settlements with gats
        descrepancies. """

        def get_gats(grv_no):
            """ get all the gats numbers for the grievance number sent as an argument.
            return a list of gats numbers as an array or return an array with one empty string. """
            sql_ = "SELECT gats_no FROM informalc_gats WHERE grv_no = '%s'" % grv_no
            result = inquire(sql_)
            gats_array = []
            if result:
                for gats in result:
                    gats_array.append(*gats)
                return gats_array
            else:
                return [""]

        def hourrate_adjuster(hourrate):
            """ multiple hour by rate and return adjusted hours """
            hourrate_split = hourrate.split("/")
            return float(hourrate_split[0]) * float(hourrate_split[1])

        if not len(self.parent.parent.search_result):  # if there are no search results
            messagebox.showerror("Report Generator",
                                 "There are no search results to display. The report was not generated.",
                                 parent=self.parent.win.topframe)
            return
        # ------------------------------------------------------------------------------------------ generate file name
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = "infc_grv_list" + "_" + stamp + ".txt"
        report = open(dir_path('infc_grv') + filename, "w")
        # ------------------------------------------------------------------------------------------- generate document
        report.write("   Monetary Remedy Summary with Gats Descrepanies\n\n")
        report.write("   Only settlements of \'monetary remedy\' or \'back pay are displayed\'\n")
        if not fullreport:
            report.write("   Only settlements where gats descrepancies are noted are displayed \n")
        report.write('\n  {:<47}{:<24}{:<24}\n'.format("", "Settlement Awards", "Gats Descrepancies"))
        report.write('  {:<18}{:<12}{:<10}{:>12}{:>12}{:>12}{:>12}\n'
                     .format("    Grievance #", "GATS #", "Docs?", "Hours", "Dollars", "Hours", "Dollars"))
        report.write("      -----------------------------------------------------------------------------------\n")

        # ----------------------------------------------------------------------------------------------- collect data
        total_hour = 0.0
        total_dollar = 0.0
        total_gats_hour = 0.0
        total_gats_dollar = 0.0
        i = 1
        monetary_remedies = []  # store all grievances where there is a monetary remedy settlement
        for sett in self.parent.parent.search_result:
            if sett[11] in ("monetary remedy", "back pay"):  # find decisions of monetary / back pay
                monetary_remedies.append(sett)
        # -------------------------------------------------------------------------------- loop for each settlement
        # create progress bar
        pb = ProgressBarDe(title="Informal C Reports", label="Generating Gats Descrepancies Report")
        pb.max_count(len(monetary_remedies))  # get count of the progress bar
        pb.start_up()  # start the progress bar
        pb_counter = 1  # initialize the count of the progress bar
        for sett in monetary_remedies:  # for each settlement with monetary remedy.
            sql = "SELECT DISTINCT carrier_name FROM informalc_awards2 WHERE grv_no='%s'" % sett[2]
            query = inquire(sql)
            query = distinctresult_to_list(query)
            query.sort()  # sort the names alphabetically
            award_hour = 0.0
            award_dollar = 0.0
            gats_hour = 0.0
            gats_dollar = 0.0
            # --------------------------------------------------------------------------------- loop for each name
            pb.move_count(pb_counter)  # increment the count of the progress bar
            pb.change_text("Reading settlement: {}".format(sett[2]))  # update the text of the progress bar
            for name in query:  # for each award in the given settlement
                carrier_hourrate_total = 0.0
                carrier_dollar_total = 0.0
                carrier_gats_hourrate_total = 0.0
                carrier_gats_dollar_total = 0.0
                dollar_array = []  # re initialize arrays
                hourrate_array = []
                gats_dollar_array = []
                gats_hourrate_array = []
                # for each gats descrepancy in the given settlement
                sql = "SELECT * FROM informalc_awards2 WHERE carrier_name='%s' AND grv_no='%s'" \
                      % (name, sett[2])
                query = inquire(sql)  # get all records of awards for that carrier.
                # sort awards/gats into list
                # -------------------------------------------------------------------- loop for each award/ descrepancy
                for rec in query:
                    if not rec[2]:  # rec[2] is the award amount
                        pass
                    elif "/" in rec[2]:  # if the award is an hour/rate
                        hourrate_array.append(rec[2])
                    else:  # if the award is a dollar value
                        dollar_array.append(rec[2])
                    if not rec[3]:  # rec[3] is the gats descrepancy
                        pass
                    elif "/" in rec[3]:  # if the gats descrepancy is an hour/rate
                        split_hourrate = rec[3].split(",")  # since the gats descrepancy can contain multiple values
                        for element in split_hourrate:  # add each of those values to the array
                            gats_hourrate_array.append(element)
                    else:  # if the gats descrepancy is a dollar value
                        split_hourrate = rec[3].split(",")  # since the gats descrepancy can contain multiple values
                        for element in split_hourrate:  # add each of those values to the array
                            gats_dollar_array.append(element)
                # ------------------------------------------------------------------------ increment carrier totals
                for dollar in dollar_array:
                    carrier_dollar_total += float(dollar)
                for hour in hourrate_array:
                    carrier_hourrate_total += hourrate_adjuster(hour)
                for g_dollar in gats_dollar_array:
                    carrier_gats_dollar_total += float(g_dollar)
                for g_hour in gats_hourrate_array:
                    carrier_gats_hourrate_total += hourrate_adjuster(g_hour)
                # --------------------------- subtract the gats total from the award total to get the true gats total
                if gats_dollar_array:
                    carrier_gats_dollar_total = max(carrier_dollar_total - carrier_gats_dollar_total, 0.0)
                if gats_hourrate_array:
                    carrier_gats_hourrate_total = max(carrier_hourrate_total - carrier_gats_hourrate_total, 0.0)
                # ------------------------------ add carrier awards/ descrepancies to settlement awards/ descrepancies
                award_hour += carrier_hourrate_total
                award_dollar += carrier_dollar_total
                gats_hour += carrier_gats_hourrate_total
                gats_dollar += carrier_gats_dollar_total
                # ----------------------------------------------- after loop has sorted and added all of the awards...
                total_hour += carrier_hourrate_total  # increment the total to appear on the bottom of the report
                total_dollar += carrier_dollar_total
                total_gats_hour += carrier_gats_hourrate_total
                total_gats_dollar += carrier_gats_dollar_total
            # --------------------------------------- format the awards to the format which will appear on the report.
            # if zero, convert to blank lines "   ---"
            award_hour = Convert(award_hour).empty_returns_str("   ----")
            if award_hour == "   ----":
                pass
            else:  # if award is not blank lines, then format the number
                award_hour = "{:,.2f}".format(float(award_hour)).lstrip('0')
            # if zero, convert to blank lines
            award_dollar = Convert(award_dollar).empty_returns_str("   ----")
            if award_dollar == "   ----":
                pass
            else:  # if award is not blank lines, then format the number
                award_dollar = "${:,.2f}".format(float(award_dollar)).lstrip('0')
            # if zero, convert to blank lines
            gats_hour = Convert(gats_hour).empty_returns_str("   ----")
            if gats_hour == "   ----":
                pass
            else:  # if award is not blank lines, then format the number
                gats_hour = "{:,.2f}".format(float(gats_hour)).lstrip('0')
            # if zero, convert to blank lines
            gats_dollar = Convert(gats_dollar).empty_returns_str("   ----")
            if gats_dollar == "   ----":
                pass
            else:  # if award is not blank lines, then format the number
                gats_dollar = "${:,.2f}".format(float(gats_dollar)).lstrip('0')
            s_gats = get_gats(sett[2])  # get all gats information from informalc_gats table
            # ------------------------------------------------------------------------------- write lines to report
            if fullreport:
                for gi in range(len(s_gats)):  # for gats_no in s_gats:
                    if gi == 0:  # for the first line
                        report.write('{:>4}  {:<14}{:<12}{:<10}{:>11}{:>12}{:>12}{:>12}\n'
                                     .format(str(i), sett[2], s_gats[gi], sett[13], award_hour, award_dollar,
                                             gats_hour, gats_dollar))
                    if gi != 0:  # if there is more than one gats number (s_gats), write them on their own line
                        report.write('{:<20}{:<12}\n'.format("", s_gats[gi]))
                    if i % 3 == 0:  # every third line, insert a line a for readability
                        report.write(
                            "      -------------------------------------------------------------------------------"
                            "----\n")
                i += 1
            elif gats_dollar != "   ----" and gats_hour != "   ----":
                for gi in range(len(s_gats)):  # for gats_no in s_gats:
                    if gi == 0:  # for the first line
                        report.write('{:>4}  {:<14}{:<12}{:<10}{:>11}{:>12}{:>12}{:>12}\n'
                                     .format(str(i), sett[2], s_gats[gi], sett[13], award_hour, award_dollar,
                                             gats_hour, gats_dollar))
                    if gi != 0:  # if there is more than one gats number (s_gats), write them on their own line
                        report.write('{:<20}{:<12}\n'.format("", s_gats[gi]))
                    if i % 3 == 0:  # every third line, insert a line a for readability
                        report.write("      ---------------------------------------------------------------------"
                                     "-------------\n")
                i += 1
            pb_counter += 1  # increment the count of the progress bar
        report.write(
            "      -----------------------------------------------------------------------------------\n")
        # --------------------------------------------------------------------------------------------- end of report
        report.write("{:<21}{:>32}\n".format("      Total Hours", "{0:,.2f}".format(total_hour)))
        report.write("{:<21}{:>44}\n".format("      Total Dollars", "${0:,.2f}".format(total_dollar)))
        report.write("{:<46}{:>31}\n"
                     .format("      Total Gats Descrepancies Hours", "{0:,.2f}".format(total_gats_hour)))
        report.write("{:<46}{:>43}\n"
                     .format("      Total Gats Descrepancies Dollars", "${0:,.2f}".format(total_gats_dollar)))
        report.close()
        pb.stop()  # stop and close the progress bar
        # --------------------------------------------------------------------------------------------- save and open
        try:
            if sys.platform == "win32":
                os.startfile(dir_path('infc_grv') + filename)
            if sys.platform == "linux":
                subprocess.call(["xdg-open", 'kb_sub/infc_grv/' + filename])
            if sys.platform == "darwin":
                subprocess.call(["open", dir_path('infc_grv') + filename])
        except PermissionError:
            messagebox.showerror("Report Generator",
                                 "The report was not generated.", parent=self.parent.win.topframe)

    def bycarriers(self):
        """ generates a text report for settlements by carriers. """
        # ------------------------------------------------------------------------- get carriers and grievance numbers
        unique_carrier = self.parent.uniquecarrier()  # get a list of distinct carrier names
        unique_grv = []  # get a list of all grv numbers in search range
        for grv in self.parent.parent.search_result:
            if grv[2] not in unique_grv:  # make a list of distinct grievance numbers
                unique_grv.append(grv[2])  # put these in "unique_grv"
        # ------------------------------------------------------------------------------------------------- name file
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = "infc_grv_list" + "_" + stamp + ".txt"
        report = open(dir_path('infc_grv') + filename, "w")
        # ----------------------------------------------------------------------------------------------- progress bar
        pb = ProgressBarDe(title="Informal C Reports", label="Standby. The report is generating.")
        pb.max_count(len(unique_carrier))  # the count of the pb is the number of carriers in unique carrier
        pb.start_up()  # start the progress bar
        # ---------------------------------------------------------------------------------------------- write to text
        report.write("Settlement Report By Carriers\n\n")
        pb_count = 1
        for name in unique_carrier:
            pb.move_count(pb_count)
            pb.change_text("Writing report for {}".format(name))
            report.write("{:<30}\n\n".format(name))
            # --------------------------------------------------------------------------------------- call award stack
            award_stack = self.AwardReports().run_carrier(name, unique_grv)
            for award in award_stack:
                report.write(award)
            report.write("\n\n\n")
            pb_count += 1
        report.close()
        pb.stop()
        # ---------------------------------------------------------------------------------------------- save and open
        try:
            if sys.platform == "win32":
                os.startfile(dir_path('infc_grv') + filename)
            if sys.platform == "linux":
                subprocess.call(["xdg-open", 'kb_sub/infc_grv/' + filename])
            if sys.platform == "darwin":
                subprocess.call(["open", dir_path('infc_grv') + filename])
        except PermissionError:
            messagebox.showerror("Report Generator",
                                 "The report was not generated.", parent=self.parent.win.topframe)

    def adjustments(self, fullreport=True):
        """ generates a report of settlements where the decision calls for an adjustment. """
        adjust_sett = []
        if fullreport:
            for s in self.parent.parent.search_result:  # loop through all results
                if s[11] == "adjustment":
                    adjust_sett.append(s)
        else:
            for s in self.parent.parent.search_result:  # loop through all results
                if s[11] == "adjustment" and s[13] in ("no", "no status"):
                    adjust_sett.append(s)
        if len(adjust_sett) == 0:
            msg = "There are no records matching your search results. "
            messagebox.showwarning("Informal C Reports", msg, parent=self.parent.win.topframe)
            return
        # ---------------------------------------------------------------------------------------------- file name
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = "infc_grv_list" + "_" + stamp + ".txt"
        report = open(dir_path('infc_grv') + filename, "w")
        # ------------------------------------------------------------------------------------------------ headers
        # report will display the elements (with indexes):
        # Grievance Number(2), date (determined by sort index), grievant (0), issue (6)
        report.write("   Adjustments \n\n")
        if fullreport:
            report.write("   Showing all settlements within search criteria where the decision requires \n"
                         "   an adjustment. \n\n")
        else:
            report.write("   Showing settlements within search criteria where the decision requires an adjustment \n"
                         "   and document status is \"no\" or \"no status\". \n\n")
        report.write('     {:<18}{:<20}{:<14}{:<22}{:<12}\n'
                     .format("    Grievance #", "Grievant", "Signing Date", "Issue", "Docs?"))
        report.write("       -----------------------------------------------------------------------------"
                     "----\n")
        i = 0
        for s in adjust_sett:
            sign = Convert(s[10]).dtstr_to_backslashstr()  # format date signed
            report.write('{:<5}{:<4}{:<14}{:<20}{:<14}{:<22}{:<12}\n'
                         .format("", str(i), s[2], s[0], sign, s[6], s[13]))
            i += 1
        report.close()
        # ---------------------------------------------------------------------------------------------- save and open
        try:
            if sys.platform == "win32":
                os.startfile(dir_path('infc_grv') + filename)
            if sys.platform == "linux":
                subprocess.call(["xdg-open", 'kb_sub/infc_grv/' + filename])
            if sys.platform == "darwin":
                subprocess.call(["open", dir_path('infc_grv') + filename])
        except PermissionError:
            messagebox.showerror("Report Generator",
                                 "The report was not generated.", parent=self.parent.win.topframe)

    def bycarrier_apply(self, names, cursor):
        """ generates a text report for a specified carrier. """
        if not cursor:
            return
        # ------------------------------------------------------------------------------------------------- name file
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = "infc_grv_list" + "_" + stamp + ".txt"
        report = open(dir_path('infc_grv') + filename, "w")
        unique_grv = []  # get a list of all grv numbers in search range
        for grv in self.parent.parent.search_result:
            if grv[2] not in unique_grv:
                unique_grv.append(grv[2])  # put these in "unique_grv"
        name = names[cursor[0]]
        report.write("Settlement Report By Carrier\n\n")
        report.write("{:<30}\n\n".format(name))
        # ----------------------------------------------------------------------------------------------- award stack
        award_stack = self.AwardReports().run_carrier(name, unique_grv)
        for row in award_stack:
            report.write(row)
        report.close()
        # --------------------------------------------------------------------------------------------- save and open
        try:
            if sys.platform == "win32":
                os.startfile(dir_path('infc_grv') + filename)
            if sys.platform == "linux":
                subprocess.call(["xdg-open", 'kb_sub/infc_grv/' + filename])
            if sys.platform == "darwin":
                subprocess.call(["open", dir_path('infc_grv') + filename])
        except PermissionError:
            messagebox.showerror("Report Generator", "The report was not generated.", parent=self.parent.win.topframe)

    def no_settlement(self):
        """ this a summary of all grievances which do not have settlement records. """
        no_settlement = []
        for s in self.parent.parent.search_result:  # loop through all results
            if not s[8]:  # if there is no grievance number in the settlement portion of the array
                no_settlement.append(s)  # add to a list of grvs with no settlement
        if len(no_settlement) == 0:
            msg = "There are no records matching your search results. "
            messagebox.showwarning("Informal C Reports", msg, parent=self.parent.win.topframe)
            return
        # ---------------------------------------------------------------------------------------------- file name
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = "infc_grv_list" + "_" + stamp + ".txt"
        report = open(dir_path('infc_grv') + filename, "w")
        # ------------------------------------------------------------------------------------------------ headers
        # report will display the elements (with indexes):
        # Grievance Number(2), date (determined by sort index), grievant (0), issue (6)
        report.write("   No Settlement Report\n\n")
        report.write("   Showing all grievances which do not have settlements. \n\n")
        report.write('{:>18}{:>14}  {:<20}{:<22}\n'
                     .format("    Grievance #", "Meeting Date", "Grievant", "Issue"))
        report.write("       --------------------------------------------------------------------------------"
                     "----\n")
        i = 0
        for r in no_settlement:
            formatted_date = Convert(r[5]).dtstr_to_backslashstr()
            report.write('{:>4}{:>14}{:>14}  {:<20}{:<22}\n'
                         .format(str(i + 1), r[2], formatted_date, r[0], r[6]))
            if i % 3 == 0:  # insert a line every third loop for visual clarity and readability
                report.write("       ----------------------------------------------------------------------"
                             "--------------\n")
            i += 1
        report.write("       --------------------------------------------------------------------------------"
                     "----\n")  # insert line at the end to close out report
        report.close()
        # ----------------------------------------------------------------------------------------- save and open
        if sys.platform == "win32":
            os.startfile(dir_path('infc_grv') + filename)
        if sys.platform == "linux":
            subprocess.call(["xdg-open", 'kb_sub/infc_grv/' + filename])
        if sys.platform == "darwin":
            subprocess.call(["open", dir_path('infc_grv') + filename])

    def delinquency(self):
        """ this a summary of all grievances which do not have settlement records. """

        def get_present_date():
            """ use simpledialog to get the present date """
            default = Convert(datetime.now()).dt_to_backslash_str()
            entered_date = askstring("Compliance Delinquency Report",
                                     "Enter the date the report is generated from", initialvalue=default)
            if entered_date is None:  # if the user selects 'cancel'
                return entered_date
            if not informalc_date_checker(self.parent.win.topframe, entered_date, "present day"):
                msgg = "Report will generate using the current day. Rerun the report to try again"
                messagebox.showinfo("Compliance Delinquency Report", msgg, parent=self.parent.win.topframe)
                return Convert(default).backslashdate_to_datetime()
            else:
                return Convert(entered_date).backslashdate_to_datetime()

        # ------------------------------------------------------------------------------------ get qualifying recs
        grace_period = 4  # number of weeks in the grace period before proof is due
        present_date = get_present_date()
        if present_date is None:  # if the user selects cancel, abort the report
            return
        needproof = []
        for r in self.parent.parent.search_result:  # loop through all results
            if r[11] in ("monetary remedy", "backpay", "adjustment"):  # if the grievance requires proof
                needproof.append(r)  # add to a list of grvs with no settlement
        over_due = []  # store records of grievances that require proof, but don't have it.
        for n in needproof:
            if n[13] in ("no",):  # only include recs where docs = 'no'
                over_due.append(n)
        if len(over_due) == 0:  # if there are no qualifying recs
            msg = "There are no records matching your search results. "
            messagebox.showwarning("Informal C Reports", msg, parent=self.parent.win.topframe)
            return
        # ---------------------------------------------------------------------------------------------- file name
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = "infc_grv_list" + "_" + stamp + ".txt"
        report = open(dir_path('infc_grv') + filename, "w")
        # ------------------------------------------------------------------------------------------------ headers
        # report will display the elements (with indexes):
        # Grievance Number(2), date (determined by sort index), grievant (0), issue (6)
        report.write("   Compliance Delinquency Report\n\n")
        report.write("   Showing all settlements where compliance is pending. \n\n")
        formatted_date = Convert(present_date).dt_to_backslash_str()
        report.write("   Current day of report: {}. \n\n".format(formatted_date))
        report.write("   If proof due date not specified, the due date is date signed, plus {} weeks. \n"
                     "   For most complete results, use \'Search All\' in the search criteria.\n\n"
                     .format(str(grace_period)))
        report.write('{:<7}{:<16}{:<14}{:<14}{:<14}{:<22}\n'
                     .format("", "Grievance #", "Level", "Date Signed", "Proof Due", "Delinquency"))
        report.write("       -----------------------------------------------------------------------------\n")
        i = 0
        for r in over_due:
            # ------------------------------------------------------------------------------------- get delinquency
            d_date = datetime(1, 1, 1, 0, 0)  # initialize and declare due date
            if r[12]:  # if there is a proof due date
                d_date = Convert(r[12]).str_to_dt()  # convert string to datetime
            elif r[10]:  # if there is a date signed date
                d_date = Convert(r[10]).str_to_dt() + timedelta(weeks=grace_period)
            # if there is no proof due nor date signed - due date can not be found
            if d_date == datetime(1, 1, 1, 0, 0):  # if due date hasn't changed.
                delinquency = "unknown"
            elif d_date < present_date:
                diff = present_date - d_date  # returns an int of days
                delinquency = "{} days delinquent".format(diff.days)
            elif present_date < d_date:
                diff = d_date - present_date
                delinquency = "{} days remaining".format(diff.days)
            elif d_date.date == present_date.date:
                delinquency = "due today"
            else:
                delinquency = "due today"
            # --------------------------------------------------------------------------------------- format text
            datesigned = "----------"
            if r[10]:
                datesigned = Convert(r[10]).dtstr_to_backslashstr()  # convert string to datetime
            proofdue = "----------"
            if r[12]:
                proofdue = Convert(r[12]).dtstr_to_backslashstr()  # convert string to datetime
            report.write('{:<7}{:<16}{:<14}{:<14}{:<14}{:<22}\n'
                         .format(str(i + 1), r[2], r[9], datesigned, proofdue, delinquency))
            if i % 3 == 0:  # insert a line every third loop for visual clarity and readability
                report.write("       -----------------------------------------------------------------------------\n")
            i += 1
        # insert line at the end to close out report
        report.write("       -----------------------------------------------------------------------------\n")
        report.close()
        # ----------------------------------------------------------------------------------------- save and open
        if sys.platform == "win32":
            os.startfile(dir_path('infc_grv') + filename)
        if sys.platform == "linux":
            subprocess.call(["xdg-open", 'kb_sub/infc_grv/' + filename])
        if sys.platform == "darwin":
            subprocess.call(["open", dir_path('infc_grv') + filename])

    def missing_awards(self):
        """ finds settlements with missing awards and writes a report showing grievances where awards have not
        been entered. """
        # --------------------------------------------------------------------------- find settlements missing awards
        needawards = []
        for r in self.parent.parent.search_result:  # loop through all results
            if r[11] in ("monetary remedy", "backpay"):  # if the grievance requires proof
                sql = "SELECT * FROM informalc_awards WHERE grv_no = '%s'" % r[2]  # search by grievance number
                result = inquire(sql)
                if not result:  # if there is no result from the sql search
                    needawards.append(r[2])  # all settlement to the list of those missing awards.
        # ---------------------------------------------------------------------------------------------- file name
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = "infc_missing_awards" + "_" + stamp + ".txt"
        report = open(dir_path('infc_grv') + filename, "w")
        # ---------------------------------------------------------------------------------------------- write report
        report.write("Missing Awards\n\n")
        report.write("    This report list all grievances settled for \'monetary remedy\' or \'backpay\' \n"
                     "    and checks if awards have been entered. \n\n")
        if not needawards:
            report.write("    No \'monetary remedy\' or \'backpay\' settlements were found missing awards.")
        else:
            report.write("    Missing Awards:\n")
            i = 1
            for na in needawards:
                report.write("    {:>5}. {}\n".format(str(i), na))
                i += 1
        # --------------------------------------------------------------------------------------- close, save and open
        report.close()
        try:
            if sys.platform == "win32":
                os.startfile(dir_path('infc_grv') + filename)
            if sys.platform == "linux":
                subprocess.call(["xdg-open", 'kb_sub/infc_grv/' + filename])
            if sys.platform == "darwin":
                subprocess.call(["open", dir_path('infc_grv') + filename])
        except PermissionError:
            messagebox.showerror("Report Generator",
                                 "The report was not generated.", parent=self.parent.win.topframe)

    def rptcarrierandid(self):
        """ generates a text report with only carrier name and employee id number. """
        carriers = self.parent.uniquecarrier()  # get a list of carrier names
        if len(carriers) == 0:
            messagebox.showerror("Report Generator",
                                 "There are no carriers in the carrier list. The report was not generated.",
                                 parent=self.parent.win.topframe)
            return
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = "infc_grv_list" + "_" + stamp + ".txt"
        report = open(dir_path('infc_grv') + filename, "w")
        report.write("Carrier List\n\n")
        i = 1
        for carrier in carriers:
            emp_id = ""
            sql = "SELECT emp_id FROM name_index WHERE kb_name = '%s'" % carrier
            result = inquire(sql)
            if result:
                emp_id = result[0][0]
            report.write("{:>4} {:<25}{:>8}\n".format(str(i), carrier, emp_id))
            i += 1
        report.close()
        try:
            if sys.platform == "win32":
                os.startfile(dir_path('infc_grv') + filename)
            if sys.platform == "linux":
                subprocess.call(["xdg-open", 'kb_sub/infc_grv/' + filename])
            if sys.platform == "darwin":
                subprocess.call(["open", dir_path('infc_grv') + filename])
        except PermissionError:
            messagebox.showerror("Report Generator",
                                 "The report was not generated.", parent=self.parent.win.topframe)

    def grv_summary(self):
        """ this a summary of all grievances as they appear on the search results screen.
        this is called by the button on the bottom of showtime. """
        if len(self.parent.search_result) == 0:
            msg = "There are no records matching your search results. "
            messagebox.showwarning("Informal C Reports", msg, parent=self.parent.win.topframe)
            return
        # --------------------------------------------------------- get the date by which the results are sorted...
        # "Start Incident Date", "End Incident Date", "Meeting Date", "Signed Date", "Proof Due"
        sortby = (3, 4, 5, 10, 12)  # store the indexes of the dates in this tuple.
        sort_index = sortby[int(self.parent.sortby.get())]  # sent by self.sortby stringvar
        # convert to backslash date or empty
        # selecteddate = Convert(self.search_result[i][sort_index]).dtstr_to_backslashstr()
        # ---------------------------------------------------------------------------------------------- file name
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = "infc_grv_list" + "_" + stamp + ".txt"
        report = open(dir_path('infc_grv') + filename, "w")
        # ------------------------------------------------------------------------------------------------ headers
        # report will display the elements (with indexes):
        # Grievance Number(2), date (determined by sort index), grievant (0), issue (6), decision (11)
        report.write("   Grievance List Summary\n\n")
        report.write("   Showing all grievances/settlements within search criteria\'\n\n")
        date_header = ("Start Date", "End Date", "Meeting Date", "Signed Date", "Proof Due")
        date_head_index = int(self.parent.sortby.get())
        report.write('{:>18}{:>14}  {:<20}{:<22}  {:<20}\n'
                     .format("    Grievance #", date_header[date_head_index], "Grievant", "Issue", "Settlement"))
        report.write("       ----------------------------------------------------------------------------------"
                     "----\n")
        i = 0
        for r in self.parent.search_result:
            formatted_date = Convert(r[sort_index]).dtstr_to_backslashstr()
            report.write('{:>4}{:>14}{:>14}  {:<20}{:<22}  {:<20}\n'
                         .format(str(i + 1), r[2], formatted_date, r[0], r[6], r[11]))
            if i % 3 == 0:  # insert a line every third loop for visual clarity and readability
                report.write("       ----------------------------------------------------------------------"
                             "----------------\n")
            i += 1
        report.write("       --------------------------------------------------------------------------------"
                     "------\n")  # insert line at the end to close out report
        report.close()
        # ----------------------------------------------------------------------------------------- save and open
        if sys.platform == "win32":
            os.startfile(dir_path('infc_grv') + filename)
        if sys.platform == "linux":
            subprocess.call(["xdg-open", 'kb_sub/infc_grv/' + filename])
        if sys.platform == "darwin":
            subprocess.call(["open", dir_path('infc_grv') + filename])

    def everything_report(self, grv_info):
        """ generates a text report for a specific grievance number.
        this is called by the buttons in showtime on the rows with 'report' and 'enter awards'. """
        grv_info = list(grv_info)  # correct for legacy problem of NULL Settlement Levels
        # ----------------------------------------------------------------------------------------------- name file
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = "infc_grv_list" + "_" + stamp + ".txt"
        report = open(dir_path('infc_grv') + filename, "w")
        # --------------------------------------------------------------------------------------- get everything stack
        everything_stack = self.EverythingReport().run(grv_info)
        for row in everything_stack:
            report.write(row)
        # ------------------------------------------------------------------------------- get index/associations report
        index_reports = self.IndexReports().run(grv_info[2])
        for ir in index_reports:
            report.write(ir)  # write index/associations line by line
        if index_reports:
            report.write("\n")
        # ------------------------------------------------------------------------------------------ get awards stack
        if grv_info[11] in ("monetary remedy", "back pay"):  # only run if settlement is monetary or back pay
            # grv_stack = self.GrvAwardReports().run(grv_info[2])
            grv_stack = self.AwardReports().run_grievance(grv_info[2])
            for row in grv_stack:
                report.write(row)
        report.close()
        # ------------------------------------------------------------------------------------------- save and open
        try:
            if sys.platform == "win32":
                os.startfile(dir_path('infc_grv') + filename)
            if sys.platform == "linux":
                subprocess.call(["xdg-open", 'kb_sub/infc_grv/' + filename])
            if sys.platform == "darwin":
                subprocess.call(["open", dir_path('infc_grv') + filename])
        except PermissionError:
            messagebox.showerror("Report Generator",
                                 "The report was not generated.", parent=self.parent.win.topframe)


class InformalCOptions:
    """ this class will generate text files for displaying options of issues and decisions for informalc.
    The class can accept a frame passed from the method so that a messagebox can be displayed. """

    def __init__(self):
        self.frame = None

    def issue_options(self, frame):
        """ this is open a text document showing a list of all issue options for informal c. Showing standard
        and custom. """
        self.frame = frame
        # --------------------------------------------------------------------------- find settlements missing awards
        standard_options = []  # array for standard issue categories
        custom_options = []  # array for custom issue categories
        sql = "SELECT * FROM informalc_issuescategories"
        result = inquire(sql)
        result = issuedecisionresult_sorter(result)  # sort results by first value
        for r in result:  # separate the issue categories into standard and custom
            if r[3] == "True":
                standard_options.append(r)
            else:
                custom_options.append(r)
        # ---------------------------------------------------------------------------------------------- file name
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = "infc_list_options" + "_" + stamp + ".txt"
        report = open(dir_path('infc_grv') + filename, "w")
        # ---------------------------------------------------------------------------------------------- write report
        report.write("Issue Options\n\n")
        report.write("    This report list all standard and custom issue options. Standard issue options \n"
                     "    can not be deleted. Indexes are used by Informal C Speedsheets as a shortcut.\n\n")
        report.write("\n    Standard Issue Options \n\n")
        report.write("    {:>5} {:>8}  {:<25}\n".format("Index", "Article", "Issue Option"))
        report.write("    -----------------------------------------\n")
        for so in standard_options:
            report.write("    {:>5} {:>8}  {:<25}\n".format(so[0], so[1], so[2]))
        report.write("\n")
        if not custom_options:
            report.write("    No custom options have been created.")
        else:
            report.write("\n    Custom Issue Options \n\n")
            report.write("    {:>5} {:>8}  {:<25}\n".format("Index", "Article", "Issue Option"))
            report.write("    -----------------------------------------\n")
            for co in custom_options:
                report.write("    {:>5} {:>8}  {:<25}\n".format(co[0], co[1], co[2]))
            report.write("\n")

        # --------------------------------------------------------------------------------------- close, save and open
        report.close()
        try:
            if sys.platform == "win32":
                os.startfile(dir_path('infc_grv') + filename)
            if sys.platform == "linux":
                subprocess.call(["xdg-open", 'kb_sub/infc_grv/' + filename])
            if sys.platform == "darwin":
                subprocess.call(["open", dir_path('infc_grv') + filename])
        except PermissionError:
            messagebox.showerror("Report Generator",
                                 "The report was not generated.", parent=self.frame)
            
    def decision_options(self, frame):
        """ this is open a text document showing a list of all decision options for informal c. Showing standard
        and custom. """
        self.frame = frame
        # --------------------------------------------------------------------------- find settlements missing awards
        standard_options = []  # array for standard decision categories
        custom_options = []  # array for custom decision categories
        sql = "SELECT * FROM informalc_decisioncategories"
        result = inquire(sql)
        result = issuedecisionresult_sorter(result)  # sort results by index
        for r in result:  # separate the decision categories into standard and custom
            if r[3] == "True":
                standard_options.append(r)
            else:
                custom_options.append(r)
        # ---------------------------------------------------------------------------------------------- file name
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = "infc_list_options" + "_" + stamp + ".txt"
        report = open(dir_path('infc_grv') + filename, "w")
        # ---------------------------------------------------------------------------------------------- write report
        report.write("Decision Options\n\n")
        report.write("    This report list all standard and custom decision options. Standard decision options \n"
                     "    can not be deleted. Indexes are used by Informal C Speedsheets as a shortcut.\n\n")
        report.write("\n    Standard Decision Options \n\n")
        report.write("    {:>5} {:>8}  {:<25}\n".format("Index", "Type", "Decision Option"))
        report.write("    -----------------------------------------\n")
        for so in standard_options:
            report.write("    {:>5} {:>8}  {:<25}\n".format(so[0], so[1], so[2]))
        report.write("\n")
        if not custom_options:
            report.write("    No custom options have been created.")
        else:
            report.write("\n    Custom Decision Options \n\n")
            report.write("    {:>5} {:>8}  {:<25}\n".format("Index", "Type", "Decision Option"))
            report.write("    -----------------------------------------\n")
            for co in custom_options:
                report.write("    {:>5} {:>8}  {:<25}\n".format(co[0], co[1], co[2]))
            report.write("\n")

        # --------------------------------------------------------------------------------------- close, save and open
        report.close()
        try:
            if sys.platform == "win32":
                os.startfile(dir_path('infc_grv') + filename)
            if sys.platform == "linux":
                subprocess.call(["xdg-open", 'kb_sub/infc_grv/' + filename])
            if sys.platform == "darwin":
                subprocess.call(["open", dir_path('infc_grv') + filename])
        except PermissionError:
            messagebox.showerror("Report Generator",
                                 "The report was not generated.", parent=self.frame)


class RptCarrierId:
    """
    Generate a spread sheet with the carrier's name and employee id for all carriers in the search criteria.
    """

    def __init__(self, parent):
        self.parent = parent
        self.wb = None  # workbook object
        self.carrierlist = None  # workbook name
        self.ws_header = None  # style
        self.input_name = None  # style
        self.input_s = None  # style
        self.col_header = None  # style
        self.i = 0  # this counts the rows/ number of carriers.
        self.no_empid = []  # an array for carriers with no employee id

    def run(self):
        """ this method is the master method for running all other methods in proper order """
        self.get_styles()
        self.build_workbook()
        self.set_dimensions()
        self.build_header()
        self.fill_body()
        self.show_noempid()
        self.save_open()

    def get_styles(self):
        """ Named styles for workbook """
        bd = Side(style='thin', color="80808080")  # defines borders
        self.ws_header = NamedStyle(name="ws_header", font=Font(bold=True, name='Arial', size=12))
        self.input_name = NamedStyle(name="input_name", font=Font(name='Arial', size=8),
                                     border=Border(left=bd, top=bd, right=bd, bottom=bd))
        self.input_s = NamedStyle(name="input_s", font=Font(name='Arial', size=8),
                                  border=Border(left=bd, top=bd, right=bd, bottom=bd),
                                  alignment=Alignment(horizontal='right'))
        self.col_header = NamedStyle(name="col_header", font=Font(bold=True, name='Arial', size=8),
                                     alignment=Alignment(horizontal='left'))

    def build_workbook(self):
        """ creates the workbook object """
        self.wb = Workbook()  # define the workbook
        self.carrierlist = self.wb.active  # create first worksheet
        self.carrierlist.title = "carrier list"  # title first worksheet
        self.carrierlist.oddFooter.center.text = "&A"

    def set_dimensions(self):
        """ adjust the height and width on the violations/ instructions page """
        self.carrierlist.column_dimensions["A"].width = 5
        self.carrierlist.column_dimensions["B"].width = 20
        self.carrierlist.column_dimensions["C"].width = 10

    def build_header(self):
        """ build the header of the spreadsheet """
        self.carrierlist.merge_cells('A1:R1')
        self.carrierlist['A1'] = "Carrier List with Employee ID Numbers"
        self.carrierlist['A1'].style = self.ws_header
        cell = self.carrierlist.cell(row=3, column=2)
        cell.value = "Carrier Name"
        cell.style = self.col_header
        cell = self.carrierlist.cell(row=3, column=3)
        cell.value = "Employee ID"
        cell.style = self.col_header

    def fill_body(self):
        """ this loop will fill the body of the spreadsheet with the carrier list """
        carriers = self.parent.uniquecarrier()  # get a list of carrier names
        self.i = 1
        for carrier in carriers:
            sql = "SELECT emp_id FROM name_index WHERE kb_name = '%s'" % carrier
            result = inquire(sql)
            if result:
                emp_id = result[0][0]
                cell = self.carrierlist.cell(row=self.i + 3, column=1)
                cell.value = str(self.i)
                cell.style = self.input_name
                cell = self.carrierlist.cell(row=self.i + 3, column=2)
                cell.value = carrier
                cell.style = self.input_name
                cell = self.carrierlist.cell(row=self.i + 3, column=3)
                cell.value = emp_id
                cell.style = self.input_s
                self.i += 1
            else:
                self.no_empid.append(carrier)

    def show_noempid(self):
        """ this will display the a list of carriers with no employee id. """
        if len(self.no_empid) == 0:
            return
        self.i += 4
        cell = self.carrierlist.cell(row=self.i, column=2)
        cell.value = "Carriers without Employee ID"
        cell.style = self.col_header
        i = 1
        self.i += 1
        for carrier in self.no_empid:
            cell = self.carrierlist.cell(row=self.i, column=1)
            cell.value = str(i)
            cell.style = self.input_name
            cell = self.carrierlist.cell(row=self.i, column=2)
            cell.value = carrier
            cell.style = self.input_name
            self.i += 1
            i += 1

    def save_open(self):
        """ save the spreadsheet and open """
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        xl_filename = "infc_grv_list" + "_" + stamp + ".xlsx"
        try:
            self.wb.save(dir_path('infc_grv') + xl_filename)
            messagebox.showinfo("Spreadsheet generator",
                                "Your spreadsheet was successfully generated. \n"
                                "File is named: {}".format(xl_filename),
                                parent=self.parent.win.topframe)
            if sys.platform == "win32":  # open the text document
                os.startfile(dir_path('infc_grv') + xl_filename)
            if sys.platform == "linux":
                subprocess.call(["xdg-open", 'kb_sub/infc_grv/' + xl_filename])
            if sys.platform == "darwin":
                subprocess.call(["open", dir_path('infc_grv') + xl_filename])
        except PermissionError:
            messagebox.showerror("Spreadsheet generator",
                                 "The spreadsheet was not generated. \n"
                                 "Suggestion: "
                                 "Make sure that identically named spreadsheets are closed "
                                 "(the file can't be overwritten while open).",
                                 parent=self.parent.win.topframe)
