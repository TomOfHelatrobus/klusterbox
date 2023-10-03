"""
a klusterbox module: The Klusterbox Equitability and Distribution Spreadsheet Generators
This file contains libraries for OT Equitability and OT Distribution classes. These are used to detect unequitable
distribution of overtime and mandates among non list carriers.
"""
# custom modules
import projvar
from kbtoolbox import commit, inquire, CarrierList, Convert, datetime, dir_path, Handler, Moves, Overtime, \
    ProgressBarDe, Quarter
from tkinter import messagebox
from datetime import timedelta
import sys
import os
import subprocess
# Spreadsheet Libraries
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment, PatternFill


class QuarterRecs:
    """
    gets records for all carriers in a quarter.
    """
    def __init__(self, carrier, startdate, enddate, station):
        self.carrier = carrier
        self.start = startdate
        self.end = enddate
        self.station = station

    def get_filtered_recs(self, ls):
        """ get a filtered set of records limited by list status. """
        lst = ls  # the ls arguement is an array with ot list types (otdl, wal, nl, aux, ptf)
        recset = self.get_recs()
        if recset is None:
            return
        for rec in recset:
            if rec[2] in lst:
                return recset

    def get_recs(self):
        """ get all records in the investigation range - the entire quarter """
        sql = "SELECT * FROM carriers WHERE carrier_name = '%s' and effective_date BETWEEN '%s' AND '%s' " \
              "ORDER BY effective_date DESC" \
              % (self.carrier, self.start, self.end)
        rec = inquire(sql)
        # get the relevant previous record (RPR)
        sql = "SELECT MAX(effective_date), carrier_name, list_status, ns_day, route_s, station " \
              "FROM carriers WHERE carrier_name = '%s' and effective_date <= '%s' " \
              "ORDER BY effective_date DESC" \
              % (self.carrier, self.start - timedelta(days=1))
        before_range = inquire(sql)
        add_it = True  # indicates that the RPR needs to be added to carrier records
        if rec:
            for r in rec:  # loop through all records in carrier records
                if r[0] == str(self.start):  # if a record is for the saturday in range
                    add_it = False  # do not add the RPR.
        if add_it:  # add the RPR if there is not sat range record
            if before_range[0] != (None, None, None, None, None, None):
                rec.append(before_range[0])
        #  filter out record sets with no station matches
        station_anchor = False
        for r in rec:  # loop through all carrier records
            if r[5] == self.station:  # check if at least one record matchs to the station
                station_anchor = True  # indicates that the carrecs are good for the current station.
        rec_set = []  # initialize array to put record sets into carrier list
        #  filter out any consecutive duplicate records
        if station_anchor:
            last_rec = ["xxx", "xxx", "xxx", "xxx", "xxx", "xxx"]
            for r in reversed(rec):
                if r[2] != last_rec[2] or r[3] != last_rec[3] or r[4] != last_rec[4] or r[5] != last_rec[5]:
                    last_rec = r
                    rec_set.insert(0, r)  # add to the front of the list
            return rec_set
        else:
            return  # will return None - NoneType


class OTEquitSpreadsheet:
    """
    generate spreadsheet for overtime equitability.
    """
    def __init__(self):
        self.frame = None
        self.pb = None  # the progress bar object
        self.pbi = None  # progress bar counter index
        self.date = None
        self.station = None
        self.year = None
        self.month = None
        self.startdate = None
        self.enddate = None
        self.quarter = None
        self.full_quarter = None
        self.startdate_index = []
        self.enddate_index = []
        self.carrierlist = []
        self.recset = []
        self.minrows = 1
        self.otcalcpref = "off_route"  # preference for overtime calculation - "off_route" or "all"
        self.carrier_overview = []  # a list of carrier's name, status and makeups
        self.date_array = []  # a list of all days in the quarter as a datetimes
        self.assignment_check = []  # multidimensional array of carriers and days with no bid assignment
        self.front_padding = 0  # number of empty triad to fill worksheet prior to startdate
        self.end_padding = 0  # number of empty trids to fill worksheet prior to enddate
        self.end_pad_indicator = 0  # shows days after last day for triad builder
        self.ringrefset = []  # multidimensional array - daily rings/refusals for each otdl carrier
        self.dates_breakdown = []  # a list of dates for display on spreadsheets
        self.week = ["w01", "w02", "w03", "w04", "w05", "w06", "w07", "w08", "w09", "w10", "w11", "w12", "w13",
                     "w14", "w15"]
        self.week_label = ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12",
                           "13", "14", "15"]
        self.triad_week_index = 0
        self.triad_column = 0
        self.triad_row = 0
        self.footer_row = 0
        self.wb = None  # workbook
        self.overview = None  # first worksheet which summarizes all the following worksheets
        self.ws = None  # worksheet
        self.instructions = None  # last worksheet which provides instructions
        self.ws_header = None  # NamedStyles for spreadsheet
        self.date_dov = None
        self.date_dov_title = None
        self.col_header = None
        self.col_center_header = None
        self.col_header_instructions = None
        self.input_name = None
        self.input_s = None
        self.input_blue = None
        self.input_center = None
        self.calcs = None
        self.ws_name = None
        self.ref_ot = None
        self.instruct_text = None

    def create(self, frame, date, station):
        """ a master method for building the spreadsheet. """
        self.frame = frame
        if not self.ask_ok():  # abort if user selects cancel from askokcancel
            return
        self.pb = ProgressBarDe(label="Building OTDL Equitability Spreadsheet")
        self.pb.max_count(100)  # set length of progress bar
        self.pb.start_up()  # start the progress bar
        self.pbi = 1
        self.pb.move_count(self.pbi)  # increment progress bar
        self.pb.change_text("Gathering Data... ")
        self.station = station
        self.date = date  # a datetime object from the quarter is passed and used as date
        self.breakdown_date()  # the passed datetime object is broken down into year and month
        self.define_quarter()  # the year and month are used to generate quarter and full quarter
        self.get_dates()  # use quarter information to get start and end date
        self.get_carrierlist()  # generate a raw list of carriers at station before or on end date.
        self.get_recsets()  # filter the carrierlist to get only otdl carriers
        self.get_settings()  # get minimum rows and ot calculation preference
        self.get_carier_overview()  # build a list of carrier's name, status and makeups
        self.carrier_overview_add()  # adds empty sets so the lenght of carrier overview = minimum rows.
        self.get_date_array()  # get a list of all days in the quarter as datetime objects
        self.get_assignment_check()  # get a multidimensional array of carriers time spent off assignment.
        self.get_front_padding()  # get number of empty triads needed to pad prior to startdate
        self.get_end_padding()  # get number of empty triads needed to pad after enddate
        self.get_ringrefset()  # build multidimensional array - daily rings/refusals for each carrier
        self.get_date_breakdown()  # build an array of dates for display on the spreadsheets
        self.get_footer_row()  # get the row where the footer will be
        self.build_workbook()  # build the spreadsheet and define the worksheets.
        self.set_dimensions_overview()  # column widths for overview sheet
        self.set_dimensions_weekly()  # column widths for weekly worksheets
        self.set_dimensions_instructions()  # column widths for instructions sheet
        self.get_styles()  # define workbook styles
        self.build_header_overview()  # build the header for overview worksheet
        self.build_columnheader_overview()  # build column headers for overview worksheet
        self.build_main_overview()  # build main body for overview worksheet
        self.build_overview_footer()  # create the footer at the bottom of the worksheet for averages and totals
        self.build_header_weeklysheets()  # build the header for the weekly worksheet
        self.build_columnheader_worksheets()  # build column headers for weekly worksheet
        self.build_main_worksheets()  # build the main parts of the worksheet save the triad groups
        self.triads_delegator()  # orders which triads to build from build triads
        self.build_worksheet_footer()
        self.build_instructions()
        self.save_open()  # save and open the spreadsheet

    def ask_ok(self):
        """ continue the process if user presses ok. """
        if messagebox.askokcancel("Spreadsheet generator",
                                  "Do you want to generate an Over Time Equitability Spreadsheet?",
                                  parent=self.frame):
            return True
        return False

    def breakdown_date(self):
        """ breakdown the date into year and month """
        self.year = int(self.date.strftime("%Y"))
        self.month = int(self.date.strftime("%m"))

    def define_quarter(self):
        """ defines the quarter """
        self.quarter = Quarter(self.month).find()  # convert the month into a quarter - 1 through 4.
        self.full_quarter = str(self.year) + "-" + str(self.quarter)  # create a string expressing the year - quarter

    def get_dates(self):
        """ gets the start and end of the quarter. """
        self.startdate_index = (datetime(self.year, 1, 1), datetime(self.year, 4, 1), datetime(self.year, 7, 1),
                                datetime(self.year, 10, 1))
        self.enddate_index = (datetime(self.year, 3, 31), datetime(self.year, 6, 30), datetime(self.year, 9, 30),
                              datetime(self.year, 12, 31))
        self.startdate = self.startdate_index[int(self.quarter) - 1]
        self.enddate = self.enddate_index[int(self.quarter) - 1]

    def starting_day(self):
        """ returns the column position of the startdate as an odd number (5 to 17) """
        days = ("Saturday", "Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday")
        i = 6  #
        for day in days:  # loop through tuple of days"
            if self.startdate.strftime("%A") == day:  # if the startdate matches the day
                return i  # returns the column of the first date
            i -= 1  # count down from Saturday

    def get_carrierlist(self):
        """ get the carrier list. """
        self.carrierlist = CarrierList(self.startdate, self.enddate, self.station).get_distinct()

    def get_recsets(self):
        """ get the clock rings for the quarter. """
        for carrier in self.carrierlist:
            otlist = ("otdl", )  # list type for carriers wanted
            rec = QuarterRecs(carrier[0], self.startdate, self.enddate, self.station).get_filtered_recs(otlist)
            if rec:
                self.recset.append(rec)

    def get_settings(self):
        """ get minimum rows and ot calculation preference """
        sql = "SELECT tolerance FROM tolerances"
        result = inquire(sql)
        self.minrows = int(result[25][0])
        self.otcalcpref = result[26][0]

    def get_status(self, recs):
        """ returns true if the carrier's last record is otdl and the station is correct. """
        if recs[0][2] == "otdl" and recs[0][5] == self.station:
            return True
        return False

    def get_pref(self, recs):
        """ get the preferences from otdl preferences table. """
        carrier = recs[0][1]
        sql = "SELECT preference FROM otdl_preference WHERE carrier_name = '%s' and quarter = '%s' and station = '%s'" \
              % (carrier, self.full_quarter, self.station)
        pref = inquire(sql)
        if not pref:  # if there is no record in the database, create one
            sql = "INSERT INTO otdl_preference (quarter, carrier_name, preference, station, makeups) " \
                  "VALUES('%s', '%s', '%s', '%s', '%s')" \
                  % (self.full_quarter, carrier, "12", self.station, "")
            commit(sql)  # enter the new record in the dbae
            return "12"  # return 12 hour preference
        else:
            return pref[0][0]  # return the pulled from the database.

    @staticmethod
    def get_status_pref(status, pref):
        """ takes status and preference to get the status mode """
        if pref == "track" and not status:  # if not on otdl and prefence is track: status is track
            return "track"
        elif pref != "track" and not status:  # if not on the otdl and prefrence is not track: status is off
            return "off"
        else:  # if on the otdl
            return pref

    def get_makeups(self, carrier):
        """ get makeup records from the otdl preference table. """
        sql = "SELECT makeups FROM otdl_preference WHERE carrier_name = '%s' and quarter = '%s' and station = '%s'" \
              % (carrier, self.full_quarter, self.station)
        makeup = inquire(sql)
        if makeup:
            return makeup[0][0]
        else:
            return ""

    def get_carier_overview(self):
        """ build a list of carrier's name, status and makeups """
        self.pbi += 1  # increment progress bar counter
        self.pb.change_text("Gathering Carrier Data... ")  # update progress bar text
        for recs in self.recset:  # loop through the recsets
            carrier = recs[0][1]  # get the carrier name
            status = self.get_status_pref(self.get_status(recs), self.get_pref(recs))
            makeup = self.get_makeups(carrier)
            add_this = (carrier, status, makeup)
            self.carrier_overview.append(add_this)

    def carrier_overview_add(self):
        """ adds empty sets so the lenght of carrier overview = minimum rows. """
        while len(self.carrier_overview) < self.minrows:
            add_this = ("", "", "")
            self.carrier_overview.append(add_this)

    def get_date_array(self):
        """ get a list of all days in the quarter as datetime objects """
        running_date = self.startdate
        while running_date <= self.enddate:
            self.date_array.append(running_date)
            running_date += timedelta(days=1)

    def get_front_padding(self):
        """ get the number of empty triads to put before startdate to fill worksheet """
        self.front_padding = 6 - self.starting_day()

    def get_end_padding(self):
        """ get number of empty triads needed after enddate to fill worksheet """
        self.end_padding = 105 - (self.front_padding + len(self.date_array))
        self.end_pad_indicator = self.front_padding + len(self.date_array)

    def get_assignment_check(self):
        """ get a multidimensional array of carriers time with no bid assignment """
        for i in range(len(self.carrier_overview)):
            if not self.carrier_overview[i][0]:  # if the carrier is not empty set for minimum rows
                self.assignment_check.append([])  # add empty into the array
            elif not self.check_for_unassigned(i):  # if there is not a record with no assignment
                self.assignment_check.append([])  # add empty into the array
            else:  # if there is a carrier and they have been unassigned at some point in the quarter
                self.assignment_check.append([])  # add to the array
                self.get_unassigned(i)  # fill it with dates in which carrier was unassigned

    def check_for_unassigned(self, i):
        """ check for any records where otdl carrier is unassigned """
        for rec in self.recset[i]:
            if rec[4] == "" or rec[4] == "0000":
                return True
        return False

    def get_unassigned(self, i):
        """ get a list of datetimes for periods where a carrier is unassigned."""
        loop = 0
        dates = self.get_unassigned_dates(i)
        for revrec in reversed(self.recset[i]):
            if not revrec[4] or revrec[4] == "0000":  # if there is no assignment for the record
                date = max(Convert(revrec[0]).dt_converter(), self.startdate)  # handle RPRs, default to date in range
                if loop + 1 != len(self.recset[i]):  # if there is at least one more record in the set
                    while date < dates[loop + 1]:  # until the date matchs the next
                        self.assignment_check[i].append(date)
                        date += timedelta(days=1)
                if loop + 1 == len(self.recset[i]):  # if this is the last record in the set
                    while date != self.enddate + timedelta(days=1):
                        self.assignment_check[i].append(date)
                        date += timedelta(days=1)
            loop += 1

    def get_unassigned_dates(self, i):
        """ get reversed list of effective dates from the recset """
        dates = []
        for revrec in reversed(self.recset[i]):
            dates.append(Convert(revrec[0]).dt_converter())
        return dates

    def get_ringrefset(self):
        """ build multidimensional array - daily rings/refusals for each otdl carrier """
        self.pb.max_count(8 + (len(self.carrier_overview)*2))  # set length of progress bar
        for i in range(len(self.carrier_overview)):
            # update progress bar text
            self.pb.change_text("Gathering Carrier Rings: {}/{} ".format(i, len(self.carrier_overview)))
            self.pbi += 1  # increment progress bar counter
            self.pb.move_count(self.pbi)  # increment progress bar
            self.ringrefset.append([])  # each carrier has an array
            self.get_daily_ringrefs(i)

    def get_overtime(self, total, moves, code, has_route):
        """ find the overtime pending ot calculation preference and ns day code """
        if self.otcalcpref == "off_route" and has_route:
            return Overtime().proper_overtime(total, moves, code)
        else:  # default to straight overtime if the carrier has no route.
            return Overtime().straight_overtime(total, code)

    def get_daily_ringrefs(self, index):
        """ the ring ref - clock rings and refusals. gets the clock rings to determine the overtime then gets
        and stores the refusals time/type in the self.ringrefset variable. """
        daily_ringref = []
        carrier = self.carrier_overview[index][0]  # get the carrier name using carrier overview md array and index
        for _ in range(self.front_padding):  # insert front padding so empty cells fill worksheet
            add_this = ["", "", ""]
            daily_ringref.append(add_this)
        for date in self.date_array:  # get the ringrefs from the database or empty if none
            has_route = True  # notes that the carrier has a bid assignment
            if date in self.assignment_check[index]:  # if the date matchs a date where the carrier had no assignment
                has_route = False  # note that the carrier had no assignment
            overtime = ""
            sql = "SELECT total, code, moves FROM rings3 WHERE rings_date = '%s' AND carrier_name = '%s'" \
                  % (date, carrier)
            results = inquire(sql)
            if results:
                total = results[0][0]
                code = results[0][1]
                moves = Moves().timeoffroute(results[0][2])  # calculate the time off route
                overtime = self.get_overtime(total, moves, code, has_route)  # find the overtime
            ref_type = ""
            ref_time = ""
            sql = "SELECT refusal_type, refusal_time FROM refusals WHERE refusal_date = '%s' AND carrier_name = '%s'" \
                  % (date, carrier)
            ref_results = inquire(sql)
            if ref_results:
                ref_type = ref_results[0][0]
                ref_time = ref_results[0][1]
            add_this = [overtime, ref_type, ref_time]
            daily_ringref.append(add_this)
        for _ in range(self.end_padding):  # insert front padding so empty cells fill worksheet
            add_this = ["", "", ""]
            daily_ringref.append(add_this)
        self.ringrefset[index] = daily_ringref

    def get_date_breakdown(self):
        """ gets the dates and defines a start - end date string. """
        date = self.startdate
        days_til_end = self.starting_day()
        if not days_til_end:
            display_date = date.strftime("%a %m/%d/%Y")
        else:
            enddate = date + timedelta(days=days_til_end)
            display_date = date.strftime("%a %m/%d/%Y") + " through " + enddate.strftime("%a %m/%d/%Y")
        self.dates_breakdown.append(display_date)
        date += timedelta(days=days_til_end + 1)
        for _ in range(14):  # loop once for each week
            enddate = date + timedelta(days=6)
            display_date = date.strftime("%m/%d/%Y") + " through " + enddate.strftime("%m/%d/%Y")
            self.dates_breakdown.append(display_date)
            date += timedelta(weeks=1)

    def get_footer_row(self):
        """ get the number of the row where the footer will go. """
        self.footer_row = (len(self.carrier_overview) * 2) + 7

    def build_workbook(self):
        """ get the workbook object. """
        self.pbi += 1  # increment progress bar counter
        self.pb.move_count(self.pbi)  # increment progress bar
        self.pb.change_text("Building Workbook... ")  # update progress bar text
        self.ws = []
        self.wb = Workbook()  # define the workbook
        self.overview = self.wb.active  # create first worksheet
        self.overview.title = "overview"  # give the first worksheet a name
        for i in range(15):  # create worksheet for remaining 15 weeks
            self.ws.append(self.wb.create_sheet(self.week[i]))  # create subsequent worksheets
            self.ws[i].title = self.week[i]  # title subsequent worksheets
        self.instructions = self.wb.create_sheet("instructions")

    def set_dimensions_overview(self):
        """ sets the widths of the columns for the overview tab. """
        self.overview.column_dimensions["A"].width = 6
        self.overview.column_dimensions["B"].width = 12
        self.overview.column_dimensions["C"].width = 6
        self.overview.column_dimensions["D"].width = 7
        self.overview.column_dimensions["E"].width = 3
        self.overview.column_dimensions["F"].width = 11
        self.overview.column_dimensions["G"].width = 12
        self.overview.column_dimensions["H"].width = 12

    def set_dimensions_weekly(self):
        """ sets the width of the columns for weekly tabs. """
        for i in range(15):
            self.ws[i].column_dimensions["A"].width = 6
            self.ws[i].column_dimensions["B"].width = 14
            self.ws[i].column_dimensions["C"].width = 5
            self.ws[i].column_dimensions["D"].width = 3
            self.ws[i].column_dimensions["E"].width = 2
            self.ws[i].column_dimensions["F"].width = 4
            self.ws[i].column_dimensions["G"].width = 2
            self.ws[i].column_dimensions["H"].width = 4
            self.ws[i].column_dimensions["I"].width = 2
            self.ws[i].column_dimensions["J"].width = 4
            self.ws[i].column_dimensions["K"].width = 2
            self.ws[i].column_dimensions["L"].width = 4
            self.ws[i].column_dimensions["M"].width = 2
            self.ws[i].column_dimensions["N"].width = 4
            self.ws[i].column_dimensions["O"].width = 2
            self.ws[i].column_dimensions["P"].width = 4
            self.ws[i].column_dimensions["Q"].width = 2
            self.ws[i].column_dimensions["R"].width = 4
            self.ws[i].column_dimensions["S"].width = 7

    def set_dimensions_instructions(self):
        """ sets the column widths for the instructions tab. """
        self.instructions.column_dimensions["A"].width = 14
        self.instructions.column_dimensions["B"].width = 14
        self.instructions.column_dimensions["C"].width = 14
        self.instructions.column_dimensions["D"].width = 14
        self.instructions.column_dimensions["E"].width = 14
        self.instructions.column_dimensions["F"].width = 14

    def get_styles(self):
        """ Named styles for workbook """
        bd = Side(style='thin', color="80808080")  # defines borders
        self.ws_header = NamedStyle(name="ws_header", font=Font(bold=True, name='Arial', size=12))
        self.date_dov = NamedStyle(name="date_dov", font=Font(name='Arial', size=8))
        self.date_dov_title = NamedStyle(name="date_dov_title", font=Font(bold=True, name='Arial', size=8),
                                         alignment=Alignment(horizontal='right'))
        self.col_header = NamedStyle(name="col_header", font=Font(bold=True, name='Arial', size=8))
        self.col_center_header = NamedStyle(name="col_center_header", font=Font(bold=True, name='Arial', size=8),
                                            alignment=Alignment(horizontal='center'))
        self.col_header_instructions = \
            NamedStyle(name="col_header_instructions", font=Font(bold=True, name='Arial', size=10))
        self.input_name = NamedStyle(name="input_name", font=Font(name='Arial', size=8),
                                     border=Border(left=bd, right=bd, top=bd, bottom=bd),
                                     alignment=Alignment(horizontal='left'))
        self.input_s = NamedStyle(name="input_s", font=Font(name='Arial', size=8),
                                  border=Border(left=bd, right=bd, top=bd, bottom=bd),
                                  alignment=Alignment(horizontal='right'))
        self.input_blue = NamedStyle(name="input_blue", font=Font(name='Arial', size=8),
                                     border=Border(left=bd, right=bd, top=bd, bottom=bd),
                                     fill=PatternFill(fgColor='e1f7f3', fill_type='solid'),
                                     alignment=Alignment(horizontal='right'))
        """fill_type: Value must be one of {'darkTrellis', 'darkGrid', 'lightVertical', 'darkDown', 'solid', 'lightUp', 
        'lightHorizontal', 'mediumGray', 'lightTrellis', 'darkHorizontal', 'darkGray', 'lightGray', 'darkVertical', 
        'gray125', 'darkUp', 'gray0625', 'lightDown', 'lightGrid'"""
        self.input_center = NamedStyle(name="input_center", font=Font(name='Arial', size=8),
                                       border=Border(left=bd, right=bd, top=bd, bottom=bd),
                                       alignment=Alignment(horizontal='center'))
        self.calcs = NamedStyle(name="calcs", font=Font(name='Arial', size=8),
                                border=Border(left=bd, right=bd, top=bd, bottom=bd),
                                fill=PatternFill(fgColor='e5e4e2', fill_type='solid'),
                                alignment=Alignment(horizontal='right'))
        self.ws_name = NamedStyle(name="ws_name", font=Font(name='Arial', size=8),
                                  border=Border(left=bd, right=bd, top=bd, bottom=bd),
                                  fill=PatternFill(fgColor='e5e4e2', fill_type='solid'),
                                  alignment=Alignment(horizontal='left'))
        self.ref_ot = NamedStyle(name="ref_ot", font=Font(name='Arial', size=8),
                                 border=Border(left=bd, right=bd, top=bd, bottom=bd),
                                 fill=PatternFill(fgColor='e5e4e2', fill_type='solid'),
                                 alignment=Alignment(horizontal='center'))
        self.instruct_text = NamedStyle(name="instruct_text", font=Font(name='Arial', size=10),
                                        alignment=Alignment(horizontal='left', vertical='top'))

    def build_header_overview(self):
        """ build the header for overview worksheet """
        cell = self.overview.cell(row=1, column=1)  # page title
        cell.value = "OTDL Equitability Worksheet"
        cell.style = self.ws_header
        self.overview.merge_cells('A1:E1')
        cell = self.overview.cell(row=2, column=1)  # date
        cell.value = "dates: "
        cell.style = self.date_dov_title
        cell = self.overview.cell(row=2, column=2)  # fill in dates
        date = self.startdate_index[self.quarter - 1].strftime("%m/%d/%Y") + " through " + \
            self.enddate_index[self.quarter - 1].strftime("%m/%d/%Y")
        cell.value = date
        cell.style = self.date_dov
        self.overview.merge_cells('B2:E2')
        cell = self.overview.cell(row=3, column=1)  # ot type label
        cell.value = "ot type: "
        cell.style = self.date_dov_title
        cell = self.overview.cell(row=3, column=2)  # fill in ot type
        cell.value = self.otcalcpref
        cell.style = self.date_dov
        cell = self.overview.cell(row=2, column=6)  # station
        cell.value = "station: "
        cell.style = self.date_dov_title
        cell = self.overview.cell(row=2, column=7)  # fill in station
        cell.value = self.station
        cell.style = self.date_dov
        self.overview.merge_cells('G2:H2')
        cell = self.overview.cell(row=3, column=6)  # number of carriers
        cell.value = "# of carriers active on otdl: "
        cell.style = self.date_dov_title
        self.overview.merge_cells('F3:G3')
        cell = self.overview.cell(row=3, column=8)  # fill in number of carriers
        lastnum = ((len(self.carrier_overview)*2)+5)
        formula = "=COUNTIF(%s!C%s:C%s, %s)+COUNTIF(%s!C%s:C%s, %s)+COUNTIF(%s!C%s:C%s, \"track\")" \
                  % ("overview", str(6), str(lastnum), str(12),
                     "overview", str(6), str(lastnum), str(10),
                     "overview", str(6), str(lastnum))
        cell.value = formula
        cell.style = self.calcs

    def build_columnheader_overview(self):
        """ build column header for the overview tab. """
        cell = self.overview.cell(row=5, column=1)  # name
        cell.value = "name"
        cell.style = self.col_header
        self.overview.merge_cells('A5:B5')
        cell = self.overview.cell(row=5, column=3)  # status
        cell.value = "status"
        cell.style = self.col_center_header
        cell = self.overview.cell(row=5, column=4)  # make up
        cell.value = "make up"
        cell.style = self.col_center_header
        cell = self.overview.cell(row=5, column=5)  # refusals/overtime
        cell.value = "refusals/overtime"
        cell.style = self.col_center_header
        self.overview.merge_cells('E5:F5')
        cell = self.overview.cell(row=5, column=7)  # opportunities
        cell.value = "opportunities"
        cell.style = self.col_center_header
        cell = self.overview.cell(row=5, column=8)  # diff from avg
        cell.value = "diff from avg"
        cell.style = self.col_center_header

    def build_main_overview(self):
        """ build the body of the overview tab. """
        self.pbi += 1  # increment progress bar counter
        self.pb.move_count(self.pbi)  # increment progress bar
        self.pb.change_text("Building Overview... ")  # update progress bar text
        row = 6
        for i in range(len(self.carrier_overview)):
            self.overview.row_dimensions[row].height = 13
            self.overview.row_dimensions[row+1].height = 13
            cell = self.overview.cell(row=row, column=1)  # name
            cell.value = self.carrier_overview[i][0]
            cell.style = self.input_name
            self.overview.merge_cells('A' + str(row) + ':' + 'B' + str(row+1))
            cell = self.overview.cell(row=row, column=3)  # status
            cell.value = Handler(self.carrier_overview[i][1]).str_to_int_or_str()
            cell.style = self.input_center
            self.overview.merge_cells('C' + str(row) + ':' + 'C' + str(row + 1))
            cell = self.overview.cell(row=row, column=4)  # make up
            cell.value = Handler(self.carrier_overview[i][2]).str_to_float_or_str()
            cell.style = self.input_center
            cell.number_format = "#,###.00;[RED]-#,###.00"
            self.overview.merge_cells('D' + str(row) + ':' + 'D' + str(row + 1))
            cell = self.overview.cell(row=row, column=5)  # refusals label
            cell.value = "ref"
            cell.style = self.ref_ot
            cell = self.overview.cell(row=row+1, column=5)  # OT label
            cell.value = "ot"
            cell.style = self.ref_ot
            cell = self.overview.cell(row=row, column=6)  # refusals
            formula = "=IF(OR(%s!C%s=12, %s!C%s=10, %s!C%s=\"track\")," \
                      "SUM(%s!S%s, %s!S%s, %s!S%s, %s!S%s, %s!S%s, " \
                      "%s!S%s, %s!S%s, %s!S%s, %s!S%s, %s!S%s, " \
                      "%s!S%s, %s!S%s, %s!S%s, %s!S%s, %s!S%s),0)" \
                      % ("overview", str(row), "overview", str(row), "overview", str(row),
                         self.week[0], str(row), self.week[1], str(row), self.week[2], str(row),
                         self.week[3], str(row), self.week[4], str(row), self.week[5], str(row),
                         self.week[6], str(row), self.week[7], str(row), self.week[8], str(row),
                         self.week[9], str(row), self.week[10], str(row), self.week[11], str(row),
                         self.week[12], str(row), self.week[13], str(row), self.week[14], str(row))
            cell.value = formula
            cell.style = self.calcs
            cell.number_format = "#,###.00;[RED]-#,###.00"
            cell = self.overview.cell(row=row + 1, column=6)  # overtime
            formula = "=IF(OR(%s!C%s=12, %s!C%s=10, %s!C%s=\"track\")," \
                      "SUM(%s!S%s, %s!S%s, %s!S%s, %s!S%s, %s!S%s, " \
                      "%s!S%s, %s!S%s, %s!S%s, %s!S%s, %s!S%s, " \
                      "%s!S%s, %s!S%s, %s!S%s, %s!S%s, %s!S%s)," \
                      "0)" \
                      % ("overview", str(row), "overview", str(row), "overview", str(row),
                         self.week[0], str(row+1), self.week[1], str(row+1), self.week[2], str(row+1),
                         self.week[3], str(row+1), self.week[4], str(row+1), self.week[5], str(row+1),
                         self.week[6], str(row+1), self.week[7], str(row+1), self.week[8], str(row+1),
                         self.week[9], str(row+1), self.week[10], str(row+1), self.week[11], str(row+1),
                         self.week[12], str(row+1), self.week[13], str(row+1), self.week[14], str(row+1))
            cell.value = formula
            cell.style = self.calcs
            cell.number_format = "#,###.00;[RED]-#,###.00"
            cell = self.overview.cell(row=row, column=7)  # opportunities
            formula = "=IF(OR(%s!C%s=12, %s!C%s=10, %s!C%s=\"track\"),(%s!F%s+%s!F%s)-%s!D%s,0)" % \
                      ("overview", str(row), "overview", str(row), "overview", str(row), "overview", str(row),
                       "overview", str(row+1), "overview", str(row))
            cell.value = formula
            cell.style = self.calcs
            cell.number_format = "#,###.00;[RED]-#,###.00"
            self.overview.merge_cells('G' + str(row) + ":" + 'G' + str(row+1))
            cell = self.overview.cell(row=row, column=8)  # difference from average
            formula = "=IF(%s!A%s=\"\",0, IF(OR(%s!C%s=12,%s!C%s=10, %s!C%s=\"track\"),%s!G%s-%s!$G$%s,\"off list\"))" \
                      % ("overview", str(row), "overview", str(row), "overview", str(row), "overview", str(row),
                         "overview", str(row), "overview", str(self.footer_row+2))  # last row is avg
            cell.value = formula
            cell.style = self.calcs
            cell.number_format = "#,###.00;[RED]-#,###.00"
            self.overview.merge_cells('H' + str(row) + ":" + 'H' + str(row + 1))
            row += 2

    def get_totalovertime_formula(self, sheet, row, column):
        """ gives formulas for totals counting skipping rows. """
        """
        the row argument is the starting row of the count,
        the column is given as a number and matched to a letter with the dictionary
        """
        column_dict = {5: "E", 6: "F", 7: "G", 9: "I", 11: "K", 13: "M", 15: "O", 17: "Q", 19: "S"}
        string = "=SUM("
        while row < self.footer_row-2:
            string += "{}!{}{},".format(sheet, column_dict[column], row)
            row += 2
        string += "{}!{}{})".format(sheet, column_dict[column], row)
        return string

    def build_overview_footer(self):
        """ create the footer at the bottom of the worksheet for averages and totals """
        cell = self.overview.cell(row=self.footer_row, column=5)  # label total overtime
        cell.value = "total overtime:"
        cell.style = self.date_dov_title
        cell = self.overview.cell(row=self.footer_row, column=6)  # calculate total overtime
        formula = self.get_totalovertime_formula("overview", 7, 6)
        cell.value = formula
        cell.style = self.calcs
        cell.number_format = "#,###.00;[RED]-#,###.00"
        cell = self.overview.cell(row=self.footer_row, column=7)  # calculate total opportunities
        formula = "=SUM(%s!G%s:G%s)" \
                  % ("overview", str(6), str(self.footer_row - 2))
        cell.value = formula
        cell.style = self.calcs
        cell.number_format = "#,###.00;[RED]-#,###.00"
        cell = self.overview.cell(row=self.footer_row, column=8)  # label total opportunities
        cell.value = "  :total opportunities"
        cell.style = self.col_header
        cell = self.overview.cell(row=self.footer_row+2, column=5)  # label average overtime
        cell.value = "average overtime:"
        cell.style = self.date_dov_title
        cell = self.overview.cell(row=self.footer_row+2, column=6)  # calculate average overtime
        formula = "=%s!F%s/%s!$H$3" % ("overview", self.footer_row, "overview")
        cell.value = formula
        cell.style = self.calcs
        cell.number_format = "#,###.00;[RED]-#,###.00"
        cell = self.overview.cell(row=self.footer_row+2, column=7)  # calculate average opportunities
        formula = "=%s!G%s/%s!$H$3" % ("overview", str(self.footer_row), "overview")
        cell.value = formula
        cell.style = self.calcs
        cell.number_format = "#,###.00;[RED]-#,###.00"
        cell = self.overview.cell(row=self.footer_row+2, column=8)  # label average opportunities
        cell.value = "  :average opportunities"
        cell.style = self.col_header

    def build_header_weeklysheets(self):
        """ build the header for the weekly worksheet """
        self.pbi += 1  # increment progress bar counter
        self.pb.move_count(self.pbi)  # increment progress bar
        self.pb.change_text("Building Weekly Worksheets - Headers ")  # update progress bar text
        for i in range(15):
            cell = self.ws[i].cell(row=1, column=1)  # page title
            cell.value = "OTDL Equitability Worksheet"
            cell.style = self.ws_header
            self.ws[i].merge_cells('A1:H1')
            cell = self.ws[i].cell(row=1, column=16)  # week
            cell.value = "week: "
            cell.style = self.date_dov_title
            self.ws[i].merge_cells('P1:R1')
            cell = self.ws[i].cell(row=1, column=19)  # fill in week
            cell.value = self.week_label[i]
            cell.style = self.date_dov
            cell = self.ws[i].cell(row=2, column=1)  # date
            cell.value = "dates: "
            cell.style = self.date_dov_title
            cell = self.ws[i].cell(row=2, column=2)  # fill in date
            cell.value = self.dates_breakdown[i]
            cell.style = self.date_dov
            self.ws[i].merge_cells('B2:J2')
            cell = self.ws[i].cell(row=3, column=1)  # ot type label
            cell.value = "ot type: "
            cell.style = self.date_dov_title
            cell = self.ws[i].cell(row=3, column=2)  # fill in ot type
            cell.value = self.otcalcpref
            cell.style = self.date_dov
            cell = self.ws[i].cell(row=2, column=11)  # station
            cell.value = "station: "
            cell.style = self.date_dov_title
            self.ws[i].merge_cells('K2:M2')
            cell = self.ws[i].cell(row=2, column=14)  # fill in station
            cell.value = self.station
            cell.style = self.date_dov
            self.ws[i].merge_cells('N2:S2')
            cell = self.ws[i].cell(row=3, column=10)  # number of carriers
            cell.value = "# of carriers active on otdl: "
            cell.style = self.date_dov_title
            self.ws[i].merge_cells('J3:R3')
            cell = self.ws[i].cell(row=3, column=19)  # calculate number of carriers
            formula = "=%s!%s%s" % ("overview", "H", "3")
            cell.value = formula
            cell.style = self.calcs

    def build_columnheader_worksheets(self):
        """ build the column headers for the weekly worksheets. """
        for i in range(15):
            cell = self.ws[i].cell(row=5, column=1)  # name
            cell.value = "name"
            cell.style = self.col_header
            self.ws[i].merge_cells('A5:B5')
            cell = self.ws[i].cell(row=5, column=3)  # status
            cell.value = "status"
            cell.style = self.col_header
            self.ws[i].merge_cells('C5:D5')
            cell = self.ws[i].cell(row=5, column=5)  # sat
            cell.value = "sat"
            cell.style = self.col_center_header
            self.ws[i].merge_cells('E5:F5')
            cell = self.ws[i].cell(row=5, column=7)  # sun
            cell.value = "sun"
            cell.style = self.col_center_header
            self.ws[i].merge_cells('G5:H5')
            cell = self.ws[i].cell(row=5, column=9)  # mon
            cell.value = "mon"
            cell.style = self.col_center_header
            self.ws[i].merge_cells('I5:J5')
            cell = self.ws[i].cell(row=5, column=11)  # tue
            cell.value = "tue"
            cell.style = self.col_center_header
            self.ws[i].merge_cells('K5:L5')
            cell = self.ws[i].cell(row=5, column=13)  # wed
            cell.value = "wed"
            cell.style = self.col_center_header
            self.ws[i].merge_cells('M5:N5')
            cell = self.ws[i].cell(row=5, column=15)  # thu
            cell.value = "thu"
            cell.style = self.col_center_header
            self.ws[i].merge_cells('O5:P5')
            cell = self.ws[i].cell(row=5, column=17)  # fri
            cell.value = "fri"
            cell.style = self.col_center_header
            self.ws[i].merge_cells('Q5:R5')
            cell = self.ws[i].cell(row=5, column=19)  # weekly
            cell.value = "weekly"
            cell.style = self.col_center_header

    def build_main_worksheets(self):
        """ build the main port of the weekly worksheets. """
        self.pbi += 1  # increment progress bar counter
        self.pb.move_count(self.pbi)  # increment progress bar
        self.pb.change_text("Building Weekly Worksheets - Body ")  # update progress bar text
        for i in range(len(self.ws)):  # for each worksheet
            row = 6
            for _ in range(len(self.carrier_overview)):  # for each carrier or empty set (get reach minimum rows)
                cell = self.ws[i].cell(row=row, column=1)  # name
                formula = "=IF(%s!A%s = \"\",\"\",%s!A%s)" % ("overview", str(row), "overview", str(row))
                cell.value = formula
                cell.style = self.ws_name
                self.ws[i].merge_cells('A' + str(row) + ":" + "B" + str(row+1))
                cell = self.ws[i].cell(row=row, column=3)  # status
                formula = "=IF(%s!C%s = \"\",\"\",%s!C%s)" % ("overview", str(row), "overview", str(row))
                cell.value = formula
                cell.style = self.ref_ot
                self.ws[i].merge_cells('C' + str(row) + ":" + "C" + str(row+1))
                cell = self.ws[i].cell(row=row, column=4)  # ref label
                cell.value = "ref"
                cell.style = self.ref_ot
                cell = self.ws[i].cell(row=row+1, column=4)  # ot label
                cell.value = "ot"
                cell.style = self.ref_ot
                cell = self.ws[i].cell(row=row, column=19)  # ref weekly total
                formula = "=SUM(%s!%s%s,%s!%s%s,%s!%s%s,%s!%s%s,%s!%s%s,%s!%s%s,%s!%s%s)" \
                          % (self.week[i], "F", str(row), self.week[i], "H", str(row), self.week[i], "J", str(row),
                             self.week[i], "L", str(row), self.week[i], "N", str(row), self.week[i], "P", str(row),
                             self.week[i], "R", str(row))
                cell.value = formula
                cell.style = self.calcs
                cell.number_format = "#,###.00;[RED]-#,###.00"
                cell = self.ws[i].cell(row=row + 1, column=19)  # overtime weekly total
                formula = "=SUM(%s!%s%s:%s%s)" % (self.week[i], "E", str(row+1), "Q", str(row+1))
                cell.value = formula
                cell.style = self.calcs
                cell.number_format = "#,###.00;[RED]-#,###.00"
                row += 2

    @staticmethod
    def get_triad_merge(row, column):
        """ returns a string for for merge in triad group """
        column -= 5
        letter = ("E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R")
        text = letter[column] + str(row+1) + ":" + letter[column+1] + str(row+1)
        return text

    def get_triad_refset(self, carrier, date):
        """ get the data to fill the triad. """
        ringrefset = ("", "", "")  # default is empty set
        if self.carrier_overview[carrier][0]:  # if the carrier overview is not an empty set
            ringrefset = self.ringrefset[carrier][date]  # get the ring ref set for that carrier and date
        return ringrefset

    def triads_delegator(self):
        """ build and fill moves triads. """
        self.triad_row = 6
        for c in range(len(self.carrier_overview)):  # for each carrier including empty sets for minimum rows
            # update progress bar text
            self.pb.change_text("Building Weekly Worksheets - Triad Groups: row {}/{}"
                                .format(c, len(self.carrier_overview)))
            self.pbi += 1  # increment progress bar counter
            self.pb.move_count(self.pbi)  # increment progress bar
            self.triad_week_index = 0  # week starts at zero
            self.triad_column = 5  # column starts at first day due to front padding
            for i in range(105):  # loop for self.front_padding + len(self.date_array) + self.end_padding
                self.build_triads(c, i)
            self.triad_row += 2

    def triad_style(self, i):
        """ fill the triads with blue if they are outside the quarter. """
        if i < self.front_padding:
            return self.input_blue
        if i >= self.end_pad_indicator:
            return self.input_blue
        return self.input_s

    def build_triads(self, c, i):
        """ builds the individual triads. """
        ringrefset = self.get_triad_refset(c, i)
        # refusal indicator field
        cell = self.ws[self.triad_week_index].cell(row=self.triad_row, column=self.triad_column)
        cell.value = ringrefset[1]
        cell.style = self.triad_style(i)
        # refusal time field
        cell = self.ws[self.triad_week_index].cell(row=self.triad_row, column=self.triad_column + 1)
        cell.value = Handler(ringrefset[2]).str_to_float_or_str()
        cell.style = self.triad_style(i)
        cell.number_format = "#,###.00;[RED]-#,###.00"
        # overtime field
        cell = self.ws[self.triad_week_index].cell(row=self.triad_row + 1, column=self.triad_column)
        cell.value = Handler(ringrefset[0]).str_to_float_or_str()
        cell.style = self.triad_style(i)
        cell.number_format = "#,###.00;[RED]-#,###.00"
        # merge overtime field
        self.ws[self.triad_week_index].merge_cells(self.get_triad_merge(self.triad_row, self.triad_column))
        self.triad_column += 2
        if self.triad_column >= 19:
            self.triad_column = 5
            self.triad_week_index += 1

    def build_worksheet_footer(self):
        """ build the footer """
        column_dict = {5: "E", 6: "F", 7: "G", 9: "I", 11: "K", 13: "M", 15: "O", 17: "Q", 19: "S"}
        for i in range(len(self.ws)):  # for each worksheet
            cell = self.ws[i].cell(row=self.footer_row, column=4)  # total overtime label
            cell.value = "total overtime:  "
            cell.style = self.date_dov_title
            cell = self.ws[i].cell(row=self.footer_row+2, column=4)  # average overtime label
            cell.value = "average overtime:  "
            cell.style = self.date_dov_title
            column_array = (5, 7, 9, 11, 13, 15, 17)
            for col in column_array:  # loop though for each column
                cell = self.ws[i].cell(row=self.footer_row, column=col)
                formula = self.get_totalovertime_formula(self.week[i], 7, col)
                cell.value = formula
                cell.style = self.calcs
                cell.number_format = "#,###.00;[RED]-#,###.00"
                merge_string = self.get_triad_merge(self.footer_row-1, col)  # subtact 1 from row to match function
                self.ws[i].merge_cells(merge_string)
                cell = self.ws[i].cell(row=self.footer_row+2, column=col)
                formula = "=%s!%s%s/%s!$S$3" % (self.week[i], column_dict[col], self.footer_row, self.week[i])
                cell.value = formula
                cell.style = self.calcs
                cell.number_format = "#,###.00;[RED]-#,###.00"
                merge_string = self.get_triad_merge(self.footer_row + 1, col)  # subtact 1 from row to match function
                self.ws[i].merge_cells(merge_string)
            cell = self.ws[i].cell(row=self.footer_row, column=19)
            formula = self.get_totalovertime_formula(self.week[i], 7, 19)
            cell.value = formula
            cell.style = self.calcs
            cell.number_format = "#,###.00;[RED]-#,###.00"
            cell = self.ws[i].cell(row=self.footer_row + 2, column=19)
            formula = "=%s!%s%s/%s!$S$3" % (self.week[i], column_dict[19], self.footer_row, self.week[i])
            cell.value = formula
            cell.style = self.calcs
            cell.number_format = "#,###.00;[RED]-#,###.00"

    def build_instructions(self):
        """ build the instructions tab. """
        self.pbi += 1  # increment progress bar counter
        self.pb.move_count(self.pbi)  # increment progress bar
        self.pb.change_text("Building Instructions... ")  # update progress bar text
        self.instructions.merge_cells('A1:F1')  # merge cells for page title
        self.instructions.merge_cells('A5:F47')  # merge cells for instructions
        cell = self.instructions.cell(row=1, column=1)  # page title
        cell.value = "OTDL Equitability Worksheet"
        cell.style = self.ws_header
        cell = self.instructions.cell(row=3, column=1)  # sub header
        cell.value = "Instructions"
        cell.style = self.col_header_instructions
        cell = self.instructions.cell(row=5, column=1)
        text = "CAUTION: Do not write in grayed out cells. These cells have formulas. Writing in " \
               "them will delete those formulas. If this happens, do a CTRL Z to undo.\n\n" \
               "1. NAME:  Enter the carrier names on the first page only. Formulas on other pages " \
               "will import the name so that you don’t have to write it 15 times.\n\n" \
               "2. STATUS:  Enter the status on the first page only. Again formulas will do the work " \
               "and copy it to other pages. Enter “12” (for 12 hour preference), “10” (10 hour " \
               "preference), “off” (if the carrier has gotten off the list) or “track” (if the carrier " \
               "is off the list, but you want to continue tracking their equitability). Leave the field " \
               "blank if there is no carrier. If a carrier name is in the “name” column and the “status” " \
               "field is blank, the default status is “off”. If the status is “off”, the carrier's rings " \
               "and refusals will not be calculated and the carrier will not be figured into the average." \
               "\n\n" \
               "3.  MAKE UP:  This applies to make up opportunities from grievance settlements. " \
               "See JCAM Article 8.5.C.2 Remedies. This value is given in hours and clicks e.g. " \
               "5.32 or 14.00.\n\n" \
               "4. REFUSALS/OVERTIME: This displays refusals and overtime worked.\n\n" \
               "5. OPPORTUNITIES:   This displays total opportunities for overtime. This is all " \
               "refusals + overtime.\n\n" \
               "6. DIFF FROM AVERAGE:  This cell uses formulas to calculate the average " \
               "overtime of all carriers and the individual carrier’s difference from that. If they have " \
               "more than average, the number will be positive otherwise it will be negative. This will " \
               "be the core of your case in your grievance for OTDL equitability violations. " \
               "There are 15 worksheets. Each on represents a service week. Start with the first " \
               "week and proceed day by day.\n\n" \
               "7. For each day and each carrier there are are groups of three cells.\n\n" \
               "TOP LEFT (smaller) CELL:  This is a one letter explanation for any refusal: You can " \
               "use your own system, but my suggestions are: “p” for preference (the carrier " \
               "refused on the grounds that they are on the 10 hour list), “a” for annual " \
               "(the carrier missed overtime opportunities due to being on annual leave), “s” for sick " \
               "leave, for non scheduled day (the carrier worked on their nonscheduled day – see JCAM " \
               "Article 8.5.C.2.d Not Counted Toward “Equitability), “x” for exceptional " \
               "circumstances (see JCAM Article 8.5.E - Exceptional Situations May Excuse " \
               "Mandatory Overtime) and “r” for refusal (the carrier tells the supervisor to go jump " \
               "in a lake).\n\n" \
               "TOP RIGHT CELL: This is the amount of overtime that the carrier refused or was " \
               "unable to work.\n\n" \
               "BOTTOM CELL: This is overtime worked. Normally these is overtime worked off " \
               "the carrier’s own route, but it could mean any overtime depending on any local " \
               "agreements with management (LMOUs).\n\n" \
               "At the very bottom , there are totals and averages for the day. These are for your " \
               "information.\n\n\n\n\n\n\n"
        cell.value = text
        cell.style = self.instruct_text
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    def save_open(self):
        """ name the excel file """
        self.pbi += 1  # increment progress bar counter
        self.pb.move_count(self.pbi)  # increment progress bar
        self.pb.change_text("Saving Workbook... ")  # update progress bar text
        quarter = self.full_quarter.replace(" ", "")
        xl_filename = "ot_equit_" + quarter + ".xlsx"
        try:
            self.wb.save(dir_path('ot_equitability') + xl_filename)
            messagebox.showinfo("Spreadsheet generator",
                                "Your spreadsheet was successfully generated. \n"
                                "File is named: {}".format(xl_filename),
                                parent=self.frame)
            if sys.platform == "win32":
                os.startfile(dir_path('ot_equitability') + xl_filename)
            if sys.platform == "linux":
                subprocess.call(["xdg-open", 'kb_sub/ot_equitability/' + xl_filename])
            if sys.platform == "darwin":
                subprocess.call(["open", dir_path('ot_equitability') + xl_filename])
        except PermissionError:
            messagebox.showerror("Spreadsheet generator",
                                 "The spreadsheet was not opened. \n"
                                 "Suggestion: "
                                 "Make sure that identically named spreadsheets are closed "
                                 "(the file can't be overwritten while open).",
                                 parent=self.frame)
        self.pb.stop()


class OTDistriSpreadsheet:
    """
    Generates a spreadsheet that shows the distribution overtime by carriers on otdl, wal, nl, aux list.
    """
    def __init__(self):
        self.frame = None
        self.pb = None  # the progress bar object
        self.pbi = None  # progress bar counter index
        self.date = None
        self.station = None
        self.rangeopt = None  # "weekly" or "quarterly"
        self.listoptions = None
        self.year = None
        self.month = None
        self.startdate = None
        self.enddate = None
        self.quarter = None
        self.full_quarter = None
        self.startdate_index = []
        self.enddate_index = []
        self.carrierlist = []
        self.recset = []
        self.minrows = 1
        self.otcalcpref = "off_route"  # preference for overtime calculation - "off_route" or "all"
        self.carrier_overview = []  # a list of carrier's name, status and makeups
        self.date_array = []  # a list of all days in the quarter as a datetimes
        self.assignment_check = []  # multidimensional array of carriers and days with no bid assignment
        self.front_padding = 0
        self.end_padding = 0
        self.end_pad_indicator = 7  # shows days after last day for triad builder
        self.ringsset = []
        self.dates_breakdown = []  # a list of dates for display on spreadsheets
        self.sheetcount = 15  # there are 15 weekly sheets in a quarterly investigation
        self.daycount = 105  # there are 105 days in a quarterly investigation
        self.week = ["w01", "w02", "w03", "w04", "w05", "w06", "w07", "w08", "w09", "w10", "w11", "w12", "w13",
                     "w14", "w15"]
        self.week_label = ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12",
                           "13", "14", "15"]
        self.footer_row = 0
        self.wb = None  # workbook
        self.overview = None  # first worksheet which summarizes all the following worksheets
        self.ws = None  # worksheet
        self.instructions = None  # last worksheet which provides instructions
        self.ws_header = None  # NamedStyles for spreadsheet
        self.date_dov = None
        self.date_dov_title = None
        self.col_header = None
        self.col_center_header = None
        self.col_header_instructions = None
        self.input_name = None
        self.input_s = None
        self.input_blue = None
        self.input_center = None
        self.calcs = None
        self.ws_name = None
        self.ref_ot = None
        self.instruct_text = None
        self.weekly_row = None
        self.weekly_week_index = None
        self.weekly_column = None

    def create(self, frame, date, station, rangeopt, listoptions):
        """ a master method for running methods in proper order. """
        self.frame = frame
        if not self.ask_ok():  # abort if user selects cancel from askokcancel
            return
        self.pb = ProgressBarDe(label="Building OT Distribution Spreadsheet")
        self.pb.max_count(100)  # set length of progress bar
        self.pb.start_up()  # start the progress bar
        self.pbi = 1
        self.pb.move_count(self.pbi)  # increment progress bar
        self.pb.change_text("Gathering Data... ")
        self.station = station
        self.rangeopt = rangeopt
        if self.rangeopt == "weekly":  # unless the investigation range is weekly
            self.sheetcount = 1  # then there is only one week.
        if self.rangeopt == "weekly":   # if the investigation range is weekly
            self.daycount = 7  # then there are 7 days
        self.listoptions = listoptions
        self.date = date  # a datetime object from the quarter is passed and used as date
        self.breakdown_date()
        self.define_quarter()
        self.get_dates()
        self.starting_day()
        self.get_carrierlist()
        self.get_recsets()  # filter the carrierlist to get only carriers with selected list statuses
        self.get_settings()  # get minimum rows and ot calculation preference
        self.get_carier_overview()  # build a list of carrier's name, status and makeups
        self.carrier_overview_add()  # adds empty sets so the lenght of carrier overview = minimum rows.
        self.get_date_array()  # get a list of all days in the quarter as datetime objects
        self.get_assignment_check()  # get a multidimensional array of carriers time spent off assignment.
        if rangeopt == "quarterly":  # front and end padding is only needed for quarterly investigations
            self.get_front_padding()
            self.get_end_padding()
        self.get_ringsset()
        self.get_date_breakdown()  # build an array of dates for display on the spreadsheets
        self.get_footer_row()  # get the row where the footer will be
        self.build_workbook()  # build the spreadsheet and define the worksheets.
        self.set_dimensions_overview()  # column widths for overview sheet
        self.set_dimensions_weekly()  # column widths for weekly worksheets
        self.set_dimensions_instructions()  # column widths for instructions sheet
        self.get_styles()  # define workbook styles
        self.build_header_overview()  # build the header for overview worksheet
        self.build_columnheader_overview()  # build column headers for overview worksheet
        self.build_main_overview()  # build main body for overview worksheet
        self.build_overview_footer()  # create the footer at the bottom of the worksheet for averages and totals
        self.build_header_weeklysheets()  # build the header for the weekly worksheet
        self.build_columnheader_worksheets()  # build column headers for weekly worksheet
        self.build_main_worksheets()  # build the main parts of the worksheet save the triad groups
        self.daily_delegator()  # orders which cells to build for weekly worksheet
        self.build_worksheet_footer()
        self.build_instructions()
        self.save_open()  # save and open the spreadsheet

    def ask_ok(self):
        """ ends the process if the user selects cancel. """
        if messagebox.askokcancel("Spreadsheet generator",
                                  "Do you want to generate an Overtime Distribution Spreadsheet?",
                                  parent=self.frame):
            return True
        return False

    def breakdown_date(self):
        """ breakdown the date into year and month """
        self.year = int(self.date.strftime("%Y"))
        self.month = int(self.date.strftime("%m"))

    def define_quarter(self):
        """ express the quarter as a number 1 to 4. """
        self.quarter = Quarter(self.month).find()  # convert the month into a quarter - 1 through 4.
        self.full_quarter = str(self.year) + "-" + str(self.quarter)  # create a string expressing the year - quarter

    def get_dates(self):
        """ get the dates of the quarter. """
        self.startdate_index = (datetime(self.year, 1, 1), datetime(self.year, 4, 1), datetime(self.year, 7, 1),
                                datetime(self.year, 10, 1))
        self.enddate_index = (datetime(self.year, 3, 31), datetime(self.year, 6, 30), datetime(self.year, 9, 30),
                              datetime(self.year, 12, 31))
        self.startdate = self.startdate_index[int(self.quarter) - 1]  # if investigation range is quarterly
        self.enddate = self.enddate_index[int(self.quarter) - 1]
        if self.rangeopt == "weekly":  # if investigation range is weekly
            self.startdate = projvar.invran_date_week[0]
            self.enddate = projvar.invran_date_week[6]

    def starting_day(self):
        """ returns the column position of the startdate as an odd number (5 to 17) """
        days = ("Saturday", "Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday")
        i = 6  #
        for day in days:  # loop through tuple of days"
            if self.startdate.strftime("%A") == day:  # if the startdate matches the day
                return i  # returns the column of the first date
            i -= 1  # count down from Saturday

    def get_carrierlist(self):
        """ get a distinct list of carrier names for the investigation range """
        self.carrierlist = CarrierList(self.startdate, self.enddate, self.station).get_distinct()

    def get_recsets(self):
        """ get the ring records for the carrier list. """
        for carrier in self.carrierlist:
            rec = QuarterRecs(carrier[0], self.startdate, self.enddate, self.station)\
                .get_filtered_recs(self.listoptions)
            if rec:
                self.recset.append(rec)

    def get_settings(self):
        """ get minimum rows and ot calculation preference """
        sql = "SELECT tolerance FROM tolerances"
        result = inquire(sql)
        self.minrows = int(result[27][0])
        self.otcalcpref = result[28][0]

    @staticmethod
    def get_status(recs):
        """ return the carrier's list status """
        add_plus = False
        status = recs[0][2]
        for i in range(len(recs)):  # loop through all recs in recset
            if recs[i][2] != status:  # check for any list statuses that differ from the first
                add_plus = True
        if add_plus:  # add a "+" if there is more than one list status
            status += " +"
        return status

    def get_carier_overview(self):
        """ build a list of carrier's name, status and makeups """
        self.pbi += 1  # increment progress bar counter
        self.pb.change_text("Gathering Carrier Data... ")  # update progress bar text
        for recs in self.recset:  # loop through the recsets
            carrier = recs[0][1]  # get the carrier name
            status = self.get_status(recs)
            add_this = (carrier, status)
            self.carrier_overview.append(add_this)

    def carrier_overview_add(self):
        """ adds empty sets so the lenght of carrier overview = minimum rows. """
        while len(self.carrier_overview) < self.minrows:
            add_this = ("", "")
            self.carrier_overview.append(add_this)

    def get_date_array(self):
        """ get a list of all days in the quarter as datetime objects """
        running_date = self.startdate
        while running_date <= self.enddate:
            self.date_array.append(running_date)
            running_date += timedelta(days=1)

    def get_front_padding(self):
        """ get the number of empty triads to put before startdate to fill worksheet """
        self.front_padding = 6 - self.starting_day()

    def get_end_padding(self):
        """ get number of empty triads needed after enddate to fill worksheet """
        self.end_padding = 105 - (self.front_padding + len(self.date_array))
        self.end_pad_indicator = self.front_padding + len(self.date_array)

    def get_assignment_check(self):
        """ get a multidimensional array of carriers time with no bid assignment """
        for i in range(len(self.carrier_overview)):
            if not self.carrier_overview[i][0]:  # if the carrier is not empty set for minimum rows
                self.assignment_check.append([])  # add empty into the array
            elif not self.check_for_unassigned(i):  # if there is not a record with no assignment
                self.assignment_check.append([])  # add empty into the array
            else:  # if there is a carrier and they have been unassigned at some point in the quarter
                self.assignment_check.append([])  # add to the array
                self.get_unassigned(i)  # fill it with dates in which carrier was unassigned

    def check_for_unassigned(self, i):
        """ check for any records where otdl carrier is unassigned """
        for rec in self.recset[i]:
            if rec[4] == "" or rec[4] == "0000":
                return True
        return False

    def get_unassigned(self, i):
        """ get a list of datetimes for periods where a carrier is unassigned. """
        loop = 0
        dates = self.get_unassigned_dates(i)
        for revrec in reversed(self.recset[i]):
            if not revrec[4] or revrec[4] == "0000":  # if there is no assignment for the record
                date = max(Convert(revrec[0]).dt_converter(), self.startdate)  # handle RPRs, default to date in range
                if loop + 1 != len(self.recset[i]):  # if there is at least one more record in the set
                    while date < dates[loop + 1]:  # until the date matches the next
                        self.assignment_check[i].append(date)
                        date += timedelta(days=1)
                if loop + 1 == len(self.recset[i]):  # if this is the last record in the set
                    while date != self.enddate + timedelta(days=1):
                        self.assignment_check[i].append(date)
                        date += timedelta(days=1)
            loop += 1

    def get_unassigned_dates(self, i):
        """ get reversed list of effective dates from the recset """
        dates = []
        for revrec in reversed(self.recset[i]):
            dates.append(Convert(revrec[0]).dt_converter())
        return dates

    def get_ringsset(self):
        """ build multidimensional array - daily rings/refusals for each otdl carrier """
        self.pb.max_count(8 + (len(self.carrier_overview)*2))  # set length of progress bar
        for i in range(len(self.carrier_overview)):
            # update progress bar text
            self.pb.change_text("Gathering Carrier Rings: {}/{} ".format(i, len(self.carrier_overview)))
            self.pbi += 1  # increment progress bar counter
            self.pb.move_count(self.pbi)  # increment progress bar
            self.ringsset.append([])  # each carrier has an array
            self.get_daily_rings(i)

    def get_overtime(self, total, moves, code, has_route):
        """ find the overtime pending ot calculation preference and ns day code """
        if self.otcalcpref == "off_route" and has_route:
            return Overtime().proper_overtime(total, moves, code)
        else:  # default to straight overtime if the carrier has no route.
            return Overtime().straight_overtime(total, code)

    def get_daily_rings(self, index):
        """ get the clock rings for the day. """
        daily_ring = []
        carrier = self.carrier_overview[index][0]  # get the carrier name using carrier overview md array and index
        for _ in range(self.front_padding):  # insert front padding so empty cells fill worksheet
            add_this = ""
            daily_ring.append(add_this)
        for date in self.date_array:  # get the ringrefs from the database or empty if none
            has_route = True  # notes that the carrier has a bid assignment
            if date in self.assignment_check[index]:  # if the date matchs a date where the carrier had no assignment
                has_route = False  # note that the carrier had no assignment
            overtime = ""
            sql = "SELECT total, code, moves FROM rings3 WHERE rings_date = '%s' AND carrier_name = '%s'" \
                  % (date, carrier)
            results = inquire(sql)
            if results:
                total = results[0][0]
                code = results[0][1]
                moves = Moves().timeoffroute(results[0][2])  # calculate the time off route
                overtime = self.get_overtime(total, moves, code, has_route)  # find the overtime
            daily_ring.append(overtime)
        for _ in range(self.end_padding):  # insert front padding so empty cells fill worksheet
            add_this = ""
            daily_ring.append(add_this)
        self.ringsset[index] = daily_ring

    def get_date_breakdown(self):
        """ gets a string showing a range of dates. """
        date = self.startdate
        days_til_end = self.starting_day()
        if not days_til_end:
            display_date = date.strftime("%a %m/%d/%Y")
        else:
            enddate = date + timedelta(days=days_til_end)
            display_date = date.strftime("%a %m/%d/%Y") + " through " + enddate.strftime("%a %m/%d/%Y")
        self.dates_breakdown.append(display_date)
        date += timedelta(days=days_til_end + 1)
        for _ in range(14):  # loop once for each week
            enddate = date + timedelta(days=6)
            display_date = date.strftime("%m/%d/%Y") + " through " + enddate.strftime("%m/%d/%Y")
            self.dates_breakdown.append(display_date)
            date += timedelta(weeks=1)

    def get_footer_row(self):
        """ get the number of the row where the footer will go. """
        self.footer_row = len(self.carrier_overview) + 7

    def build_workbook(self):
        """ build the workbook object. """
        self.pbi += 1  # increment progress bar counter
        self.pb.move_count(self.pbi)  # increment progress bar
        self.pb.change_text("Building Workbook... ")  # update progress bar text
        self.ws = []
        self.wb = Workbook()  # define the workbook
        self.overview = self.wb.active  # create first worksheet
        self.overview.title = "overview"  # give the first worksheet a name
        for i in range(self.sheetcount):  # create worksheet for remaining weeks (15 or 1)
            self.ws.append(self.wb.create_sheet(self.week[i]))  # create subsequent worksheets
            self.ws[i].title = self.week[i]  # title subsequent worksheets
        self.instructions = self.wb.create_sheet("instructions")

    def set_dimensions_overview(self):
        """ get the width of the columns for the overview. """
        self.overview.column_dimensions["A"].width = 6
        self.overview.column_dimensions["B"].width = 14
        self.overview.column_dimensions["C"].width = 6
        self.overview.column_dimensions["D"].width = 7
        self.overview.column_dimensions["E"].width = 11
        self.overview.column_dimensions["F"].width = 11
        self.overview.column_dimensions["G"].width = 12
        self.overview.column_dimensions["H"].width = 12

    def set_dimensions_weekly(self):
        """ get the width of the column for the weekly tabs. """
        for i in range(self.sheetcount):
            self.ws[i].column_dimensions["A"].width = 6
            self.ws[i].column_dimensions["B"].width = 14
            self.ws[i].column_dimensions["C"].width = 6
            self.ws[i].column_dimensions["D"].width = 5
            self.ws[i].column_dimensions["E"].width = 5
            self.ws[i].column_dimensions["F"].width = 5
            self.ws[i].column_dimensions["G"].width = 5
            self.ws[i].column_dimensions["H"].width = 5
            self.ws[i].column_dimensions["I"].width = 5
            self.ws[i].column_dimensions["J"].width = 5
            self.ws[i].column_dimensions["K"].width = 7
            self.ws[i].column_dimensions["L"].width = 4
            self.ws[i].column_dimensions["M"].width = 2
            self.ws[i].column_dimensions["N"].width = 4
            self.ws[i].column_dimensions["O"].width = 2
            self.ws[i].column_dimensions["P"].width = 4
            self.ws[i].column_dimensions["Q"].width = 2
            self.ws[i].column_dimensions["R"].width = 4
            self.ws[i].column_dimensions["S"].width = 7

    def set_dimensions_instructions(self):
        """ get the width of the columns for the instructions page."""
        self.instructions.column_dimensions["A"].width = 14
        self.instructions.column_dimensions["B"].width = 14
        self.instructions.column_dimensions["C"].width = 14
        self.instructions.column_dimensions["D"].width = 14
        self.instructions.column_dimensions["E"].width = 14
        self.instructions.column_dimensions["F"].width = 14

    def get_styles(self):
        """ Named styles for workbook """
        bd = Side(style='thin', color="80808080")  # defines borders
        self.ws_header = NamedStyle(name="ws_header", font=Font(bold=True, name='Arial', size=12))
        self.date_dov = NamedStyle(name="date_dov", font=Font(name='Arial', size=8))
        self.date_dov_title = NamedStyle(name="date_dov_title", font=Font(bold=True, name='Arial', size=8),
                                         alignment=Alignment(horizontal='right'))
        self.col_header = NamedStyle(name="col_header", font=Font(bold=True, name='Arial', size=8))
        self.col_center_header = NamedStyle(name="col_center_header", font=Font(bold=True, name='Arial', size=8),
                                            alignment=Alignment(horizontal='center'))
        self.col_header_instructions = \
            NamedStyle(name="col_header_instructions", font=Font(bold=True, name='Arial', size=10))
        self.input_name = NamedStyle(name="input_name", font=Font(name='Arial', size=8),
                                     border=Border(left=bd, right=bd, top=bd, bottom=bd),
                                     alignment=Alignment(horizontal='left'))
        self.input_s = NamedStyle(name="input_s", font=Font(name='Arial', size=8),
                                  border=Border(left=bd, right=bd, top=bd, bottom=bd),
                                  alignment=Alignment(horizontal='right'))
        self.input_blue = NamedStyle(name="input_blue", font=Font(name='Arial', size=8),
                                     border=Border(left=bd, right=bd, top=bd, bottom=bd),
                                     fill=PatternFill(fgColor='e1f7f3', fill_type='solid'),
                                     alignment=Alignment(horizontal='right'))
        """fill_type: Value must be one of {'darkTrellis', 'darkGrid', 'lightVertical', 'darkDown', 'solid', 'lightUp', 
        'lightHorizontal', 'mediumGray', 'lightTrellis', 'darkHorizontal', 'darkGray', 'lightGray', 'darkVertical', 
        'gray125', 'darkUp', 'gray0625', 'lightDown', 'lightGrid'"""
        self.input_center = NamedStyle(name="input_center", font=Font(name='Arial', size=8),
                                       border=Border(left=bd, right=bd, top=bd, bottom=bd),
                                       alignment=Alignment(horizontal='center'))
        self.calcs = NamedStyle(name="calcs", font=Font(name='Arial', size=8),
                                border=Border(left=bd, right=bd, top=bd, bottom=bd),
                                fill=PatternFill(fgColor='e5e4e2', fill_type='solid'),
                                alignment=Alignment(horizontal='right'))
        self.ws_name = NamedStyle(name="ws_name", font=Font(name='Arial', size=8),
                                  border=Border(left=bd, right=bd, top=bd, bottom=bd),
                                  fill=PatternFill(fgColor='e5e4e2', fill_type='solid'),
                                  alignment=Alignment(horizontal='left'))
        self.ref_ot = NamedStyle(name="ref_ot", font=Font(name='Arial', size=8),
                                 border=Border(left=bd, right=bd, top=bd, bottom=bd),
                                 fill=PatternFill(fgColor='e5e4e2', fill_type='solid'),
                                 alignment=Alignment(horizontal='center'))
        self.instruct_text = NamedStyle(name="instruct_text", font=Font(name='Arial', size=10),
                                        alignment=Alignment(horizontal='left'))

    def build_header_overview(self):
        """ build the header for overview worksheet """
        cell = self.overview.cell(row=1, column=1)  # page title
        cell.value = "Overtime Distribution Worksheet"
        cell.style = self.ws_header
        self.overview.merge_cells('A1:E1')
        cell = self.overview.cell(row=2, column=1)  # date
        cell.value = "dates: "
        cell.style = self.date_dov_title
        cell = self.overview.cell(row=2, column=2)  # fill in dates
        date = self.startdate.strftime("%m/%d/%Y") + " through " + \
            self.enddate.strftime("%m/%d/%Y")
        cell.value = date
        cell.style = self.date_dov
        self.overview.merge_cells('B2:D2')
        cell = self.overview.cell(row=3, column=1)  # ot type label
        cell.value = "ot type: "
        cell.style = self.date_dov_title
        cell = self.overview.cell(row=3, column=2)  # fill in ot type
        date = self.otcalcpref
        cell.value = date
        cell.style = self.date_dov
        cell = self.overview.cell(row=2, column=5)  # station
        cell.value = "station: "
        cell.style = self.date_dov_title
        cell = self.overview.cell(row=2, column=6)  # fill in station
        cell.value = self.station
        cell.style = self.date_dov
        self.overview.merge_cells('F2:H2')
        cell = self.overview.cell(row=3, column=3)  # number of carriers
        cell.value = "# of carriers without restrictions: "
        cell.style = self.date_dov_title
        self.overview.merge_cells('C3:E3')
        cell = self.overview.cell(row=3, column=6)  # fill in number of carriers
        lastnum = (len(self.carrier_overview) + 6)
        formula = "=COUNTA(%s!A%s:A%s)-COUNTA(%s!D%s:D%s)" \
                  % ("overview", str(6), str(lastnum),
                     "overview", str(6), str(lastnum))
        cell.value = formula
        cell.style = self.calcs

    def build_columnheader_overview(self):
        """ build the column headers for the overview page. """
        cell = self.overview.cell(row=5, column=1)  # name
        cell.value = "name"
        cell.style = self.col_header
        self.overview.merge_cells('A5:B5')
        cell = self.overview.cell(row=5, column=3)  # list status
        cell.value = "list"
        cell.style = self.col_center_header
        cell = self.overview.cell(row=5, column=4)  # medical
        cell.value = "medical"
        cell.style = self.col_center_header
        cell = self.overview.cell(row=5, column=5)  # overtime
        cell.value = "overtime"
        cell.style = self.col_center_header
        cell = self.overview.cell(row=5, column=6)  # diff from avg
        cell.value = "diff from avg"
        cell.style = self.col_center_header

    def build_main_overview(self):
        """ build the overview page. """
        self.pbi += 1  # increment progress bar counter
        self.pb.move_count(self.pbi)  # increment progress bar
        self.pb.change_text("Building Overview... ")  # update progress bar text
        row = 6
        for i in range(len(self.carrier_overview)):
            self.overview.row_dimensions[row].height = 13
            self.overview.row_dimensions[row+1].height = 13
            cell = self.overview.cell(row=row, column=1)  # name
            cell.value = self.carrier_overview[i][0]
            cell.style = self.input_name
            self.overview.merge_cells('A' + str(row) + ':' + 'B' + str(row))
            cell = self.overview.cell(row=row, column=3)  # list status
            cell.value = Handler(self.carrier_overview[i][1]).str_to_int_or_str()
            cell.style = self.input_center
            cell = self.overview.cell(row=row, column=4)  # medical
            cell.value = ""
            cell.style = self.input_center
            cell = self.overview.cell(row=row, column=5)  # overtime
            if self.rangeopt == "quarterly":
                formula = "=IF(%s!A%s=\"\",\"\"," \
                          "IF(%s!D%s=\"\"," \
                          "SUM(%s!K%s, %s!K%s, %s!K%s, %s!K%s, %s!K%s, " \
                          "%s!K%s, %s!K%s, %s!K%s, %s!K%s, %s!K%s, " \
                          "%s!K%s, %s!K%s, %s!K%s, %s!K%s, %s!K%s,0)))" \
                          % ("overview", str(row), "overview", str(row),
                             self.week[0], str(row), self.week[1], str(row), self.week[2], str(row),
                             self.week[3], str(row), self.week[4], str(row), self.week[5], str(row),
                             self.week[6], str(row), self.week[7], str(row), self.week[8], str(row),
                             self.week[9], str(row), self.week[10], str(row), self.week[11], str(row),
                             self.week[12], str(row), self.week[13], str(row), self.week[14], str(row))
            else:
                formula = "=IF(%s!A%s=\"\",\"\",IF(%s!D%s=\"\",%s!K%s,0))" \
                          % ("overview", str(row), "overview", str(row), self.week[0], str(row))
            cell.value = formula
            cell.style = self.calcs
            cell.number_format = "#,###.00;[RED]-#,###.00"
            cell = self.overview.cell(row=row, column=6)  # difference from average
            formula = "=IF(%s!A%s=\"\",\"\", IF(%s!D%s=\"\",%s!E%s-%s!$E$%s,\"restrictions\"))" \
                      % ("overview", str(row), "overview", str(row),
                         "overview", str(row), "overview", str(self.footer_row+2))  # last row is avg
            cell.value = formula
            cell.style = self.calcs
            cell.number_format = "#,###.00;[RED]-#,###.00"
            row += 1

    def build_overview_footer(self):
        """ create the footer at the bottom of the worksheet for averages and totals """
        cell = self.overview.cell(row=self.footer_row, column=4)  # label total overtime
        cell.value = "total overtime:"
        cell.style = self.date_dov_title
        cell = self.overview.cell(row=self.footer_row, column=5)  # calculate total overtime
        formula = "=SUM(%s!E6:E%s)" % ("overview", self.footer_row - 2)
        cell.value = formula
        cell.style = self.calcs
        cell.number_format = "#,###.00;[RED]-#,###.00"
        cell = self.overview.cell(row=self.footer_row + 2, column=4)  # label average overtime
        cell.value = "average overtime:"
        cell.style = self.date_dov_title
        cell = self.overview.cell(row=self.footer_row + 2, column=5)  # calculate average overtime
        formula = "=%s!E%s/%s!$F$3" % ("overview", self.footer_row, "overview")
        cell.value = formula
        cell.style = self.calcs
        cell.number_format = "#,###.00;[RED]-#,###.00"

    def build_header_weeklysheets(self):
        """ build the header for the weekly worksheet """
        self.pbi += 1  # increment progress bar counter
        self.pb.move_count(self.pbi)  # increment progress bar
        self.pb.change_text("Building Weekly Worksheets - Headers ")  # update progress bar text
        for i in range(self.sheetcount):
            cell = self.ws[i].cell(row=1, column=1)  # page title
            cell.value = "Overtime Distribution Worksheet"
            cell.style = self.ws_header
            self.ws[i].merge_cells('A1:G1')
            if self.rangeopt == "quarterly":
                cell = self.ws[i].cell(row=1, column=8)  # week (label)
                cell.value = "week: "
                cell.style = self.date_dov_title
                self.ws[i].merge_cells('H1:I1')
                cell = self.ws[i].cell(row=1, column=10)  # fill in week
                cell.value = self.week_label[i]
                cell.style = self.date_dov
            cell = self.ws[i].cell(row=2, column=1)  # date (label)
            cell.value = "dates: "
            cell.style = self.date_dov_title
            cell = self.ws[i].cell(row=2, column=2)  # fill in date
            cell.value = self.dates_breakdown[i]
            cell.style = self.date_dov
            self.ws[i].merge_cells('B2:E2')
            cell = self.ws[i].cell(row=2, column=8)  # station (label)
            cell.value = "station: "
            cell.style = self.date_dov_title
            self.ws[i].merge_cells('H2:I2')
            cell = self.ws[i].cell(row=2, column=10)  # fill in station
            cell.value = self.station
            cell.style = self.date_dov
            self.ws[i].merge_cells('J2:M2')
            cell = self.ws[i].cell(row=3, column=1)  # ot type label
            cell.value = "ot type: "
            cell.style = self.date_dov_title
            cell = self.ws[i].cell(row=3, column=2)  # fill in ot type
            date = self.otcalcpref
            cell.value = date
            cell.style = self.date_dov
            cell = self.ws[i].cell(row=3, column=5)  # number of carriers
            cell.value = "# of carriers without restrictions: "
            cell.style = self.date_dov_title
            self.ws[i].merge_cells('E3:I3')
            cell = self.ws[i].cell(row=3, column=10)  # calculate number of carriers
            formula = "=%s!%s%s" % ("overview", "F", "3")
            cell.value = formula
            cell.style = self.calcs
            self.ws[i].merge_cells('J3:K3')

    def build_columnheader_worksheets(self):
        """ build the column headers for the weekly worksheets. """
        for i in range(self.sheetcount):
            cell = self.ws[i].cell(row=5, column=1)  # name
            cell.value = "name"
            cell.style = self.col_header
            self.ws[i].merge_cells('A5:B5')
            cell = self.ws[i].cell(row=5, column=3)  # status
            cell.value = "list"
            cell.style = self.col_header
            # self.ws[i].merge_cells('C5:D5')
            cell = self.ws[i].cell(row=5, column=4)  # sat
            cell.value = "sat"
            cell.style = self.col_center_header
            cell = self.ws[i].cell(row=5, column=5)  # sun
            cell.value = "sun"
            cell.style = self.col_center_header
            cell = self.ws[i].cell(row=5, column=6)  # mon
            cell.value = "mon"
            cell.style = self.col_center_header
            cell = self.ws[i].cell(row=5, column=7)  # tue
            cell.value = "tue"
            cell.style = self.col_center_header
            cell = self.ws[i].cell(row=5, column=8)  # wed
            cell.value = "wed"
            cell.style = self.col_center_header
            cell = self.ws[i].cell(row=5, column=9)  # thu
            cell.value = "thu"
            cell.style = self.col_center_header
            cell = self.ws[i].cell(row=5, column=10)  # fri
            cell.value = "fri"
            cell.style = self.col_center_header
            cell = self.ws[i].cell(row=5, column=11)  # weekly
            cell.value = "weekly"
            cell.style = self.col_center_header

    def build_main_worksheets(self):
        """ build the main part of the worksheets. """
        self.pbi += 1  # increment progress bar counter
        self.pb.move_count(self.pbi)  # increment progress bar
        self.pb.change_text("Building Weekly Worksheets - Body ")  # update progress bar text
        for i in range(len(self.ws)):  # for each worksheet
            row = 6
            for _ in range(len(self.carrier_overview)):  # for each carrier or empty set (get reach minimum rows)
                cell = self.ws[i].cell(row=row, column=1)  # name
                formula = "=IF(%s!A%s = \"\",\"\",%s!A%s)" % ("overview", str(row), "overview", str(row))
                cell.value = formula
                cell.style = self.ws_name
                self.ws[i].merge_cells('A' + str(row) + ":" + "B" + str(row))
                cell = self.ws[i].cell(row=row, column=3)  # status
                formula = "=IF(%s!C%s = \"\",\"\",%s!C%s)" % ("overview", str(row), "overview", str(row))
                cell.value = formula
                cell.style = self.ref_ot
                self.ws[i].merge_cells('C' + str(row) + ":" + "C" + str(row))
                cell = self.ws[i].cell(row=row, column=11)  # overtime weekly total
                formula = "=SUM(%s!%s%s:%s%s)" % (self.week[i], "D", str(row), "J", str(row))
                cell.value = formula
                cell.style = self.calcs
                cell.number_format = "#,###.00;[RED]-#,###.00"
                row += 1

    def daily_delegator(self):
        """ orders which cells to build for weekly worksheet """
        self.weekly_row = 6
        for c in range(len(self.carrier_overview)):  # for each carrier including empty sets for minimum rows
            # update progress bar text
            self.pb.change_text("Building Weekly Worksheets - Daily Cells: row {}/{}"
                                .format(c, len(self.carrier_overview)))
            self.pbi += 1  # increment progress bar counter
            self.pb.move_count(self.pbi)  # increment progress bar
            self.weekly_week_index = 0  # week starts at zero
            self.weekly_column = 4  # column starts at first day due to front padding
            for i in range(self.daycount):  # loop for self.front_padding + len(self.date_array) + self.end_padding
                self.build_daily(c, i)
            self.weekly_row += 1

    def daily_style(self, i):
        """ shows the cells in blue if they fall outside the quarter. """
        if i < self.front_padding:
            return self.input_blue
        if i >= self.end_pad_indicator:
            return self.input_blue
        return self.input_s

    def build_daily(self, c, i):
        """ build and fill the cell for the day. """
        ringsset = self.get_daily_refset(c, i)
        # overtime field
        cell = self.ws[self.weekly_week_index].cell(row=self.weekly_row, column=self.weekly_column)
        cell.value = Handler(ringsset).str_to_float_or_str()
        cell.style = self.daily_style(i)
        cell.number_format = "#,###.00;[RED]-#,###.00"
        self.weekly_column += 1
        if self.weekly_column >= 11:
            self.weekly_column = 4
            self.weekly_week_index += 1

    def get_daily_refset(self, carrier, date):
        """ get the daily data. """
        ringsset = ""  # default is empty string
        if self.carrier_overview[carrier][0]:  # if the carrier overview is not an empty set
            ringsset = self.ringsset[carrier][date]  # get the ring ref set for that carrier and date
        return ringsset

    def build_worksheet_footer(self):
        """ build the footer for the worksheet. """
        column_dict = {4: "D", 5: "E", 6: "F", 7: "G", 8: "H", 9: "I", 10: "J", 11: "K", 12: "L", 13: "M", 15: "O",
                       17: "Q", 19: "S"}
        for i in range(len(self.ws)):  # for each worksheet
            cell = self.ws[i].cell(row=self.footer_row, column=3)  # total overtime label
            cell.value = "total overtime:  "
            cell.style = self.date_dov_title
            cell = self.ws[i].cell(row=self.footer_row+2, column=3)  # average overtime label
            cell.value = "average overtime:  "
            cell.style = self.date_dov_title
            column_array = (4, 5, 6, 7, 8, 9, 10)
            for col in column_array:  # loop though for each column
                cell = self.ws[i].cell(row=self.footer_row, column=col)
                formula = "=SUM(%s!%s%s:%s%s)" % \
                          (self.week[i], column_dict[col], 6, column_dict[col], self.footer_row - 2)
                cell.value = formula
                cell.style = self.calcs
                cell.number_format = "#,###.00;[RED]-#,###.00"
                cell = self.ws[i].cell(row=self.footer_row+2, column=col)
                formula = "=%s!%s%s/%s!$J$3" % (self.week[i], column_dict[col], self.footer_row, self.week[i])
                cell.value = formula
                cell.style = self.calcs
                cell.number_format = "#,###.00;[RED]-#,###.00"
            cell = self.ws[i].cell(row=self.footer_row, column=11)
            formula = "=SUM(%s!K%s:K%s)" % (self.week[i], 6, self.footer_row-2)
            cell.value = formula
            cell.style = self.calcs
            cell.number_format = "#,###.00;[RED]-#,###.00"
            cell = self.ws[i].cell(row=self.footer_row + 2, column=11)
            formula = "=%s!K%s/%s!$J$3" % (self.week[i], self.footer_row, self.week[i])
            cell.value = formula
            cell.style = self.calcs
            cell.number_format = "#,###.00;[RED]-#,###.00"

    def build_instructions(self):
        """ build the instructions page. """
        self.pbi += 1  # increment progress bar counter
        self.pb.move_count(self.pbi)  # increment progress bar
        self.pb.change_text("Building Instructions... ")  # update progress bar text
        self.instructions.merge_cells('A1:F1')  # merge cells for title
        self.instructions.merge_cells('A5:F47')  # merge cells for instructions
        cell = self.instructions.cell(row=1, column=1)  # page title
        cell.value = "Overtime Distribution Worksheet"
        cell.style = self.ws_header
        cell = self.instructions.cell(row=3, column=1)  # sub header
        cell.value = "Instructions"
        cell.style = self.col_header_instructions
        cell = self.instructions.cell(row=5, column=1)
        text = "CAUTION: Do not write in grayed out cells. These cells have formulas. Writing in " \
               "them will delete those formulas. If this happens, do a CTRL Z to undo.\n\n" \
               "1. NAME:  Enter the carrier names on the first page only. Formulas on other pages " \
               "will import the name so that you don’t have to repeat it.\n\n" \
               "2. LIST:  Enter the status on the first page only. Again formulas will do the work " \
               "and copy it to other pages. Options are \"otdl\", \"wal\", \"nl\", \"aux\" or \"ptf\". " \
               "A \"+\" indicates that the carrier was on additional list during the investigation period. " \
               "Leave the field blank if there is no carrier. This field does not affect any formulas " \
               "and is for the user\'s information only. " \
               "\n\n" \
               "3. MEDICAL: This field shows if the carrier has medical restrictions. If the carrier has " \
               "medical restrictions, enter \"yes\" into the cell, although any word or character will work. " \
               "Although the column in titled \"medical\" you can use it for any situation where you want the " \
               "carrier\'s times to be removed from the totals and averages. This column does not pull data from " \
               "Klusterbox and must be manually entered in on the spreadsheet as Klusterbox does not track " \
               "medical restrictions. " \
               "Entering anything into this cell will cause formulas to change the average number of carriers, " \
               "zero out the overtime cell and will put \"restrictions\" into the \"diff from avg\" cell. \n\n" \
               "4. OVERTIME: This displays refusals and overtime worked. Information is pulled from the following " \
               "weekly worksheets. \n\n" \
               "5. DIFF FROM AVERAGE:  This cell uses formulas to calculate the average " \
               "overtime for individual carriers. It pulls the total average overtime from the bottom of the " \
               "page and calculates the difference for each individual carrier. If the carrier has " \
               "more than average, the number will be positive otherwise it will be negative. \n\n" \
               "6. WEEKLY WORKSHEETS: There are 15 worksheets for quarterly investigations and one worksheet " \
               "for weekly investigations. Each worksheet represents a service week. Start with the first week " \
               "and proceed day by day. If generated by Klusterbox, carrier clock rings will be pulled from " \
               "the database to calculate the carrier\'s overtime.\n\n" \
               "7. INPUT CELLS IN WEEKLY WORKSHEETS: Check the top of the page for \"ot type\", this will either " \
               "be \"off_route\" (the default), or \"all\" when the spreadsheet is generated by Klusterbox. " \
               "For each day and each carrier there are is one cell. This cell shows the amount of overtime " \
               "worked. If the overtime calculations preference is \"off route\" then only overtime worked off the " \
               "carrier\'s route will be shown, if the preference is \"all\" then all of the carrier\'s overtime " \
               "both off and on their own route will be shown. If the carrier has no route, the calculation will " \
               "default to \"all\" during the time the carrier is unassigned.\n\n" \
               "At the very bottom , there are totals and averages for the day. These are for your " \
               "information.\n"
        cell.value = text
        cell.style = self.instruct_text
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    def save_open(self):
        """ name the excel file """
        self.pbi += 1  # increment progress bar counter
        self.pb.move_count(self.pbi)  # increment progress bar
        self.pb.change_text("Saving Workbook... ")  # update progress bar text
        quarter = self.full_quarter.replace(" ", "")
        rang = "_q"
        if self.rangeopt == "weekly":
            rang = "_w"
        xl_filename = "ot_dist_" + quarter + rang + ".xlsx"
        try:
            self.wb.save(dir_path('ot_distribution') + xl_filename)
            messagebox.showinfo("Spreadsheet generator",
                                "Your spreadsheet was successfully generated. \n"
                                "File is named: {}".format(xl_filename),
                                parent=self.frame)
            if sys.platform == "win32":
                os.startfile(dir_path('ot_distribution') + xl_filename)
            if sys.platform == "linux":
                subprocess.call(["xdg-open", 'kb_sub/ot_distribution/' + xl_filename])
            if sys.platform == "darwin":
                subprocess.call(["open", dir_path('ot_distribution') + xl_filename])
        except PermissionError:
            messagebox.showerror("Spreadsheet generator",
                                 "The spreadsheet was not opened. \n"
                                 "Suggestion: "
                                 "Make sure that identically named spreadsheets are closed "
                                 "(the file can't be overwritten while open).",
                                 parent=self.frame)
        self.pb.stop()
