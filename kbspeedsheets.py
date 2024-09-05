"""
a klusterbox module: Klusterbox Speedsheets Generator, Verifications and Input
klusterbox classes for the generation, checking and input of speedsheets.
"""
# custom libraries
import projvar  # defines project variables used in all modules.
from kbtoolbox import CarrierList, CarrierRecFilter, CarrierRecSet, commit, Convert, dir_path, Handler, inquire, \
    MovesChecker, NameChecker, ProgressBarDe, Rings, RingTimeChecker, RouteChecker, SpeedSettings
# standard libraries
from tkinter import messagebox
from datetime import timedelta
from operator import itemgetter
import os
import sys
import subprocess
# non standard libraries
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment, PatternFill, Protection


class SpeedSheetGen:
    """
    this class generates speedsheets.
    """
    def __init__(self, frame, full_report):
        self.frame = frame
        self.full_report = full_report  # true - all inclusive, false - carrier recs only
        self.pb = ProgressBarDe(label="Building SpeedSheet")  # create the progress bar object
        self.db = SpeedSettings()  # calls values from tolerance table
        self.date = projvar.invran_date
        self.day_array = (str(projvar.invran_date.strftime("%a")).lower(),)  # if invran_weekly_span == False
        self.range = "day"
        if projvar.invran_weekly_span:  # if investigation range is weekly
            self.day_array = ("sat", "sun", "mon", "tue", "wed", "thu", "fri")
            self.range = "week"
        self.rotate_mode = self.db.speedcell_ns_rotate_mode  # NS day mode preference: True-rotating or False-fixed
        self.ns_pref = "r"  # "r" for rotating
        if not self.rotate_mode:
            self.ns_pref = "f"  # "f" for fixed
        # speedsheet move triads are configured to 'time/time/route' for false or 'route/time/time for true
        self.triad_routefirst = self.db.triad_routefirst  # move notation preference: True - route first
        self.dlsn_dict = {"sat": "sat", "mon": "mon", "tue": "tue", "wed": "wed", "thu": "thu", "fri": "fri",
                          "rsat": "sat", "rmon": "mon", "rtue": "tue", "rwed": "wed", "rthu": "thu", "rfri": "fri",
                          "fsat": "sat", "fmon": "mon", "ftue": "tue", "fwed": "wed", "fthu": "thu", "ffri": "fri",
                          "  ": "none", "": "none"}
        self.id_recset = []
        self.car_recs = []
        self.speedcell_count = 0
        self.ws_list = []
        self.ws_titles = []
        self.wb = Workbook()  # define the workbook
        self.ws_header = ""
        self.tourrings_mode = None
        self.list_header = ""  # spreadsheet styles
        self.date_dov = ""
        self.date_dov_title = ""
        self.car_col_header = ""
        self.bold_name = ""
        self.input_name = ""
        self.col_header = ""
        self.input_s = ""
        self.input_ns = ""
        self.filename = ""

    def gen(self):
        """ this is the master method. """
        self.get_id_recset()  # get carrier list and format for speedsheets
        self.get_car_recs()  # sort carrier list by worksheet
        self.speedcell_count = self.count()  # get a count of rows for progress bar
        self.make_workbook_object()  # generate and open the workbook
        self.name_styles()  # define the spreadsheet styles
        self.get_tourring_mode()  # if the tourrings mode is true, there are two extra column
        self.make_workbook()  # generate and open the workbook
        self.stopsaveopen()  # stop, save and open

    def get_id_recset(self):
        """ get filtered/ condensed record set and employee id """
        if projvar.invran_weekly_span:  # if the investigation range is weekly
            start = projvar.invran_date_week[0]  # use sat - fri to curate the carrier list
            end = projvar.invran_date_week[6]
        else:  # if the investigation range is one day
            start = projvar.invran_date  # use the one day to curate the carrier list
            end = projvar.invran_date
        carriers = CarrierList(start, end, projvar.invran_station).get()  # first get a carrier list
        for c in carriers:
            # filter out any recs where list status is unchanged
            filtered_recs = CarrierRecFilter(c, projvar.invran_date_week[0]).filter_nonlist_recs()
            # condense multiple recs into format used by speedsheets
            condensed_recs = CarrierRecFilter(filtered_recs, projvar.invran_date_week[0]).condense_recs(
                self.db.speedcell_ns_rotate_mode)
            self.id_recset.append(self.add_id(condensed_recs))  # merge carriers with emp id

    @staticmethod
    def add_id(recset):
        """ put the employee id and carrier records together in a list """
        carrier = recset[1]
        sql = "SELECT emp_id FROM name_index WHERE kb_name = '%s'" % carrier
        result = inquire(sql)
        if len(result) == 1:
            addthis = (result[0][0], recset)
        else:
            addthis = ("", recset)  # if there is no employee id, insert an empty string
        return addthis

    def get_car_recs(self):
        """ sort carrier records by the worksheets they will be put on """
        self.car_recs = [self.order_by_id()]  # combine the id_rec arrays for emp id and alphabetical
        if not self.db.abc_breakdown:
            order_abc = self.order_alphabetically()  # sort the id_recset alphabetically
        else:
            order_abc = self.order_by_abc_breakdown()  # sort the id_recset w/o emp id by abc breakdown
        for abc in order_abc:
            self.car_recs.append(abc)

    def order_by_id(self):
        """ order id_recset by employee id """
        ordered_recs = []
        for rec in self.id_recset:  # loop through the carrier list
            if rec[0] != "":  # if the item for employee id is not empty
                ordered_recs.append(rec)  # add the record set to the array
        ordered_recs.sort(key=itemgetter(0))  # sort the array by the employee id
        return ordered_recs

    def order_alphabetically(self):
        """ order id recset alphabetically into one tab """
        alpha_recset = ["alpha_array", ]
        alpha_recset[0] = []
        for rec in self.id_recset:
            if rec[0] == "":
                alpha_recset[0].append(rec)
        return alpha_recset

    def order_by_abc_breakdown(self):
        """ sort id recset alphabetically into multiple tabs """
        abc_recset = ["a_array", "b_array", "cd_array", "efg_array", "h_array", "ijk_array", "m_array",
                      "nop_array", "qr_array", "s_array", "tuv_array", "w_array", "xyz_array"]
        for i in range(len(abc_recset)):
            abc_recset[i] = []
        for rec in self.id_recset:
            if rec[0] == "":
                if rec[1][1][0] == "a":  # sort names without emp ids into lettered arrays
                    abc_recset[0].append(rec)
                elif rec[1][1][0] == "b":
                    abc_recset[1].append(rec)
                elif rec[1][1][0] == "rec" or rec[1][1][0] == "d":
                    abc_recset[2].append(rec)
                elif rec[1][1][0] == "e" or rec[1][1][0] == "f" or rec[1][1][0] == "g":
                    abc_recset[3].append(rec)
                elif rec[1][1][0] == "h":
                    abc_recset[4].append(rec)
                elif rec[1][1][0] == "i" or rec[1][1][0] == "j" or rec[1][1][0] == "k":
                    abc_recset[5].append(rec)
                elif rec[1][1][0] == "m":
                    abc_recset[6].append(rec)
                elif rec[1][1][0] == "n" or rec[1][1][0] == "o" or rec[1][1][0] == "p":
                    abc_recset[7].append(rec)
                elif rec[1][1][0] == "q" or rec[1][1][0] == "r":
                    abc_recset[8].append(rec)
                elif rec[1][1][0] == "s":
                    abc_recset[9].append(rec)
                elif rec[1][1][0] == "t" or rec[1][1][0] == "u" or rec[1][1][0] == "v":
                    abc_recset[10].append(rec)
                elif rec[1][1][0] == "w":
                    abc_recset[11].append(rec)
                else:
                    abc_recset[12].append(rec)
        return abc_recset

    def count_minrow_array(self):
        """ gets a count of minimum row info for each SpeedSheet tab """
        minrow_array = [self.db.min_empid, ]  # get minimum rows for employee id sheet
        if not self.db.abc_breakdown:
            minrow_array.append(self.db.min_alpha)  # get minimum rows for alphabetical sheet
        else:
            for _ in range(len(self.car_recs) - 1):  # get minimum rows for abc breakdown sheets
                minrow_array.append(self.db.min_abc)
        return minrow_array

    def count_car_recs(self):
        """ gets a count of carrier records for each SpeedSheet tab """
        car_recs = [len(self.car_recs[0]), ]  # get count of carrier recs for employee id sheet
        if not self.db.abc_breakdown:
            car_recs.append(len(self.car_recs[1]))  # get count of carriers for alphabetical sheet
        else:
            for i in range(1, len(self.car_recs)):  # get count of carriers for abc breakdown
                car_recs.append(len(self.car_recs[i]))
        return car_recs

    def count(self):
        """ compare the minimum row and carrier records arrays to get the number of SpeedCells """
        speedcell_count = 0  # initialized the count
        minrows = self.count_minrow_array()  # get minimum row count
        carrecs = self.count_car_recs()  # get count of carriers
        for i in range(len(minrows)):  # loop through results
            speedcell_count += max(minrows[i], carrecs[i])  # take the larger of the two
        return speedcell_count  # return total number of speedcells to be generated

    def mv_to_speed(self, triset):
        """ format mv triads for output to speedsheets """
        if triset == "":
            return triset  # do nothing if blank
        triset = self.triad_reorder(triset)
        return self.mv_format(triset)  # send to mv_format for formating if not blank

    def triad_reorder(self, triset):
        """ reorders the triad to the 'route, time, time' order if self.triad_routefirst is true. """
        if not self.triad_routefirst:  # if the triad route first setting is off
            return triset  # skip the process, return original string
        mv_array = triset.split(",")  # split by commas
        fixed_array = []
        for i in range(0, len(mv_array), 3):  # loop through all elements in groups of three
            fixed_array.append(mv_array[i + 2])  # the last shall be first
            fixed_array.append(mv_array[i])  # and the first shall be second from last
            fixed_array.append(mv_array[i + 1])  # the second if pushed to the last place
        return Convert(fixed_array).array_to_string()  # convert the array back into a string and return it.

    @staticmethod
    def mv_format(triset):
        """ format mv triads for output to speedsheets """
        mv_array = triset.split(",")  # split by commas
        mv_str = ""  # the move string
        i = 1  # initiate counter
        for mv in mv_array:
            mv_str += mv
            if i % 3 != 0:
                mv_str += "+"  # put + between items in move triads
            elif i % 3 == 0 and i != len(mv_array):
                mv_str += "/"  # put / between move triads
            else:
                mv_str += ""  # if at the end
            i += 1  # increment counter
        return mv_str

    def make_workbook_object(self):
        """ make the workbook object """
        if not self.db.abc_breakdown:
            self.ws_list = ["emp_id", "alphabet"]
            self.ws_titles = ["by employee id", "alphabetically"]
        else:
            self.ws_list = ["emp_id", "a", "b", "cd", "efg", "h", "ijk", "m", "nop", "qr", "s", "tuv", "w", "xyz"]
            self.ws_titles = ["by employee id", "a", "b", "c,d", "e,f,g", "h", "i,j,k",
                              "m", "n,o,p", "q,r,", "s", "t,u,v", "w", "x,y,z"]
        self.ws_list[0] = self.wb.active  # create first worksheet
        self.ws_list[0].title = self.ws_titles[0]  # title first worksheet
        self.ws_list[0].protection.sheet = True
        for i in range(1, len(self.ws_list)):  # loop to create all other worksheets
            self.ws_list[i] = self.wb.create_sheet(self.ws_titles[i])
            # self.ws_list[i].protection.sheet = True

    def title(self):
        """ generate title and filename """
        if self.full_report and self.range == "week":
            title = "Speedsheet - All Inclusive Weekly"
            self.filename = "speed_" + str(format(projvar.invran_date_week[0], "%y_%m_%d")) + "_all_w" + ".xlsx"
        elif self.full_report and self.range == "day":
            title = "Speedsheet - All Inclusive Daily"
            self.filename = "speed_" + str(format(projvar.invran_date, "%y_%m_%d")) + "_all_d" + ".xlsx"
        elif not self.full_report and self.range == "week":
            title = "Speedsheet - Carriers"
            self.filename = "speed_" + str(format(projvar.invran_date_week[0], "%y_%m_%d")) + "_carrier_w" + ".xlsx"
        else:  # if not self.full_report and self.range == "day":
            title = "Speedsheet - Carriers"
            self.filename = "speed_" + str(format(projvar.invran_date, "%y_%m_%d")) + "_carrier_d" + ".xlsx"
        return title

    def name_styles(self):
        """ Named styles for workbook """
        bd = Side(style='thin', color="80808080")  # defines borders
        self.ws_header = NamedStyle(name="ws_header", font=Font(bold=True, name='Arial', size=12))
        self.list_header = NamedStyle(name="list_header", font=Font(bold=True, name='Arial', size=10))
        self.date_dov = NamedStyle(name="date_dov", font=Font(name='Arial', size=8))
        self.date_dov_title = NamedStyle(name="date_dov_title", font=Font(bold=True, name='Arial', size=8),
                                         alignment=Alignment(horizontal='right'))
        # other color options: yellow: faf818, blue: 18fafa, green: 18fa20, grey: ababab
        if self.full_report:  # color carrier cells
            self.car_col_header = NamedStyle(name="car_col_header", font=Font(bold=True, name='Arial', size=8),
                                             fill=PatternFill(fgColor='18fafa', fill_type='solid'),
                                             border=Border(left=bd, top=bd, right=bd, bottom=bd),
                                             alignment=Alignment(horizontal='left'))
            self.bold_name = NamedStyle(name="bold_name", font=Font(name='Arial', size=8, bold=True),
                                        fill=PatternFill(fgColor='18fafa', fill_type='solid'),
                                        border=Border(left=bd, top=bd, right=bd, bottom=bd))
            self.input_name = NamedStyle(name="input_name", font=Font(name='Arial', size=8),
                                         fill=PatternFill(fgColor='18fafa', fill_type='solid'),
                                         border=Border(left=bd, top=bd, right=bd, bottom=bd))
        else:  # do not color carrier cells
            self.car_col_header = NamedStyle(name="car_col_header", font=Font(bold=True, name='Arial', size=8),
                                             border=Border(left=bd, top=bd, right=bd, bottom=bd),
                                             alignment=Alignment(horizontal='left'))
            self.bold_name = NamedStyle(name="bold_name", font=Font(name='Arial', size=8, bold=True),
                                        border=Border(left=bd, top=bd, right=bd, bottom=bd))
            self.input_name = NamedStyle(name="input_name", font=Font(name='Arial', size=8),
                                         border=Border(left=bd, top=bd, right=bd, bottom=bd))
        self.col_header = NamedStyle(name="col_header", font=Font(bold=True, name='Arial', size=8),
                                     border=Border(left=bd, top=bd, right=bd, bottom=bd),
                                     alignment=Alignment(horizontal='left'))

        self.input_s = NamedStyle(name="input_s", font=Font(name='Arial', size=8),
                                  border=Border(left=bd, top=bd, right=bd, bottom=bd),
                                  alignment=Alignment(horizontal='right'))
        self.input_ns = NamedStyle(name="input_ns", font=Font(bold=True, name='Arial', size=8, color='ff0000'),
                                   border=Border(left=bd, top=bd, right=bd, bottom=bd),
                                   alignment=Alignment(horizontal='right'))

    def get_tourring_mode(self):
        """ get the tourrings mode. if the tourrings mode is True then the worksheet will include BT and ET.
        Also, columns will be shortened to accomodate the extra columns. """
        sql = "SELECT tolerance FROM tolerances WHERE category = '%s'" % "tourrings"
        results = inquire(sql)
        self.tourrings_mode = int(results[0][0])

    def make_workbook(self):
        """ build the workbook, cell by cell"""
        pi = 0
        empty_sc = 0
        width = 8  # set the widths of the colums
        ringwidth = 8
        if self.tourrings_mode:  # if bt/et are displayed, columns are shorter
            width = 7
            ringwidth = 6
        self.pb.max_count(self.speedcell_count)  # set length of progress bar
        self.pb.start_up()  # start the progress bar
        for i in range(len(self.ws_list)):
            # format cell widths
            self.ws_list[i].oddFooter.center.text = "&A"
            self.ws_list[i].column_dimensions["A"].width = ringwidth
            self.ws_list[i].column_dimensions["B"].width = ringwidth
            self.ws_list[i].column_dimensions["C"].width = ringwidth
            self.ws_list[i].column_dimensions["D"].width = width
            self.ws_list[i].column_dimensions["E"].width = width
            self.ws_list[i].column_dimensions["F"].width = width
            self.ws_list[i].column_dimensions["G"].width = width
            self.ws_list[i].column_dimensions["H"].width = ringwidth
            self.ws_list[i].column_dimensions["I"].width = ringwidth
            self.ws_list[i].column_dimensions["J"].width = width
            self.ws_list[i].column_dimensions["K"].width = width
            self.ws_list[i].column_dimensions["L"].width = width
            # hide BT/ET column if self.tourrings_mode is False
            if not self.tourrings_mode:
                self.ws_list[i].column_dimensions["C"].hidden = True
                self.ws_list[i].column_dimensions["I"].hidden = True
            cell = self.ws_list[i].cell(column=1, row=1)
            cell.value = self.title()
            cell.style = self.ws_header
            self.ws_list[i].merge_cells('A1:J1')
            # create date/ pay period/ station header
            cell = self.ws_list[i].cell(row=2, column=1)  # date label
            cell.value = "Date:  "
            cell.style = self.date_dov_title
            cell = self.ws_list[i].cell(row=2, column=2)  # date
            # if investigation range is daily
            cell.value = "{}".format(projvar.invran_date.strftime("%m/%d/%Y"))
            if projvar.invran_weekly_span:
                cell.value = "{} through {}".format(projvar.invran_date_week[0].strftime("%m/%d/%Y"),
                                                    projvar.invran_date_week[6].strftime("%m/%d/%Y"))
            cell.style = self.date_dov
            self.ws_list[i].merge_cells('B2:E2')
            cell = self.ws_list[i].cell(row=2, column=6)  # pay period label
            cell.value = "PP:  "
            cell.style = self.date_dov_title
            cell = self.ws_list[i].cell(row=2, column=7)  # pay period
            cell.value = projvar.pay_period
            cell.style = self.date_dov
            cell = self.ws_list[i].cell(row=2, column=9)  # station label
            cell.value = "Station:"
            cell.style = self.date_dov_title
            self.ws_list[i].merge_cells('I2:J2')
            cell = self.ws_list[i].cell(row=2, column=11)  # station
            cell.value = projvar.invran_station
            cell.style = self.date_dov
            self.ws_list[i].merge_cells('K2:L2')
            # apply title - show how carriers are sorted
            cell = self.ws_list[i].cell(row=3, column=1)
            if i == 0:
                cell.value = "Carriers listed by Employee ID"
            else:
                cell.value = "Carriers listed Alphabetically: {}".format(self.ws_titles[i])
            cell.style = self.list_header
            self.ws_list[i].merge_cells('A3:F3')
            row = 3
            if i == 0:  # only execute on the first sheet of the workbook
                cell = self.ws_list[i].cell(row=row, column=7)  # ns day preference title and field
                cell.value = "ns day preference (r=rotating/f=fixed): "  # ns day preference
                cell.style = self.date_dov_title
                self.ws_list[i].merge_cells('G' + str(row) + ':' + 'K' + str(row))
                cell = self.ws_list[i].cell(row=row, column=12)  #
                cell.value = self.ns_pref
                cell.style = self.date_dov
            row += 1
            # only display the moves notation preference if the speedsheet is 'all inclusive'
            if i == 0 and self.title() != "Speedsheet - Carriers":
                cell = self.ws_list[i].cell(row=row, column=7)  # move notation title and field
                cell.value = "move notation - route first: "  # title
                cell.style = self.date_dov_title
                self.ws_list[i].merge_cells('G' + str(row) + ':' + 'K' + str(row))
                cell = self.ws_list[i].cell(row=row, column=12)  # value is True or False
                cell.value = str(self.triad_routefirst)
                cell.style = self.date_dov
                row += 1
            # Headers for Carrier List
            cell = self.ws_list[i].cell(row=row, column=1)  # header day
            cell.value = "Days"
            cell.style = self.car_col_header
            cell = self.ws_list[i].cell(row=row, column=2)  # header carrier name
            cell.value = "Carrier Name"
            cell.style = self.car_col_header
            self.ws_list[i].merge_cells('B' + str(row) + ':' + 'E' + str(row))
            cell = self.ws_list[i].cell(row=row, column=6)  # header list type
            cell.value = "List"
            cell.style = self.car_col_header
            cell = self.ws_list[i].cell(row=row, column=7)  # header ns day
            cell.value = "NS Day"
            cell.style = self.car_col_header
            cell = self.ws_list[i].cell(row=row, column=8)  # header route
            cell.value = "Route/s"
            cell.style = self.car_col_header
            self.ws_list[i].merge_cells('H' + str(row) + ':' + 'K' + str(row))
            cell = self.ws_list[i].cell(row=row, column=12)  # header emp id
            cell.value = "Emp id"
            cell.style = self.car_col_header
            row += 1  # start at row 5 after the page header display
            if self.full_report:  # only include rings headers on all inclusive not carrier only
                # Headers for Rings
                cell = self.ws_list[i].cell(row=row, column=1)  # header day
                cell.value = "Day"
                cell.style = self.col_header
                cell = self.ws_list[i].cell(row=row, column=2)  # header 5200
                cell.value = "5200"
                cell.style = self.col_header
                cell = self.ws_list[i].cell(row=row, column=3)  # header BT
                cell.value = "BT"
                cell.style = self.col_header
                cell = self.ws_list[i].cell(row=row, column=4)  # header MOVES
                cell.value = "MOVES"
                cell.style = self.col_header
                self.ws_list[i].merge_cells('D' + str(row) + ':' + 'G' + str(row))
                cell = self.ws_list[i].cell(row=row, column=8)  # header RS
                cell.value = "RS"
                cell.style = self.col_header
                cell = self.ws_list[i].cell(row=row, column=9)  # header ET
                cell.value = "ET"
                cell.style = self.col_header
                cell = self.ws_list[i].cell(row=row, column=10)  # header codes
                cell.value = "CODE"
                cell.style = self.col_header
                cell = self.ws_list[i].cell(row=row, column=11)  # header leave type
                cell.value = "LV type"
                cell.style = self.col_header
                cell = self.ws_list[i].cell(row=row, column=12)  # header leave time
                cell.value = "LV time"
                cell.style = self.col_header
                row += 1  # update start at row 7 after the page header display if all inclusive
            # freeze panes
            self.ws_list[i].freeze_panes = self.ws_list[i].cell(row=row, column=1)  # ['A6'] or ['A7']
            if i == 0:
                rowcount = self.db.min_empid  # get minimum speedcell count for employee id tab
            elif i != 0 and not self.db.abc_breakdown:
                rowcount = self.db.min_alpha  # get minimum speedcell count for alphabetical tab
            else:
                rowcount = self.db.min_abc  # get minimum speedcell count for each abc breakdown tab
            rowcounter = max(rowcount, len(self.car_recs[i]))
            for r in range(rowcounter):
                if r < len(self.car_recs[i]):  # if the carrier records are not exhausted
                    eff_date = self.car_recs[i][r][1][0]  # carrier effective date
                    car_name = self.car_recs[i][r][1][1]  # carrier name
                    car_list = self.car_recs[i][r][1][2]  # carrier list status
                    car_ns = self.car_recs[i][r][1][3]  # carrier ns day
                    car_route = self.car_recs[i][r][1][4]  # carrier route
                    car_empid = self.car_recs[i][r][0]  # carrier employee id number
                    self.pb.move_count(pi)  # increment progress bar
                    self.pb.change_text("Formatting Speedcell for {}".format(car_name))
                else:
                    eff_date = ""  # enter blanks once records are exhausted
                    car_name = ""
                    car_list = ""
                    car_ns = ""
                    car_route = ""
                    car_empid = ""
                    self.pb.move_count(pi)  # increment progress bar
                    empty_sc += 1  # increment counter for empty speedcells
                    self.pb.change_text("Formatting empty Speedcell #{}".format(empty_sc))
                pi += 1  # progress bar counter
                cell = self.ws_list[i].cell(row=row, column=1)  # carrier effective date
                cell.value = eff_date
                cell.style = self.input_name
                cell.protection = Protection(locked=False)
                cell = self.ws_list[i].cell(row=row, column=2)  # carrier name
                cell.value = car_name
                cell.style = self.bold_name
                cell.number_format = "@"
                cell.protection = Protection(locked=False)
                self.ws_list[i].merge_cells('B' + str(row) + ':' + 'E' + str(row))
                cell = self.ws_list[i].cell(row=row, column=6)  # carrier list status
                cell.value = car_list
                cell.style = self.input_name
                cell.number_format = "@"
                cell.protection = Protection(locked=False)
                cell = self.ws_list[i].cell(row=row, column=7)  # carrier ns day
                cell.value = car_ns
                cell.style = self.input_name
                cell.number_format = "@"
                cell.protection = Protection(locked=False)
                cell = self.ws_list[i].cell(column=8, row=row)  # carrier route
                cell.value = car_route
                cell.style = self.input_name
                cell.number_format = "@"
                cell.protection = Protection(locked=False)
                self.ws_list[i].merge_cells('H' + str(row) + ':' + 'K' + str(row))
                cell = self.ws_list[i].cell(column=12, row=row)  # carrier emp id
                cell.value = car_empid
                cell.style = self.input_name
                cell.number_format = "@"
                cell.protection = Protection(locked=False)
                row += 1
                ring_recs = []
                if self.full_report:
                    if r < len(self.car_recs[i]):  # if the carrier records are not exhausted
                        if self.range == "day":  # if the investigation range is for a day
                            ring_recs = Rings(car_name, self.date).get_for_day()  # get the rings for the carrier
                        else:  # if the investigation range is for a week
                            ring_recs = Rings(car_name, self.date).get_for_week()  # get the rings for the carrier
                    for d in range(len(self.day_array)):
                        if r < len(self.car_recs[i]) and ring_recs[d] != []:  # if the carrier records are not exhausted
                            ring_5200 = Convert(ring_recs[d][2]).str_to_floatoremptystr()  # rings 5200
                            ring_move = self.mv_to_speed(ring_recs[d][5])  # format rings MOVES
                            ring_rs = Convert(ring_recs[d][3]).str_to_floatoremptystr()  # rings RS
                            ring_code = Convert(ring_recs[d][4]).empty_not_none()  # rings CODES
                            ring_lvty = Convert(ring_recs[d][6]).empty_not_none()  # rings LEAVE TYPE
                            ring_lvtm = Convert(ring_recs[d][7]).str_to_floatoremptystr()  # rings LEAVE TIME
                            ring_bt = Convert(ring_recs[d][9]).str_to_floatoremptystr()  # rings BT
                            ring_et = Convert(ring_recs[d][10]).str_to_floatoremptystr()  # rings ET
                        else:
                            ring_5200 = ""  # rings 5200
                            ring_move = ""  # rings MOVES
                            ring_rs = ""  # rings RS
                            ring_code = ""  # rings CODES
                            ring_lvty = ""  # rings LEAVE TYPE
                            ring_lvtm = ""  # rings LEAVE TIME
                            ring_bt = ""  # rings BT
                            ring_et = ""  # rings ET
                        cell = self.ws_list[i].cell(column=1, row=row)  # rings day
                        cell.value = self.day_array[d]
                        if self.day_array[d] == self.dlsn_dict[car_ns.lower()]:  # if it is the nsday
                            cell.style = self.input_ns  # display it red and bold
                        else:
                            cell.style = self.input_s
                        cell = self.ws_list[i].cell(column=2, row=row)  # rings 5200
                        cell.value = ring_5200
                        cell.style = self.input_s
                        cell.number_format = "#,###.00;[RED]-#,###.00"
                        cell.protection = Protection(locked=False)
                        cell = self.ws_list[i].cell(column=3, row=row)  # rings BT
                        cell.value = ring_bt
                        cell.style = self.input_s
                        cell.number_format = "#,###.00;[RED]-#,###.00"
                        cell.protection = Protection(locked=False)
                        cell = self.ws_list[i].cell(column=4, row=row)  # rings moves
                        cell.value = ring_move
                        cell.style = self.input_s
                        cell.protection = Protection(locked=False)
                        self.ws_list[i].merge_cells('D' + str(row) + ':' + 'G' + str(row))
                        cell = self.ws_list[i].cell(column=8, row=row)  # rings RS
                        cell.value = ring_rs
                        cell.style = self.input_s
                        cell.number_format = "#,###.00;[RED]-#,###.00"
                        cell.protection = Protection(locked=False)
                        cell = self.ws_list[i].cell(column=9, row=row)  # rings ET
                        cell.value = ring_et
                        cell.style = self.input_s
                        cell.number_format = "#,###.00;[RED]-#,###.00"
                        cell.protection = Protection(locked=False)
                        cell = self.ws_list[i].cell(column=10, row=row)  # rings code
                        cell.value = ring_code
                        cell.style = self.input_s
                        cell.protection = Protection(locked=False)
                        cell = self.ws_list[i].cell(column=11, row=row)  # rings lv type
                        cell.value = ring_lvty
                        cell.style = self.input_s
                        cell.protection = Protection(locked=False)
                        cell = self.ws_list[i].cell(column=12, row=row)  # rings lv time
                        cell.value = ring_lvtm
                        cell.style = self.input_s
                        cell.number_format = "#,###.00;[RED]-#,###.00"
                        cell.protection = Protection(locked=False)
                        row += 1
        self.pb.stop()

    def stopsaveopen(self):
        """ save and open the speedsheet. """
        try:
            self.wb.save(dir_path('speedsheets') + self.filename)
            messagebox.showinfo("Speedsheet Generator",
                                "Your speedsheet was successfully generated. \n"
                                "File is named: {}".format(self.filename),
                                parent=self.frame)
            if sys.platform == "win32":
                os.startfile(dir_path('speedsheets') + self.filename)
            if sys.platform == "linux":
                subprocess.call(["xdg-open", 'kb_sub/speedsheets/' + self.filename])
            if sys.platform == "darwin":
                subprocess.call(["open", dir_path('speedsheets') + self.filename])
        except PermissionError:
            messagebox.showerror("Speedsheet generator",
                                 "The speedsheet was not generated. \n"
                                 "Suggestion: \n"
                                 "Make sure that identically named speedsheets are closed \n"
                                 "(the file can't be overwritten while open).\n",
                                 parent=self.frame)


class OpenText:
    """
    This is a class used by the About Klusterbox to open files for viewing.
    """
    def __init__(self):
        self.frame = None

    def open_docs(self, frame, doc):
        """ opens docs in the about_klusterbox() function """
        self.frame = frame
        try:
            if sys.platform == "win32":
                if projvar.platform == "py":
                    try:
                        path = doc
                        os.startfile(path)  # in IDE the files are in the project folder
                    except FileNotFoundError:
                        path = os.path.join(os.path.sep, os.getcwd(), 'kb_sub', doc)
                        os.startfile(path)  # in KB legacy the files are in the kb_sub folder
                if projvar.platform == "winapp":
                    path = os.path.join(os.path.sep, os.getcwd(), doc)
                    os.startfile(path)
            if sys.platform == "linux":
                subprocess.call(doc)
            if sys.platform == "darwin":
                if projvar.platform == "macapp":
                    path = os.path.join(os.path.sep, 'Applications', 'klusterbox.app', 'Contents', 'Resources', doc)
                    subprocess.call(["open", path])
                if projvar.platform == "py":
                    subprocess.call(["open", doc])
        except FileNotFoundError:
            messagebox.showerror("Project Documents",
                                 "The document was not opened or found.",
                                 parent=self.frame)


class SpeedCarrierCheck:
    """
    accepts carrier records from SpeedSheets
    """

    def __init__(self, parent, sheet, row, name, day, list_stat, nsday, route, empid):
        self.parent = parent  # get objects from SpeedSheetCheck
        self.sheet = sheet  # input here is coming directly from the speedcell
        self.row = str(row)
        self.name = name  # get information passed from SpeedCell
        self.day = day
        self.list_stat = list_stat
        self.nsday = nsday.lower()
        self.route = route
        self.empid = empid
        self.tacs_name = ""  # get names and employee id numbers from name index
        self.kb_name = ""
        self.index_id = ""
        sql = "SELECT * FROM name_index WHERE kb_name = '%s'" % self.name  # access dbase to check emp id
        result = inquire(sql)
        if result:
            self.tacs_name = result[0][0]
            self.kb_name = result[0][1]
            self.index_id = result[0][2]
        self.filtered_recset = []
        self.onrec_date = ""  # get carrier information "on record" from the database
        self.onrec_name = ""
        self.onrec_list = ""
        self.onrec_nsday = ""
        self.onrec_route = ""
        self.addday = []  # checked input formatted for entry into database
        self.addlist = ["empty"]
        self.addnsday = "empty"
        self.addroute = "empty"
        self.addempid = ""
        self.parent.allowaddrecs = True  # if False, records will not be added to database
        self.error_array = []  # arrays for error, fyi and add reports
        self.fyi_array = []
        self.attn_array = []
        self.add_array = []
        self.ns_dict = \
            {"s": "sat", "m": "mon", "tu": "tue", "u": "tue", "w": "wed", "th": "thu", "h": "thu", "f": "fri",
             "fs": "sat", "fm": "mon", "ftu": "tue", "fu": "tue", "fw": "wed", "fth": "thu", "fh": "thu", "ff": "fri",
             "rs": "sat", "rm": "mon", "rtu": "tue", "ru": "tue", "rw": "wed", "rth": "thu", "rh": "thu", "rf": "fri",
             "sat": "sat", "mon": "mon", "tue": "tue", "wed": "wed", "thu": "thu", "fri": "fri",
             "rsat": "sat", "rmon": "mon", "rtue": "tue", "rwed": "wed", "rthu": "thu", "rfri": "fri",
             "fsat": "sat", "fmon": "mon", "ftue": "tue", "fwed": "wed", "fthu": "thu", "ffri": "fri"}

    def check_all(self):
        """ master method to run other methods. """
        self.get_carrec()  # get carrier records and condense them into one array
        self.check_name()  # check for errors with the carrier name
        self.check_employee_id_format()
        self.check_employee_id_situation()
        self.check_employee_id_use()
        self.check_list_status()
        self.check_ns()
        self.check_route()
        if self.parent.interject:  # True = add to database/ False = pre-check
            self.add_recs()
        self.generate_report()

    def get_carrec(self):
        """ get carrier records and condense them into one array """
        carrec = CarrierRecSet(self.name, self.parent.start_date, self.parent.end_date, self.parent.station).get()
        self.filtered_recset = CarrierRecFilter(carrec, self.parent.start_date).filter_nonlist_recs()
        carrec = CarrierRecFilter(self.filtered_recset, self.parent.start_date).condense_recs_ns()
        self.onrec_date = carrec[0]
        self.onrec_name = carrec[1]
        self.onrec_list = carrec[2]
        self.onrec_nsday = carrec[3]
        self.onrec_route = carrec[4]

    def check_name(self):
        """ check for errors with the carrier name """
        if self.name == self.onrec_name:
            return
        if not NameChecker(self.name).check_characters():
            error = "     ERROR: Carrier name can not contain numbers or most special characters\n"
            self.error_array.append(error)
            self.parent.allowaddrecs = False  # do not allow this speedcell be be input into database
        if not NameChecker(self.name).check_length():
            error = "     ERROR: Carrier name must not exceed 42 characters\n"
            self.error_array.append(error)
            self.parent.allowaddrecs = False  # do not allow this speedcell be be input into database
        if not NameChecker(self.name).check_comma():
            error = "     ERROR: Carrier name must contain one comma to separate last name and first initial\n"
            self.error_array.append(error)
            self.parent.allowaddrecs = False  # do not allow this speedcell be be input into database
        if not NameChecker(self.name).check_initial():
            attn = "     ATTENTION: Carrier name should must contain one initial ideally, \n" \
                   "                unless more are needed to create a distinct carrier name.\n"
            self.attn_array.append(attn)

    def check_employee_id_situation(self):
        """ checks the employee id. """
        if self.index_id == "" and self.empid == "":  # if both emp id and name index are blank
            pass
        elif self.index_id == self.empid:  # if the emp id from the name index and the speedsheet match
            pass
        elif self.index_id != "" and self.empid == "":  # if value in name index but spdcell is blank
            attn = "     ATTENTION: employee id can not be deleted from speedsheet\n"
            self.attn_array.append(attn)  # place this on "addition" report for user's information
            return
        elif self.index_id == "" and self.empid != "":  # if name index blank and spd cell has a value
            self.addempid = self.empid
            attn = "     ATTENTION: Possible new employee id\n"  # report
            self.attn_array.append(attn)
        else:
            error = "     ERROR: Employee id contridiction. \n" \
                    "            You can not change employee id with speedsheet\n"  # report
            self.error_array.append(error)
            self.parent.allowaddrecs = False  # do not allow this speedcell be be input into database

    def check_employee_id_format(self):
        """ verifies the employee id """
        if self.empid == "":  # allow empty strings
            pass
        elif str(self.empid).isnumeric():  # allow integers and numeric strings
            self.empid = str(self.empid).zfill(8)  # change self.empid to string and zero fill to 8 places
            pass
        else:  # don't allow anything else
            error = "     ERROR: employee id is not numeric\n"  # report
            self.error_array.append(error)
            self.parent.allowaddrecs = False
            return

    def check_employee_id_use(self):
        """ make sure the employee id is not being used by another carrier """
        kb_name = ""
        emp_id = ""
        if self.empid != "":
            sql = "SELECT * FROM name_index WHERE emp_id = '%s'" % self.empid
            result = inquire(sql)
            if result:
                kb_name = result[0][1]
                emp_id = result[0][2]
        if emp_id == "":
            return
        elif kb_name == self.name:
            pass
        else:
            error = "     ERROR: employee id is in use by another carrier\n"
            self.error_array.append(error)
            self.parent.allowaddrecs = False

    def add_list_status(self, dlsn_array, dlsn_day_array):
        """ checks for list status.
        enters dynamic list status notation array and dynamic list status notation day array. """
        if not self.filtered_recset:  # if the carrier is new
            self.addlist = dlsn_array
            self.addday = dlsn_day_array
            fyi = "     FYI: New List status will be entered: {}\n".format(dlsn_array)
            self.fyi_array.append(fyi)
        elif self.onrec_list != Convert(dlsn_array).array_to_string():  # if the list has changed
            self.addlist = dlsn_array
            self.addday = dlsn_day_array
            fyi = "     FYI: List status will be updated to: {}\n".format(dlsn_array)
            self.fyi_array.append(fyi)
        elif self.onrec_date != Convert(dlsn_day_array).array_to_string():  # if the days have changed
            self.addlist = dlsn_array
            self.addday = dlsn_day_array
            fyi = "     FYI: List status will be updated to: {}\n".format(dlsn_array)
            self.fyi_array.append(fyi)
        else:  # if there has been no change, do not change add___ vars.
            pass

    def check_list_status(self):
        """ adds list status. """
        self.list_stat = str(self.list_stat)
        self.list_stat = self.list_stat.strip()
        if self.list_stat == "":  # if the list_stat is empty
            self.add_list_status(["nl"], [])
            return
        dlsn_array = []  # dynamic list status notation array
        if self.list_stat != "":
            dlsn_array = Convert(self.list_stat).string_to_array()
        if len(dlsn_array) > 6:  # check number of list status changes
            error = "     ERROR: More than six changes in list status are not allowed\n"
            self.error_array.append(error)
            self.parent.allowaddrecs = False
            return
        for ls in dlsn_array:  # check for any input that does not conform with list status notation
            ls = ls.strip()  # strip any whitespace
            ls = ls.lower()  # make lowercase
            if ls in ("n", "w", "o", "a", "p", "c"):  # acceptable values
                pass
            elif ls in ("nl", "wal", "otdl", "odl", "aux", "cca", "ptf"):  # acceptable values
                pass
            else:
                error = "     ERROR: No such list status or list status notation {}\n".format(ls)
                self.error_array.append(error)
                self.parent.allowaddrecs = False
                return
        dlsn_array = self.dlsn_baseready(dlsn_array)  # format the list status/es for database
        # check days
        self.day = str(self.day)
        self.day = self.day.strip()
        dlsn_day_array = []  # dynamic list status notation day array
        if self.day != "":
            dlsn_day_array = Convert(self.day).string_to_array()
        if len(dlsn_day_array) > 7:
            error = "     ERROR: More than seven changes in days are not allowed\n"
            self.error_array.append(error)
            self.parent.allowaddrecs = False
        if len(dlsn_day_array) == 0 and len(dlsn_array) == 0:
            return
        elif len(dlsn_day_array) + 1 > len(dlsn_array):
            error = "     ERROR: Too many days compared to the list status {}\n" \
                    "            (hint: SpeedCell notation does not mention the \n" \
                    "            first day.) \n".format(self.day)
            self.error_array.append(error)
            self.parent.allowaddrecs = False
            return
        elif len(dlsn_day_array) + 1 < len(dlsn_array):
            error = "     ERROR: Too many list statuses compared to days {}\n" \
                    "            (SpeedCell notation requires that list status \n" \
                    "            changes be accompanied by the day of the change.) \n".format(self.day)
            self.error_array.append(error)
            self.parent.allowaddrecs = False
            return
        else:
            pass
        for d in dlsn_day_array:
            d = d.strip()  # strip any whitespace
            d = d.lower()  # make lowercase
            if d in ("s", "m", "tu", "u", "w", "th", "h", "f"):
                pass
            elif d in ("sat", "mon", "tue", "wed", "thu", "fri"):
                pass
            else:
                error = "     ERROR: No such day or day notation {}\n".format(d)
                self.error_array.append(error)
                self.parent.allowaddrecs = False
                return
        dlsn_day_array = self.day_baseready(dlsn_day_array)  # format the day/s for the database
        if self.check_day_sequence(dlsn_day_array) is False:  # check days for correct sequence
            return
        self.add_list_status(dlsn_array, dlsn_day_array)

    @staticmethod
    def dlsn_baseready(array):
        """ format dynamic list status notation into database ready """
        new = []
        for ls in array:  # for each list status
            if ls in ("nl", "n"):
                new.append("nl")
            if ls in ("wal", "w"):
                new.append("wal")
            if ls in ("otdl", "odl", "o"):
                new.append("otdl")
            if ls in ("aux", "a", "cca", "c"):
                new.append("aux")
            if ls in ("ptf", "p"):
                new.append("ptf")
        return new

    def check_day_sequence(self, array):
        """ check the day/s for correct sequence """
        sequence = ("sat", "sun", "mon", "tue", "wed", "thu", "fri")
        past = []
        for a in array:
            if a in past:
                error = "     ERROR: Days are out of sequence {}\n".format(self.day)
                self.error_array.append(error)
                self.parent.allowaddrecs = False
                return False
            for s in sequence:
                if s == a:
                    past.append(s)
                    break
                past.append(s)

    @staticmethod
    def day_baseready(array):
        """ format dynamic list status notation into database ready """
        new = []
        for d in array:
            if d in ("sat", "s"):
                new.append("sat")
            if d in ("mon", "m"):
                new.append("mon")
            if d in ("tue", "tu", "u"):
                new.append("tue")
            if d in ("wed", "w"):
                new.append("wed")
            if d in ("thu", "th", "h"):
                new.append("thu")
            if d in ("fri", "f"):
                new.append("fri")
        return new

    def ns_baseready(self, ns, mode):
        """ formats provided ns day into a fixed or rotating ns day for database input """
        baseready = self.parent.ns_true_rev[ns]  # if True is passed use rotate mode
        if not mode:  # if False is passed use fixed mode
            baseready = self.parent.ns_false_rev[ns]
        return baseready

    def add_ns(self, baseready):
        """ add ns day """
        if self.onrec_nsday == baseready:
            pass  # keep value of addnsday var as "empty"
        else:
            fyi = "     FYI: New or updated nsday: {}.\n".format(self.parent.ns_custom[baseready])  # report
            self.fyi_array.append(fyi)
            self.addnsday = baseready

    def check_ns(self):
        """ self.parent.ns_rotate_mode: True for rotate, False for fixed """
        ns = "none"  # initialize ns variable
        if not self.nsday:  # if string is empty
            self.add_ns(ns)  # ns day is "none"
        if self.nsday in ("sat", "mon", "tue", "wed", "thu", "fri"):
            baseready = self.ns_baseready(self.nsday, self.parent.ns_rotate_mode)  # format for dbase input
        elif self.nsday in ("s", "m", "tu", "u", "w", "th", "h", "f"):
            ns = self.ns_dict[self.nsday]  # translate the notation
            baseready = self.ns_baseready(ns, self.parent.ns_rotate_mode)
        elif self.nsday == "  ":  # if the string is almost empty
            baseready = ns  # ns day is "none"
        elif self.nsday in ("rsat", "rmon", "rtue", "rwed", "rthu", "rfri",
                            "rs", "rm", "rtu", "ru", "rw", "rth", "rh", "rf"):
            ns = self.ns_dict[self.nsday]  # use dictionary to get the day
            baseready = self.ns_baseready(ns, True)  # use ns rotate mode to get correct dictionary for day
        elif self.nsday in ("fsat", "fmon", "ftue", "fwed", "fthu", "ffri",
                            "fs", "fm", "ftu", "fu", "fw", "fth", "fh", "ff"):
            ns = self.ns_dict[self.nsday]
            baseready = self.ns_baseready(ns, False)  # use ns rotate mode to get correct dictionary for day
        else:
            error = "     ERROR: No such nsday: \"{}\"\n".format(self.nsday)  # report
            self.error_array.append(error)
            self.parent.allowaddrecs = False  # do not allow speedcell to be input into dbase
            return
        self.add_ns(baseready)

    def add_route(self):
        """ add route """
        if self.route == self.onrec_route:
            pass  # retain "empty" value for addroute variable
        else:
            fyi = "     FYI: New or updated route: {}\n".format(self.route)
            self.fyi_array.append(fyi)
            self.addroute = self.route  # save to input to dbase

    def check_route(self):
        """ check route """
        self.route = str(self.route)
        self.route = self.route.strip()
        if self.route == "":
            self.add_route()
        elif 4 > len(self.route) > 0:  # zero fill any inputs with between 0 and 4 digits
            self.route = self.route.zfill(4)
        if not RouteChecker(self.route).check_all():
            error = "     ERROR: Improper route formatting\n"  # report
            self.error_array.append(error)
            self.parent.allowaddrecs = False  # do not allow speedcell to be input into dbase
            return
        else:
            self.route = Handler(self.route).routes_adj()
            self.add_route()

    def add_recs(self):
        """ add records using the add___ vars. """
        chg_these = []
        if not self.parent.allowaddrecs:  # if all checks passed
            return
        self.name = self.name.lower()  # make sure the name is lowercase
        if self.addlist != ["empty"]:
            add = "     INPUT: List Status added or updated to database >>{}\n" \
                .format(Convert(self.addlist).array_to_string())  # report
            self.add_array.append(add)
            chg_these.append("list")
            list_place = self.addlist
        else:
            list_place = Convert(self.onrec_list).string_to_array()
        if self.addnsday != "empty":
            add = "     INPUT: Nsday added or updated to database >>{}\n".format(self.addnsday)  # report
            self.add_array.append(add)
            chg_these.append("ns")
            ns_place = self.addnsday
        else:
            ns_place = self.onrec_nsday
        if self.addroute != "empty":
            add = "     INPUT: Route added or updated to database >>{}\n".format(self.addroute)  # report
            self.add_array.append(add)
            chg_these.append("route")
            route_place = self.addroute
        else:
            route_place = self.onrec_route
        if self.addempid != "":
            sql = "INSERT INTO name_index (tacs_name, kb_name, emp_id) VALUES('%s', '%s', '%s')" \
                  % ("", self.name, str(self.empid).zfill(8))
            commit(sql)
            add = "     INPUT: Employee id added or updated to database >>{}\n".format(self.addempid)  # report
            self.add_array.append(add)
        # is the earliest car rec a Relevent Preceeding Record or a sat range:
        rpr = True  # Relevent Preceeding Record
        if self.filtered_recset:
            lastrec = self.filtered_recset.pop()  # get the earliest rec from rec set
            if lastrec[0] == str(self.parent.start_date):  # if last rec is the saturday in range
                rpr = False  # then there is no RPR
        if len(chg_these) != 0:  # build the first rec
            if rpr:  # insert the first rec
                sql = "INSERT INTO carriers(effective_date, carrier_name, list_status, ns_day, route_s, " \
                      "station) VALUES('%s','%s','%s','%s','%s','%s')" \
                      % (self.parent.start_date, self.name, list_place[0], ns_place, route_place, self.parent.station)
            else:  # update the first rec to replace pre existing record.
                sql = "UPDATE carriers SET list_status = '%s', ns_day = '%s', route_s = '%s', station = '%s'" \
                      "WHERE carrier_name = '%s' and effective_date = '%s'" \
                      % (list_place[0], ns_place, route_place, self.parent.station, self.name, self.parent.start_date)
            commit(sql)
        if self.addlist != ["empty"] and "list" in chg_these:
            second_date = self.parent.start_date + timedelta(days=1)
            seventh_date = self.parent.end_date  # delete all dates in service week except sat range
            sql = "DELETE FROM carriers WHERE carrier_name = '%s' and effective_date BETWEEN '%s' and '%s'" % \
                  (self.name, second_date, seventh_date)
            commit(sql)  # delete any records in investigation range except saturday
            for i in range(len(self.addlist)):
                if i == 0:
                    pass  # the first rec has already been entered
                else:
                    date = Convert(self.addday[i - 1]).day_to_datetime_str(self.parent.start_date)
                    sql = "INSERT INTO carriers(effective_date, carrier_name, list_status, ns_day, route_s, " \
                          "station) VALUES('%s','%s','%s','%s','%s','%s')" \
                          % (date, self.name, list_place[i], ns_place, route_place, self.parent.station)
                    commit(sql)

    def generate_report(self):
        """ generate a report """
        self.parent.fatal_rpt += len(self.error_array)
        self.parent.add_rpt += len(self.add_array)
        self.parent.fyi_rpt += len(self.fyi_array)
        if not self.parent.interject:
            master_array = self.error_array + self.attn_array  # use these reports for precheck
            if self.parent.fullreport:  # if the full report option is selected...
                master_array += self.fyi_array   # include the fyi messages.
        else:
            master_array = self.error_array + self.attn_array  # use these reports for input
            if self.parent.fullreport:  # if the full report option is selected...
                master_array += self.add_array  # include the adds messages.
        if len(master_array) > 0:
            if not self.parent.name_mentioned:
                self.parent.report.write("\n{}\n".format(self.name))
                self.parent.name_mentioned = True
            self.parent.report.write("   >>> sheet: \"{}\" --> row: \"{}\"  <<<\n".format(self.sheet, self.row))
            if not self.parent.allowaddrecs:
                self.parent.report.write("     SPEEDCELL ENTRY PROHIBITED: Correct errors!\n")
                # self.parent.fatal_rpt += 1
            for rpt in master_array:  # write all reports that have been keep in arrays.
                self.parent.report.write(rpt)


class SpeedRingCheck:
    """
    accepts carrier rings from SpeedSheets
    """

    def __init__(self, parent, sheet, row, day, hours, bt, moves, rs, et, codes, lv_type, lv_time):
        self.parent = parent
        self.sheet = sheet
        self.row = row
        self.day = day
        self.hours = hours
        self.bt = bt
        self.moves = moves
        self.rs = rs
        self.et = et
        self.codes = codes
        self.lv_type = lv_type
        self.lv_time = lv_time
        self.allowaddrings = True
        self.error_array = []
        self.fyi_array = []
        self.attn_array = []
        self.add_array = []
        self.onrec_list = ""  # get carrier information "on record" from the database
        self.onrec_nsday = ""
        self.onrec_route = ""
        self.onrec_date = ""  # get rings information "on record" from the database
        self.onrec_name = ""
        self.onrec_5200 = ""
        self.onrec_bt = ""
        self.onrec_moves = ""
        self.onrec_rs = ""
        self.onrec_et = ""
        self.onrec_codes = ""
        self.onrec_leave_type = ""
        self.onrec_leave_time = ""
        self.adddate = "empty"  # checked input formatted for entry into database
        self.add5200 = "empty"
        self.addbt = "empty"
        self.addrs = "empty"
        self.addet = "empty"
        self.addcode = "empty"
        self.addmoves = "empty"
        self.addlvtype = "empty"
        self.addlvtime = "empty"
        self.exist5200 = False
        self.existbt = False
        self.auto_et = False

    def check(self):
        """ master method for running methods in sequence. """
        if self.check_day():  # if the day is a valid day
            self.get_onrecs()  # get existing "on record" records from the database
            self.check_5200()  # check 5200/ hours
            self.check_leave_time()  # check leave time
            if not self.check_empty():  # checks if the record should be deleted
                self.check_bt()  # check "begin tour"
                self.check_et()  # check "end tour"
                self.check_rs()  # check "return to station"
                self.check_codes()  # check the codes/notes
                self.check_leave_type()  # check leave type
                self.check_moves()  # check moves
                if self.parent.interject:  # if user wants to update database
                    self.add_recs()  # format and input rings into database
        self.generate_report()

    def get_day_as_datetime(self):
        """ get the datetime object for the day in use """
        day = Convert(self.day).day_to_datetime_str(self.parent.start_date)
        self.adddate = day
        return day

    def get_onrecs(self):
        """ gets the records already in the database ie on record. """
        carrec = CarrierRecSet(self.parent.name, self.parent.start_date, self.parent.end_date,
                               self.parent.station).get()
        if carrec:
            self.onrec_list = carrec[0][2]  # get carrier information "on record" from the database
            self.onrec_nsday = carrec[0][3]
            self.onrec_route = carrec[0][4]
            ringrec = Rings(self.parent.name, self.get_day_as_datetime()).get_for_day()
            if ringrec[0]:  # if there is a result for clock rings on that day
                self.onrec_date = ringrec[0][0]  # get rings information "on record" from the database
                self.onrec_name = ringrec[0][1]
                self.onrec_5200 = ringrec[0][2]
                self.onrec_rs = ringrec[0][3]
                self.onrec_codes = ringrec[0][4]
                self.onrec_moves = ringrec[0][5]
                self.onrec_leave_type = ringrec[0][6]
                self.onrec_leave_time = ringrec[0][7]
                self.onrec_bt = ringrec[0][9]
                self.onrec_et = ringrec[0][10]

    def check_day(self):
        """ checks the day. """
        days = ("sat", "sun", "mon", "tue", "wed", "thu", "fri")
        self.day = self.day.strip()
        self.day = str(self.day)
        self.day = self.day.lower()
        if self.day not in days:
            error = "     ERROR: Rings day is not correctly formatted. Acceptable values: sat, sun \n" \
                    "     mon, tue, wed, thu, or fri. Got instead \"{}\": \n".format(self.day)
            self.error_array.append(error)
            self.allowaddrings = False  # do not allow speedcell to be input into dbase
            return False
        return True

    def check_empty(self):
        """ determine conditions where existing record is deleted """
        permitted_codes = ("no call", "ns day", "annual", "sick", "excused")
        if not self.hours:
            if not self.lv_time:
                if self.codes not in permitted_codes:
                    if self.onrec_date:  # if there is an existing record to delete
                        self.delete_recs()  # delete any pre existing record
                    return True
        return False

    def add_5200(self):
        """ adds 5200 time to an add5200 array which will add values to database. """
        if self.hours == "0.0" and self.onrec_5200 in ("0", "0.00", "0.0", "", 0, 0.0):
            pass
        elif self.hours != self.onrec_5200:  # compare 5200 time against 5200 from database,
            self.add5200 = self.hours  # if different, the add
            fyi = "     FYI: New or updated 5200 time: {}\n".format(self.hours)
            self.fyi_array.append(fyi)

    def check_5200(self):
        """ checks the 5200 time """
        if type(self.hours) == str and not self.hours:  # pass if value is an empty string
            self.add_5200()
            return
        ring = RingTimeChecker(self.hours).make_float()  # returns float or False
        if ring is not False:
            self.hours = ring  # convert the item to a float, if not already
        else:  # if fail, create error msg and return
            error = "     ERROR: 5200 time must be a number. Got instead \"{}\": \n".format(self.hours)
            self.error_array.append(error)
            self.allowaddrings = False
            return
        if not RingTimeChecker(self.hours).over_24():
            error = "     ERROR: 5200 time can not exceed 24.00. Got instead \"{}\": \n".format(self.hours)
            self.error_array.append(error)
            self.allowaddrings = False
            return
        if not RingTimeChecker(self.hours).less_than_zero():
            error = "     ERROR: 5200 time can not be negative. Got instead \"{}\": \n".format(self.hours)
            self.error_array.append(error)
            self.allowaddrings = False
            return
        if not RingTimeChecker(self.hours).count_decimals_place():
            error = "     ERROR 5200 time can have no more than two decimal places. Got instead \"{}\": \n" \
                .format(self.hours)
            self.error_array.append(error)
            self.allowaddrings = False
            return
        self.hours = str(self.hours)  # convert float back to string
        self.hours = Convert(self.hours).hundredths()  # make number a string with 2 decimal places
        self.exist5200 = self.hours
        self.add_5200()

    def add_bt(self):
        """ defines the addbt var and writes report """
        if self.bt == "0.0" and self.onrec_bt in ("0", "0.00", "0.0", "", 0, 0.0):
            pass
        elif self.bt != self.onrec_bt:  # compare 5200 time against 5200 from database,
            self.addbt = self.bt  # if different, the add
            fyi = "     FYI: New or updated begin tour: {}\n".format(self.bt)
            self.fyi_array.append(fyi)

    def check_bt(self):
        """ check the begin tour """
        if type(self.bt) == str and not self.bt:  # pass if value is an empty string
            self.add_bt()
            return
        ring = RingTimeChecker(self.bt).make_float()  # returns float or False
        if ring is not False:
            self.bt = ring  # convert the attribute to a float, if not already
        else:  # if fail, create error msg and return
            error = "     ERROR: BT must be a number. Got instead \"{}\": \n".format(self.bt)
            self.error_array.append(error)
            self.allowaddrings = False
            return
        if not RingTimeChecker(self.bt).over_24():
            error = "     ERROR: BT time can not exceed 24.00. Got instead \"{}\": \n".format(self.bt)
            self.error_array.append(error)
            self.allowaddrings = False
            return
        if not RingTimeChecker(self.bt).less_than_zero():
            error = "     ERROR: BT time can not be negative. Got instead \"{}\": \n".format(self.bt)
            self.error_array.append(error)
            self.allowaddrings = False
            return
        if not RingTimeChecker(self.bt).count_decimals_place():
            error = "     ERROR: BT time can have no more than two decimal places. Got instead \"{}\": \n". \
                format(self.bt)
            self.error_array.append(error)
            self.allowaddrings = False
            return
        self.bt = str(self.bt)  # convert float back to string
        self.bt = Convert(self.bt).hundredths()  # make number a string with 2 decimal places
        self.existbt = self.bt
        self.add_bt()

    def add_et(self):
        """ defines the addet var and writes report """
        if self.et in ("0", "0.00", "0.0", "", 0, 0.0) and self.existbt and self.exist5200:
            endtour = self.auto_endtour()
            self.addet = endtour  # if different, the add
            fyi = "     FYI: Automated end tour generated: {}\n".format(endtour)
            self.fyi_array.append(fyi)
            self.auto_et = True  # enables message in text report to show auto endtour was used.
        elif self.et == "0.0" and self.onrec_et in ("0", "0.00", "0.0", "", 0, 0.0):
            pass
        elif self.et != self.onrec_et:  # compare 5200 time against 5200 from database,
            self.addet = self.et  # if different, the add
            fyi = "     FYI: New or updated end tour: {}\n".format(self.et)
            self.fyi_array.append(fyi)

    def auto_endtour(self):
        """ add 50 clicks to the begin tour and 5200 time """
        if float(self.exist5200) >= 6:
            auto_et = float(self.existbt) + float(self.exist5200) + .50
        else:
            auto_et = float(self.existbt) + float(self.exist5200)
        if auto_et >= 24:
            auto_et -= 24
        return "{:.2f}".format(auto_et)

    def check_et(self):
        """ check the end tour """
        if type(self.et) == str and not self.et:  # pass if value is an empty string
            self.add_et()
            return
        ring = RingTimeChecker(self.et).make_float()  # returns float or False
        if ring is not False:
            self.et = ring  # convert the attribute to a float, if not already
        else:  # if fail, create error msg and return
            error = "     ERROR: ET must be a number. Got instead \"{}\": \n".format(self.et)
            self.error_array.append(error)
            self.allowaddrings = False
            return
        if not RingTimeChecker(self.et).over_24():
            error = "     ERROR: ET time can not exceed 24.00. Got instead \"{}\": \n".format(self.et)
            self.error_array.append(error)
            self.allowaddrings = False
            return
        if not RingTimeChecker(self.et).less_than_zero():
            error = "     ERROR: ET time can not be negative. Got instead \"{}\": \n".format(self.et)
            self.error_array.append(error)
            self.allowaddrings = False
            return
        if not RingTimeChecker(self.et).count_decimals_place():
            error = "     ERROR: ET time can have no more than two decimal places. Got instead \"{}\": \n". \
                format(self.et)
            self.error_array.append(error)
            self.allowaddrings = False
            return
        self.et = str(self.et)  # convert float back to string
        self.et = Convert(self.et).hundredths()  # make number a string with 2 decimal places
        self.add_et()

    def add_rs(self):
        """ defines the addrs var and writes report. """
        if self.rs == "0.0" and self.onrec_rs in ("0", "0.00", "0.0", "", 0, 0.0):
            pass
        elif self.rs != self.onrec_rs:  # compare 5200 time against 5200 from database,
            self.addrs = self.rs  # if different, the add
            fyi = "     FYI: New or updated return to station: {}\n".format(self.rs)
            self.fyi_array.append(fyi)

    def check_rs(self):
        """ check the return to station. """
        if type(self.rs) == str and not self.rs:  # pass if value is an empty string
            self.add_rs()
            return
        ring = RingTimeChecker(self.rs).make_float()  # returns float or False
        if ring is not False:
            self.rs = ring  # convert the attribute to a float, if not already
        else:  # if fail, create error msg and return
            error = "     ERROR: RS must be a number. Got instead \"{}\": \n".format(self.rs)
            self.error_array.append(error)
            self.allowaddrings = False
            return
        if not RingTimeChecker(self.rs).over_24():
            error = "     ERROR: RS time can not exceed 24.00. Got instead \"{}\": \n".format(self.rs)
            self.error_array.append(error)
            self.allowaddrings = False
            return
        if not RingTimeChecker(self.rs).less_than_zero():
            error = "     ERROR: RS time can not be negative. Got instead \"{}\": \n".format(self.rs)
            self.error_array.append(error)
            self.allowaddrings = False
            return
        if not RingTimeChecker(self.rs).count_decimals_place():
            error = "     ERROR: RS time can have no more than two decimal places. Got instead \"{}\": \n". \
                format(self.rs)
            self.error_array.append(error)
            self.allowaddrings = False
            return
        self.rs = str(self.rs)  # convert float back to string
        self.rs = Convert(self.rs).hundredths()  # make number a string with 2 decimal places
        self.add_rs()

    def add_moves(self, baseready):
        """ defines the addmoves variable and writes to report. """
        if baseready != self.onrec_moves:  # if the moves are different from on record moves from dbase,
            self.addmoves = baseready  # add the moves
            fyi = "     FYI: New or updated moves: {}\n".format(baseready)
            self.fyi_array.append(fyi)

    def check_moves(self):
        """ checks the moves. """
        self.moves = str(self.moves)
        self.moves = self.moves.strip()
        if type(self.moves) == str and not self.moves:
            self.add_moves("")
            return
        self.moves = self.moves.replace("+", ",").replace("/", ",").replace("//", ",") \
            .replace("-", ",").replace("*", ",")  # replace all delimiters with commas
        moves_array = Convert(self.moves).string_to_array()  # convert the moves string to an array
        if not MovesChecker(moves_array).length():  # check number of items is multiple of three
            error = "     ERROR: Moves must be given in multiples of three. Got instead \"{}\": \n" \
                .format(len(moves_array))
            self.error_array.append(error)
            self.allowaddrings = False
            return
        for i in range(len(moves_array)):
            if self.triad_elem_istime(i):  # method returns True is the element is in a 'time' place
                move_ring = RingTimeChecker(moves_array[i]).make_float()  # try to convert moves_array[i] to a float.
                if move_ring is not False:  # if fail, create error msg and return
                    moves_array[i] = move_ring  # convert the item to a float, if not already
                else:
                    error = "     ERROR: Move times must be a number. Got instead \"{}\": \n".format(moves_array[i])
                    self.error_array.append(error)
                    self.allowaddrings = False
                    return
                if not RingTimeChecker(moves_array[i]).over_24():
                    error = "     ERROR: Move time can not exceed 24.00. Got instead \"{}\": ".format(moves_array[i])
                    self.error_array.append(error)
                    self.allowaddrings = False
                    return
                if not RingTimeChecker(moves_array[i]).less_than_zero():
                    error = "     ERROR: Move time can not be negative. Got instead \"{}\": \n".format(moves_array[i])
                    self.error_array.append(error)
                    self.allowaddrings = False
                    return
                if not RingTimeChecker(moves_array[i]).count_decimals_place():
                    error = "     ERROR: Move time can have no more than two decimal places. Got instead \"{}\": \n" \
                        .format(moves_array[i])
                    self.error_array.append(error)
                    self.allowaddrings = False
                    return
            #  if (i + 1) % 3 == 0:  # check the route component of the move triad
            if not self.triad_elem_istime(i):  # method returns false is the element is in the 'route' place
                if not RouteChecker(moves_array[i]).check_numeric():
                    error = "     ERROR: Routes in move triads must be numeric. Got instead \"{}\": \n" \
                        .format(moves_array[i])
                    self.error_array.append(error)
                    self.allowaddrings = False
                    return
                if not RouteChecker(moves_array[i]).check_length():
                    error = "     ERROR: Routes in move triads must have 4 or 5 digits. Got instead \"{}\": \n" \
                        .format(moves_array[i])
                    self.error_array.append(error)
                    self.allowaddrings = False
                    return
        moves_array = self.triad_restoreorder(moves_array)  # if self.triad_routefirst is true - reorder the array
        for i in range(0, len(moves_array), 3):
            if moves_array[i] > moves_array[i + 1]:
                error = "     ERROR: first time \"{}\" must be lesser than the second \n" \
                        "            time \"{}\" in moves.\n".format(moves_array[i], moves_array[i + 1])
                self.error_array.append(error)
                self.allowaddrings = False
                return
            else:  # convert the items back into strings with 2 decimal places
                moves_array[i] = str(moves_array[i])
                moves_array[i] = Convert(moves_array[i]).hundredths()
                moves_array[i + 1] = str(moves_array[i + 1])
                moves_array[i + 1] = Convert(moves_array[i + 1]).hundredths()
        baseready = Convert(moves_array).array_to_string()  # convert the moves array to a baseready string
        self.add_moves(baseready)

    def triad_elem_istime(self, i):
        """ finds if the triad element is a time or a route - returns true for time. """
        if not self.parent.triad_routefirst:  # if the triad route first setting is off
            if i % 3 == 0 or (i + 2) % 3 == 0:  # if the first or second value
                return True
            if (i + 1) % 3 == 0:  # if the third value
                return False
        else:  # if the triad route first setting is on
            if i % 3 == 0:  # if the first value
                return False
            if (i + 2) % 3 == 0 or (i + 1) % 3 == 0:  # if the second or third value.
                return True

    def triad_restoreorder(self, moves_array):
        """ restores the triad to the 'time, time, route' order if self.parent.triad_routefirst is true. """
        if not self.parent.triad_routefirst:  # if the triad route first setting is off
            return moves_array  # skip the process
        fixed_array = []
        for i in range(0, len(moves_array), 3):
            fixed_array.append(moves_array[i + 1])
            fixed_array.append(moves_array[i + 2])
            fixed_array.append(moves_array[i])
        return fixed_array

    def add_codes(self):
        """ adds to the codes varible and writes to the report. """
        if self.codes == self.onrec_codes:  # compare 5200 time against 5200 from database,
            pass
        else:
            self.addcode = self.codes  # if different, the add
            fyi = "     FYI: New or updated code/note: {}\n".format(self.codes)
            self.fyi_array.append(fyi)

    def check_codes(self):
        """ checks the codes. """
        all_codes = ("none", "ns day", "no call", "light", "sch chg", "annual", "sick", "excused")
        self.codes = self.codes.strip()  # strip any whitespace away from the code valve
        self.codes = str(self.codes)  # convert any code value into a string datatype
        self.codes = self.codes.lower()  # convert any code value to lowercase
        if not self.codes:  # if the code is blank, then return
            self.codes = "none"
            self.add_codes()
            return
        if self.codes not in all_codes:
            error = "     ERROR: There is no such code/note. Got instead: \"{}\" \n" \
                .format(self.codes)
            self.error_array.append(error)
            self.allowaddrings = False
            return
        if self.onrec_list in ("nl", "wal"):
            if self.codes in ("no call", "light", "sch chg", "annual", "sick", "excused"):
                attn = "     ATTENTION: The code/note you entered is not consistant with the list status \n" \
                       "                for the day. Only \"none\" and \"ns day\" are useful for {} carriers. \n" \
                       "                Got instead: {}\n"\
                       "                Ignore this message if there are multiple list statuses for the week\n"\
                    .format(self.onrec_list, self.codes)  # report
                self.attn_array.append(attn)
        # deleted otdl from list below. as of version 4.003 otdl carrier are allowed the ns day code.
        if self.onrec_list in ("aux", "ptf"):
            if self.codes in ("ns day",):
                attn = "     ATTENTION: The code/note you entered is not consistant with the list status \n" \
                       "                for the day. Only \"none\", \"no call\", \"light\", \"sch chg\", \n" \
                       "                \"annual\", \"sick\", \"excused\" are useful for {} carriers. \n" \
                       "                Got instead: {}\n" \
                       "                Ignore this message if there are multiple list statuses for the week\n"\
                       .format(self.onrec_list, self.codes)
                self.attn_array.append(attn)
        self.add_codes()

    def add_lvtype(self):
        """ store the leave type if it has changed and passes checks """
        if self.lv_type == self.onrec_leave_type:  # compare 5200 time against 5200 from database,
            pass  # take no action if they are the same
        else:
            self.addlvtype = self.lv_type  # if different, the add
            fyi = "     FYI: New or updated leave type: {}\n".format(self.lv_type)
            self.fyi_array.append(fyi)

    def check_leave_type(self):
        """ check the leave type """
        all_codes = ("none", "annual", "sick", "holiday", "other", "combo")
        self.lv_type = str(self.lv_type)  # make sure lv type is a string
        self.lv_type = self.lv_type.strip()  # remove whitespace
        self.lv_type = self.lv_type.lower()  # force lv type to be lowercase
        if not self.lv_type:
            self.lv_type = "none"
            self.add_lvtype()  # store the leave type if it has changed and passes checks
            return
        if self.lv_type not in all_codes:
            error = "     ERROR: There is no such leave type. Acceptable types are: \"none\", \n" \
                    "            \"annual\", \"sick\", \"holiday\", \"other\" \n" \
                    "            Got instead: \"{}\"\n".format(self.lv_type)
            self.error_array.append(error)
            self.allowaddrings = False
            return
        self.add_lvtype()  # store the leave type if it has changed and passes checks

    def add_leave_time(self):
        """ add to the leave time variable. """
        if self.lv_time == "0.0" and self.onrec_leave_time in ("0", "0.00", "0.0", "", 0, 0.0):
            pass  # if new and old lv times are both empty, take no action
        elif self.lv_time != self.onrec_leave_time:  # compare lv type time against lv type from database,
            self.addlvtime = self.lv_time  # if different, the add
            fyi = "     FYI: New or updated leave time: {}\n".format(self.lv_time)
            self.fyi_array.append(fyi)

    def check_leave_time(self):
        """ checks the leave time. """
        if type(self.lv_time) == str and not self.lv_time:  # pass if value is an empty string
            self.add_leave_time()
            return
        ring = RingTimeChecker(self.lv_time).make_float()  # try to convert moves_array[i] to a float.
        if ring is not False:  # if fail, create error msg and return
            self.lv_time = ring  # convert the item to a float, if not already
        else:
            error = "     ERROR: Leave time must be a number. Got instead \"{}\": \n".format(self.lv_time)
            self.error_array.append(error)
            self.allowaddrings = False
            return
        if not RingTimeChecker(self.lv_time).over_8():
            error = "     ERROR: Leave time can not exceed 8.00. Got instead \"{}\": \n".format(self.lv_time)
            self.error_array.append(error)
            self.allowaddrings = False
            return
        if not RingTimeChecker(self.lv_time).less_than_zero():
            error = "     ERROR: Leave time can not be negative. Got instead \"{}\": \n".format(self.lv_time)
            self.error_array.append(error)
            self.allowaddrings = False
            return
        if not RingTimeChecker(self.lv_time).count_decimals_place():
            error = "     ERROR: Leave time can have no more than two decimal places. Got instead \"{}\": \n" \
                .format(self.lv_time)
            self.error_array.append(error)
            self.allowaddrings = False
            return
        self.lv_time = str(self.lv_time)  # make lv time back into a string
        self.lv_time = Convert(self.lv_time).hundredths()  # make lv time into a string number with 2 decimal places
        self.add_leave_time()

    def delete_recs(self):
        """ delete any pre existing record """
        if not self.parent.interject:
            fyi = "     FYI: Clock Rings record will be deleted from database\n"
            self.fyi_array.append(fyi)
            return
        sql = "DELETE FROM rings3 WHERE rings_date = '%s' and carrier_name = '%s'" % (self.adddate, self.parent.name)
        commit(sql)
        add = "     DELETE: Clock Rings record deleted from database\n"  # report
        self.add_array.append(add)

    def add_recs(self):
        """ adds the records to the database"""
        chg_these = []
        if not self.allowaddrings:
            return
        # determine conditions where existing record is deleted
        permitted_codes = ("no call", "ns day", "annual", "sick", "excused")
        if not self.hours:
            if not self.lv_time:
                if self.codes not in permitted_codes:
                    if self.onrec_date:  # if there is an existing record to delete
                        self.delete_recs()  # delete any pre existing record
                        return
        # contruct the sql command to commit to the database.
        if self.add5200 != "empty":  # 5200 place of sql command
            add = "     INPUT: 5200 time added or updated to database >>{}\n".format(self.add5200)  # report
            self.add_array.append(add)
            chg_these.append("hours")
            hours_place = self.add5200
        else:
            hours_place = self.onrec_5200
        if self.addbt != "empty":  # bt place of sql command
            add = "     INPUT: BT time added or updated to database >>{}\n".format(self.addbt)  # report
            self.add_array.append(add)
            chg_these.append("bt")
            bt_place = self.addbt
        else:
            bt_place = self.onrec_bt
        if self.addet != "empty":  # et place of sql command
            if self.auto_et:
                add = "     INPUT: ET time added to database automatically >>{}\n".format(self.addet)  # report
            else:
                add = "     INPUT: ET time added or updated to database >>{}\n".format(self.addet)  # report
            self.add_array.append(add)
            chg_these.append("et")
            et_place = self.addet
        else:
            et_place = self.onrec_et
        if self.addrs != "empty":  # rs place of sql command
            add = "     INPUT: RS time added or updated to database >>{}\n".format(self.addrs)  # report
            self.add_array.append(add)
            chg_these.append("rs")
            rs_place = self.addrs
        else:
            rs_place = self.onrec_rs
        if self.addcode != "empty":  # code place of sql command
            add = "     INPUT: Code/note added or updated to database >>{}\n".format(self.addcode)  # report
            self.add_array.append(add)
            chg_these.append("code")
            code_place = self.addcode
        else:
            code_place = self.onrec_codes
        if self.addmoves != "empty":  # moves place of sql command
            add = "     INPUT: Moves added or updated to database >>{}\n".format(self.addmoves)  # report
            self.add_array.append(add)
            chg_these.append("moves")
            moves_place = self.addmoves
        else:
            moves_place = self.onrec_moves
        if self.addlvtype != "empty":  # lv type place of sql command
            add = "     INPUT: Leave type added or updated to database >>{}\n".format(self.addlvtype)  # report
            self.add_array.append(add)
            chg_these.append("lv type")
            lv_type_place = self.addlvtype
        else:
            lv_type_place = self.onrec_leave_type
        if self.addlvtime != "empty":  # lv time place of sql command
            add = "     INPUT: Leave time added or updated to database >>{}\n".format(self.addlvtime)  # report
            self.add_array.append(add)
            chg_these.append("lv time")
            lv_time_place = self.addlvtime
        else:
            lv_time_place = self.onrec_leave_time
        # if there are items to change, construct the sql command
        if chg_these:
            if not self.onrec_date:  # if there is no rings record for the date
                sql = "INSERT INTO rings3(rings_date, carrier_name, total, rs, code, " \
                      "moves, leave_type, leave_time, bt, et) " \
                      "VALUES('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s')" \
                      % (self.adddate, self.parent.name, hours_place, rs_place, code_place, moves_place,
                         lv_type_place, lv_time_place, bt_place, et_place)
            else:  # if a record already exist
                sql = "UPDATE rings3 SET total = '%s', rs = '%s', code = '%s', moves = '%s', leave_type = '%s', " \
                      "leave_time = '%s', bt = '%s', et = '%s' WHERE rings_date = '%s' and carrier_name = '%s'" % \
                      (hours_place, rs_place, code_place, moves_place, lv_type_place, lv_time_place, bt_place,
                       et_place, self.adddate, self.parent.name)
            commit(sql)

    def generate_report(self):
        """ generate a report """
        self.parent.rings_fatal_rpt += len(self.error_array)
        self.parent.rings_add_rpt += len(self.add_array)
        self.parent.rings_fyi_rpt += len(self.fyi_array)
        if not self.parent.interject:
            master_array = self.error_array + self.attn_array  # use these reports for precheck
            if self.parent.fullreport:  # if the full report option is selected...
                master_array += self.fyi_array  # also include the fyi messages.
        else:
            master_array = self.error_array + self.attn_array  # use these reports for input
            if self.parent.fullreport:  # if the full report option is selected...
                master_array += self.add_array   # also include the adds messages.
        if len(master_array) > 0:
            if not self.parent.name_mentioned:
                self.parent.report.write("\n{}\n".format(self.parent.name))
                self.parent.name_mentioned = True
            self.parent.report.write("   >>> sheet: \"{}\" --> row: \"{}\" <<<\n".format(self.sheet, self.row))
            if not self.allowaddrings:
                self.parent.report.write("     CLOCK RINGS ENTRY PROHIBITED: Correct errors!\n")
                # self.parent.fatal_rpt += 1
            for rpt in master_array:  # write all reports that have been keep in arrays.
                self.parent.report.write(rpt)
