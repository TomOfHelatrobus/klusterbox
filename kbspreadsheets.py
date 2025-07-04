"""
a klusterbox module: Klusterbox Improper Mandates and 12 and 60 Hour Violations Spreadsheets Generator.
klusterbox classes for spreadsheets: the improper mandate worksheet and the 12 and 60 hour violations spreadsheets
"""
import projvar  # custom libraries
from kbtoolbox import inquire, CarrierList, dir_path, isfloat, Convert, Rings, ProgressBarDe, Moves, DateHandler
# standard libraries
from tkinter import messagebox
import os
import sys
import subprocess
from datetime import datetime, timedelta
# non standard libraries
from openpyxl import Workbook
from openpyxl.worksheet.pagebreak import Break
from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment, PatternFill


class ImpManSpreadsheet:
    """
    This generates the famous klusterbox spreadsheets breaking down availability and off route mandates.
    """
    def __init__(self):
        self.frame = None  # the frame of parent
        self.pb = None  # progress bar object
        self.pbi = 0  # progress bar count index
        self.daily_loop_range = 7
        if not projvar.invran_weekly_span:
            self.daily_loop_range = 1
        self.startdate = None  # start date of the investigation
        self.enddate = None  # ending date of the investigation
        self.dates = []  # all days of the investigation
        self.carrierlist = []  # all carriers in carrier list
        self.carrier_breakdown = []  # all carriers in carrier list broken down into appropiate list
        self.wb = None  # the workbook object
        self.ws_list = []  # "saturday", "sunday", "monday", "tuesday", "wednesday", "thursday", "friday"
        self.summary = None  # worksheet for summary page
        self.reference = None  # worksheet for reference page
        self.remedy = None  # worksheet for remedy page
        self.remedy_10hr = None  # worksheet for 10 hour/ letter carrier paragraph remedy page
        self.ws_header = None  # style
        self.list_header = None  # style
        self.date_dov = None  # style
        self.remedy_style = None  # style
        self.date_dov_title = None  # style
        self.col_header = None  # style
        self.input_name = None  # style
        self.input_s = None  # style
        self.calcs = None  # style
        self.instruct_text = None  # style
        self.min_ss_nl = 0  # minimum rows for "no list"
        self.min_ss_wal = 0  # minimum rows for work assignment list
        self.min_ss_otdl = 0  # minimum rows for overtime desired list
        self.min_ss_aux = 0  # minimum rows for auxiliary
        self.show_remedy = None  # setting to show the remedy tab
        self.remedy_rate = None  # setting for remedy hourly rate
        self.remedy_tolerance = None  # setting to remedy tolerance
        self.max_pivot = 0.0  # the maximum allowed pivot.
        self.day = None  # build worksheet - loop once for each day
        self.i = 0  # build worksheet loop iteration
        self.lsi = 0  # list loop iteration
        self.pref = ("nl", "wal", "otdl", "aux")
        self.ot_list = ("No List Carriers", "Work Assignment Carriers", "Overtime Desired List Carriers",
                        "Auxiliary Assistance")  # list loop iteration
        self.row = 0  # list loop iteration/ the row placement
        self.mod_carrierlist = []  # carrier list with empty recs added to reach minimum row quantity
        self.carrier = ""  # carrier name
        self.list_ = ""
        self.nsday = ""
        self.rings = []  # carrier rings queried from database
        self.totalhours = ""  # carrier rings - 5200 time
        self.codes = ""  # carrier rings - code/note
        self.rs = ""  # carrier rings - return to station
        self.moves = ""  # carrier rings - moves on and off route with route
        self.lvtype = ""  # carrier rings - leave type
        self.lvtime = ""  # carrier rings - leave time
        self.movesarray = []
        self.avail_max = 0.0  # the maximum about of availability for a carrier on a given day
        self.cum_hr_dict = {}  # a dictionary to hold cumulative hours for a specific carrier
        self.cum_ot_dict = {}  # a dictionary to hold cumulative overtime hours for a specific carrier
        self.avail_ot_dict = {}  # a dictionary that holds prior available ot for the previous day.
        self.a_max_dict = {}  # a dictionary for holding a max values of every carrier for each day.
        self.offbid_dict = {}  # a dictionary holding off bid data for every carrier for each day.
        self.odlr_indicator = []  # indicates that carrier is odlr for at least one day
        self.odln_indicator = []  # indicates that carrier is odln for at least one day
        self.move_i = 0  # increments rows for multiple move functionality
        self.tol_ot_ownroute = 0.0  # tolerance for ot on own route
        self.tol_ot_offroute = 0.0  # tolerance for ot off own route
        self.tol_availability = 0.0  # tolerance for availability
        self.pb_nl_wal = True  # page break between no list and work assignment
        self.pb_wal_otdl = True  # page break between work assignment and otdl
        self.pb_otdl_aux = True  # page break between otdl and auxiliary
        self.day_of_week = []  # seven day array for weekly investigations/ one day array for daily investigations
        self.mandates_own_route = []  # stores cell location on each sheet for no list own route overtime
        self.mandates_all = []  # stores cell location for total mandates for each sheet
        self.availability_10 = []  # stores cell location for total availability to 10 hrs for each sheet
        self.availability_max = []  # stores cell location for total maximum availabilityfor each sheet
        self.remedy_summary_row = 1  # stores the row of the remedy sheet totals for the summary
        self.remedy10_summary_row = 1  # stores the row of the 10 hr remedy sheet totals for the summary
        self.first_row = 0  # stores the first row for each list, re initialized at end of list
        self.last_row = 0  # stores the last row for each list, re initialized at end of list
        self.subtotal_loc_holder = []  # stores the cell location of a subtotal for total mandates/ availability
        self.remedy_array = []  # an array that holds: list, day, name and cell for violations
        self.remedy_row = 0  # holds the value of the row on the remedy page
        self.remedy10_row = 0  # holds the value of the row on the remedy page
        self.remedy_start_row = 0  # will hold the start row of the list which was last started. updates for each list
        self.remedy_footer_row = []
        self.remedy_equalizer_rows = []
        self.remedy_cum = []  # hold 4 tuples of two values - start and end of totals columns.
        self.remedy_incrementor = 0
        self.remedy_blank_nl = [[], [], [], [], [], [], []]
        self.remedy_blank_wal = [[], [], [], [], [], [], []]
        self.remedy_blank_otdl = [[], [], [], [], [], [], []]
        self.remedy_blank_aux = [[], [], [], [], [], [], []]
        self.order_nl_blanks = []
        self.order_wal_blanks = []
        self.order_otdl_blanks = []
        self.order_aux_blanks = []
        self.remedy10_array = []  # an array that holds: list, day, name and cell for violations
        self.remedy10_row = 0  # holds the value of the row on the remedy page
        self.remedy10_row = 0  # holds the value of the row on the remedy page
        self.remedy10_start_row = 0  # will hold the start row of the list which was last started. updates for each list
        self.remedy10_footer_row = []
        self.remedy10_equalizer_rows = []
        self.remedy10_cum = []  # hold 4 tuples of two values - start and end of totals columns.
        self.remedy10_incrementor = 0
        self.remedy10_blank_nl = [[], [], [], [], [], [], []]
        self.remedy10_blank_otdl = [[], [], [], [], [], [], []]
        self.remedy10_blank_aux = [[], [], [], [], [], [], []]
        self.order10_nl_blanks = []
        self.order10_otdl_blanks = []
        self.order10_aux_blanks = []
        self.cell_tol_ot_ownroute = 'reference!C$5'  #
        self.cell_tol_ot_offroute = 'reference!C$6'
        self.cell_tol_availability = 'reference!C$7'

    def create(self, frame):
        """ master method for calling all methods in class """
        self.frame = frame
        if not self.ask_ok():  # abort if user selects cancel from askokcancel
            return
        self.pb = ProgressBarDe(label="Building Improper Mandates Spreadsheet")
        self.pb.max_count(100)  # set length of progress bar
        self.pb.start_up()  # start the progress bar
        self.pbi = 1
        self.pb.move_count(self.pbi)  # increment progress bar
        self.pb.change_text("Gathering Data... ")
        self.get_dates()
        self.get_pb_max_count()
        self.get_carrierlist()
        self.get_carrier_breakdown()  # breakdown carrier list into no list, wal, otdl, aux
        self.get_tolerances()  # get tolerances, minimum rows and page break preferences from tolerances table
        self.get_styles()
        self.build_workbook()
        self.set_dimensions()
        self.build_ws_loop()  # calls list loop and carrier loop
        self._build_remedy_blank_arrays()  # for both 12 and 10 hour
        self._equalize_remedy_blank_arrays()  # for both 12 and 10 hour
        self._order_remedy_blank_arrays()  # for both 12 and 10 hour
        self._build_remedy_header()
        self._build_remedy()
        self._build_remedy10()
        self.build_summary()
        self.build_refs()
        self.save_open()

    def ask_ok(self):
        """ ends process if user cancels """
        if messagebox.askokcancel("Spreadsheet generator",
                                  "Do you want to generate an Improper Mandates Spreadsheet?",
                                  parent=self.frame):
            return True
        return False

    def get_dates(self):
        """ get the dates from the project variables """
        self.startdate = projvar.invran_date  # set daily investigation range as default - get start date
        self.enddate = projvar.invran_date  # get end date
        self.dates = [projvar.invran_date, ]  # create an array of days - only one day if daily investigation range
        if projvar.invran_weekly_span:  # if the investigation range is weekly
            date = projvar.invran_date_week[0]
            self.startdate = projvar.invran_date_week[0]
            self.enddate = projvar.invran_date_week[6]
            self.dates = []
            for _ in range(7):  # create an array with all the days in the weekly investigation range
                self.dates.append(date)
                date += timedelta(days=1)

    def get_pb_max_count(self):
        """ set length of progress bar """
        self.pb.max_count((len(self.dates)*4)+3)  # once for each list in each day, plus reference, summary and saving

    def get_carrierlist(self):
        """ get record sets for all carriers """
        self.carrierlist = CarrierList(self.startdate, self.enddate, projvar.invran_station).get()

    def get_carrier_breakdown(self):
        """ breakdown carrier list into no list, wal, otdl, aux """
        timely_rec = []
        for day in self.dates:
            nl_array = []
            wal_array = []
            otdl_array = []
            aux_array = []
            for carrier in self.carrierlist:
                for rec in reversed(carrier):
                    if Convert(rec[0]).dt_converter() <= day:
                        timely_rec = rec
                if timely_rec[2] == "nl":
                    nl_array.append(timely_rec)
                if timely_rec[2] == "wal":
                    wal_array.append(timely_rec)
                if timely_rec[2] == "otdl":
                    otdl_array.append(timely_rec)
                if timely_rec[2] == "odlr":  # for odl regular day only -
                    if timely_rec[1] not in self.odlr_indicator:  # add name to odlr indicator array
                        self.odlr_indicator.append(timely_rec[1])
                    # if it is sunday or their ns day, put the record in the nl array
                    if day.strftime("%a") == projvar.ns_code[timely_rec[3]] or day.strftime("%a") == "Sun":
                        nl_array.append(timely_rec)
                    else:  # if it is a sunday or their ns day, put record in no list array.
                        otdl_array.append(timely_rec)
                if timely_rec[2] == "odln":  # for odl non scheduled day only
                    if timely_rec[1] not in self.odln_indicator:  # add name to odln indicator array
                        self.odln_indicator.append(timely_rec[1])
                    # if it is sunday or their ns day, put the record in the otdl array
                    if day.strftime("%a") == projvar.ns_code[timely_rec[3]] or day.strftime("%a") == "Sun":
                        otdl_array.append(timely_rec)
                    else:
                        nl_array.append(timely_rec)
                if timely_rec[2] == "aux" or timely_rec[2] == "ptf":
                    aux_array.append(timely_rec)
            daily_breakdown = [nl_array, wal_array, otdl_array, aux_array]
            self.carrier_breakdown.append(daily_breakdown)

    def get_tolerances(self):
        """ get spreadsheet tolerances, row minimums and page break prefs from tolerance table """
        sql = "SELECT tolerance FROM tolerances"
        result = inquire(sql)
        self.tol_ot_ownroute = float(result[0][0])  # overtime on own route tolerance
        self.tol_ot_offroute = float(result[1][0])  # overtime off own route tolerance
        self.tol_availability = float(result[2][0])  # availability tolerance
        self.min_ss_nl = int(result[3][0])  # minimum rows for no list
        self.min_ss_wal = int(result[4][0])  # mimimum rows for work assignment
        self.min_ss_otdl = int(result[5][0])  # minimum rows for otdl
        self.min_ss_aux = int(result[6][0])  # minimum rows for auxiliary
        self.pb_nl_wal = Convert(result[21][0]).str_to_bool()  # page break between no list and work assignment
        self.pb_wal_otdl = Convert(result[22][0]).str_to_bool()  # page break between work assignment and otdl
        self.pb_otdl_aux = Convert(result[23][0]).str_to_bool()  # page break between otdl and auxiliary
        self.show_remedy = Convert(result[48][0]).str_to_bool()
        self.remedy_rate = Convert(result[49][0]).hundredths()
        self.remedy_tolerance = float(Convert(result[54][0]).hundredths())
        self.max_pivot = float(result[42][0])  # the maximum allowed pivot which will be displayed.

    def get_styles(self):
        """ Named styles for workbook """
        bd = Side(style='thin', color="80808080")  # defines borders
        self.ws_header = NamedStyle(name="ws_header", font=Font(bold=True, name='Arial', size=12))
        self.list_header = NamedStyle(name="list_header", font=Font(bold=True, name='Arial', size=10))
        self.date_dov = NamedStyle(name="date_dov", font=Font(name='Arial', size=8))
        self.date_dov_title = NamedStyle(name="date_dov_title", font=Font(bold=True, name='Arial', size=8),
                                         alignment=Alignment(horizontal='right'))
        self.remedy_style = NamedStyle(name="remedy_style", font=Font(name='Arial', size=8),
                                       alignment=Alignment(horizontal='left'))
        self.col_header = NamedStyle(name="col_header", font=Font(bold=True, name='Arial', size=8),
                                     alignment=Alignment(horizontal='right'))
        self.input_name = NamedStyle(name="input_name", font=Font(name='Arial', size=8),
                                     border=Border(left=bd, top=bd, right=bd, bottom=bd))
        self.input_s = NamedStyle(name="input_s", font=Font(name='Arial', size=8),
                                  border=Border(left=bd, top=bd, right=bd, bottom=bd),
                                  alignment=Alignment(horizontal='right'))
        self.calcs = NamedStyle(name="calcs", font=Font(name='Arial', size=8),
                                border=Border(left=bd, top=bd, right=bd, bottom=bd),
                                fill=PatternFill(fgColor='e5e4e2', fill_type='solid'),
                                alignment=Alignment(horizontal='right'))
        self.instruct_text = NamedStyle(name="instruct_text", font=Font(name='Arial', size=9),
                                        alignment=Alignment(horizontal='left', vertical='top'))

    def build_workbook(self):
        """ build the workbook object """
        day_finder = ["Sat", "Sun", "Mon", "Tue", "Wed", "Thu", "Fri"]
        day_of_week = ["saturday", "sunday", "monday", "tuesday", "wednesday", "thursday", "friday"]
        i = 0
        self.wb = Workbook()  # define the workbook
        if not projvar.invran_weekly_span:  # if investigation range is daily
            for ii in range(len(day_finder)):
                if projvar.invran_date.strftime("%a") == day_finder[ii]:  # find the correct day
                    i = ii
            self.ws_list.append(self.wb.active)  # create first worksheet
            self.ws_list[0].title = day_of_week[i]  # title first worksheet
            self.day_of_week.append(day_of_week[i])  # create self.day_of_week array with one day
        if projvar.invran_weekly_span:  # if investigation range is weekly
            for day in day_of_week:
                self.day_of_week.append(day)  # create self.day_of_week array with seven days
            self.ws_list.append(self.wb.active)  # create first worksheet
            self.ws_list[0].title = "saturday"  # title first worksheet
            for i in range(1, 7):  # create worksheet for remaining six days
                self.ws_list.append(self.wb.create_sheet(day_of_week[i]))  # create subsequent worksheets
                self.ws_list[i].title = day_of_week[i]  # title subsequent worksheets
        self.remedy = self.wb.create_sheet("remedy")
        self.remedy_10hr = self.wb.create_sheet("remedy_10hr")  # remedy for letter carrier paragraph
        self.summary = self.wb.create_sheet("summary")
        self.reference = self.wb.create_sheet("reference")

    def set_dimensions(self):
        """ set the dimensions of the workbook """
        for i in range(len(self.dates)):
            self.ws_list[i].oddFooter.center.text = "&A"
            self.ws_list[i].column_dimensions["A"].width = 20
            self.ws_list[i].column_dimensions["B"].width = 5
            self.ws_list[i].column_dimensions["C"].width = 6
            self.ws_list[i].column_dimensions["D"].width = 6
            self.ws_list[i].column_dimensions["E"].width = 6
            self.ws_list[i].column_dimensions["F"].width = 6
            self.ws_list[i].column_dimensions["G"].width = 6
            self.ws_list[i].column_dimensions["H"].width = 7
            self.ws_list[i].column_dimensions["I"].width = 6
            self.ws_list[i].column_dimensions["J"].width = 6
            self.ws_list[i].column_dimensions["K"].width = 7
        self.summary.column_dimensions["A"].width = 14
        self.summary.column_dimensions["B"].width = 9
        self.summary.column_dimensions["C"].width = 9
        self.summary.column_dimensions["D"].width = 9
        self.summary.column_dimensions["E"].width = 2
        self.summary.column_dimensions["F"].width = 9
        self.summary.column_dimensions["G"].width = 9
        self.summary.column_dimensions["H"].width = 9
        self.reference.column_dimensions["A"].width = 14
        self.reference.column_dimensions["B"].width = 8
        self.reference.column_dimensions["C"].width = 8
        self.reference.column_dimensions["D"].width = 2
        self.reference.column_dimensions["E"].width = 50
        self.remedy.oddFooter.center.text = "&A"
        self.remedy.column_dimensions["A"].width = 21
        self.remedy.column_dimensions["B"].width = 6
        self.remedy.column_dimensions["C"].width = 6
        self.remedy.column_dimensions["D"].width = 6
        self.remedy.column_dimensions["E"].width = 6
        self.remedy.column_dimensions["F"].width = 6
        self.remedy.column_dimensions["G"].width = 6
        self.remedy.column_dimensions["H"].width = 6
        self.remedy.column_dimensions["I"].width = 7
        self.remedy.column_dimensions["J"].width = 10
        self.remedy_10hr.oddFooter.center.text = "&A"
        self.remedy_10hr.column_dimensions["A"].width = 21
        self.remedy_10hr.column_dimensions["B"].width = 6
        self.remedy_10hr.column_dimensions["C"].width = 6
        self.remedy_10hr.column_dimensions["D"].width = 6
        self.remedy_10hr.column_dimensions["E"].width = 6
        self.remedy_10hr.column_dimensions["F"].width = 6
        self.remedy_10hr.column_dimensions["G"].width = 6
        self.remedy_10hr.column_dimensions["H"].width = 6
        self.remedy_10hr.column_dimensions["I"].width = 7
        self.remedy_10hr.column_dimensions["J"].width = 10
        
    def build_ws_loop(self):
        """ this loops once for each list. """
        self.i = 0
        for day in self.dates:
            self.day = day
            self.build_ws_headers()
            self.list_loop()  # loops four times. once for each list.
            self.i += 1

    def build_ws_headers(self):
        """ worksheet headers """
        cell = self.ws_list[self.i].cell(row=1, column=1)
        cell.value = "Improper Mandate Worksheet"
        cell.style = self.ws_header
        self.ws_list[self.i].merge_cells('A1:E1')
        cell = self.ws_list[self.i].cell(row=3, column=1)
        cell.value = "Date:  "  # create date/ pay period/ station header
        cell.style = self.date_dov_title
        cell = self.ws_list[self.i].cell(row=3, column=2)
        cell.value = format(self.day, "%A  %m/%d/%y")
        cell.style = self.date_dov
        self.ws_list[self.i].merge_cells('B3:D3')
        cell = self.ws_list[self.i].cell(row=3, column=5)
        cell.value = "Pay Period:  "
        cell.style = self.date_dov_title
        self.ws_list[self.i].merge_cells('E3:F3')
        cell = self.ws_list[self.i].cell(row=3, column=7)
        cell.value = projvar.pay_period
        cell.style = self.date_dov
        self.ws_list[self.i].merge_cells('G3:H3')
        cell = self.ws_list[self.i].cell(row=4, column=1)
        cell.value = "Station:  "
        cell.style = self.date_dov_title
        cell = self.ws_list[self.i].cell(row=4, column=2)
        cell.value = projvar.invran_station
        cell.style = self.date_dov
        self.ws_list[self.i].merge_cells('B4:D4')

    def increment_progbar(self):
        """ move the progress bar, update with info on what is being done """
        lst = ("No List", "Work Assignment", "Overtime Desired", "Auxiliary")
        self.pbi += 1
        self.pb.move_count(self.pbi)  # increment progress bar
        self.pb.change_text("Building day {}: list: {}".format(self.day.strftime("%A"), lst[self.lsi]))

    def list_loop(self):
        """ loops four times. once for each list. """
        self.lsi = 0  # iterations of the list loop method
        self.row = 6
        for _ in self.ot_list:  # loops for nl, wal, otdl and aux
            self.list_and_column_headers()  # builds headers for list and column
            self.carrierlist_mod()  # add empty carrier records to carrier list until quantity matches minrows pref
            self.get_first_row()
            self.carrierloop()  # loop for each carrier
            self.build_footer()  # insert the footer
            self.pagebreak()  # insert a pagebreak if consistent with settings
            self.increment_progbar()  # increment progress bar
            self.lsi += 1

    def list_and_column_headers(self):
        """ builds headers for list and column """
        cell = self.ws_list[self.i].cell(row=self.row, column=1)
        cell.value = self.ot_list[self.lsi]  # "No List Carriers",
        cell.style = self.list_header
        if self.pref[self.lsi] in ("nl", "wal"):
            self.row += 1
        else:
            self.row += 2
        cell = self.ws_list[self.i].cell(row=self.row, column=1)  # column headers for any list
        cell.value = "Name"
        cell.style = self.col_header
        cell = self.ws_list[self.i].cell(row=self.row, column=2)
        cell.value = "note"
        cell.style = self.col_header
        cell = self.ws_list[self.i].cell(row=self.row, column=3)
        cell.value = "5200"
        cell.style = self.col_header
        cell = self.ws_list[self.i].cell(row=self.row, column=4)
        cell.value = "RS"
        cell.style = self.col_header
        if self.pref[self.lsi] in ("nl", "wal"):
            self.column_header_non()  # column headers specific for non otdl
        else:
            self.column_header_ot()  # column headers specific for otdl or aux

    def column_header_non(self):
        """ column headers specific for non otdl """
        cell = self.ws_list[self.i].cell(row=self.row, column=5)
        cell.value = "MV off"
        cell.style = self.col_header
        cell = self.ws_list[self.i].cell(row=self.row, column=6)
        cell.value = "MV on"
        cell.style = self.col_header
        cell = self.ws_list[self.i].cell(row=self.row, column=7)
        cell.value = "Route"
        cell.style = self.col_header
        cell = self.ws_list[self.i].cell(row=self.row, column=8)
        cell.value = "MV total"
        cell.style = self.col_header
        cell = self.ws_list[self.i].cell(row=self.row, column=9)
        cell.value = "OT own"
        cell.style = self.col_header
        cell = self.ws_list[self.i].cell(row=self.row, column=10)
        cell.value = "off rt"
        cell.style = self.col_header
        cell = self.ws_list[self.i].cell(row=self.row, column=11)
        cell.value = "OT off"
        cell.style = self.col_header
        self.row += 1

    def column_header_ot(self):
        """ column headers specific for otdl or aux """
        cell = self.ws_list[self.i].cell(row=self.row - 1, column=5)
        cell.value = "Availability to:"
        cell.style = self.col_header
        to_what = "to 11.5"
        if self.pref[self.lsi] == "otdl":
            to_what = "to 12"
        cell = self.ws_list[self.i].cell(row=self.row, column=5)
        cell.value = "to 10"
        cell.style = self.col_header
        cell = self.ws_list[self.i].cell(row=self.row, column=6)
        cell.value = to_what
        cell.style = self.col_header
        self.row += 1

    def carrierlist_mod(self):
        """ add empty carrier records to carrier list until quantity matches minrows preference """
        self.mod_carrierlist = self.carrier_breakdown[self.i][self.lsi]
        if self.pref[self.lsi] in ("nl",):  # if "no list"
            minrows = self.min_ss_nl
        elif self.pref[self.lsi] in ("wal",):  # if "work assignment list"
            minrows = self.min_ss_wal
        elif self.pref[self.lsi] in ("otdl",):  # if "overtime desired list"
            minrows = self.min_ss_otdl
        else:  # if "auxiliary"
            minrows = self.min_ss_aux
        while len(self.mod_carrierlist) < minrows:  # until carrier list quantity matches minrows
            add_this = ('', '', '', '', '', '')
            self.mod_carrierlist.append(add_this)  # append empty recs to carrier list

    def get_first_row(self):
        """ record the number of the first row for totals formulas in footers """
        self.first_row = self.row

    def _build_remedy_array(self):
        """ builds the remedy_array which holds the list, day, name and cell coordinates as a
        tuple (sheet, column, row) """
        _list = {0: "nl", 1: "wal", 2: "otdl", 3: "aux"}
        day = format(self.day, "%a").lower()  # format datetime to abbreviated day of week e.g. "sat", "sun", "mon"
        column = "K"  # column for nl and wal is K
        if _list[self.lsi] in ("otdl", "aux"):
            column = "F"  # column for otdl and aux is F
        coordinates = (self.day_of_week[self.i], column, self.row)
        add_this = (_list[self.lsi], day, self.carrier, coordinates)
        self.remedy_array.append(add_this)

    def _build_remedy10_array(self):
        """ builds the remedy10_array which holds the list, day, name and cell coordinates as a
        tuple (sheet, column, row) """
        _list = {0: "nl", 1: "wal", 2: "otdl", 3: "aux"}
        day = format(self.day, "%a").lower()  # format datetime to abbreviated day of week e.g. "sat", "sun", "mon"
        column = "I"  # column for nl is I (work assignment can not have a violation)
        if _list[self.lsi] in ("otdl", "aux"):
            column = "E"  # column for otdl and aux is E
        coordinates = (self.day_of_week[self.i], column, self.row)
        add_this = (_list[self.lsi], day, self.carrier, coordinates)
        self.remedy10_array.append(add_this)

    def carrierloop(self):
        """ loop for each carrier """
        for carrier in self.mod_carrierlist:
            self.get_last_row()  # record the number of the last row for total formulas in footers
            self.carrier = carrier[1]  # current iteration of carrier list is assigned self.carrier
            self.list_ = carrier[2]  # get the list status of the carrier
            self.nsday = carrier[3]
            # fill empty with a designator which starts with 'zzz!_' and a number
            self.get_rings()  # get individual carrier rings for the day
            if carrier[1]:  # if the carrier data set is not empty (for blank rows)
                self.build_availability_dict()
                self.calc_max_availability()
                self.offbid_dict[self.carrier].append(self.find_offbid())  # add true or false to the offbid dict
            self.display_recs()  # put the carrier and the first part of rings into the spreadsheet
            if self.pref[self.lsi] in ("nl", "wal"):  # if the list is no list or work assignment
                self.get_movesarray()  # get the moves
                self.display_moves()  # display the moves
                self.display_formulas_non()  # display the formulas for nl/wal
            else:  # if otdl or aux
                self.display_formulas_ot()  # display formulas for otdl/aux
            self._build_remedy_array()  # add to remedy array to hold coordinates for remedy sheet
            self._build_remedy10_array()  # add to remedy 10 array to hold coordinates for 10 hr remedy sheet
            self.increment_rows()

    def get_last_row(self):
        """ record the number of the last row for totals formulas in footers """
        self.last_row = self.row

    def increment_rows(self):
        """ increment the rows counter """
        self.row += 1
        self.row += self.move_i  # add 1 plus any the added rows from multiple moves
        self.move_i = 0  # reset the row incrementor for multiple move functionality

    def get_rings(self):
        """ get individual carrier rings for the day """
        self.rings = Rings(self.carrier, self.dates[self.i]).get_for_day()  # assign as self.rings
        self.totalhours = ""  # set default as an empty string
        self.rs = ""
        self.codes = ""
        self.moves = ""
        self.lvtype = ""
        self.lvtime = ""
        if self.rings[0]:  # if rings record is not blank
            self.totalhours = self.rings[0][2]
            self.rs = self.rings[0][3]
            self.codes = self.rings[0][4]
            self.moves = self.rings[0][5]
            self.lvtype = self.rings[0][6]
            self.lvtime = self.rings[0][7]

    def build_availability_dict(self):
        """ add the carrier's name to the availability dictionaries on the first loop of days """
        if self.i == 0:
            self.cum_hr_dict[self.carrier] = 0.0
            self.cum_ot_dict[self.carrier] = 0.0
            self.avail_ot_dict[self.carrier] = 20.0
            self.a_max_dict[self.carrier] = []
            self.offbid_dict[self.carrier] = []

    def calc_max_availability(self):
        """ get the maximum availability for the day for the given carrier
        this takes into account: weekly hours to 60, weekly ot hours to 20, daily limit to 12 or 11.50, leave,
        ns day """
        totalhours = Convert(self.totalhours).str_to_float()
        lv_time = Convert(self.lvtime).str_to_float()
        # cumulative hours for the week
        cum_hr = (lv_time + totalhours) + float(self.cum_hr_dict[self.carrier])  # cumulative ot hours for the week
        cum_ot = max(totalhours - 8, 0) + float(self.cum_ot_dict[self.carrier])
        if self.codes in ("no call", "ns day"):  # if it is the carrier's ns day
            cum_ot = totalhours + float(self.cum_ot_dict[self.carrier])
        if self.codes == "ns day":  # if ns day, then full day is added to cumulative ot.
            cum_ot = totalhours + float(self.cum_ot_dict[self.carrier])
        avail_wkly = max(60 - cum_hr, 0)  # the weekly availability is 60 - weekly cumulative
        avail_ot = max(20 - cum_ot, 0)  # the weekly ot availability is 20 - weekly ot cumulative
        avail_daily = max(11.50 - totalhours, 0)  # daily availability is 11.50 minus daily work hours
        if self.list_ == "otdl":  # except if the carrier is on the otdl
            avail_daily = max(12 - totalhours, 0)  # then daily availability is 12 minus daily work hours
        if self.list_ == "odln":
            if self.day.strftime("%a") == projvar.ns_code[self.nsday] or self.day.strftime("%a") == "Sun":
                avail_daily = max(8 - totalhours, 0)  # then daily availability is 8 minus daily work hours
        avail_leave = 12  # availability is zeroed out if the carrier takes leave
        if self.lvtype not in ("", "none"):  # zero out if lvtype is empty or 'none'
            avail_leave = 0
        prior_avail_ot = 20  # this is the available ot from the prior day, default is 20
        if self.i != 0:  # if this is not the first day
            prior_avail_ot = self.avail_ot_dict[self.carrier]  # get the value from the dictionary
        avail_ns = avail_ot  # this code will zero out availability if the carrier can not work 8 hours on an ns day.
        # if it is the ns day and 8 hours are not available
        if self.codes in ("ns day", "no call") and prior_avail_ot < 8:
            avail_ns = 0  # zero out availability
        avail_codes = avail_ot
        if self.codes in ("light", "excused", "sch chg", "annual", "sick"):  # if carrier excused for day
            avail_codes = 0  # if any of the listed codes are in self.codes - zero availability
        # select the lowest value from all criteria.
        self.avail_max = min(avail_wkly, avail_ot, avail_daily, avail_leave, avail_ns, avail_codes)
        self.update_availability_dict(cum_hr, cum_ot, avail_ot)

    def update_availability_dict(self, cum_hr, cum_ot, avail_ot):
        """ update the 3 availability dictionaries used to find max availablity , takes 3 arguments """
        self.cum_hr_dict.update({self.carrier: cum_hr})
        self.cum_ot_dict.update({self.carrier: cum_ot})
        self.avail_ot_dict.update({self.carrier: avail_ot})
        avail_max = Convert(self.avail_max).hundredths_float()  # convert the avail max to a float with 2 decimal places
        self.a_max_dict[self.carrier].append(avail_max)

    def find_offbid(self):
        """ do calculations to determine off route, on route and violation values.
        returns True if there is a violation. Adds violation boolean to self.offbid_dict[self.carrier]. """
        offroute = 0.0  # this is the total time spent off the carrier's route
        if not self.totalhours:  # if the total hours is zero - the violation is zero
            return False
        if self.codes == "ns day":  # if it is the carrier's ns day - violation is zero
            return False
        if not self.moves:  # if the moves is empty, then the violation is zero
            return False
        index = 0  # set the index to 1. This will point to an element in the moves array.
        moves = Convert(self.moves).string_to_array()  # simplify the variable name
        while index < len(moves):  # calculate the total time off route
            offroute += float(moves[index+1]) - float(moves[index])
            index += 3
        ownroute = max(float(self.totalhours) - offroute, 0)   # calculate the total time spent on route
        violation = max(8 - ownroute, 0)  # calculate the total violation
        if violation > self.max_pivot:
            return True
        return False

    def display_recs(self):
        """ put the carrier and the first part of rings into the spreadsheet """
        cell = self.ws_list[self.i].cell(row=self.row, column=1)  # name
        cell.value = self.carrier
        if self.list_ in ("odlr", "odln"):
            cell.value = self.carrier + " (" + self.list_ + ")"
        cell.style = self.input_name
        cell = self.ws_list[self.i].cell(row=self.row, column=2)  # code
        cell.value = Convert(self.codes).empty_not_none()
        cell.style = self.input_s
        cell = self.ws_list[self.i].cell(row=self.row, column=3)  # 5200
        cell.value = Convert(self.totalhours).str_to_floatoremptystr()
        cell.style = self.input_s
        cell.number_format = "#,###.00;[RED]-#,###.00"
        cell = self.ws_list[self.i].cell(row=self.row, column=4)  # return to station
        cell.value = Convert(self.rs).str_to_floatoremptystr()
        cell.style = self.input_s
        cell.number_format = "#,###.00;[RED]-#,###.00"

    def get_movesarray(self):
        """ builds sets of moves for each triad """
        multiple_sets = False  # is there more than one triad?
        self.movesarray = []  # re initialized - a list of tuples of move sets
        moves_array = []  # initialized - the moves string converted into an array
        move_off = ""  # if empty set, use default values
        move_back = ""
        move_route = ""
        formula = "=SUM(%s!F%s - %s!E%s)" % (self.day_of_week[self.i], self.row, self.day_of_week[self.i], self.row)
        if not self.moves:  # if string is empty
            pass  # use default values
        else:  # if the string is not empty
            moves_array = Convert(self.moves).string_to_array()
            if len(moves_array)/3 == 1:  # if there is only one set of moves
                move_off = moves_array[0]
                move_back = moves_array[1]
                move_route = moves_array[2]
            else:  # if there are multiple move sets
                multiple_sets = True
                move_off = "*"
                move_back = "*"
                move_route = "*"
                formula = "=SUM(%s!H%s:H%s)" % \
                          (self.day_of_week[self.i], self.row + 1, int(self.row + len(moves_array) / 3))
        add_this = (move_off, move_back, move_route, formula)
        self.movesarray.append(add_this)
        if multiple_sets:  # if multiple sets are detected
            i = 0
            formula_row_i = 1  # increment the row in the formula
            for move in moves_array:
                if (i + 3) % 3 == 0:
                    move_off = move
                if (i + 2) % 3 == 0:
                    move_back = move
                if (i + 1) % 3 == 0:
                    move_route = move
                    formula = "=SUM(%s!F%s - %s!E%s)" % (self.day_of_week[self.i], self.row + formula_row_i,
                                                         self.day_of_week[self.i], self.row + formula_row_i)
                    add_this = (move_off, move_back, move_route, formula)
                    self.movesarray.append(add_this)
                    formula_row_i += 1  # increment the row in the formula after each moves_set
                i += 1  # increment i

    def display_moves(self):
        """ fill the mv off, mv on and route columns. """
        for move_set in self.movesarray:
            for move_cell in range(4):
                move = move_set[move_cell]
                cell = self.ws_list[self.i].cell(row=self.row + self.move_i, column=5 + move_cell)
                if move_cell in (0, 1):  # format move times as floats or empty strings
                    cell.value = Convert(move).empty_not_zerofloat()  # insert an iteration of self.movesarray
                else:  # do not alter route or formula elements of move sets
                    cell.value = move  # insert an iteration of self.movesarray
                cell.style = self.input_s  # assign worksheet style for MV off, MV on and Route
                if move_cell == 3:
                    cell.style = self.calcs  # use alternate style for Moves Total
                if move_cell != 2:  # do not apply to routes column
                    cell.number_format = "#,###.00;[RED]-#,###.00"
            self.move_i += 1
        self.move_i -= 1  # correction

    def display_formulas_non(self):
        """ fill the formulas columns for non list carriers. """
        # ot_formula = "=IF(MAX(MAX(%s!C%s-8,0)-%s!K%s)>reference!C$3, MAX(MAX(%s!C%s-8,0)-%s!K%s),0)" \
        #              % (self.day_of_week[self.i], str(self.row), self.day_of_week[self.i], str(self.row),
        #                 self.day_of_week[self.i], str(self.row), self.day_of_week[self.i], str(self.row))
        ot_formula = "=IF(MAX(MAX(%s!C%s-8,0)-%s!K%s)>%s, MAX(MAX(%s!C%s-8,0)-%s!K%s),0)" \
                     % (self.day_of_week[self.i], str(self.row), self.day_of_week[self.i], str(self.row),
                        self.cell_tol_ot_ownroute,
                        self.day_of_week[self.i], str(self.row), self.day_of_week[self.i], str(self.row))
        off_rt_formula = "=%s!H%s" % (self.day_of_week[self.i], str(self.row))  # copy data from column H/ MV total
        ot_off_rt_formula = "=IF(%s!C%s=\"\",0, " \
                            "IF(OR(%s!B%s=\"ns day\",%s!J%s>=%s!C%s),%s!C%s, " \
                            "IF(%s!C%s<=8+%s,0, " \
                            "MIN(MAX(%s!C%s-8,0), " \
                            "IF(%s!J%s<=%s,0,%s!J%s)))))" \
                            % (self.day_of_week[self.i], str(self.row), self.day_of_week[self.i], str(self.row),
                               self.day_of_week[self.i], str(self.row), self.day_of_week[self.i], str(self.row),
                               self.day_of_week[self.i], str(self.row), self.day_of_week[self.i], str(self.row),
                               self.cell_tol_ot_offroute,
                               self.day_of_week[self.i], str(self.row), self.day_of_week[self.i], str(self.row),
                               self.cell_tol_ot_offroute,
                               self.day_of_week[self.i], str(self.row))
        formulas = (ot_formula, off_rt_formula, ot_off_rt_formula)
        column_i = 0
        for formula in formulas:
            cell = self.ws_list[self.i].cell(row=self.row, column=9 + column_i)
            cell.value = formula  # insert an iteration of formulas
            cell.style = self.calcs  # assign worksheet style
            cell.number_format = "#,###.00;[RED]-#,###.00"
            column_i += 1

    def display_formulas_ot(self):
        """ fill the formula carrier for otdl carriers. """
        max_hrs = 12  # maximum hours for otdl carriers
        ten_hrs = 10
        if self.list_ == "aux":  # alter formula by list preference
            max_hrs = 11.5  # maximux hours for auxiliary carriers
        if self.list_ == "odln":  # alter formula by list preference
            max_hrs = 8  # maximux hours for auxiliary carriers
            ten_hrs = 8
        formula_ten = "=IF(OR(%s!B%s = \"light\",%s!B%s = \"excused\", %s!B%s = \"sch chg\", %s!B%s = \"annual\", " \
                      "%s!B%s = \"sick\", %s!C%s >= %s - %s), 0, IF(%s!B%s = \"no call\", %s, " \
                      "IF(%s!C%s = 0, 0, MAX(%s - %s!C%s, 0))))" % \
                      (self.day_of_week[self.i], str(self.row), self.day_of_week[self.i], str(self.row),
                       self.day_of_week[self.i], str(self.row), self.day_of_week[self.i], str(self.row),
                       self.day_of_week[self.i], str(self.row), self.day_of_week[self.i], str(self.row),
                       ten_hrs, self.cell_tol_availability,
                       self.day_of_week[self.i], str(self.row),
                       ten_hrs, self.day_of_week[self.i], str(self.row),
                       ten_hrs, self.day_of_week[self.i], str(self.row))
        formula_max = "=IF(OR(%s!B%s = \"light\",%s!B%s = \"excused\", %s!B%s = \"sch chg\", %s!B%s = \"annual\", " \
                      "%s!B%s = \"sick\", %s!C%s >= %s - %s), 0, IF(%s!B%s = \"no call\", %s, " \
                      "IF(%s!C%s = 0, 0, MAX(%s - %s!C%s, 0))))" % \
                      (self.day_of_week[self.i], str(self.row), self.day_of_week[self.i], str(self.row),
                       self.day_of_week[self.i], str(self.row), self.day_of_week[self.i], str(self.row),
                       self.day_of_week[self.i], str(self.row), self.day_of_week[self.i], str(self.row),
                       max_hrs, self.cell_tol_availability,
                       self.day_of_week[self.i], str(self.row),
                       max_hrs, self.day_of_week[self.i], str(self.row),
                       max_hrs, self.day_of_week[self.i], str(self.row))
        formulas = (formula_ten, formula_max)
        column_i = 0
        for formula in formulas:
            cell = self.ws_list[self.i].cell(row=self.row, column=5 + column_i)
            cell.value = formula  # insert an iteration formulas
            cell.style = self.calcs  # assign worksheet style
            cell.number_format = "#,###.00;[RED]-#,###.00"
            column_i += 1

    def build_footer(self):
        """ call the footer depending on the list. """
        if self.pref[self.lsi] == "nl":
            self.nl_footer()
        elif self.pref[self.lsi] == "wal":
            self.wal_footer()
        elif self.pref[self.lsi] == "otdl":
            self.otdl_footer()
        else:
            self.aux_footer()
            
    def nl_footer(self):
        """ build the non list footer. """
        self.row += 1
        cell = self.ws_list[self.i].cell(row=self.row, column=8)  # totals for no list overtime
        cell.value = "Total NL Overtime"
        cell.style = self.col_header
        cell = self.ws_list[self.i].cell(row=self.row, column=9)  # OT
        formula = "=SUM(%s!I%s:I%s)" % (self.day_of_week[self.i], self.first_row, self.last_row)
        cell.value = formula
        cell.style = self.calcs
        cell.number_format = "#,###.00;[RED]-#,###.00"
        location_nl_totals = (self.day_of_week[self.i], "I", self.row)  # save location for totals after wal
        self.mandates_own_route.append(location_nl_totals)  # collect totals for summary
        self.row += 2
        cell = self.ws_list[self.i].cell(row=self.row, column=10)  # totals for no list mandates
        cell.value = "Total NL Mandates"
        cell.style = self.col_header
        cell = self.ws_list[self.i].cell(row=self.row, column=11)  # OT off route
        formula = "=SUM(%s!K%s:K%s)" % (self.day_of_week[self.i], self.first_row, self.last_row)
        cell.value = formula
        cell.style = self.calcs
        cell.number_format = "#,###.00;[RED]-#,###.00"
        self.subtotal_loc_holder.append(self.row)  # collect subtotal location for total after wal
        self.row += 1

    def wal_footer(self):
        """ build the work assignment list footer. """
        self.row += 1
        cell = self.ws_list[self.i].cell(row=self.row, column=10)
        cell.value = "Total WAL Mandates"
        cell.style = self.col_header
        formula = "=SUM(%s!K%s:K%s)" % (self.day_of_week[self.i], self.first_row, self.last_row)
        cell = self.ws_list[self.i].cell(row=self.row, column=11)
        cell.value = formula  # OT off route
        cell.style = self.calcs
        cell.number_format = "#,###.00;[RED]-#,###.00"
        self.subtotal_loc_holder.append(self.row)  # collect subtotal location for total after wal
        self.row += 2
        formula = "=SUM(%s!K%s + %s!K%s)" % (self.day_of_week[self.i], self.subtotal_loc_holder[0],
                                             self.day_of_week[self.i], self.subtotal_loc_holder[1])
        cell = self.ws_list[self.i].cell(row=self.row, column=10)
        cell.value = "Total Mandates"
        cell.style = self.col_header
        cell = self.ws_list[self.i].cell(row=self.row, column=11)
        cell.value = formula  # total ot off route for nl and wal
        cell.style = self.calcs
        cell.number_format = "#,###.00;[RED]-#,###.00"
        add_this = (self.day_of_week[self.i], "K", self.row)
        self.mandates_all.append(add_this)
        self.subtotal_loc_holder = []  # empty out the subtotal location holder for future use with otdl/aux
        self.row += 1

    def otdl_footer(self):
        """ build the ot desired list footer. """
        self.row += 1
        cell = self.ws_list[self.i].cell(row=self.row, column=4)  # header
        cell.value = "Total OTDL Availability"
        cell.style = self.col_header
        cell = self.ws_list[self.i].cell(row=self.row, column=5)  # availability to 10
        formula = "=SUM(%s!E%s:E%s)" % (self.day_of_week[self.i], self.first_row, self.last_row)
        cell.value = formula
        cell.style = self.calcs
        cell.number_format = "#,###.00;[RED]-#,###.00"
        cell = self.ws_list[self.i].cell(row=self.row, column=6)  # availability to 12
        formula = "=SUM(%s!F%s:F%s)" % (self.day_of_week[self.i], self.first_row, self.last_row)
        cell.value = formula
        cell.style = self.calcs
        cell.number_format = "#,###.00;[RED]-#,###.00"
        self.subtotal_loc_holder.append(self.row)  # collect subtotal location for total after aux
        self.row += 1

    def aux_footer(self):
        """ build the auxiliary list footer. """
        self.row += 1
        cell = self.ws_list[self.i].cell(row=self.row, column=4)
        cell.value = "Total AUX Availability"
        cell.style = self.col_header
        formula = "=SUM(%s!E%s:E%s)" % (self.day_of_week[self.i], self.first_row, self.last_row)
        cell = self.ws_list[self.i].cell(row=self.row, column=5)
        cell.value = formula  # availability to 10
        cell.style = self.calcs
        cell.number_format = "#,###.00;[RED]-#,###.00"
        formula = "=SUM(%s!F%s:F%s)" % (self.day_of_week[self.i], self.first_row, self.last_row)
        cell = self.ws_list[self.i].cell(row=self.row, column=6)
        cell.value = formula  # availability to 11.5
        cell.style = self.calcs
        cell.number_format = "#,###.00;[RED]-#,###.00"
        self.subtotal_loc_holder.append(self.row)  # collect subtotal location for total after aux
        self.row += 2
        cell = self.ws_list[self.i].cell(row=self.row, column=4)
        cell.value = "Total Availability"
        cell.style = self.col_header
        cell = self.ws_list[self.i].cell(row=self.row, column=5)  # availability for otdl and aux
        formula = "=SUM(%s!E%s + %s!E%s)" % (self.day_of_week[self.i], self.subtotal_loc_holder[0],
                                             self.day_of_week[self.i], self.subtotal_loc_holder[1])
        cell.value = formula
        cell.style = self.calcs
        cell.number_format = "#,###.00;[RED]-#,###.00"
        cell = self.ws_list[self.i].cell(row=self.row, column=6)  # availability to for otdl
        formula = "=%s!F%s" % (self.day_of_week[self.i], self.subtotal_loc_holder[0])
        cell.value = formula
        cell.style = self.calcs
        cell.number_format = "#,###.00;[RED]-#,###.00"
        add_this = (self.day_of_week[self.i], "E", self.row)  # location of total availability to 10
        self.availability_10.append(add_this)  # collect location of totals for summary, put in array
        add_this = (self.day_of_week[self.i], "F", self.row)  # location of total availability to max
        self.availability_max.append(add_this)  # collect location of totals for summary, put in array
        self.subtotal_loc_holder = []  # empty out the subtotal location holder for future use with otdl/aux
        self.row += 1

    def pagebreak(self):
        """ create a page break if consistant with user preferences """
        if self.pref[self.lsi] == "nl" and not self.pb_nl_wal:
            self.row += 1
            return
        if self.pref[self.lsi] == "wal" and not self.pb_wal_otdl:
            self.row += 1
            return
        if self.pref[self.lsi] == "otdl" and not self.pb_otdl_aux:
            self.row += 1
            return
        if self.pref[self.lsi] == "aux":
            self.row += 1
            return
        try:
            self.ws_list[self.i].page_breaks.append(Break(id=self.row))  # effective for mac
        except AttributeError:
            self.ws_list[self.i].row_breaks.append(Break(id=self.row))  # effective for windows
        self.row += 1
            
    def _build_remedy_header(self):
        """ remedy headers for 12 hour and 10 hour remedies"""
        sheet = (self.remedy, self.remedy_10hr)
        title = ("Remedies for Improperly Mandated Overtime OFF Own Route",
                 "Remedies for Improperly Mandated Overtime ON Own Route")
        text = ("This sheet shows potential violations for off route overtime improperly worked by no list "
                "and work assignment carriers while otdl carriers had not been worked to the full extent. "
                "Management must utilize otdl carriers to the fullest extent (12 hours for otdl carriers and "
                "11.50 hours for auxiliary carriers) before mandating no list or work assignment carriers to "
                "work overtime off their assignments.  Management can use Auxiliary carriers to provide "
                "assistance instead of otdl carriers. Equalize the potential violations to find the remedy. "
                "Availability is adjusted for weekly overtime and total hour limits for weekly (but not daily) "
                "investigations. Off bid violations are not included as potential violations. Availability hours "
                "for otdl and auxiliary carriers overlaps for both 10 and 12 hour availability. ",
                "This sheet shows potential violations for overtime improperly worked by no list carriers on "
                "their own assignment while otdl and auxiliary carriers had not been worked to the full extent "
                "(10 hours for both otdl and auxiliary carriers). Management must utilize otdl and auxiliary "
                "carriers to the fullest extent before mandating no list carriers to work overtime on their "
                "assignments per the Memorandum on the Letter Carrier Paragraph. Equalize the potential violations "
                "to find the remedy. Availability is adjusted for weekly overtime and total hour limits for "
                "weekly (but not daily) investigations. Off bid violations are not included as potential violations. "
                "Availability hours for otdl and auxiliary carriers overlaps for both 10 and 12 hour availability. "
                )
        for i in range(len(sheet)):
            cell = sheet[i].cell(row=1, column=1)
            cell.value = title[i]
            cell.style = self.ws_header
            sheet[i].merge_cells('A1:J1')
            cell = sheet[i].cell(row=3, column=1)
            cell.value = "Date:  "  # create date/ pay period/ station header
            cell.style = self.date_dov_title
            cell = sheet[i].cell(row=3, column=2)
            cell.value = self.dates[0].strftime("%x")
            if projvar.invran_weekly_span:
                cell.value = self.dates[0].strftime("%x") + " - " + self.dates[6].strftime("%x")
            cell.style = self.date_dov
            sheet[i].merge_cells('B3:D3')
            cell = sheet[i].cell(row=3, column=6)
            cell.value = "Pay Period:  "
            cell.style = self.date_dov_title
            sheet[i].merge_cells('F3:G3')
            cell = sheet[i].cell(row=3, column=8)
            cell.value = projvar.pay_period
            cell.style = self.date_dov
            sheet[i].merge_cells('H3:I3')
            cell = sheet[i].cell(row=4, column=1)
            cell.value = "Station:  "
            cell.style = self.date_dov_title
            cell = sheet[i].cell(row=4, column=2)
            cell.value = projvar.invran_station
            cell.style = self.date_dov
            sheet[i].merge_cells('B4:D4')
            cell = sheet[i].cell(row=4, column=5)
            cell.value = "Remedy Tolerance:  "  # label for the remedy tolerance
            cell.style = self.date_dov_title
            sheet[i].merge_cells('E4:G4')
            cell = sheet[i].cell(row=4, column=8)  # the $ value of the remedy
            cell.value = self.remedy_tolerance
            cell.style = self.remedy_style
            cell.number_format = "#,###.00;[RED]-#,###.00"
            if self.show_remedy:
                cell = sheet[i].cell(row=5, column=6)
                cell.value = "Remedy Rate:  "  # label for the remedy rate
                cell.style = self.date_dov_title
                sheet[i].merge_cells('F5:G5')
                cell = sheet[i].cell(row=5, column=8)  # the $ value of the remedy
                cell.value = float(self.remedy_rate)
                cell.style = self.remedy_style
                cell.number_format = "[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00"
            else:
                sheet[i].row_dimensions[5].hidden = True
            cell = sheet[i].cell(row=7, column=1)
            cell.value = text[i]
            cell.style = self.instruct_text
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            sheet[i].row_dimensions[7].height = 90
            sheet[i].merge_cells('A7:J7')

    def _remedy_list_header(self, _list):
        """ create headers for each list on the remedy page """
        list_title = {"nl": "No List Carriers",
                      "wal": "Work Assignment Carriers ",
                      "otdl": "Overtime Desired List Carriers ",
                      "aux": "Auxiliary Assistance"}
        default_percent = {"nl": ".50", "wal": ".50 ", "otdl": "1.50", "aux": "1.50"}
        cell = self.remedy.cell(row=self.remedy_row, column=1)  # list section header
        cell.value = list_title[_list]
        cell.style = self.list_header
        self.remedy.merge_cells('G' + str(self.remedy_row) + ':I' + str(self.remedy_row))
        if self.show_remedy:
            cell = self.remedy.cell(row=self.remedy_row, column=7)  # remedy percentage label
            cell.value = "remedy percentage:"
            cell.style = self.date_dov_title
            cell = self.remedy.cell(row=self.remedy_row, column=10)  # remedy percentage input
            cell.value = default_percent[_list]
            cell.style = self.date_dov
            cell.number_format = "#,###.00;[RED]-#,###.00"
        self.remedy_row += 1
        sub_header = ("sat", "sun", "mon", "tue", "wed", "thu", "fri", "total")
        if self.show_remedy:
            sub_header = ("sat", "sun", "mon", "tue", "wed", "thu", "fri", "total", "remedy")
        for i in range(len(sub_header)):
            cell = self.remedy.cell(row=self.remedy_row, column=i+2)
            cell.value = sub_header[i]
            cell.style = self.col_header
        self.remedy_row += 1

    def _remedy_list_footer(self, _list):
        """ this will place a footter at the end of each list section"""
        list_title = {"nl": "No List Mandates:  ",
                      "wal": "Work Assignment Mandates:  ",
                      "otdl": "OTDL Availability:  ",
                      "aux": "Auxiliary Availability:  "}
        self.remedy_row += 1
        self.remedy_footer_row.append(self.remedy_row)  # save the row of the footers for equalization rows
        cell = self.remedy.cell(row=self.remedy_row, column=1)  # title row for section footer
        cell.value = list_title[_list]
        cell.style = self.date_dov_title
        column = ("B", "C", "D", "E", "F", "G", "H")
        for i in range(7):
            formula = "=SUM(%s!%s%s:%s!%s%s)" % ("remedy", column[i], self.remedy_start_row,
                                                 "remedy", column[i],  self.remedy_row-2)
            cell = self.remedy.cell(row=self.remedy_row, column=i+2)
            cell.value = formula
            cell.style = self.calcs
            cell.number_format = "#,###.00;[RED]-#,###.00"
        add_this = (self.remedy_start_row, self.remedy_row-2)  # capture the section for cumulative total at end
        self.remedy_cum.append(add_this)
        self.remedy_row += 1

    def _remedy_equalization(self):
        """ create rows at the bottom of the sheet for total mandates, total availability, and equalization """
        column = ("B", "C", "D", "E", "F", "G", "H")
        self.remedy_row += 1
        cell = self.remedy.cell(row=self.remedy_row, column=1)  # list section header
        cell.value = "Equalization"
        cell.style = self.list_header
        if self.show_remedy:
            self.remedy.merge_cells('G' + str(self.remedy_row) + ':I' + str(self.remedy_row))
            cell = self.remedy.cell(row=self.remedy_row, column=7)  # remedy percentage label
            cell.value = "cumulative remedy:"
            cell.style = self.date_dov_title
            formula = "=SUM(%s!J%s:%s!J%s)+SUM(%s!J%s:%s!J%s)+SUM(%s!J%s:%s!J%s)+SUM(%s!J%s:%s!J%s)" \
                      % ("remedy", self.remedy_cum[0][0], "remedy", self.remedy_cum[0][1],
                         "remedy", self.remedy_cum[1][0], "remedy", self.remedy_cum[1][1],
                         "remedy", self.remedy_cum[2][0], "remedy", self.remedy_cum[2][1],
                         "remedy", self.remedy_cum[3][0], "remedy", self.remedy_cum[3][1])
            cell = self.remedy.cell(row=self.remedy_row, column=10)  # remedy percentage input
            cell.value = formula
            cell.style = self.calcs
            cell.number_format = "[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00"
        self.remedy_row += 2
        # -------------------------------------------------------------------------------------------------- mandates
        self.remedy_equalizer_rows.append(self.remedy_row)  # save row number for equalization formula
        cell = self.remedy.cell(row=self.remedy_row, column=1)  # title row for section footer
        cell.value = "Mandates:  "
        cell.style = self.date_dov_title
        for i in range(7):
            formula = "=SUM(%s!%s%s+%s!%s%s)" % ("remedy", column[i], self.remedy_footer_row[0],
                                                 "remedy", column[i], self.remedy_footer_row[1])
            cell = self.remedy.cell(row=self.remedy_row, column=i+2)
            cell.value = formula
            cell.style = self.calcs
            cell.number_format = "#,###.00;[RED]-#,###.00"
        self.remedy_row += 1
        # ----------------------------------------------------------------------------------------------- availability
        self.remedy_equalizer_rows.append(self.remedy_row)  # save row number for equalization formula
        cell = self.remedy.cell(row=self.remedy_row, column=1)  # title row for section footer
        cell.value = "Availability:  "
        cell.style = self.date_dov_title
        for i in range(7):
            formula = "=%s!%s%s" % ("remedy", column[i], self.remedy_footer_row[2])
            cell = self.remedy.cell(row=self.remedy_row, column=i+2)
            cell.value = formula
            cell.style = self.calcs
            cell.number_format = "#,###.00;[RED]-#,###.00"
        self.remedy_row += 2
        # -------------------------------------------------------------------------------------------- estimated hours
        self.remedy.merge_cells('G' + str(self.remedy_row) + ':I' + str(self.remedy_row))
        cell = self.remedy.cell(row=self.remedy_row, column=7)  # remedy percentage label
        cell.value = "estimated hours:"
        cell.style = self.date_dov_title
        formula = "=(SUM(MIN(%s!B%s,%s!B%s)+MIN(%s!C%s,%s!C%s)+MIN(%s!D%s,%s!D%s)+" \
                  "MIN(%s!E%s,%s!E%s)+MIN(%s!F%s,%s!F%s)+MIN(%s!G%s,%s!G%s)+MIN(%s!H%s,%s!H%s))*2)" % \
                  ("remedy", int(self.remedy_row - 3), "remedy", int(self.remedy_row - 2),
                   "remedy", int(self.remedy_row - 3), "remedy", int(self.remedy_row - 2),
                   "remedy", int(self.remedy_row - 3), "remedy", int(self.remedy_row - 2),
                   "remedy", int(self.remedy_row - 3), "remedy", int(self.remedy_row - 2),
                   "remedy", int(self.remedy_row - 3), "remedy", int(self.remedy_row - 2),
                   "remedy", int(self.remedy_row - 3), "remedy", int(self.remedy_row - 2),
                   "remedy", int(self.remedy_row - 3), "remedy", int(self.remedy_row - 2))
        cell = self.remedy.cell(row=self.remedy_row, column=10)  # remedy percentage input
        cell.value = formula
        cell.style = self.calcs
        cell.number_format = "#,###.00;[RED]-#,###.00"
        self.remedy_row += 1
        # -------------------------------------------------------------------------------------------- estimated remedy
        if self.show_remedy:
            self.remedy.merge_cells('G' + str(self.remedy_row) + ':I' + str(self.remedy_row))
            cell = self.remedy.cell(row=self.remedy_row, column=7)  # remedy percentage label
            cell.value = "estimated remedy:"
            cell.style = self.date_dov_title
            formula = "=(%s!J%s * %s!H5)" % ("remedy", self.remedy_row - 1, "remedy")
            cell = self.remedy.cell(row=self.remedy_row, column=10)  # remedy percentage input
            cell.value = formula
            cell.style = self.calcs
            cell.number_format = "[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00"
            self.remedy_row += 1
        self.remedy_row += 1
        # ----------------------------------------------------------------------------------------------- equalization
        cell = self.remedy.cell(row=self.remedy_row, column=1)  # title row for section footer
        cell.value = "Equalization:  "
        cell.style = self.date_dov_title
        for i in range(7):
            formula = "=SUM(%s!%s%s-%s!%s%s)" % ("remedy", column[i], self.remedy_equalizer_rows[0],
                                                 "remedy", column[i], self.remedy_equalizer_rows[1])
            cell = self.remedy.cell(row=self.remedy_row, column=i + 2)
            cell.value = formula
            cell.style = self.calcs
            cell.number_format = "#,###.00;[RED]-#,###.00"
        self.remedy.merge_cells('I' + str(self.remedy_row) + ':J' + str(self.remedy_row))
        cell = self.remedy.cell(row=self.remedy_row, column=9)  # remedy percentage input
        cell.value = " <- adjust to zero"
        cell.style = self.date_dov_title
        self.remedy_row += 1
        self.remedy.merge_cells('B' + str(self.remedy_row) + ':H' + str(self.remedy_row))
        cell = self.remedy.cell(row=self.remedy_row, column=2)  # row for exposition on equalization
        cell.value = "\n" \
                     "1. Using the OTDL Weekly Availability Worksheet, alter/delete availability from the OTDL " \
                     "section if there is no availability. \n" \
                     "2. If value is positive, subtract/delete from No List and Work Assignment sections to " \
                     "equalize. \n" \
                     "3. If the value is negative, subtract/delete from OTDL and Auxiliary sections to equalize. \n"
        cell.style = self.instruct_text
        self.remedy['B' + str(self.remedy_row)].alignment = Alignment(wrap_text=True, vertical='top',
                                                                      shrink_to_fit=False)
        self.remedy.row_dimensions[self.remedy_row].height = 100

    @staticmethod
    def _remedy_violation_cell(_list, violation_cell, a_max, _12hour=True):
        """ accepts an empty string or a tuple. returns empty string for empty string
         returns a formula for a tuple. """
        sheet = "remedy"
        if not _12hour:
            sheet = "remedy_10hr"
        formula = ""  # the default is an empty string
        if type(violation_cell) == tuple:
            if _list in ("nl", "wal"):
                formula = "=IF(%s!%s%s>=%s!H4,%s!%s%s,\"\") " % \
                          (violation_cell[0], violation_cell[1], violation_cell[2], sheet,
                           violation_cell[0], violation_cell[1], violation_cell[2])
            else:
                formula = "=IF(MIN(%s!%s%s,%s)>=%s!H4,MIN(%s!%s%s,%s),\"\") " % \
                          (violation_cell[0], violation_cell[1], violation_cell[2], a_max, sheet,
                           violation_cell[0], violation_cell[1], violation_cell[2], a_max)

        return formula

    def _display_odl_mod_name(self, name):
        """ return a modified name is the names is in the odlr or odln indicator arrays. """
        if name in self.odlr_indicator and name in self.odln_indicator:
            return name + " (odl+)"
        elif name in self.odlr_indicator:
            return name + " (odlr)"
        elif name in self.odln_indicator:
            return name + " (odln)"
        else:
            return name

    def _display_remedy_row(self, name, _list, violation_cells, offbid):
        """ display the name, daily violations, total and remedy for each name - will fill one row of remedy sheet """
        days = ("saturday", "sunday", "monday", "tuesday", "wednesday", "thursday", "friday")
        if not projvar.invran_weekly_span:
            violation_cells = [x for x in violation_cells if x]  # remove all empty elements from violation cells array
        try:
            a_max_array = self.a_max_dict[name]
        except KeyError:
            a_max_array = [12, 12, 12, 12, 12, 12, 12]
        cell = self.remedy.cell(row=self.remedy_row, column=1)
        cell.value = self._display_odl_mod_name(name)
        cell.style = self.input_name
        cell.number_format = '@'
        if projvar.invran_weekly_span:
            for i in range(len(self.dates)):  # display violations
                cell = self.remedy.cell(row=self.remedy_row, column=i+2)
                cell.value = self._remedy_violation_cell(_list, violation_cells[i], a_max_array[i])  # get the formula
                if i in offbid:
                    cell.value = ""
                cell.style = self.input_s
                cell.number_format = "#,###.00;[RED]-#,###.00"
        else:  # for a daily investigation
            for i in range(7):  # display violations
                cell = self.remedy.cell(row=self.remedy_row, column=i+2)
                cell.value = ""
                if days[i] == violation_cells[0][0]:  # if the correct day-display
                    # get the formula/ 12 hour kwarg is sent to designate the OT ON route sheet.
                    cell.value = self._remedy_violation_cell(_list, violation_cells[0], a_max_array[0],
                                                             _12hour=False)
                    if 0 in offbid:
                        cell.value = ""
                cell.style = self.input_s
                cell.number_format = "#,###.00;[RED]-#,###.00"
        # display totals cell at the end of the row
        formula_a = "=IF(SUM(%s!B%s:%s!H%s)>0, SUM(%s!B%s:%s!H%s), \"\")" % \
                    ('remedy', self.remedy_row, 'remedy', self.remedy_row, 'remedy', self.remedy_row, 'remedy',
                     self.remedy_row)
        cell = self.remedy.cell(row=self.remedy_row, column=9)
        cell.value = formula_a
        cell.style = self.calcs
        cell.number_format = "#,###.00;[RED]-#,###.00"
        if self.show_remedy:  # display remedy cell for a dollar amount remedy
            formula_a = "=IF(AND(%s!I%s<>\"\", %s!I%s<>0),(%s!H$5 * %s!J$%s) * %s!I%s,\"\")" % \
                        ('remedy', str(self.remedy_row), 'remedy', str(self.remedy_row),
                         'remedy', 'remedy', str(self.remedy_start_row-2), 'remedy', str(self.remedy_row))
            cell = self.remedy.cell(row=self.remedy_row, column=10)
            cell.value = formula_a
            cell.style = self.calcs
            cell.number_format = "[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00"

    def _build_remedy_blank_arrays(self):
        """ use the remedy array to find all blank rows and build the blank remedy arrays """
        day = {"sat": 0, "sun": 1, "mon": 2, "tue": 3, "wed": 4, "thu": 5, "fri": 6}
        for rec in self.remedy_array:
            if not rec[2]:  # if there is not a name in the array
                if rec[0] == "nl":  # place array in no list blank array
                    self.remedy_blank_nl[day[rec[1]]].append(rec[3])
                if rec[0] == "wal":  # place array in work assignment blank array
                    self.remedy_blank_wal[day[rec[1]]].append(rec[3])
                if rec[0] == "otdl":  # place array in overtime desired list blank array
                    self.remedy_blank_otdl[day[rec[1]]].append(rec[3])
                if rec[0] == "aux":  # place array in auxiliary blank array
                    self.remedy_blank_aux[day[rec[1]]].append(rec[3])
        for rec in self.remedy10_array:
            if not rec[2]:  # if there is not a name in the array
                if rec[0] == "nl":  # place array in no list blank array
                    self.remedy10_blank_nl[day[rec[1]]].append(rec[3])
                # because wal is not used in 10 hour remedy, there is no such array
                if rec[0] == "otdl":  # place array in overtime desired list blank array
                    self.remedy10_blank_otdl[day[rec[1]]].append(rec[3])
                if rec[0] == "aux":  # place array in auxiliary blank array
                    self.remedy10_blank_aux[day[rec[1]]].append(rec[3])
    
    @staticmethod        
    def _get_array_length(array):
        """ get the maximum number of elements among the 7 sub-arrays"""
        max_count = 0
        for sub_array in array:
            if len(sub_array) > max_count:
                max_count = len(sub_array)
        return max_count
            
    def _equalize_remedy_blank_arrays(self):
        """ ensure that all remedy blank arrays have equal count of elements in all 7 of their sub-arrays """
        max_count = self._get_array_length(self.remedy_blank_nl)  # blank nl array
        for i in range(7):
            while len(self.remedy_blank_nl[i]) < max_count:
                self.remedy_blank_nl[i].append("")
                self.remedy10_blank_nl[i].append("")
        max_count = self._get_array_length(self.remedy_blank_wal)  # blank wal array
        for i in range(7):
            while len(self.remedy_blank_wal[i]) < max_count:
                self.remedy_blank_wal[i].append("")
                # there is no remedy10_blank_wal array
        max_count = self._get_array_length(self.remedy_blank_otdl)  # blank otdl array
        for i in range(7):
            while len(self.remedy_blank_otdl[i]) < max_count:
                self.remedy_blank_otdl[i].append("")
                self.remedy10_blank_otdl[i].append("")
        max_count = self._get_array_length(self.remedy_blank_aux)  # blank aux array
        for i in range(7):
            while len(self.remedy_blank_aux[i]) < max_count:
                self.remedy_blank_aux[i].append("")
                self.remedy10_blank_aux[i].append("")

    def _order_remedy_blank_arrays(self):
        """ sort the remedy blank arrays so that the unnamed user has 7 coordinates """
        max_count = len(self.remedy_blank_nl[0])  # no list
        for i in range(max_count):
            new_array = []
            new10_array = []
            for ii in range(7):  # for each day of the week
                new_array.append(self.remedy_blank_nl[ii][i])
                new10_array.append(self.remedy10_blank_nl[ii][i])
            self.order_nl_blanks.append(new_array)
            self.order10_nl_blanks.append(new10_array)
        max_count = len(self.remedy_blank_wal[0])  # work asignment
        for i in range(max_count):
            new_array = []
            for ii in range(7):  # for each day of the week
                new_array.append(self.remedy_blank_wal[ii][i])
            self.order_wal_blanks.append(new_array)
        max_count = len(self.remedy_blank_otdl[0])  # otdl
        self.order_otdl_blanks = []
        for i in range(max_count):
            new_array = []
            new10_array = []
            for ii in range(7):  # for each day of the week
                new_array.append(self.remedy_blank_otdl[ii][i])
                new10_array.append(self.remedy10_blank_otdl[ii][i])
            self.order_otdl_blanks.append(new_array)
            self.order10_otdl_blanks.append(new10_array)
        max_count = len(self.remedy_blank_aux[0])  # auxiliary
        self.order_aux_blanks = []
        for i in range(max_count):
            new_array = []
            new10_array = []
            for ii in range(7):  # for each day of the week
                new_array.append(self.remedy_blank_aux[ii][i])
                new10_array.append(self.remedy10_blank_aux[ii][i])
            self.order_aux_blanks.append(new_array)
            self.order10_aux_blanks.append(new10_array)

    def _build_remedy_blanks(self, _list, _12hr=True):
        """ build the remedy blanks on the remedy page """
        if _list == "nl":
            for i in range(len(self.order_nl_blanks)):
                name = ""
                if _12hr:
                    self._display_remedy_row(name, _list, self.order_nl_blanks[i], [])
                    self.remedy_row += 1
                else:
                    self._display_remedy10_row(name, _list, self.order10_nl_blanks[i], [])
                    self.remedy10_row += 1
        if _list == "wal":
            for i in range(len(self.order_wal_blanks)):
                name = ""
                if _12hr:
                    self._display_remedy_row(name, _list, self.order_wal_blanks[i], [])
                    self.remedy_row += 1
        if _list == "otdl":
            for i in range(len(self.order_otdl_blanks)):
                name = ""
                if _12hr:
                    self._display_remedy_row(name, _list, self.order_otdl_blanks[i], [])
                    self.remedy_row += 1
                else:
                    self._display_remedy10_row(name, _list, self.order10_otdl_blanks[i], [])
                    self.remedy10_row += 1
        if _list == "aux":
            for i in range(len(self.order_aux_blanks)):
                name = ""
                if _12hr:
                    self._display_remedy_row(name, _list, self.order_aux_blanks[i], [])
                    self.remedy_row += 1
                else:
                    self._display_remedy10_row(name, _list, self.order10_aux_blanks[i], [])
                    self.remedy10_row += 1

    def _build_remedy(self):
        """ build the remedy page """
        self.remedy_row = 9
        _list_array = ("nl", "wal", "otdl", "aux")
        day_array = ("sat", "sun", "mon", "tue", "wed", "thu", "fri")
        if not projvar.invran_weekly_span:
            day_array = (projvar.invran_date.strftime("%a").lower(), )
        for _list in _list_array:  # sort names by list
            temp_names = []
            for rec in self.remedy_array:  # create a distinct list of names for each list
                if rec[0] == _list:  # if the rec cooresponds to the list.
                    if rec[2]:  # if the name element is not empty
                        if rec[2] not in temp_names:  # avoid duplicates
                            temp_names.append(rec[2])  # add the name
            temp_names = sorted(temp_names)  # sort the list
            self._remedy_list_header(_list)  # create headers for each list on the remedy page
            self.remedy_start_row = self.remedy_row  # capture the starting row for sum formula
            for name in temp_names:
                violation_cells = []  # list with an element for each day, holds cell coordinates or empty string.
                offbid = []
                for i in range(len(self.dates)):
                    add_this = ""
                    for r in self.remedy_array:
                        if r[0] == _list and r[1] == day_array[i] and r[2] == name:
                            add_this = r[3]
                            if self.offbid_dict[name][i] and _list in ("nl", "wal"):  # ignore offbid violations
                                offbid.append(i)  # put index of offbid days in an array
                            break
                    violation_cells.append(add_this)
                self._display_remedy_row(name, _list, violation_cells, offbid)
                self.remedy_row += 1  # after block of name/remedies - add a blank row for readability
            # solutions for blank rows - insert here
            self._build_remedy_blanks(_list)
            self._remedy_list_footer(_list)
            if _list is not "aux":  # insert page breaks for all list except the last
                try:
                    self.remedy.page_breaks.append(Break(id=self.remedy_row))
                except AttributeError:
                    self.remedy.row_breaks.append(Break(id=self.remedy_row))  # effective for windows
            self.remedy_row += 1
        self._remedy_equalization()  # write the rows for the end of the sheet

    def _remedy10_list_header(self, _list):
        """ create headers for each list on the remedy page """
        list_title = {"nl": "No List Carriers",
                      "otdl": "Overtime Desired List Carriers ",
                      "aux": "Auxiliary Assistance"}
        default_percent = {"nl": ".50", "otdl": "1.50", "aux": "1.50"}
        cell = self.remedy_10hr.cell(row=self.remedy10_row, column=1)  # list section header
        cell.value = list_title[_list]
        cell.style = self.list_header
        self.remedy_10hr.merge_cells('G' + str(self.remedy10_row) + ':I' + str(self.remedy10_row))
        if self.show_remedy:
            cell = self.remedy_10hr.cell(row=self.remedy10_row, column=7)  # remedy percentage label
            cell.value = "remedy percentage:"
            cell.style = self.date_dov_title
            cell = self.remedy_10hr.cell(row=self.remedy10_row, column=10)  # remedy percentage input
            cell.value = default_percent[_list]
            cell.style = self.date_dov
            cell.number_format = "#,###.00;[RED]-#,###.00"
        self.remedy10_row += 1
        sub_header = ("sat", "sun", "mon", "tue", "wed", "thu", "fri", "total")
        if self.show_remedy:
            sub_header = ("sat", "sun", "mon", "tue", "wed", "thu", "fri", "total", "remedy")
        for i in range(len(sub_header)):
            cell = self.remedy_10hr.cell(row=self.remedy10_row, column=i+2)
            cell.value = sub_header[i]
            cell.style = self.col_header
        self.remedy10_row += 1

    def _remedy10_list_footer(self, _list):
        """ this will place a footter at the end of each list section"""
        list_title = {"nl": "No List Mandates:  ",
                      "otdl": "OTDL Availability:  ",
                      "aux": "Auxiliary Availability:  "}
        self.remedy10_row += 1
        self.remedy10_footer_row.append(self.remedy10_row)  # save the row of the footers for equalization rows
        cell = self.remedy_10hr.cell(row=self.remedy10_row, column=1)  # title row for section footer
        cell.value = list_title[_list]
        cell.style = self.date_dov_title
        column = ("B", "C", "D", "E", "F", "G", "H")
        for i in range(7):
            formula = "=SUM(%s!%s%s:%s!%s%s)" % ("remedy_10hr", column[i], self.remedy10_start_row,
                                                 "remedy_10hr", column[i],  self.remedy10_row-2)
            cell = self.remedy_10hr.cell(row=self.remedy10_row, column=i+2)
            cell.value = formula
            cell.style = self.calcs
            cell.number_format = "#,###.00;[RED]-#,###.00"
        add_this = (self.remedy10_start_row, self.remedy10_row-2)  # capture the section for cumulative total at end
        self.remedy10_cum.append(add_this)
        self.remedy10_row += 1

    def _remedy10_equalization(self):
        """ create rows at the bottom of the sheet for total mandates, total availability, and equalization """
        column = ("B", "C", "D", "E", "F", "G", "H")
        self.remedy10_row += 1
        cell = self.remedy_10hr.cell(row=self.remedy10_row, column=1)  # list section header
        cell.value = "Equalization"
        cell.style = self.list_header
        if self.show_remedy:
            self.remedy_10hr.merge_cells('G' + str(self.remedy10_row) + ':I' + str(self.remedy10_row))
            cell = self.remedy_10hr.cell(row=self.remedy10_row, column=7)  # remedy percentage label
            cell.value = "cumulative remedy:"
            cell.style = self.date_dov_title
            formula = "=SUM(%s!J%s:%s!J%s)+SUM(%s!J%s:%s!J%s)+SUM(%s!J%s:%s!J%s)" \
                      % ("remedy_10hr", self.remedy10_cum[0][0], "remedy_10hr", self.remedy10_cum[0][1],
                         "remedy_10hr", self.remedy10_cum[1][0], "remedy_10hr", self.remedy10_cum[1][1],
                         "remedy_10hr", self.remedy10_cum[2][0], "remedy_10hr", self.remedy10_cum[2][1])
            cell = self.remedy_10hr.cell(row=self.remedy10_row, column=10)  # remedy percentage input
            cell.value = formula
            cell.style = self.calcs
            cell.number_format = "[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00"
        self.remedy10_row += 2
        # ----------------------------------------------------------------------------------------- own route mandates
        self.remedy10_equalizer_rows.append(self.remedy10_row)  # save row number for equalization formula
        cell = self.remedy_10hr.cell(row=self.remedy10_row, column=1)  # title row for section footer
        cell.value = "Own Route Mandates:  "
        cell.style = self.date_dov_title
        for i in range(7):
            formula = "=SUM(%s!%s%s)" % ("remedy_10hr", column[i], self.remedy10_footer_row[0])
            cell = self.remedy_10hr.cell(row=self.remedy10_row, column=i+2)
            cell.value = formula
            cell.style = self.calcs
            cell.number_format = "#,###.00;[RED]-#,###.00"
        self.remedy10_row += 1
        # ----------------------------------------------------------------------------------------------- availability
        self.remedy10_equalizer_rows.append(self.remedy10_row)  # save row number for equalization formula
        cell = self.remedy_10hr.cell(row=self.remedy10_row, column=1)  # title row for section footer
        cell.value = "Availability to 10 Hours:  "
        cell.style = self.date_dov_title
        for i in range(7):
            formula = "=%s!%s%s+%s!%s%s" % ("remedy_10hr", column[i], self.remedy10_footer_row[1],
                                            "remedy_10hr", column[i], self.remedy10_footer_row[2])
            cell = self.remedy_10hr.cell(row=self.remedy10_row, column=i+2)
            cell.value = formula
            cell.style = self.calcs
            cell.number_format = "#,###.00;[RED]-#,###.00"
        self.remedy10_row += 2
        # -------------------------------------------------------------------------------------------- estimated hours
        self.remedy_10hr.merge_cells('G' + str(self.remedy10_row) + ':I' + str(self.remedy10_row))
        cell = self.remedy_10hr.cell(row=self.remedy10_row, column=7)  # remedy percentage label
        cell.value = "estimated hours:"
        cell.style = self.date_dov_title
        formula = "=(SUM(MIN(%s!B%s,%s!B%s)+MIN(%s!C%s,%s!C%s)+MIN(%s!D%s,%s!D%s)+" \
                  "MIN(%s!E%s,%s!E%s)+MIN(%s!F%s,%s!F%s)+MIN(%s!G%s,%s!G%s)+MIN(%s!H%s,%s!H%s))*2)" % \
                  ("remedy_10hr", int(self.remedy10_row - 3), "remedy_10hr", int(self.remedy10_row - 2),
                   "remedy_10hr", int(self.remedy10_row - 3), "remedy_10hr", int(self.remedy10_row - 2),
                   "remedy_10hr", int(self.remedy10_row - 3), "remedy_10hr", int(self.remedy10_row - 2),
                   "remedy_10hr", int(self.remedy10_row - 3), "remedy_10hr", int(self.remedy10_row - 2),
                   "remedy_10hr", int(self.remedy10_row - 3), "remedy_10hr", int(self.remedy10_row - 2),
                   "remedy_10hr", int(self.remedy10_row - 3), "remedy_10hr", int(self.remedy10_row - 2),
                   "remedy_10hr", int(self.remedy10_row - 3), "remedy_10hr", int(self.remedy10_row - 2))
        cell = self.remedy_10hr.cell(row=self.remedy10_row, column=10)  # remedy percentage input
        cell.value = formula
        cell.style = self.calcs
        cell.number_format = "#,###.00;[RED]-#,###.00"
        self.remedy10_row += 1
        # -------------------------------------------------------------------------------------------- estimated remedy
        if self.show_remedy:
            self.remedy_10hr.merge_cells('G' + str(self.remedy10_row) + ':I' + str(self.remedy10_row))
            cell = self.remedy_10hr.cell(row=self.remedy10_row, column=7)  # remedy percentage label
            cell.value = "estimated remedy:"
            cell.style = self.date_dov_title
            formula = "=(%s!J%s * %s!H5)" % ("remedy_10hr", self.remedy10_row - 1, "remedy_10hr")
            cell = self.remedy_10hr.cell(row=self.remedy10_row, column=10)  # remedy percentage input
            cell.value = formula
            cell.style = self.calcs
            cell.number_format = "[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00"
            self.remedy10_row += 1
        self.remedy10_row += 1
        # ----------------------------------------------------------------------------------------------- equalization
        cell = self.remedy_10hr.cell(row=self.remedy10_row, column=1)  # title row for section footer
        cell.value = "Equalization:  "
        cell.style = self.date_dov_title
        for i in range(7):
            formula = "=SUM(%s!%s%s-%s!%s%s)" % ("remedy_10hr", column[i], self.remedy10_equalizer_rows[0],
                                                 "remedy_10hr", column[i], self.remedy10_equalizer_rows[1])
            cell = self.remedy_10hr.cell(row=self.remedy10_row, column=i + 2)
            cell.value = formula
            cell.style = self.calcs
            cell.number_format = "#,###.00;[RED]-#,###.00"
        self.remedy_10hr.merge_cells('I' + str(self.remedy10_row) + ':J' + str(self.remedy10_row))
        cell = self.remedy_10hr.cell(row=self.remedy10_row, column=9)  # remedy percentage input
        cell.value = " <- adjust to zero"
        cell.style = self.date_dov_title
        self.remedy10_row += 1
        self.remedy_10hr.merge_cells('B' + str(self.remedy10_row) + ':H' + str(self.remedy10_row))
        cell = self.remedy_10hr.cell(row=self.remedy10_row, column=2)  # row for exposition on equalization
        cell.value = "\n" \
                     "1. Using the OTDL Weekly Availability Worksheet, alter/delete availability from the OTDL " \
                     "section if there is no availability. \n" \
                     "2. If value is positive, subtract/delete from No List and Work Assignment sections to " \
                     "equalize. \n" \
                     "3. If the value is negative, subtract/delete from OTDL and Auxiliary sections to equalize. \n"
        cell.style = self.instruct_text
        self.remedy_10hr['B' + str(self.remedy10_row)].alignment = Alignment(wrap_text=True, vertical='top',
                                                                             shrink_to_fit=False)
        self.remedy_10hr.row_dimensions[self.remedy10_row].height = 100

    def _display_remedy10_row(self, name, _list, violation_cells, offbid):
        """ display the name, daily violations, total and remedy for each name - will fill one row of remedy sheet """
        days = ("saturday", "sunday", "monday", "tuesday", "wednesday", "thursday", "friday")
        if not projvar.invran_weekly_span:
            violation_cells = [x for x in violation_cells if x]  # remove all empty elements from violation cells array
        try:
            a_max_array = self.a_max_dict[name]
        except KeyError:
            a_max_array = [10, 10, 10, 10, 10, 10, 10]
        cell = self.remedy_10hr.cell(row=self.remedy10_row, column=1)
        cell.value = self._display_odl_mod_name(name)  # display the name of the carrier
        cell.style = self.input_name
        cell.number_format = '@'
        if projvar.invran_weekly_span:  # for a weekly investigation
            for i in range(7):  # display violations
                cell = self.remedy_10hr.cell(row=self.remedy10_row, column=i+2)
                # get the formula/ 12 hour kwarg is sent to designate the OT ON route sheet.
                cell.value = self._remedy_violation_cell(_list, violation_cells[i], a_max_array[i], _12hour=False)
                if i in offbid:
                    cell.value = ""
                cell.style = self.input_s
                cell.number_format = "#,###.00;[RED]-#,###.00"
        else:  # for a daily investigation
            for i in range(7):  # display violations
                cell = self.remedy_10hr.cell(row=self.remedy10_row, column=i+2)
                cell.value = ""
                if days[i] == violation_cells[0][0]:  # if the correct day-display
                    # get the formula/ 12 hour kwarg is sent to designate the OT ON route sheet.
                    cell.value = self._remedy_violation_cell(_list, violation_cells[0], a_max_array[0],
                                                             _12hour=False)
                    if i in offbid:
                        cell.value = ""
                cell.style = self.input_s
                cell.number_format = "#,###.00;[RED]-#,###.00"
        # display totals cell at the end of the row
        formula_a = "=IF(SUM(%s!B%s:%s!H%s)>0, SUM(%s!B%s:%s!H%s), \"\")" % \
                    ('remedy_10hr', self.remedy10_row, 'remedy_10hr', self.remedy10_row, 'remedy_10hr',
                     self.remedy10_row, 'remedy_10hr', self.remedy10_row)
        cell = self.remedy_10hr.cell(row=self.remedy10_row, column=9)
        cell.value = formula_a
        cell.style = self.calcs
        cell.number_format = "#,###.00;[RED]-#,###.00"
        if self.show_remedy:  # display remedy cell for a dollar amount remedy
            formula_a = "=IF(AND(%s!I%s<>\"\", %s!I%s<>0),(%s!H$5 * %s!J$%s) * %s!I%s,\"\")" % \
                        ('remedy_10hr', str(self.remedy10_row), 'remedy_10hr', str(self.remedy10_row),
                         'remedy_10hr', 'remedy_10hr', str(self.remedy10_start_row-2), 'remedy_10hr',
                         str(self.remedy10_row))
            cell = self.remedy_10hr.cell(row=self.remedy10_row, column=10)
            cell.value = formula_a
            cell.style = self.calcs
            cell.number_format = "[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00"

    def _build_remedy10(self):
        """ build the remedy page """
        self.remedy10_row = 9
        _list_array = ("nl", "otdl", "aux")
        day_array = ("sat", "sun", "mon", "tue", "wed", "thu", "fri")
        if not projvar.invran_weekly_span:
            day_array = (projvar.invran_date.strftime("%a").lower(), )
        for _list in _list_array:  # sort names by list
            temp_names = []
            for rec in self.remedy10_array:  # create a distinct list of names for each list
                if rec[0] == _list:  # if the rec cooresponds to the list.
                    if rec[2]:  # if the name element is not empty
                        if rec[2] not in temp_names:  # avoid duplicates
                            temp_names.append(rec[2])  # add the name
            temp_names = sorted(temp_names)  # sort the list
            self._remedy10_list_header(_list)  # create headers for each list on the remedy page
            self.remedy10_start_row = self.remedy10_row  # capture the starting row for sum formula
            for name in temp_names:
                violation_cells = []  # list with an element for each day, holds cell coordinates or empty string.
                offbid = []
                for i in range(len(self.dates)):
                    add_this = ""
                    for r in self.remedy10_array:
                        if r[0] == _list and r[1] == day_array[i] and r[2] == name:
                            add_this = r[3]
                            if self.offbid_dict[name][i] and _list in ("nl", "wal"):  # ignore offbid violations
                                offbid.append(i)  # put index of offbid days in an array
                            break
                    violation_cells.append(add_this)
                self._display_remedy10_row(name, _list, violation_cells, offbid)
                self.remedy10_row += 1  # after block of name/remedies - add a blank row for readability
            self._build_remedy_blanks(_list, _12hr=False)  # solutions for blank rows
            self._remedy10_list_footer(_list)
            if _list is not "aux":  # insert page breaks for all list except the last
                try:
                    self.remedy_10hr.page_breaks.append(Break(id=self.remedy10_row))
                except AttributeError:
                    self.remedy_10hr.row_breaks.append(Break(id=self.remedy10_row))  # effective for windows
            self.remedy10_row += 1
        self._remedy10_equalization()  # write the rows for the end of the sheet

    def build_summary(self):
        """ build the summary page. """
        self.pbi += 1
        self.pb.move_count(self.pbi)  # increment progress bar
        self.pb.change_text("Building day Summary...")
        daily_dict = {"saturday": 0, "sunday": 1, "monday": 2, "tuesday": 3, "wednesday": 4, "thursday": 5, "friday": 6}

        def get_formula(section_name, column, ii):
            """ returns the appropriate formula given the args """
            day = daily_dict[self.dates[ii].strftime("%A").lower()]
            day_column = ("B", "C", "D", "E", "F", "G", "H")
            if section_name == "unadjusted":
                if column == "B":  # availability to 10 hr formula
                    location = self.availability_10[ii]  # location of the total from the worksheet from the array
                    return "=%s!%s%s" % (location[0], location[1], location[2])
                if column == "C":  # mandates on own route formula
                    location = self.mandates_own_route[ii]  # location of the total from the worksheet from the array
                    return "=%s!%s%s" % (location[0], location[1], location[2])
                if column == "F":  # availability to 12 hr formula
                    location = self.availability_max[ii]  # location of the total from the worksheet from the array
                    return "=%s!%s%s" % (location[0], location[1], location[2])
                if column == "G":  # off route mandates formula
                    location = self.mandates_all[ii]  # get location of total mandates from worksheet from the array
                    return "=%s!%s%s" % (location[0], location[1], location[2])  # total mandates
            if section_name == "adjusted":
                if column == "B":  # availability to 10 hr formula
                    return "=%s!%s%s" % ("remedy_10hr", day_column[day], self.remedy10_equalizer_rows[1])
                if column == "C":  # mandates on own route formula
                    return "=%s!%s%s" % ("remedy_10hr", day_column[day], self.remedy10_equalizer_rows[0])
                if column == "F":  # availability to 12 hr formula
                    return "=%s!%s%s" % ("remedy", day_column[day], self.remedy_equalizer_rows[1])
                if column == "G":  # off route mandates formula
                    return "=%s!%s%s" % ("remedy", day_column[day], self.remedy_equalizer_rows[0])

        self.summary['A1'] = "Improper Mandate Summary"
        self.summary['A1'].style = self.ws_header
        self.summary.merge_cells('A1:E1')
        self.summary['A3'] = "Pay Period:  "
        self.summary['A3'].style = self.date_dov_title
        self.summary['B3'] = projvar.pay_period
        self.summary['B3'].style = self.date_dov
        self.summary.merge_cells('B5:D5')
        self.summary['A4'] = "Station:  "
        self.summary['A4'].style = self.date_dov_title
        self.summary['B4'] = projvar.invran_station
        self.summary['B4'].style = self.date_dov
        self.summary.merge_cells('B4:D4')
        sections = ("Unadjusted Totals", "Totals Adjusted for Availability, Off Bid Violations and Tolerances")
        section = ("unadjusted", "adjusted")
        row = 6
        for section_count in range(len(sections)):
            self.summary['A' + str(row)] = sections[section_count]
            self.summary['A' + str(row)].style = self.ws_header
            self.summary.merge_cells('A' + str(row) + ':H' + str(row))
            row += 1
            self.summary['B' + str(row)] = "No list"
            self.summary['B' + str(row)].style = self.date_dov_title
            self.summary['B' + str(row + 1)] = "overtime"
            self.summary['B' + str(row + 1)].style = self.date_dov_title
            self.summary['C' + str(row)] = "Availability"
            self.summary['C' + str(row)].style = self.date_dov_title
            self.summary['C' + str(row + 1)] = "to 10"
            self.summary['C' + str(row + 1)].style = self.date_dov_title
            self.summary['D' + str(row + 1)] = "violations"
            self.summary['D' + str(row + 1)].style = self.date_dov_title
            self.summary['F' + str(row)] = "Off route"
            self.summary['F' + str(row)].style = self.date_dov_title
            self.summary['F' + str(row + 1)] = "mandates"
            self.summary['F' + str(row + 1)].style = self.date_dov_title
            self.summary['G' + str(row)] = "Availability"
            self.summary['G' + str(row)].style = self.date_dov_title
            self.summary['G' + str(row + 1)] = "to 12"
            self.summary['G' + str(row + 1)].style = self.date_dov_title
            self.summary['H' + str(row + 1)] = "violations"
            self.summary['H' + str(row + 1)].style = self.date_dov_title
            row += 2
            for i in range(len(self.dates)):
                self.summary['A' + str(row)].value = format(self.dates[i], "%m/%d/%y %a")
                self.summary['A' + str(row)].style = self.date_dov_title
                # ----------------------------------------------------------------------------------- no list overtime
                self.summary['B' + str(row)] = get_formula(section[section_count], "C", i)
                self.summary['B' + str(row)].style = self.input_s
                self.summary['B' + str(row)].number_format = "#,###.00;[RED]-#,###.00"
                # --------------------------------------------------------------------------------- availability to 10
                self.summary['C' + str(row)] = get_formula(section[section_count], "B", i)
                self.summary['C' + str(row)].style = self.input_s
                self.summary['C' + str(row)].number_format = "#,###.00;[RED]-#,###.00"
                # --------------------------------------------------------------------------------- 10 hour violations
                formula = "=IF(%s!B%s<%s!C%s,%s!B%s,%s!C%s)" \
                          % ('summary', str(row), 'summary', str(row), 'summary',
                             str(row), 'summary', str(row))
                self.summary['D' + str(row)] = formula
                self.summary['D' + str(row)].style = self.calcs
                self.summary['D' + str(row)].number_format = "#,###.00;[RED]-#,###.00"
                # ---------------------------------------------------------------------------------- off route mandates
                self.summary['F' + str(row)] = get_formula(section[section_count], "G", i)
                self.summary['F' + str(row)].style = self.input_s
                self.summary['F' + str(row)].number_format = "#,###.00;[RED]-#,###.00"
                # --------------------------------------------------------------------------------- availability to 12
                self.summary['G' + str(row)] = get_formula(section[section_count], "F", i)
                self.summary['G' + str(row)].style = self.input_s
                self.summary['G' + str(row)].number_format = "#,###.00;[RED]-#,###.00"
                # ---------------------------------------------------------------------------------- 12 hour violations
                formula = "=IF(%s!F%s<%s!G%s,%s!F%s,%s!G%s)" \
                          % ('summary', str(row), 'summary', str(row), 'summary',
                             str(row), 'summary', str(row))
                self.summary['H' + str(row)] = formula
                self.summary['H' + str(row)].style = self.calcs
                self.summary['H' + str(row)].number_format = "#,###.00;[RED]-#,###.00"
                row += 2

    def build_refs(self):
        """ build the references page. This shows tolerances and defines labels. """
        self.pbi += 1
        self.pb.move_count(self.pbi)  # increment progress bar
        self.pb.change_text("Building Reference Page")
        # tolerances
        row = 1
        cell = self.reference.cell(row=row, column=1)
        cell.value = "Improper Mandate Reference Sheet"
        cell.style = self.ws_header
        self.reference.merge_cells('A1:E1')
        row += 2
        self.reference['B' + str(row)].style = self.list_header
        self.reference['B' + str(row)] = "Tolerances"
        row += 1
        text = "These tolerances will apply to the OT on route, OT off route and Availability totals on the daily " \
               "worksheets. They can be changed below or in the Klusterbox app (Management> Tolerances). \n" \
               "Remedy tolerances can be set on the Remedies Worksheet or on the Klusterbox app (Management> " \
               "Spreadsheet settings> Improper Mandate Spreadsheets> Remedy Settings)."
        cell = self.reference.cell(row=row, column=3)
        cell.value = text
        cell.style = self.instruct_text
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        self.reference.merge_cells('C4:E4')
        self.reference.row_dimensions[4].height = 75
        row += 1
        self.reference['C' + str(row)] = self.tol_ot_ownroute  # overtime on own route tolerance
        self.reference['C' + str(row)].style = self.input_s
        self.reference['C' + str(row)].number_format = "#,###.00;[RED]-#,###.00"
        self.reference['E' + str(row)] = "overtime on own route"
        row += 1
        self.reference['C' + str(row)] = self.tol_ot_offroute  # overtime off own route tolerance
        self.reference['C' + str(row)].style = self.input_s
        self.reference['C' + str(row)].number_format = "#,###.00;[RED]-#,###.00"
        self.reference['E' + str(row)] = "overtime off own route"
        row += 1
        self.reference['C' + str(row)] = self.tol_availability  # availability tolerance
        self.reference['C' + str(row)].style = self.input_s
        self.reference['C' + str(row)].number_format = "#,###.00;[RED]-#,###.00"
        self.reference['E' + str(row)] = "availability tolerance"
        row += 2
        # note guide
        self.reference['B' + str(row)].style = self.list_header
        self.reference['B' + str(row)] = "Note Guide"
        row += 1
        self.reference['C' + str(row)] = "ns day"
        self.reference['C' + str(row)].style = self.input_s
        self.reference['E' + str(row)] = "Carrier worked on their non scheduled day"
        row += 2
        self.reference['C' + str(row)] = "no call"
        self.reference['C' + str(row)].style = self.input_s
        self.reference['E' + str(row)] = "Carrier was not scheduled for overtime"
        row += 1
        self.reference['C' + str(row)] = "light"
        self.reference['C' + str(row)].style = self.input_s
        self.reference['E' + str(row)] = "Carrier on light duty and unavailable for overtime"
        row += 1
        self.reference['C' + str(row)] = "sch chg"
        self.reference['C' + str(row)].style = self.input_s
        self.reference['E' + str(row)] = "Schedule change: unavailable for overtime"
        row += 1
        self.reference['C' + str(row)] = "annual"
        self.reference['C' + str(row)].style = self.input_s
        self.reference['E' + str(row)] = "Annual leave"
        row += 1
        self.reference['C' + str(row)] = "sick"
        self.reference['C' + str(row)].style = self.input_s
        self.reference['E' + str(row)] = "Sick leave"
        row += 1
        self.reference['C' + str(row)] = "excused"
        self.reference['C' + str(row)].style = self.input_s
        self.reference['E' + str(row)] = "Carrier excused from mandatory overtime"
        row += 2
        # column headers
        self.reference['B' + str(row)].style = self.list_header
        self.reference['B' + str(row)] = "Column Headers"
        row += 1
        self.reference['C' + str(row)] = "Name"
        self.reference['C' + str(row)].style = self.input_s
        self.reference['E' + str(row)] = "The name of the carrier. "
        row += 1
        self.reference['C' + str(row)] = "Note"
        self.reference['C' + str(row)].style = self.input_s
        self.reference['E' + str(row)] = "Special circumstances. See note guide above."
        row += 1
        self.reference['C' + str(row)] = "5200"
        self.reference['C' + str(row)].style = self.input_s
        self.reference['E' + str(row)] = "Total hours worked"
        row += 1
        self.reference['C' + str(row)] = "RS"
        self.reference['C' + str(row)].style = self.input_s
        self.reference['E' + str(row)] = "Return to station time."
        row += 1
        self.reference['C' + str(row)] = "MV off"
        self.reference['C' + str(row)].style = self.input_s
        self.reference['E' + str(row)] = "Time moved off own route"
        row += 1
        self.reference['C' + str(row)] = "MV on"
        self.reference['C' + str(row)].style = self.input_s
        self.reference['E' + str(row)] = "Time moved on/returned to own route"
        row += 1
        self.reference['C' + str(row)] = "Route"
        self.reference['C' + str(row)].style = self.input_s
        self.reference['E' + str(row)] = "Route of overtime/pivot"
        row += 1
        self.reference['C' + str(row)] = "MV Total"
        self.reference['C' + str(row)].style = self.input_s
        self.reference['E' + str(row)] = "Time spent on overtime/pivot off route"
        row += 1
        self.reference['C' + str(row)] = "OT own"
        self.reference['C' + str(row)].style = self.input_s
        self.reference['E' + str(row)] = "Daily overtime on the carrier's own route"
        row += 1
        self.reference['C' + str(row)] = "Off rte"
        self.reference['C' + str(row)].style = self.input_s
        self.reference['E' + str(row)] = "Total daily time spent off route"
        row += 1
        self.reference['C' + str(row)] = "OT off"
        self.reference['C' + str(row)].style = self.input_s
        self.reference['E' + str(row)] = "Daily overtime off route"
        row += 2
        self.reference['C' + str(row)] = "to 10"
        self.reference['C' + str(row)].style = self.input_s
        self.reference['E' + str(row)] = "Total availability to 10 hours"
        row += 1
        self.reference['C' + str(row)] = "to 12"
        self.reference['C' + str(row)].style = self.input_s
        self.reference['E' + str(row)] = "Total availability to 12 hours"

    def save_open(self):
        """ name and open the excel file """
        self.pbi += 1
        self.pb.move_count(self.pbi)  # increment progress bar
        self.pb.change_text("Saving...")
        self.pb.stop()
        r = "_w"
        if not projvar.invran_weekly_span:  # if investigation range is daily
            r = "_d"
        xl_filename = "kb" + str(format(self.dates[0], "_%y_%m_%d")) + r + ".xlsx"
        try:
            self.wb.save(dir_path('spreadsheets') + xl_filename)
            messagebox.showinfo("Spreadsheet generator",
                                "Your spreadsheet was successfully generated. \n"
                                "File is named: {}".format(xl_filename),
                                parent=self.frame)
            if sys.platform == "win32":
                os.startfile(dir_path('spreadsheets') + xl_filename)
            if sys.platform == "linux":
                subprocess.call(["xdg-open", 'kb_sub/spreadsheets/' + xl_filename])
            if sys.platform == "darwin":
                subprocess.call(["open", dir_path('spreadsheets') + xl_filename])
        except PermissionError:
            messagebox.showerror("Spreadsheet generator",
                                 "The spreadsheet was not opened. \n"
                                 "Suggestion: "
                                 "Make sure that identically named spreadsheets are closed "
                                 "(the file can't be overwritten while open).",
                                 parent=self.frame)


class OvermaxSpreadsheet:
    """
    This generates the 12 and 60 hour violations worksheet. This spreadsheeet is a klusterbox original and is the
    most comprehensive spreadsheet of its kind.
    """
    def __init__(self):
        self.frame = None
        self.pb = None  # progress bar object
        self.pbi = 0  # progress bar count index
        self.carrier_list = []
        self.wb = None  # workbook object
        self.violations = None  # workbook object sheet
        self.instructions = None  # workbook object sheet
        self.summary = None  # workbook object sheet
        self.startdate = None  # start of the investigation range
        self.enddate = None  # ending day of the investigation range
        self.dates = []  # all the dates of the investigation range
        self.rings = []  # all rings for all carriers in the carrier list
        self.min_rows = 0  # the minimum number of rows set by user preferences
        self.non_otdl_violation = 11.5  # Hours that non-otdl carriers work before 12/60 violation
        self.wal_12hour = None
        self.wal_12hr_mod = ""  # text inserted into formulas which varies depending on wal_12hour setting
        self.wal_dec_exempt = None  # work assignment list december exemption - true or false
        self.wal_dec_exempt_mod = ""  # text inserted into formulas which varies depending on wal_dec_exempt setting
        self.show_remedy = ""  # show the remedy on the summary sheet
        self.remedy_rate = ""  # the hourly pay rate of the remedy.
        self.remedy_tolerance = ""  # the tolerance for the remedy. e.g. ".50" - 50 clicks
        self.ws_header = None  # style
        self.date_dov = None  # style
        self.date_dov_title = None  # style
        self.remedy_style = None  # style
        self.col_header = None  # style
        self.col_center_header = None  # style
        self.vert_header = None  # style
        self.input_name = None  # style
        self.input_s = None  # style
        self.calcs = None  # style
        self.vert_calcs = None  # style
        self.instruct_text = None  # style
        self.violation_recsets = []  # carrier info, daily hours, leavetypes and leavetimes

    def create(self, frame):
        """ master method for calling methods"""
        self.frame = frame
        if not self.ask_ok():
            return
        self.pb = ProgressBarDe(label="Building Improper Mandates Spreadsheet")
        self.pb.max_count(100)  # set length of progress bar
        self.pb.start_up()  # start the progress bar
        self.pbi = 1
        self.pb.move_count(self.pbi)  # increment progress bar
        self.pb.change_text("Gathering Data... ")
        self.get_dates()
        self.get_carrierlist()
        self.get_rings()
        self.get_minrows()
        self.set_wal12hrmod()
        self.set_waldecexempt()
        self.get_styles()
        self.build_workbook()
        self.set_dimensions()
        self.build_summary()
        self.build_violations()
        self.build_instructions()
        self.violated_recs()
        self.get_pb_len()
        self.show_violations()
        self.save_open()

    def ask_ok(self):
        """ continue if user selects ok. """
        if messagebox.askokcancel("Spreadsheet generator",
                                  "Do you want to generate an Over Max Spreadsheet for violations "
                                  "of the 12 and 60 Hour Rule?",
                                  parent=self.frame):
            return True
        return False

    def get_dates(self):
        """ get the dates of the investigation range from the project variables. """
        date = projvar.invran_date_week[0]
        self.startdate = projvar.invran_date_week[0]
        self.enddate = projvar.invran_date_week[6]
        for _ in range(7):
            self.dates.append(date)
            date += timedelta(days=1)

    def get_carrierlist(self):
        """ call the carrierlist class from kbtoolbox module to get the carrier list """
        carrierlist = CarrierList(self.startdate, self.enddate, projvar.invran_station).get()
        for carrier in carrierlist:
            self.carrier_list.append(carrier[0])  # add the first record for each carrier in rec set

    def get_rings(self):
        """ get clock rings from the rings table """
        sql = "SELECT * FROM rings3 WHERE rings_date BETWEEN '%s' AND '%s' ORDER BY rings_date, carrier_name" \
              % (projvar.invran_date_week[0], projvar.invran_date_week[6])
        self.rings = inquire(sql)

    def get_minrows(self):
        """ get minimum rows and other settings. """
        sql = "SELECT tolerance FROM tolerances"
        result = inquire(sql)
        self.min_rows = int(result[14][0])
        #  get the work assignment list 12 hour violation setting -
        #  if True: violation occurs after 12 hours.
        #  if False: violation occurs after 11.50 hours
        self.wal_12hour = Convert(result[44][0]).str_to_bool()
        #  get the work assignment list december exemption setting -
        #  if True: treat wal carriers same as otdl carriers for december exemption.
        self.wal_dec_exempt = Convert(result[45][0]).str_to_bool()
        self.show_remedy = Convert(result[50][0]).str_to_bool()
        self.remedy_rate = Convert(result[51][0]).hundredths()
        self.remedy_tolerance = Convert(result[55][0]).hundredths()

    def set_wal12hrmod(self):
        """ if the wal_12hour setting is True, don't include 'wal' in formula"""
        self.wal_12hr_mod = "nl"  #
        if not self.wal_12hour:
            self.wal_12hr_mod = "wal"

    def set_waldecexempt(self):
        """ if the wal_dec_exempt is True, include 'wal' in formula for total violations.
        This will treat wal carriers same as otdl carriers durning december."""
        self.wal_dec_exempt_mod = "otdl"
        if self.wal_dec_exempt:
            self.wal_dec_exempt_mod = "wal"

    def get_styles(self):
        """ Named styles for workbook """
        bd = Side(style='thin', color="80808080")  # defines borders
        self.ws_header = NamedStyle(name="ws_header", font=Font(bold=True, name='Arial', size=12))
        self.date_dov = NamedStyle(name="date_dov", font=Font(name='Arial', size=8))
        self.date_dov_title = NamedStyle(name="date_dov_title", font=Font(bold=True, name='Arial', size=8),
                                         alignment=Alignment(horizontal='right'))
        self.remedy_style = NamedStyle(name="remedy_style", font=Font(name='Arial', size=8),
                                       alignment=Alignment(horizontal='left'))
        self.col_header = NamedStyle(name="col_header", font=Font(bold=True, name='Arial', size=8),
                                     border=Border(left=bd, right=bd, top=bd, bottom=bd),
                                     alignment=Alignment(horizontal='left'))
        self.col_center_header = NamedStyle(name="col_center_header", font=Font(bold=True, name='Arial', size=8),
                                            alignment=Alignment(horizontal='center'),
                                            border=Border(left=bd, right=bd, top=bd, bottom=bd))
        self.vert_header = NamedStyle(name="vert_header", font=Font(bold=True, name='Arial', size=8),
                                      border=Border(left=bd, right=bd, top=bd, bottom=bd),
                                      alignment=Alignment(horizontal='right', text_rotation=90))
        self.input_name = NamedStyle(name="input_name", font=Font(name='Arial', size=8),
                                     border=Border(left=bd, right=bd, top=bd, bottom=bd),
                                     alignment=Alignment(horizontal='left'))
        self.input_s = NamedStyle(name="input_s", font=Font(name='Arial', size=8),
                                  border=Border(left=bd, right=bd, top=bd, bottom=bd),
                                  alignment=Alignment(horizontal='right'))
        self.calcs = NamedStyle(name="calcs", font=Font(name='Arial', size=8),
                                border=Border(left=bd, right=bd, top=bd, bottom=bd),
                                fill=PatternFill(fgColor='e5e4e2', fill_type='solid'),
                                alignment=Alignment(horizontal='right'))
        self.vert_calcs = NamedStyle(name="vert_calcs", font=Font(name='Arial', size=8),
                                     border=Border(left=bd, right=bd, top=bd, bottom=bd),
                                     fill=PatternFill(fgColor='e5e4e2', fill_type='solid'),
                                     alignment=Alignment(horizontal='right', text_rotation=90))
        self.instruct_text = NamedStyle(name="instruct_text", font=Font(name='Arial', size=8),
                                        alignment=Alignment(horizontal='left', vertical='top'))

    def build_workbook(self):
        """ creates the workbook object """
        self.pb.change_text("Building workbook...")
        self.wb = Workbook()  # define the workbook
        self.violations = self.wb.active  # create first worksheet
        self.violations.title = "violations"  # title first worksheet
        self.violations.oddFooter.center.text = "&A"
        self.summary = self.wb.create_sheet("summary")
        self.summary.oddFooter.center.text = "&A"
        self.instructions = self.wb.create_sheet("instructions")
        self.instructions.oddFooter.center.text = "&A"

    def set_dimensions(self):
        """ adjust the height and width on the violations/ instructions page """
        sheets = (self.violations, self.instructions)
        for sheet in sheets:
            sheet.column_dimensions["A"].width = 13
            sheet.column_dimensions["B"].width = 4
            sheet.column_dimensions["C"].width = 5
            sheet.column_dimensions["D"].width = 4
            sheet.column_dimensions["E"].width = 2
            sheet.column_dimensions["F"].width = 4
            sheet.column_dimensions["G"].width = 2
            sheet.column_dimensions["H"].width = 4
            sheet.column_dimensions["I"].width = 2
            sheet.column_dimensions["J"].width = 4
            sheet.column_dimensions["K"].width = 2
            sheet.column_dimensions["L"].width = 4
            sheet.column_dimensions["M"].width = 2
            sheet.column_dimensions["N"].width = 4
            sheet.column_dimensions["O"].width = 2
            sheet.column_dimensions["P"].width = 4
            sheet.column_dimensions["Q"].width = 2
            sheet.column_dimensions["R"].width = 4
            sheet.column_dimensions['R'].hidden = True
            sheet.column_dimensions["S"].width = 5
            sheet.column_dimensions["T"].width = 5
            sheet.column_dimensions["U"].width = 2
            sheet.column_dimensions["V"].width = 2
            sheet.column_dimensions["W"].width = 2
            sheet.column_dimensions["X"].width = 5

    def build_summary(self):
        """ summary worksheet - format cells """
        self.pb.change_text("Building summary")
        self.summary.merge_cells('A1:R1')
        self.summary['A1'] = "12 and 60 Hour Violations Summary"
        self.summary['A1'].style = self.ws_header
        self.summary.column_dimensions["A"].width = 15
        self.summary.column_dimensions["B"].width = 8
        self.summary['A3'] = "Date: "
        self.summary['A3'].style = self.date_dov_title
        self.summary.merge_cells('B3:D3')  # blank field for date
        self.summary['B3'] = self.dates[0].strftime("%x") + " - " + self.dates[6].strftime("%x")
        self.summary['B3'].style = self.date_dov
        self.summary.merge_cells('K3:N3')
        self.summary['F3'] = "Pay Period: "  # Pay Period Header
        self.summary['F3'].style = self.date_dov_title
        self.summary.merge_cells('G3:I3')  # blank field for pay period
        self.summary['G3'] = projvar.pay_period
        self.summary['G3'].style = self.date_dov
        self.summary['A4'] = "Station: "  # Station Header
        self.summary['A4'].style = self.date_dov_title
        self.summary.merge_cells('B4:D4')  # blank field for station
        self.summary['B4'] = projvar.invran_station
        self.summary['B4'].style = self.date_dov
        self.summary.merge_cells('E4:F4')  # Remedy Rate (optional)
        self.summary['E4'] = "Remedy Tolerance: "
        self.summary['E4'].style = self.date_dov_title
        self.summary['G4'] = float(self.remedy_tolerance)
        self.summary['G4'].style = self.remedy_style
        self.summary['G4'].number_format = "#,##0.00;[RED]-#,##0.00"
        if self.show_remedy:
            self.summary.merge_cells('E5:F5')  # Remedy Rate (optional)
            self.summary['E5'] = "Remedy Rate: "
            self.summary['E5'].style = self.date_dov_title
            self.summary['G5'] = float(self.remedy_rate)
            self.summary['G5'].style = self.remedy_style
        self.summary['G5'].number_format = "[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00"
        self.summary['A7'] = "name"
        self.summary['A7'].style = self.col_center_header
        self.summary['B7'] = "violation"
        self.summary['B7'].style = self.col_center_header
        if self.show_remedy:
            self.summary['C7'] = "remedy"
            self.summary['C7'].style = self.col_center_header

    def build_violations(self):
        """ self.violations worksheet - format cells """
        self.pb.change_text("Building violations...")
        for x in (2, 3, 4, 5, 7, 8, 9):
            self.violations.row_dimensions[x].height = 10  # adjust all row height
        self.violations.row_dimensions[6].height = 30  # adjust all row height
        self.violations.merge_cells('A1:R1')
        self.violations['A1'] = "12 and 60 Hour Violations Worksheet"
        self.violations['A1'].style = self.ws_header
        self.violations['A3'] = "Date:"
        self.violations['A3'].style = self.date_dov_title
        self.violations.merge_cells('B3:J3')  # blank field for date
        self.violations['B3'] = self.dates[0].strftime("%x") + " - " + self.dates[6].strftime("%x")
        self.violations['B3'].style = self.date_dov
        self.violations.merge_cells('K3:N3')
        self.violations['K3'] = "Pay Period:"
        self.violations['k3'].style = self.date_dov_title
        self.violations.merge_cells('O3:S3')  # blank field for pay period
        self.violations['O3'] = projvar.pay_period
        self.violations['O3'].style = self.date_dov
        self.violations['A4'] = "Station:"
        self.violations['A4'].style = self.date_dov_title
        self.violations.merge_cells('B4:J4')  # blank field for station
        self.violations['B4'] = projvar.invran_station
        self.violations['B4'].style = self.date_dov
        self.violations.merge_cells('D7:Q7')
        self.violations.merge_cells('K4:N4')
        self.violations['K4'] = "Dec Exception:"  # December exception
        self.violations['k4'].style = self.date_dov_title
        self.violations.merge_cells('O4:S4')  # blank field for pay period
        self.violations['O4'] = "no"  # enter yes or no
        self.violations['O4'].style = self.date_dov
        self.violations['D7'] = "Daily Paid Leave times with type"
        self.violations['D7'].style = self.col_center_header
        self.violations.merge_cells('D8:Q8')
        self.violations['D8'] = "Daily 5200 times"
        self.violations['D8'].style = self.col_center_header
        self.violations['A9'] = "name"
        self.violations['A9'].style = self.col_header
        self.violations['B9'] = "list"
        self.violations['B9'].style = self.col_header
        self.violations.merge_cells('C5:C9')
        self.violations['C5'] = "Weekly \n5200"
        self.violations['C5'].style = self.vert_header
        self.violations.merge_cells('D9:E9')
        self.violations['D9'] = "sat"
        self.violations['D9'].style = self.col_center_header
        self.violations.merge_cells('F9:G9')
        self.violations['F9'] = "sun"
        self.violations['F9'].style = self.col_center_header
        self.violations.merge_cells('H9:I9')
        self.violations['H9'] = "mon"
        self.violations['H9'].style = self.col_center_header
        self.violations.merge_cells('J9:K9')
        self.violations['J9'] = "tue"
        self.violations['J9'].style = self.col_center_header
        self.violations.merge_cells('L9:M9')
        self.violations['L9'] = "wed"
        self.violations['L9'].style = self.col_center_header
        self.violations.merge_cells('N9:O9')
        self.violations['N9'] = "thr"
        self.violations['N9'].style = self.col_center_header
        self.violations.merge_cells('P9:Q9')
        self.violations['P9'] = "fri"
        self.violations['P9'].style = self.col_center_header
        self.violations.merge_cells('S5:S9')
        self.violations['S5'] = " Weekly\nViolation"
        self.violations['S5'].style = self.vert_header
        self.violations.merge_cells('T5:T9')
        self.violations['T5'] = "Daily\nViolation"
        self.violations['T5'].style = self.vert_header
        self.violations.merge_cells('U5:U9')
        self.violations['U5'] = "Wed Adj"
        self.violations['U5'].style = self.vert_header
        self.violations.merge_cells('V5:V9')
        self.violations['V5'] = "Thr Adj"
        self.violations['V5'].style = self.vert_header
        self.violations.merge_cells('W5:W9')
        self.violations['W5'] = "Fri Adj"
        self.violations['W5'].style = self.vert_header
        self.violations.merge_cells('X5:X9')
        self.violations['X5'] = "Total\nViolation"
        self.violations['X5'].style = self.vert_header

    def build_instructions(self):
        """ format the instructions cells """
        self.pb.change_text("Building instructions")
        self.instructions.merge_cells('A1:R1')
        self.instructions['A1'] = "12 and 60 Hour Violations Instructions"
        self.instructions['A1'].style = self.ws_header
        self.instructions.row_dimensions[3].height = 290
        self.instructions['A3'].style = self.instruct_text
        self.instructions.merge_cells('A3:X3')
        self.instructions['A3'] = \
            "Caution for Mac Users: \n" \
            "Using the Apple Numbers Spreadsheet program is not recommended. Apple Numbers " \
            "does not support vertical text or hidden fields, both of which are used in the " \
            "12 and 60 Hour Violations Spreadsheet. If you are using Mac, you can download " \
            "Libre Office Calc, which is recommended, for free. Microsoft Excel or Google Docs " \
            "will also work properly. \n\n" \
            "December Exemption Setting: \n" \
            "Enter \"yes\" in this cell (use lowercase only) to exempt otdl carriers from " \
            "violations during the month of December. The default is \"no\". " \
            "Turning WAL December Exemption to \'on\' in Spreadsheet Setting in Klusterbox " \
            "will modify the formulas to include \"wal\" carriers in the exemption.\n" \
            "\tWAL 12 Hour Violation Setting is {}\n\n" \
            "Instructions: \n" \
            "1. Fill in the name \n" \
            "2. Fill in the list. Enter either \"otdl\",\"wal\",\"nl\",\"aux\" or \"ptf\" in list " \
            "columns. Use only lowercase. \n" \
            "   If you do not enter anything, the default is \"otdl\". \n" \
            "\totdl = overtime desired list\n" \
            "\twal = work assignment list\n" \
            "\tnl = no list \n" \
            "\taux = auxiliary (this would be a cca or city carrier assistant).\n" \
            "\tptf = part time flexible \n" \
            "3. Fill in the weekly 5200 time in field C if it exceeds 60 hours " \
            "or if the sum of all daily non 5200 times (all fields D) plus \n" \
            "   the weekly 5200 time (field C) will  exceed 60 hours.\n" \
            "4. Fill in any daily non 5200 times and types in fields D and E. " \
            "Enter only paid leave types such as sick leave, annual\n" \
            "   leave and holiday leave. Do not enter unpaid leave types such as LWOP " \
            "(leave without pay) or AWOL (absent \n" \
            "   without leave).\n" \
            "5. Fill in any daily 5200 times which exceed 12 hours for otdl carriers " \
            "or 11.50 hours for any other carrier in fields F.\n" \
            "   Failing to fill out the daily values for Wednesday, Thursday and Friday " \
            "could cause errors in calculating the adjustments,\n" \
            "   so fill those in.\n" \
            "6. The gray fields will fill automatically. Do not enter an information in " \
            "these fields as it will delete the formulas.\n" \
            "7. Field O will show the violation in hours which you should seek a remedy " \
            "for. \n".format(Convert(self.wal_12hour).bool_to_onoff())
        self.instructions['A3'].alignment = Alignment(wrap_text=True, vertical='top')
        self.instructions.row_dimensions[4].height = 10
        self.instructions.row_dimensions[5].height = 30
        for x in range(6, 12):
            self.instructions.row_dimensions[x].height = 10  # adjust all row height
        self.instructions.merge_cells('D6:Q6')
        self.instructions['D6'] = "Daily Paid Leave times with type"
        self.instructions['D6'].style = self.col_center_header
        self.instructions.merge_cells('D7:Q7')
        self.instructions['D7'] = "Daily 5200 times"
        self.instructions['D7'].style = self.col_center_header
        self.instructions['A8'] = "name"
        self.instructions['A8'].style = self.col_header
        self.instructions['B8'] = "list"
        self.instructions['B8'].style = self.col_header
        self.instructions.merge_cells('C4:C8')
        self.instructions['C4'] = "Weekly \n5200"
        self.instructions['C4'].style = self.vert_header
        self.instructions.merge_cells('D8:E8')
        self.instructions['D8'] = "sat"
        self.instructions['D8'].style = self.col_center_header
        self.instructions.merge_cells('F8:G8')
        self.instructions['F8'] = "sun"
        self.instructions['F8'].style = self.col_center_header
        self.instructions.merge_cells('H8:I8')
        self.instructions['H8'] = "mon"
        self.instructions['H8'].style = self.col_center_header
        self.instructions.merge_cells('J8:K8')
        self.instructions['J8'] = "tue"
        self.instructions['J8'].style = self.col_center_header
        self.instructions.merge_cells('L8:M8')
        self.instructions['L8'] = "wed"
        self.instructions['L8'].style = self.col_center_header
        self.instructions.merge_cells('N8:O8')
        self.instructions['N8'] = "thr"
        self.instructions['N8'].style = self.col_center_header
        self.instructions.merge_cells('P8:Q8')
        self.instructions['P8'] = "fri"
        self.instructions['P8'].style = self.col_center_header
        self.instructions.merge_cells('S4:S8')
        self.instructions['S4'] = " Weekly\nViolation"
        self.instructions['S4'].style = self.vert_header
        self.instructions.merge_cells('T4:T8')
        self.instructions['T4'] = "Daily\nViolation"
        self.instructions['T4'].style = self.vert_header
        self.instructions.merge_cells('U4:U8')
        self.instructions['U4'] = "Wed Adj"
        self.instructions['U4'].style = self.vert_header
        self.instructions.merge_cells('V4:V8')
        self.instructions['V4'] = "Thr Adj"
        self.instructions['V4'].style = self.vert_header
        self.instructions.merge_cells('W4:W8')
        self.instructions['W4'] = "Fri Adj"
        self.instructions['W4'].style = self.vert_header
        self.instructions.merge_cells('X4:X8')
        self.instructions['X4'] = "Total\nViolation"
        self.instructions['X4'].style = self.vert_header
        self.instructions['A9'] = "A"
        self.instructions['A9'].style = self.col_center_header
        self.instructions['B9'] = "B"
        self.instructions['B9'].style = self.col_center_header
        self.instructions['C9'] = "C"
        self.instructions['C9'].style = self.col_center_header
        self.instructions['D9'] = "D"
        self.instructions['D9'].style = self.col_center_header
        self.instructions['E9'] = "E"
        self.instructions['E9'].style = self.col_center_header
        self.instructions['F9'] = "G"
        self.instructions['F9'].style = self.col_center_header
        self.instructions.merge_cells('F9:G9')
        self.instructions['H9'] = "D"
        self.instructions['H9'].style = self.col_center_header
        self.instructions['I9'] = "E"
        self.instructions['I9'].style = self.col_center_header
        self.instructions['J9'] = "D"
        self.instructions['J9'].style = self.col_center_header
        self.instructions['K9'] = "E"
        self.instructions['K9'].style = self.col_center_header
        self.instructions['L9'] = "D"
        self.instructions['L9'].style = self.col_center_header
        self.instructions['M9'] = "E"
        self.instructions['M9'].style = self.col_center_header
        self.instructions['N9'] = "D"
        self.instructions['N9'].style = self.col_center_header
        self.instructions['O9'] = "E"
        self.instructions['O9'].style = self.col_center_header
        self.instructions['P9'] = "D"
        self.instructions['P9'].style = self.col_center_header
        self.instructions['Q9'] = "E"
        self.instructions['Q9'].style = self.col_center_header
        self.instructions['S9'] = "J"
        self.instructions['S9'].style = self.col_center_header
        self.instructions['T9'] = "K"
        self.instructions['T9'].style = self.col_center_header
        self.instructions['U9'] = "L"
        self.instructions['U9'].style = self.col_center_header
        self.instructions['V9'] = "M"
        self.instructions['V9'].style = self.col_center_header
        self.instructions['W9'] = "N"
        self.instructions['W9'].style = self.col_center_header
        self.instructions['X9'] = "O"
        self.instructions['X9'].style = self.col_center_header
        i = 10
        # instructions name
        self.instructions.merge_cells('A' + str(i) + ':A' + str(i + 1))  # merge box for name
        self.instructions['A10'] = "kubrick, s"
        self.instructions['A10'].style = self.input_name
        self.instructions['A11'].style = self.input_name
        # instructions list
        self.instructions.merge_cells('B' + str(i) + ':B' + str(i + 1))  # merge box for list type input
        self.instructions['B10'] = "wal"
        self.instructions['B10'].style = self.input_s
        self.instructions['B11'].style = self.input_s
        # instructions weekly
        self.instructions.merge_cells('C' + str(i) + ':C' + str(i + 1))  # merge box for weekly input
        self.instructions['C10'] = 75.00
        self.instructions['C10'].style = self.input_s
        self.instructions['C11'].style = self.input_s
        self.instructions['C10'].number_format = "#,###.00;[RED]-#,###.00"
        # instructions saturday
        self.instructions.merge_cells('D' + str(i + 1) + ':E' + str(i + 1))  # merge box for sat 5200
        self.instructions['D' + str(i)] = ""  # leave time
        self.instructions['D' + str(i)].style = self.input_s
        self.instructions['D' + str(i)].number_format = "#,###.00;[RED]-#,###.00"
        self.instructions['E' + str(i)] = ""  # leave type
        self.instructions['E' + str(i)].style = self.input_s
        self.instructions['D' + str(i + 1)] = 13.00  # 5200 time
        self.instructions['D' + str(i + 1)].style = self.input_s
        self.instructions['D' + str(i + 1)].number_format = "#,###.00;[RED]-#,###.00"
        # instructions sunday
        self.instructions.merge_cells('F' + str(i + 1) + ':G' + str(i + 1))  # merge box for sun 5200
        self.instructions['F' + str(i)] = ""  # leave time
        self.instructions['F' + str(i)].style = self.input_s
        self.instructions['F' + str(i)].number_format = "#,###.00;[RED]-#,###.00"
        self.instructions['G' + str(i)] = ""  # leave type
        self.instructions['G' + str(i)].style = self.input_s
        self.instructions['F' + str(i + 1)] = ""  # 5200 time
        self.instructions['F' + str(i + 1)].style = self.input_s
        self.instructions['F' + str(i + 1)].number_format = "#,###.00;[RED]-#,###.00"
        # instructions monday
        self.instructions.merge_cells('H' + str(i + 1) + ':I' + str(i + 1))  # merge box for mon 5200
        self.instructions['H' + str(i)] = 8  # leave time
        self.instructions['H' + str(i)].style = self.input_s
        self.instructions['H' + str(i)].number_format = "#,###.00;[RED]-#,###.00"
        self.instructions['I' + str(i)] = "h"  # leave type
        self.instructions['I' + str(i)].style = self.input_s
        self.instructions['H' + str(i + 1)] = ""  # 5200 time
        self.instructions['H' + str(i + 1)].style = self.input_s
        self.instructions['H' + str(i + 1)].number_format = "#,###.00;[RED]-#,###.00"
        # instructions tuesday
        self.instructions.merge_cells('J' + str(i + 1) + ':K' + str(i + 1))  # merge box for tue 5200
        self.instructions['J' + str(i)] = ""  # leave time
        self.instructions['J' + str(i)].style = self.input_s
        self.instructions['J' + str(i)].number_format = "#,###.00;[RED]-#,###.00"
        self.instructions['K' + str(i)] = ""  # leave type
        self.instructions['K' + str(i)].style = self.input_s
        self.instructions['J' + str(i + 1)] = 14  # 5200 time
        self.instructions['J' + str(i + 1)].style = self.input_s
        self.instructions['J' + str(i + 1)].number_format = "#,###.00;[RED]-#,###.00"
        # instructions wednesday
        self.instructions.merge_cells('L' + str(i + 1) + ':M' + str(i + 1))  # merge box for wed 5200
        self.instructions['L' + str(i)] = ""  # leave time
        self.instructions['L' + str(i)].style = self.input_s
        self.instructions['L' + str(i)].number_format = "#,###.00;[RED]-#,###.00"
        self.instructions['M' + str(i)] = ""  # leave type
        self.instructions['M' + str(i)].style = self.input_s
        self.instructions['L' + str(i + 1)] = 14  # 5200 time
        self.instructions['L' + str(i + 1)].style = self.input_s
        self.instructions['M' + str(i + 1)].number_format = "#,###.00;[RED]-#,###.00"
        # instructions thursday
        self.instructions.merge_cells('N' + str(i + 1) + ':O' + str(i + 1))  # merge box for thr 5200
        self.instructions['N' + str(i)] = ""  # leave time
        self.instructions['N' + str(i)].style = self.input_s
        self.instructions['N' + str(i)].number_format = "#,###.00;[RED]-#,###.00"
        self.instructions['O' + str(i)] = ""  # leave type
        self.instructions['O' + str(i)].style = self.input_s
        self.instructions['N' + str(i + 1)] = 13  # 5200 time
        self.instructions['N' + str(i + 1)].style = self.input_s
        self.instructions['N' + str(i + 1)].number_format = "#,###.00;[RED]-#,###.00"
        # instructions friday
        self.instructions.merge_cells('P' + str(i + 1) + ':Q' + str(i + 1))  # merge box for fri 5200
        self.instructions['P' + str(i)] = ""  # leave time
        self.instructions['P' + str(i)].style = self.input_s
        self.instructions['P' + str(i)].number_format = "#,###.00;[RED]-#,###.00"
        self.instructions['Q' + str(i)] = ""  # leave type
        self.instructions['Q' + str(i)].style = self.input_s
        self.instructions['P' + str(i + 1)] = 13  # 5200 time
        self.instructions['P' + str(i + 1)].style = self.input_s
        self.instructions['P' + str(i + 1)].number_format = "#,###.00;[RED]-#,###.00"
        # instructions hidden columns
        page = "instructions"
        formula = "=SUM(%s!D%s:P%s)+%s!D%s + %s!H%s + %s!J%s + %s!L%s + " \
                  "%s!N%s + %s!P%s" % (page, str(i + 1), str(i + 1),
                                       page, str(i), page, str(i), page, str(i),
                                       page, str(i), page, str(i), page, str(i))
        self.instructions['R' + str(i)] = formula
        self.instructions['R' + str(i)].style = self.calcs
        self.instructions['R' + str(i)].number_format = "#,###.00;[RED]-#,###.00"
        formula = "=SUM(%s!C%s+%s!D%s+%s!H%s+%s!J%s+%s!L%s+%s!N%s+%s!P%s)" % \
                  (page, str(i), page, str(i), page, str(i),
                   page, str(i), page, str(i), page, str(i),
                   page, str(i))
        self.instructions['R' + str(i + 1)] = formula
        self.instructions['R' + str(i + 1)].style = self.calcs
        self.instructions['R' + str(i + 1)].number_format = "#,###.00;[RED]-#,###.00"
        # instructions weekly self.violations
        self.instructions.merge_cells('S' + str(i) + ':S' + str(i + 1))  # merge box for weekly violation
        formula = "=IF(OR(%s!B%s = \"aux\",%s!B%s = \"ptf\"),0," \
                  "MAX(IF(%s!R%s>%s!R%s,MAX(%s!R%s-60,0),MAX(%s!R%s-60)),0))" \
                  % (page, str(i), page, str(i), page, str(i), page, str(i + 1), page, str(i),
                     page, str(i + 1))
        self.instructions['S10'] = formula
        self.instructions['S10'].style = self.calcs
        self.instructions['S11'].style = self.calcs
        self.instructions['S10'].number_format = "#,###.00;[RED]-#,###.00"
        # instructions daily self.violations
        formula_d = "=IF(OR(%s!B%s=\"%s\",%s!B%s=\"nl\",%s!B%s=\"ptf\",%s!B%s=\"aux\")," \
                    "(SUM(IF(%s!D%s>%s,%s!D%s-%s,0)+IF(%s!H%s>%s,%s!H%s-%s,0)+IF(%s!J%s>%s,%s!J%s-%s,0)" \
                    "+IF(%s!L%s>%s,%s!L%s-%s,0)+IF(%s!N%s>%s,%s!N%s-%s,0)+IF(%s!P%s>%s,%s!P%s-%s,0)))," \
                    "(SUM(IF(%s!D%s>12,%s!D%s-12,0)+IF(%s!H%s>12,%s!H%s-12,0)+IF(%s!J%s>12,%s!J%s-12,0)" \
                    "+IF(%s!L%s>12,%s!L%s-12,0)+IF(%s!N%s>12,%s!N%s-12,0)+IF(%s!P%s>12,%s!P%s-12,0))))" \
                    % (page, str(i), self.wal_12hr_mod,
                       page, str(i), page, str(i), page, str(i),
                       page, str(i + 1), str(self.non_otdl_violation),
                       page, str(i + 1), str(self.non_otdl_violation),
                       page, str(i + 1), str(self.non_otdl_violation),
                       page, str(i + 1), str(self.non_otdl_violation),
                       page, str(i + 1), str(self.non_otdl_violation),
                       page, str(i + 1), str(self.non_otdl_violation),
                       page, str(i + 1), str(self.non_otdl_violation),
                       page, str(i + 1), str(self.non_otdl_violation),
                       page, str(i + 1), str(self.non_otdl_violation),
                       page, str(i + 1), str(self.non_otdl_violation),
                       page, str(i + 1), str(self.non_otdl_violation),
                       page, str(i + 1), str(self.non_otdl_violation),
                       page, str(i + 1),
                       page, str(i + 1), page, str(i + 1),
                       page, str(i + 1), page, str(i + 1), page, str(i + 1),
                       page, str(i + 1), page, str(i + 1), page, str(i + 1),
                       page, str(i + 1), page, str(i + 1), page, str(i + 1))
        self.instructions['T' + str(i)] = formula_d
        self.instructions.merge_cells('T' + str(i) + ':T' + str(i + 1))  # merge box for daily violation
        self.instructions['T' + str(i)].style = self.calcs
        self.instructions['T' + str(i+1)].style = self.calcs
        self.instructions['T' + str(i)].number_format = "#,###.00"
        # instructions wed adjustment
        self.instructions.merge_cells('U' + str(i) + ':U' + str(i + 1))  # merge box for wed adj
        formula_e = "=IF(OR(%s!B%s=\"%s\",%s!B%s=\"nl\",%s!B%s=\"ptf\",%s!B%s=\"aux\")," \
                    "IF(AND(%s!S%s-(%s!N%s+%s!N%s+%s!P%s+%s!P%s)>0,%s!L%s>%s)," \
                    "IF(%s!S%s-(%s!N%s+%s!N%s+%s!P%s+%s!P%s)>%s!L%s-%s,%s!L%s-%s,%s!S%s-" \
                    "(%s!N%s+%s!N%s+%s!P%s+%s!P%s)),0)," \
                    "IF(AND(%s!S%s-(%s!N%s+%s!N%s+%s!P%s+%s!P%s)>0,%s!L%s>12)," \
                    "IF(%s!S%s-(%s!N%s+%s!N%s+%s!P%s+%s!P%s)>%s!L%s-12,%s!L%s-12,%s!S%s-" \
                    "(%s!N%s+%s!N%s+%s!P%s+%s!P%s)),0))" \
                    % (page, str(i), self.wal_12hr_mod,
                       page, str(i), page, str(i), page, str(i),
                       page, str(i), page, str(i + 1), page, str(i),
                       page, str(i + 1), page, str(i), page, str(i + 1),
                       str(self.non_otdl_violation),
                       page, str(i), page, str(i + 1), page, str(i),
                       page, str(i + 1), page, str(i), page, str(i + 1),
                       str(self.non_otdl_violation),
                       page, str(i + 1), str(self.non_otdl_violation),
                       page, str(i), page, str(i + 1),
                       page, str(i), page, str(i + 1), page, str(i),
                       page, str(i), page, str(i + 1), page, str(i),
                       page, str(i + 1), page, str(i), page, str(i + 1),
                       page, str(i), page, str(i + 1), page, str(i),
                       page, str(i + 1), page, str(i), page, str(i + 1),
                       page, str(i + 1), page, str(i), page, str(i + 1),
                       page, str(i), page, str(i + 1), page, str(i))
        self.instructions['U' + str(i)] = formula_e
        self.instructions['U' + str(i)].style = self.vert_calcs
        self.instructions['U' + str(i+1)].style = self.vert_calcs
        self.instructions['U' + str(i)].number_format = "#,###.00"
        # instructions thr adjustment
        formula_f = "=IF(OR(%s!B%s=\"%s\",%s!B%s=\"nl\",%s!B%s=\"ptf\",%s!B%s=\"aux\")," \
                    "IF(AND(%s!S%s-(%s!P%s+%s!P%s)>0,%s!N%s>%s)," \
                    "IF(%s!S%s-(%s!P%s+%s!P%s)>%s!N%s-%s,%s!N%s-%s,%s!S%s-(%s!P%s+%s!P%s)),0)," \
                    "IF(AND(%s!S%s-(%s!P%s+%s!P%s)>0,%s!N%s>12)," \
                    "IF(%s!S%s-(%s!P%s+%s!P%s)>%s!N%s-12,%s!N%s-12,%s!S%s-(%s!P%s+%s!P%s)),0))" \
                    % (page, str(i), self.wal_12hr_mod,
                       page, str(i), page, str(i), page, str(i),
                       page, str(i), page, str(i + 1), page, str(i),
                       page, str(i + 1), str(self.non_otdl_violation),
                       page, str(i), page, str(i + 1), page, str(i),
                       page, str(i + 1), str(self.non_otdl_violation),
                       page, str(i + 1), str(self.non_otdl_violation),
                       page, str(i), page, str(i + 1), page, str(i),
                       page, str(i), page, str(i + 1), page, str(i),
                       page, str(i + 1), page, str(i), page, str(i + 1), page, str(i),
                       page, str(i + 1), page, str(i + 1), page, str(i),
                       page, str(i + 1), page, str(i)
                       )
        self.instructions.merge_cells('V' + str(i) + ':V' + str(i + 1))  # merge box for thr adj
        self.instructions['V' + str(i)] = formula_f
        self.instructions['V' + str(i)].style = self.vert_calcs
        self.instructions['V' + str(i+1)].style = self.vert_calcs
        self.instructions['V' + str(i)].number_format = "#,###.00"
        # instructions fri adjustment
        self.instructions.merge_cells('W' + str(i) + ':W' + str(i + 1))  # merge box for fri adj
        formula_g = "=IF(OR(%s!B%s=\"%s\",%s!B%s=\"nl\",%s!B%s=\"aux\",%s!B%s=\"ptf\")," \
                    "IF(AND(%s!S%s>0,%s!P%s>%s)," \
                    "IF(%s!S%s>%s!P%s-%s,%s!P%s-%s,%s!S%s),0)," \
                    "IF(AND(%s!S%s>0,%s!P%s>12)," \
                    "IF(%s!S%s>%s!P%s-12,%s!P%s-12,%s!S%s),0))" \
                    % (page, str(i), self.wal_12hr_mod,
                       page, str(i), page, str(i), page, str(i),
                       page, str(i), page, str(i + 1), str(self.non_otdl_violation),
                       page, str(i), page, str(i + 1), str(self.non_otdl_violation),
                       page, str(i + 1), str(self.non_otdl_violation),
                       page, str(i), page, str(i), page, str(i + 1), page, str(i),
                       page, str(i + 1), page, str(i + 1), page, str(i))
        self.instructions['W' + str(i)] = formula_g
        self.instructions['W' + str(i)].style = self.vert_calcs
        self.instructions['W' + str(i+1)].style = self.vert_calcs
        self.instructions['W' + str(i)].number_format = "#,###.00"
        # instructions total violation
        self.instructions.merge_cells('X' + str(i) + ':X' + str(i + 1))  # merge box for total violation
        formula_h = "=SUM(%s!S%s:T%s)-(%s!U%s+%s!V%s+%s!W%s)" \
                    % (page, str(i), str(i), page, str(i),
                       page, str(i), page, str(i))
        self.instructions['X' + str(i)] = formula_h
        self.instructions['X' + str(i)].style = self.calcs
        self.instructions['X' + str(i+1)].style = self.calcs
        self.instructions['X' + str(i)].number_format = "#,###.00"
        self.instructions['D12'] = "F"
        self.instructions['D12'].style = self.col_center_header
        self.instructions.merge_cells('D12:E12')
        self.instructions['F12'] = "F"
        self.instructions['F12'].style = self.col_center_header
        self.instructions.merge_cells('F12:G12')
        self.instructions['H12'] = "F"
        self.instructions['H12'].style = self.col_center_header
        self.instructions.merge_cells('H12:I12')
        self.instructions['J12'] = "F"
        self.instructions['J12'].style = self.col_center_header
        self.instructions.merge_cells('J12:K12')
        self.instructions['L12'] = "F"
        self.instructions['L12'].style = self.col_center_header
        self.instructions.merge_cells('L12:M12')
        self.instructions['N12'] = "F"
        self.instructions['N12'].style = self.col_center_header
        self.instructions.merge_cells('N12:O12')
        self.instructions['P12'] = "F"
        self.instructions['P12'].style = self.col_center_header
        self.instructions.merge_cells('P12:Q12')
        # legend section
        # create text for daily violations:
        text_k_a = "wal, "
        text_k_b = ""
        if self.wal_12hr_mod:
            text_k_a = ""
            text_k_b = " and wal"
        self.instructions.row_dimensions[14].height = 190
        self.instructions['A14'].style = self.instruct_text
        self.instructions.merge_cells('A14:X14')
        self.instructions['A14'] = \
            "Legend: \n" \
            "A.  Name \n" \
            "B.  List: Either otdl, wal, nl, ptf or aux (always use lowercase to preserve " \
            "operation of the formulas).\n" \
            "C.  Weekly 5200 Time: Enter the 5200 time for the week. \n" \
            "D.  Daily Non 5200 Time: Enter daily hours for either holiday, annual sick leave or " \
            "other type of paid leave.\n" \
            "E.  Daily Non 5200 Type: Enter “a” for annual, “s” for sick, “h” for holiday, etc. \n" \
            "F.  Daily 5200 Hours: Enter 5200 hours or hours worked for the day. \n" \
            "G.  No value allowed: No non 5200 times allowed for Sundays.\n" \
            "J.  Weekly Violations: This is the total of self.violations over 60 hours in a week.\n" \
            "K.  Daily Violations: This is the total of daily violations which have exceeded 11.50 " \
            "(for {}nl, ptf or aux)\n" \
            "     or 12 hours in a day (for otdl{}).\n" \
            "L.  Wednesday Adjustment: In cases were the 60 hour limit is reached " \
            "and a daily violation happens (on Wednesday),\n" \
            "     this column deducts one of the violations so to provide a correct remedy.\n" \
            "M.  Thursday Adjustment: In cases were the 60 hour limit is reached and " \
            "a daily violation happens (on Thursday), \n" \
            "     this column deducts one of the violations so to provide a correct remedy.\n" \
            "N.  Friday Adjustment: In cases were the 60 hour limit is reached and " \
            "a daily violation happens (on Friday),\n" \
            "     this column deducts one of the violations so to provide a correct remedy.\n" \
            "O.  Total Violation: This field is the end result of the calculation. " \
            "This is the addition of the total daily  " \
            "violations and the\n" \
            "     weekly violation, it shows the sum of the two. " \
            "This is the value which the steward should seek a remedy for.".format(text_k_a, text_k_b)
        self.instructions['A14'].alignment = Alignment(wrap_text=True, vertical='top')

    def violated_recs(self):
        """
        The violation record set is appended if the carrier has a daily violation or a weekly violation of
        over 60 hours in a week. It consist of 4 arrays: 1. carrier info (name and list), 2. daily hours array,
        3. daily leavetypes and 4 daily leavetimes. The carrier list the status on Saturday.
        """
        twelvehourlimit = ("otdl",)  # the 12 hour limit only applies to otdl carriers
        if self.wal_12hour:  # unless the WAL 12 Hour Violation setting is "on"
            twelvehourlimit = ("otdl", "wal")  # then the 12 hour limit applies to otdl and wal carriers
        i = 0
        while i <= len(self.carrier_list)-1:
            totals_array = ["", "", "", "", "", "", ""]  # daily hours
            leavetype_array = ["", "", "", "", "", "", ""]  # daily leave types
            leavetime_array = ["", "", "", "", "", "", ""]  # daily leave times
            carrier_rings = []
            total = 0.0
            grandtotal = 0.0
            # carrier name, list status for Saturday, and total weekly hours worked
            carrier_array = [self.carrier_list[i][1], self.carrier_list[i][2], 0.0]
            cc = 0
            daily_violation = False
            for day in self.dates:
                for ring in self.rings:
                    if ring[0] == str(day) and ring[1] == self.carrier_list[i][1]:  # find rings for carrier
                        carrier_rings.append(ring)  # add any rings to an array
                        if isfloat(ring[2]):
                            totals_array[cc] = float(ring[2])  # if hours worked is a number, add it as a number
                            if float(ring[2]) > 12 and self.carrier_list[i][2] in twelvehourlimit:
                                daily_violation = True
                            if float(ring[2]) > self.non_otdl_violation and \
                                    self.carrier_list[i][2] not in twelvehourlimit:
                                daily_violation = True
                        else:
                            totals_array[cc] = ring[2]  # if hours worked is empty string, add empty string
                        if ring[6] == "annual":
                            leavetype_array[cc] = "A"
                        if ring[6] == "sick":
                            leavetype_array[cc] = "S"
                        if ring[6] == "holiday":
                            leavetype_array[cc] = "H"
                        if ring[6] == "other":
                            leavetype_array[cc] = "O"
                        if ring[7] == "0.0" or ring[7] == "0":
                            leavetime_array[cc] = ""
                        elif isfloat(ring[7]):
                            leavetime_array[cc] = float(ring[7])
                        else:
                            leavetime_array[cc] = ring[7]
                cc += 1
            for item in carrier_rings:
                if item[2] == "":  # convert empty 5200 strings to zero
                    t = 0.0
                else:
                    t = float(item[2])
                if item[7] == "":  # convert leave time strings to zero
                    lv = 0.0
                else:
                    lv = float(item[7])
                total += t
                grandtotal = grandtotal + t + lv
            carrier_array[2] = total  # append total weekly hours worked to carrier array
            if grandtotal > 60 or daily_violation:  # only append violation recset, if there has been a violation
                violation_recset = [carrier_array, totals_array, leavetype_array, leavetime_array]  # build recset
                self.violation_recsets.append(violation_recset)  # append violations record set
            i += 1
        while len(self.violation_recsets) < self.min_rows:  # if minimum rows haven't been reached
            carrier_array = ["", "", 0.0]  # carrier information plus total weekly hours worked
            totals_array = ["", "", "", "", "", "", ""]  # daily hours
            leavetype_array = ["", "", "", "", "", "", ""]  # daily leave types
            leavetime_array = ["", "", "", "", "", "", ""]  # daily leave times
            violation_recset = [carrier_array, totals_array, leavetype_array, leavetime_array]  # combine
            self.violation_recsets.append(violation_recset)  # append these empty recs into the violations rec sets

    def get_pb_len(self):
        """ get the lenght of the progress bar. """
        self.pb.max_count(len(self.violation_recsets))  # set length of progress bar

    def show_violations(self):
        """ generates the rows of the violations and the summary worksheets. """
        summary_i = 8
        i = 10
        for line in self.violation_recsets:
            carrier_name = line[0][0]
            self.pbi += 1
            self.pb.move_count(self.pbi)  # increment progress bar
            self.pb.change_text("Building display for {}".format(carrier_name))
            carrier_list = line[0][1]
            total = line[0][2]
            totals_array = line[1]
            leavetype_array = line[2]
            leavetime_array = line[3]
            self.violations.row_dimensions[i].height = 10  # adjust all row height
            self.violations.row_dimensions[i + 1].height = 10
            self.violations.merge_cells('A' + str(i) + ':A' + str(i + 1))
            self.violations['A' + str(i)] = carrier_name  # name
            self.violations['A' + str(i)].style = self.input_name
            self.violations['A' + str(i+1)].style = self.input_name
            self.violations.merge_cells('B' + str(i) + ':B' + str(i + 1))  # merge box for list
            self.violations['B' + str(i)] = carrier_list  # list
            self.violations['B' + str(i)].style = self.input_s
            self.violations['B' + str(i+1)].style = self.input_s
            self.violations.merge_cells('C' + str(i) + ':C' + str(i + 1))  # merge box for weekly 5200
            self.violations['C' + str(i)] = Convert(total).empty_not_zerofloat()  # total
            self.violations['C' + str(i)].style = self.input_s
            self.violations['C' + str(i+1)].style = self.input_s
            self.violations['C' + str(i)].number_format = "#,###.00;[RED]-#,###.00"
            # saturday
            self.violations.merge_cells('D' + str(i + 1) + ':E' + str(i + 1))  # merge box for sat 5200
            self.violations['D' + str(i)] = leavetime_array[0]  # leave time
            self.violations['D' + str(i)].style = self.input_s
            self.violations['D' + str(i)].number_format = "#,###.00;[RED]-#,###.00"
            self.violations['E' + str(i)] = leavetype_array[0]  # leave type
            self.violations['E' + str(i)].style = self.input_s
            self.violations['D' + str(i + 1)] = totals_array[0]  # 5200 time
            self.violations['D' + str(i + 1)].style = self.input_s
            self.violations['D' + str(i + 1)].number_format = "#,###.00;[RED]-#,###.00"
            # sunday
            self.violations.merge_cells('F' + str(i + 1) + ':G' + str(i + 1))  # merge box for sun 5200
            self.violations['F' + str(i)] = leavetime_array[1]  # leave time
            self.violations['F' + str(i)].style = self.input_s
            self.violations['F' + str(i)].number_format = "#,###.00;[RED]-#,###.00"
            self.violations['G' + str(i)] = leavetype_array[1]  # leave type
            self.violations['G' + str(i)].style = self.input_s
            self.violations['F' + str(i + 1)] = totals_array[1]  # 5200 time
            self.violations['F' + str(i + 1)].style = self.input_s
            self.violations['F' + str(i + 1)].number_format = "#,###.00;[RED]-#,###.00"
            # monday
            self.violations.merge_cells('H' + str(i + 1) + ':I' + str(i + 1))  # merge box for mon 5200
            self.violations['H' + str(i)] = leavetime_array[2]  # leave time
            self.violations['H' + str(i)].style = self.input_s
            self.violations['H' + str(i)].number_format = "#,###.00;[RED]-#,###.00"
            self.violations['I' + str(i)] = leavetype_array[2]  # leave type
            self.violations['I' + str(i)].style = self.input_s
            self.violations['H' + str(i + 1)] = totals_array[2]  # 5200 time
            self.violations['H' + str(i + 1)].style = self.input_s
            self.violations['H' + str(i + 1)].number_format = "#,###.00;[RED]-#,###.00"
            # tuesday
            self.violations.merge_cells('J' + str(i + 1) + ':K' + str(i + 1))  # merge box for tue 5200
            self.violations['J' + str(i)] = leavetime_array[3]  # leave time
            self.violations['J' + str(i)].style = self.input_s
            self.violations['J' + str(i)].number_format = "#,###.00;[RED]-#,###.00"
            self.violations['K' + str(i)] = leavetype_array[3]  # leave type
            self.violations['K' + str(i)].style = self.input_s
            self.violations['J' + str(i + 1)] = totals_array[3]  # 5200 time
            self.violations['J' + str(i + 1)].style = self.input_s
            self.violations['J' + str(i + 1)].number_format = "#,###.00;[RED]-#,###.00"
            # wednesday
            self.violations.merge_cells('L' + str(i + 1) + ':M' + str(i + 1))  # merge box for wed 5200
            self.violations['L' + str(i)] = leavetime_array[4]  # leave time
            self.violations['L' + str(i)].style = self.input_s
            self.violations['L' + str(i)].number_format = "#,###.00;[RED]-#,###.00"
            self.violations['M' + str(i)] = leavetype_array[4]  # leave type
            self.violations['M' + str(i)].style = self.input_s
            self.violations['L' + str(i + 1)] = totals_array[4]  # 5200 time
            self.violations['L' + str(i + 1)].style = self.input_s
            self.violations['L' + str(i + 1)].number_format = "#,###.00;[RED]-#,###.00"
            # thursday
            self.violations.merge_cells('N' + str(i + 1) + ':O' + str(i + 1))  # merge box for thr 5200
            self.violations['N' + str(i)] = leavetime_array[5]  # leave time
            self.violations['N' + str(i)].style = self.input_s
            self.violations['N' + str(i)].number_format = "#,###.00;[RED]-#,###.00"
            self.violations['O' + str(i)] = leavetype_array[5]  # leave type
            self.violations['O' + str(i)].style = self.input_s
            self.violations['N' + str(i + 1)] = totals_array[5]  # 5200 time
            self.violations['N' + str(i + 1)].style = self.input_s
            self.violations['N' + str(i + 1)].number_format = "#,###.00;[RED]-#,###.00"
            # friday
            self.violations.merge_cells('P' + str(i + 1) + ':Q' + str(i + 1))  # merge box for fri 5200
            self.violations['P' + str(i)] = leavetime_array[6]  # leave time
            self.violations['P' + str(i)].style = self.input_s
            self.violations['P' + str(i)].number_format = "#,###.00;[RED]-#,###.00"
            self.violations['Q' + str(i)] = leavetype_array[6]  # leave type
            self.violations['Q' + str(i)].style = self.input_s
            self.violations['P' + str(i + 1)] = totals_array[6]  # 5200 time
            self.violations['P' + str(i + 1)].style = self.input_s
            self.violations['P' + str(i + 1)].number_format = "#,###.00;[RED]-#,###.00"
            # calculated fields
            # hidden columns
            formula_a = "=SUM(%s!D%s:P%s)+%s!D%s + %s!F%s + %s!H%s + %s!J%s + %s!L%s + " \
                        "%s!N%s + %s!P%s" % ("violations", str(i + 1), str(i + 1),
                                             "violations", str(i), "violations", str(i), "violations", str(i),
                                             "violations", str(i), "violations", str(i), "violations", str(i),
                                             "violations", str(i))
            self.violations['R' + str(i)] = formula_a
            self.violations['R' + str(i)].style = self.calcs
            self.violations['R' + str(i)].number_format = "#,###.00;[RED]-#,###.00"
            formula_b = "=SUM(%s!C%s+%s!D%s+%s!F%s+%s!H%s+%s!J%s+%s!L%s+%s!N%s+%s!P%s)" % \
                        ("violations", str(i), "violations", str(i), "violations", str(i),
                         "violations", str(i), "violations", str(i), "violations", str(i),
                         "violations", str(i), "violations", str(i))
            self.violations['R' + str(i + 1)] = formula_b
            self.violations['R' + str(i + 1)].style = self.calcs
            self.violations['R' + str(i + 1)].number_format = "#,###.00;[RED]-#,###.00"
            # weekly violation
            self.violations.merge_cells('S' + str(i) + ':S' + str(i + 1))  # merge box for weekly violation
            formula_c = "=IF(OR(%s!B%s = \"aux\",%s!B%s = \"ptf\"),0," \
                        "MAX(IF(%s!R%s>%s!R%s,MAX(%s!R%s-60,0),MAX(%s!R%s-60)),0))" \
                        % ("violations", str(i), "violations", str(i), "violations", str(i),
                           "violations", str(i + 1), "violations", str(i), "violations", str(i + 1))
            self.violations['S' + str(i)] = formula_c
            self.violations['S' + str(i)].style = self.calcs
            self.violations['S' + str(i+1)].style = self.calcs
            self.violations['S' + str(i)].number_format = "#,###.00;[RED]-#,###.00"
            # daily violation
            formula_d = "=IF(OR(%s!B%s=\"%s\",%s!B%s=\"nl\",%s!B%s=\"ptf\",%s!B%s=\"aux\")," \
                        "(SUM(IF(%s!D%s>%s,%s!D%s-%s,0)+IF(%s!F%s>%s,%s!F%s-%s,0)" \
                        "+IF(%s!H%s>%s,%s!H%s-%s,0)+" \
                        "IF(%s!J%s>%s,%s!J%s-%s,0)" \
                        "+IF(%s!L%s>%s,%s!L%s-%s,0)+IF(%s!N%s>%s,%s!N%s-%s,0)+" \
                        "IF(%s!P%s>%s,%s!P%s-%s,0)))," \
                        "(SUM(IF(%s!D%s>12,%s!D%s-12,0)+IF(%s!F%s>12,%s!F%s-12,0)+IF(%s!H%s>12,%s!H%s-12,0)" \
                        "+IF(%s!J%s>12,%s!J%s-12,0)" \
                        "+IF(%s!L%s>12,%s!L%s-12,0)+IF(%s!N%s>12,%s!N%s-12,0)+IF(%s!P%s>12,%s!P%s-12,0))))"\
                        % ("violations", str(i), self.wal_12hr_mod,
                           "violations", str(i), "violations", str(i), "violations", str(i),
                           "violations", str(i + 1), str(self.non_otdl_violation),
                           "violations", str(i + 1), str(self.non_otdl_violation),
                           "violations", str(i + 1), str(self.non_otdl_violation),
                           "violations", str(i + 1), str(self.non_otdl_violation),
                           "violations", str(i + 1), str(self.non_otdl_violation),
                           "violations", str(i + 1), str(self.non_otdl_violation),
                           "violations", str(i + 1), str(self.non_otdl_violation),
                           "violations", str(i + 1), str(self.non_otdl_violation),
                           "violations", str(i + 1), str(self.non_otdl_violation),
                           "violations", str(i + 1), str(self.non_otdl_violation),
                           "violations", str(i + 1), str(self.non_otdl_violation),
                           "violations", str(i + 1), str(self.non_otdl_violation),
                           "violations", str(i + 1), str(self.non_otdl_violation),
                           "violations", str(i + 1), str(self.non_otdl_violation),
                           "violations", str(i + 1), "violations", str(i + 1), "violations", str(i + 1),
                           "violations", str(i + 1), "violations", str(i + 1), "violations", str(i + 1),
                           "violations", str(i + 1), "violations", str(i + 1), "violations", str(i + 1),
                           "violations", str(i + 1), "violations", str(i + 1), "violations", str(i + 1),
                           "violations", str(i + 1), "violations", str(i + 1))
            self.violations['T' + str(i)] = formula_d
            self.violations.merge_cells('T' + str(i) + ':T' + str(i + 1))  # merge box for daily violation
            self.violations['T' + str(i)].style = self.calcs
            self.violations['T' + str(i+1)].style = self.calcs
            self.violations['T' + str(i)].number_format = "#,###.00"
            # wed adjustment
            self.violations.merge_cells('U' + str(i) + ':U' + str(i + 1))  # merge box for wed adj
            formula_e = "=IF(OR(%s!B%s=\"%s\",%s!B%s=\"nl\",%s!B%s=\"ptf\",%s!B%s=\"aux\")," \
                        "IF(AND(%s!S%s-(%s!N%s+%s!N%s+%s!P%s+%s!P%s)>0,%s!L%s>%s)," \
                        "IF(%s!S%s-(%s!N%s+%s!N%s+%s!P%s+%s!P%s)>%s!L%s-%s,%s!L%s-%s,%s!S%s-" \
                        "(%s!N%s+%s!N%s+%s!P%s+%s!P%s)),0)," \
                        "IF(AND(%s!S%s-(%s!N%s+%s!N%s+%s!P%s+%s!P%s)>0,%s!L%s>12)," \
                        "IF(%s!S%s-(%s!N%s+%s!N%s+%s!P%s+%s!P%s)>%s!L%s-12,%s!L%s-12,%s!S%s-" \
                        "(%s!N%s+%s!N%s+%s!P%s+%s!P%s)),0))" \
                        % ("violations", str(i), self.wal_12hr_mod,
                           "violations", str(i), "violations", str(i), "violations", str(i),
                           "violations", str(i), "violations", str(i + 1), "violations", str(i),
                           "violations", str(i + 1), "violations", str(i), "violations", str(i + 1),
                           str(self.non_otdl_violation),
                           "violations", str(i), "violations", str(i + 1), "violations", str(i),
                           "violations", str(i + 1), "violations", str(i), "violations", str(i + 1),
                           str(self.non_otdl_violation),
                           "violations", str(i + 1),
                           str(self.non_otdl_violation),
                           "violations", str(i), "violations", str(i + 1),
                           "violations", str(i), "violations", str(i + 1), "violations", str(i),
                           "violations", str(i), "violations", str(i + 1), "violations", str(i),
                           "violations", str(i + 1), "violations", str(i), "violations", str(i + 1),
                           "violations", str(i), "violations", str(i + 1), "violations", str(i),
                           "violations", str(i + 1), "violations", str(i), "violations", str(i + 1),
                           "violations", str(i + 1), "violations", str(i), "violations", str(i + 1),
                           "violations", str(i), "violations", str(i + 1), "violations", str(i))
            self.violations['U' + str(i)] = formula_e
            self.violations['U' + str(i)].style = self.vert_calcs
            self.violations['U' + str(i+1)].style = self.vert_calcs
            self.violations['U' + str(i)].number_format = "#,###.00"
            # thr adjustment
            formula_f = "=IF(OR(%s!B%s=\"%s\",%s!B%s=\"nl\",%s!B%s=\"ptf\",%s!B%s=\"aux\")," \
                        "IF(AND(%s!S%s-(%s!P%s+%s!P%s)>0,%s!N%s>%s)," \
                        "IF(%s!S%s-(%s!P%s+%s!P%s)>%s!N%s-%s,%s!N%s-%s,%s!S%s-(%s!P%s+%s!P%s)),0)," \
                        "IF(AND(%s!S%s-(%s!P%s+%s!P%s)>0,%s!N%s>12)," \
                        "IF(%s!S%s-(%s!P%s+%s!P%s)>%s!N%s-12,%s!N%s-12,%s!S%s-(%s!P%s+%s!P%s)),0))" \
                        % ("violations", str(i), self.wal_12hr_mod,
                           "violations", str(i), "violations", str(i), "violations", str(i),
                           "violations", str(i), "violations", str(i + 1), "violations", str(i),
                           "violations", str(i + 1), str(self.non_otdl_violation),
                           "violations", str(i), "violations", str(i + 1), "violations", str(i),
                           "violations", str(i + 1), str(self.non_otdl_violation),
                           "violations", str(i + 1), str(self.non_otdl_violation),
                           "violations", str(i),
                           "violations", str(i + 1), "violations", str(i),
                           "violations", str(i), "violations", str(i + 1), "violations", str(i),
                           "violations", str(i + 1),
                           "violations", str(i), "violations", str(i + 1), "violations", str(i),
                           "violations", str(i + 1), "violations", str(i + 1), "violations", str(i),
                           "violations", str(i + 1), "violations", str(i)
                           )
            self.violations.merge_cells('V' + str(i) + ':V' + str(i + 1))  # merge box for thr adj
            self.violations['V' + str(i)] = formula_f
            self.violations['V' + str(i)].style = self.vert_calcs
            self.violations['V' + str(i+1)].style = self.vert_calcs
            self.violations['V' + str(i)].number_format = "#,###.00"
            # fri adjustment
            self.violations.merge_cells('W' + str(i) + ':W' + str(i + 1))  # merge box for fri adj
            formula_g = "=IF(OR(%s!B%s=\"%s\",%s!B%s=\"nl\",%s!B%s=\"ptf\",%s!B%s=\"aux\")," \
                        "IF(AND(%s!S%s>0,%s!P%s>%s)," \
                        "IF(%s!S%s>%s!P%s-%s,%s!P%s-%s,%s!S%s),0)," \
                        "IF(AND(%s!S%s>0,%s!P%s>12)," \
                        "IF(%s!S%s>%s!P%s-12,%s!P%s-12,%s!S%s),0))" \
                        % ("violations", str(i), self.wal_12hr_mod,
                           "violations", str(i), "violations", str(i),
                           "violations", str(i),
                           "violations", str(i), "violations", str(i + 1),
                           str(self.non_otdl_violation),
                           "violations", str(i), "violations", str(i + 1),
                           str(self.non_otdl_violation),
                           "violations", str(i + 1),
                           str(self.non_otdl_violation),
                           "violations", str(i), "violations", str(i), "violations",  str(i + 1), "violations", str(i),
                           "violations", str(i + 1), "violations", str(i + 1), "violations", str(i))
            self.violations['W' + str(i)] = formula_g
            self.violations['W' + str(i)].style = self.vert_calcs
            self.violations['W' + str(i+1)].style = self.vert_calcs
            self.violations['W' + str(i)].number_format = "#,###.00"
            # total violation
            self.violations.merge_cells('X' + str(i) + ':X' + str(i + 1))  # merge box for total violation
            formula_h = "=IF(AND(%s!O4=\"yes\"," \
                        "OR(%s!B%s=\"otdl\", %s!B%s=\"%s\")),\"exempt\",SUM(%s!S%s:T%s)-(%s!U%s+%s!V%s+%s!W%s))" \
                        % ("violations", "violations", str(i),
                           "violations", str(i), self.wal_dec_exempt_mod,
                           "violations", str(i), str(i), "violations", str(i),
                           "violations", str(i), "violations", str(i))
            self.violations['X' + str(i)] = formula_h
            self.violations['X' + str(i)].style = self.calcs
            self.violations['X' + str(i+1)].style = self.calcs
            self.violations['X' + str(i)].number_format = "#,###.00"
            # ----------------------------------------------------------------------------------fill cells for summary
            formula_i = "=IF(%s!A%s = 0,\"\",%s!A%s)" % ("violations", str(i), "violations", str(i))
            self.summary['A' + str(summary_i)] = formula_i
            self.summary['A' + str(summary_i)].style = self.input_name
            formula_j = "= IF(violations!X%s = \"exempt\", \"\", " \
                        "IF(violations!X%s >= summary!G$4,violations!X%s, \"\"))" \
                        % (str(i), str(i), str(i))
            self.summary['B' + str(summary_i)] = formula_j
            self.summary['B' + str(summary_i)].style = self.input_s
            self.summary['B' + str(summary_i)].number_format = "#,###.00"
            self.summary.row_dimensions[summary_i].height = 10  # adjust all row height
            if self.show_remedy:  # optional Super remedy solution
                formula_k = "=IF(OR(summary!B%s=\"exempt\", summary!B%s=\"\"),\"\",summary!B%s*summary!G5)" \
                            % (str(summary_i), str(summary_i), str(summary_i))
                self.summary['C' + str(summary_i)] = formula_k
                self.summary['C' + str(summary_i)].style = self.calcs
                self.summary['C' + str(summary_i)].number_format = "[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00"
                self.summary.row_dimensions[summary_i].height = 10  # adjust all row height
            i += 2
            summary_i += 1
        i += 1
        self.violations.merge_cells('L' + str(i) + ':T' + str(i))  # label for cumulative violations
        self.violations['L' + str(i)] = "Cumulative Total Violations:  "
        self.violations['L' + str(i)].style = self.date_dov_title
        self.violations.merge_cells('U' + str(i) + ':X' + str(i))  # total violation summary at bottom of page
        formula_h = "=SUM(%s!X%s:X%s)" \
                    % ("violations", "9", str(i-2))
        self.violations['U' + str(i)] = formula_h
        self.violations['U' + str(i)].style = self.calcs
        self.violations['X' + str(i)].style = self.calcs
        self.violations['U' + str(i)].number_format = "#,###.00"
        self.violations.row_dimensions[i].height = 20  # adjust all row height
        self.violations.merge_cells('L' + str(i) + ':T' + str(i))  # label for cumulative violations
        if self.show_remedy:  # optional Super remedy solution
            summary_i += 1
            self.summary['A' + str(summary_i)] = "Cumulative Total Remedy:  "
            self.summary['A' + str(summary_i)].style = self.date_dov_title
            self.summary.merge_cells('A' + str(summary_i) + ':B' + str(summary_i))  # total violation summary
            formula_i = "=SUM(%s!C%s:C%s)" \
                        % ("summary", "7", str(summary_i-2))
            self.summary['C' + str(summary_i)] = formula_i
            self.summary['C' + str(summary_i)].style = self.calcs
            self.summary['C' + str(summary_i)].number_format = "[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00"
            self.summary.row_dimensions[summary_i].height = 20  # adjust all row height

    def save_open(self):
        """ save the spreadsheet and open """
        self.pbi += 1
        self.pb.move_count(self.pbi)  # increment progress bar
        self.pb.change_text("Saving...")
        self.pb.stop()
        xl_filename = "kb_om" + str(format(projvar.invran_date_week[0], "_%y_%m_%d")) + ".xlsx"
        try:
            self.wb.save(dir_path('over_max_spreadsheet') + xl_filename)
            messagebox.showinfo("Spreadsheet generator",
                                "Your spreadsheet was successfully generated. \n"
                                "File is named: {}".format(xl_filename),
                                parent=self.frame)
            if sys.platform == "win32":  # open the text document
                os.startfile(dir_path('over_max_spreadsheet') + xl_filename)
            if sys.platform == "linux":
                subprocess.call(["xdg-open", 'kb_sub/over_max_spreadsheet/' + xl_filename])
            if sys.platform == "darwin":
                subprocess.call(["open", dir_path('over_max_spreadsheet') + xl_filename])
        except PermissionError:
            messagebox.showerror("Spreadsheet generator",
                                 "The spreadsheet was not generated. \n"
                                 "Suggestion: "
                                 "Make sure that identically named spreadsheets are closed "
                                 "(the file can't be overwritten while open).",
                                 parent=self.frame)


class ImpManSpreadsheet4:
    """
    This is a spreadsheet built to placate step B for Region 4.
    """
    def __init__(self):
        self.frame = None  # the frame of parent
        self.pb = None  # progress bar object
        self.pbi = 0  # progress bar count index
        self.startdate = None  # start date of the investigation
        self.enddate = None  # ending date of the investigation
        self.dates = []  # all days of the investigation
        self.carrierlist = []  # all carriers in carrier list
        self.carrier_breakdown = []  # all carriers in carrier list broken down into appropiate list
        self.mod_carrierlist = []
        self.tol_ot_ownroute = 0.0  # get tolerances from tolerances table.
        self.tol_ot_offroute = 0.0
        self.tol_availability = 0.0
        self.min_man4_nl = 0  # get minimum rows from tolerances table
        self.min_man4_wal = 0
        self.min_man4_otdl = 0
        self.min_man4_aux = 0
        self.wb = None  # the workbook object
        self.ws_list = []  # "saturday", "sunday", "monday", "tuesday", "wednesday", "thursday", "friday"
        self.day_of_week = []  # seven day array for weekly investigations/ one day array for daily investigations
        # styles for worksheet
        self.ws_header = None  # style
        self.list_header = None  # style
        self.date_dov = None  # style
        self.date_dov_title = None  # style
        self.col_header = None  # style
        self.input_name = None  # style
        self.input_s = None  # style
        self.calcs = None  # style
        self.quad_top = None  # style
        self.quad_bottom = None  # style
        self.quad_left = None  # style
        self.quad_right = None  # style
        self.col_header_left = None  # style
        self.col_header = None  # style
        self.footer_left = None  # style
        self.footer_right = None  # style
        self.footer_mid = None  # style
        self.day = None  # build worksheet - loop once for each day
        self.i = 0  # build worksheet loop iteration
        self.lsi = 0  # list loop iteration
        self.pref = ("nl", "wal", "otdl", "aux")
        self.row = 0
        # cell for summary quadrants page
        self.cellc9 = None  # non otdl own route violations
        self.cellf9 = None  # non otdl off route violations
        self.cellf11 = None  # wal off route violations
        self.cellj9 = None  # aux availability to 10 hours
        self.cellm9 = None  # aux availability to 11.5 hours
        self.cellj11 = None  # otdl availability to 10 hours
        self.cellm11 = None  # otdl availability to 12 hours
        self.cellf16 = None  # carriers out past dispatch of value
        self.celln16 = None  # otdl/aux availability to DOV

        self.ot_list = ("NON OTDL", "Work Assignment", "Auxiliary", "OTDL")  # list loop iteration
        self.page_titles = ("NON-OTDL Employees that worked overtime",
                            "Work Assignment Employees that worked off their assignment",
                            "Auxiliary Employees who were available to work OT",
                            "OTDL Employees who were available to work OT")
        self.pref = ("nl", "wal", "aux", "otdl")
        self.pb4_nl_wal = True  # page break between no list and work assignment
        self.pb4_wal_aux = True  # page break between work assignment and otdl
        self.pb4_aux_otdl = True  # page break between otdl and auxiliary
        self.min4_ss_nl = 0  # minimum rows for "no list"
        self.min4_ss_wal = 0  # minimum rows for work assignment list
        self.min4_ss_otdl = 0  # minimum rows for overtime desired list
        self.min4_ss_aux = 0  # minimum rows for auxiliary
        self.row_number = 1
        self.carrier = None  # current iteration of carrier's name is assigned self.carrier
        self.list_ = None  # current iteration of carrier's list status is assigned self.carrier
        self.route = None  # current iteration of carrier's route is assigned self.carrier
        self.nsday = None
        self.rings = []  # assign as self.rings
        self.totalhours = 0.0  # set default as an empty string
        self.bt = ""
        self.rs = ""
        self.et = ""
        self.codes = ""
        self.moves = ""
        self.overtime = 0.0  # the amount of overtime worked by the carrier
        self.onroute = 0.0  # the amount of overworked on the carrier's own route.
        self.offroute = 0.0  # empty string or calculated time that carrier spent off their assignment
        self.offroute_adj = 0.0  # self.offroute adjusted for pivot time, ns days, and whole days off bid assignment
        self.otherroute_array = []  # a list of routes where carrier worked off assignment
        self.odlr_indicator = []  # indicates that carrier is odlr for at least one day
        self.odln_indicator = []  # indicates that carrier is odln for at least one day
        self.otherroute = ""  # the off assignment route the carrier worked on - formated for the cell
        self.avail_10 = 0.0  # otdl/aux availability to 10 hours
        self.avail_115 = 0.0  # aux availability to 11.50 hours
        self.avail_12 = 0.0  # otdl availability to 12 hours.
        self.lvtype = ""
        self.lvtime = ""
        self.first_row = 0  # record the number of the first row for totals formulas in footers
        self.last_row = 0  # record the number of the last row for totals formulas in footers
        # build a dictionary for displaying list statuses on spreadsheet
        self.list_dict = {'': '', 'nl': 'non list', 'wal': 'wal', 'otdl': 'otdl', 'aux': 'cca', 'ptf': 'ptf',
                          'odlr': 'odl regular', 'odln': 'odl nsday'}
        self.display_limiter = "show all"  # show all, only workdays, only mandates
        self.display_counter = 0  # count the number of rows displayed per list loop
        self.listrange = []  # records the first row, last row and summary row of each list
        self.dayrange = []  # records the listranges for the day by appending listranges after each listloop.
        self.dovarray = []  # build a list of 7 dov times. One for each day.

    def create(self, frame):
        """ a master method for running other methods in proper order."""
        self.frame = frame
        if not self.ask_ok():  # abort if user selects cancel from askokcancel
            return
        self.pb = ProgressBarDe(label="Building Improper Mandates Spreadsheet")
        self.pb.max_count(100)  # set length of progress bar
        self.pb.start_up()  # start the progress bar
        self.pbi = 1
        self.pb.move_count(self.pbi)  # increment progress bar
        self.pb.change_text("Gathering Data... ")
        self.get_dates()
        self.get_settings()  # get tolerances, minimum rows and page breaks from tolerances table
        self.get_pb_max_count()  # set the length of the progress bar
        self.get_carrierlist()
        self.get_carrier_breakdown()  # breakdown carrier list into no list, wal, otdl, aux
        self.get_tolerances()  # get tolerances, minimum rows and page break preferences from tolerances table
        self.get_dov()  # get the dispatch of value for each day
        self.get_styles()
        self.build_workbook()
        self.set_dimensions()
        self.build_ws_loop()  # loop once for each day
        self.save_open()

    def ask_ok(self):
        """ ends process if user cancels """
        if messagebox.askokcancel("Spreadsheet generator",
                                  "Do you want to generate an \nImproper Mandates No. 4 Spreadsheet?",
                                  parent=self.frame):
            return True
        return False

    def get_dates(self):
        """ get the dates from the project variables """
        self.startdate = projvar.invran_date  # set daily investigation range as default - get start date
        self.enddate = projvar.invran_date  # get end date
        self.dates = [projvar.invran_date, ]  # create an array of days - only one day if daily investigation range
        if projvar.invran_weekly_span:  # if the investigation range is weekly
            date = projvar.invran_date_week[0]
            self.startdate = projvar.invran_date_week[0]
            self.enddate = projvar.invran_date_week[6]
            self.dates = []
            for _ in range(7):  # create an array with all the days in the weekly investigation range
                self.dates.append(date)
                date += timedelta(days=1)

    def get_settings(self):
        """ get spreadsheet tolerances, row minimums and page break prefs from tolerance table """
        sql = "SELECT tolerance FROM tolerances"
        result = inquire(sql)
        self.tol_ot_ownroute = float(result[0][0])  # overtime on own route tolerance
        self.tol_ot_offroute = float(result[1][0])  # overtime off own route tolerance
        self.tol_availability = float(result[2][0])  # availability tolerance
        self.min4_ss_nl = int(result[32][0])  # minimum rows for no list
        self.min4_ss_wal = int(result[33][0])  # mimimum rows for work assignment
        self.min4_ss_otdl = int(result[34][0])  # minimum rows for otdl
        self.min4_ss_aux = int(result[35][0])  # minimum rows for auxiliary
        self.pb4_nl_wal = Convert(result[36][0]).str_to_bool()  # page break between no list and work assignment
        self.pb4_wal_aux = Convert(result[37][0]).str_to_bool()  # page break between work assignment and otdl
        self.pb4_aux_otdl = Convert(result[38][0]).str_to_bool()  # page break between otdl and auxiliary
        self.display_limiter = result[39][0]

    def get_pb_max_count(self):
        """ set length of progress bar """
        self.pb.max_count((len(self.dates)*4)+1)  # once for each list in each day, plus saving

    def get_carrierlist(self):
        """ get record sets for all carriers """
        self.carrierlist = CarrierList(self.startdate, self.enddate, projvar.invran_station).get()

    def get_carrier_breakdown(self):
        """ breakdown carrier list into no list, wal, otdl, aux """
        timely_rec = []
        for day in self.dates:
            nl_array = []
            wal_array = []
            aux_array = []
            otdl_array = []
            for carrier in self.carrierlist:
                for rec in reversed(carrier):
                    if Convert(rec[0]).dt_converter() <= day:
                        timely_rec = rec
                if timely_rec[2] == "nl":
                    nl_array.append(timely_rec)
                if timely_rec[2] == "wal":
                    wal_array.append(timely_rec)
                if timely_rec[2] == "otdl":
                    otdl_array.append(timely_rec)
                if timely_rec[2] == "odlr":  # for odl regular day only -
                    if timely_rec[1] not in self.odlr_indicator:  # add name to odlr indicator array
                        self.odlr_indicator.append(timely_rec[1])
                    # if it is sunday or their ns day, put the record in the nl array
                    if day.strftime("%a") == projvar.ns_code[timely_rec[3]] or day.strftime("%a") == "Sun":
                        nl_array.append(timely_rec)
                    else:  # if it is a sunday or their ns day, put record in no list array.
                        otdl_array.append(timely_rec)
                if timely_rec[2] == "odln":  # for odl non scheduled day only
                    if timely_rec[1] not in self.odln_indicator:  # add name to odln indicator array
                        self.odln_indicator.append(timely_rec[1])
                    # if it is sunday or their ns day, put the record in the otdl array
                    if day.strftime("%a") == projvar.ns_code[timely_rec[3]] or day.strftime("%a") == "Sun":
                        otdl_array.append(timely_rec)
                    else:
                        nl_array.append(timely_rec)
                if timely_rec[2] == "aux" or timely_rec[2] == "ptf":
                    aux_array.append(timely_rec)
            daily_breakdown = [nl_array, wal_array, aux_array, otdl_array]
            self.carrier_breakdown.append(daily_breakdown)

    def get_tolerances(self):
        """ get spreadsheet tolerances, row minimums and page break prefs from tolerance table """
        sql = "SELECT tolerance FROM tolerances"
        result = inquire(sql)
        self.tol_ot_ownroute = float(result[0][0])  # overtime on own route tolerance
        self.tol_ot_offroute = float(result[1][0])  # overtime off own route tolerance
        self.tol_availability = float(result[2][0])  # availability tolerance
        self.min_man4_nl = int(result[3][0])  # minimum rows for no list
        self.min_man4_wal = int(result[4][0])  # mimimum rows for work assignment
        self.min_man4_otdl = int(result[5][0])  # minimum rows for otdl
        self.min_man4_aux = int(result[6][0])  # minimum rows for auxiliary

    def get_dov(self):
        """ get the dov records currently in the database """
        days = ("sat", "sun", "mon", "tue", "wed", "thu", "fri")
        for i in range(len(days)):
            sql = "SELECT * FROM dov WHERE eff_date <= '%s' AND station = '%s' AND day = '%s' " \
                  "ORDER BY eff_date DESC" % \
                  (projvar.invran_date_week[0], projvar.invran_station, days[i])
            result = inquire(sql)
            for rec in result:
                if rec[0] == Convert(projvar.invran_date_week[0]).dt_to_str():
                    self.dovarray.append(rec[3])
                    break
                elif rec[4] == "False":
                    self.dovarray.append(rec[3])
                    break
                else:
                    continue

    def get_styles(self):
        """ Named styles for workbook """
        bd = Side(style='thin', color="80808080")  # defines borders
        self.ws_header = NamedStyle(name="ws_header", font=Font(bold=True, name='Arial', size=12))
        self.list_header = NamedStyle(name="list_header", font=Font(bold=True, name='Arial', size=10))
        self.date_dov = NamedStyle(name="date_dov", font=Font(name='Arial', size=8))
        self.date_dov_title = NamedStyle(name="date_dov_title", font=Font(bold=True, name='Arial', size=8),
                                         alignment=Alignment(horizontal='right'))
        self.col_header_left = NamedStyle(name="col_header_left", font=Font(bold=True, name='Arial', size=8),
                                          alignment=Alignment(horizontal='left', vertical='bottom', wrap_text=True))
        self.col_header = NamedStyle(name="col_header", font=Font(bold=True, name='Arial', size=8),
                                     alignment=Alignment(horizontal='center', vertical='bottom', wrap_text=True))
        self.input_name = NamedStyle(name="input_name", font=Font(name='Arial', size=8),
                                     border=Border(left=bd, top=bd, right=bd, bottom=bd))
        self.input_s = NamedStyle(name="input_s", font=Font(name='Arial', size=8),
                                  border=Border(left=bd, top=bd, right=bd, bottom=bd),
                                  alignment=Alignment(horizontal='right', wrap_text=True))
        self.calcs = NamedStyle(name="calcs", font=Font(name='Arial', size=8),
                                border=Border(left=bd, top=bd, right=bd, bottom=bd),
                                fill=PatternFill(fgColor='e5e4e2', fill_type='solid'),
                                alignment=Alignment(horizontal='right'))

        self.quad_top = NamedStyle(name="quad_top", font=Font(name='Arial', size=10),
                                   alignment=Alignment(horizontal='left', vertical='top', wrap_text=True),
                                   border=Border(left=bd, top=bd, right=bd))
        self.quad_bottom = NamedStyle(name="quad_bottom", font=Font(name='Arial', size=10),
                                      alignment=Alignment(horizontal='right', wrap_text=True),
                                      border=Border(left=bd, bottom=bd, right=bd))
        self.quad_left = NamedStyle(name="quad_left", font=Font(name='Arial', size=10),
                                    alignment=Alignment(horizontal='left', vertical='top', wrap_text=True),
                                    border=Border(left=bd, bottom=bd, top=bd))
        self.quad_right = NamedStyle(name="quad_right", font=Font(name='Arial', size=10),
                                     alignment=Alignment(horizontal='right', vertical='top', wrap_text=True),
                                     border=Border(top=bd, bottom=bd, right=bd))
        self.footer_left = NamedStyle(name="footer_left", font=Font(bold=True, name='Arial', size=8),
                                      alignment=Alignment(horizontal='left'),
                                      border=Border(left=bd, bottom=bd, top=bd))
        self.footer_right = NamedStyle(name="footer_right", font=Font(bold=True, name='Arial', size=8),
                                       alignment=Alignment(horizontal='right'),
                                       border=Border(top=bd, bottom=bd, right=bd))
        self.footer_mid = NamedStyle(name="footer_mid", font=Font(bold=True, name='Arial', size=8),
                                     alignment=Alignment(horizontal='right'),
                                     border=Border(top=bd, bottom=bd))

    def build_workbook(self):
        """ build the workbook object """
        day_finder = ["Sat", "Sun", "Mon", "Tue", "Wed", "Thu", "Fri"]
        day_of_week = ["saturday", "sunday", "monday", "tuesday", "wednesday", "thursday", "friday"]
        i = 0
        self.wb = Workbook()  # define the workbook
        if not projvar.invran_weekly_span:  # if investigation range is daily
            for ii in range(len(day_finder)):
                if projvar.invran_date.strftime("%a") == day_finder[ii]:  # find the correct day
                    i = ii
            self.ws_list.append(self.wb.active)  # create first worksheet
            self.ws_list[0].title = day_of_week[i]  # title first worksheet
            self.day_of_week.append(day_of_week[i])  # create self.day_of_week array with one day
        if projvar.invran_weekly_span:  # if investigation range is weekly
            for day in day_of_week:
                self.day_of_week.append(day)  # create self.day_of_week array with seven days
            self.ws_list.append(self.wb.active)  # create first worksheet
            self.ws_list[0].title = "saturday"  # title first worksheet
            for i in range(1, 7):  # create worksheet for remaining six days
                self.ws_list.append(self.wb.create_sheet(day_of_week[i]))  # create subsequent worksheets
                self.ws_list[i].title = day_of_week[i]  # title subsequent worksheets

    def set_dimensions(self):
        """ set the orientation and dimensions of the workbook """
        for i in range(len(self.dates)):
            self.ws_list[i].set_printer_settings(paper_size=1, orientation='landscape')  # set orientation
            self.ws_list[i].oddFooter.center.text = "&A"  # include the footer
            self.ws_list[i].column_dimensions["A"].width = 4  # column width
            self.ws_list[i].column_dimensions["B"].width = 4
            self.ws_list[i].column_dimensions["C"].width = 9
            self.ws_list[i].column_dimensions["D"].width = 9
            self.ws_list[i].column_dimensions["E"].width = 5
            self.ws_list[i].column_dimensions["F"].width = 4
            self.ws_list[i].column_dimensions["G"].width = 9
            self.ws_list[i].column_dimensions["H"].width = 9
            self.ws_list[i].column_dimensions["I"].width = 9
            self.ws_list[i].column_dimensions["J"].width = 9
            self.ws_list[i].column_dimensions["K"].width = 9
            self.ws_list[i].column_dimensions["L"].width = 5
            self.ws_list[i].column_dimensions["M"].width = 4
            self.ws_list[i].column_dimensions["N"].width = 9
            self.ws_list[i].column_dimensions["O"].width = 9
            self.ws_list[i].column_dimensions["P"].width = 9
            self.ws_list[i].row_dimensions[8].height = 45  # adjust row height
            self.ws_list[i].row_dimensions[10].height = 45  # adjust row height

    def build_ws_loop(self):
        """ this loops once for each list. """
        self.i = 0
        for day in self.dates:
            self.dayrange = []  # initialize array for holding all start/stop/summary rows for all four list.
            self.day = day
            self.build_ws_headers()
            self.build_ws_quads()
            self.pagebreak(force=True)  # force the page break
            self.list_loop()  # loops four times. once for each list.
            self.fill_quads()  # write formulas for the quadrants on the cover sheet.
            self.i += 1

    def build_ws_headers(self):
        """ worksheet headers """
        cell = self.ws_list[self.i].cell(row=1, column=3)
        cell.value = "Improper Mandate Worksheet"
        cell.style = self.ws_header
        self.ws_list[self.i].merge_cells('C1:O1')
        cell = self.ws_list[self.i].cell(row=3, column=3)  # create date label
        cell.value = "Date:  "
        cell.style = self.date_dov_title
        cell = self.ws_list[self.i].cell(row=3, column=4)  # display date
        cell.value = format(self.day, "%A  %m/%d/%y")
        cell.style = self.date_dov
        self.ws_list[self.i].merge_cells('D3:H3')
        cell = self.ws_list[self.i].cell(row=3, column=9)  # pay period label
        cell.value = "Pay Period:  "
        cell.style = self.date_dov_title
        self.ws_list[self.i].merge_cells('I3:J3')
        cell = self.ws_list[self.i].cell(row=3, column=11)  # display pay period
        cell.value = projvar.pay_period
        cell.style = self.date_dov
        self.ws_list[self.i].merge_cells('K3:O3')
        cell = self.ws_list[self.i].cell(row=4, column=3)  # station label
        cell.value = "Station:  "
        cell.style = self.date_dov_title
        cell = self.ws_list[self.i].cell(row=4, column=4)  # display station
        cell.value = projvar.invran_station
        cell.style = self.date_dov
        self.ws_list[self.i].merge_cells('D4:H4')

    def build_ws_quads(self):
        """ build the two quadrants and other elements on the coversheet """
        # Violations Quadrants
        # Top Left
        self.ws_list[self.i].merge_cells('C8:E8')
        cell = self.ws_list[self.i].cell(row=8, column=3)  # NON-OTDL Violations on own route Page 1
        cell.value = "NON-OTDL \nViolations on own route \nPage 1"
        cell.style = self.quad_top
        self.ws_list[self.i].merge_cells('C9:E9')
        self.cellc9 = self.ws_list[self.i].cell(row=9, column=3)
        self.cellc9.value = ""
        self.cellc9.style = self.quad_bottom
        self.cellc9.number_format = "#,###.00;[RED]-#,###.00"
        # Top Right
        cell = self.ws_list[self.i].cell(row=8, column=6)  # NON-OTDL Violations off own route Page 1
        cell.value = "NON-OTDL \nViolations off own route \nPage 1"
        cell.style = self.quad_top
        self.ws_list[self.i].merge_cells('F8:H8')
        self.cellf9 = self.ws_list[self.i].cell(row=9, column=6)  # value filled later
        self.cellf9.style = self.quad_bottom
        self.ws_list[self.i].merge_cells('F9:H9')
        self.cellf9.number_format = "#,###.00;[RED]-#,###.00"
        # Bottom Left
        self.ws_list[self.i].merge_cells('C10:E10')
        cell = self.ws_list[self.i].cell(row=10, column=3)  # Blank
        cell.value = ""
        cell.style = self.quad_top
        cell = self.ws_list[self.i].cell(row=11, column=3)
        cell.value = ""
        cell.style = self.quad_bottom
        self.ws_list[self.i].merge_cells('C11:E11')
        # Bottom Right
        cell = self.ws_list[self.i].cell(row=10, column=6)  # Work assignment Violations off own route Page 2
        cell.value = "Work Assignment \nViolations off own route \nPage 2"
        cell.style = self.quad_top
        self.ws_list[self.i].merge_cells('F10:H10')
        self.cellf11 = self.ws_list[self.i].cell(row=11, column=6)
        self.cellf11.value = ""
        self.cellf11.style = self.quad_bottom
        self.ws_list[self.i].merge_cells('F11:H11')
        self.cellf11.number_format = "#,###.00;[RED]-#,###.00"
        # Totals Left
        cell = self.ws_list[self.i].cell(row=12, column=3)
        formula = "= %s!C%s" % (self.day_of_week[self.i], "9")
        cell.value = formula
        cell.style = self.quad_bottom
        self.ws_list[self.i].merge_cells('C12:E12')
        cell.number_format = "#,###.00;[RED]-#,###.00"
        # Totals Right
        cell = self.ws_list[self.i].cell(row=12, column=6)
        formula = "= SUM(%s!F%s + %s!F%s)" % (self.day_of_week[self.i], "9", self.day_of_week[self.i], "11")
        cell.value = formula
        cell.style = self.quad_bottom
        self.ws_list[self.i].merge_cells('F12:H12')
        cell.number_format = "#,###.00;[RED]-#,###.00"
        # Availability Quadrants
        # Top Left
        cell = self.ws_list[self.i].cell(row=8, column=10)  # CCAs Available to 10 hours Page 3
        cell.value = "CCAs \nAvailable to 10 hours \nPage 3"
        cell.style = self.quad_top
        self.ws_list[self.i].merge_cells('J8:L8')
        self.cellj9 = self.ws_list[self.i].cell(row=9, column=10)  # value filled later
        self.cellj9.style = self.quad_bottom
        self.ws_list[self.i].merge_cells('J9:L9')
        self.cellj9.number_format = "#,###.00;[RED]-#,###.00"
        # Top Right
        cell = self.ws_list[self.i].cell(row=8, column=13)  # CCAs Available to 11.5 hours Page 3
        cell.value = "CCAs \nAvailable to 11.5 hours \nPage 3"
        cell.style = self.quad_top
        self.ws_list[self.i].merge_cells('M8:O8')
        self.cellm9 = self.ws_list[self.i].cell(row=9, column=13)  # value filled later
        self.cellm9.style = self.quad_bottom
        self.ws_list[self.i].merge_cells('M9:O9')
        self.cellm9.number_format = "#,###.00;[RED]-#,###.00"
        # Bottom Left
        cell = self.ws_list[self.i].cell(row=10, column=10)  # OTDL Available to 10 hours Page 4
        cell.value = "OTDL \nAvailable to 10 hours \nPage 4"
        cell.style = self.quad_top
        self.ws_list[self.i].merge_cells('J10:L10')
        self.cellj11 = self.ws_list[self.i].cell(row=11, column=10)  # value filled later
        self.cellj11.style = self.quad_bottom
        self.ws_list[self.i].merge_cells('J11:L11')
        self.cellj11.number_format = "#,###.00;[RED]-#,###.00"
        # Bottom Right
        cell = self.ws_list[self.i].cell(row=10, column=13)  # OTDL Available to 12 hours Page 4
        cell.value = "OTDL \nAvailable to 12 hours \nPage 4"
        cell.style = self.quad_top
        self.ws_list[self.i].merge_cells('M10:O10')
        self.cellm11 = self.ws_list[self.i].cell(row=11, column=13)  # value filled later
        self.cellm11.style = self.quad_bottom
        self.ws_list[self.i].merge_cells('M11:O11')
        self.cellm11.number_format = "#,###.00;[RED]-#,###.00"
        # Totals Left
        cell = self.ws_list[self.i].cell(row=12, column=10)
        formula = "= SUM(%s!J%s + %s!J%s)" % (self.day_of_week[self.i], "9", self.day_of_week[self.i], "11")
        cell.value = formula
        cell.style = self.quad_bottom
        self.ws_list[self.i].merge_cells('J12:L12')
        cell.number_format = "#,###.00;[RED]-#,###.00"
        # Totals Right
        cell = self.ws_list[self.i].cell(row=12, column=13)
        formula = "= SUM(%s!M%s + %s!M%s)" % (self.day_of_week[self.i], "9", self.day_of_week[self.i], "11")
        cell.value = formula
        cell.style = self.quad_bottom
        self.ws_list[self.i].merge_cells('M12:O12')
        cell.number_format = "#,###.00;[RED]-#,###.00"
        # DOV box
        cell = self.ws_list[self.i].cell(row=14, column=3)  # Dispatch of Value
        cell.value = "Dispatch of Value"
        cell.style = self.quad_top
        self.ws_list[self.i].merge_cells('C14:E15')
        cell = self.ws_list[self.i].cell(row=16, column=3)  # aquired by get_dov()
        cell.value = self.dovarray[self.i]
        cell.style = self.quad_bottom
        self.ws_list[self.i].merge_cells('C16:E16')
        cell.number_format = "#,###.00;[RED]-#,###.00"

        # DOV Violations box
        cell = self.ws_list[self.i].cell(row=14, column=6)  # Dispatch of Value
        cell.value = "Carriers out past \nDispatch of Value"
        cell.style = self.quad_top
        self.ws_list[self.i].merge_cells('F14:H15')
        self.cellf16 = self.ws_list[self.i].cell(row=16, column=6)  # value filled later - # of carriers out past DOV
        self.cellf16.style = self.quad_bottom
        self.ws_list[self.i].merge_cells('F16:H16')
        self.cellf16.number_format = "#,##0"

        # Straight Time Available:
        cell = self.ws_list[self.i].cell(row=14, column=10)  # Straight Time Available:
        cell.value = "Straight Time Available:"
        cell.style = self.quad_left
        self.ws_list[self.i].merge_cells('J14:M14')
        cell = self.ws_list[self.i].cell(row=14, column=14)  # straight time available formula
        cell.value = ""
        cell.style = self.quad_right
        self.ws_list[self.i].merge_cells('N14:O14')

        # Available to DOV:
        cell = self.ws_list[self.i].cell(row=16, column=10)  # Available to DOV:
        cell.value = "Available to DOV:"
        cell.style = self.quad_left
        self.ws_list[self.i].merge_cells('J16:M16')

        self.celln16 = self.ws_list[self.i].cell(row=16, column=14)  # avaiable to dov formula
        self.celln16.style = self.quad_right
        self.ws_list[self.i].merge_cells('N16:O16')
        self.celln16.number_format = "#,###.00;[RED]-#,###.00"
        self.row = 19  # starts on row 19 to give room to the quadrants

    def list_loop(self):
        """ loops four times. once for each list. """
        self.lsi = 0  # iterations of the list loop method
        for _ in self.ot_list:  # loops for nl, wal, otdl and aux
            self.list_and_column_headers()  # builds headers for list and columns
            self.carrierlist_mod()
            self.get_first_row()
            self.row_number = 1  # initialize the row number that appears on the far left column
            self.carrierloop()  # loop once to fill a row with carrier rings data
            self.fill_for_minrows()  # fill in blank rows to fullfill minrows requirement
            self.get_listrange()
            self.dayrange.append(self.listrange)  # get rows information for each list loop.
            self.build_footer()
            self.pagebreak()
            self.increment_progbar()
            self.lsi += 1
        self.lsi = 0  # reset list loop iteration

    def list_and_column_headers(self):
        """ builds headers for list and column """
        n_14heads = ("On Own \nRoute", "Off Route", "Available \nto 10", "Available \nto 10*")
        o_15heads = ("Off Route", "Other Route", "Available \nto 11.50", "Available \nto 12.00*")
        p_16heads = ("Other Route", "", "Available \nto DOV", "Available \nto DOV")

        cell = self.ws_list[self.i].cell(row=self.row, column=1)
        cell.value = self.page_titles[self.lsi]  # Displays the page title for each list,
        cell.style = self.list_header
        self.ws_list[self.i].merge_cells('A' + str(self.row) + ':O' + str(self.row))
        cell = self.ws_list[self.i].cell(row=self.row, column=16)
        cell.value = "Page {}".format(self.lsi+1)  # Displays the page title for each list,
        cell.style = self.list_header
        self.row += 2
        self.ws_list[self.i].row_dimensions[self.row].height = 30  # adjust row height
        cell = self.ws_list[self.i].cell(row=self.row, column=2)  # Name Header
        cell.value = "Name"
        cell.style = self.col_header_left
        self.ws_list[self.i].merge_cells('B' + str(self.row) + ':D' + str(self.row))

        cell = self.ws_list[self.i].cell(row=self.row, column=5)  # OTDL List/Status
        cell.value = "OTDL \nList"
        if self.ot_list[self.lsi] == "Auxiliary":  # Aux has header variation
            cell.value = "status"
        cell.style = self.col_header_left
        self.ws_list[self.i].merge_cells('E' + str(self.row) + ':F' + str(self.row))

        cell = self.ws_list[self.i].cell(row=self.row, column=7)  # Route Assigned
        cell.value = "Route \nAssigned"
        cell.style = self.col_header_left

        cell = self.ws_list[self.i].cell(row=self.row, column=8)  # BT
        cell.value = "BT"
        cell.style = self.col_header

        cell = self.ws_list[self.i].cell(row=self.row, column=9)  # MV
        cell.value = "MV"
        cell.style = self.col_header

        cell = self.ws_list[self.i].cell(row=self.row, column=10)  # RS
        cell.value = "RS"
        cell.style = self.col_header

        cell = self.ws_list[self.i].cell(row=self.row, column=11)  # ET
        cell.value = "ET"
        cell.style = self.col_header

        cell = self.ws_list[self.i].cell(row=self.row, column=12)  # Overtime Worked
        cell.value = "Overtime \nWorked"
        cell.style = self.col_header_left
        self.ws_list[self.i].merge_cells('L' + str(self.row) + ':M' + str(self.row))

        cell = self.ws_list[self.i].cell(row=self.row, column=14)  # On/off route, avail to 10
        cell.value = n_14heads[self.lsi]
        cell.style = self.col_header_left

        cell = self.ws_list[self.i].cell(row=self.row, column=15)  # off/other route, avail to 11.5/12
        cell.value = o_15heads[self.lsi]
        cell.style = self.col_header_left

        if self.lsi != 1:  # do not display this column for work assignment carriers
            cell = self.ws_list[self.i].cell(row=self.row, column=16)  # other route, avail to DOV
            cell.value = p_16heads[self.lsi]
            cell.style = self.col_header_left
        self.row += 1  # increment the row so first name starts on fresh line.

    def carrierlist_mod(self):
        """ get the carrier list appropriate to the day and list status """
        self.mod_carrierlist = self.carrier_breakdown[self.i][self.lsi]

    def get_first_row(self):
        """ record the number of the first row for totals formulas in footers """
        self.first_row = self.row

    def carrierloop(self):
        """ loop for each carrier """
        self.display_counter = 0  # count the number of rows displayed
        for carrier in self.mod_carrierlist:
            self.carrier = carrier[1]  # current iteration of carrier list is assigned self.carrier
            self.list_ = carrier[2]  # get the list status of the carrier
            self.route = carrier[4]  # get the route of the carrier
            self.display_route()  # will alter self.route if the route is a floater/t6 string.
            self.get_rings()  # get individual carrier rings for the day
            self.number_crunching()  # do calculations to get overtime and availability
            if self.qualify():  # test the rings to see if they need to be displayed
                self.display_recs()  # build the carrier and the rings row into the spreadsheet
                self.display_counter += 1
                self.row += 1

    def fill_for_minrows(self):
        """ fill in blank rows to fullfill minrows requirement. """
        self.carrier = ""  # current iteration of carrier list is assigned self.carrier
        self.list_ = ""  # get the list status of the carrier
        self.route = ""  # get the route of the carrier
        self.totalhours = 0.0  # set default as an empty string
        self.bt = ""  # begin tour
        self.rs = ""  # return to station
        self.et = ""  # end tour
        self.codes = ""  # codes from carrier rings
        self.moves = ""  # moves from carrier rings
        self.overtime = 0.0  # the total overtime worked
        self.onroute = 0.0  # the amount of overtime worked on the carrier's own route.
        self.offroute = 0.0  # total time spend off route
        self.offroute_adj = 0.0  # self.offroute adjusted for pivot time, ns days, and whole days off bid assignment
        self.otherroute_array = []  # a list of routes where carrier worked off assignment
        self.otherroute = ""  # a formatted display for routes worked off assignment
        self.avail_10 = 0.0  # otdl/aux availability to 10 hours
        self.avail_115 = 0.0  # aux availability to 11.50 hours
        self.avail_12 = 0.0  # otdl availability to 12 hours.
        blank_lines = []  # make an array for blank lines
        if self.pref[self.lsi] in ("nl",):  # if "no list"
            minrows = self.min4_ss_nl
        elif self.pref[self.lsi] in ("wal",):  # if "work assignment list"
            minrows = self.min4_ss_wal
        elif self.pref[self.lsi] in ("otdl",):  # if "overtime desired list"
            minrows = self.min4_ss_otdl
        else:  # if "auxiliary"
            minrows = self.min4_ss_aux
        while self.display_counter < minrows:  # until carrier list quantity matches minrows
            add_this = ('', '', '', '', '', '')
            blank_lines.append(add_this)  # append empty recs to carrier list
            self.display_counter += 1
        for _ in blank_lines:
            self.display_recs()  # put the carrier and the first part of rings into the spreadsheet
            self.row += 1

    def display_route(self):
        """ formats route number for floater/t6 strings into a short version """
        if self.route:
            route = self.route.split("/")
            if len(route) == 5:
                self.route = "T6: {} +".format(route[0])

    def get_rings(self):
        """ get individual carrier rings for the day """
        self.rings = Rings(self.carrier, self.dates[self.i]).get_for_day()  # assign as self.rings
        self.totalhours = 0.0  # set default as an empty string
        self.bt = ""
        self.rs = ""
        self.et = ""
        self.codes = ""
        self.moves = ""
        self.lvtype = ""
        self.lvtime = ""
        if self.rings[0]:  # if rings record is not blank
            self.totalhours = float(Convert(self.rings[0][2]).zero_not_empty())
            self.bt = self.rings[0][9]
            self.rs = self.rings[0][3]
            self.et = self.rings[0][10]
            self.codes = self.rings[0][4]
            self.moves = self.rings[0][5]
            self.lvtype = self.rings[0][6]
            self.lvtime = self.rings[0][7]

    def number_crunching(self):
        """ crunch numbers to get overtime, off route, other route and availability"""
        self.overtime = 0.0  # the total overtime worked
        self.onroute = 0.0  # the amount of overtime worked on the carrier's own route.
        self.offroute = 0.0  # total time spend off route
        self.offroute_adj = 0.0
        self.otherroute_array = []  # a list of routes where carrier worked off assignment
        self.otherroute = ""  # a formatted display for routes worked off assignment
        self.avail_10 = 0.0  # otdl/aux availability to 10 hours
        self.avail_115 = 0.0  # aux availability to 11.50 hours
        self.avail_12 = 0.0  # otdl availability to 12 hours.
        self.calc_overtime()  # calculate the amount of overtime worked
        if self.pref[self.lsi] in ("nl", "wal"):
            if self.moves:
                self.calc_offroute()  # calculate the time that the carrier spent off their route and get other route
                self.format_otherroute()  # format the self.other route so that if fits in the spreadsheet cell
            self.calc_offroute_adj()  # adj for pivot time or if code is nsday or whole day spent off route
        if self.pref[self.lsi] == "nl":
            self.calc_onroute()  # calculate the overtime worked on carrier's own route.
        if self.pref[self.lsi] in ("otdl", "aux"):
            self.calc_availability()
            self.moves = ""  # empty self.moves for otdl and aux carriers.

    def calc_overtime(self):
        """ calculates the amount of overtime worked. if it is the carrier's ns day, then the full day is overtime. """
        if self.codes == "ns day":
            self.overtime = self.totalhours
        else:
            self.overtime = max(self.totalhours - 8, 0)

    def calc_offroute(self):
        """ calculate the time that the carrier spent off their route assignment, get other route """
        moves = self.moves.split(",")
        move_sets = int(len(moves)/3)  # get the number of triads in the moves array
        count = 0
        for _ in range(move_sets):
            offroute = float(moves[count+1]) - float(moves[count])  # calculate off route time per triad
            self.offroute += offroute  # add triad time off route
            self.otherroute_array.append(moves[count+2])
            count += 3
        self.offroute = round(self.offroute, 2)
        if self.offroute >= self.totalhours:
            self.offroute = self.totalhours
        self.moves = moves[0]  # replace moves with the first time moved off route

    def calc_offroute_adj(self):
        """ calculate the off route overtime for ns days or if the whole day is spent off own route. """
        self.offroute_adj = min(self.overtime, self.offroute)  # will adjust for pivot time
        if self.codes == "ns day":  # if it is the ns day, then whole day is off route
            self.offroute_adj = self.totalhours
            self.otherroute_array.append("ns day")
            self.otherroute = "ns day"
            self.moves = self.bt
        if self.totalhours:
            # if self.totalhours <= self.offroute:  # if the whole day is off route
            if self.offroute == self.totalhours:  # if the whole day is off route
                self.offroute_adj = self.totalhours
                self.otherroute = "off bid"

    def calc_onroute(self):
        """ calculate the overtime the carrier worked on their own route. """
        if self.codes == "ns day":
            self.onroute = 0
        else:
            self.onroute = max(self.overtime - self.offroute, 0)

    def format_otherroute(self):
        """ format the self.other route so that if fits in the spreadsheet cell. """
        if self.otherroute_array:
            if len(self.otherroute_array) > 1:
                # format like "1024 + 1"
                self.otherroute = self.otherroute_array[0] + "+" + str(len(self.otherroute_array) - 1)
            else:
                # format like "1024"
                self.otherroute = self.otherroute_array[0]

    def calc_availability(self):
        """ calculate otdl and aux availability """
        if not self.totalhours and self.codes == "no call":  # if the carrier was not scheduled for the day
            self.avail_10 = 10  # otdl/aux availability to 10 hours
            self.avail_115 = 11.5  # aux availability to 11.50 hours
            self.avail_12 = 12  # otdl availability to 12 hours.
            if self.list_ == "odln":  # odln carriers are available for only 8 hours when working on their ns day.
                self.avail_10 = 8  # odln availability to 8 hours
                self.avail_12 = 8  # odln availability to 8 hours.
            return
        if self.codes in ("light", "excused", "sch chg", "annual", "sick"):  # if carrier excused for day
            self.avail_10 = 0  # otdl/aux availability to 0 hours
            self.avail_115 = 0  # aux availability to 0 hours
            self.avail_12 = 0  # otdl availability to 0 hours.
            return
        if not self.totalhours:
            return
        self.avail_10 = max(10 - self.totalhours, 0)
        self.avail_115 = max(11.5 - self.totalhours, 0)
        self.avail_12 = max(12 - self.totalhours, 0)
        if self.list_ == "odln":  # odln carriers are available for only 8 hours when working on their ns day.
            self.avail_10 = max(8 - self.totalhours, 0)  # odln availability to 8 hours
            self.avail_12 = max(8 - self.totalhours, 0)  # odln availability to 8 hours

    def qualify(self):
        """ check to see if the carrier information needs to be displayed. """
        if self.pref[self.lsi] in ("otdl", "aux"):  # display all for otdl and aux
            return True
        if self.display_limiter == "show all":  # display all if the limiter is set to "show all"
            return True
        if self.display_limiter == "only workdays":  # display only days when the carrier worked.
            if self.totalhours:
                return True
            return False
        if self.display_limiter == "only mandates":
            if self.pref[self.lsi] == "nl":
                if self.overtime or self.offroute_adj:
                    return True
                return False
            if self.pref[self.lsi] == "wal":
                if self.offroute_adj:
                    return True
                return False

    def _display_odl_mod_name(self, name):
        """ return a modified name is the names is in the odlr or odln indicator arrays. """
        if name in self.odlr_indicator and name in self.odln_indicator:
            return name + " (odl+)"
        elif name in self.odlr_indicator:
            return name + " (odlr)"
        elif name in self.odln_indicator:
            return name + " (odln)"
        else:
            return name

    def display_recs(self):
        """ put the carrier and the first part of rings into the spreadsheet - it's show time! """
        cell = self.ws_list[self.i].cell(row=self.row, column=1)  # row number
        cell.value = "{}.".format(self.row_number)
        cell.style = self.input_s
        self.row_number += 1  # increment the row number
        cell = self.ws_list[self.i].cell(row=self.row, column=2)  # name
        cell.value = self._display_odl_mod_name(self.carrier)
        cell.style = self.input_name
        self.ws_list[self.i].merge_cells('B' + str(self.row) + ':D' + str(self.row))
        cell = self.ws_list[self.i].cell(row=self.row, column=5)  # list status
        cell.value = self.list_dict[self.list_]
        cell.style = self.input_s
        self.ws_list[self.i].merge_cells('E' + str(self.row) + ':F' + str(self.row))
        cell = self.ws_list[self.i].cell(row=self.row, column=7)  # route
        cell.value = self.route
        cell.style = self.input_s
        cell = self.ws_list[self.i].cell(row=self.row, column=8)  # begin tour
        cell.value = Convert(self.bt).str_to_floatoremptystr()
        cell.style = self.input_s
        cell.number_format = "#,###.00;[RED]-#,###.00"
        cell = self.ws_list[self.i].cell(row=self.row, column=9)  # move
        cell.value = Convert(self.moves).str_to_floatoremptystr()
        cell.style = self.input_s
        cell.number_format = "#,###.00;[RED]-#,###.00"
        cell = self.ws_list[self.i].cell(row=self.row, column=10)  # return to station
        cell.value = Convert(self.rs).str_to_floatoremptystr()
        cell.style = self.input_s
        cell.number_format = "#,###.00;[RED]-#,###.00"
        cell = self.ws_list[self.i].cell(row=self.row, column=11)  # end tour
        cell.value = Convert(self.et).str_to_floatoremptystr()
        cell.style = self.input_s
        cell.number_format = "#,###.00;[RED]-#,###.00"
        cell = self.ws_list[self.i].cell(row=self.row, column=12)  # overtime worked
        cell.value = Convert(self.overtime).str_to_floatoremptystr()
        cell.style = self.input_s
        cell.number_format = "#,###.00;[RED]-#,###.00"
        self.ws_list[self.i].merge_cells('L' + str(self.row) + ':M' + str(self.row))
        column = 14
        if self.pref[self.lsi] == "nl":
            cell = self.ws_list[self.i].cell(row=self.row, column=column)  # on route
            cell.value = Convert(self.onroute).str_to_floatoremptystr()
            cell.style = self.input_s
            cell.number_format = "#,###.00;[RED]-#,###.00"
            column += 1
        if self.pref[self.lsi] in ("nl", "wal"):
            cell = self.ws_list[self.i].cell(row=self.row, column=column)  # off route
            cell.value = Convert(self.offroute_adj).str_to_floatoremptystr()
            cell.style = self.input_s
            cell.number_format = "#,###.00;[RED]-#,###.00"
            column += 1
        if self.pref[self.lsi] in ("nl", "wal"):
            cell = self.ws_list[self.i].cell(row=self.row, column=column)  # other route
            cell.value = self.otherroute
            cell.style = self.input_s
            cell.number_format = "#,###.00;[RED]-#,###.00"
        if self.pref[self.lsi] in ("otdl", "aux"):
            cell = self.ws_list[self.i].cell(row=self.row, column=column)  # availability to 10
            cell.value = self.avail_10
            cell.style = self.input_s
            cell.number_format = "#,###.00;[RED]-#,###.00"
            column += 1
            if self.pref[self.lsi] == "otdl":  # change value dependant on otdl or aux
                value = self.avail_12
            else:
                value = self.avail_115
            cell = self.ws_list[self.i].cell(row=self.row, column=column)  # availability to 12 or 11.5
            cell.value = value
            cell.style = self.input_s
            cell.number_format = "#,###.00;[RED]-#,###.00"
            column += 1
            formula = "=IF(%s!J%s = \"\", \"\", IF(%s!C%s = \"\", \"no dov\", MAX(%s!C%s-%s!J%s, 0)))" % \
                      (self.day_of_week[self.i], str(self.row),
                       self.day_of_week[self.i], "16",
                       self.day_of_week[self.i], "16",
                       self.day_of_week[self.i], str(self.row))
            cell = self.ws_list[self.i].cell(row=self.row, column=column)  # availability to DOV
            cell.value = formula
            cell.style = self.input_s
            cell.number_format = "#,###.00;[RED]-#,###.00"
        self.get_last_row()  # record the number of the last row for total formulas in footers

    def get_last_row(self):
        """ record the number of the last row for totals formulas in footers """
        self.last_row = self.row

    def get_listrange(self):
        """ get the first row and last row of the current list and store them into an array called listrange.
        later, the summary row will be added."""
        self.listrange = []
        self.listrange.append(self.first_row)
        self.listrange.append(self.last_row)
        self.listrange.append(self.row)  # put the 3rd and final row number into the listrange - summary row

    def build_footer(self):
        """ call the footer depending on the list. """
        if self.pref[self.lsi] == "nl":
            self.nl_footer()
        elif self.pref[self.lsi] == "wal":
            self.wal_footer()
        elif self.pref[self.lsi] == "otdl":
            self.otdl_footer()
        else:
            self.aux_footer()
        self.row += 1

    def nl_footer(self):
        """ build the non list footer. """
        self.ws_list[self.i].row_dimensions[self.row].height = 20  # adjust row height
        cell = self.ws_list[self.i].cell(row=self.row, column=2)  # totals label
        cell.value = "     Totals:"
        cell.style = self.footer_left
        self.ws_list[self.i].merge_cells('B' + str(self.row) + ':M' + str(self.row))
        cell = self.ws_list[self.i].cell(row=self.row, column=14)  # totals own route
        formula = "=SUM(%s!N%s:N%s)" % (self.day_of_week[self.i], self.listrange[0], self.listrange[1])
        cell.value = formula
        cell.style = self.footer_mid
        cell.number_format = "#,###.00;[RED]-#,###.00"
        cell = self.ws_list[self.i].cell(row=self.row, column=15)  # totals off route
        formula = "=SUM(%s!O%s:O%s)" % (self.day_of_week[self.i], self.listrange[0], self.listrange[1])
        cell.value = formula
        cell.style = self.footer_mid
        cell.number_format = "#,###.00;[RED]-#,###.00"
        cell = self.ws_list[self.i].cell(row=self.row, column=16)  # blank formated cell
        cell.value = ""
        cell.style = self.footer_right

    def wal_footer(self):
        """ build the work assignment footer """
        self.ws_list[self.i].row_dimensions[self.row].height = 20  # adjust row height
        cell = self.ws_list[self.i].cell(row=self.row, column=2)  # totals label
        cell.value = "     Totals:"
        cell.style = self.footer_left
        self.ws_list[self.i].merge_cells('B' + str(self.row) + ':M' + str(self.row))

        cell = self.ws_list[self.i].cell(row=self.row, column=14)  # totals off route
        formula = "=SUM(%s!N%s:N%s)" % (self.day_of_week[self.i], self.listrange[0], self.listrange[1])
        cell.value = formula
        cell.style = self.footer_mid
        cell.number_format = "#,###.00;[RED]-#,###.00"
        cell = self.ws_list[self.i].cell(row=self.row, column=15)  # blank formated cell
        cell.style = self.footer_right

    def aux_footer(self):
        """ build the footer for auxiliary - cca and ptf - carriers. """
        self.ws_list[self.i].row_dimensions[self.row].height = 20  # adjust row height
        cell = self.ws_list[self.i].cell(row=self.row, column=2)  # totals label
        cell.value = "     Totals:"
        cell.style = self.footer_left
        self.ws_list[self.i].merge_cells('B' + str(self.row) + ':M' + str(self.row))
        cell = self.ws_list[self.i].cell(row=self.row, column=14)  # availability to 10
        formula = "=SUM(%s!N%s:N%s)" % (self.day_of_week[self.i], self.listrange[0], self.listrange[1])
        cell.value = formula
        cell.style = self.footer_mid
        cell.number_format = "#,###.00;[RED]-#,###.00"
        cell = self.ws_list[self.i].cell(row=self.row, column=15)  # availability to 11.50
        formula = "=SUM(%s!O%s:O%s)" % (self.day_of_week[self.i], self.listrange[0], self.listrange[1])
        cell.value = formula
        cell.style = self.footer_mid
        cell.number_format = "#,###.00;[RED]-#,###.00"
        cell = self.ws_list[self.i].cell(row=self.row, column=16)  # availability to DOV
        formula = "=SUM(%s!P%s:P%s)" % (self.day_of_week[self.i], self.listrange[0], self.listrange[1])
        cell.value = formula
        cell.style = self.footer_right
        cell.number_format = "#,###.00;[RED]-#,###.00"

    def otdl_footer(self):
        """ build the overtime desired list footer. """
        self.ws_list[self.i].row_dimensions[self.row].height = 20  # adjust row height
        cell = self.ws_list[self.i].cell(row=self.row, column=2)  # totals label
        cell.value = "     Totals:"
        cell.style = self.footer_left
        self.ws_list[self.i].merge_cells('B' + str(self.row) + ':M' + str(self.row))
        cell = self.ws_list[self.i].cell(row=self.row, column=14)  # availability to 10
        formula = "=SUM(%s!N%s:N%s)" % (self.day_of_week[self.i], self.listrange[0], self.listrange[1])
        cell.value = formula
        cell.style = self.footer_mid
        cell.number_format = "#,###.00;[RED]-#,###.00"
        cell = self.ws_list[self.i].cell(row=self.row, column=15)  # availability to 12
        formula = "=SUM(%s!O%s:O%s)" % (self.day_of_week[self.i], self.listrange[0], self.listrange[1])
        cell.value = formula
        cell.style = self.footer_mid
        cell.number_format = "#,###.00;[RED]-#,###.00"
        cell = self.ws_list[self.i].cell(row=self.row, column=16)  # availability to DOV
        formula = "=SUM(%s!P%s:P%s)" % (self.day_of_week[self.i], self.listrange[0], self.listrange[1])
        cell.value = formula
        cell.style = self.footer_right
        cell.number_format = "#,###.00;[RED]-#,###.00"
        self.row += 1
        self.ws_list[self.i].row_dimensions[self.row].height = 20  # adjust row height
        cell = self.ws_list[self.i].cell(row=self.row, column=2)  # totals label
        cell.value = "     * odln carriers are available for only 8 hours when working on their ns day."
        cell.style = self.date_dov
        self.ws_list[self.i].merge_cells('B' + str(self.row) + ':P' + str(self.row))

    def fill_quads(self):
        """ write formulas for the quadrants on the cover sheet. %s!J%s"""
        formula = "=%s!N%s" % (self.day_of_week[self.i], self.dayrange[0][2])  # cell N+nl summary
        self.cellc9.value = formula  # non otdl own route violations
        formula = "=%s!O%s" % (self.day_of_week[self.i], self.dayrange[0][2])  # cell O+nl summary
        self.cellf9.value = formula  # non otdl off route violations
        formula = "=%s!N%s" % (self.day_of_week[self.i], self.dayrange[1][2])  # cell O+wal summary
        self.cellf11.value = formula  # wal off route violations
        formula = "=%s!N%s" % (self.day_of_week[self.i], self.dayrange[2][2])  # cell N+aux summary
        self.cellj9.value = formula  # aux availability to 10 hours
        formula = "=%s!O%s" % (self.day_of_week[self.i], self.dayrange[2][2])  # cell O+aux summary
        self.cellm9.value = formula  # aux availability to 11.5 hours
        formula = "=%s!N%s" % (self.day_of_week[self.i], self.dayrange[3][2])  # cell N+otdl summary
        self.cellj11.value = formula  # otdl availability to 10 hours
        formula = "=%s!O%s" % (self.day_of_week[self.i], self.dayrange[3][2])  # cell O+otdl summary
        self.cellm11.value = formula  # otdl availability to 12 hours
        formula = "=SUM(%s!P%s + %s!P%s)" \
                  % (self.day_of_week[self.i], self.dayrange[2][2],  # cell P+aux summary
                     self.day_of_week[self.i], self.dayrange[3][2])  # cell P+otdl summary
        self.celln16.value = formula
        formula = "=MAX(" \
                  "COUNTIF(%s!J%s:J%s, \">\"&%s!C%s) + " \
                  "COUNTIF(%s!J%s:J%s, \">\"&%s!C%s) + " \
                  "COUNTIF(%s!J%s:J%s, \">\"&%s!C%s) + " \
                  "COUNTIF(%s!J%s:J%s, \">\"&%s!C%s)," \
                  "0)" \
                  % (self.day_of_week[self.i], self.dayrange[0][0], self.dayrange[0][1], self.day_of_week[self.i], "16",
                     self.day_of_week[self.i], self.dayrange[1][0], self.dayrange[1][1], self.day_of_week[self.i], "16",
                     self.day_of_week[self.i], self.dayrange[2][0], self.dayrange[2][1], self.day_of_week[self.i], "16",
                     self.day_of_week[self.i], self.dayrange[3][0], self.dayrange[3][1], self.day_of_week[self.i], "16")
        self.cellf16.value = formula  # carriers out past dispatch of value

    def pagebreak(self, force=False):
        """ create a page break if consistant with user preferences. If page break is True, then the page
        break can not be skipped. This ensures that there is always a page break after the summary page."""
        if self.pref[self.lsi] == "nl" and not self.pb4_nl_wal:
            if not force:
                self.row += 1
                return
        if self.pref[self.lsi] == "wal" and not self.pb4_wal_aux:
            if not force:
                self.row += 1
                return
        if self.pref[self.lsi] == "aux" and not self.pb4_aux_otdl:
            if not force:
                self.row += 1
                return
        if self.pref[self.lsi] == "otdl":
            if not force:
                self.row += 1
                return
        try:
            self.ws_list[self.i].page_breaks.append(Break(id=self.row))
        except AttributeError:
            self.ws_list[self.i].row_breaks.append(Break(id=self.row))  # effective for windows
        self.row += 1

    def increment_progbar(self):
        """ move the progress bar, update with info on what is being done """
        lst = ("No List", "Work Assignment", "Overtime Desired", "Auxiliary")
        self.pbi += 1
        self.pb.move_count(self.pbi)  # increment progress bar
        self.pb.change_text("Building day {}: list: {}".format(self.day.strftime("%A"), lst[self.lsi]))

    def save_open(self):
        """ name and open the excel file """
        self.pbi += 1
        self.pb.move_count(self.pbi)  # increment progress bar
        self.pb.change_text("Saving...")
        self.pb.stop()
        r = "_w"
        if not projvar.invran_weekly_span:  # if investigation range is daily
            r = "_d"
        xl_filename = "man4" + str(format(self.dates[0], "_%y_%m_%d")) + r + ".xlsx"
        try:
            self.wb.save(dir_path('mandates_4') + xl_filename)
            messagebox.showinfo("Spreadsheet generator",
                                "Your spreadsheet was successfully generated. \n"
                                "File is named: {}".format(xl_filename),
                                parent=self.frame)
            if sys.platform == "win32":
                os.startfile(dir_path('mandates_4') + xl_filename)
            if sys.platform == "linux":
                subprocess.call(["xdg-open", 'kb_sub/mandates_4/' + xl_filename])
            if sys.platform == "darwin":
                subprocess.call(["open", dir_path('mandates_4') + xl_filename])
        except PermissionError:
            messagebox.showerror("Spreadsheet generator",
                                 "The spreadsheet was not opened. \n"
                                 "Suggestion: "
                                 "Make sure that identically named spreadsheets are closed "
                                 "(the file can't be overwritten while open).",
                                 parent=self.frame)


class ImpManSpreadsheet5:
    """ this table will create spreadsheets which can be copy/pasted into improper mandate grievance contentions """

    def __init__(self):
        self.frame = None  # the frame of parent
        self.pb = None  # progress bar object
        self.pbi = 0  # progress bar count index
        self.startdate = None  # start date of the investigation
        self.enddate = None  # ending date of the investigation
        self.dates = []  # all days of the investigation
        self.carrierlist = []  # all carriers in carrier list
        self.carrier_breakdown = []  # all carriers in carrier list broken down into appropiate list
        self.mod_carrierlist = []
        self.remedy_tolerance = 0.0  # get tolerances from tolerances table.self.remedy_tolerance
        self.max_pivot = 0.0
        self.impman5_fullreport = False
        self.impman5_report = True
        self.wb = None  # the workbook object
        self.remedy_sheet = None  # the work sheet for remedies
        self.award_sheet = None  # the work sheet for awards
        self.report = None  # the text document
        self.report_filename = None  # the file name of the report
        self.ws_list = []  # "saturday", "sunday", "monday", "tuesday", "wednesday", "thursday", "friday"
        self.day_of_week = []  # seven day array for weekly investigations/ one day array for daily investigations
        # styles for worksheet
        self.ws_header = None  # style
        self.list_header = None  # style
        self.date_dov = None  # style
        self.date_dov_title = None  # style
        self.col_header = None  # style
        self.input_name = None  # style
        self.input_list = None  # style
        self.instruct_text = None  # style
        self.input_s = None  # style
        self.calcs = None  # style
        self.col_header_left = None  # style
        self.col_header_right = None  # style
        self.col_header = None  # style
        self.footer_left = None  # style
        self.footer_right = None  # style
        self.footer_mid = None  # style
        self.day = None  # build worksheet - loop once for each day
        self.i = 0  # build worksheet loop iteration
        self.lsi = 0  # list loop iteration
        self.row = 1
        self.page_titles = ("NON-OTDL and Work Assignment Employees that worked overtime",
                            "OTDL/Auxiliary Employees who were available to work overtime")
        self.carrier = None  # current iteration of carrier's name is assigned self.carrier
        self.list_ = None  # current iteration of carrier's list status is assigned self.carrier
        self.nsday = None  # current iteration of carriers non scheduled day is assigned to self.carrier
        self.route = None  # current iteration of carrier's route is assigned self.carrier
        self.ot_available = False  # is it the carrier's non scheduled day?
        self.rings = []  # assign as self.rings
        self.totalhours = 0.0  # set default as an empty string
        self.bt = ""
        self.rs = ""
        self.et = ""
        self.codes = ""
        self.moves = ""
        self.mandate_names = [[], [], [], [], [], [], []]  # multiple array, contains names of mandated carriers
        self.available_names = [[], [], [], [], [], [], []]  # multiple array, contains names of available carriers
        self.otdl_names = [[], [], [], [], [], [], []]  # multiple array, contains names of available carriers
        self.odlr_names = [[], [], [], [], [], [], []]  # multiple array, contains names of available carriers
        self.odln_names = [[], [], [], [], [], [], []]  # multiple array, contains names of available carriers
        self.aux_names = [[], [], [], [], [], [], []]  # multiple array, contains names of available carriers
        self.mandate_totals = [0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0]  # contains totals of mandated carriers
        self.available_totals = [0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0]  # contains totals of available carriers
        self.otdl_totals = [0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0]  # contains totals of available carriers
        self.aux_totals = [0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0]  # contains totals of available carriers
        self.mentioned_names = []  # gives a list of carriers mentioned in the grievance
        self.remedy_array = []  # hold arrays with data about possible violations
        self.award_array = []  # hold name, list and row of carriers in awards array
        self.cum_hr_dict = {}  # a dictionary to hold cumulative hours for a specific carrier
        self.cum_ot_dict = {}  # a dictionary to hold cumulative overtime hours for a specific carrier
        self.avail_ot_dict = {}  # a dictionary that holds prior available ot for the previous day.
        self.odlr_indicator = []  # indicates that carrier is odlr for at least one day
        self.odln_indicator = []  # indicates that carrier is odln for at least one day
        self.avail_max = 0  # the maximum amount of availability for a carrier on a given day
        self.overtime = 0.0  # the amount of overtime worked by the carrier
        self.onroute = 0.0  # the amount of overworked on the carrier's own route.
        self.offroute = 0.0  # empty string or calculated time that carrier spent off their assignment
        self.offroute_adj = 0.0  # self.offroute adjusted for pivot time, ns days, and whole days off bid assignment
        self.otherroute_array = []  # a list of routes where carrier worked off assignment
        self.otherroute = ""  # the off assignment route the carrier worked on - formated for the cell
        self.overtime_rate = 0.0  # otdl and aux available for ot rate remedy
        self.penalty_rate = 0.0  # otdl and aux available for penalty remedy
        self.lvtype = ""
        self.lvtime = ""
        # build a dictionary for displaying list statuses on spreadsheet
        self.list_dict = {'': '', 'nl': 'non list', 'wal': 'wal', 'otdl': 'otdl', 'aux': 'cca', 'ptf': 'ptf',
                          'odlr': 'odlr', 'odln': 'odln'}
        self.rem_man_row_end = 0
        self.rem_avail_row_start = 0
        self.dovarray = []  # build a list of 7 dov times. One for each day.
        self.inv_range = 1  # number of days in the investigation range - either 1 or 7
        self.date_index = None  # a number for the day of the week ie sat = 0, sun = 1, mon = 2, etc

    def create(self, frame):
        """ a master method for running other methods in proper order."""
        self.frame = frame
        if not self.ask_ok():  # abort if user selects cancel from askokcancel
            return
        self.pb = ProgressBarDe(label="Building Improper Mandates Spreadsheet")
        self.pb.max_count(100)  # set length of progress bar
        self.pb.start_up()  # start the progress bar
        self.pbi = 1
        self.pb.move_count(self.pbi)  # increment progress bar
        self.pb.change_text("Gathering Data... ")
        self.get_dates()
        self.get_settings()
        self.get_pb_max_count()  # set the length of the progress bar
        self.get_carrierlist()
        self.get_carrier_breakdown()  # breakdown carrier list into non-otdl and available
        self.get_dov()  # get the dispatch of value for each day
        self.get_styles()
        self.build_workbook()
        self.build_text_doc()
        self.set_dimensions()
        self.build_ws_loop()  # loop once for each day
        self.build_remedy()  # build the worksheet for remedies
        self.build_awards()  # build the worksheet for awards
        self.write_mentioned_names()  # write all names mentioned at the end of the text report
        self.save_open()
        self.save_open_report()

    def ask_ok(self):
        """ ends process if user cancels """
        if messagebox.askokcancel("Spreadsheet generator",
                                  "Do you want to generate an \nImproper Mandates No. 5 Spreadsheet?",
                                  parent=self.frame):
            return True
        return False

    def get_dates(self):
        """ get the dates from the project variables """
        self.startdate = projvar.invran_date  # set daily investigation range as default - get start date
        self.enddate = projvar.invran_date  # get end date
        self.dates = [projvar.invran_date, ]  # create an array of days - only one day if daily investigation range
        self.date_index = DateHandler(projvar.invran_date).find_index()
        if projvar.invran_weekly_span:  # if the investigation range is weekly
            date = projvar.invran_date_week[0]
            self.startdate = projvar.invran_date_week[0]
            self.enddate = projvar.invran_date_week[6]
            self.dates = []
            for _ in range(7):  # create an array with all the days in the weekly investigation range
                self.dates.append(date)
                date += timedelta(days=1)
            self.inv_range = 7

    def get_settings(self):
        """ get the tolerances for the spreadsheet from the database """
        sql = "SELECT tolerance FROM tolerances"
        result = inquire(sql)
        self.remedy_tolerance = float(Convert(result[56][0]).hundredths())
        self.max_pivot = float(Convert(result[57][0]).hundredths())
        self.impman5_fullreport = Convert(result[58][0]).str_to_bool()
        self.impman5_report = Convert(result[59][0]).str_to_bool()

    def get_pb_max_count(self):
        """ set length of progress bar """
        self.pb.max_count((len(self.dates)*2)+1+1)  # once for each list in each day and remedy plus saving

    def get_carrierlist(self):
        """ get record sets for all carriers """
        self.carrierlist = CarrierList(self.startdate, self.enddate, projvar.invran_station).get()

    def get_carrier_breakdown(self):
        """ breakdown carrier list into no list, wal, otdl, aux """
        timely_rec = []
        for day in self.dates:
            non_otdl_array = []  # non otdl and work assignment carriers mandates to work overtime
            available_array = []  # otdl and auxiliary carriers available to carrier overtime
            for carrier in self.carrierlist:
                for rec in reversed(carrier):
                    if Convert(rec[0]).dt_converter() <= day:
                        timely_rec = rec
                if timely_rec[2] == "nl":
                    non_otdl_array.append(timely_rec)
                if timely_rec[2] == "wal":
                    non_otdl_array.append(timely_rec)
                if timely_rec[2] == "otdl":
                    available_array.append(timely_rec)
                if timely_rec[2] == "odlr":  # for odl regular day only -
                    if timely_rec[1] not in self.odlr_indicator:  # add name to odlr indicator array
                        self.odlr_indicator.append(timely_rec[1])
                    # if it is sunday or their ns day, put the record in the nl array
                    if day.strftime("%a") == projvar.ns_code[timely_rec[3]] or day.strftime("%a") == "Sun":
                        non_otdl_array.append(timely_rec)
                    else:  # if it is a sunday or their ns day, put record in no list array.
                        available_array.append(timely_rec)
                if timely_rec[2] == "odln":  # for odl non scheduled day only
                    if timely_rec[1] not in self.odln_indicator:  # add name to odln indicator array
                        self.odln_indicator.append(timely_rec[1])
                    # if it is sunday or their ns day, put the record in the otdl array
                    if day.strftime("%a") == projvar.ns_code[timely_rec[3]] or day.strftime("%a") == "Sun":
                        available_array.append(timely_rec)
                    else:
                        non_otdl_array.append(timely_rec)
                if timely_rec[2] == "aux" or timely_rec[2] == "ptf":
                    available_array.append(timely_rec)
            daily_breakdown = [non_otdl_array, available_array]
            self.carrier_breakdown.append(daily_breakdown)

    def get_dov(self):
        """ get the dov records currently in the database """
        days = ("sat", "sun", "mon", "tue", "wed", "thu", "fri")
        for i in range(len(days)):
            sql = "SELECT * FROM dov WHERE eff_date <= '%s' AND station = '%s' AND day = '%s' " \
                  "ORDER BY eff_date DESC" % \
                  (projvar.invran_date_week[0], projvar.invran_station, days[i])
            result = inquire(sql)
            for rec in result:
                if rec[0] == Convert(projvar.invran_date_week[0]).dt_to_str():
                    self.dovarray.append(rec[3])
                    break
                elif rec[4] == "False":
                    self.dovarray.append(rec[3])
                    break
                else:
                    continue

    def get_styles(self):
        """ Named styles for workbook """
        bd = Side(style='thin', color="80808080")  # defines borders
        self.ws_header = NamedStyle(name="ws_header", font=Font(bold=True, name='Arial', size=12))
        self.list_header = NamedStyle(name="list_header", font=Font(bold=True, name='Arial', size=10))
        self.date_dov = NamedStyle(name="date_dov", font=Font(name='Arial', size=8))
        self.date_dov_title = NamedStyle(name="date_dov_title", font=Font(bold=True, name='Arial', size=8),
                                         alignment=Alignment(horizontal='right'))
        self.col_header_left = NamedStyle(name="col_header_left", font=Font(bold=True, name='Arial', size=8),
                                          alignment=Alignment(horizontal='left', vertical='bottom'))
        self.col_header_right = NamedStyle(name="col_header_right", font=Font(bold=True, name='Arial', size=8),
                                           alignment=Alignment(horizontal='right', vertical='bottom'))
        self.col_header = NamedStyle(name="col_header", font=Font(bold=True, name='Arial', size=8),
                                     alignment=Alignment(horizontal='center', vertical='bottom'))
        self.input_name = NamedStyle(name="input_name", font=Font(name='Arial', size=8),
                                     border=Border(left=bd, top=bd, bottom=bd))
        self.input_list = NamedStyle(name="input_list", font=Font(name='Arial', size=8),
                                     border=Border(top=bd, right=bd, bottom=bd),
                                     alignment=Alignment(horizontal='right'))
        self.input_s = NamedStyle(name="input_s", font=Font(name='Arial', size=8),
                                  border=Border(left=bd, top=bd, right=bd, bottom=bd),
                                  alignment=Alignment(horizontal='right'))
        self.calcs = NamedStyle(name="calcs", font=Font(name='Arial', size=8),
                                border=Border(left=bd, top=bd, right=bd, bottom=bd),
                                fill=PatternFill(fgColor='e5e4e2', fill_type='solid'),
                                alignment=Alignment(horizontal='right'))
        self.instruct_text = NamedStyle(name="instruct_text", font=Font(name='Arial', size=9),
                                        alignment=Alignment(horizontal='left', vertical='top'))

    def build_workbook(self):
        """ build the workbook object """
        day_finder = ["Sat", "Sun", "Mon", "Tue", "Wed", "Thu", "Fri"]
        day_of_week = ["saturday", "sunday", "monday", "tuesday", "wednesday", "thursday", "friday"]
        i = 0
        self.wb = Workbook()  # define the workbook
        if not projvar.invran_weekly_span:  # if investigation range is daily
            for ii in range(len(day_finder)):
                if projvar.invran_date.strftime("%a") == day_finder[ii]:  # find the correct day
                    i = ii
            self.ws_list.append(self.wb.active)  # create first worksheet
            self.ws_list[0].title = day_of_week[i]  # title first worksheet
            self.day_of_week.append(day_of_week[i])  # create self.day_of_week array with one day
        if projvar.invran_weekly_span:  # if investigation range is weekly
            for day in day_of_week:
                self.day_of_week.append(day)  # create self.day_of_week array with seven days
            self.ws_list.append(self.wb.active)  # create first worksheet
            self.ws_list[0].title = "saturday"  # title first worksheet
            for i in range(1, 7):  # create worksheet for remaining six days
                self.ws_list.append(self.wb.create_sheet(day_of_week[i]))  # create subsequent worksheets
                self.ws_list[i].title = day_of_week[i]  # title subsequent worksheets
        self.remedy_sheet = self.wb.create_sheet("remedy")
        self.award_sheet = self.wb.create_sheet("awards")

    def build_text_doc(self):
        """ build the text document for the list of names and totals.  """
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.report_filename = "impropermandating5" + "_" + stamp + ".txt"
        self.report = open(dir_path('report') + self.report_filename, "w")
        self.report.write("\nImproper Mandating Report\n")

    def set_dimensions(self):
        """ set the orientation and dimensions of the workbook """
        for i in range(len(self.dates)):
            self.ws_list[i].oddFooter.center.text = "&A"  # include the footer
            self.ws_list[i].column_dimensions["A"].width = 16  # column width
            self.ws_list[i].column_dimensions["B"].width = 6
            self.ws_list[i].column_dimensions["C"].width = 20
            self.ws_list[i].column_dimensions["D"].width = 20
            self.ws_list[i].column_dimensions["E"].width = 20
        self.remedy_sheet.oddFooter.center.text = "&A"  # include the footer
        self.remedy_sheet.column_dimensions["A"].width = 15  # set dimensions for remedy worksheet
        column_tuple = ("B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P")
        for col in column_tuple:
            self.remedy_sheet.column_dimensions[col].width = 4
        self.remedy_sheet.column_dimensions["Q"].width = 6  # column width
        self.remedy_sheet.column_dimensions["R"].width = 6  # column width
        self.award_sheet.column_dimensions["A"].width = 16  # column width awards
        self.award_sheet.column_dimensions["B"].width = 6  # column width

    def build_ws_loop(self):
        """ this loops once for each day. """
        self.i = 0
        for day in self.dates:
            self.build_ws_headers()
            self.day = day
            self.list_loop()  # loops four times. once for each list.
            self.i += 1
            self.row = 1

    def build_ws_headers(self):
        """ worksheet headers """
        cell = self.ws_list[self.i].cell(row=self.row, column=1)
        cell.value = "Improper Mandate Worksheet"
        cell.style = self.ws_header
        # self.ws_list[self.i].merge_cells('A1:E1')
        self.row += 2
        cell = self.ws_list[self.i].cell(row=self.row, column=1)
        cell.value = "Date:  "  # create date/ pay period/ station header
        cell.style = self.date_dov_title
        cell = self.ws_list[self.i].cell(row=self.row, column=2)
        cell.value = format(self.dates[self.i], "%A  %m/%d/%y")
        cell.style = self.date_dov
        self.ws_list[self.i].merge_cells('B3:C3')
        cell = self.ws_list[self.i].cell(row=self.row, column=4)
        cell.value = "Pay Period:  "
        cell.style = self.date_dov_title
        # self.ws_list[self.i].merge_cells('E3:F3')
        cell = self.ws_list[self.i].cell(row=self.row, column=5)
        cell.value = projvar.pay_period
        cell.style = self.date_dov
        # self.ws_list[self.i].merge_cells('G3:H3')
        self.row += 1
        cell = self.ws_list[self.i].cell(row=self.row, column=1)
        cell.value = "Station:  "
        cell.style = self.date_dov_title
        cell = self.ws_list[self.i].cell(row=self.row, column=2)
        cell.value = projvar.invran_station
        cell.style = self.date_dov
        self.row += 2
        self.ws_list[self.i].merge_cells('B4:C4')

    def build_remedy_ws_headers(self):
        """ remedy worksheet header """
        cell = self.remedy_sheet.cell(row=self.row, column=1)
        cell.value = "Improper Mandate Worksheet"
        cell.style = self.ws_header
        self.remedy_sheet.merge_cells('A1:Q1')
        self.row += 2
        cell = self.remedy_sheet.cell(row=self.row, column=1)  # date label
        cell.value = "Date:  "
        cell.style = self.date_dov_title
        cell = self.remedy_sheet.cell(row=self.row, column=2)
        cell.value = self.dates[0].strftime("%x")  # date
        if projvar.invran_weekly_span:
            cell.value = self.dates[0].strftime("%x") + " - " + self.dates[6].strftime("%x")
        cell.style = self.date_dov
        self.remedy_sheet.merge_cells('B3:E3')
        cell = self.remedy_sheet.cell(row=self.row, column=9)  # pay period label
        cell.value = "Pay Period:  "
        cell.style = self.date_dov_title
        self.remedy_sheet.merge_cells('I3:K3')
        cell = self.remedy_sheet.cell(row=self.row, column=12)  # pay period
        cell.value = projvar.pay_period
        cell.style = self.date_dov
        self.remedy_sheet.merge_cells('L3:N3')
        self.row += 1
        cell = self.remedy_sheet.cell(row=self.row, column=1)  # station label
        cell.value = "Station:  "
        cell.style = self.date_dov_title
        cell = self.remedy_sheet.cell(row=self.row, column=2)  # station
        cell.value = projvar.invran_station
        cell.style = self.date_dov
        self.remedy_sheet.merge_cells('B4:E4')
        self.row += 1

    def list_loop(self):
        """ loops two times. once for each group. """
        self.lsi = 0  # iterations of the list loop method - 1st nl and wal - 2nd otdl and aux
        for _ in range(2):  # loops for nl, wal, otdl and aux
            self.list_and_column_headers()  # builds headers for list and columns
            self.carrierlist_mod()
            self.carrierloop()  # loop once to fill a row with carrier rings data
            self.update_report()  # write contentions for the day to the report document
            self.increment_progbar()
            self.lsi += 1
        self.lsi = 0  # reset list loop iteration

    def list_and_column_headers(self):
        """ builds headers for list and column """
        c_3headers = ("Own route/ string", "OT worked")
        d_4headers = ("Mandated route", "Available at OT rate")
        e_5headers = ("Mandated OT", "Available at penalty rate")
        self.row += 1
        cell = self.ws_list[self.i].cell(row=self.row, column=1)
        cell.value = self.page_titles[self.lsi]  # Displays the table title for each list,
        cell.style = self.list_header
        self.ws_list[self.i].merge_cells('A' + str(self.row) + ':D' + str(self.row))
        if self.lsi == 0:
            cell = self.ws_list[self.i].cell(row=self.row, column=5)
            cell.value = self.dates[self.i].strftime("%A  %m/%d/%y")
            cell.style = self.col_header_right
        self.row += 1
        cell = self.ws_list[self.i].cell(row=self.row, column=1)  # Name Header
        cell.value = "Name"
        cell.style = self.col_header_left
        self.ws_list[self.i].merge_cells('A' + str(self.row) + ':B' + str(self.row))
        cell = self.ws_list[self.i].cell(row=self.row, column=3)  # own route or ot worked
        cell.value = c_3headers[self.lsi]
        cell.style = self.col_header_left
        cell = self.ws_list[self.i].cell(row=self.row, column=4)  # mandated route or available at OT rate
        cell.value = d_4headers[self.lsi]
        cell.style = self.col_header_left
        cell = self.ws_list[self.i].cell(row=self.row, column=5)  # mandated OT or available at penalty rate
        cell.value = e_5headers[self.lsi]
        cell.style = self.col_header_left
        self.row += 1

    def carrierlist_mod(self):
        """ get the carrier list appropriate to the day and list status """
        self.mod_carrierlist = self.carrier_breakdown[self.i][self.lsi]

    def carrierloop(self):
        """ loop for each carrier """
        for carrier in self.mod_carrierlist:
            self.carrier = carrier[1]  # current iteration of carrier list is assigned self.carrier
            self.list_ = carrier[2]  # get the list status of the carrier
            self.nsday = carrier[3]
            self.route = carrier[4]  # get the route of the carrier
            if self.list_ in ("otdl", "ptf", "aux"):
                self.ot_available = True
            elif self.list_ == "odlr":  # for odl regular day only -
                if self.dates[self.i].strftime("%a") == projvar.ns_code[self.nsday] \
                        or self.dates[self.i].strftime("%a") == "Sun":  # if it is sunday or their ns day
                    self.ot_available = False
                else:  # if it is a sunday or their ns day, put record in no list array.
                    self.ot_available = True
            elif self.list_ == "odln":  # for odl nonscheduled day only -
                if self.dates[self.i].strftime("%a") == projvar.ns_code[self.nsday] \
                        or self.dates[self.i].strftime("%a") == "Sun":  # if it is sunday or their ns day
                    self.ot_available = True
                else:  # if it is a sunday or their ns day, put record in no list array.
                    self.ot_available = False
            else:
                self.ot_available = False
            self.build_availability_dict()  # build three dictionaries related to availability
            self.get_rings()  # get individual carrier rings for the day
            self.number_crunching()  # do calculations to get overtime and availability
            if self.qualify():  # test the rings to see if they need to be displayed
                self.display_recs()  # build the carrier and the rings row into the spreadsheet
                self.increment_names_array()  # build a list of mandated/ available carriers
                self.increment_mentioned_names()  # build a list of all carriers mentioned
                self.increment_totals_array()  # build a list of mandated/available totals
                self.append_remedy_array()  # add the possible violation to the remedy array
                self.row += 1

    def update_report(self):
        """ write contentions for the day to the report document.
        includes list of mandated carrriers, mandate hour totals, list of available carriers
        and available hour totals."""
        if self.lsi == 1:
            self.report.write("\n{}\n".format(self.dates[self.i].strftime("%x %A")))
            contention1 = "On {}, the following letter carriers are not on any overtime desired list for the " \
                          "[quarter] of [year] at the [station]: {}. This is documented by the absence of these " \
                          "letter carrier’s names on the current overtime desired list  sign-up sheet in the " \
                          "case file."\
                .format(self.dates[self.i].strftime("%x %A"),
                        Convert(self.mandate_names[self.i]).array_to_string_withand())
            contention2 = "The following letter carriers are on the 10/12-hour overtime desired list for the " \
                          "[quarter] of [year] at the [station]: {}. This is documented by the Overtime Desired " \
                          "List sign-up sheet included in the case file."\
                .format(Convert(self.otdl_names[self.i]).array_to_string_withand())
            contention3 = "The following letter carriers are ptf or cca carriers at the [station] during the week of " \
                          "[investigation range start]-[investigation range end]: {}. This is documented by the " \
                          "employee everything report included in the case file. Part Time Flexible (ptf) are " \
                          "designated the D/A code 43-4 and City Carrier Assistant (cca or aux) are designated by " \
                          "the D/A code 84-4."\
                .format(Convert(self.aux_names[self.i]).array_to_string_withand())
            contention5_1 = "The tables below show the distribution of overtime worked/denied for {}. " \
                            "The first table below shows overtime hours worked by non-otdl and work assignment list " \
                            "carriers, their regular route, and the route on which the overtime was worked. " \
                            "The second table shows the otdl and auxiliary carriers, the number of overtime hours " \
                            "worked, the number of hours they were available at the regular overtime rate, and the " \
                            "number of hours they were available at the penalty overtime rate. All data included in " \
                            "the tables is documented by the TACS Employee Everything reports included in the case " \
                            "file. "\
                .format(self.dates[self.i].strftime("%x %A"))
            # contention 5_2 has four components - otdl availability, aux availablity, mandated carriers and remedy.
            otdl_availability = "there was no otdl availability"
            if self.otdl_totals[self.i]:
                otdl_availability = "otdl: {} were available for {:.2f} combined hours at the overtime and " \
                                    "penalty overtime rate"\
                    .format(Convert(self.otdl_names[self.i]).array_to_string_withand(),
                            self.otdl_totals[self.i],)
            aux_availability = "There was no auxiliary availability"
            if self.aux_totals[self.i]:
                aux_availability = "Auxiliary: {} were available for {:.2f} combined hours at the overtime and " \
                                   "penalty overtime rate"\
                    .format(Convert(self.aux_names[self.i]).array_to_string_withand(),
                            self.aux_totals[self.i],)
            mandates = "there were no mandated carriers"
            if self.mandate_totals[self.i]:
                mandates = "non otdl and work assignment: {} were mandated for {:.2f} hours"\
                    .format(Convert(self.mandate_names[self.i]).array_to_string_withand(),
                            self.mandate_totals[self.i])
            remedy = "There was no violation"
            remedy_total = min(self.otdl_totals[self.i], self.mandate_totals[self.i])
            all_ = ""
            if self.otdl_totals[self.i] >= self.mandate_totals[self.i]:
                all_ = "all "
            if remedy_total:
                remedy = " Otdl and/or auxiliary letter carrier(s) should have been assigned {}{:.2f} " \
                         "hours of the overtime worked by non otdl and work assignment list carriers"\
                    .format(all_, remedy_total)
            contention5_2 = "As the table above illustrates, on {} {}. {}. " \
                            "On the same date, {}. {}."\
                .format(self.dates[self.i].strftime("%a %x"), otdl_availability, aux_availability, mandates, remedy)
            if self.impman5_report:
                if self.impman5_fullreport:
                    self.report.write("\n{}\n".format(contention1))
                    self.report.write("\n{}\n".format(contention2))
                    self.report.write("\n{}\n".format(contention3))
                self.report.write("\n{}\n".format(contention5_1))
                self.report.write("\n{}\n".format(contention5_2))

    def get_rings(self):
        """ get individual carrier rings for the day """
        self.rings = Rings(self.carrier, self.dates[self.i]).get_for_day()  # assign as self.rings
        self.totalhours = 0.0  # set default as an empty string
        self.bt = ""
        self.rs = ""
        self.et = ""
        self.codes = ""
        self.moves = ""
        self.lvtype = ""
        self.lvtime = ""
        if self.rings[0]:  # if rings record is not blank
            self.totalhours = float(Convert(self.rings[0][2]).zero_not_empty())
            self.bt = self.rings[0][9]
            self.rs = self.rings[0][3]
            self.et = self.rings[0][10]
            self.codes = self.rings[0][4]
            if self.day_of_week[self.i] == "Sunday":
                if self.list_ in ("otdl", "odln") and not self.totalhours:
                    self.codes = "no call"
            self.moves = self.rings[0][5]
            self.lvtype = self.rings[0][6]
            self.lvtime = self.rings[0][7]

    def build_availability_dict(self):
        """ add the carrier's name to the availability dictionaries on the first loop of days """
        if self.i == 0:
            self.cum_hr_dict[self.carrier] = 0.0
            self.cum_ot_dict[self.carrier] = 0.0
            self.avail_ot_dict[self.carrier] = 20.0

    def calc_max_availability(self):
        """ get the maximum availability for the day for the given carrier
        this takes into account: weekly hours to 60, weekly ot hours to 20, daily limit to 12 or 11.50, leave,
        ns day """
        totalhours = Convert(self.totalhours).str_to_float()
        lv_time = Convert(self.lvtime).str_to_float()
        # cumulative hours for the week
        cum_hr = (lv_time + totalhours) + float(self.cum_hr_dict[self.carrier])  # cumulative ot hours for the week
        cum_ot = max(totalhours - 8, 0) + float(self.cum_ot_dict[self.carrier])
        if self.codes in ("no call", "ns day"):  # if it is the carrier's ns day
            cum_ot = totalhours + float(self.cum_ot_dict[self.carrier])
        if self.codes == "ns day":  # if ns day, then full day is added to cumulative ot.
            cum_ot = totalhours + float(self.cum_ot_dict[self.carrier])
        avail_wkly = max(60 - cum_hr, 0)  # the weekly availability is 60 - weekly cumulative
        avail_ot = max(20 - cum_ot, 0)  # the weekly ot availability is 20 - weekly ot cumulative
        avail_daily = max(11.50 - totalhours, 0)  # daily availability is 11.50 minus daily work hours
        if self.list_ == "otdl":  # except if the carrier is on the otdl
            avail_daily = max(12 - totalhours, 0)  # then daily availability is 12 minus daily work hours
        if self.list_ == "odln" and self.ot_available:
            avail_daily = max(8 - totalhours, 0)  # then daily availability is 8 minus daily work hours
        avail_leave = 12  # availability is zeroed out if the carrier takes leave
        if self.lvtype not in ("", "none"):  # zero out if lvtype is empty or 'none'
            avail_leave = 0
        prior_avail_ot = 20  # this is the available ot from the prior day, default is 20
        if self.i != 0:  # if this is not the first day
            prior_avail_ot = self.avail_ot_dict[self.carrier]  # get the value from the dictionary
        avail_ns = avail_ot  # this code will zero out availability if the carrier can not work 8 hours on an ns day.
        # if it is the ns day and 8 hours are not available
        if self.codes in ("ns day", "no call") and prior_avail_ot < 8:
            avail_ns = 0  # zero out availability
        avail_codes = avail_ot
        if self.codes in ("light", "excused", "sch chg", "annual", "sick"):  # if carrier excused for day
            avail_codes = 0  # if any of the listed codes are in self.codes - zero availability
        # select the lowest value from all criteria.
        self.avail_max = min(avail_wkly, avail_ot, avail_daily, avail_leave, avail_ns, avail_codes)
        self.update_availability_dict(cum_hr, cum_ot, avail_ot)

    def update_availability_dict(self, cum_hr, cum_ot, avail_ot):
        """ update the 3 availability dictionaries used to find max availablity , takes 3 arguments """
        self.cum_hr_dict.update({self.carrier: cum_hr})
        self.cum_ot_dict.update({self.carrier: cum_ot})
        self.avail_ot_dict.update({self.carrier: avail_ot})

    def number_crunching(self):
        """ crunch numbers to get overtime, off route, other route and availability"""
        def mandated_crunching():
            """ get possible mandating violations for no list, wal, odlr and odln """
            if self.moves:
                self.calc_offroute()  # calculate the time that the carrier spent off their route and get other route
                self.format_otherroute()  # format the self.other route so that if fits in the spreadsheet cell
            self.calc_offroute_adj()  # adj for pivot time or if code is nsday or whole day spent off route

        def avaiability_crunching():
            """ get possible availability violations for no list, wal, odlr and odln.  calculate the
            availability at the ot rate and the penalty rate
            pen_max = maximum penalty rate for the day
            ot_max = maximum overtime rate for the day
            pen_srt = penalty start: after how many hours does the penalty rate start to apply. """
            self.penalty_rate = 0.0  # initialize
            self.overtime_rate = 0.0
            pen_max, ot_max, pen_srt = 2.0, 2.0, 10.0
            if self.codes in ("ns day", "no call"):
                pen_max, ot_max, pen_srt = 4.0, 8.0, 8.0
            if self.list_ in ("aux", "ptf"):
                pen_max, ot_max, pen_srt = 1.5, 2.0, 10.0
            if self.list_ in ("odln", ):  # for odl ns day only carriers working on their ns day.
                pen_max, ot_max, pen_srt = 0.0, 8.0, 8.0
            wk = self.totalhours
            a_max = self.avail_max  # available maximum
            a_ceil = wk + a_max  # available ceiling
            pen_adj = max(a_ceil - pen_srt, 0)
            pen_rate = min([a_max, pen_max, pen_adj])  # penalty overtime rate maximum
            ot_rate = min(a_max - pen_rate, ot_max)
            self.penalty_rate = pen_rate  # assign remedy rates
            self.overtime_rate = ot_rate

        self.overtime = 0.0  # the total overtime worked
        self.onroute = 0.0  # the amount of overtime worked on the carrier's own route.
        self.offroute = 0.0  # total time spend off route
        self.offroute_adj = 0.0
        self.otherroute_array = []  # a list of routes where carrier worked off assignment
        self.otherroute = ""  # display routes worked off assignment
        self.calc_max_availability()  # get maximum availability and store in self.avail_max
        self.calc_overtime()  # calculate the amount of overtime worked
        if self.ot_available:
            avaiability_crunching()
        else:
            mandated_crunching()

    def calc_overtime(self):
        """ calculates the amount of overtime worked. if it is the carrier's ns day, then the full day is overtime. """
        if self.codes == "ns day":
            self.overtime = self.totalhours
        else:
            self.overtime = max(self.totalhours - 8, 0)

    def calc_offroute(self):
        """ calculate the time that the carrier spent off their route assignment, get other route """
        moves = self.moves.split(",")
        move_sets = int(len(moves)/3)  # get the number of triads in the moves array
        count = 0
        for _ in range(move_sets):
            offroute = float(moves[count+1]) - float(moves[count])  # calculate off route time per triad
            self.offroute += offroute  # add triad time off route
            self.otherroute_array.append(moves[count+2])
            count += 3
        self.offroute = round(self.offroute, 2)
        if self.offroute >= self.totalhours:  # if the carrier took lunch, off route could be greater than total hours
            self.offroute = self.totalhours

    def format_otherroute(self):
        """ format the self.other route. format like '1024, 1008, 0935' . do not allow duplicates"""
        unique_routes = []
        for route in self.otherroute_array:
            if route not in unique_routes:
                unique_routes.append(route)
        self.otherroute = Convert(unique_routes).array_to_string()

    def calc_offroute_adj(self):
        """ calculate the off route overtime for ns days or if the whole day is spent off own route. """
        self.offroute_adj = min(self.overtime, self.offroute)  # will adjust for pivot time
        if self.codes == "ns day":  # if it is the ns day, then whole day is off route
            self.offroute_adj = self.totalhours
            self.otherroute_array.append("ns day")
            self.otherroute = "ns day"
            self.moves = self.bt
        if self.totalhours:  # detect off bid violations use max pivot from off bid spreadsheet settings.
            ownroute = max(self.totalhours - self.offroute, 0)  # calculate the total time spent on route
            violation = max(8 - ownroute, 0)  # calculate the total violation
            if violation > self.max_pivot and self.codes != "ns day":
                self.otherroute = "off bid"

    def calc_onroute(self):
        """ calculate the overtime the carrier worked on their own route. """
        if self.codes == "ns day":
            self.onroute = 0
        else:
            self.onroute = max(self.overtime - self.offroute, 0)

    def qualify(self):
        """ check to see if the carrier information needs to be displayed. """
        if self.otherroute == "off bid" and not self.ot_available:  # if violation is concurrent with off bid violation
            return False
        if self.list_ in ("aux", "ptf"):  # do not count aux carriers who miss days
            if self.totalhours == 0.0:
                return False
        if self.ot_available and not self.rings[0]:  # if the otdl carrier has no record for the day - return false
            return False
        if not self.ot_available:  # if there is any overtime worked off route
            if self.offroute_adj >= self.remedy_tolerance:
                return True
        if self.ot_available:  # implement tolerances
            if self.penalty_rate + self.overtime_rate >= self.remedy_tolerance:
                return True
        return False

    def display_recs(self):
        """ put the carrier and the first part of rings into the spreadsheet - it's show time! """
        cell = self.ws_list[self.i].cell(row=self.row, column=1)  # name
        cell.value = self.carrier
        cell.style = self.input_name
        cell = self.ws_list[self.i].cell(row=self.row, column=2)  # list status
        cell.value = self.list_dict[self.list_]
        cell.style = self.input_list
        cell.number_format = "#,###.00;[RED]-#,###.00"
        cell = self.ws_list[self.i].cell(row=self.row, column=3)  # own route or overtime worked
        cell.value = self.route  # default, the carrier is no list or wal
        if self.lsi == 1:  # if the carrier is an otdl or aux carrier
            cell.value = self.overtime
        cell.style = self.input_s
        cell.number_format = "#,###.00;[RED]-#,###.00"
        cell = self.ws_list[self.i].cell(row=self.row, column=4)  # mandated route or available at ot rate
        cell.value = self.otherroute  # default, the carrier is no list or wal
        if self.ot_available:  # if the carrier is an otdl
            cell.value = self.overtime_rate
        cell.style = self.input_s
        cell.number_format = "#,###.00;[RED]-#,###.00"
        cell = self.ws_list[self.i].cell(row=self.row, column=5)  # mandated route or available at penalty rate
        cell.value = self.offroute_adj  # default, the carrier is no list or wal
        if self.ot_available:  # if the carrier is an auxiliary carrier
            cell.value = self.penalty_rate
        cell.style = self.input_s
        cell.number_format = "#,###.00;[RED]-#,###.00"

    def increment_names_array(self):
        """ build a list of mandated/ available carriers """
        if not self.ot_available:
            self.mandate_names[self.i].append(self.carrier)
        else:
            self.available_names[self.i].append(self.carrier)
            if self.list_ in ("otdl", "odlr", "odln"):  # if on otdl or odl+ and available for ot
                self.otdl_names[self.i].append(self.carrier)
            else:
                self.aux_names[self.i].append(self.carrier)

    def append_remedy_array(self):
        """ builds a multidimensional array e.g. [["weeks, t", "mandate" 2, .88],]
        elements are 1. name, 2. list, 3. day, 4. possible violation - will be ot and penalty for availability """
        list_ = "mandated"  # default values for non list and work assignment
        poss_violation_1 = self.offroute_adj
        poss_violation_2 = 0.0
        if self.ot_available:
            list_ = "available"
            poss_violation_1 = self.overtime_rate
            poss_violation_2 = self.penalty_rate
        to_add = [self.carrier, list_, self.i, poss_violation_1, poss_violation_2]
        self.remedy_array.append(to_add)

    def increment_mentioned_names(self):
        """ builds a list of all names mentioned in the investigation """
        if self.i != 1:  # do not include sundays
            if self.carrier not in self.mentioned_names:
                self.mentioned_names.append(self.carrier)

    def increment_totals_array(self):
        """ build a list of mandated/ available totals  """
        if self.list_ in ("nl", "wal"):
            self.mandate_totals[self.i] += self.offroute_adj
        elif self.list_ in ("odlr", "odln") and not self.ot_available:
            self.mandate_totals[self.i] += self.offroute_adj
        elif self.list_ in ("otdl", ):
            avail_total = self.overtime_rate + self.penalty_rate
            self.available_totals[self.i] += avail_total
            self.otdl_totals[self.i] += avail_total
        elif self.list_ in ("odlr", "odln") and self.ot_available:
            avail_total = self.overtime_rate + self.penalty_rate
            self.available_totals[self.i] += avail_total
            self.otdl_totals[self.i] += avail_total
        else:
            avail_total = self.overtime_rate + self.penalty_rate
            self.available_totals[self.i] += avail_total
            self.aux_totals[self.i] += avail_total

    def build_remedy(self):
        """ build the worksheet for finding a remedy. """
        self.row = 1  # initialize the row of the remedy worksheet
        self.lsi = 2
        self.increment_progbar()
        self.build_remedy_ws_headers()
        self.remedy_headers("mandated")
        all_remedy_arrays = self.sort_remedy_list()  # first sort the remedy array into two list
        remedy_category = ("mandated", "available")
        for i in range(len(all_remedy_arrays)):  # cycle through sorted remedy array
            if i == 1:
                self.remedy_headers("available")  # once available remedies are run - create headers
                self.rem_avail_row_start = self.row  # hold this value for the equalization totals
            for name in all_remedy_arrays[i]:  # start with mandated, then available
                remedy_array = []  # capture all remedies for each carrier
                for array in self.remedy_array:
                    if name == array[0] and array[1] == remedy_category[i]:  # if the correct name and list - display
                        remedy_array.append(array)
                self.display_remedy(remedy_array)
        self.remedy_equalization()

    def sort_remedy_list(self):
        """ first sort the remedy array into two list """
        all_mandated = []
        all_available = []
        for d in range(len(self.dates)):
            for name in self.mandate_names[d]:
                if name not in all_mandated:
                    all_mandated.append(name)
            for name in self.available_names[d]:
                if name not in all_available:
                    all_available.append(name)
        all_mandated.sort()
        all_available.sort()
        all_remedy_arrays = (all_mandated, all_available)
        return all_remedy_arrays

    def display_remedy(self, remedy_array):
        """ create the cells on the remedy worksheet"""
        if remedy_array[0][1] == "mandated":
            self.display_mandated_remedy(remedy_array)
            to_add = (remedy_array[0][0], remedy_array[0][1], self.row - 1)
            self.award_array.append(to_add)
        if remedy_array[0][1] == "available":
            self.display_available_remedy(remedy_array)
            to_add = (remedy_array[0][0], remedy_array[0][1], self.row - 2)
            self.award_array.append(to_add)

    def remedy_headers(self, list_):
        """ put headers on the top of the mandated remedies """
        cell = self.remedy_sheet.cell(row=self.row, column=1)  # white space
        cell.value = ""
        self.remedy_sheet.merge_cells('A' + str(self.row) + ':Q' + str(self.row))
        self.row += 1
        cell = self.remedy_sheet.cell(row=self.row, column=1)  # possible violation
        cell.value = "NON-OTDL and Work Assignment Employees that worked overtime"
        if list_ == "available":
            cell.value = "OTDL/Auxiliary Employees who were available to work overtime"
        cell.style = self.list_header
        self.remedy_sheet.merge_cells('A' + str(self.row) + ':Q' + str(self.row))
        self.row += 1
        headers = ("Name", "sat", "sun", "mon", "tue", "wed", "thu", "fri", "total", "")
        if list_ == "available":
            headers = ("Name", "sat", "sun", "mon", "tue", "wed", "thu", "fri", "ot rate", "penalty")
        column_array = (1, 3, 5, 7, 9, 11, 13, 15, 17, 18)
        merge1_tuple = ("A", "C", "E", "G", "I", "K", "M", "O", "Q")
        merge2_tuple = ("B", "D", "F", "H", "J", "L", "N", "P", "R")
        for i in range(10):
            cell = self.remedy_sheet.cell(row=self.row, column=column_array[i])  # possible violation
            cell.value = headers[i]
            cell.style = self.col_header_left
            if i < 8:
                self.remedy_sheet.merge_cells(merge1_tuple[i] + str(self.row) + ':' + merge2_tuple[i] + str(self.row))
        self.row += 1

    def display_mandated_remedy(self, remedy_array):
        """ display remedy cells for mandated carriers """
        def get_mandated_violation(ii):
            """ find the appropriate violation for the appropriate day """
            for array in remedy_array:  # cycle through all arrays in remedy array
                if ii == array[2]:
                    return array[3]
            return ""

        def modify_name(name):
            """ add odl indicator to name """
            if name in self.odln_indicator:
                return name + " (odln)"
            if name in self.odlr_indicator:
                return name + " (odlr)"
            return name

        column_array = (3, 5, 7, 9, 11, 13, 15)
        merge1_tuple = ("C", "E", "G", "I", "K", "M", "O")
        merge2_tuple = ("D", "F", "H", "J", "L", "N", "P")
        if not projvar.invran_weekly_span:
            column_array = (column_array[self.date_index], )
            merge1_tuple = (merge1_tuple[self.date_index], )
            merge2_tuple = (merge2_tuple[self.date_index], )
        cell = self.remedy_sheet.cell(row=self.row, column=1)  # name
        cell.value = modify_name(remedy_array[0][0])
        cell.style = self.input_name
        self.remedy_sheet.merge_cells('A' + str(self.row) + ':B' + str(self.row))
        for i in range(self.inv_range):
            cell = self.remedy_sheet.cell(row=self.row, column=column_array[i])  # possible violation
            cell.value = get_mandated_violation(i)
            cell.style = self.input_s
            cell.number_format = "#,###.00;[RED]-#,###.00"
            self.remedy_sheet.merge_cells(merge1_tuple[i] + str(self.row) + ':' + merge2_tuple[i] + str(self.row))
        cell = self.remedy_sheet.cell(row=self.row, column=17)  # total
        formula = "=SUM(%s!C%s:O%s)" % ("remedy", self.row, self.row)
        cell.value = formula
        cell.style = self.calcs
        cell.number_format = "#,###.00;[RED]-#,###.00"
        self.row += 1

    def display_available_remedy(self, remedy_array):
        """ display remedy cells for mandated carriers """
        def get_aux_status(name):
            """ determine if the carrier is an auxiliary carrier by checking self.aux_names multi array """
            for ii in range(self.inv_range):
                if name in self.aux_names[ii]:
                    return True
            return False

        def modify_name(name):
            """ add '(aux)' to the name of any ptf or aux carriers """
            if aux_status:
                return name + " (aux)"
            if name in self.odln_indicator:
                return name + " (odln)"
            if name in self.odlr_indicator:
                return name + " (odlr)"
            return name

        def get_label(label):
            """ returns 'ot', 'pen' or 'void' if the carrier is an aux carrier """
            if aux_status:
                return "void"
            return label

        def get_otrate_violation(ii):
            """ find the appropriate overtime rate violation for the appropriate day """
            for array in remedy_array:  # cycle through all arrays in remedy array
                if ii == array[2]:
                    return array[3]
            return ""

        def get_penrate_violation(ii):
            """ find the appropriate penalty rate violation for the appropriate day """
            for array in remedy_array:  # cycle through all arrays in remedy array
                if ii == array[2]:
                    return array[4]
            return ""

        aux_status = get_aux_status(remedy_array[0][0])
        rates_column_array = (3, 5, 7, 9, 11, 13, 15)
        together_column_array = (4, 6, 8, 10, 12, 14, 16)
        rates_tuple = ("C", "E", "G", "I", "K", "M", "O")
        together_tuple = ("D", "F", "H", "J", "L", "N", "P")
        if not projvar.invran_weekly_span:
            rates_column_array = (rates_column_array[self.date_index], )
            together_column_array = (together_column_array[self.date_index], )
            rates_tuple = (rates_tuple[self.date_index], )
            together_tuple = (together_tuple[self.date_index], )
        cell = self.remedy_sheet.cell(row=self.row, column=1)  # name
        cell.value = modify_name(remedy_array[0][0])
        cell.style = self.input_name
        self.remedy_sheet.merge_cells('A' + str(self.row) + ':A' + str(self.row + 1))
        cell = self.remedy_sheet.cell(row=self.row, column=2)  # ot label
        cell.value = get_label("ot")
        cell.style = self.input_s
        cell = self.remedy_sheet.cell(row=self.row + 1, column=2)  # pen label
        cell.value = get_label("pen")
        cell.style = self.calcs
        for i in range(self.inv_range):
            cell = self.remedy_sheet.cell(row=self.row, column=rates_column_array[i])  # ot rate availability
            otrate_violation = get_otrate_violation(i)
            formula = "=IF(%s!%s%s=\"void\",0,%s)" % ("remedy", "B", self.row, otrate_violation)
            cell.value = formula
            cell.style = self.input_s
            cell.number_format = "#,###.00;[RED]-#,###.00"
            cell = self.remedy_sheet.cell(row=self.row + 1, column=rates_column_array[i])  # pen rate availability
            penrate_violation = get_penrate_violation(i)
            formula = "=IF(%s!%s%s=\"void\",0,%s)" % ("remedy", "B", self.row + 1, penrate_violation)
            cell.value = formula
            cell.style = self.calcs
            cell.number_format = "#,###.00;[RED]-#,###.00"
            cell = self.remedy_sheet.cell(row=self.row, column=together_column_array[i])  # sum availability
            formula = "=IF(SUM(%s!%s%s:%s%s)>0,SUM(%s!%s%s:%s%s),\"\"" \
                      % ("remedy", rates_tuple[i], self.row, rates_tuple[i], self.row + 1,
                         "remedy", rates_tuple[i], self.row, rates_tuple[i], self.row + 1)
            cell.value = formula
            cell.style = self.input_s
            cell.number_format = "#,###.00;[RED]-#,###.00"
            self.remedy_sheet.merge_cells(together_tuple[i] + str(self.row) + ':' +
                                          together_tuple[i] + str(self.row + 1))
        cell = self.remedy_sheet.cell(row=self.row, column=17)  # sum ot rate
        formula = "=IF(SUM(%s!C%s+%s!E%s+%s!G%s+%s!I%s+%s!K%s+%s!M%s+%s!O%s)>0," \
                  "SUM(%s!C%s+%s!E%s+%s!G%s+%s!I%s+%s!K%s+%s!M%s+%s!O%s),\"\"" \
                  % ("remedy", self.row, "remedy", self.row, "remedy", self.row, "remedy", self.row,
                     "remedy", self.row, "remedy", self.row, "remedy", self.row, "remedy", self.row,
                     "remedy", self.row, "remedy", self.row, "remedy", self.row, "remedy", self.row,
                     "remedy", self.row, "remedy", self.row)
        cell.value = formula
        cell.style = self.input_s
        cell.number_format = "#,###.00;[RED]-#,###.00"
        self.remedy_sheet.merge_cells("Q" + str(self.row) + ':' + "Q" + str(self.row + 1))
        cell = self.remedy_sheet.cell(row=self.row, column=18)  # sum penalty rate
        formula = "=IF(SUM(%s!C%s+%s!E%s+%s!G%s+%s!I%s+%s!K%s+%s!M%s+%s!O%s)>0," \
                  "SUM(%s!C%s+%s!E%s+%s!G%s+%s!I%s+%s!K%s+%s!M%s+%s!O%s),\"\"" \
                  % ("remedy", self.row + 1, "remedy", self.row + 1, "remedy", self.row + 1, "remedy", self.row + 1,
                     "remedy", self.row + 1, "remedy", self.row + 1, "remedy", self.row + 1, "remedy", self.row + 1,
                     "remedy", self.row + 1, "remedy", self.row + 1, "remedy", self.row + 1, "remedy", self.row + 1,
                     "remedy", self.row + 1, "remedy", self.row + 1)
        cell.value = formula
        cell.style = self.calcs
        cell.number_format = "#,###.00;[RED]-#,###.00"
        self.remedy_sheet.merge_cells("R" + str(self.row) + ':' + "R" + str(self.row + 1))
        self.row += 2

    def remedy_equalization(self):
        """ create rows at the bottom of the sheet for total mandates, total availability, and equalization """
        a_tuple = ("C", "E", "G", "I", "K", "M", "O", "Q")
        b_tuple = ("D", "F", "H", "J", "L", "N", "P", "R")
        a_column = (3, 5, 7, 9, 11, 13, 15)
        self.row += 1
        cell = self.remedy_sheet.cell(row=self.row, column=1)  # list section header
        cell.value = "Equalization"
        cell.style = self.list_header
        self.row += 1
        # -------------------------------------------------------------------------------------------------- mandates
        cell = self.remedy_sheet.cell(row=self.row, column=1)  # title row for section footer
        cell.value = "Mandates:  "
        cell.style = self.date_dov_title
        for i in range(7):
            formula = "=SUM(%s!%s%s:%s%s)" % ("remedy", a_tuple[i], 8, a_tuple[i], self.rem_avail_row_start - 4)
            cell = self.remedy_sheet.cell(row=self.row, column=a_column[i])
            cell.value = formula
            cell.style = self.calcs
            cell.number_format = "#,###.00;[RED]-#,###.00"
            self.remedy_sheet.merge_cells(a_tuple[i] + str(self.row) + ':' + b_tuple[i] + str(self.row))
        self.row += 1
        # ----------------------------------------------------------------------------------------------- availability
        cell = self.remedy_sheet.cell(row=self.row, column=1)  # title row for section footer
        cell.value = "Availability:  "
        cell.style = self.date_dov_title
        for i in range(7):
            formula = "=SUM(%s!%s%s:%s%s)" % ("remedy", b_tuple[i], self.rem_avail_row_start,
                                              b_tuple[i], self.row - 2)
            cell = self.remedy_sheet.cell(row=self.row, column=a_column[i])
            cell.value = formula
            cell.style = self.calcs
            cell.number_format = "#,###.00;[RED]-#,###.00"
            self.remedy_sheet.merge_cells(a_tuple[i] + str(self.row) + ':' + b_tuple[i] + str(self.row))
        self.row += 2
        # ----------------------------------------------------------------------------------------------- equalization
        cell = self.remedy_sheet.cell(row=self.row, column=1)  # title row for section footer
        cell.value = "Equalization:  "
        cell.style = self.date_dov_title
        for i in range(7):
            formula = "=SUM(%s!%s%s-%s!%s%s)" % ("remedy", a_tuple[i], self.row - 3,
                                                 "remedy", a_tuple[i], self.row - 2)
            cell = self.remedy_sheet.cell(row=self.row, column=a_column[i])
            cell.value = formula
            cell.style = self.calcs
            cell.number_format = "#,###.00;[RED]-#,###.00"
            self.remedy_sheet.merge_cells(a_tuple[i] + str(self.row) + ':' + b_tuple[i] + str(self.row))
        cell = self.remedy_sheet.cell(row=self.row, column=17)  # remedy percentage input
        cell.value = " <- adjust to zero"
        cell.style = self.date_dov_title
        self.remedy_sheet.merge_cells('Q' + str(self.row) + ':R' + str(self.row))
        self.row += 1
        self.remedy_sheet.merge_cells('B' + str(self.row) + ':Q' + str(self.row))
        cell = self.remedy_sheet.cell(row=self.row, column=2)  # row for exposition on equalization
        cell.value = "\n" \
                     "1. Using the OTDL Weekly Availability Worksheet, alter/delete availability from the OTDL " \
                     "section if there is no availability. \n" \
                     "2. If value is positive, subtract/delete from No List and Work Assignment sections to " \
                     "equalize. \n" \
                     "3. If the value is negative, subtract/delete from OTDL and Auxiliary sections to equalize. \n"
        cell.style = self.instruct_text
        self.remedy_sheet['B' + str(self.row)].alignment = Alignment(wrap_text=True, vertical='top',
                                                                     shrink_to_fit=False)
        self.remedy_sheet.row_dimensions[self.row].height = 100

    def build_awards(self):
        """ build the awards worksheet """
        penalty_array = []
        start_availability = False
        self.row = 1
        cell = self.award_sheet.cell(row=self.row, column=1)  # section header
        cell.value = "Improperly Mandated Carrier Awards"
        cell.style = self.list_header
        self.award_sheet.merge_cells('A' + str(self.row) + ':D' + str(self.row))
        self.row += 2
        for award in self.award_array:
            if not start_availability and award[1] == "available":
                self.row += 1
                start_availability = True
                cell = self.award_sheet.cell(row=self.row, column=1)  # section header
                cell.value = "Overtime Rate Awards"
                cell.style = self.list_header
                self.award_sheet.merge_cells('A' + str(self.row) + ':D' + str(self.row))
                self.row += 2
            if start_availability:
                penalty_array.append(award)
            cell = self.award_sheet.cell(row=self.row, column=1)  # name
            cell.value = award[0]
            cell.style = self.input_name
            cell = self.award_sheet.cell(row=self.row, column=2)  # possible ot rate award
            formula = "=%s!%s%s" % ("remedy", "Q", str(award[2]))
            cell.value = formula
            cell.style = self.input_s
            cell.number_format = "#,###.00;[RED]-#,###.00"
            self.row += 1
        self.row += 1
        cell = self.award_sheet.cell(row=self.row, column=1)  # section header for penalty overtime
        cell.value = "Penalty Overtime Rate Awards"
        cell.style = self.list_header
        self.award_sheet.merge_cells('A' + str(self.row) + ':D' + str(self.row))
        self.row += 2
        for award in penalty_array:
            cell = self.award_sheet.cell(row=self.row, column=1)  # name
            cell.value = award[0]
            cell.style = self.input_name
            cell = self.award_sheet.cell(row=self.row, column=2)  # possible ot rate award
            formula = "=%s!%s%s" % ("remedy", "R", str(award[2]))
            cell.value = formula
            cell.style = self.input_s
            cell.number_format = "#,###.00;[RED]-#,###.00"
            self.row += 1

    def increment_progbar(self):
        """ move the progress bar, update with info on what is being done """
        lst = ("Finding Mandates", "Finding Availability", "Calculate Remedy")
        self.pbi += 1
        self.pb.move_count(self.pbi)  # increment progress bar
        self.pb.change_text("Building day {}: list: {}".format(self.day.strftime("%A"), lst[self.lsi]))

    def write_mentioned_names(self):
        """ before ending the class, write all the mentioned names to the report """
        if not self.mentioned_names:
            self.report.write("There are no names mentioned in this investigation.\n")
            return
        self.mentioned_names.sort()
        self.report.write("\n")
        self.report.write("Mentioned Names: \n")
        for name in self.mentioned_names:
            self.report.write("{}\n".format(name))

    def save_open(self):
        """ name and open the excel file """
        self.pbi += 1
        self.pb.move_count(self.pbi)  # increment progress bar
        self.pb.change_text("Saving...")
        self.pb.stop()
        r = "_w"
        if not projvar.invran_weekly_span:  # if investigation range is daily
            r = "_d"
        xl_filename = "man5" + str(format(self.dates[0], "_%y_%m_%d")) + r + ".xlsx"
        try:
            self.wb.save(dir_path('mandates_5') + xl_filename)
            messagebox.showinfo("Spreadsheet generator",
                                "Your spreadsheet was successfully generated. \n"
                                "File is named: {}".format(xl_filename),
                                parent=self.frame)
            if sys.platform == "win32":
                os.startfile(dir_path('mandates_5') + xl_filename)
            if sys.platform == "linux":
                subprocess.call(["xdg-open", 'kb_sub/mandates_5/' + xl_filename])
            if sys.platform == "darwin":
                subprocess.call(["open", dir_path('mandates_5') + xl_filename])
        except PermissionError:
            messagebox.showerror("Spreadsheet generator",
                                 "The spreadsheet was not opened. \n"
                                 "Suggestion: "
                                 "Make sure that identically named spreadsheets are closed "
                                 "(the file can't be overwritten while open).",
                                 parent=self.frame)

    def save_open_report(self):
        """ name and open the text file """
        self.report.close()
        if sys.platform == "win32":  # open the text document
            os.startfile(dir_path('report') + self.report_filename)
        if sys.platform == "linux":
            subprocess.call(["xdg-open", 'kb_sub/report/' + self.report_filename])
        if sys.platform == "darwin":
            subprocess.call(["open", dir_path('report') + self.report_filename])


class OffbidSpreadsheet:
    """
    Create a spreadsheet for calculating and detecting situations where carriers do no get 8 hours of work
    on their own bid assignments due to off route assignments.
    """
    def __init__(self):
        self.frame = None  # the frame of parent
        self.pb = None  # progress bar object
        self.pbi = 0  # progress bar count index
        self.startdate = None  # start date of the investigation
        self.enddate = None  # ending date of the investigation
        self.dates = []  # all days of the investigation
        self.carrierlist = []  # all carriers in carrier list
        self.wb = None  # the workbook object
        self.row = 1
        self.violation_number = 0
        self.ws_header = None  # style
        self.date_dov = None  # style
        self.date_dov_title = None  # style
        self.name_header = None  # style
        self.name_header_left = None  # style
        self.col_header = None  # style
        self.input_name = None  # style
        self.input_s = None  # style
        self.calcs = None  # style
        self.list_header = None  # style
        self.summary_name = None  # style
        self.offbid = None  # worksheet for the analysis of off bid violations
        self.summary = None  # worksheet for summary and maybe remedies
        self.carrier = ""  # carrier name
        self.route = ""  # carrier route
        self.rings = []  # carrier rings queried from database
        self.totalhours = ""  # carrier rings - 5200 time
        self.codes = ""  # carrier rings - code/note
        self.moves = ""  # carrier rings - moves on and off route with route
        self.move_i = 0  # increments rows for multiple move functionality
        self.i = 0  # the day being investigated as a number 0 - 6.
        self.max_pivot = 0.0  # the maximum allowed pivot.
        self.distinct_pages = None
        self.show_remedy = None  # if True, show the remedy
        self.show_sunday = None  # if False, do not display Sunday
        self.hourly_remedy = 0.0  # the hourly rate of the remedy in dollars
        self.move_set = 0  # extra rows used by the multiple moves display
        self.no_qualifiers = True  # is True as long as no violations have been found for any carriers.
        self.summary_array = []  # an array that holds: name, day and cell for violations
        self.summary_row = None  # hold the number of the row for the summary sheet

    def create(self, frame):
        """ a master method for running the other methods in proper order. """
        self.frame = frame
        if not self.ask_ok():  # abort if user selects cancel from askokcancel
            return
        self.pb = ProgressBarDe(label="Building Off Bid Assignment Spreadsheet")
        self.pb.max_count(1000)  # set length of progress bar
        self.pb.start_up()  # start the progress bar
        self.pbi = 1
        self.pb.move_count(self.pbi)  # increment progress bar
        self.pb.change_text("Initializing... ")
        self.get_dates()
        self.get_carrierlist()
        self.get_pb_max_count()
        self._get_settings()
        self.get_styles()
        self.build_workbook()
        self.carrierloop()
        self._build_summary()
        if self.no_qualifiers:
            self.no_violations()
        self.save_open()

    def ask_ok(self):
        """ ends process if user cancels """
        if messagebox.askokcancel("Spreadsheet generator",
                                  "Do you want to generate an \n"
                                  "Off Bid Assignment Spreadsheet?",
                                  parent=self.frame):
            return True
        return False

    def get_dates(self):
        """ get the dates from the project variables """
        self.startdate = projvar.invran_date  # set daily investigation range as default - get start date
        self.enddate = projvar.invran_date  # get end date
        self.dates = [projvar.invran_date, ]  # create an array of days - only one day if daily investigation range
        if projvar.invran_weekly_span:  # if the investigation range is weekly
            date = projvar.invran_date_week[0]
            self.startdate = projvar.invran_date_week[0]
            self.enddate = projvar.invran_date_week[6]
            self.dates = []
            for _ in range(7):  # create an array with all the days in the weekly investigation range
                self.dates.append(date)
                date += timedelta(days=1)

    def get_carrierlist(self):
        """ get record sets for all carriers """
        self.carrierlist = CarrierList(self.startdate, self.enddate, projvar.invran_station).get()

    def _get_settings(self):
        """ get the maximum pivot and distinct page value from the database. """
        sql = "SELECT tolerance FROM tolerances WHERE category = 'offbid_maxpivot'"
        result = inquire(sql)
        self.max_pivot = float(result[0][0])
        sql = "SELECT tolerance FROM tolerances WHERE category = 'offbid_distinctpage'"
        result = inquire(sql)
        self.distinct_pages = Convert(result[0][0]).str_to_bool()
        sql = "SELECT tolerance FROM tolerances WHERE category = 'offbid_show_remedy'"
        result = inquire(sql)
        self.show_remedy = Convert(result[0][0]).str_to_bool()
        sql = "SELECT tolerance FROM tolerances WHERE category = 'offbid_remedy'"
        result = inquire(sql)
        self.hourly_remedy = float(result[0][0])
        sql = "SELECT tolerance FROM tolerances WHERE category = 'offbid_show_sunday'"
        result = inquire(sql)
        self.show_sunday = Convert(result[0][0]).str_to_bool()

    def get_pb_max_count(self):
        """ set length of progress bar """
        self.pb.max_count(len(self.carrierlist)+1)  # once for each list in each day, plus reference, summary and saving

    def get_styles(self):
        """ Named styles for workbook """
        bd = Side(style='thin', color="80808080")  # defines borders
        self.ws_header = NamedStyle(name="ws_header", font=Font(bold=True, name='Arial', size=12))
        self.summary_name = NamedStyle(name="summary_name", font=Font(name='Arial', size=8),
                                       border=Border(left=bd, top=bd, right=bd, bottom=bd),
                                       alignment=Alignment(horizontal='left'))
        self.date_dov = NamedStyle(name="date_dov", font=Font(name='Arial', size=8),
                                   alignment=Alignment(horizontal='left'))
        self.date_dov_title = NamedStyle(name="date_dov_title", font=Font(bold=True, name='Arial', size=8),
                                         alignment=Alignment(horizontal='right'))
        self.name_header = NamedStyle(name="name_header", font=Font(bold=True, name='Arial', size=8, color='666666'),
                                      alignment=Alignment(horizontal='right'))
        self.name_header_left = NamedStyle(name="name_header_left",
                                           font=Font(bold=True, name='Arial', size=8, color='666666'),
                                           alignment=Alignment(horizontal='left'))
        self.col_header = NamedStyle(name="col_header", font=Font(bold=True, name='Arial', size=8, color='666666'),
                                     alignment=Alignment(horizontal='center'),
                                     border=Border(left=bd, top=bd, right=bd, bottom=bd))
        self.input_name = NamedStyle(name="input_name", font=Font(bold=True, name='Arial', size=10),
                                     alignment=Alignment(horizontal='left'))
        self.input_s = NamedStyle(name="input_s", font=Font(name='Arial', size=8),
                                  border=Border(left=bd, top=bd, right=bd, bottom=bd),
                                  alignment=Alignment(horizontal='right'))
        self.calcs = NamedStyle(name="calcs", font=Font(name='Arial', size=8),
                                border=Border(left=bd, top=bd, right=bd, bottom=bd),
                                fill=PatternFill(fgColor='e5e4e2', fill_type='solid'),
                                alignment=Alignment(horizontal='right'))
        self.list_header = NamedStyle(name="list_header", font=Font(bold=True, name='Arial', size=10))

    def build_workbook(self):
        """ creates the workbook object """
        self.wb = Workbook()  # define the workbook
        self.offbid = self.wb.active  # create first worksheet
        self.offbid.title = "off bid"  # title first worksheet
        self.offbid.oddFooter.center.text = "&A"
        self.summary = self.wb.create_sheet("summary")  # create summary worksheet
        self.summary.oddFooter.center.text = "&A"

    def carrierloop(self):
        """ loop for each carrier """
        for carrier in self.carrierlist:
            self.carrier = carrier[0][1]  # current iteration of carrier list is assigned self.carrier
            self.pbi += 1
            self.pb.move_count(self.pbi)  # increment progress bar
            self.pb.change_text("Checking: {}".format(self.carrier))
            self.route = carrier[0][4]  # get the route of the carrier
            self.get_rings()  # get individual carrier rings for the day - define self.rings
            if self.qualify():  # test the rings to see if there is a violation during the week
                self.no_qualifiers = False
                self.conditional_header()  # insert the header if proper conditions apply
                self.violation_number += 1  # increment the row number
                self.display_recs()  # build the carrier and the rings row into the spreadsheet
                self.conditional_pagebreak()

    def conditional_header(self):
        """ insert the header on certain conditions"""
        if self.violation_number == 0 or self.distinct_pages:
            self.build_headers()

    def build_headers(self, summary=False):
        """ worksheet headers """
        sheet = self.offbid
        row = self.row
        if summary:
            sheet = self.summary
            row = 1
        cell = sheet.cell(row=row, column=1)
        cell.value = "Off Bid Assignment Worksheet"
        if summary:
            cell.value = "Off Bid Assignment Summary"
        cell.style = self.ws_header
        sheet.merge_cells('A' + str(row) + ':E' + str(row))
        row += 2
        cell = sheet.cell(row=row, column=1)
        cell.value = "Date:  "  # create date/ pay period/ station header
        cell.style = self.date_dov_title
        cell = sheet.cell(row=row, column=2)
        date_string = self.dates[0].strftime("%x")
        # The date can be one day or a service week (a range of 7 days)
        if len(self.dates) > 1:
            date_string = self.dates[0].strftime("%x") + " - " + self.dates[6].strftime("%x")
        cell.value = date_string  # fill in the date/s
        cell.style = self.date_dov
        sheet.merge_cells('B' + str(row) + ':D' + str(row))
        cell = sheet.cell(row=row, column=5)
        cell.value = "Pay Period:  "
        cell.style = self.date_dov_title
        sheet.merge_cells('E' + str(row) + ':F' + str(row))
        cell = sheet.cell(row=row, column=7)
        cell.value = projvar.pay_period
        cell.style = self.date_dov
        sheet.merge_cells('G' + str(row) + ':H' + str(row))
        row += 1
        cell = sheet.cell(row=row, column=1)
        cell.value = "Station:  "
        cell.style = self.date_dov_title
        cell = sheet.cell(row=row, column=2)
        cell.value = projvar.invran_station
        cell.style = self.date_dov
        sheet.merge_cells('B' + str(row) + ':D' + str(row))
        if summary and self.show_remedy:
            cell = sheet.cell(row=row, column=5)
            cell.value = "Remedy Rate:  "
            cell.style = self.date_dov_title
            cell = sheet.cell(row=row, column=7)
            cell.value = self.hourly_remedy
            cell.style = self.date_dov
            cell.number_format = "[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00"
            sheet.merge_cells('E' + str(row) + ':F' + str(row))
        if not summary:
            self.row += 5

    def get_rings(self):
        """ get individual carrier rings for the day - define self.rings"""
        self.rings = []
        for date in self.dates:
            rings = Rings(self.carrier, date).get_for_day()  # assign as rings
            totalhours = 0.0  # set default as an empty string
            codes = ""
            moves = ""
            if rings[0]:  # if rings record is not blank
                totalhours = float(Convert(rings[0][2]).zero_not_empty())
                codes = rings[0][4]
                moves = rings[0][5]
            to_add = [date, totalhours, codes, moves]
            self.rings.append(to_add)

    def qualify(self):
        """ test to see if the carrier/rings need to be displayed. """
        qualify = False
        for i in range(len(self.rings)):  # loops for each day in the investigation range
            if self.number_crunching(i):  # returns True if there is a violation
                qualify = True
            else:  # if there is no violations for the day - insert False into self.rings
                self.rings[i].append(False)
        if qualify:  # if there is a violation for at least one day
            return True
        return False

    def number_crunching(self, i):
        """ do calculations to determine off route, on route and violation values.
            returns True if there is a violation. Adds violation boolean to self.rings. """
        offroute = 0.0  # this is the total time spent off the carrier's route
        if not self.rings[i][1]:  # if the total hours is zero - the violation is zero
            return False
        if self.rings[i][2] == "ns day":  # if it is the carrier's ns day - violation is zero
            return False
        if not self.rings[i][3]:  # if the moves is empty, then the violation is zero
            return False
        # if the setting for show sunday is false/off, do not show any day which is sunday
        if not self.show_sunday and self.rings[i][0].strftime("%a") == "Sun":
            return False
        index = 0  # set the index to 1. This will point to an element in the moves array.
        totalhours = self.rings[i][1]  # simplify the variable name
        moves = Convert(self.rings[i][3]).string_to_array()  # simplify the variable name
        while index < len(moves):  # calculate the total time off route
            offroute += float(moves[index+1]) - float(moves[index])
            index += 3
        ownroute = max(totalhours - offroute, 0)   # calculate the total time spent on route
        violation = max(8 - ownroute, 0)  # calculate the total violation
        if violation > self.max_pivot:
            self.rings[i].append(True)
            return True
        return False

    def display_recs(self):
        """ build the carrier and ring recs into the spreadsheet. """
        cell = self.offbid.cell(row=self.row, column=1)  # carrier label
        cell.value = "carrier:  "
        cell.style = self.name_header
        cell = self.offbid.cell(row=self.row, column=2)  # carrier name input
        cell.value = self.carrier
        cell.style = self.input_name
        self.offbid.merge_cells('B' + str(self.row) + ':E' + str(self.row))
        cell = self.offbid.cell(row=self.row, column=6)  # route label
        cell.value = "route:  "
        cell.style = self.name_header
        cell = self.offbid.cell(row=self.row, column=7)  # route input
        cell.value = self.route
        cell.style = self.input_name
        self.offbid.merge_cells('G' + str(self.row) + ':J' + str(self.row))
        self.row += 1
        # use loops and an array to build the column headers
        column_headers = ("day", "date", "5200", "mv off", "mv on", "route", "off rt", "on rt", "violation")
        for i in range(9):
            cell = self.offbid.cell(row=self.row, column=i + 2)  # column headers
            cell.value = column_headers[i]
            cell.style = self.col_header
        self. row += 1
        self.display_daily()  # create a row/s to display the daily information on the violation
        self.row += 1

    def _build_summary_array(self, i):
        """ builds the summary_array which holds the name, day and cell coordinates as a
        tuple (sheet, column, row) """
        column = "J"  # column for all violations is J
        coordinates = (self.offbid.title, column, self.row)
        add_this = (self.carrier, self.rings[i][0].strftime("%a"), self.rings[i][0].strftime("%m/%d/%Y"), coordinates)
        self.summary_array.append(add_this)

    def display_daily(self):
        """ display the daily ring recs for the carrier. """
        for i in range(len(self.rings)):
            if self.rings[i][4]:
                cell = self.offbid.cell(row=self.row, column=2)  # day
                cell.value = self.rings[i][0].strftime("%a")
                cell.style = self.input_s
                cell = self.offbid.cell(row=self.row, column=3)  # date
                cell.value = self.rings[i][0].strftime("%m/%d/%Y")
                cell.style = self.input_s
                cell = self.offbid.cell(row=self.row, column=4)  # 5200
                cell.value = self.rings[i][1]
                cell.style = self.input_s
                cell.number_format = "#,###.00;[RED]-#,###.00"
                self.display_moves(i)
                cell = self.offbid.cell(row=self.row, column=9)  # on route
                formula = "=MAX(D" + str(self.row) + "-H" + str(self.row) + ",0)"
                cell.value = formula
                cell.style = self.calcs
                cell.number_format = "#,###.00;[RED]-#,###.00"
                cell = self.offbid.cell(row=self.row, column=10)  # violation
                formula = "=IF(AND(D" + str(self.row) + ">0,H" + str(self.row) + ">0),8-I" + str(self.row) + ",0)"
                cell.value = formula
                cell.style = self.calcs
                cell.number_format = "#,###.00;[RED]-#,###.00"
                self._build_summary_array(i)
                self.row += (self.move_set - 1)  # correct for increment after last move set.
                self.row += 1

    def display_moves(self, i):
        """ display the moves. include contingencies for multiple moves.
            also displays the off route formula column as that changes with multiple moves"""
        moves = Convert(self.rings[i][3]).string_to_array()
        set_count = Moves().count_movesets(moves)
        if len(moves) > 3:
            moves = ["*", "*", "*"] + moves
        move_place = 0
        self.move_set = 0  # extra rows used to display multiple moves/ incremented in self.display_moves()
        for move in moves:
            cell = self.offbid.cell(row=self.row + self.move_set, column=5 + move_place)  # move off
            cell.value = move
            cell.style = self.input_s
            if move_place == 2:
                formulacell = self.offbid.cell(row=self.row + self.move_set, column=8)  # formula cell
                formulacell.style = self.calcs
                if not self.move_set and set_count > 1:  # if this is the first row of a multiple row
                    formula = "=SUM(H" + str(self.row + 1) + ":H" + str(self.row + set_count) + ")"
                else:
                    formula = "=SUM(F" + str(self.row + self.move_set) + "-E" + str(self.row + self.move_set) + ")"
                formulacell.value = formula
                formulacell.number_format = "#,###.00;[RED]-#,###.00"
                move_place = 0
                self.move_set += 1
            else:
                move_place += 1
                cell.number_format = "#,###.00;[RED]-#,###.00"

    def conditional_pagebreak(self):
        """ insert a page break if the correct conditions apply """
        if self.distinct_pages:
            try:
                self.offbid.page_breaks.append(Break(id=self.row))
                self.row += 1
            except AttributeError:
                self.offbid.row_breaks.append(Break(id=self.row))  # effective for windows
                self.row += 1

    def _build_summary_columnheaders(self):
        """ create column headers for the summary sheet """
        summary_row = 6
        titles = ("name", "day", "date", "violation", "total", "remedy")
        column = ("2", "4", "5", "6", "7", "8")
        column_range = 5
        if self.show_remedy:
            column_range = 6
        for i in range(column_range):
            cell = self.summary.cell(row=summary_row, column=int(column[i]))  # carrier name input
            cell.value = titles[i]
            if i == 0:
                cell.style = self.name_header_left
            else:
                cell.style = self.name_header

    def _build_summary_body(self):
        """ builds the body of the summary sheet, showing name, day, date, violation, total and maybe remedy. """
        name_array = []
        self.summary_row = 6
        for name in self.summary_array:
            if name[0] not in name_array:
                name_occurances = 0
                for array in self.summary_array:  # count the number of violations for each name
                    if array[0] == name[0]:  # if the carrier name == the name in the array
                        name_occurances += 1
                self.summary_row += 1
                cell = self.summary.cell(row=self.summary_row, column=2)  # carrier name input
                cell.value = name[0]
                cell.style = self.summary_name
                self.summary.merge_cells('B' + str(self.summary_row) + ':C' + str(self.summary_row))
                cell = self.summary.cell(row=self.summary_row, column=7)  # total violation formula column
                formula = "=SUM(F%s:F%s)" % (str(self.summary_row), str(self.summary_row + (name_occurances - 1)))
                cell.value = formula
                cell.style = self.input_s
                cell.number_format = "#,###.00;[RED]-#,###.00"
                if self.show_remedy:
                    cell = self.summary.cell(row=self.summary_row, column=8)  # cumulative remedy formula
                    formula = "=%s!G%s*%s!G4" \
                              % ("summary", str(self.summary_row), "summary")
                    cell.value = formula
                    cell.style = self.calcs
                    cell.number_format = "[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00"
                if name[0] not in name_array:
                    name_array.append(name[0])
            cell = self.summary.cell(row=self.summary_row, column=4)  # day column
            cell.value = name[1]
            cell.style = self.input_s
            cell = self.summary.cell(row=self.summary_row, column=5)  # date column
            cell.value = name[2]
            cell.style = self.input_s
            cell = self.summary.cell(row=self.summary_row, column=6)  # formula column
            formula = '=\'%s\'!J%s' % (name[3][0], str(name[3][2]))
            cell.value = formula
            cell.style = self.input_s
            cell.number_format = "#,###.00;[RED]-#,###.00"
            self.summary_row += 1

    def _build_summary_footer(self):
        """ builds the cumulative total and maybe the cummulative remedy for the summary sheet. """
        self.summary_row += 1
        cell = self.summary.cell(row=self.summary_row, column=4)  # cumulative total label
        cell.value = "Cumulative Violations: "
        cell.style = self.date_dov_title
        self.summary.merge_cells('D' + str(self.summary_row) + ':F' + str(self.summary_row))
        cell = self.summary.cell(row=self.summary_row, column=7)  # cumulative total formula
        formula = "=SUM(G7:G%s)" % (str(self.summary_row - 2))
        cell.value = formula
        cell.style = self.calcs
        cell.number_format = "#,###.00;[RED]-#,###.00"
        if self.show_remedy:
            cell = self.summary.cell(row=self.summary_row + 2, column=4)  # cumulative remedy label
            cell.value = "Cumulative Remedy: "
            cell.style = self.date_dov_title
            self.summary.merge_cells('D' + str(self.summary_row + 2) + ':F' + str(self.summary_row + 2))
            cell = self.summary.cell(row=self.summary_row + 2, column=8)  # cumulative remedy formula
            formula = "=SUM(H7:H%s)" % (str(self.summary_row - 2))
            cell.value = formula
            cell.style = self.calcs
            cell.number_format = "[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00"

    def _build_summary(self):
        """ fill the contents of the summary with the results of the summary array """
        self.build_headers(summary=True)  # worksheet headers - 'summry=True' generates header for summary
        if not self.no_qualifiers:  # if there are violations to show..
            self._build_summary_columnheaders()  # create column headers for the summary sheet
            self._build_summary_body()  # builds the body of the summary
            self._build_summary_footer()  # builds the cumulative total and maybe the cummulative remedy

    def no_violations(self):
        """ if self.no_qualifiers is True after all carriers have been checked, this will display a message
            saying that no violations occured. """
        self.build_headers()
        self.build_headers(summary=True)
        for sheet in (self.offbid, self.summary):
            cell = sheet.cell(row=6, column=2)  # No off bid violation found
            cell.value = "No off bid violations found"
            cell.style = self.list_header
            sheet.merge_cells('B6' + ':J6')

    def save_open(self):
        """ name and open the excel file """
        self.pbi += 1
        self.pb.move_count(self.pbi)  # increment progress bar
        self.pb.change_text("Saving...")
        self.pb.stop()
        r = "_w"
        if not projvar.invran_weekly_span:  # if investigation range is daily
            r = "_d"
        xl_filename = "kb_offbid_" + str(format(self.dates[0], "_%y_%m_%d")) + r + ".xlsx"
        try:
            self.wb.save(dir_path('off_bid') + xl_filename)
            messagebox.showinfo("Spreadsheet generator",
                                "Your spreadsheet was successfully generated. \n"
                                "File is named: {}".format(xl_filename),
                                parent=self.frame)
            if sys.platform == "win32":
                os.startfile(dir_path('off_bid') + xl_filename)
            if sys.platform == "linux":
                subprocess.call(["xdg-open", 'kb_sub/off_bid/' + xl_filename])
            if sys.platform == "darwin":
                subprocess.call(["open", dir_path('off_bid') + xl_filename])
        except PermissionError:
            messagebox.showerror("Spreadsheet generator",
                                 "The spreadsheet was not opened. \n"
                                 "Suggestion: "
                                 "Make sure that identically named spreadsheets are closed "
                                 "(the file can't be overwritten while open).",
                                 parent=self.frame)


class OtAvailSpreadsheet:
    """ this will generate a spreadsheet that will display the hours and paid leave an otdl carrier has worked, 
    and use this to tally how much availability that carrier has day to day.
    This was modified to include odln and odln"""
    def __init__(self):
        self.frame = None
        self.pb = None  # progress bar object
        self.pbi = 0  # progress bar count index
        self.carrier_list = []  # build a carrier list
        self.nsday_dict = {}
        self.ot_carrier = None
        self.wb = None  # workbook object
        self.availability = None  # workbook object sheet
        self.startdate = None  # start of the investigation range
        self.enddate = None  # ending day of the investigation range
        self.dates = []  # all the dates of the investigation range
        self.rings = []  # all rings for all carriers in the carrier list
        self.ws_header = None  # style
        self.date_dov = None  # style
        self.date_dov_title = None  # style
        self.name_header = None  # style
        self.col_header = None  # style
        self.input_name = None  # style
        self.input_s = None  # style
        self.calcs = None  # style
        self.list_header = None  # style
        self.violation_recsets = []  # carrier info, daily hours, leavetypes and leavetimes
        self.row = 1

    def create(self, frame):
        """ master method for calling methods"""
        self.frame = frame
        if not self.ask_ok():
            return
        self.get_dates()
        self.get_carrierlist()
        self.pb = ProgressBarDe(label="OTDL Weekly Availability Spreadsheet")
        self.pb.max_count(len(self.carrier_list))  # set length of progress bar
        self.pb.start_up()  # start the progress bar
        self.pbi = 1
        self.pb.move_count(self.pbi)  # increment progress bar
        self.pb.change_text("Gathering Data... ")
        self.build_workbook()
        self.get_styles()
        self.set_dimensions()
        self.build_headers()
        self.carrierloop()
        self.save_open()

    def ask_ok(self):
        """ continue if user selects ok. """
        if messagebox.askokcancel("Spreadsheet generator",
                                  "Do you want to generate a spreadsheet for OTDL availability?",
                                  parent=self.frame):
            return True
        return False
        
    def get_dates(self):
        """ get the dates of the investigation range from the project variables. """
        date = projvar.invran_date_week[0]
        self.startdate = projvar.invran_date_week[0]
        self.enddate = projvar.invran_date_week[6]
        for _ in range(7):
            self.dates.append(date)
            date += timedelta(days=1)

    def get_carrierlist(self):
        """ call the carrierlist class from kbtoolbox module to get the carrier list """
        carrierlist = CarrierList(self.startdate, self.enddate, projvar.invran_station).get()
        for carrier in carrierlist:
            for rec in carrier:
                if rec[2] in ("otdl", "odlr", "odln"):
                    self.carrier_list.append(carrier[0])  # add record for each otdl carrier in carrier list
                    self.nsday_dict.setdefault(rec[1], projvar.ns_code[rec[3]].lower())  # put ns day in dictionary
                    break

    def get_styles(self):
        """ Named styles for workbook """
        bd = Side(style='thin', color="80808080")  # defines borders
        self.ws_header = NamedStyle(name="ws_header", font=Font(bold=True, name='Arial', size=12))
        self.date_dov = NamedStyle(name="date_dov", font=Font(name='Arial', size=8))
        self.date_dov_title = NamedStyle(name="date_dov_title", font=Font(bold=True, name='Arial', size=8),
                                         alignment=Alignment(horizontal='right'))
        self.name_header = NamedStyle(name="name_header", font=Font(bold=True, name='Arial', size=8, color='666666'),
                                      alignment=Alignment(horizontal='right'))
        self.col_header = NamedStyle(name="col_header", font=Font(bold=True, name='Arial', size=8, color='666666'),
                                     alignment=Alignment(horizontal='center'),
                                     border=Border(left=bd, top=bd, right=bd, bottom=bd))
        self.input_name = NamedStyle(name="input_name", font=Font(bold=True, name='Arial', size=10),
                                     alignment=Alignment(horizontal='left'))
        self.input_s = NamedStyle(name="input_s", font=Font(name='Arial', size=8),
                                  border=Border(left=bd, top=bd, right=bd, bottom=bd),
                                  alignment=Alignment(horizontal='right'))
        self.calcs = NamedStyle(name="calcs", font=Font(name='Arial', size=8),
                                border=Border(top=bd, right=bd, bottom=bd, left=bd),
                                fill=PatternFill(fgColor='e5e4e2', fill_type='solid'),
                                alignment=Alignment(horizontal='right'))
        self.list_header = NamedStyle(name="list_header", font=Font(bold=True, name='Arial', size=10))

    def build_headers(self):
        """ self.availability worksheet header - format cells """
        self.pb.change_text("Building Availability...")
        self.availability.merge_cells('A1:O1')
        self.availability['A1'] = "OTDL Weekly Availability Worksheet"
        self.availability['A1'].style = self.ws_header
        self.availability['A3'] = "Date: "  # date label
        self.availability['A3'].style = self.date_dov_title
        self.availability.merge_cells('B3:H3')  # blank field for date
        self.availability['B3'] = self.dates[0].strftime("%x") + " - " + self.dates[6].strftime("%x")
        self.availability['B3'].style = self.date_dov
        self.availability.merge_cells('I3:L3')  # pay period label
        self.availability['I3'] = "Pay Period: "
        self.availability['I3'].style = self.date_dov_title  # blank field for pay period
        self.availability.merge_cells('M3:O3')
        self.availability['M3'] = projvar.pay_period
        self.availability['M3'].style = self.date_dov
        self.availability['A4'] = "Station: "  # station label
        self.availability['A4'].style = self.date_dov_title
        self.availability.merge_cells('B4:O4')  # blank field for station
        self.availability['B4'] = projvar.invran_station
        self.availability['B4'].style = self.date_dov

    def build_workbook(self):
        """ creates the workbook object """
        self.pb.change_text("Building workbook...")
        self.wb = Workbook()  # define the workbook
        self.availability = self.wb.active  # create first worksheet
        self.availability.title = "availability"  # title first worksheet
        self.availability.oddFooter.center.text = "&A"
        
    def set_dimensions(self):
        """ adjust the height and width on the violations/ instructions page """
        for x in range(2, 4):
            self.availability.row_dimensions[x].height = 10  # adjust all row height
        sheets = (self.availability, )
        for sheet in sheets:
            sheet.column_dimensions["A"].width = 19
            sheet.column_dimensions["B"].width = 6
            sheet.column_dimensions["C"].width = 2
            sheet.column_dimensions["D"].width = 6
            sheet.column_dimensions["E"].width = 2
            sheet.column_dimensions["F"].width = 6
            sheet.column_dimensions["G"].width = 2
            sheet.column_dimensions["H"].width = 6
            sheet.column_dimensions["I"].width = 2
            sheet.column_dimensions["J"].width = 6
            sheet.column_dimensions["K"].width = 2
            sheet.column_dimensions["L"].width = 6
            sheet.column_dimensions["M"].width = 2
            sheet.column_dimensions["N"].width = 6
            sheet.column_dimensions["O"].width = 2

    def carrierloop(self):
        """ loop for each carrier """
        self.row = 6  # allow space for headers
        first_page = True
        carriers_displayed = 0
        for carrier in self.carrier_list:
            self.pbi += 1
            self.pb.move_count(self.pbi)  # increment progress bar
            self.pb.change_text("Checking: {}".format(self.ot_carrier))
            if carrier[2] in ("otdl", "odlr", "odln"):
                self.ot_carrier = carrier[1]  # current iteration of carrier list is assigned self.carrier
                self.display_recs()
                carriers_displayed += 1
            if first_page and carriers_displayed == 4:  # allow only five carriers per page.
                self.make_pagebreak()  # insert a page break
                carriers_displayed = 0  # reinitialize the counter
                first_page = False
            if not first_page and carriers_displayed == 5:
                self.make_pagebreak()  # insert a page break
                carriers_displayed = 0  # reinitialize the counter

    def display_recs(self):
        """ build the carrier and ring recs into the spreadsheet. """
        ns_day_array = self.get_nsday()
        full_day_dict = {"sat": "saturday", "sun": "sunday", "mon": "monday", "tue": "tuesday", "wed": "wednesday",
                         "thu": "thursday", "fri": "friday", "  ": ""}
        merge_first = ("B", "D", "F", "H", "J", "L", "N")
        merge_second = ("C", "E", "G", "I", "K", "M", "O")
        col_increment = 2
        cell = self.availability.cell(row=self.row, column=1)  # carrier label
        cell.value = "carrier:  "
        cell.style = self.name_header

        cell = self.availability.cell(row=self.row, column=2)  # carrier name input
        cell.value = self.ot_carrier
        cell.style = self.input_name
        self.availability.merge_cells('B' + str(self.row) + ':I' + str(self.row))
        cell = self.availability.cell(row=self.row, column=10)  # carrier ns day label
        cell.value = "ns day: "
        cell.style = self.name_header
        self.availability.merge_cells('J' + str(self.row) + ':K' + str(self.row))

        cell = self.availability.cell(row=self.row, column=12)  # carrier ns day input
        cell.value = full_day_dict[self.nsday_dict[self.ot_carrier]]
        cell.style = self.input_name
        self.availability.merge_cells('L' + str(self.row) + ':O' + str(self.row))

        self.row += 1
        # row headers
        cell = self.availability.cell(column=1, row=self.row + 1)  # paid leave label
        cell.value = "paid leave/ type: "
        cell.style = self.name_header
        self.availability.row_dimensions[self.row + 1].height = 12  # adjust all row height
        cell = self.availability.cell(column=1, row=self.row + 2)  # hours worked label
        cell.value = "hours worked/ ns day: "
        cell.style = self.name_header
        self.availability.row_dimensions[self.row + 2].height = 12  # adjust all row height
        cell = self.availability.cell(column=1, row=self.row + 3)  # cumulative hours label
        cell.value = "cumulative hours: "
        cell.style = self.name_header
        self.availability.row_dimensions[self.row + 3].height = 12  # adjust all row height
        cell = self.availability.cell(column=1, row=self.row + 4)  # cumulative overtime hours label
        cell.value = "cumulative overtime: "
        cell.style = self.name_header
        self.availability.row_dimensions[self.row + 4].height = 12  # adjust all row height
        cell = self.availability.cell(column=1, row=self.row + 5)  # weekly availability label
        cell.value = "available weekly: "
        cell.style = self.name_header
        self.availability.row_dimensions[self.row + 5].height = 12  # adjust all row height
        cell = self.availability.cell(column=1, row=self.row + 6)  # overtime availability label
        cell.value = "available overtime: "
        cell.style = self.name_header
        self.availability.row_dimensions[self.row + 6].height = 12  # adjust all row height
        cell = self.availability.cell(column=1, row=self.row + 7)  # cumulative hours label
        cell.value = "available daily: "
        cell.style = self.name_header
        self.availability.row_dimensions[self.row + 7].height = 12  # adjust all row height
        # use loops and an array to build the column headers
        column_headers = ("sat", "sun", "mon", "tue", "wed", "thu", "fri")
        for i in range(7):
            # get the total hours, leave type and leave hours for the carrier
            rings = self.get_rings(self.dates[i])
            # ------------------------------------------------------------------------------------- column headers row
            cell = self.availability.cell(column=i + col_increment, row=self.row)
            cell.value = column_headers[i]
            cell.style = self.col_header
            self.availability.merge_cells(str(merge_first[i]) + str(self.row) + ":" +
                                          str(merge_second[i]) + str(self.row))
            # ----------------------------------------------------------------------------------------- paid leave row
            cell = self.availability.cell(column=i + col_increment, row=self.row + 1)  # display paid leave hours
            cell.value = self.format_time(rings[2])  # format and display leave time
            cell.style = self.input_s
            cell.number_format = "#,###.00;[RED]-#,###.00"
            cell = self.availability.cell(column=i + 1 + col_increment, row=self.row + 1)  # display leave code
            cell.value = self.leave_code(rings[1])  # format and display leave code
            cell.style = self.col_header
            # ----------------------------------------------------------------------------------------- 5200 hours row
            cell = self.availability.cell(column=i + col_increment, row=self.row + 2)  # display 5200 hours
            cell.value = self.format_time(rings[0])
            cell.style = self.input_s
            cell.number_format = "#,###.00;[RED]-#,###.00"
            cell = self.availability.cell(column=i + 1 + col_increment, row=self.row + 2)  # display ns day indicator
            cell.value = ""  # display indicator for ns day
            if i in ns_day_array:
                cell.value = "N"
            cell.style = self.col_header
            # --------------------------------------------------------------------------------------- cumulative hours
            cell = self.availability.cell(column=i + col_increment, row=self.row + 3)
            cell.value = self.cum_formula(i, self.row)  # get the formula for the cell
            cell.style = self.calcs
            cell.number_format = "#,###.00;[RED]-#,###.00"
            self.availability.merge_cells(str(merge_first[i]) + str(self.row + 3) + ":" +
                                          str(merge_second[i]) + str(self.row + 3))
            # ------------------------------------------------------------------------------------- cumulative overtime
            cell = self.availability.cell(column=i + col_increment, row=self.row + 4)
            cell.value = self.cum_ot_formula(i, self.row)  # get the formula for the cell
            cell.style = self.calcs
            cell.number_format = "#,###.00;[RED]-#,###.00"
            self.availability.merge_cells(str(merge_first[i]) + str(self.row + 4) + ":" +
                                          str(merge_second[i]) + str(self.row + 4))
            # ------------------------------------------------------------------------------------- weekly availability
            cell = self.availability.cell(column=i + col_increment, row=self.row + 5)
            cell.value = self.avail_formula(i, self.row)  # get the formula for the cell
            cell.style = self.calcs
            cell.number_format = "#,###.00;[RED]-#,###.00"
            self.availability.merge_cells(str(merge_first[i]) + str(self.row + 5) + ":" +
                                          str(merge_second[i]) + str(self.row + 5))
            # ----------------------------------------------------------------------------------- overtime availability
            cell = self.availability.cell(column=i + col_increment, row=self.row + 6)
            cell.value = self.avail_ot_formula(i, self.row)  # get the formula for the cell
            cell.style = self.calcs
            cell.number_format = "#,###.00;[RED]-#,###.00"
            self.availability.merge_cells(str(merge_first[i]) + str(self.row + 6) + ":" +
                                          str(merge_second[i]) + str(self.row + 6))
            # --------------------------------------------------------------------------------------daily availability
            cell = self.availability.cell(column=i + col_increment, row=self.row + 7)
            cell.value = self.avail_daily(i, self.row)  # get the formula for the cell
            cell.style = self.calcs
            cell.number_format = "#,###.00;[RED]-#,###.00"
            self.availability.merge_cells(str(merge_first[i]) + str(self.row + 7) + ":" +
                                          str(merge_second[i]) + str(self.row + 7))
            col_increment += 1  # move over two columns
        self.row += 9

    def get_rings(self, date):
        """ get individual carrier rings for the day - define self.rings"""
        sql = "SELECT total, leave_type, leave_time FROM rings3 WHERE carrier_name = '%s' " \
              "AND rings_date = '%s' ORDER BY rings_date, carrier_name" % (self.ot_carrier, date)
        rings = inquire(sql)
        totalhours = 0.0  # set default as an empty string
        lv_type = ""
        lv_hours = ""
        if rings:  # if rings record is not blank
            totalhours = float(Convert(rings[0][0]).zero_not_empty())
            lv_type = rings[0][1]
            lv_hours = rings[0][2]
        return [totalhours, lv_type, lv_hours]

    @staticmethod
    def leave_code(leave):
        """ converts the leave type to a one letter code. """
        if leave == "annual":
            return "A"
        elif leave == "sick":
            return "S"
        elif leave == "holiday":
            return "H"
        elif leave == "other":
            return "O"
        elif leave == "combo":
            return "C"
        elif leave == "none":
            return ""
        else:
            return ""

    def get_nsday(self):
        """ get the nsday from self.nsday_dict and the self.get_rings() data.
        The method returns a list with a day of the week as an index. Sunday (1) is included in the list a default. """
        ns_days = ("sat", "sun", "mon", "tue", "wed", "thu", "fri")
        ns_day = [1, ]  # sunday is a default ns day
        for i in range(7):
            if i == 1:  # skip sunday
                continue
            sql = "SELECT code FROM rings3 WHERE carrier_name = '%s' " \
                  "AND rings_date = '%s' ORDER BY rings_date, carrier_name" % (self.ot_carrier, self.dates[i])
            rings = inquire(sql)
            if rings:
                if rings[0][0] in ("ns day", "no call"):
                    ns_day.append(i)
        if len(ns_day) > 1:
            return ns_day
        # if no ns day was found other than sunday, then use other method
        if self.ot_carrier in self.nsday_dict:  # if the carrier's name is in the nsday dictionary...
            if self.nsday_dict[self.ot_carrier] != '  ':  # if the carrier's nsday is not none
                ns_day.append(ns_days.index(self.nsday_dict[self.ot_carrier]))
        return ns_day

    @staticmethod
    def format_time(time):
        """ format the time for leave time and total time """
        if time == "0.0" or time == "0":
            return ""
        elif isfloat(time):
            return float(time)
        else:
            return time

    @staticmethod
    def cum_formula(day, row):
        """ return a formula for cumulative hours """
        if day == 0:  # if the day is saturday
            return "=SUM(%s!B%s+B%s)" % ('availability', str(row + 1), str(row + 2))
        if day == 1:  # if the day is sunday
            return "=SUM(%s!B%s+D%s+D%s)" % ('availability', str(row + 3), str(row + 1), str(row + 2))
        if day == 2:  # if the day is monday
            return "=SUM(%s!D%s+F%s+F%s)" % ('availability', str(row + 3), str(row + 1), str(row + 2))
        if day == 3:  # if the day is tuesday
            return "=SUM(%s!F%s+H%s+H%s)" % ('availability', str(row + 3), str(row + 1), str(row + 2))
        if day == 4:  # if the day is wednesday
            return "=SUM(%s!H%s+J%s+J%s)" % ('availability', str(row + 3), str(row + 1), str(row + 2))
        if day == 5:  # if the day is thursday
            return "=SUM(%s!J%s+L%s+L%s)" % ('availability', str(row + 3), str(row + 1), str(row + 2))
        if day == 6:  # if the day is friday
            return "=SUM(%s!L%s+N%s+N%s)" % ('availability', str(row + 3), str(row + 1), str(row + 2))

    @staticmethod
    def cum_ot_formula(day, row):
        """ return a formula for cumulative hours """
        if day == 0:  # if the day is saturday
            return "=IF(availability!C%s=\"\",MAX(availability!B%s-8,0),availability!B%s)" \
                   % (str(row + 2), str(row + 2), str(row + 2))
        if day == 1:  # if the day is sunday
            return "=SUM(IF(availability!E%s=\"\",MAX(availability!D%s-8,0),availability!D%s)+availability!B%s" \
                   % (str(row + 2), str(row + 2), str(row + 2), str(row + 4))
        if day == 2:  # if the day is monday
            return "=SUM(IF(availability!G%s=\"\",MAX(availability!F%s-8,0),availability!F%s)+availability!D%s" \
                   % (str(row + 2), str(row + 2), str(row + 2), str(row + 4))
        if day == 3:  # if the day is tuesday
            return "=SUM(IF(availability!I%s=\"\",MAX(availability!H%s-8,0),availability!H%s)+availability!F%s" \
                   % (str(row + 2), str(row + 2), str(row + 2), str(row + 4))
        if day == 4:  # if the day is wednesday
            return "=SUM(IF(availability!K%s=\"\",MAX(availability!J%s-8,0),availability!J%s)+availability!H%s" \
                   % (str(row + 2), str(row + 2), str(row + 2), str(row + 4))
        if day == 5:  # if the day is thursday
            return "=SUM(IF(availability!M%s=\"\",MAX(availability!L%s-8,0),availability!L%s)+availability!J%s" \
                   % (str(row + 2), str(row + 2), str(row + 2), str(row + 4))
        if day == 6:  # if the day is friday
            return "=SUM(IF(availability!O%s=\"\",MAX(availability!N%s-8,0),availability!N%s)+availability!L%s" \
                   % (str(row + 2), str(row + 2), str(row + 2), str(row + 4))

    @staticmethod
    def avail_formula(day, row):
        """ return a formula for cumulative hours """
        if day == 0:  # if the day is saturday
            return "=MAX(%s-%s!B%s, 0)" % (str(60), 'availability', str(row + 3))
        if day == 1:  # if the day is sunday
            return "=MAX(%s-%s!D%s, 0)" % (str(60), 'availability', str(row + 3))
        if day == 2:  # if the day is monday
            return "=MAX(%s-%s!F%s, 0)" % (str(60), 'availability', str(row + 3))
        if day == 3:  # if the day is tuesday
            return "=MAX(%s-%s!H%s, 0)" % (str(60), 'availability', str(row + 3))
        if day == 4:  # if the day is wednesday
            return "=MAX(%s-%s!J%s, 0)" % (str(60), 'availability', str(row + 3))
        if day == 5:  # if the day is thursday
            return "=MAX(%s-%s!L%s, 0)" % (str(60), 'availability', str(row + 3))
        if day == 6:  # if the day is friday
            return "=MAX(%s-%s!N%s, 0)" % (str(60), 'availability', str(row + 3))

    @staticmethod
    def avail_ot_formula(day, row):
        """ return a formula for cumulative hours """
        if day == 0:  # if the day is saturday
            return "=MAX(%s-%s!B%s, 0)" % (str(20), 'availability', str(row + 4))
        if day == 1:  # if the day is sunday
            return "=MAX(%s-%s!D%s, 0)" % (str(20), 'availability', str(row + 4))
        if day == 2:  # if the day is monday
            return "=MAX(%s-%s!F%s, 0)" % (str(20), 'availability', str(row + 4))
        if day == 3:  # if the day is tuesday
            return "=MAX(%s-%s!H%s, 0)" % (str(20), 'availability', str(row + 4))
        if day == 4:  # if the day is wednesday
            return "=MAX(%s-%s!J%s, 0)" % (str(20), 'availability', str(row + 4))
        if day == 5:  # if the day is thursday
            return "=MAX(%s-%s!L%s, 0)" % (str(20), 'availability', str(row + 4))
        if day == 6:  # if the day is friday
            return "=MAX(%s-%s!N%s, 0)" % (str(20), 'availability', str(row + 4))

    @staticmethod
    def avail_daily(day, row):
        """ return a formula for cumulative hours """
        if day == 0:  # if the day is saturday
            return "=MIN(IF(availability!C%s=\"\",MIN(MAX(12-availability!B%s,0),availability!B%s),0)," \
                   "availability!B%s)" \
                   % (str(row + 1), str(row + 2), str(row + 5), str(row + 6))
        if day == 1:  # if the day is sunday
            #        % (str(row + 1), str(row + 2), str(row + 5), str(row + 6))
            return "=MIN(IF(availability!E%s<>\"\",0,MIN(MAX(12-availability!D%s,0),availability!D%s))," \
                   "IF(AND(availability!E%s<>\"\",availability!B%s<8),0,availability!D%s))" \
                   % (str(row + 1), str(row + 2), str(row + 5), str(row + 2), str(row + 6), str(row + 6))
        if day == 2:  # if the day is monday
            return "=MIN(IF(availability!G%s<>\"\",0,MIN(MAX(12-availability!F%s,0),availability!F%s))," \
                   "IF(AND(availability!G%s<>\"\",availability!D%s<8),0,availability!F%s))" \
                   % (str(row + 1), str(row + 2), str(row + 5), str(row + 2), str(row + 6), str(row + 6))
        if day == 3:  # if the day is tuesday
            return "=MIN(IF(availability!I%s<>\"\",0,MIN(MAX(12-availability!H%s,0),availability!H%s))," \
                   "IF(AND(availability!I%s<>\"\",availability!F%s<8),0,availability!H%s))" \
                   % (str(row + 1), str(row + 2), str(row + 5), str(row + 2), str(row + 6), str(row + 6))
        if day == 4:  # if the day is wednesday
            return "=MIN(IF(availability!K%s<>\"\",0,MIN(MAX(12-availability!J%s,0),availability!J%s))," \
                   "IF(AND(availability!K%s<>\"\",availability!H%s<8),0,availability!J%s))" \
                   % (str(row + 1), str(row + 2), str(row + 5), str(row + 2), str(row + 6), str(row + 6))
        if day == 5:  # if the day is thursday
            return "=MIN(IF(availability!M%s<>\"\",0,MIN(MAX(12-availability!L%s,0),availability!L%s))," \
                   "IF(AND(availability!M%s<>\"\",availability!J%s<8),0,availability!L%s))" \
                   % (str(row + 1), str(row + 2), str(row + 5), str(row + 2), str(row + 6), str(row + 6))
        if day == 6:  # if the day is friday
            return "=MIN(IF(availability!O%s<>\"\",0,MIN(MAX(12-availability!N%s,0),availability!N%s))," \
                   "IF(AND(availability!O%s<>\"\",availability!L%s<8),0,availability!N%s))" \
                   % (str(row + 1), str(row + 2), str(row + 5), str(row + 2), str(row + 6), str(row + 6))

    def make_pagebreak(self):
        """ create a page break """
        try:
            self.availability.page_breaks.append(Break(id=self.row))
            self.row += 1
        except AttributeError:
            self.availability.row_breaks.append(Break(id=self.row))  # effective for windows
            self.row += 1

    def save_open(self):
        """ save the spreadsheet and open """
        self.pbi += 1
        self.pb.move_count(self.pbi)  # increment progress bar
        self.pb.change_text("Saving...")
        self.pb.stop()
        xl_filename = "kb_wa" + str(format(projvar.invran_date_week[0], "_%y_%m_%d")) + ".xlsx"
        try:
            self.wb.save(dir_path('weekly_availability') + xl_filename)
            messagebox.showinfo("Spreadsheet generator",
                                "Your spreadsheet was successfully generated. \n"
                                "File is named: {}".format(xl_filename),
                                parent=self.frame)
            if sys.platform == "win32":  # open the text document
                os.startfile(dir_path('weekly_availability') + xl_filename)
            if sys.platform == "linux":
                subprocess.call(["xdg-open", 'kb_sub/weekly_availability/' + xl_filename])
            if sys.platform == "darwin":
                subprocess.call(["open", dir_path('weekly_availability') + xl_filename])
        except PermissionError:
            messagebox.showerror("Spreadsheet generator",
                                 "The spreadsheet was not generated. \n"
                                 "Suggestion: "
                                 "Make sure that identically named spreadsheets are closed "
                                 "(the file can't be overwritten while open).",
                                 parent=self.frame)
    