"""
Klusterbox: The main module
Copyright 2019 Thomas Weeks

Caution: To ensure proper operation of Legacy Klusterbox outside Program Files (Windows) or Applications (mac OS),
make sure to keep the Klusterbox.exe and the kb_sub folder in the same folder.

For the newest version of Klusterbox, visit www.klusterbox.com/download.
Visit https://github.com/TomOfHelatrobus/klusterbox for the most recent source code.

This version of Klusterbox is being released under the GNU General Public License version 3.
"""
# custom modules
import projvar
from kbreports import InformalCIndex, Reports, Messenger, CheatSheet, Archive, InformalCReports, RptCarrierId, \
    InformalCOptions
from kbtoolbox import commit, inquire, Convert, Handler, dir_filedialog, dir_path, gen_ns_dict, \
    informalc_date_checker, isfloat, isint, macadj, MakeWindow, MinrowsChecker, NsDayDict, \
    ProgressBarDe, BackSlashDateChecker, CarrierList, CarrierRecFilter, dir_path_check, dt_converter, \
    find_pp, gen_carrier_list, Quarter, RingTimeChecker, Globals, \
    SpeedSettings, titlebar_icon, RefusalTypeChecker, ReportName, DateChecker, NameChecker, \
    RouteChecker, BuildPath, EmpIdChecker, SeniorityChecker, DateTimeChecker, GrievanceChecker, \
    IndexArticleChecker, IssueDecisionChecker, DecisionTypeChecker, distinctresult_to_list, \
    issuedecisionresult_sorter, AwardsChecker, AwardsFormatting, save_all
from kbspreadsheets import OvermaxSpreadsheet, ImpManSpreadsheet, ImpManSpreadsheet4, OffbidSpreadsheet, \
    OtAvailSpreadsheet
from kbdatabase import DataBase, setup_plaformvar, setup_dirs_by_platformvar, DovBase, DataBaseFix
from kbspeedsheets import SpeedSheetGen, OpenText, SpeedCarrierCheck, SpeedRingCheck
from kbequitability import QuarterRecs, OTEquitSpreadsheet, OTDistriSpreadsheet
from kbcsv_repair import CsvRepair
from kbcsv_reader import MaxHr, ee_skimmer
from kbpdfhandling import PdfConverter
from kbenterrings import EnterRings
from kbinformalc import InfcSpeedSheetGen, InfcSpeedWorkBookGet, Awards, informalc_gen_clist, \
    informalc_date_converter
from kbfixes import Fixes
# PDF Converter Libraries
from PyPDF2 import PdfFileReader, PdfFileWriter
# Standard Libraries
from tkinter import messagebox, filedialog, BooleanVar, Button, Checkbutton, \
    DISABLED, E, Entry, FALSE, Frame, IntVar, Label, LEFT, Menu, OptionMenu, Radiobutton, RIDGE, StringVar, \
    TclError, Tk, W, BOTH, BOTTOM, Canvas, END, Listbox, RIGHT, Scrollbar, VERTICAL, Y, ttk
from datetime import datetime, timedelta
import sqlite3
from operator import itemgetter
import os
# import shutil
from shutil import copyfile, rmtree
from csv import reader
import sys
import subprocess
import time
from math import ceil
from webbrowser import open_new  # for hyper link at about_klusterbox()
from threading import Thread  # run load workbook while progress bar runs
# Pillow Library
from PIL import ImageTk, Image  # Pillow Library
# Spreadsheet Libraries
from openpyxl import load_workbook

__author__ = "Thomas Weeks"
__author_email__ = "tomweeks@klusterbox.com"

# version variables
version = 6.00  # version number must be convertable to a float and should increase for Fixes()
release_date = "Dec 6, 2023"  # format is Jan 1, 2022


class ProgressBarIn:
    """ Indeterminate Progress Bar """

    def __init__(self, title="", label="", text=""):
        self.title = title
        self.label = label
        self.text = text
        self.pb_root = Tk()  # create a window for the progress bar
        self.pb_label = Label(self.pb_root, text=self.label)  # make label for progress bar
        self.pb = ttk.Progressbar(self.pb_root, length=400, mode="indeterminate")  # create progress bar
        self.pb_text = Label(self.pb_root, text=self.text, anchor="w")

    def start_up(self):
        """ starts up the progress bar. """
        titlebar_icon(self.pb_root)  # place icon in titlebar
        self.pb_root.title(self.title)
        self.pb_label.grid(row=0, column=0, sticky="w")
        self.pb.grid(row=1, column=0, sticky="w")
        self.pb_text.grid(row=2, column=0, sticky="w")
        while pb_flag:  # use global as a flag. stop loop when flag is False
            projvar.root.update()
            self.pb['value'] += 1
            time.sleep(.01)

    def stop(self):
        """ stops and destroys the progress bar. """
        self.pb.stop()  # stop and destroy the progress bar
        self.pb_text.destroy()
        self.pb_label.destroy()  # destroy the label for the progress bar
        self.pb.destroy()
        self.pb_root.destroy()


class InformalC:
    """
    This is the home page of the Informal C program.
    """
    def __init__(self):
        self.win = None  # tkinter frame object for the main window
        self.nav = None  # tkinter frame object for the mac navigation button.
        # station
        self.stationvar = None  # this is the stringvar for the station
        self.station_options = []  # the list of stations.
        self.station = None  # the station options
        # misc
        self.companion_root = None  # holds the root Tk for the informalc_root()
        self.listbox_fill = None  # used by several methods to carry list for listboxes # v
        # sql search
        self.grv_sql = ""  # the sql to search for distinct grievances
        self.set_sql = ""  # the sql to search for distinct settlements
        self.gats_sql = "SELECT DISTINCT grv_no FROM informalc_gats"  # get all grievances with gats numbers on record
        self.search_result = []  # var for the grievances search result
        self.search_grv_result = []  # a list of distinct grv numbers of grievances that match the search criteria
        self.search_set_result = []  # a list of distinct grv numbers of settlements that match the search criteria
        self.search_gat_result = []  # a list of distinct grv numbers that have associated gats numbers
        self.blank_criteria = True  # if all search criteria are blank, do not include 'where' in sql
        # grievance
        self.src_grievance = None  # stringvar used for search of grievance number # v
        # grievant
        self.grvent = []
        self.grvent_entry = []
        self.grvent_del = []
        self.grvent_showlist = None
        # incident date
        self.option_incidentdate = None  # stringvar used for search criteria option menu
        self.incident_start = None
        self.incident_end = None
        # meeting date
        self.option_meetingdate = None  # stringvar used for search criteria option menu
        self.meeting_start = None
        self.meeting_end = None
        # issue
        self.src_issue = []
        self.src_issue_entry = []
        self.src_issue_del = []
        self.src_issue_showlist = None
        self.issue_description = []  # a list of issues for the option menu
        self.issue_article = []  # a list of articles which corrosponds to the issues
        # level
        self.option_level = None  # stringvar used for search criteria option menu
        self.level_listbox = None
        self.level_options = (
            "informal a",
            "formal a",
            "step b",
            "pre arb",
            "arbitration"
        )
        # date signed
        self.option_signeddate = None  # stringvar used for search criteria option menu
        self.signed_start = None
        self.signed_end = None
        # decision
        self.decision = []
        self.decision_entry = []
        self.decision_del = []
        self.decision_showlist = None
        self.decision_description = []  # a list of decisions from the db.
        # proof due
        self.proofdue_start = None
        self.proofdue_end = None
        self.option_proofduedate = None  # stringvar used for search criteria option menu
        # docs
        self.option_docs = None  # stringvar used for search criteria option menu
        self.docs_listbox = None
        self.doc_options = (  # vars for document list/ option menus
            "no status",
            "non-applicable",
            "no",
            "yes",
            "unknown",
            "yes-not paid",
            "yes-in part",
            "yes-verified",
            "no-moot",
            "no-ignore"
        )
        # gats
        self.gats = None
        # option menu tuple for all date searchs.
        self.date_options = (
            "include all",
            "within specified range",
            "within last 6 months",
            "within last year",
            "within last two years",
            "within last three years"
        )
        # sort by
        self.sortby = None
        self.sort_order = None
        # vars for add awards
        self.var_id = None  # vars for addawards_screen()
        self.var_name = None  # vars for addawards_screen()
        self.var_award = None  # vars for addawards_screen()
        self.var_gats = None  # an array of stringvars for addawards_screen() gats discrepancies.
        self.award_gats_entry = []  # an array for holding entry widgets for gats discrepancies
        self.award_gats_del = []  # an array for holding delete buttons for gats discrepancies
        # vars for showtime nav bar function
        self.current_page = 1  # the current page of results to display
        self.rec_display_limit = 50  # the number of records displayed before a new page is needed

    def informalc(self, frame):
        """ a master method for running the other methods in proper sequence. """

        def clear_tempfolders():
            """ try clear contents of temp folder. ignore is the a file from the folder is in use. """
            try:
                if os.path.isdir(dir_path_check('infc_grv')):
                    rmtree(dir_path_check('infc_grv'))
            except PermissionError:
                pass

        def get_station():
            """ this sets the station to what was used for the klusterbox investigation range. """
            if projvar.invran_station:
                self.station = projvar.invran_station

        def get_station_options():
            """ this will get the station options ona place them in self.station_options"""
            for station in projvar.list_of_stations:
                self.station_options.append(station)
            if "out of station" in self.station_options:
                self.station_options.remove("out of station")

        def get_display_limit():
            """ fetch the informalc result limit from the tolerances table. """
            sql = "SELECT tolerance FROM tolerances WHERE category = '%s'" % "informalc_result_limit"
            result = inquire(sql)
            self.rec_display_limit = int(result[0][0])

        def get_issuecats():
            """ fetch the issue categories from the informalc_issuescategories table of the db
            and place them in arrays. """
            self.issue_description = []  # re initialize list
            self.issue_article = []  # re initialize list
            sql = "SELECT * FROM informalc_issuescategories"
            results = inquire(sql)
            results = issuedecisionresult_sorter(results)
            for r in results:
                self.issue_description.append(r[2])
                self.issue_article.append(r[1])

        def get_decisioncats():
            """ fetch the decision categories from the informalc_decisioncategories table of the db and place them in
             arrays """
            self.decision_description = []  # re initialize list
            sql = "SELECT * FROM informalc_decisioncategories"
            results = inquire(sql)
            results = issuedecisionresult_sorter(results)
            for r in results:
                self.decision_description.append(r[2])

        def station_screen_autorouting():
            """ this will automatically route the user depending on the amount of station options.
            One station option will automatically chose that option,
            Zero station options will show an error message and exit informal c. """
            if not self.station_options:
                messagebox.showerror("No Stations in Database",
                                     "There are no stations in the Klusterbox Database.\n"
                                     "Proper function of Informal C requires at least one "
                                     "station to be entered into the Klusterbox Database. \n"
                                     "Please return to Klusterbox and enter a station.\n\n"
                                     "> Management > List of Stations > Enter New Station",
                                     parent=frame)
                return True
            if len(self.station_options) == 1:
                self.station = self.station_options[0]
                self.menu_screen(frame)
                return True
            return False

        clear_tempfolders()  # clear contents of temp folder
        get_station()  # this uses the investigation range station as the default
        get_station_options()  # this gets the list of stations
        get_display_limit()  # get the number of search results per page to be displayed
        get_issuecats()  # gets all from informalc_issuescategories table and puts it in self.issue_description
        get_decisioncats()  # get all from informalc_decisioncategories and puts it in self.decision_description
        if not station_screen_autorouting():
            if not self.station:
                self.station_screen(frame)  # this allows the user to change/select the station
            else:
                self.menu_screen(frame)  # this fills the screen with widgets.

    def mac_navigation(self):
        """ create a screen for navigation to be used instead of the pulldown menu for macOS """
        self.nav = MakeWindow()
        self.nav.create(self.win.topframe)
        # options with a first value of 0 are enabled buttons.
        # options with a first value of 1 will be disabled if the investigation range is not set.
        # options with a first value of 2 are labels."""

        options = (
            (2, "Informal C Operations ______________________"),
            (0, "Open Archive", lambda: Archive().file_dialogue(dir_path('informalc_speedsheets'))),
            (0, "Clear Archive", lambda: Archive().remove_file_var(self.nav.topframe, 'informalc_speedsheets')),
            (0, "Generate New Grievances", lambda: InfcSpeedSheetGen(self.nav.topframe, self.station, "new").new()),
            (0, "Generate All Grievances", lambda: InfcSpeedSheetGen(self.nav.topframe, self.station, "all").all()),
            (0, "Pre-check", lambda: InfcSpeedWorkBookGet().open_file(self.nav.topframe, False)),
            (0, "Input to Database", lambda: InfcSpeedWorkBookGet().open_file(self.nav.topframe, True)),
            (0, "Speedsheet Guide", lambda: InformalCIndex().speedsheet_guide()),
            (0, "Grievant Guide", lambda: InformalCIndex().grievant_guide(self.station)),
        )
        i = 0
        row = 0
        for _ in range(len(options)):
            if options[i][0] == 3:  # if the option is a delete button
                button = Button(self.nav.body, text=options[i][1], width=5, anchor="w", padx=5,
                                activebackground="grey", highlightcolor="red", command=options[i][2])
                button.grid(row=row - 1, column=2, sticky="w")
            elif options[i][0] == 2:  # if the option is a header
                label = Label(self.nav.body, text=options[i][1], fg="blue", width=26, anchor="w")
                label.grid(row=row, column=1, pady=5, sticky="w", columnspan=3)
                row += 1
            else:  # the option is a button
                button = Button(self.nav.body, text=options[i][1], width=21, anchor="w", padx=5,
                                activebackground="grey", highlightcolor="red", command=options[i][2])
                button.grid(row=row, column=1, sticky="w")
                if not projvar.invran_day and options[i][0]:  # disable the button until invran is set.
                    button.config(state=DISABLED)
                row += 1
            i += 1
        Button(self.nav.buttons, text="Quit", width=macadj(13, 13), command=projvar.root.destroy).pack(side=LEFT)
        Button(self.nav.buttons, text="Go Back", width=15,
               command=lambda: self.informalc(frame=self.nav.topframe)).pack(side=LEFT)
        self.nav.finish()

    def pulldown_menu(self):
        """ create a pulldown menu, and add it to the menu bar """
        menubar = Menu(self.win.topframe)
        # speedsheeet menu
        speed_menu = Menu(menubar, tearoff=0)
        speed_menu.add_command(label="Open Archive",
                               command=lambda: Archive().file_dialogue(dir_path('informalc_speedsheets')))
        speed_menu.add_command(label="Clear Archive",
                               command=lambda: Archive().remove_file_var(self.win.topframe, 'informalc_speedsheets'))
        speed_menu.add_command(label="Generate New Grievances",
                               command=lambda: InfcSpeedSheetGen(self.win.topframe, self.station, "new").new())
        # speed_menu.add_command(label="Generate Selected Grievances",
        #                        command=lambda: InfcSpeedSheetGen(self.win.topframe, self.station, "selected")
        #                        .selected())
        speed_menu.add_command(label="Generate All Grievances",
                               command=lambda: InfcSpeedSheetGen(self.win.topframe, self.station, "all").all())
        speed_menu.add_command(label="Pre-check",
                               command=lambda: InfcSpeedWorkBookGet().open_file(self.win.topframe, False))
        speed_menu.add_command(label="Input to Database",
                               command=lambda: InfcSpeedWorkBookGet().open_file(self.win.topframe, True))
        speed_menu.add_command(label="Speedsheet Guide",
                               command=lambda: InformalCIndex().speedsheet_guide())
        speed_menu.add_command(label="Grievant Guide",
                               command=lambda: InformalCIndex().grievant_guide(self.station))
        menubar.add_cascade(label="Speedsheet", menu=speed_menu)
        projvar.root.config(menu=menubar)
        projvar.root.update()  # root update

    def station_screen(self, frame):
        """ this allows the user to change/ select the station """

        def station_screen_submit():
            """ this will update the station and route the user to the main menu
            or if no selection is made, there will be an error message and the screen will refresh. """
            if self.stationvar.get() == "Select a Station":
                messagebox.showerror("Prohibited Action",
                                     "Please select a station.",
                                     parent=self.win.body)
                self.station_screen(self.win.topframe)  # return to and refresh the station screen
            else:
                self.station = self.stationvar.get()
                self.menu_screen(self.win.topframe)

        self.win = MakeWindow()
        self.win.create(frame)  # creates the screen object
        self.stationvar = StringVar(self.win.body)
        row = 0
        Label(self.win.body, text="Informal C", font=macadj("bold", "Helvetica 18")).grid(row=row, sticky="w")
        row += 1
        Label(self.win.body, text="The C is for Compliance").grid(row=row, sticky="w")
        row += 1
        Label(self.win.body, text="").grid(row=row)
        row += 1
        Label(self.win.body, text="Please select a Station: ", background=macadj("gray95", "white"),
              fg=macadj("black", "black"), width=macadj(22, 20), anchor="w", height=macadj(1, 1)). \
            grid(row=row, column=0, sticky="w")
        Label(self.win.body, text="", height=macadj(1, 2)).grid(row=row, column=1)
        row += 1
        self.stationvar.set("Select a Station")
        station_om = OptionMenu(self.win.body, self.stationvar, *self.station_options)
        station_om.config(width=macadj(40, 34))
        station_om.grid(row=row, column=0, columnspan=2, sticky="e")
        # configure the submit button
        button_submit = Button(self.win.buttons)
        button_submit.config(text="Submit", width=20, command=lambda: station_screen_submit())
        if sys.platform == "win32":
            button_submit.config(anchor="w")
        button_submit.grid(row=0, column=1)
        # configure the "quit" button
        button_back = Button(self.win.buttons)
        button_back.config(text="Quit Informal C", width=20, command=lambda: MainFrame().start(self.win.topframe))
        if sys.platform == "win32":
            button_back.config(anchor="w")
        button_back.grid(row=0, column=0)
        self.win.finish()  # this commands the window to loop and persist.

    def menu_screen(self, frame):
        """ the main screen for informal c. """
        self.win = MakeWindow()
        self.win.create(frame)  # creates the screen object
        if not projvar.mac_navigation:
            self.pulldown_menu()  # create a pulldown menu, and add it to the menu bar
        Label(self.win.body, text="Informal C", font=macadj("bold", "Helvetica 18")).grid(row=0, sticky="w")
        Label(self.win.body, text="The C is for Compliance").grid(row=1, sticky="w")
        Label(self.win.body, text="").grid(row=2)
        row = 3
        if projvar.mac_navigation:  # conditional on user preference for navigation
            Button(self.win.body, text="Navigation", width=30,
                   command=lambda: self.mac_navigation()).grid(row=row, pady=5)
            row += 1
        Button(self.win.body, text=" Enter New Grievance", width=30,
               command=lambda: self.GrievanceInput(self).informalc_new(self.win.topframe)).grid(row=row, pady=5)
        row += 1
        Button(self.win.body, text="Grievance Tracker", width=30,
               command=lambda: self.master_search(self.win.topframe)).grid(row=row, pady=5)
        row += 1
        Button(self.win.body, text="Payout Entry", width=30,
               command=lambda: self.PayoutEntry(self).poe_search(self.win.topframe)).grid(row=row, pady=5)
        row += 1
        Button(self.win.body, text="Payout Report", width=30,
               command=lambda: self.PayoutReport(self).informalc_por(self.win.topframe)).grid(row=row, pady=5)
        row += 1
        Label(self.win.body, text="", width=70).grid(row=row)
        # configure the "quit" button
        button_back = Button(self.win.buttons)
        button_back.config(text="Quit Informal C", width=20, command=lambda: MainFrame().start(self.win.topframe))
        if sys.platform == "win32":
            button_back.config(anchor="w")
        button_back.grid(row=0, column=0)
        self.win.finish()  # this commands the window to loop and persist.

    def master_search(self, frame):
        """ master method for running other methods in proper order. """
        self.win = MakeWindow()
        self.win.create(frame)
        self.get_stringvars()
        self.initialize_listbox_components()  # initialize list holding strvars and widgets
        self.build_search_screen()
        # self.build_buttons()
        self.win.finish()

    def get_stringvars(self):
        """ initialize varibles """
        self.src_grievance = StringVar(self.win.topframe)  #
        self.incident_start = StringVar(self.win.topframe)  #
        self.incident_end = StringVar(self.win.topframe)  #
        self.meeting_start = StringVar(self.win.topframe)  #
        self.meeting_end = StringVar(self.win.topframe)  #
        self.signed_start = StringVar(self.win.topframe)  #
        self.signed_end = StringVar(self.win.topframe)  #
        self.proofdue_start = StringVar(self.win.topframe)  #
        self.proofdue_end = StringVar(self.win.topframe)  #
        self.gats = StringVar(self.win.topframe)  #
        self.option_incidentdate = StringVar(self.win.topframe)  #
        self.option_meetingdate = StringVar(self.win.topframe)  #
        self.option_signeddate = StringVar(self.win.topframe)  #
        self.option_proofduedate = StringVar(self.win.topframe)  #
        self.option_level = StringVar(self.win.topframe)  #
        self.option_docs = StringVar(self.win.topframe)  #
        self.sortby = StringVar(self.win.topframe)
        self.sort_order = StringVar(self.win.topframe)

    def initialize_listbox_components(self):
        """ initialize list holding stringvars, entry and button widgets for listbox/ entry fields for grievant,
        issue and decisions """
        self.grvent = []  # initialize a list holding stringvars
        self.grvent_entry = []  # list holding entry widgets
        self.grvent_del = []  # list holding button widgets
        self.src_issue = []  # initialize a list holding stringvars
        self.src_issue_entry = []  # list holding entry widgets
        self.src_issue_del = []  # list holding button widgets
        self.decision = []  # initialize a list holding stringvars
        self.decision_entry = []  # list holding entry widgets
        self.decision_del = []  # list holding button widgets

    def build_search_screen(self):
        """ builds page for searching grievance settlements. """

        def re_initialize_search_vars():
            """ initialize list for holding search results """
            self.grv_sql = ""  # the sql to search for distinct grievances
            self.set_sql = ""  # the sql to search for distinct settlements
            self.search_grv_result = []
            self.search_set_result = []
            self.search_gat_result = []
            self.search_result = []
            self.current_page = 1
            self.blank_criteria = True

        re_initialize_search_vars()  # in the event that a search is run more than once, re initialize
        button_alignment = macadj("w", "center")
        row = 0
        if not projvar.mac_navigation:
            self.pulldown_menu()  # create a pulldown menu, and add it to the menu bar
        Label(self.win.body, text="Grievance Search Criteria", font=macadj("bold", "Helvetica 18")) \
            .grid(row=row, columnspan=6, sticky="w")
        row += 1
        Label(self.win.body, text="").grid(row=row, columnspan=6)
        row += 1
        # ---------------------------------------------------------------------------------------------- mac navigation
        nav_button = Button(self.win.body, text="Navigation", width=23, anchor="center",
                            command=lambda: self.mac_navigation())
        if projvar.mac_navigation:  # conditional on user preference for navigation
            nav_button.grid(row=row, column=5, pady=3)  # display the navigation button
            row += 1
        # ---------------------------------------------------------------------------------------------- search for all
        Label(self.win.body, width=29, anchor="w", text="Search for all in {}".format(self.station))\
            .grid(row=row, column=0, columnspan=5, sticky="w")
        Button(self.win.body, text="Search All", width=23, anchor="center",
               command=lambda: self.search_all_apply(self.win.topframe)).grid(row=row, column=5, pady=3)
        row += 1
        # separator
        separator = ttk.Separator(self.win.body, orient='horizontal')
        separator.grid(column=0, row=row, columnspan=6, sticky="nesw", pady=5)
        row += 1
        # --------------------------------------------------------------------------------------------------- grievance
        grvframe = Frame(self.win.body)
        grvframe.grid(row=row, column=0, columnspan=6, sticky="w")
        Label(grvframe, text="Search by grievance number:", width=29, anchor="w")\
            .grid(row=0, column=0, columnspan=4, sticky="w")
        Entry(grvframe, textvariable=self.src_grievance, width=macadj(27, 22), justify='right') \
            .grid(row=0, column=5)
        Button(grvframe, text="Search", width=macadj(23, 23), anchor="center",
               command=lambda: self.search_grv_apply(self.win.topframe)).grid(row=1, column=5, pady=3)
        row += 1
        # separator
        separator = ttk.Separator(self.win.body, orient='horizontal')
        separator.grid(column=0, row=row, columnspan=6, sticky="nesw", pady=5)
        row += 1
        # -----------------------------------------------------------------------------------------------------grievant
        # grievant entry/ listbox / new frame
        grievantframe = Frame(self.win.body)
        grievantframe.grid(row=row, column=0, columnspan=6, sticky="w")
        Label(grievantframe, text="Search by grievant:", width=29, anchor="w") \
            .grid(row=0, column=0, columnspan=4, sticky="w")
        add_stringvar = StringVar(self.win.body)
        self.grvent.append(add_stringvar)
        # create entry/delete button widgets for grievant and add them to arrays
        grvent = Entry(grievantframe, textvariable=self.grvent[0], width=macadj(21, 16), justify='right')
        grvent.grid(row=0, column=4)
        self.grvent_entry.append(grvent)  # add this to an array of entry widgets for grievant
        del_ = Button(grievantframe, text="add", width=4, anchor=button_alignment,
                      command=lambda: self.add_grvent_field(grievantframe))
        del_.grid(row=0, column=5)
        self.grvent_del.append(del_)  # add this to an array of widgets of delete buttons
        # configure the show list button this will update location with self.add_grvent_field() command
        self.grvent_showlist = Button(grievantframe, text="show list", width=23, anchor="center",
                                      command=lambda: self.informalc_root("selectcarrier", childframe=grievantframe))
        self.grvent_showlist.grid(row=len(self.grvent)+1, column=4, columnspan=2, pady=3)
        Label(self.win.body, text=" ").grid(row=row, columnspan=6)
        row += 1
        # separator
        separator = ttk.Separator(self.win.body, orient='horizontal')
        separator.grid(column=0, row=row, columnspan=6, sticky="nesw", pady=5)
        row += 1
        # ----------------------------------------------------------------------------------------------- incident date
        incidentframe = Frame(self.win.body)
        incidentframe.grid(row=row, column=0, columnspan=6, sticky="w")

        def callback_incident(*args):
            """ watch the incident date option menu for changes using trace. """
            if args:  # do something with args to prevent an error with pycharm
                projvar.try_absorber = False
            if self.option_incidentdate.get() == "within specified range":
                hiddenafterlabel_incident.grid(row=1, column=4, sticky="w")
                hiddenbeforelabel_incident.grid(row=1, column=5, sticky="w")
                hiddenafterentry_incident.grid(row=2, column=4)
                hiddenbeforeentry_incident.grid(row=2, column=5)
            if self.option_incidentdate.get() != "within specified range":
                hiddenafterlabel_incident.grid_remove()
                hiddenbeforelabel_incident.grid_remove()
                hiddenafterentry_incident.grid_remove()
                hiddenbeforeentry_incident.grid_remove()
            # bind the expanding frome to the canvas and scrollregion
            incidentframe.bind('<Configure>', lambda e: self.win.c.configure(scrollregion=self.win.c.bbox("all")))
            self.win.topframe.bind("<Configure>", self.win.detect_resize)  # track when the window changes size
            projvar.root.update()
        Label(incidentframe, text="Search by incident date:", width=29, anchor="w") \
            .grid(row=0, column=0, columnspan=4, sticky="w")
        om = OptionMenu(incidentframe, self.option_incidentdate, *self.date_options)
        om.config(width=macadj(22, 20))
        om.grid(row=0, column=4, columnspan=2, sticky="e")
        self.option_incidentdate.set("include all")
        self.option_incidentdate.trace("w", callback_incident)
        hiddenafterlabel_incident = Label(incidentframe, text="After", fg="grey", justify="left")
        hiddenafterlabel_incident.grid_remove()
        hiddenbeforelabel_incident = Label(incidentframe, text="Before", fg="grey", justify="left")
        hiddenbeforelabel_incident.grid_remove()
        hiddenafterentry_incident = Entry(incidentframe, textvariable=self.incident_start,
                                          width=macadj(14, 8), justify='right')
        hiddenafterentry_incident.grid_remove()
        hiddenbeforeentry_incident = Entry(incidentframe, textvariable=self.incident_end,
                                           width=macadj(14, 8), justify='right')
        hiddenbeforeentry_incident.grid_remove()
        row += 1
        # separator
        separator = ttk.Separator(self.win.body, orient='horizontal')
        separator.grid(column=0, row=row, columnspan=6, sticky="nesw", pady=5)
        row += 1
        # ----------------------------------------------------------------------------------------------- meeting date
        meetingframe = Frame(self.win.body)
        meetingframe.grid(row=row, column=0, columnspan=6, sticky="w")

        def callback_meeting(*args):
            """ watch the meeting date option menu for changes using trace. """
            if args:  # do something with args to prevent an error with pycharm
                projvar.try_absorber = False
            if self.option_meetingdate.get() == "within specified range":
                hiddenafterlabel_meeting.grid(row=1, column=4, sticky="w")
                hiddenbeforelabel_meeting.grid(row=1, column=5, sticky="w")
                hiddenafterentry_meeting.grid(row=2, column=4)
                hiddenbeforeentry_meeting.grid(row=2, column=5)
            if self.option_meetingdate.get() != "within specified range":
                hiddenafterlabel_meeting.grid_remove()
                hiddenbeforelabel_meeting.grid_remove()
                hiddenafterentry_meeting.grid_remove()
                hiddenbeforeentry_meeting.grid_remove()
            # bind the expanding frome to the canvas and scrollregion
            meetingframe.bind('<Configure>', lambda e: self.win.c.configure(scrollregion=self.win.c.bbox("all")))
            self.win.topframe.bind("<Configure>", self.win.detect_resize)  # track when the window changes size
            projvar.root.update()
        Label(meetingframe, text="Search by meeting date:", width=29, anchor="w") \
            .grid(row=0, column=0, columnspan=4, sticky="w")
        om = OptionMenu(meetingframe, self.option_meetingdate, *self.date_options)
        om.config(width=macadj(22, 20))
        om.grid(row=0, column=4, columnspan=2, sticky="e")
        self.option_meetingdate.set("include all")
        self.option_meetingdate.trace("w", callback_meeting)
        hiddenafterlabel_meeting = Label(meetingframe, text="After", fg="grey", justify="left")
        hiddenafterlabel_meeting.grid_remove()
        hiddenbeforelabel_meeting = Label(meetingframe, text="Before", fg="grey", justify="left")
        hiddenbeforelabel_meeting.grid_remove()
        hiddenafterentry_meeting = Entry(meetingframe, textvariable=self.meeting_start,
                                         width=macadj(14, 8), justify='right')
        hiddenafterentry_meeting.grid_remove()
        hiddenbeforeentry_meeting = Entry(meetingframe, textvariable=self.meeting_end,
                                          width=macadj(14, 8), justify='right')
        hiddenbeforeentry_meeting.grid_remove()
        row += 1
        # separator
        separator = ttk.Separator(self.win.body, orient='horizontal')
        separator.grid(column=0, row=row, columnspan=6, sticky="nesw", pady=5)
        row += 1
        # ------------------------------------------------------------------------------------------------------- issue
        # issue entry/ listbox / new frame
        issueframe = Frame(self.win.body)
        issueframe.grid(row=row, column=0, columnspan=6, sticky="w")
        Label(issueframe, text="Search by issue:", width=29, anchor="w") \
            .grid(row=0, column=0, columnspan=4, sticky="w")
        add_stringvar = StringVar(self.win.body)
        self.src_issue.append(add_stringvar)
        # create entry/delete button widgets for issue and add them to arrays
        src_issue = Entry(issueframe, textvariable=self.src_issue[0], width=macadj(21, 16), justify='right')
        src_issue.grid(row=0, column=4)
        self.src_issue_entry.append(src_issue)  # add this to an array of entry widgets for issue
        del_ = Button(issueframe, text="add", width=4, anchor=button_alignment,
                      command=lambda: self.add_src_issue_field(issueframe))
        del_.grid(row=0, column=5)
        self.src_issue_del.append(del_)  # add this to an array of widgets of delete buttons
        # configure the show list button this will update location with self.add_src_issue_field() command
        self.src_issue_showlist = Button(issueframe, text="show list", width=23, anchor="center",
                                         command=lambda: self.informalc_root("selectissue", childframe=issueframe))
        self.src_issue_showlist.grid(row=len(self.src_issue) + 1, column=4, columnspan=2, pady=3)
        Label(self.win.body, text=" ").grid(row=row, columnspan=6)
        row += 1
        # separator
        separator = ttk.Separator(self.win.body, orient='horizontal')
        separator.grid(column=0, row=row, columnspan=6, sticky="nesw", pady=5)
        row += 1
        # ----------------------------------------------------------------------------------------------------- level
        levelframe = Frame(self.win.body)
        levelframe.grid(row=row, column=0, columnspan=6, sticky="w")

        def callback_level(*args):
            """ watch the level option menu for changes using trace. """
            if args:  # do something with args to prevent an error with pycharm
                projvar.try_absorber = False
            if self.option_level.get() == "selection":
                self.level_listbox.grid(row=1, column=4, columnspan=2, sticky="w")
            if self.option_level.get() == "include all":
                self.level_listbox.selection_clear(0, 'end')
                self.level_listbox.grid_remove()
            # bind the expanding frome to the canvas and scrollregion
            levelframe.bind('<Configure>', lambda e: self.win.c.configure(scrollregion=self.win.c.bbox("all")))
            self.win.topframe.bind("<Configure>", self.win.detect_resize)  # track when the window changes size
            projvar.root.update()

        Label(levelframe, text="Search by level:", width=29, anchor="w") \
            .grid(row=0, column=0, columnspan=4, sticky="w")
        om = OptionMenu(levelframe, self.option_level, "include all", "selection")
        om.config(width=macadj(22, 20))
        om.grid(row=0, column=4, columnspan=2, sticky="e")
        self.option_level.set("include all")
        self.option_level.trace("w", callback_level)
        self.level_listbox = Listbox(levelframe, height=len(self.level_options), width=macadj(28, 23),
                                     selectmode="multiple", exportselection=False)
        self.level_listbox.grid_remove()
        for i in range(len(self.level_options)):
            self.level_listbox.insert(i, self.level_options[i])
        row += 1
        # separator
        separator = ttk.Separator(self.win.body, orient='horizontal')
        separator.grid(column=0, row=row, columnspan=6, sticky="nesw", pady=5)
        row += 1
        # ----------------------------------------------------------------------------------------------- signed date
        signedframe = Frame(self.win.body)
        signedframe.grid(row=row, column=0, columnspan=6, sticky="w")

        def callback_signed(*args):
            """ watch the signed date option menu for changes using trace. """
            if args:  # do something with args to prevent an error with pycharm
                projvar.try_absorber = False
            if self.option_signeddate.get() == "within specified range":
                hiddenafterlabel_signed.grid(row=1, column=4, sticky="w")
                hiddenbeforelabel_signed.grid(row=1, column=5, sticky="w")
                hiddenafterentry_signed.grid(row=2, column=4)
                hiddenbeforeentry_signed.grid(row=2, column=5)
            if self.option_signeddate.get() != "within specified range":
                hiddenafterlabel_signed.grid_remove()
                hiddenbeforelabel_signed.grid_remove()
                hiddenafterentry_signed.grid_remove()
                hiddenbeforeentry_signed.grid_remove()
            # bind the expanding frome to the canvas and scrollregion
            signedframe.bind('<Configure>', lambda e: self.win.c.configure(scrollregion=self.win.c.bbox("all")))
            self.win.topframe.bind("<Configure>", self.win.detect_resize)  # track when the window changes size
            projvar.root.update()

        Label(signedframe, text="Search by signed date:", width=29, anchor="w") \
            .grid(row=0, column=0, columnspan=4, sticky="w")
        om = OptionMenu(signedframe, self.option_signeddate, *self.date_options)
        om.config(width=macadj(22, 20))
        om.grid(row=0, column=4, columnspan=2, sticky="e")
        self.option_signeddate.set("include all")
        self.option_signeddate.trace("w", callback_signed)
        hiddenafterlabel_signed = Label(signedframe, text="After", fg="grey", justify="left")
        hiddenafterlabel_signed.grid_remove()
        hiddenbeforelabel_signed = Label(signedframe, text="Before", fg="grey", justify="left")
        hiddenbeforelabel_signed.grid_remove()
        hiddenafterentry_signed = Entry(signedframe, textvariable=self.signed_start,
                                        width=macadj(14, 8), justify='right')
        hiddenafterentry_signed.grid_remove()
        hiddenbeforeentry_signed = Entry(signedframe, textvariable=self.signed_end,
                                         width=macadj(14, 8), justify='right')
        hiddenbeforeentry_signed.grid_remove()
        row += 1
        # separator
        separator = ttk.Separator(self.win.body, orient='horizontal')
        separator.grid(column=0, row=row, columnspan=6, sticky="nesw", pady=5)
        row += 1
        # -------------------------------------------------------------------------------------------------- decision
        # decision entry/ listbox / new frame
        decisionframe = Frame(self.win.body)
        decisionframe.grid(row=row, column=0, columnspan=6, sticky="w")
        Label(decisionframe, text="Search by decision:", width=29, anchor="w") \
            .grid(row=0, column=0, columnspan=4, sticky="w")
        add_stringvar = StringVar(self.win.body)
        self.decision.append(add_stringvar)
        # create entry/delete button widgets for decision and add them to arrays
        decision = Entry(decisionframe, textvariable=self.decision[0], width=macadj(21, 16), justify='right')
        decision.grid(row=0, column=4)
        self.decision_entry.append(decision)  # add this to an array of entry widgets for decision
        del_ = Button(decisionframe, text="add", width=4, anchor=button_alignment,
                      command=lambda: self.add_decision_field(decisionframe))
        del_.grid(row=0, column=5)
        self.decision_del.append(del_)  # add this to an array of widgets of delete buttons
        # configure the show list button this will update location with self.add_decision_field() command
        self.decision_showlist = Button(decisionframe, text="show list", width=23, anchor="center",
                                        command=lambda: self.informalc_root("selectdecision", childframe=decisionframe))
        self.decision_showlist.grid(row=len(self.decision)+1, column=4, columnspan=2, pady=3)
        Label(self.win.body, text=" ").grid(row=row, columnspan=6)
        row += 1
        # separator
        separator = ttk.Separator(self.win.body, orient='horizontal')
        separator.grid(column=0, row=row, columnspan=6, sticky="nesw", pady=5)
        row += 1

        # ----------------------------------------------------------------------------------------------- proofdue date
        proofdueframe = Frame(self.win.body)
        proofdueframe.grid(row=row, column=0, columnspan=6, sticky="w")

        def callback_proofdue(*args):
            """ watch the proofdue date option menu for changes using trace. """
            if args:  # do something with args to prevent an error with pycharm
                projvar.try_absorber = False
            if self.option_proofduedate.get() == "within specified range":
                hiddenafterlabel_proofdue.grid(row=1, column=4, sticky="w")
                hiddenbeforelabel_proofdue.grid(row=1, column=5, sticky="w")
                hiddenafterentry_proofdue.grid(row=2, column=4)
                hiddenbeforeentry_proofdue.grid(row=2, column=5)
            if self.option_proofduedate.get() != "within specified range":
                hiddenafterlabel_proofdue.grid_remove()
                hiddenbeforelabel_proofdue.grid_remove()
                hiddenafterentry_proofdue.grid_remove()
                hiddenbeforeentry_proofdue.grid_remove()
            # bind the expanding frome to the canvas and scrollregion
            proofdueframe.bind('<Configure>', lambda e: self.win.c.configure(scrollregion=self.win.c.bbox("all")))
            self.win.topframe.bind("<Configure>", self.win.detect_resize)  # track when the window changes size
            projvar.root.update()

        Label(proofdueframe, text="Search by proofdue date:", width=29, anchor="w") \
            .grid(row=0, column=0, columnspan=4, sticky="w")
        om = OptionMenu(proofdueframe, self.option_proofduedate, *self.date_options)
        om.config(width=macadj(22, 20))
        om.grid(row=0, column=4, columnspan=2, sticky="e")
        self.option_proofduedate.set("include all")
        self.option_proofduedate.trace("w", callback_proofdue)
        hiddenafterlabel_proofdue = Label(proofdueframe, text="After", fg="grey", justify="left")
        hiddenafterlabel_proofdue.grid_remove()
        hiddenbeforelabel_proofdue = Label(proofdueframe, text="Before", fg="grey", justify="left")
        hiddenbeforelabel_proofdue.grid_remove()
        hiddenafterentry_proofdue = Entry(proofdueframe, textvariable=self.proofdue_start,
                                          width=macadj(14, 8), justify='right')
        hiddenafterentry_proofdue.grid_remove()
        hiddenbeforeentry_proofdue = Entry(proofdueframe, textvariable=self.proofdue_end,
                                           width=macadj(14, 8), justify='right')
        hiddenbeforeentry_proofdue.grid_remove()
        row += 1
        # separator
        separator = ttk.Separator(self.win.body, orient='horizontal')
        separator.grid(column=0, row=row, columnspan=6, sticky="nesw", pady=5)
        row += 1
        # ----------------------------------------------------------------------------------------------------- docs
        docsframe = Frame(self.win.body)
        docsframe.grid(row=row, column=0, columnspan=6, sticky="w")

        def callback_docs(*args):
            """ watch the docs option menu for changes using trace. """
            if args:  # do something with args to prevent an error with pycharm
                projvar.try_absorber = False
            if self.option_docs.get() == "selection":
                self.docs_listbox.grid(row=1, column=4, columnspan=2, sticky="w")
            if self.option_docs.get() == "include all":
                self.docs_listbox.selection_clear(0, 'end')
                self.docs_listbox.grid_remove()
            # bind the expanding frome to the canvas and scrollregion
            docsframe.bind('<Configure>', lambda e: self.win.c.configure(scrollregion=self.win.c.bbox("all")))
            self.win.topframe.bind("<Configure>", self.win.detect_resize)  # track when the window changes size
            projvar.root.update()

        Label(docsframe, text="Search by docs:", width=29, anchor="w") \
            .grid(row=0, column=0, columnspan=4, sticky="w")
        om = OptionMenu(docsframe, self.option_docs, "include all", "selection")
        om.config(width=macadj(22, 20))
        om.grid(row=0, column=4, columnspan=2, sticky="e")
        self.option_docs.set("include all")
        self.option_docs.trace("w", callback_docs)
        self.docs_listbox = Listbox(docsframe, height=len(self.doc_options), width=macadj(28, 23),
                                    selectmode="multiple", exportselection=False)
        self.docs_listbox.grid_remove()
        for i in range(len(self.doc_options)):
            self.docs_listbox.insert(i, self.doc_options[i])
        row += 1
        # separator
        separator = ttk.Separator(self.win.body, orient='horizontal')
        separator.grid(column=0, row=row, columnspan=6, sticky="nesw", pady=5)
        row += 1
        # ----------------------------------------------------------------------------------------------------- gats
        gatframe = Frame(self.win.body)
        gatframe.grid(row=row, column=0, columnspan=6, sticky="w")
        Label(gatframe, text="Search by gats:", width=29, anchor="w") \
            .grid(row=0, column=0, columnspan=4, sticky="w")
        om = OptionMenu(gatframe, self.gats, "include all", "yes", "no")
        om.config(width=macadj(22, 20))
        om.grid(row=0, column=4, columnspan=2, sticky="e")
        self.gats.set("include all")
        row += 1
        # separator
        separator = ttk.Separator(self.win.body, orient='horizontal')
        separator.grid(column=0, row=row, columnspan=6, sticky="nesw", pady=5)
        row += 1
        # --------------------------------------------------------------------------------------------------- sort by
        sortframe = Frame(self.win.body)
        sortframe.grid(row=row, column=0, columnspan=6, sticky="w")
        Label(sortframe, text="Sort by:", width=29, anchor="w") \
            .grid(row=0, column=0, columnspan=4, sticky="w")
        radiotext = ["Start Incident Date", "End Incident Date", "Meeting Date", "Signed Date", "Proof Due"]
        for i in range(5):
            r = Radiobutton(sortframe, text=radiotext[i], variable=self.sortby, value=i)
            r.grid(row=i, column=4, columnspan=2, sticky="w")
        self.sortby.set(2)
        row += 1
        # separator
        separator = ttk.Separator(self.win.body, orient='horizontal')
        separator.grid(column=0, row=row, columnspan=6, sticky="nesw", pady=5)
        row += 1
        # ------------------------------------------------------------------------------------------------- sort order
        sortorderframe = Frame(self.win.body)
        sortorderframe.grid(row=row, column=0, columnspan=6, sticky="w")
        Label(sortorderframe, text="Sort Order:", width=29, anchor="w") \
            .grid(row=0, column=0, columnspan=4, sticky="w")
        radiotext = ["Earliest to Recent", "Recent to Earliest"]
        for i in range(2):
            r = Radiobutton(sortorderframe, text=radiotext[i], variable=self.sort_order, value=i)
            r.grid(row=i, column=4, columnspan=2, sticky="w")
        self.sort_order.set(1)
        # ------------------------------------------------------------------------------- buttons bottom of the screen
        button_alignment = macadj("w", "center")
        Button(self.win.buttons, text="Go Back", width=macadj(20, 27), anchor=button_alignment,
               command=lambda: (self.destroy_companion(), self.informalc(self.win.topframe))).grid(row=0, column=0)
        Button(self.win.buttons, text="Search", width=macadj(20, 26), anchor=button_alignment,
               command=lambda: (self.destroy_companion(), self.search_apply(self.win.topframe))).grid(row=0, column=1)

    def add_grvent_field(self, childframe, carrier=None):
        """ added fields for search criteria - grievant"""
        add_stringvar = StringVar(self.win.body)
        if carrier:
            add_stringvar.set(carrier)
        self.grvent.append(add_stringvar)  # add this to an array of stringvars for non compliance
        grvent = Entry(childframe, textvariable=self.grvent[len(self.grvent)-1], justify='right', width=macadj(20, 15))
        grvent.grid(row=len(self.grvent)-1, column=4, sticky="w")
        self.grvent_entry.append(grvent)  # add this to an array of entry widgets for non compliance
        del_ = Button(childframe, text="del", width=macadj(4, 3), anchor="center",
                      command=lambda x=len(self.grvent)-1: self.del_grvent_field(x))
        del_.grid(row=len(self.grvent)-1, column=5)
        self.grvent_del.append(del_)  # add this to an array of widgets of delete buttons
        self.grvent_showlist.grid(row=len(self.grvent) + 1, column=4)
        projvar.root.update()
        # bind the expanding frome to the canvas and scrollregion
        childframe.bind('<Configure>', lambda e: self.win.c.configure(scrollregion=self.win.c.bbox("all")))
        self.win.topframe.bind("<Configure>", self.win.detect_resize)  # track when the window changes size

    def del_grvent_field(self, x):
        """ delete a field from search criteria - grievant, entry widgets as well as the delete button.
        set the value of the corresponding stringvar to an empty string. """
        self.grvent_entry[x].grid_remove()
        self.grvent_del[x].grid_remove()
        self.grvent[x].set("")  # set the value of the stringvar to empty string

    def add_src_issue_field(self, childframe, issue=None):
        """ added fields for search criteria - issue"""
        add_stringvar = StringVar(self.win.body)
        if issue:
            add_stringvar.set(issue)
        self.src_issue.append(add_stringvar)  # add this to an array of stringvars for non compliance
        src_issue = Entry(childframe, textvariable=self.src_issue[len(self.src_issue) - 1],
                          justify='right', width=macadj(20, 15))
        src_issue.grid(row=len(self.src_issue)-1, column=4, sticky="w")
        self.src_issue_entry.append(src_issue)  # add this to an array of entry widgets for non compliance
        del_ = Button(childframe, text="del", width=macadj(4, 3), anchor="center",
                      command=lambda x=len(self.src_issue)-1: self.del_src_issue_field(x))
        del_.grid(row=len(self.src_issue)-1, column=5)
        self.src_issue_del.append(del_)  # add this to an array of widgets of delete buttons
        self.src_issue_showlist.grid(row=len(self.src_issue) + 1, column=4)
        projvar.root.update()
        # bind the expanding frome to the canvas and scrollregion
        childframe.bind('<Configure>', lambda e: self.win.c.configure(scrollregion=self.win.c.bbox("all")))
        self.win.topframe.bind("<Configure>", self.win.detect_resize)  # track when the window changes size

    def del_src_issue_field(self, x):
        """ delete a field from the search criteria - issue, entry widgets as well as the delete button.
        set the value of the corresponding stringvar to an empty string. """
        self.src_issue_entry[x].grid_remove()
        self.src_issue_del[x].grid_remove()
        self.src_issue[x].set("")  # set the value of the stringvar to empty string

    def add_decision_field(self, childframe, decision_1=None):
        """ added fields for search criteria - decision"""
        add_stringvar = StringVar(self.win.body)
        if decision_1:
            add_stringvar.set(decision_1)
        self.decision.append(add_stringvar)  # add this to an array of stringvars for non compliance
        decision = Entry(childframe, textvariable=self.decision[len(self.decision)-1], justify='right',
                         width=macadj(20, 15))
        decision.grid(row=len(self.decision)-1, column=4, sticky="w")
        self.decision_entry.append(decision)  # add this to an array of entry widgets for non compliance
        del_ = Button(childframe, text="del", width=macadj(4, 3), anchor="center",
                      command=lambda x=len(self.decision)-1: self.del_decision_field(x))
        del_.grid(row=len(self.decision)-1, column=5)
        self.decision_del.append(del_)  # add this to an array of widgets of delete buttons
        self.decision_showlist.grid(row=len(self.decision) + 1, column=4)
        projvar.root.update()
        # bind the expanding frome to the canvas and scrollregion
        childframe.bind('<Configure>', lambda e: self.win.c.configure(scrollregion=self.win.c.bbox("all")))
        self.win.topframe.bind("<Configure>", self.win.detect_resize)  # track when the window changes size

    def del_decision_field(self, x):
        """ delete a field from search criteria - decision, entry widgets as well as the delete button.
        set the value of the corresponding stringvar to an empty string. """
        self.decision_entry[x].grid_remove()
        self.decision_del[x].grid_remove()
        self.decision[x].set("")  # set the value of the stringvar to empty string

    def search_apply(self, frame):
        """
        gather all stringvars needed from build_search_screen().
        conduct checks on all input - return if there are any errors.
        generate sql for the db search and store it in case it is needed again.
        commit the sql search and store the results.
        go to the search results screen.
        """
        #  statement builders for grievance sql
        grvent_sql = ""  # initialize the sql statement builder
        incident_sql = ""  # initialize the sql statement builder
        meeting_sql = ""  # initialize the sql statement builder
        issue_sql = ""  # initialize the sql statement builder
        # statement builders for settlement sql
        level_sql = ""  # initialize the sql statment
        signed_sql = ""  # initialize the sql statement builder
        decision_sql = ""  # initialize the sql statement builder
        proofdue_sql = ""  # initialize the sql statement builder
        docs_sql = ""  # initialize the sql statement builder
        gats_sql = ""  # initialize the sql statement builder
        # --------------------------------------------------------------------------------------------------- grievant
        grv_array = []
        for grvent in self.grvent:  # loop for all elements in list
            grvent = grvent.get()  # get the value from the stringvar
            grvent = grvent.lower().strip()
            if grvent and grvent not in grv_array:  # if it is not empty and not a duplicate...
                grv_array.append(grvent)  # put it in the array
        if grv_array:
            grvent_sql = "("  # start the sql statment
            i = 0
            for grvent in grv_array:
                if not NameChecker(grvent).check_characters():
                    msg = "Grievant name can not contain numbers or most special characters\n"
                    messagebox.showerror("Invalid Data Entry", msg, parent=self.win.topframe)
                    return
                grvent_sql += "grievant = '{}'".format(grvent)
                # if the list has more than one and instance is not last
                if len(grv_array) > 1 and i + 1 < len(grv_array):
                    grvent_sql += " OR "  # add 'or' to the statement
                i += 1
            grvent_sql += ")"
            self.blank_criteria = False  # if all criteria are blank, don't include 'where' in sql
        # ------------------------------------------------------------------------------------------------------ dates
        now = datetime.now().date()  # get the current date as a datetime object
        sixmonthsago = now - timedelta(weeks=26)  # datetime from six months ago
        oneyearago = now - timedelta(weeks=52)  # datetime from six months ago
        twoyearsago = now - timedelta(weeks=52*2)  # datetime from six months ago
        threeyearsago = now - timedelta(weeks=52*3)  # datetime from six months ago
        default_start = datetime(1000, 1, 1)  # default start is jan 1, 1000 AD
        default_end = datetime(9000, 1, 1)  # default end is jan 1 9000 AD
        date_types = ["incident", "meeting", "signed", "proofdue"]
        # a string of descriptions for the criteria for use in error messages
        dates_names = [
            ["incident start date", "incident end date"],
            ["meeting start date", "meeting end date"],
            ["signed start date", "signed end date"],
            ["proof due start date", "proof due end date"]
        ]
        # put the values of the date stringvars for the search criteria in a list of four pairs
        dates_input = (
            (self.incident_start.get(), self.incident_end.get()),
            (self.meeting_start.get(), self.meeting_end.get()),
            (self.signed_start.get(), self.signed_end.get()),
            (self.proofdue_start.get(), self.proofdue_end.get())
        )
        # put the values of the date option menus for the search criteria in a list of four pairs
        date_options_input = [self.option_incidentdate.get(), self.option_meetingdate.get(),
                              self.option_signeddate.get(), self.option_proofduedate.get()]
        for i in range(len(dates_input)):  # will loop four times
            start = default_start
            end = default_end
            for ii in range(2):  # loop twice for each pair in dates input
                if date_options_input[i] == "within specified range":
                    if not dates_input[i][ii]:  # if the date was left empty
                        if not ii & 1:  # if ii is an even number, it is the first of the pair
                            start = default_start  # put in default start
                        else:  # if ii is an odd number, it is the second of the pair
                            end = default_end  # put in default end
                    else:  # if the user put something in the date field
                        if not informalc_date_checker(frame, dates_input[i][ii], dates_names[i][ii]):  # run check
                            return
                        date_split = dates_input[i][ii].split("/")
                        # convert into a datetime object
                        if not ii & 1 and dates_input[i][ii]:  # if ii is an even number, it is the first of the pair
                            start = datetime(int(date_split[2]), int(date_split[0]), int(date_split[1]))
                        if ii & 1 and dates_input[i][ii]:  # if ii is an odd number, it is the second of the pair
                            end = datetime(int(date_split[2]), int(date_split[0]), int(date_split[1]))
                    if start > end:  # check that start date comes before end date
                        messagebox.showerror("Invalid Data Entry",
                                             "Your starting incident date must be earlier than your "
                                             "ending incident date.",
                                             parent=frame)
                        return
            # define the start and end dates for the search criteria
            if date_options_input[i] == "within last 6 months":
                start = sixmonthsago
                end = default_end
            if date_options_input[i] == "within last year":
                start = oneyearago
                end = default_end
            if date_options_input[i] == "within last two years":
                start = twoyearsago
                end = default_end
            if date_options_input[i] == "within last three years":
                start = threeyearsago
                end = default_end
            if date_options_input[i] != "include all":
                self.blank_criteria = False  # if all criteria are blank, don't include 'where' in sql
                # write 'where' statement for incident date sql - incident dates get different treatment
                if date_types[i] == "incident":
                    if date_options_input[i] in ("within last 6 months", "within last year", "within last two years",
                                                 "within last three years"):
                        incident_sql = "(startdate BETWEEN '{}' AND '{}')".format(start, end)
                    if date_options_input[i] == "within specified range":
                        incident_sql = "('{}' <= enddate AND startdate <= '{}'  )".format(start, end)
                # write 'where' sql statement for meeting, signed or proof due
                elif date_types[i] == "meeting":
                    meeting_sql = "(meetingdate BETWEEN '{}' AND '{}')".format(start, end)
                elif date_types[i] == "signed":
                    signed_sql = "(date_signed BETWEEN '{}' AND '{}')".format(start, end)
                elif date_types[i] == "proofdue":
                    proofdue_sql = "(proofdue BETWEEN '{}' AND '{}')".format(start, end)
        # ------------------------------------------------------------------------------------------------------ issue
        issue_array = []
        for issue in self.src_issue:  # loop for all elements in list
            issue = issue.get()  # get the value from the stringvar
            issue = issue.lower().strip()
            if issue and issue not in issue_array:  # if it is not empty and not a duplicate...
                issue_array.append(issue)  # put it in the array
        if issue_array:  # if not empty, then write the sql 'where' statement
            issue_sql = "("  # start the sql statment
            i = 0
            for issue in issue_array:
                issue_sql += "issue = '{}'".format(issue)
                # if the list has more than one and instance is not last
                if len(issue_array) > 1 and i + 1 < len(issue_array):
                    issue_sql += " OR "  # add 'or' to the statement
                i += 1
            issue_sql += ")"
            self.blank_criteria = False  # if all criteria are blank, don't include 'where' in sql
        # ----------------------------------------------------------------------------------------------------- level
        if self.option_level.get() == "selection":  # shows all options in list
            level_selections = []
            for index in self.level_listbox.curselection():
                level_selections.append(self.level_options[index])
            if level_selections:  # if not empty, write the sql 'where' statement
                level_sql = "("  # start the sql statment
                i = 0
                for level in level_selections:
                    level_sql += "level = '{}'".format(level)
                    # if the list has more than one and instance is not last
                    if len(level_selections) > 1 and i + 1 < len(level_selections):
                        level_sql += " OR "  # add 'or' to the statement
                    i += 1
                level_sql += ")"
                self.blank_criteria = False  # if all criteria are blank, don't include 'where' in sql
        # ---------------------------------------------------------------------------------------------------- decision
        decision_array = []
        for decision in self.decision:  # loop for all elements in list
            decision = decision.get()  # get the value from the stringvar
            decision = decision.lower().strip()
            if decision and decision not in decision_array:  # if it is not empty and not a duplicate...
                decision_array.append(decision)  # put it in the array
        if decision_array:  # if not empty, write the sql 'where' statement
            decision_sql = "("  # start the sql statment
            i = 0
            for decision in decision_array:
                decision_sql += "decision = '{}'".format(decision)
                # if the list has more than one and instance is not last
                if len(decision_array) > 1 and i + 1 < len(decision_array):
                    decision_sql += " OR "  # add 'or' to the statement
                i += 1
            decision_sql += ")"
            self.blank_criteria = False  # if all criteria are blank, don't include 'where' in sql
        # ------------------------------------------------------------------------------------------------------- docs
        if self.option_docs.get() == "selection":  # shows all options in list
            docs_selections = []
            for index in self.docs_listbox.curselection():
                if self.doc_options[index] == 'no status':  # replace 'no status' with an empty string
                    docs_selections.append('')
                else:
                    docs_selections.append(self.doc_options[index])
            if docs_selections:  # if not empty, write the sql 'where' statement
                docs_sql = "("  # start the sql statment
                i = 0
                for docs in docs_selections:
                    docs_sql += "docs = '{}'".format(docs)
                    # if the list has more than one and instance is not last
                    if len(docs_selections) > 1 and i + 1 < len(docs_selections):
                        docs_sql += " OR "  # add 'or' to the statement
                    i += 1
                docs_sql += ")"
                self.blank_criteria = False  # if all criteria are blank, don't include 'where' in sql
        # --------------------------------------------------------------------------------------------------- gats sql
        # Since gats information comes from a different table i.e. the informalc_gats table, a different
        # process is used. A list of grievance numbers with gats numbers is built. Later that list is either a filter
        # for the search results ('yes' - selected) or the list is removed from the search results ('no' - selected).
        if self.gats.get() != "include all":  # if not empty (aka 'include all') then get gats search results.
            result = inquire(self.gats_sql)  # inquire from informalc_gats table
            if result:  # if the result is not empty
                for r in result:  # build a list...
                    self.search_gat_result.append(r[0])  # of grievances with gats numbers

        # ----------------------------------------------------------------------------------------------- grievance sql
        where = ""  # initialize the sql statement builder
        where_array = []
        for sql in (grvent_sql, incident_sql, meeting_sql, issue_sql):  # find criteria that is not empty
            if sql:
                where_array.append(sql)
        i = 0
        for array in where_array:  # using criteria that is not empty
            where += array  # add that criteria i.e. "grievant - 'weeks, t'"
            if len(where_array) > 1 and i + 1 < len(where_array):  # if more than one and not last
                where += " AND "  # insert 'AND' at the end
            i += 1
        if where:  # running a search with an empty search criteria will cause an error
            self.grv_sql = "SELECT DISTINCT grv_no FROM informalc_grievances WHERE {}".format(where)
            self.search_grv_result = inquire(self.grv_sql)
        else:
            self.search_grv_result = []
        # ----------------------------------------------------------------------------------------- ---- settlement sql
        where = ""  # initialize the sql statement builder
        where_array = []
        for sql in (level_sql, signed_sql, decision_sql, proofdue_sql, docs_sql, gats_sql):
            if sql:
                where_array.append(sql)
        i = 0
        for array in where_array:
            where += array
            if len(where_array) > 1 and i + 1 < len(where_array):
                where += " AND "
            i += 1
        if where:  # running a search with an empty search criteria will cause an error
            self.set_sql = "SELECT DISTINCT grv_no FROM informalc_settlements WHERE {}".format(where)
            self.search_set_result = inquire(self.set_sql)
        # ----------------------------------------------------------------------------------------- no search criteria
        if self.blank_criteria:  # if no search critera was given,
            self.search_all_apply(frame)  # go to search all
        else:
            if not self.search_grv_result and not self.search_set_result:  # no results
                msg = "There is no record for any grievances in the database matching the search criteria."
                messagebox.showerror("Records Not Found", msg, parent=self.win.topframe)
                return
            self.merge_search_results(frame)

    def search_grv_apply(self, frame):
        """ search for the grievance number from self.build_search_screen() """
        def check_grievance_number():
            """ check the grievance number for empty value or invalid characters. """
            if not GrievanceChecker(grievance_number).has_value():
                msgg = "The grievance number must not be blank."
                messagebox.showerror("Invalid Data Entry", msgg, parent=self.win.topframe)
                return False
            if not GrievanceChecker(grievance_number).check_characters():
                msgg = "The grievance number can only contain numbers and letters."
                messagebox.showerror("Invalid Data Entry", msgg, parent=self.win.topframe)
                return False
            return True
        grievance_number = self.src_grievance.get()
        if not check_grievance_number():
            return
        self.grv_sql = "SELECT DISTINCT grv_no FROM informalc_grievances WHERE grv_no = '%s' and station = '%s'" % \
                       (grievance_number, self.station)
        self.search_grv_result = inquire(self.grv_sql)
        if not self.search_grv_result:
            msg = "There is no record for this grievance in the database: {}".format(grievance_number)
            messagebox.showerror("Record Not Found", msg, parent=self.win.topframe)
            return
        self.merge_search_results(frame)

    def search_all_apply(self, frame):
        """ search for all grievances in the station from self.build_search_screen() """
        self.grv_sql = "SELECT DISTINCT grv_no FROM informalc_grievances WHERE station = '%s'" % self.station
        self.search_grv_result = inquire(self.grv_sql)
        self.set_sql = "SELECT DISTINCT grv_no FROM informalc_settlements"
        self.search_set_result = inquire(self.set_sql)
        if not self.search_grv_result and not self.search_set_result:
            msg = "There is no record for any grievances in the database"
            messagebox.showerror("Records Not Found", msg, parent=self.win.topframe)
            return
        self.merge_search_results(frame)

    def refresh_search(self, frame, reroute=True):
        """ when changes are made the the db, it is necessary to update the search results, as some things
        might have changed."""
        self.search_result = []  # empty and reinitialize search results
        self.search_grv_result = inquire(self.grv_sql)
        self.search_set_result = inquire(self.set_sql)
        self.search_gat_result = []  # initialize the search gat results array
        result = inquire(self.gats_sql)  # inquire from informalc_gats table
        if result:
            for r in result:
                self.search_gat_result.append(r[0])  # build list of grievances with gats numbers
        if not self.search_grv_result and not self.search_set_result:
            if reroute:
                msg = "There is no record for any grievances in the database matching the search criteria. "
                messagebox.showerror("Records Not Found", msg, parent=frame)
                self.master_search(frame)  # return to the search criteria screen
        else:
            self.merge_search_results(frame, showtime=False)

    def join_list(self, grvrecs, setrecs):
        """ this method works for the merge search results to combine the recs from the grievance table and
        recs from the settlement table. """
        if self.blank_criteria:  # if there is no search criteria
            return grvrecs  # return all grievance records
        elif not grvrecs and not setrecs:  # if there are no grv records not settlement records
            return []
        elif grvrecs and not setrecs:  # if there are grv recs but not set recs
            return grvrecs
        elif setrecs and not grvrecs:  # if there are set recs but no grv recs
            return setrecs
        else:  # if there are both grv AND set recs, make joint array where elements are in both arrays.
            return [x for x in grvrecs if x in setrecs]

    def merge_search_results(self, frame, showtime=True):
        """ search results for grievances and settlements need to be combined to show grievance recs and
        settlement recs as one record. showtime=False is passed from self.refresh_search,
        so that self.showtime() isn't run .
        the search_result index is formatted as follows:
                    [0]grievant - grievance records
                    [1]station
                    [2] grv_no
                    [3] startdate
                    [4] enddate
                    [5] meetingdate
                    [6] issue
                    [7] article
                    [8] grv_no - settlement records
                    [9] level
                    [10] date_signed
                    [11] decision (change to 'in batch') if a part of a batch settlement
                    [12] proofdue
                    [13] docs """
        grvrecs = []
        setrecs = []
        if self.search_grv_result:
            for grv in self.search_grv_result:
                grvrecs.append(grv[0])
        if self.search_set_result:
            for sett in self.search_set_result:
                setrecs.append(sett[0])
        # merge both list of distinct grievence/settlement grv numbers into one joint list
        joint = self.join_list(grvrecs, setrecs)
        if self.gats.get() == "yes":  # if self.gats is 'yes' revise list to show only those with gats reports
            joint = [x for x in joint if x in self.search_gat_result]
        if self.gats.get() == "no":  # if self.gats is 'no' revise list to show only those with no gats reports
            joint = [x for x in joint if x not in self.search_gat_result]
        for number in joint:  # for each number search grievance and settlement tables
            sql = "SELECT * FROM informalc_grievances WHERE grv_no = '%s'" % number
            results_raw = inquire(sql)
            results = [list(x) for x in results_raw]
            default_set = ['', '', '', '', '', '', '']  # # if there is no settlement record, use this default
            sql = "SELECT * FROM informalc_settlements WHERE grv_no = '%s'" % number
            results_raw = inquire(sql)
            set_results = [list(x) for x in results_raw]
            if set_results:  # merge the records of those searches into one record then add it to search results.
                self.search_result.append(results[0] + set_results[0])
            else:  # if there are is no record for a settlement.
                sql = "SELECT main FROM informalc_batchindex WHERE sub = '%s'" % number  # search if batch settlement
                results_raw = inquire(sql)
                batch_results = [list(x) for x in results_raw]
                if batch_results:  # if there is a record in the batch index
                    batch_set = ['', '', '', 'in batch: ' + str(batch_results[0][0]), '', '', '']  # if there is a rec
                    self.search_result.append(results[0] + batch_set)
                else:  # if there is no record in the batch index, go with empty strings
                    self.search_result.append(results[0] + default_set)
        sortby = (3, 4, 5, 10, 12)  # startdate: 3, enddate: 4, meetingdate: 5, date_signed: 10, proofdue:12
        # recent to earliest: 0, earliest to recent: 1 (True or False)
        # sort by selected date (self.sortby) in selected order (self.sort_order)
        itemget_index = sortby[int(self.sortby.get())]  # gets the index for the date to sort the search result
        reverse_index = bool(int(self.sort_order.get()))  # get the 'true/false' for reverse for sort.
        self.search_result.sort(key=itemgetter(itemget_index), reverse=reverse_index)
        if showtime:
            self.showtime(frame)  # display the results

    def showtime(self, frame, turnpage=False):
        """ shows the results for the specified range."""
        # if turnpage is false, then initize all nav bar vars
        if not turnpage:
            self.current_page = 1  # the current page of results to display
            # self.rec_display_limit = 50  # the number of records displayed before a new page is needed
        # get the range of recs on this page. 's' is the start and 'e' is the end range
        p = self.current_page  # the current page/screen
        dl = self.rec_display_limit  # the number of records displayed per page
        r = len(self.search_result)  # the total number of records
        s = ((p - 1) * dl)  # start: find the start of the range
        e = (p * dl) - 1  # end: find the end of the range
        if e > r:
            e = r  # handle index issues on last page
        self.win = MakeWindow()
        self.win.create(frame)
        Label(self.win.body, text="Informal C Search Results", font=macadj("bold", "Helvetica 18")) \
            .grid(row=0, column=0, columnspan=4, sticky="w")
        Label(self.win.body, text="").grid(row=1)
        if not self.search_result:  # if there is nothing in the search results
            Label(self.win.body, text="The search has no results.").grid(row=3, column=0, columnspan=4)
        else:  # if there is something in the search results
            self.navigation_bar(self.win.topframe, row=2)  # navigation bar
            Label(self.win.body, text="Grievance Number", fg="grey", anchor="w").grid(row=3, column=1, sticky="w")
            datetext = ["Start Incident", "End Incident", "Meeting Date", "Signed Date", "Proof Due"]
            sort_index = int(self.sortby.get())
            Label(self.win.body, text=datetext[sort_index], fg="grey", anchor="w").grid(row=3, column=2, sticky="w")
            Label(self.win.body, text="Settlement", fg="grey", anchor="w").grid(row=3, column=3, sticky="w")
        row = 4
        ii = s
        for i in range(s, e + 1):
            if i == r:
                break
            column = 0  # re initialize the column
            if ii & 1:  # this alternates the colors of the rows between two colors
                color = "light yellow"
            else:
                color = "white"
            # Show search results. loop once for each settlement.
            # the count at the right margin
            Label(self.win.body, text=str(ii + 1), anchor="w", width=macadj(4, 2), bg=macadj(color, "white"))\
                .grid(row=row, column=column)
            column += 1
            # the grievance number
            Button(self.win.body, text=" " + self.search_result[i][2], anchor="w", width=macadj(14, 12),
                   relief=RIDGE, bg=macadj(color, "white")).grid(row=row, column=column)
            column += 1
            # "Start Incident Date", "End Incident Date", "Meeting Date", "Signed Date", "Proof Due"
            sortby = (3, 4, 5, 10, 12)
            sort_index = sortby[int(self.sortby.get())]  # sent by self.sortby stringvar
            # convert to backslash date or empty
            selecteddate = Convert(self.search_result[i][sort_index]).dtstr_to_backslashstr()
            # the date
            Button(self.win.body, text=selecteddate, width=macadj(11, 10),
                   anchor="w", relief=RIDGE, bg=macadj(color, "white")).grid(row=row, column=column)
            column += 1
            # the settlement
            Button(self.win.body, text=self.search_result[i][11], width=macadj(25, 20), anchor="w",
                   relief=RIDGE, bg=macadj(color, "white")).grid(row=row, column=column)
            column += 1
            Button(self.win.body, text="Edit", width=macadj(7, 6), relief=RIDGE, bg=macadj(color, "white"),
                   command=lambda x=self.search_result[i][2]:
                   self.GrievanceInput(self).informalc_edit(self.win.topframe, x))\
                .grid(row=row, column=column)
            column += 1
            Button(self.win.body, text="Report", width=macadj(6, 5), relief=RIDGE, bg=macadj(color, "white"),
                   command=lambda x=self.search_result[i]: InformalCReports(self).everything_report(x))\
                .grid(row=row, column=column)
            column += 1
            Button(self.win.body, text=macadj("Enter Awards", "Awards"), width=macadj(10, 6), relief=RIDGE,
                   bg=macadj(color, "white"),
                   command=lambda x=self.search_result[i][2]: self.addawards_screen(self.win.topframe, x)) \
                .grid(row=row, column=column)
            row += 1
            ii += 1
        if self.search_result:  # only show the nav bar if there are search results.
            self.navigation_bar(self.win.topframe, row=row)  # navigation bar
        # define the buttons at the bottom of the page:
        Button(self.win.buttons, text="Go Back", width=macadj(16, 13),
               command=lambda: self.master_search(self.win.topframe)).grid(row=0, column=0)
        Button(self.win.buttons, text="Report", width=macadj(16, 11),
               command=lambda: self.AwardsReports(self).reports_screen(self.win.topframe)).grid(row=0, column=1)
        Button(self.win.buttons, text="Summary", width=macadj(16, 11),
               command=lambda: InformalCReports(self).grv_summary()).grid(row=0, column=2)
        self.win.finish()

    def navigation_bar(self, frame, row=0, ):
        """ this navigation bar will allow the user to control the index of the search results, so
        that a limited number of search results appear on a screen. The user can manipulate the
        nav bar to prompt other sections of the search results. """

        def selectpage(selection):
            """ when the button is pressed, the passed selection determines what happens """
            self.current_page = selection
            self.showtime(frame, turnpage=True)

        mn = 1  # minimum: the first page is always '1'
        c = self.current_page  # current: the current page/screen
        dl = self.rec_display_limit
        mx = ceil(len(self.search_result)/dl)  # maximum: the last page/screen
        formula = [c - 1, mn, mn + 1, c - 3, c - 2, c - 1, c + 1, c + 2, c + 3, 
                   mx - 1, mx, c + 1]
        text = ["<", 1, 2, c-3, c-2, c-1, c+1, c+2, c+3, mx-1, mx, ">", ]
        if mx == 1:  # do not use is there is only one page
            return
        navframe = Frame(self.win.body, borderwidth="1", relief="ridge")
        navframe.grid(row=row, column=0, columnspan=7, pady=5)
        Label(navframe, text="Page {} of {}:".format(c, mx), fg="grey").grid(row=0, column=0)
        for i in range(12):
            hide = False
            navbuttons = Button(navframe)
            # general rules
            if formula[i] < 1:  # if value is less than 1
                hide = True
            if formula[i] == c:  # if value is same as the current page
                hide = True
            if formula[i] > mx:  # if the value is greater than last page
                hide = True
            if i in (3, 4, 5, 6, 7, 8):  # hide buttons with duplicate numbers
                if formula[i] in (mn, mn + 1, mx - 1, mx):
                    hide = True
            if i in (1, 2):
                if formula[i] in (mx - 1, mx):
                    hide = True

            navbuttons.config(text=text[i], command=lambda x=formula[i]: selectpage(x), fg="black",
                              width=4, anchor="center", relief="flat")
            navbuttons.grid(row=0, column=i+1)
            if hide:
                navbuttons.grid_remove()

    def informalc_root(self, mode, topframe=None, grv_no=None, childframe=None):
        """ creates a companion window for selecting carrier names.
        mode is 'selectcarrier', 'award' or 'selectissue'. grv_no is used for editing grievance information. """

        def get_listbox_carrriers():
            """ pull options of carriers directly from the informalc grievances table. """
            sql = "SELECT DISTINCT grievant FROM informalc_grievances WHERE station = '%s'" % self.station
            results = inquire(sql)
            unique_carrier = []
            for carrier in results:
                if carrier[0] not in unique_carrier and carrier[0] != "class action":
                    if carrier[0]:
                        unique_carrier.append(carrier[0])
            unique_carrier.sort()
            self.listbox_fill = ["class action", ] + unique_carrier

        def get_listbox_award_carrriers():
            """ get a list of issues for the listbox in self.informalc_root. if no grv_no is given, the
            issue list for year 1000 AD through 9000 AD will be the range of the issue list. """
            start = '1000-01-01 00:00:00'
            end = '9000-01-01 00:00:00'
            if grv_no:  # if a grievance number is passed, use it to get the carrier list.
                sql = "SELECT startdate, enddate FROM informalc_grievances WHERE grv_no='%s'" % grv_no
                results = inquire(sql)
                if results:
                    if results[0][0]:
                        start = results[0][0]
                    if results[0][1]:
                        end = results[0][1]
            start = dt_converter(start)
            end = dt_converter(end)
            # get a list of carriers given the search criteria.
            self.listbox_fill = ["class action", ] + informalc_gen_clist(start, end, self.station)

        def get_listbox_issues():
            """ get a list of issues for the listbox in self.informalc_root from the informalc_issuescategories
            table of the db"""
            sql = "SELECT DISTINCT issue FROM informalc_grievances WHERE station = '%s'" % self.station
            results = inquire(sql)
            unique_issue = []
            for issue in results:
                if issue[0] not in unique_issue:
                    unique_issue.append(issue[0])
            unique_issue.sort()
            self.listbox_fill = ["class action", ] + unique_issue

        def get_listbox_decisions():
            """ get a list of decisions for the listbox in self.informalc_root from the informalc_decisionscategories
            table of the db"""
            sql = "SELECT DISTINCT decision FROM informalc_settlements"
            results = inquire(sql)
            unique_decision = []
            for decision in results:
                if decision[0] not in unique_decision:
                    unique_decision.append(decision[0])
            unique_decision.sort()
            self.listbox_fill = ["class action", ] + unique_decision

        def addnames():
            """ sets the grievant field by setting the stringvar self.grvent using an index from the
             listbox and an array generated in informalc root. """
            for index in listbox.curselection():
                carrier_name = self.listbox_fill[index]
                if not self.grvent[0].get():  # if there is nothing in the first stringvar
                    self.grvent[0].set(carrier_name)  # add the name to that empty stringvar
                else:
                    self.add_grvent_field(childframe, carrier=carrier_name)  # create a new stringvar and widgets

        def addissue():
            """ sets the issue field by setting the stringvar src_issue using an index from the
             listbox and an array generated in informalc root. """
            for index in listbox.curselection():
                issue_name = self.listbox_fill[index]
                if not self.src_issue[0].get():
                    self.src_issue[0].set(issue_name)
                else:
                    self.add_src_issue_field(childframe, issue=issue_name)

        def adddecision():
            """ sets the decision field by setting the stringvar self.decision using an index from the
             listbox and an array generated in informalc root. """
            for index in listbox.curselection():
                decision_name = self.listbox_fill[index]
                if not self.decision[0].get():
                    self.decision[0].set(decision_name)
                else:
                    self.add_decision_field(childframe, decision_1=decision_name)

        def add_awardnames():
            """ inserts names into informal c awards table. """
            for index in listbox.curselection():
                sql = "INSERT INTO informalc_awards2 (grv_no, carrier_name, award, gats_discrepancy) " \
                      "VALUES('%s','%s','%s','%s')" \
                      % (grv_no, self.listbox_fill[int(index)], '', '')
                commit(sql)

        self.destroy_companion()  # destroy other companion windows if they exist
        self.companion_root = Tk()
        self.companion_root.title("KLUSTERBOX")
        titlebar_icon(self.companion_root)  # place icon in titlebar
        x_position = projvar.root.winfo_x() + 450
        y_position = projvar.root.winfo_y() - 25
        self.companion_root.geometry("%dx%d+%d+%d" % (240, 600, x_position, y_position))
        rootframe = Frame(self.companion_root)
        rootframe.pack()
        buttons = Canvas(rootframe)  # button bar
        buttons.pack(fill=BOTH, side=BOTTOM)
        text = "Add"
        if mode in ('selectissue', ):
            text = "Add Issues"
        if mode in ('selectcarrier', 'award'):
            text = "Add Carriers"
        if mode in ('selectdecision', ):
            text = "Add Decisions"
        Label(rootframe, text=text, font=macadj("bold", "Helvetica 18")).pack(anchor="w")
        Label(rootframe, text="").pack()
        scrollbar = Scrollbar(rootframe, orient=VERTICAL)
        listbox = Listbox(rootframe, selectmode="multiple", yscrollcommand=scrollbar.set)
        listbox.config(height=100, width=50)
        if mode == 'selectcarrier':  # use a list of carriers.
            get_listbox_carrriers()
        if mode == 'award':
            get_listbox_award_carrriers()
        if mode == 'selectissue':  # use a list of issue from self.issue_description
            get_listbox_issues()
        if mode == 'selectdecision':  # use a list of issue from self.issue_description
            get_listbox_decisions()
        for name in self.listbox_fill:  # fill the listbox
            listbox.insert(END, name)
        scrollbar.config(command=listbox.yview)
        scrollbar.pack(side=RIGHT, fill=Y)
        listbox.pack(side=LEFT, expand=1)
        if mode == 'selectcarrier':
            Button(buttons, text="Add Carriers", width=macadj(10, 10),
                   command=lambda: addnames()).pack(side=LEFT, anchor="w")
        if mode == 'award':
            Button(buttons, text="Add Carriers", width=macadj(10, 10),
                   command=lambda: (add_awardnames(),
                                    self.addawards_screen(topframe, grv_no))).pack(side=LEFT, anchor="w")
        if mode == "selectissue":
            Button(buttons, text="Add Issue", width=macadj(10, 10),
                   command=lambda: addissue()).pack(side=LEFT, anchor="w")
        if mode == "selectdecision":
            Button(buttons, text="Add Decision", width=macadj(10, 10),
                   command=lambda: adddecision()).pack(side=LEFT, anchor="w")
        # to destroy and re create itself.
        Button(buttons, text="Clear", width=macadj(10, 5),
               command=lambda: (self.destroy_companion(),
                                self.informalc_root(mode, topframe=topframe, grv_no=grv_no, childframe=childframe)))\
            .pack(side=LEFT, anchor="w")
        Button(buttons, text="Close", width=macadj(10, 9),
               command=lambda: (self.destroy_companion())).pack(side=LEFT, anchor="w")

    def destroy_companion(self):
        """ exit out of a screen with a companion root. Destroy the companion window if it still exist. """
        try:
            self.companion_root.destroy()  # destroy the tkinter root object
            self.companion_root = None  # re initialize the variable
        except (TclError, AttributeError):
            pass

    def addawards_screen(self, frame, grv_no):
        """ creates a screen which allows a user to adds the awards to a settlement. """
        award_frame = []  # each award is given its own frame so that multiple gats_discrepancies can displayed.
        self.award_gats_entry = []  # an array that holds entry widgets for gats discrepancies
        self.award_gats_del = []  # an array that holds delete button widgets for gats discrepancies

        def deletename(ids):
            """ deletes records from informal c awards. self.win.topframe, grv_no, ident"""
            sql_del = "DELETE FROM informalc_awards2 WHERE rowid='%s'" % ids
            commit(sql_del)
            self.addawards_screen(self.win.topframe, grv_no)

        def add_gats_field(x, childframe):
            """ added fields for gats discrepancies """
            add_stringvar = StringVar(self.win.body)
            self.var_gats[x].append(add_stringvar)  # add this to an array of stringvars for gats discrepancies
            gats_entry = Entry(childframe, textvariable=self.var_gats[x][len(self.var_gats[x]) - 1],
                               width=macadj(16, 11))
            gats_entry.grid(row=len(self.var_gats[x]) - 1, column=4, padx=2)
            self.award_gats_entry[x].append(gats_entry)  # add this to an array of entry widgets for gats discrepancies
            del_ = Button(childframe, text="-", anchor="center", width=2,
                          command=lambda xx=x, y=len(self.var_gats[x]) - 1: del_gats_field(xx, y))
            del_.grid(row=len(self.var_gats[x]) - 1, column=5)
            self.award_gats_del[x].append(del_)  # add this to an array of widgets of delete buttons
            projvar.root.update()
            # bind the expanding frome to the canvas and scrollregion
            childframe.bind('<Configure>', lambda e: self.win.c.configure(scrollregion=self.win.c.bbox("all")))
            self.win.topframe.bind("<Configure>", self.win.detect_resize)  # track when the window changes size

        def del_gats_field(x, y):
            """ delete a field from search criteria - grievant, entry widgets as well as the delete button.
            set the value of the corresponding stringvar to an empty string. """
            self.award_gats_entry[x][y].grid_remove()
            self.award_gats_del[x][y].grid_remove()
            self.var_gats[x][y].set("")  # set the value of the stringvar to empty string

        self.win = MakeWindow()
        self.win.create(frame)
        self.informalc_root("award", grv_no=grv_no, topframe=self.win.topframe)
        Label(self.win.body, text="Add/Update Settlement Awards", font=macadj("bold", "Helvetica 18")) \
            .grid(row=0, column=0, sticky="w", columnspan=4)
        Label(self.win.body, text="   Grievance Number: {}".format(grv_no), fg="blue") \
            .grid(row=1, column=0, sticky="w", columnspan=4)
        Label(self.win.body, text="Instructions/ Help ").grid(row=2, column=1, columnspan=2, sticky="e")
        Button(self.win.body, text=" read ", width=macadj(8, 4),
               command=lambda: Awards().award_instructions(self.win.topframe)).grid(row=2, column=3, sticky="w")
        sql = "SELECT grv_no,rowid,carrier_name,award,gats_discrepancy FROM informalc_awards2 WHERE grv_no ='%s' " \
              "ORDER BY carrier_name" % grv_no
        result = inquire(sql)
        # initialize arrays for names
        self.var_id = []
        self.var_name = []
        self.var_award = []
        self.var_gats = []
        if len(result) == 0:
            Label(self.win.body, text="No records in database").grid(row=3)
        else:
            Label(self.win.body, text="Carrier", fg="grey", anchor="w", width=macadj(17, 15))\
                .grid(row=3, column=0, sticky="w")
            Label(self.win.body, text="Award", fg="grey", anchor="w", width=macadj(14, 12))\
                .grid(row=3, column=1, sticky="w")
            Label(self.win.body, text="Gats discrepancy", fg="grey", anchor="w", width=macadj(13, 12))\
                .grid(row=3, column=2, sticky="w")
            Label(self.win.body, text="", fg="grey", anchor="w", width=9).grid(row=3, column=3, sticky="w")
            i = 0
            r = 4
            for res in result:
                # ----------------------------------------------------------------------------------------------- frame
                award_frame.append(Frame(self.win.body))
                award_frame[i].grid(row=r, sticky="w", columnspan=4)
                # ------------------------------------------------------------------------------------------------- id
                self.var_id.append(StringVar(self.win.topframe))  # add to arrays
                self.var_id[i].set(res[1])  # set the textvariables
                # ----------------------------------------------------------------------------------------------- name
                self.var_name.append(StringVar(self.win.topframe))
                Label(award_frame[i], text=res[2], anchor="w", width=macadj(16, 14)) \
                    .grid(row=0, column=0, sticky="w", padx=2)  # display name widget
                self.var_name[i].set(res[2])
                # ---------------------------------------------------------------------------------------------- award
                self.var_award.append(StringVar(self.win.topframe))
                Entry(award_frame[i], textvariable=self.var_award[i], width=macadj(16, 11)) \
                    .grid(row=0, column=3, padx=2)  # display award widget
                self.var_award[i].set(res[3])
                # ------------------------------------------------------------------------------------ gats discrepancy
                self.var_gats.append([])  # add an array to hold the stringvars
                self.award_gats_entry.append([])  # add an array to hold the gats discrepancies entry widgets
                self.award_gats_del.append([])  # add an array to hold the gats discrepancies delete buttons.
                gats_result = Convert(res[4]).string_to_array()  # the gats results are a string in the db
                for ii in range(len(gats_result)):  # create a separate stringvar for each element
                    self.var_gats[i].append(StringVar(self.win.topframe))
                    self.var_gats[i][ii].set(gats_result[ii])
                    gat_entry = Entry(award_frame[i], textvariable=self.var_gats[i][ii], width=macadj(16, 11))
                    gat_entry.grid(row=0 + ii, column=4, padx=2)  # display gats discrepancy widget
                    self.award_gats_entry[i].append(gat_entry)  # add to array of entry widgets
                    del_but = Button(award_frame[i], text="+", width=2,  # the first button will add fields
                                     command=lambda x=i: add_gats_field(x, award_frame[x]))
                    if ii > 0:  # if the button is not the first, redefine to a delete button
                        del_but = Button(award_frame[i], text="-", width=2,  # this button will delete fields
                                         command=lambda xx=i, y=len(self.var_gats[i]) - 1: del_gats_field(xx, y))
                    del_but.grid(row=0 + ii, column=5, padx=2)  # display the delete button
                    self.award_gats_del[i].append(del_but)  # add to array of delete button widgets
                # --------------------------------------------------------------------------------------- delete button
                Button(award_frame[i], text="delete",
                       command=lambda ident=res[1]: deletename(ident)) \
                    .grid(row=0, column=6, padx=2)  # display the delete button
                r += 1
                i += 1
        Button(self.win.buttons, text="Go Back", width=15,
               command=lambda: (self.destroy_companion(), self.showtime(self.win.topframe, turnpage=True))) \
            .grid(row=0, column=0)
        Button(self.win.buttons, text="Apply", width=15,
               command=lambda: self.addaward_apply(self.win.topframe, grv_no)).grid(row=0, column=1)
        self.win.finish()

    def addaward_apply(self, topframe, grv_no):
        """ checks and adds records to the informal c add awards table. """
        sql_queue = []
        for i in range(len(self.var_id)):
            id_no = self.var_id[i].get()  # simplify variable names
            carrier = self.var_name[i].get()  # this is a stringvar
            award = self.var_award[i].get().strip()  # this is a stringvar
            gats_discrepancy = self.var_gats[i]  # this is a list of stringvars
            if not AwardsChecker().check_all(topframe, carrier, award, gats_discrepancy):
                return
            award_add = AwardsFormatting().format_for_db(award)  # add decimal points and hundredths.
            gats_add = Convert(gats_discrepancy).strvarlist_to_string()  # make list of stringvars into a string
            gats_add = AwardsFormatting().format_for_db(gats_add)  # add decimal points and hundredths.
            sql = "UPDATE informalc_awards2 SET award='%s',gats_discrepancy='%s' WHERE rowid='%s'" % (
                award_add, gats_add, id_no)  # write sql command
            sql_queue.append(sql)  # put all sql commands in queue until all checks pass.
            self.win.buttons.update()  # update the progress bar
        pb_label = Label(self.win.buttons, text="Updating Changes: ")  # make label for progress bar
        pb_label.grid(row=0, column=2)
        pb = ttk.Progressbar(self.win.buttons, length=200, mode="determinate")  # create progress bar
        pb.grid(row=0, column=3)
        pb["maximum"] = len(self.var_id)  # set length of progress bar
        pb.start()
        ii = 0
        for sql in sql_queue:
            pb["value"] = ii  # increment progress bar
            commit(sql)
            ii += 1
        pb.stop()  # stop and destroy the progress bar
        pb_label.destroy()  # destroy the label for the progress bar
        pb.destroy()
        self.addawards_screen(topframe, grv_no)

    class AwardsReports:
        """ generate reports for settlement awards """
        def __init__(self, parent):
            self.parent = parent
            self.win = None

        def reports_screen(self, frame):
            """ a screen where users can choose from a list of reports. """
            self.win = MakeWindow()
            self.win.create(frame)
            Label(self.win.body, text="Informal C Reports", font=macadj("bold", "Helvetica 18")) \
                .grid(row=0, column=0, sticky="w")
            Label(self.win.body, text="").grid(row=1)
            row = 3
            Button(self.win.body, text="Grievance Everything", width=30,
                   command=lambda: InformalCReports(self).everything_all_report()).grid(row=row, column=0, pady=5)
            row += 1
            Button(self.win.body, text="Monetary Awards", width=30,
                   command=lambda: InformalCReports(self).monetary_sum()).grid(row=row, column=0, pady=5)
            row += 1
            Button(self.win.body, text="Gats Discrepancies", width=30,
                   command=lambda: InformalCReports(self).gats_discrepancies()).grid(row=row, column=0, pady=5)
            row += 1
            Button(self.win.body, text="Only Discrepancies", width=30,
                   command=lambda: InformalCReports(self).gats_discrepancies(fullreport=False)) \
                .grid(row=row, column=0, pady=5)
            row += 1
            Button(self.win.body, text="All Adjustments", width=30,
                   command=lambda: InformalCReports(self).adjustments())\
                .grid(row=row, column=0, pady=5)
            row += 1
            Button(self.win.body, text="No Adjustments", width=30,
                   command=lambda: InformalCReports(self).adjustments(fullreport=False))\
                .grid(row=row, column=0, pady=5)
            row += 1
            Button(self.win.body, text="Awards by Carriers", width=30,
                   command=lambda: InformalCReports(self).bycarriers()).grid(row=row, column=0, pady=5)
            row += 1
            Button(self.win.body, text="Awards by Carrier", width=30,
                   command=lambda: self.bycarrier(self.win.topframe)).grid(row=row, column=0, pady=5)
            row += 1
            Button(self.win.body, text="No Settlement", width=30,
                   command=lambda: InformalCReports(self).no_settlement()).grid(row=row, column=0, pady=5)
            row += 1
            Button(self.win.body, text="Compliance Delinquency", width=30,
                   command=lambda: InformalCReports(self).delinquency()) \
                .grid(row=row, column=0, pady=5)
            row += 1
            Button(self.win.body, text="Missing Awards", width=30,
                   command=lambda: InformalCReports(self).missing_awards())\
                .grid(row=row, column=0, pady=5)
            row += 1
            Button(self.win.body, text="Employee ID (spreadsheet)", width=30,
                   command=lambda: RptCarrierId(self).run()).grid(row=row, column=0, pady=5)
            row += 1
            Button(self.win.body, text="Employee ID (text file)", width=30,
                   command=lambda: InformalCReports(self).rptcarrierandid())\
                .grid(row=row, column=0, pady=5)
            row += 1
            Label(self.win.body, text="", width=70).grid(row=row)  # widen the column so buttons appear center
            # define the buttons at the bottom of the page:
            Button(self.win.buttons, text="Go Back", width=macadj(16, 13),
                   command=lambda: self.parent.showtime(self.win.topframe, turnpage=True)).grid(row=0, column=0)
            self.win.finish()

        def uniquecarrier(self):
            """ gets the awards for a carrier from the informalc awards table. """
            unique_grv = []
            for grv in self.parent.search_result:
                if grv[2] not in unique_grv:
                    unique_grv.append(grv[2])
            unique_carrier = []
            for each in unique_grv:
                sql = "SELECT * FROM informalc_awards2 WHERE grv_no='%s'" % each
                results = inquire(sql)
                for r in results:
                    if r[1] not in unique_carrier:
                        unique_carrier.append(r[1])
            unique_carrier.sort()
            return unique_carrier

        def bycarrier(self, frame):
            """ builds a screen that allows a user to select a carrier and generate a text report of settlements. """
            unique_carrier = self.uniquecarrier()
            self.win = MakeWindow()
            self.win.create(frame)
            Label(self.win.body, text="Informal C: Select Carrier", font=macadj("bold", "Helvetica 18")) \
                .pack(anchor="w")
            Label(self.win.body, text="").pack()
            scrollbar = Scrollbar(self.win.body, orient=VERTICAL)
            listbox = Listbox(self.win.body, selectmode="single", yscrollcommand=scrollbar.set)
            listbox.config(height=30, width=50)
            for name in unique_carrier:
                listbox.insert(END, name)
            scrollbar.config(command=listbox.yview)
            scrollbar.pack(side=RIGHT, fill=Y)
            listbox.pack(side=LEFT, expand=1)
            Button(self.win.buttons, text="Go Back", width=20,
                   command=lambda: self.reports_screen(self.win.topframe)).pack(side=LEFT)
            Button(self.win.buttons, text="Report", width=20,
                   command=lambda: InformalCReports(self).bycarrier_apply(unique_carrier, listbox.curselection()))\
                .pack(side=LEFT)
            self.win.finish()

    class GrievanceInput:
        """
        Allows the user to create new records of grievances.
        """
        def __init__(self, parent):
            self.parent = parent
            self.newentry = False  # newentry is True or False
            self.win = None  # the window object
            self.edit_grv_no = ""  # if the grievance is being edited, the grievance number is passed.
            self.row = 0  # the row of the body of the window
            self.msg = ""  # a message displayed when grievances are inserted/ updated
            self.grv_changesmade = False  # flag to update grievance record if changes are made
            self.set_changesmade = False  # flag to update settlement record if changes are made
            self.issue_description = []  # initialize and empty issue description list
            self.issue_article = []  # initialize and empty article description list
            self.issue_description = self.parent.issue_description[:]  # a list of issues for the option menu
            self.issue_article = self.parent.issue_article[:]  # a list of corrosponding articles for the list of issues
            # a list of decisions for the option menu
            self.decision_description = []  # initialize and empty decision description list
            self.decision_description = ["no decision", ] + self.parent.decision_description[:]
            self.companion_root = None  # a auxiliary window that allows users to select carriers/ issues
            self.listbox_fill = None

            #  define the stringvars
            self.grievant = None  # 1
            self.station = None  # 1.5
            self.grv_no = None  # 2
            self.startdate = None  # 3
            self.enddate = None  # 4
            self.meetingdate = None  # 5
            self.issue = None  # 6
            self.article = None  # 7
            self.non_c = None  # 8  is the grievance a non compliance grievance?
            self.reman = None  # 9  is the grievance a remanded grievance?
            self.lvl = None  # 10 the level of the settlement
            self.date_signed = None  # 11 the date the settlement was signed
            self.decision = None  # 12 the decision of the settlement
            self.proof_due = None  # 13 the date that the prooof of the remedy is due, if applicable
            self.docs = None  # 14 the status of any documentation needed for proof of compliance
            self.batset = None  # 16 is the settlement part of a batch settlement?
            self.batgat = None  # 17  is the settlement part of a gats reports?

            # get the values of the grievance/settlements on record
            self.onrec_grievance = False
            self.onrec_grievant = ""  # 1
            self.onrec_grv_no = ""  # 2
            self.onrec_startdate = ""  # 3
            self.onrec_enddate = ""  # 4
            self.onrec_meetingdate = ""  # 5
            self.onrec_issue = ""  # 6
            self.onrec_article = ""  # 7
            self.index_onrecs = []  # holds onrecs for 8 nonc, 9 remanded, 16 batset and 17 batgat
            self.onrec_settlement = False
            self.onrec_lvl = ""  # 10 the level of the settlement
            self.onrec_date_signed = ""  # 11 the date the settlement was signed
            self.onrec_decision = ""  # 12 the decision of the settlement
            self.onrec_proof_due = ""  # 13 the date that the prooof of the remedy is due, if applicable
            self.onrec_docs = ""  # 14 the status of any documentation needed for proof of compliance

            # store widgets for display and deletion
            self.nonc_entry = []  # this array store the entry fields of the non compliance indexes
            self.nonc_del = []  # this array stores the delete buttons for the non compliance entry widgets
            self.reman_entry = []  # this array store the entry fields of the remanded indexes
            self.reman_del = []  # this array stores the delete buttons for the remanded widgets
            self.batset_entry = []  # this array store the entry fields of the batch settlement indexes
            self.batset_del = []  # this array stores the delete buttons for the batch settlement entry widgets
            self.batgat_entry = []  # this array store the entry fields of the gats reports indexes
            self.batgat_del = []  # this array stores the delete buttons for the gats reports entry widgets
            self.nonc_frame = None  # this is a frame for the associations/indexes
            self.reman_frame = None  # this is a frame for the associations/indexes
            self.batset_frame = None  # this is a frame for the associations/indexes
            self.batgat_frame = None  # this is a frame for the associations/indexes

            # check elements of grievances and settlements
            self.check_grievant = None  # 1
            self.check_grv_no = None  # 2
            self.check_startdate = None  # 3
            self.check_enddate = None  # 4
            self.check_meetingdate = None  # 5
            self.check_dates = []  # array to hold startdate, enddate and meetingdate - form in check_dates()
            self.check_issue = None  # 6
            self.check_article = None  # 7
            self.check_lvl = None  # 10 the level of the settlement
            self.check_datesigned = None  # 11 the date the settlement was signed
            self.check_decision = None  # 12 the decision of the settlement
            self.check_proofdue = None  # 13 the date that the prooof of the remedy is due, if applicable
            self.check_docs = None  # 14 the status of any documentation needed for proof of compliance
            self.check_indexes = []  # multidimensional list to store indexes - nonc, remanded, batch set, gats reports
            self.add_indexes = []
            self.del_indexes = []

            # flags updates/inserts for grievance, settlement and indexes
            self.reporter_grv = False  # flag for grievances
            self.reporter_set = False  # flag for settlement
            self.reporter_index = False  # flag for indexes
            self.msg_label = None

        def informalc_new(self, frame, msg=""):
            """ master method for running other methods in proper order."""
            if msg:  # if this being reloaded after apply
                self.__init__(self.parent)  # re initialize the class
            self.newentry = True  # this is a new entry
            self.msg = msg
            self.win = MakeWindow()
            self.get_stringvars()
            self.win.create(frame)
            self.build_screen()
            self.win.finish()

        def informalc_edit(self, frame, grv_no, msg=""):
            """ master method for running other methods in proper order."""
            if msg:  # if this being reloaded after apply
                self.__init__(self.parent)  # re initialize the class
            self.newentry = False  # this is not a new entry
            self.edit_grv_no = grv_no  # the grievance to be edited
            self.msg = msg
            self.win = MakeWindow()
            self.get_onrecs(grv_no)
            self.get_stringvars()
            self.set_stringvars()
            self.win.create(frame)
            self.build_screen()
            self.win.finish()

        def get_stringvars(self):
            """ initialize the stringvars """
            self.grievant = StringVar(self.win.body)
            self.station = StringVar(self.win.body)
            self.grv_no = StringVar(self.win.body)
            self.startdate = StringVar(self.win.body)
            self.enddate = StringVar(self.win.body)
            self.meetingdate = StringVar(self.win.body)
            self.issue = StringVar(self.win.body)
            self.article = StringVar(self.win.body)
            self.non_c = [StringVar(self.win.body), ]
            self.reman = [StringVar(self.win.body), ]
            self.lvl = StringVar(self.win.body)
            self.date_signed = StringVar(self.win.body)
            self.decision = StringVar(self.win.body)
            self.proof_due = StringVar(self.win.body)
            self.docs = StringVar(self.win.body)
            self.batset = [StringVar(self.win.body), ]
            self.batgat = [StringVar(self.win.body), ]

        def get_onrecs(self, grv_number):
            """ check if there is an existing record for the grievance number in the informalc grievances table.
                   if so, store the values in the self.onrec variables. if not, the self.onrec variables
                   default to empty. """
            self.onrec_grievance = False  # make sure that onrec_grievances is re initialized.
            self.onrec_settlement = False  # make sure that onrec_settlement is re initialized.
            onrec_non_c = []
            onrec_reman = []
            onrec_batset = []
            onrec_batgat = []
            sql = "SELECT * FROM informalc_grievances WHERE grv_no = '%s' and station = '%s'" \
                  % (grv_number, self.parent.station)
            results = inquire(sql)
            if results:
                self.onrec_grievance = True
                self.onrec_grievant = results[0][0]
                # skip station as that is held in self.parent.station and is part of the search criteria
                # skip grievance number as that is self.edit_grv_no and is part of the search criteria
                self.onrec_startdate = results[0][3]
                self.onrec_enddate = results[0][4]
                self.onrec_meetingdate = results[0][5]
                self.onrec_issue = results[0][6]
                self.onrec_article = results[0][7]
            sql = "SELECT * FROM informalc_settlements WHERE grv_no = '%s'" % self.edit_grv_no
            results = inquire(sql)
            if results:
                self.onrec_settlement = True
                # skip grievance number as that is self.edit_grv_no and is part of the search criteria
                self.onrec_lvl = results[0][1]
                self.onrec_date_signed = results[0][2]
                self.onrec_decision = results[0][3]
                self.onrec_proof_due = results[0][4]
                self.onrec_docs = results[0][5]
            # use arrays and loops to get search results for all the grievances in the grv_list array.
            # search these tables
            tables_array = ("informalc_noncindex", "informalc_remandindex",
                            "informalc_batchindex", "informalc_gats")
            # search these columns in the tables
            search_criteria_array = ("followup", "refiling", "main", "grv_no")
            for i in range(len(tables_array)):  # loop for each table
                sql = "SELECT * FROM %s WHERE %s = '%s'" % \
                      (tables_array[i], search_criteria_array[i], self.edit_grv_no)
                result = inquire(sql)
                if tables_array[i] == "informalc_noncindex":  # get the onrecs for non compliance index
                    if result:  # if there is a result
                        for r in result:  # there can be multiple results
                            onrec_non_c.append(r)  # add record to the array
                if tables_array[i] == "informalc_remandindex":  # get the onrecs for informalc_remandindex
                    if result:
                        for r in result:
                            onrec_reman.append(r)
                if tables_array[i] == "informalc_batchindex":  # get the onrecs for informalc_batchindex
                    if result:
                        for r in result:
                            onrec_batset.append(r)
                if tables_array[i] == "informalc_gats":  # get the onrecs for informalc_batchindex
                    if result:
                        for r in result:
                            onrec_batgat.append(r)
            self.index_onrecs = [onrec_non_c, onrec_reman, onrec_batset, onrec_batgat]

        def set_stringvars(self):
            """ use the data from onrecs to set the stringvar values
            the index stringvars are set after the widgets are generated from self.index_onrecs"""
            if self.newentry:
                self.grievant.set("")  # 1
                self.grv_no.set("")  # 2
                self.startdate.set("")  # 3
                self.enddate.set("")  # 4
                self.meetingdate.set("")  # 5
                self.issue.set("")  # 6
                self.article.set("")  # 7
                self.lvl.set("")  # 10 the level of the settlement
                self.date_signed.set("")  # 11 the date the settlement was signed
                self.decision.set("")  # 12 the decision of the settlement
                self.proof_due.set("")  # 13 the date that the prooof of the remedy is due
                self.docs.set("")  # 14 the status of any documentation needed for proof of compliance
            else:
                self.grievant.set(self.onrec_grievant)  # 1
                self.grv_no.set(self.edit_grv_no)  # 2
                self.startdate.set(Convert(self.onrec_startdate).dtstr_to_backslashstr())  # 3
                self.enddate.set(Convert(self.onrec_enddate).dtstr_to_backslashstr())  # 4
                self.meetingdate.set(Convert(self.onrec_meetingdate).dtstr_to_backslashstr())  # 5
                self.issue.set(self.onrec_issue)  # 6
                self.article.set(self.onrec_article)  # 7
                if self.onrec_lvl:  # 10 the level of the settlement
                    self.lvl.set(self.onrec_lvl)
                else:  # if there is nothing in self.onrec_lvl
                    self.lvl.set("no status")  # enter 'no status'
                # 11 the date the settlement was signed
                self.date_signed.set(Convert(self.onrec_date_signed).dtstr_to_backslashstr())
                # if self.onrec_decision:  # 12 the decision of the settlement
                self.decision.set(self.onrec_decision)
                # 13 the date that the prooof of the remedy is due
                self.proof_due.set(Convert(self.onrec_proof_due).dtstr_to_backslashstr())
                if self.onrec_docs:  # 14 the status of any documentation needed for proof of compliance
                    self.docs.set(self.onrec_docs)
                else:
                    self.docs.set("no status")

        def build_screen(self):
            """ screen for entering in settlements. before leaving the frame, attempt to destroy self.companion_root
             with the self.destroy_companion() """
            self.row = 0
            self.build_grievanceinfo()  # this is basic grievance information
            self.build_nonc_assocs()  # for non compliance associations
            self.build_remand_assoc()  # for remanded grievance associations
            self.build_settlement()  # this area of the screen is for settlement information
            self.build_batset_assocs()  # this area of the screen is for batch settlements
            self.build_batgat_assocs()  # this area of the screen is for gats reports
            self.fillin_index_onrecs()  # this will set stringvars for indexes and populate fields
            self.build_buttons()  # configure buttons on the bottom of the screen

        def build_grievanceinfo(self):
            """ insert the header """
            text = "Enter New Grievance"  # alternate header for new grievances
            if not self.newentry:
                text = "Edit Grievance"  # alternate header for edit grievances.
            Label(self.win.body, text=text, font=macadj("bold", "Helvetica 18")) \
                .grid(row=self.row, column=0, columnspan=2, sticky="w")
            self.row += 1
            Label(self.win.body, text="").grid(row=self.row, column=0, sticky="w")
            self.row += 1
            Label(self.win.body, text="Grievant: ", background=macadj("gray95", "white"),
                  fg=macadj("black", "black"), width=macadj(24, 22), anchor="w", height=macadj(1, 1)) \
                .grid(row=self.row, column=0, sticky="w")
            Entry(self.win.body, textvariable=self.grievant, justify='right', width=macadj(20, 16)) \
                .grid(row=self.row, column=1, sticky="w")
            self.row += 1
            # button to activate companion window to select carriers.
            Button(self.win.body, text="list", width=macadj(4, 3), anchor="center",
                   command=lambda: (self.destroy_companion(), self.informalc_root("selectcarrier", self.edit_grv_no)))\
                .grid(row=self.row, column=1, sticky="e")
            self.row += 1
            Label(self.win.body, text="Grievance Number: ", background=macadj("gray95", "white"),
                  fg=macadj("black", "black"), width=macadj(24, 22), anchor="w", height=macadj(1, 1)) \
                .grid(row=self.row, column=0, sticky="w")
            Entry(self.win.body, textvariable=self.grv_no, justify='right', width=macadj(20, 16)) \
                .grid(row=self.row, column=1, sticky="w")
            self.row += 1
            if not self.newentry:
                # pass the old grv number (string), then new grievance number (stringvar) to grvchange
                Button(self.win.body, width=9, text="update", command=lambda:
                       self.grvchange(self.edit_grv_no, self.grv_no))\
                    .grid(row=self.row, column=1, sticky="e")
                self.row += 1
            # start and end dates
            Label(self.win.body, text="Incident Date").grid(row=self.row, column=0, sticky="w")
            self.row += 1
            # -------------------------------------------------------------------------------------------- start date
            Label(self.win.body, text="  Start (mm/dd/yyyy): ", background=macadj("gray95", "white"),
                  fg=macadj("black", "black"), width=macadj(24, 22), anchor="w") \
                .grid(row=self.row, column=0, sticky="w")
            Entry(self.win.body, textvariable=self.startdate, justify='right', width=macadj(20, 16)) \
                .grid(row=self.row, column=1, sticky="w")
            self.row += 1
            # ------------------------------------------------------------------------------------------------ end date
            Label(self.win.body, text="  End (mm/dd/yyyy): ", background=macadj("gray95", "white"),
                  fg=macadj("black", "black"), width=macadj(24, 22), anchor="w", height=macadj(1, 1)) \
                .grid(row=self.row, column=0, sticky="w")
            Entry(self.win.body, textvariable=self.enddate, justify='right', width=macadj(20, 16)) \
                .grid(row=self.row, column=1, sticky="w")
            self.row += 1
            # -------------------------------------------------------------------------------------------- meeting date
            Label(self.win.body, text="Meeting Date (mm/dd/yyyy): ", background=macadj("gray95", "white"),
                  fg=macadj("black", "black"), width=macadj(24, 22), anchor="w", height=macadj(1, 1)) \
                .grid(row=self.row, column=0, sticky="w")
            Entry(self.win.body, textvariable=self.meetingdate, justify='right', width=macadj(20, 16)) \
                .grid(row=self.row, column=1, sticky="w")
            self.row += 1
            # -------------------------------------------------------------------------------------------------- issue
            Label(self.win.body, text="Issue: ", background=macadj("gray95", "white"),
                  fg=macadj("black", "black"), width=macadj(24, 18), anchor="w", height=macadj(1, 1)) \
                .grid(row=self.row, column=0, sticky="w")
            Button(self.win.body, text="list", width=macadj(4, 3), anchor="center",
                   command=lambda: (self.destroy_companion(), self.informalc_root("selectissue", self.edit_grv_no))) \
                .grid(row=self.row, column=1, sticky="e")
            self.row += 1
            Entry(self.win.body, textvariable=self.issue, width=macadj(49, 42), justify='right') \
                .grid(row=self.row, column=0, sticky="w", columnspan=3)
            self.row += 1
            # ------------------------------------------------------------------------------------------------- article
            Label(self.win.body, text="Article: ", background=macadj("gray95", "white"),
                  fg=macadj("black", "black"), width=macadj(24, 20), anchor="w", height=macadj(1, 1)) \
                .grid(row=self.row, column=0, sticky="w")
            Entry(self.win.body, textvariable=self.article, justify='right', width=macadj(20, 16)) \
                .grid(row=self.row, column=1, sticky="w")
            self.row += 1
            # ------------------------------------------------------------------------------------------- delete button
            if not self.newentry:  # disable if this is a new entry
                Label(self.win.body, text="Delete Grievance: ", background=macadj("gray95", "white"),
                      fg=macadj("black", "black"), width=macadj(24, 22), anchor="w", height=macadj(1, 1)) \
                    .grid(row=self.row, column=0, sticky="w")
                Button(self.win.body, text="delete", width=macadj(17, 15), fg=macadj("white", "red"),
                       bg=macadj("red", "white"), anchor="center",
                       command=lambda: (self.destroy_companion(),
                                        self.delete_grievance(self.win.topframe, self.edit_grv_no))) \
                    .grid(row=self.row, column=1, sticky="e", pady=5)
            self.row += 1

        def build_nonc_assocs(self):
            """ these are widgets for non compliance grievance indexes. Uses separate frames to allow for
                        expanding/colapsing fields """
            text = macadj("Non Compliance Associations ___________________________",
                          "Non Compliance Associations __________________________")
            Label(self.win.body, text=text, anchor="w",
                  fg="blue").grid(row=self.row, column=0, columnspan=3, sticky="w", pady=10)
            self.row += 1
            self.nonc_frame = Frame(self.win.body)
            self.nonc_frame.grid(row=self.row, column=0, sticky="w", columnspan=2)
            Label(self.nonc_frame, text="Overdue Grievances", background=macadj("gray95", "white"),
                  fg=macadj("black", "black"), width=macadj(19, 22), anchor="w", height=macadj(1, 1)) \
                .grid(row=0, column=0, sticky="w")
            nonc = Entry(self.nonc_frame, textvariable=self.non_c[0], justify='right', width=macadj(20, 15))
            nonc.grid(row=0, column=1, sticky="w")
            self.nonc_entry.append(nonc)  # add this to an array of entry widgets for non compliance
            del_ = Button(self.nonc_frame, text="add", width=macadj(4, 3), anchor="center",
                          command=lambda: self.add_nonc_field(self.nonc_frame))
            del_.grid(row=0, column=3)
            self.nonc_del.append(del_)  # add this to an array of widgets of delete buttons
            self.row += 1

        def build_remand_assoc(self):
            """ these are widgets for remanded grievance indexes. Uses separate frames to allow for
            expanding/colapsing fields """
            text = macadj("Remanded  Associations _________________________________",
                          "Remanded Associations _______________________________")
            Label(self.win.body, text=text, anchor="w",
                  fg="blue").grid(row=self.row, column=0, columnspan=3, sticky="w", pady=10)
            self.row += 1
            self.reman_frame = Frame(self.win.body)
            self.reman_frame.grid(row=self.row, column=0, sticky="w", columnspan=2)
            Label(self.reman_frame, text="Remanded Grievances", background=macadj("gray95", "white"),
                  fg=macadj("black", "black"), width=macadj(19, 22), anchor="w", height=macadj(1, 1)) \
                .grid(row=0, column=0, sticky="w")
            reman = Entry(self.reman_frame, textvariable=self.reman[0], justify='right', width=macadj(20, 15))
            reman.grid(row=0, column=1, sticky="w")
            self.reman_entry.append(reman)  # add this to an array of entry widgets for non compliance
            del_ = Button(self.reman_frame, text="add", width=macadj(4, 3), anchor="center",
                          command=lambda: self.add_reman_field(self.reman_frame))
            del_.grid(row=0, column=3)
            self.reman_del.append(del_)  # add this to an array of widgets of delete buttons
            self.row += 1

        def build_settlement(self):
            """ this area of the screen is for settlement information. """
            text = macadj("Settlement _______________________________________________",
                          "Settlement _________________________________________")
            Label(self.win.body, text=text, anchor="w",
                  fg="blue").grid(row=self.row, column=0, columnspan=2, sticky="w")
            self.row += 1
            # -------------------------------------------------------------------------------- level of the settlement
            Label(self.win.body, text="Settlement Level: ", background=macadj("gray95", "white"),
                  fg=macadj("black", "black"), width=macadj(24, 22), anchor="w", height=macadj(1, 1)) \
                .grid(row=self.row, column=0, sticky="w")  # select settlement level
            lvl_options = ("no status", "informal a", "formal a", "step b", "pre arb", "arbitration")
            lvl_om = OptionMenu(self.win.body, self.lvl, *lvl_options)
            lvl_om.config(width=macadj(14, 13))
            lvl_om.grid(row=self.row, column=1)
            if not self.lvl:  # if the stringvar was not updated in onrecs...
                self.lvl.set("no status")
            self.row += 1
            # -------------------------------------------------------------------------------------------- date signed
            Label(self.win.body, text="Date Signed (mm/dd/yyyy): ", background=macadj("gray95", "white"),
                  fg=macadj("black", "black"), width=macadj(24, 25), anchor="w", height=macadj(1, 1)) \
                .grid(row=self.row, column=0, sticky="w")
            Entry(self.win.body, textvariable=self.date_signed, justify='right', width=macadj(20, 16)) \
                .grid(row=self.row, column=1, sticky="w")
            self.row += 1
            # ----------------------------------------------------------------------------------------------- decision
            Label(self.win.body, text="Decision: ", background=macadj("gray95", "white"),
                  fg=macadj("black", "black"), width=macadj(24, 11), anchor="w", height=macadj(1, 1)) \
                .grid(row=self.row, column=0, sticky="w")
            Button(self.win.body, text="list", width=macadj(4, 3), anchor="center",
                   command=lambda: (self.destroy_companion(),
                                    self.informalc_root("selectdecision", self.edit_grv_no))) \
                .grid(row=self.row, column=1, sticky="e")
            self.row += 1
            Entry(self.win.body, textvariable=self.decision, width=macadj(49, 41), justify='right') \
                .grid(row=self.row, column=0, sticky="w", columnspan=3)
            self.row += 1
            # ----------------------------------------------------------------------------------------------- proof due
            Label(self.win.body, text="Proof Due (mm/dd/yyyy): ", background=macadj("gray95", "white"),
                  fg=macadj("black", "black"), width=macadj(24, 22), anchor="w", height=macadj(1, 1)) \
                .grid(row=self.row, column=0, sticky="w")
            Entry(self.win.body, textvariable=self.proof_due, justify='right', width=macadj(20, 16)) \
                .grid(row=self.row, column=1, sticky="w")
            self.row += 1
            # ---------------------------------------------------------------------------------------------------- docs
            Label(self.win.body, text="Docs: ", background=macadj("gray95", "white"),
                  fg=macadj("black", "black"), width=macadj(24, 22), anchor="w", height=macadj(1, 1)) \
                .grid(row=self.row, column=0, sticky="w")  # select docs
            doc_om = OptionMenu(self.win.body, self.docs, *self.parent.doc_options)
            doc_om.config(width=macadj(14, 13))
            doc_om.grid(row=self.row, column=1)
            self.row += 1
            if not self.newentry:
                Label(self.win.body, text="Delete Settlement: ", background=macadj("gray95", "white"),
                      fg=macadj("black", "black"), width=macadj(24, 24), anchor="w", height=macadj(1, 1)) \
                    .grid(row=self.row, column=0, sticky="w")
                Button(self.win.body, text="delete", width=macadj(17, 15), fg=macadj("white", "red"),
                       bg=macadj("red", "white"), anchor="center",
                       command=lambda: (self.destroy_companion(), self.settlement_delete())) \
                    .grid(row=self.row, column=1, sticky="e", pady=5)
                self.row += 1

        def build_batset_assocs(self):
            """ create a gui for BATch SETtlement associations.  """
            text = macadj("Batch Settlement Associations ___________________________",
                          "Batch Settlement Associations _________________________")
            Label(self.win.body, text=text, anchor="w",
                  fg="blue").grid(row=self.row, column=0, columnspan=3, sticky="w", pady=10)
            self.row += 1
            self.batset_frame = Frame(self.win.body)
            self.batset_frame.grid(row=self.row, column=0, sticky="w", columnspan=2)
            Label(self.batset_frame, text="Included Grievances", background=macadj("gray95", "white"),
                  fg=macadj("black", "black"), width=macadj(19, 22), anchor="w", height=macadj(1, 1)) \
                .grid(row=0, column=0, sticky="w")
            batset = Entry(self.batset_frame, textvariable=self.batset[0], justify='right', width=macadj(20, 15))
            batset.grid(row=0, column=1, sticky="w")
            self.batset_entry.append(batset)  # add this to an array of entry widgets for non compliance
            del_ = Button(self.batset_frame, text="add", width=macadj(4, 3), anchor="center",
                          command=lambda: self.add_batset_field(self.batset_frame))
            del_.grid(row=0, column=3)
            self.batset_del.append(del_)  # add this to an array of widgets of delete buttons
            self.row += 1

        def build_batgat_assocs(self):
            """ create a gui for BATch GATs associations. """
            text = macadj("Gats Report/s ____________________________________________",
                          "Gats Report/s _______________________________________")
            Label(self.win.body, text=text, anchor="w",
                  fg="blue").grid(row=self.row, column=0, columnspan=3, sticky="w", pady=10)
            self.row += 1
            self.batgat_frame = Frame(self.win.body)
            self.batgat_frame.grid(row=self.row, column=0, sticky="w", columnspan=2)
            Label(self.batgat_frame, text="Gats Number", background=macadj("gray95", "white"),
                  fg=macadj("black", "black"), width=macadj(19, 22), anchor="w", height=macadj(1, 1)) \
                .grid(row=0, column=0, sticky="w")
            batgat = Entry(self.batgat_frame, textvariable=self.batgat[0], justify='right', width=macadj(20, 15))
            batgat.grid(row=0, column=1, sticky="w")
            self.batgat_entry.append(batgat)  # add this to an array of entry widgets for non compliance
            del_ = Button(self.batgat_frame, text="add", width=macadj(4, 3), anchor="center",
                          command=lambda: self.add_batgat_field(self.batgat_frame))
            del_.grid(row=0, column=3)
            self.batgat_del.append(del_)  # add this to an array of widgets of delete buttons
            self.row += 1

        def build_buttons(self):
            """ configure buttons on the bottom of the screen """
            button_alignment = macadj("w", "center")
            Button(self.win.buttons, text="Submit", width=macadj(13, 14), anchor=button_alignment,
                   command=lambda: (self.destroy_companion(), self.apply(goback=True))).grid(row=0, column=0)
            Button(self.win.buttons, text="Apply", width=macadj(13, 14), anchor=button_alignment,
                   command=lambda: (self.destroy_companion(), self.apply())).grid(row=0, column=1)
            if self.newentry:  # new entries go back to informal c main menu
                Button(self.win.buttons, text="Go Back", width=macadj(13, 14), anchor=button_alignment,
                       command=lambda: (self.destroy_companion(), self.parent.informalc(self.win.topframe)))\
                    .grid(row=0, column=2)
            else:  # edits go back to showtime
                Button(self.win.buttons, text="Go Back", width=macadj(13, 14), anchor=button_alignment,
                       command=lambda: (self.destroy_companion(),
                                        self.parent.showtime(self.win.topframe, turnpage=True)))\
                    .grid(row=0, column=2)
            # a label on the button bar gives notification of past actions ie insert, delete, etc
            Label(self.win.buttons, text=self.msg, fg="red")
            self.msg_label = Label(self.win.buttons, text=self.msg, fg="red")
            self.msg_label.grid(row=0, column=3)

        def add_nonc_field(self, frame, onrec=None):
            """ added fields for compliance index"""
            add_stringvar = StringVar(self.win.body)
            if onrec:  # when the onrec field is not None, it has been called by self.fillin_index_onrecs
                add_stringvar.set(onrec)  # this sets the stringvar with data from the the db
            self.non_c.append(add_stringvar)  # add this to an array of stringvars for non compliance
            nonc = Entry(frame, textvariable=self.non_c[len(self.non_c)-1], justify='right', width=macadj(20, 15))
            nonc.grid(row=len(self.non_c)-1, column=1, sticky="w")
            self.nonc_entry.append(nonc)  # add this to an array of entry widgets for non compliance
            del_ = Button(frame, text="del", width=macadj(4, 3), anchor="center",
                          command=lambda x=len(self.non_c)-1: self.del_nonc_field(x))
            del_.grid(row=len(self.non_c)-1, column=3)
            self.nonc_del.append(del_)  # add this to an array of widgets of delete buttons
            # bind the expanding frome to the canvas and scrollregion
            frame.bind('<Configure>', lambda e: self.win.c.configure(scrollregion=self.win.c.bbox("all")))
            self.win.topframe.bind("<Configure>", self.win.detect_resize)  # track when the window changes size

        def del_nonc_field(self, x):
            """ delete a field from the non compliance entry widgets as well as the delete button.
            set the value of the corresponding stringvar to an empty string. """
            self.nonc_entry[x].grid_remove()
            self.nonc_del[x].grid_remove()
            self.non_c[x].set("")  # set the value of the stringvar to empty string

        def add_reman_field(self, frame, onrec=None):
            """ added fields for compliance index"""
            add_stringvar = StringVar(self.win.body)
            if onrec:
                add_stringvar.set(onrec)
            self.reman.append(add_stringvar)  # add this to an array of stringvars for remanded
            reman = Entry(frame, textvariable=self.reman[len(self.reman)-1], justify='right', width=macadj(20, 15))
            reman.grid(row=len(self.reman)-1, column=1, sticky="w")
            self.reman_entry.append(reman)  # add this to an array of entry widgets for remanded
            del_ = Button(frame, text="del", width=macadj(4, 3), anchor="center",
                          command=lambda x=len(self.reman)-1: self.del_reman_field(x))
            del_.grid(row=len(self.reman)-1, column=3)
            self.reman_del.append(del_)  # add this to an array of widgets of delete buttons
            # bind the expanding frome to the canvas and scrollregion
            frame.bind('<Configure>', lambda e: self.win.c.configure(scrollregion=self.win.c.bbox("all")))
            self.win.topframe.bind("<Configure>", self.win.detect_resize)  # track when the window changes size

        def del_reman_field(self, x):
            """ delete a field from the remanded entry widgets as well as the delete button.
            set the value of the corresponding stringvar to an empty string. """
            self.reman_entry[x].grid_remove()
            self.reman_del[x].grid_remove()
            self.reman[x].set("")  # set the value of the stringvar to empty string

        def add_batset_field(self, frame, onrec=None):
            """ added fields for compliance index"""
            add_stringvar = StringVar(self.win.body)
            if onrec:
                add_stringvar.set(onrec)
            self.batset.append(add_stringvar)  # add this to an array of stringvars for non compliance
            batset = Entry(frame, textvariable=self.batset[len(self.batset)-1], justify='right', width=macadj(20, 15))
            batset.grid(row=len(self.batset)-1, column=1, sticky="w")
            self.batset_entry.append(batset)  # add this to an array of entry widgets for non compliance
            del_ = Button(frame, text="del", width=macadj(4, 3), anchor="center",
                          command=lambda x=len(self.batset)-1: self.del_batset_field(x))
            del_.grid(row=len(self.batset)-1, column=3)
            self.batset_del.append(del_)  # add this to an array of widgets of delete buttons
            # bind the expanding frome to the canvas and scrollregion
            frame.bind('<Configure>', lambda e: self.win.c.configure(scrollregion=self.win.c.bbox("all")))
            self.win.topframe.bind("<Configure>", self.win.detect_resize)  # track when the window changes size

        def del_batset_field(self, x):
            """ delete a field from the non compliance entry widgets as well as the delete button.
            set the value of the corresponding stringvar to an empty string. """
            self.batset_entry[x].grid_remove()
            self.batset_del[x].grid_remove()
            self.batset[x].set("")  # set the value of the stringvar to empty string

        def add_batgat_field(self, frame, onrec=None):
            """ added fields for compliance index"""
            add_stringvar = StringVar(self.win.body)
            if onrec:
                add_stringvar.set(onrec)
            self.batgat.append(add_stringvar)  # add this to an array of stringvars for non compliance
            batgat = Entry(frame, textvariable=self.batgat[len(self.batgat)-1], justify='right', width=macadj(20, 15))
            batgat.grid(row=len(self.batgat)-1, column=1, sticky="w")
            self.batgat_entry.append(batgat)  # add this to an array of entry widgets for non compliance
            del_ = Button(frame, text="del", width=macadj(4, 3), anchor="center",
                          command=lambda x=len(self.batgat)-1: self.del_batgat_field(x))
            del_.grid(row=len(self.batgat)-1, column=3)
            self.batgat_del.append(del_)  # add this to an array of widgets of delete buttons
            # bind the expanding frome to the canvas and scrollregion
            frame.bind('<Configure>', lambda e: self.win.c.configure(scrollregion=self.win.c.bbox("all")))
            self.win.topframe.bind("<Configure>", self.win.detect_resize)  # track when the window changes size

        def del_batgat_field(self, x):
            """ delete a field from the non compliance entry widgets as well as the delete button.
            set the value of the corresponding stringvar to an empty string. """
            self.batgat_entry[x].grid_remove()
            self.batgat_del[x].grid_remove()
            self.batgat[x].set("")  # set the value of the stringvar to empty string

        def fillin_index_onrecs(self):
            """ this will set the stringvars for associations/index fields by passing an 'onrec' argument to
            'add_ _field()' method for each index. by looping through self.index_onrecs. """
            tables_array = ("informalc_noncindex", "informalc_remandindex",
                            "informalc_batchindex", "informalc_gats")
            for i in range(len(self.index_onrecs)):  # this will loop 4 times. once per table.
                if tables_array[i] == "informalc_noncindex":
                    for rec in self.index_onrecs[i]:  # for each record
                        if not self.non_c[0].get():  # if there is nothing in the first stringvar
                            self.non_c[0].set(rec[1])  # add this to an array of stringvars for non compliance
                        else:
                            self.add_nonc_field(self.nonc_frame, onrec=rec[1])  # build a field and a del button
                if tables_array[i] == "informalc_remandindex":
                    for rec in self.index_onrecs[i]:  # for each record
                        if not self.reman[0].get():  # if there is nothing in the first stringvar
                            self.reman[0].set(rec[1])  # add this to an array of stringvars for non compliance
                        else:
                            self.add_reman_field(self.reman_frame, onrec=rec[1])  # build a field and a del button
                if tables_array[i] == "informalc_batchindex":
                    for rec in self.index_onrecs[i]:  # for each record
                        if not self.batset[0].get():  # if there is nothing in the first stringvar
                            self.batset[0].set(rec[1])  # add this to an array of stringvars for non compliance
                        else:
                            self.add_batset_field(self.batset_frame, onrec=rec[1])  # build a field and a del button
                if tables_array[i] == "informalc_gats":
                    for rec in self.index_onrecs[i]:  # for each record
                        if not self.batgat[0].get():  # if there is nothing in the first stringvar
                            self.batgat[0].set(rec[1])  # add this to an array of stringvars for non compliance
                        else:
                            self.add_batgat_field(self.batgat_frame, onrec=rec[1])  # build a field and a del button

        def informalc_root(self, mode, grv_no=None):
            """ creates a companion window for selecting carrier names.
            mode is 'selectcarrier' or 'selectissue'. grv_no is used for editing grievance information. """
            # self.destroy_companion()  # destroy other companion windows if they exist
            self.companion_root = Tk()
            self.companion_root.title("KLUSTERBOX")
            titlebar_icon(self.companion_root)  # place icon in titlebar
            x_position = projvar.root.winfo_x() + 450
            y_position = projvar.root.winfo_y() - 25
            self.companion_root.geometry("%dx%d+%d+%d" % (240, 600, x_position, y_position))
            topframe = Frame(self.companion_root)
            topframe.pack()
            buttons = Canvas(topframe)  # button bar
            buttons.pack(fill=BOTH, side=BOTTOM)
            text = "Add Issue"
            if mode == 'selectcarrier':
                text = "Add Carrier"
            if mode == 'selectissue':
                text = "Add Issue"
            if mode == 'selectdecision':
                text = "Add Decision"
            Label(topframe, text=text, font=macadj("bold", "Helvetica 18")).pack(anchor="w")
            Label(topframe, text="").pack()
            scrollbar = Scrollbar(topframe, orient=VERTICAL)
            listbox = Listbox(topframe, selectmode="browse", yscrollcommand=scrollbar.set)
            listbox.config(height=100, width=50)
            if mode == 'selectcarrier':  # use a list of carriers.
                self.get_listbox_carrriers()
            if mode == 'selectissue':  # use a list of issue from self.issue_description
                self.listbox_fill = self.issue_description[:]
            if mode == 'selectdecision':  # use a list of issues from self.decision_description
                self.listbox_fill = self.decision_description[:]
            for name in self.listbox_fill:  # fill the listbox
                listbox.insert(END, name)
            scrollbar.config(command=listbox.yview)
            scrollbar.pack(side=RIGHT, fill=Y)
            listbox.pack(side=LEFT, expand=1)
            if mode == 'selectcarrier':
                Button(buttons, text="Add Carrier", width=macadj(10, 10),
                       command=lambda: (self.addnames(listbox.curselection()))).pack(side=LEFT, anchor="w")
            if mode == 'selectissue':
                Button(buttons, text="Add Issue", width=macadj(10, 10),
                       command=lambda: (self.addissue(listbox.curselection()))).pack(side=LEFT, anchor="w")
            if mode == 'selectdecision':
                Button(buttons, text="Add Decision", width=macadj(10, 10),
                       command=lambda: (self.adddecision(listbox.curselection()))).pack(side=LEFT, anchor="w")
            # to destroy and re create itself.
            Button(buttons, text="Clear", width=macadj(10, 5),
                   command=lambda: (self.destroy_companion(), self.informalc_root(mode, grv_no))) \
                .pack(side=LEFT, anchor="w")
            Button(buttons, text="Close", width=macadj(10, 9),
                   command=lambda: (self.destroy_companion())).pack(side=LEFT, anchor="w")

        def get_listbox_carrriers(self):
            """ get a list of carriers for the listbox in self.informalc_root. if no grv_no is given, the
            carrier list for year 1000 AD through 9000 AD will be the range of the carrier list. """
            start = '1000-01-01 00:00:00'
            end = '9000-01-01 00:00:00'
            start = dt_converter(start)
            end = dt_converter(end)
            # get a list of carriers given the search criteria.
            self.listbox_fill = ["class action", ] + informalc_gen_clist(start, end, self.parent.station)

        def addnames(self, listbox):
            """ sets the grievant field by setting the stringvar self.grievant using an index from the
             listbox and an array generated in informalc root. """
            for index in listbox:
                self.grievant.set(self.listbox_fill[index])

        def addissue(self, listbox):
            """ sets the issue field by setting the stringvar self.issue using an index from the
             listbox and an array generated in informalc root. """
            for index in listbox:
                self.issue.set(self.listbox_fill[index])
                self.article.set(self.issue_article[index])

        def adddecision(self, listbox):
            """ sets the decision field by setting the stringvar self.decision using an index from the
             listbox and an array generated in informalc root. """
            for index in listbox:
                self.decision.set(self.listbox_fill[index])

        def destroy_companion(self):
            """ exit out of a screen with a companion root. Destroy the companion window if it still exist. """
            try:
                self.companion_root.destroy()  # destroy the tkinter root object
                self.companion_root = None  # re initialize the variable
            except (TclError, AttributeError):
                pass

        def settlement_delete(self):
            """ delete the settlement """
            if not messagebox.\
                    askokcancel("Grievance Number Change",
                                "This will delete the settlement for  grievance number {} to  in all records. "
                                "Are you sure you want to proceed?".format(self.edit_grv_no),
                                parent=self.win.topframe):
                return
            # use loops and arrays to commit changes to db
            tables = ("informalc_batchindex", "informalc_gats", "informalc_awards2", "informalc_settlements")
            fields = ("main", "grv_no", "grv_no", "grv_no")
            for i in range(4):
                sql = "DELETE FROM %s WHERE %s='%s'" % (tables[i], fields[i], self.edit_grv_no)
                commit(sql)
            self.parent.refresh_search(self.win.topframe)
            self.parent.showtime(self.win.topframe)

        def grvchange(self, old_number, new_stringvar):
            """ change the grievance number. check grv number and input it into the informalc tables. """
            new_number = new_stringvar.get()
            if not messagebox.\
                    askokcancel("Grievance Number Change",
                                "This will change the grievance number from {} to {} in all records. "
                                "Are you sure you want to proceed?".format(old_number, new_number),
                                parent=self.win.topframe):
                return
            if not self.checking_grv_number(chg_number=True, new_number=new_number):
                return  # make sure the new number passes standards
            new_number = self.reformat_grv_no(chg_number=True, new_number=new_number)  # reformat new number
            # use loops and arrays to commit changes to db
            tables = ("informalc_batchindex", "informalc_batchindex", "informalc_gats", "informalc_gats",
                      "informalc_grievances", "informalc_noncindex", "informalc_noncindex", "informalc_remandindex",
                      "informalc_remandindex", "informalc_awards2", "informalc_settlements")
            fields = ("main", "sub", "grv_no", "gats_no", "grv_no", "overdue", "followup", "remanded", "refiling",
                      "grv_no", "grv_no")
            for i in range(11):
                sql = "UPDATE %s SET %s = '%s' WHERE %s = '%s'" % \
                      (tables[i], fields[i], new_number, fields[i], old_number)
                commit(sql)
            l_passed_result = [list(x) for x in self.parent.search_result]  # chg tuple of tuples to list of lists
            for record in l_passed_result:
                if record[0] == old_number:
                    record[0] = new_number
            self.parent.search_result = l_passed_result[:]
            msg = "Grievance number changed."
            self.informalc_edit(self.win.topframe, new_number, msg=msg)

        def delete_grievance(self, frame, grv_no):
            """ deletes a record and associated records for a grievance. """
            if not messagebox.askokcancel("Delete Grievance",
                                          "Are you sure you want to delete his grievance and all the "
                                          "data associated with it?",
                                          parent=self.win.topframe):
                return
            # use loops and arrays to commit changes to db
            tables = ("informalc_batchindex", "informalc_batchindex", "informalc_gats", "informalc_gats",
                      "informalc_grievances", "informalc_noncindex", "informalc_noncindex", "informalc_remandindex",
                      "informalc_remandindex", "informalc_awards2", "informalc_settlements")
            fields = ("main", "sub", "grv_no", "gats_no", "grv_no", "followup", "overdue", "refiling", "remanded",
                      "grv_no", "grv_no")
            for i in range(11):
                sql = "DELETE FROM %s WHERE %s='%s'" % (tables[i], fields[i], grv_no)
                commit(sql)
            # delete grievance from self.parent.search_result
            self.parent.refresh_search(self.win.topframe)
            self.parent.showtime(frame)

        def apply(self, goback=False):
            """
            check the inputs one by one. if there are any errors, a messagebox will show the error and the checks will
            stop. otherwise all inputs are put into a variable to be saved for entry into the database. .
            """
            # get the values from the stringvars
            self.check_grievant = self.grievant.get()  # 1
            self.check_grv_no = self.grv_no.get()  # 2
            self.check_startdate = self.startdate.get()  # 3
            self.check_enddate = self.enddate.get()  # 4
            self.check_meetingdate = self.meetingdate.get()  # 5
            self.check_issue = self.issue.get()  # 6
            self.check_article = self.article.get()  # 7
            self.check_lvl = self.lvl.get()  # 10 the level of the settlement
            self.check_datesigned = self.date_signed.get()  # 11 the date the settlement was signed
            self.check_decision = self.decision.get()  # 12 the decision of the settlement
            self.check_proofdue = self.proof_due.get()  # 13 the date that the prooof of the remedy is due
            self.check_docs = self.docs.get()  # 14 the status of any documentation needed for proof of compliance
            # get stringvar values for indexes and place them in a multidimentional list
            # - non compliance, remanded, batch settlements and gats reports
            if self.newentry:  # get onrecs to ensure that existing grievances aren't added to db.
                self.get_onrecs(self.check_grv_no)
            for index in [self.non_c, self.reman, self.batset, self.batgat]:
                array = []
                for element in index:
                    if element.get():  # only append values that are not empty
                        to_add = element.get()
                        to_add = to_add.strip().lower()
                        array.append(to_add)
                self.check_indexes.append(array)
            if not self.checking_grievant():  # check input for errors
                self.apply_fail()  # empty the message in the button frame
                return
            if not self.checking_grv_number():  # check input for errors
                self.apply_fail()  # empty the message in the button frame
                return
            if not self.checking_dates():  # check input for errors
                self.apply_fail()  # empty the message in the button frame
                return
            if not self.checking_issue():  # check input for errors
                self.apply_fail()  # empty the message in the button frame
                return
            if not self.checking_article():  # check input for errors
                self.apply_fail()  # empty the message in the button frame
                return
            # since level, decision and docs values come from an option menu, there is no need to check them.
            # entries of 'no status' and 'no decision' will be converted to empty strings
            self.checking_optionmenus()
            if not self.checking_indexes():
                self.check_indexes = []  # re initialize array for indexes
                self.apply_fail()  # empty the message in the button frame
                return
            self.get_grv_changesmade()  # check if changes necessitate an update
            self.get_set_changesmade()  # check if changes necessitate an update
            self.get_index_changesmade()  # check if changes necessitate an update
            self.add_grv_recs()  # add grievance recs to db
            self.add_set_recs()  # add settlement recs to db
            self.add_index_recs()  # add or delete index recs to the db
            # after the database is updated, regenerate the page.
            if self.newentry:
                if not goback:
                    self.informalc_new(self.win.topframe, msg=self.report())
                else:
                    self.parent.informalc(self.win.topframe)
            else:  # if coming from the Edit Grievance screen
                if self.grv_changesmade or self.set_changesmade:  # if changes have been made
                    self.grv_changesmade, self.set_changesmade = False, False  # re initialize
                    self.parent.refresh_search(self.win.topframe, reroute=False)  # refresh search results
                if not goback:
                    self.informalc_edit(self.win.topframe, self.edit_grv_no, msg=self.report())
                else:
                    self.parent.showtime(self.win.topframe, turnpage=True)

        def apply_fail(self):
            """ if there is an error in self.apply, empty the message in the button frame. """
            self.msg_label.config(text="")
            projvar.root.update()

        def checking_grievant(self):
            """ a method for checking the grievant.  check the grievant input. this is either 'class action'
            or a carrier name. it can be blank. """
            self.check_grievant = self.check_grievant.lower().strip()
            if not self.check_grievant or self.check_grievant == "class action":  # if empty or class action
                return True  # skip checks
            if not NameChecker(self.check_grievant).check_characters():
                msg = "Grievant name can not contain numbers or most special characters\n"
                messagebox.showerror("Invalid Data Entry", msg, parent=self.win.topframe)
                return False
            if not NameChecker(self.check_grievant).check_length():
                msg = "Grievant name must not exceed 42 characters\n"
                messagebox.showerror("Invalid Data Entry", msg, parent=self.win.topframe)
                return False
            if not NameChecker(self.check_grievant).check_comma():
                msg = "Grievant name must contain one comma to separate last name and first initial\n"
                messagebox.showerror("Invalid Data Entry", msg, parent=self.win.topframe)
                return False
            if not NameChecker(self.check_grievant).check_initial():
                msg = "Grievant name should must contain one initial ideally, \n" \
                     "unless more are needed to create a distinct carrier name.\n"
                messagebox.showerror("Invalid Data Entry", msg, parent=self.win.topframe)
            return True

        def checking_grv_number(self, chg_number=False, new_number=""):
            """ check the grievance number input.
            chg_number is evoked by grvchange to check the grv number before it is changed. """
            grievance_number = self.check_grv_no
            if chg_number:
                grievance_number = new_number
            if chg_number:
                sql = "SELECT * FROM informalc_grievances WHERE grv_no = '%s'" % grievance_number
                result = inquire(sql)
                if result:
                    msg = "This grievance number is already being used in the database."
                    messagebox.showerror("Invalid Data Entry", msg, parent=self.win.topframe)
                    return False
            if self.newentry and self.onrec_grievance:
                msg = "There is already a record for this grievance in the database."
                messagebox.showerror("Invalid Data Entry", msg, parent=self.win.topframe)
                return False
            if not GrievanceChecker(grievance_number).has_value():
                msg = "The grievance number must not be blank."
                messagebox.showerror("Invalid Data Entry", msg, parent=self.win.topframe)
                return False
            if not GrievanceChecker(grievance_number).check_characters():
                msg = "The grievance number can only contain numbers and letters."
                messagebox.showerror("Invalid Data Entry", msg, parent=self.win.topframe)
                return False
            if not GrievanceChecker(grievance_number).min_lenght():
                msg = "The grievance number must contain at least 4 characters."
                messagebox.showerror("Invalid Data Entry", msg, parent=self.win.topframe)
                return False
            if not GrievanceChecker(grievance_number).max_lenght():
                msg = "The grievance number can not contain more than 20 characters."
                messagebox.showerror("Invalid Data Entry", msg, parent=self.win.topframe)
                return False
            if not chg_number:
                self.reformat_grv_no()  # clean up the grievance number for database entry
            return True

        def reformat_grv_no(self, chg_number=False, new_number=""):
            """ reformat the grievance number to all lowercase, no whitespaces, no dashes.
            this is modified to handle new grv numbers from grvchange"""
            grievance_number = self.check_grv_no
            if chg_number:
                grievance_number = new_number
            grievance_number = grievance_number.lower()  # convert grievance number to lowercas
            grievance_number = grievance_number.strip()  # strip whitespace from start and end of the string.
            grievance_number = grievance_number.replace('-', '')  # remove any dashes
            grievance_number = grievance_number.replace(' ', '')  # remove any whitespace
            if chg_number:
                return grievance_number
            else:
                self.check_grv_no = grievance_number

        def marry_startend_dates(self):
            """ since we do not want blank start or end dates """

        def checking_dates(self):
            """ check the startdate, enddate and meetingdate.
             since these are all dates with similiar criteria, use a loop to check them.
             sometimes, openpyxl sends the dates as strings of datetime objects, instead of the mm/dd/yyyy formated
             dates, the DateTimeChecker() will identify these and skip the checks. """
            self.check_dates = [self.check_startdate, self.check_enddate, self.check_meetingdate,
                                self.check_datesigned, self.check_proofdue]
            for i in range(5):
                if not self.check_date_loop(i):
                    return False
            return True

        def check_date_loop(self, i):
            """ loop from check dates """
            _type = ("incident start", "incident end", "meeting", "signed", "proof due")
            if self.check_dates[i].strip() == "":  # if the value is blank, skip all the checks
                return True
            # if the value is a valid dt object, skip all the checks
            if DateTimeChecker().check_dtstring(self.check_dates[i]):
                return True
            date_object = BackSlashDateChecker(self.check_dates[i])  # first create the date_object
            if not date_object.count_backslashes():  # this checks that there are 2 backslashes in the date
                msg = "The date for the {} date must have two backslashes. Got instead: {}\n" \
                    .format(_type[i], self.check_dates[i])
                messagebox.showerror("Invalid Data Entry", msg, parent=self.win.topframe)
                return False
            date_object.breaker()  # this breaks the object into month, day and year elements.
            if not date_object.check_numeric():  # check each element in the date to ensure they are numeric
                msg = "The month, day and year for the {} date must be numeric\n".format(_type[i])
                messagebox.showerror("Invalid Data Entry", msg, parent=self.win.topframe)
                return False
            if not date_object.check_minimums():  # check each element in the date to ensure they are greater than zero
                msg = "The month, day and year for the {} date must be greater than zero.\n" \
                    .format(_type[i])
                messagebox.showerror("Invalid Data Entry", msg, parent=self.win.topframe)
                return False
            if not date_object.check_month():  # returns False if the month is greater than 12.
                msg = "The month for the {} date must less than 13.\n".format(_type[i])
                messagebox.showerror("Invalid Data Entry", msg, parent=self.win.topframe)
                return False
            if not date_object.check_day():  # return False if the day is greater than 31.
                msg = "The day entered for the {} date is must be less than 32.\n".format(_type[i])
                messagebox.showerror("Invalid Data Entry", msg, parent=self.win.topframe)
                return False
            if not date_object.check_year():  # returns False if the year does not have 4 digits.
                msg = "The year entered for the {} date must have 4 digits.\n".format(_type[i])
                messagebox.showerror("Invalid Data Entry", msg, parent=self.win.topframe)
                return False
            if not date_object.valid_date():  # returns False if the date is not a valid date
                msg = "The date entered for the {} date is not a valid date.\n".format(_type[i])
                messagebox.showerror("Invalid Data Entry", msg, parent=self.win.topframe)
                return False
            # this removes white space from the date and each element of the date.
            self.check_dates[i] = self.reformat_date(i)
            # convert the input date into a string of a datetime object.
            self.check_dates[i] = Convert(self.check_dates[i]).backslashdate_to_dtstring()
            return True

        def reformat_date(self, i):
            """ this removes white space from the date and each element of the date. """
            breakdown = self.check_dates[i].strip()
            breakdown = breakdown.split("/")
            month = breakdown[0].strip()
            day = breakdown[1].strip()
            year = breakdown[2].strip()
            return "{}/{}/{}".format(month, day, year)

        def checking_issue(self):
            """ check the issue input """
            self.check_issue = self.check_issue.strip()  # strip out any whitespace before or after the string
            if self.check_issue == "":  # accept blank entries
                return True
            # check if issue is in list of issues, if so update the article
            if self.issue in self.issue_description:
                index = self.issue_description.index(self.check_issue)
                self.check_article = self.issue_article[index]
            return True

        def checking_article(self):
            """ check the article input """
            self.check_article = self.check_article.strip()
            if not self.check_article:
                return True
            if not isint(self.check_article):
                msg = "The number the article must be a whole number. Got: {}\n".format(self.issue)
                messagebox.showerror("Invalid Data Entry", msg, parent=self.win.topframe)
                return False
            return True

        def checking_optionmenus(self):
            """ since level, decision and docs are option menus, just covert 'no status' entry into an empty string """
            if self.check_lvl == "no status":
                self.check_lvl = ""
            if self.check_decision == "no decision":
                self.check_decision = ""
            if self.check_docs == "no status":
                self.check_docs = ""

        def checking_indexes(self):
            """ check all of the indexes. indexes are called 'associations' in the gui."""

            def grv_exist(check_this):
                """ check to against the informalc_grievances table to verify grievance is in the db. """
                sql = "SELECT COUNT(*) FROM informalc_grievances WHERE grv_no = '%s'" % check_this
                result = inquire(sql)
                if not result[0][0]:  # if there is not a record of the grievance in the grievances table
                    return False
                return True
            indexes = ("non compliance", "remanded", "batch settlements", "gats reports")
            for i in range(4):  # loop once for each of the index tables
                is_gats = False
                if i == 3:
                    is_gats = True
                mentioned_grv = []
                for ii in range(len(self.check_indexes[i])):
                    if self.check_indexes[i][ii] == self.check_grv_no:
                        msg = "The {} grievance number for {} association can not be identical to the grievance " \
                              "number being entered/edited: {}\n" \
                            .format(Handler(ii + 1).make_ordinal(), indexes[i], self.edit_grv_no)
                        messagebox.showerror("Invalid Data Entry", msg, parent=self.win.topframe)
                        return False
                    if not is_gats:  # do not do this check for gats numbers.
                        if not grv_exist(self.check_indexes[i][ii]):  # check if grv_no is in the db.
                            msg = "There is no record of the {} grievance number for {} association: {}\n" \
                                  "Grievances in the associations must first be entered as grievances."\
                                .format(Handler(ii + 1).make_ordinal(), indexes[i], self.check_indexes[i][ii])
                            messagebox.showerror("Invalid Data Entry", msg, parent=self.win.topframe)
                            return False
                    if self.check_indexes[i][ii] in mentioned_grv:
                        number_type = "grievance"
                        if is_gats:
                            number_type = "gats"  # change text message if checking gats numbers
                        msg = "The {} {} number for {} association was entered more than once: {}. " \
                              "Duplicates are not allowed. \n" \
                            .format(Handler(ii + 1).make_ordinal(), number_type, indexes[i], self.edit_grv_no)
                        messagebox.showerror("Invalid Data Entry", msg, parent=self.win.topframe)
                        return False
                    mentioned_grv.append(self.check_indexes[i][ii])  # add grv num to a list of mentioned grv numbers.
            return True

        def get_grv_changesmade(self):
            """ this will determine if the sql commit should be an update or an insert. """
            chg_these = []
            # get grievant place
            if self.check_grv_no != self.edit_grv_no:
                chg_these.append("grievance number")
            if self.check_grievant != self.onrec_grievant:  # check grievant
                chg_these.append("grievant")
            # get date places using loop
            onrec_date = [self.onrec_startdate, self.onrec_enddate, self.onrec_meetingdate]
            chg_notation = ("startdate", "enddate", "meetingdate")
            for i in range(0, 3):  # only check the first three elements - startdate, enddate and meetingdate
                if self.check_dates[i] != onrec_date[i]:
                    chg_these.append(chg_notation[i])
            if self.check_issue != self.onrec_issue:
                chg_these.append("issue")
            if self.check_article != self.onrec_article:
                chg_these.append("article")
            if len(chg_these):  # if change these is not empty
                self.grv_changesmade = True  # then update status is True.

        def get_set_changesmade(self):
            """ this will determine if the sql commit should be an update or an insert.). """
            chg_these = []
            if self.check_lvl != self.onrec_lvl:  # check level
                chg_these.append("level")
            # get date places using loop
            check_dates = [self.check_dates[3], self.check_dates[4]]  # input from date signed and proof due.
            onrec_date = [self.onrec_date_signed, self.onrec_proof_due]
            chg_notation = ("date signed", "proof due")
            for i in range(2):  # only check the last two elements - date signed and proof due.
                if check_dates[i] != onrec_date[i]:
                    chg_these.append(chg_notation[i])
            if self.check_decision != self.onrec_decision:  # check decision
                chg_these.append("decision")
            if self.check_docs != self.onrec_docs:  # check docs
                chg_these.append("docs")
            if len(chg_these):  # if change these is not empty
                self.set_changesmade = True  # then update status is True.

        def get_index_changesmade(self):
            """ this will determine if the sql commit should be an update or a delete) """
            # create a list of secondary values of the onrec pair
            secondary_onrecs = []
            for i in range(len(self.index_onrecs)):  # loop through nonc, remand, batset and batgat
                array = []  # initialize
                for ii in range(len(self.index_onrecs[i])):  # loop through all onrecs from the table
                    array.append(self.index_onrecs[i][ii][1])  # capture all 2nd values of the pair
                secondary_onrecs.append(array)  # store all results in an array with 4 arrays inside.
            master_set = []
            for i in range(4):
                array = list(set(secondary_onrecs[i] + self.check_indexes[i]))
                master_set.append(array)
            for i in range(len(master_set)):  # loop through the 4 arrays in the master set
                add_array = []  # initialize array for adding recs
                del_array = []  # initialize array for deleting recs
                for rec in master_set[i]:  # loop through all grievance numbers in the array
                    if rec not in secondary_onrecs[i]:
                        add_array.append(rec)
                    if rec not in self.check_indexes[i]:
                        del_array.append(rec)
                self.add_indexes.append(add_array)
                self.del_indexes.append(del_array)

        def add_grv_recs(self):
            """ insert, update or ignore record for grievance table. """
            if self.newentry and self.grv_changesmade:  # if this is a new entry...
                self.insert_grv()  # insert a record into the grievance database
            elif not self.newentry and self.grv_changesmade:  # if this is an edited grievance with changes...
                self.update_grv()  # update the record in the grievance table
            else:  # if this is an edited grievance with no changes...
                self.msg = "NO INPUT: Grievance Not Added."
                # self.informalc_new(self.win.topframe)

        def insert_grv(self):
            """ insert a record into the grievance table """
            sql = "INSERT INTO informalc_grievances (grievant, station, grv_no, startdate, enddate, " \
                  "meetingdate, issue, article) " \
                  "VALUES('%s','%s','%s','%s','%s','%s','%s','%s')" % \
                  (self.check_grievant, self.parent.station, self.check_grv_no, self.check_dates[0],
                   self.check_dates[1], self.check_dates[2], self.check_issue, self.check_article)
            commit(sql)
            self.reporter_grv = True

        def update_grv(self):
            """ update the record in the grievance table """
            sql = "UPDATE informalc_grievances SET grievant = '%s', startdate = '%s', enddate = '%s', " \
                  "meetingdate = '%s', issue = '%s', article = '%s' WHERE grv_no = '%s'" % \
                  (self.check_grievant, self.check_dates[0], self.check_dates[1], self.check_dates[2],
                   self.check_issue, self.check_article, self.check_grv_no)
            commit(sql)
            self.reporter_grv = True

        def add_set_recs(self):
            """ insert, update or ignore records for the settlement table """
            # if there is no record for this settlement and changes have been made..
            if not self.onrec_settlement and self.set_changesmade:
                self.insert_set()  # insert a record into the settlement table
            # if there is a pre-existing settlement record and changes have been made...
            elif self.onrec_settlement and self.set_changesmade:
                self.update_set()

        def insert_set(self):
            """ insert a record into the settlement table """
            sql = "INSERT INTO informalc_settlements (grv_no, level, date_signed, decision, proofdue, " \
                  "docs) " \
                  "VALUES('%s','%s','%s','%s','%s','%s')" % \
                  (self.check_grv_no, self.check_lvl, self.check_dates[3], self.check_decision,
                   self.check_dates[4], self.check_docs)
            commit(sql)
            self.reporter_set = True

        def update_set(self):
            """ update a record in the settlement table """
            sql = "UPDATE informalc_settlements SET level = '%s', date_signed = '%s', decision = '%s', " \
                  "proofdue = '%s', docs = '%s' WHERE grv_no = '%s'" % \
                  (self.check_lvl, self.check_dates[3], self.check_decision, self.check_dates[4],
                   self.check_docs, self.check_grv_no)
            commit(sql)
            self.reporter_set = True

        def add_index_recs(self):
            """  insert, delete or ignore records for the indexes, no recs are updated. there is only
            inserting and deleting. """
            tables = ("informalc_noncindex", "informalc_remandindex", "informalc_batchindex", "informalc_gats")
            index_columns = [
                ["followup", "overdue"],  # non compliance index
                ["refiling", "remanded"],  # remanded index
                ["main", "sub"],  # batch settlement index
                ["grv_no", "gats_no"]]  # gats reports index
            for i in range(len(self.add_indexes)):  # this will loop 4 times. once for each index
                for rec in self.add_indexes[i]:
                    # check if the record already exist in the database
                    sql = "SELECT * FROM %s WHERE %s = '%s' AND %s = '%s'" % \
                          (tables[i], index_columns[i][0], self.check_grv_no, index_columns[i][1], rec)
                    if not inquire(sql):  # if there is no result
                        sql = "INSERT INTO %s(%s, %s) VALUES('%s','%s')" % \
                              (tables[i], index_columns[i][0], index_columns[i][1], self.check_grv_no, rec)
                        commit(sql)
                        self.reporter_index = True
                for rec in self.del_indexes[i]:
                    sql = "DELETE FROM %s WHERE %s='%s' and %s='%s'" \
                          % (tables[i], index_columns[i][0], self.check_grv_no, index_columns[i][1], rec)
                    commit(sql)
                    self.reporter_index = True

        def report(self):
            """ this will generate a message showing inserts/updates/ deletions. """
            if self.newentry:
                if any((self.reporter_grv, self.reporter_set, self.reporter_index)):
                    return "Grievance {} entered".format(self.check_grv_no)
            else:
                if any((self.reporter_grv, self.reporter_set)):
                    return "Grievance {} updated".format(self.edit_grv_no)
                if self.reporter_set:
                    return "Settlement {} entered/updated".format(self.edit_grv_no)
                if self.reporter_index:
                    return "Index for {} entered/updated".format(self.edit_grv_no)

    class PayoutEntry:
        """
        this class allows users to enter carrier payout
        """

        def __init__(self, parent):
            self.parent = parent
            self.win = None
            self.add_pay_periods = None
            self.add_hours = None
            self.add_rate = None
            self.add_amount = None
            self.poe_listbox_ = None  # holds the list box object

        def poe_search(self, frame):
            """ creates a screen that allows user to update payouts for carriers. """
            self.win = MakeWindow()
            self.win.create(frame)
            the_year = StringVar(self.win.topframe)
            the_station = StringVar(self.win.topframe)
            station_options = projvar.list_of_stations
            if "out of station" in station_options:
                station_options.remove("out of station")
            the_station.set("undefined")
            backdate = StringVar(self.win.topframe)
            backdate.set("1")
            Label(self.win.body, text="Informal C: Payout Entry Criteria", font=macadj("bold", "Helvetica 18")) \
                .grid(row=0, column=0, sticky="w", columnspan=4)
            Label(self.win.body, text="").grid(row=1)
            Label(self.win.body, text="Enter the year and the station to be updated.") \
                .grid(row=2, column=0, columnspan=4, sticky="w")
            Label(self.win.body, text="\t\t\tYear: ").grid(row=3, column=1, sticky="e")
            Entry(self.win.body, textvariable=the_year, width=12).grid(row=3, column=2, sticky="w")
            Label(self.win.body, text="Station").grid(row=4, column=1, sticky="e")
            om_station = OptionMenu(self.win.body, the_station, *station_options)
            om_station.config(width=28)
            om_station.grid(row=4, column=2, columnspan=2)
            Label(self.win.body, text="Build the carrier list by going back how many years?") \
                .grid(row=5, column=0, columnspan=4, sticky="w")
            Label(self.win.body, text="Back Date: ").grid(row=6, column=1, sticky="w")
            om_backdate = OptionMenu(self.win.body, backdate, "1", "2", "3", "4", "5", "6", "7", "8", "9", "10")
            om_backdate.config(width=5)
            om_backdate.grid(row=6, column=2, sticky="w")
            button_alignment = macadj("w", "center")
            Button(self.win.buttons, text="Go Back", width=20, anchor=button_alignment,
                   command=lambda: self.parent.informalc(self.win.topframe)) \
                .grid(row=0, column=1, sticky="w")
            Button(self.win.buttons, text="Apply", width=20, anchor=button_alignment,
                   command=lambda: self.apply_search(self.win.topframe, the_year,
                                                     the_station, backdate)) \
                .grid(row=0, column=2, sticky="w")
            self.win.finish()

        def apply_search(self, frame, year, station, backdate):
            """ pay out entry - search the args for acceptable values. """
            if year.get().strip() == "":
                messagebox.showerror("Data Entry Error",
                                     "You must enter a year.",
                                     parent=self.win.topframe)
                return
            if "." in year.get():
                messagebox.showerror("Data Entry Error",
                                     "The year can not contain decimal points.",
                                     parent=self.win.topframe)
                return
            if not year.get().isnumeric():
                messagebox.showerror("Data Entry Error",
                                     "The year must numeric without any letters or special characters.",
                                     parent=self.win.topframe)
                return
            if float(year.get()) > 9999 or float(year.get()) < 2:
                messagebox.showerror("Data Entry Error",
                                     "The year must be between the year 2 and 9999.\nI think I'm being "
                                     "reasonable.",
                                     parent=frame)
                return
            if station.get() == "undefined":
                messagebox.showerror("Data Entry Error",
                                     "You must select a station.",
                                     parent=frame)
                return
            weeks = int(backdate.get()) * 52
            dt_year = datetime(int(year.get()), int(1), int(1))
            dt_start = dt_year - timedelta(weeks=weeks)
            year = year.get()
            array = []
            selection = "none"
            msg = ""
            self.poe_listbox(dt_year, station, dt_start, year)
            self.add(frame, array, selection, year, msg)

        def apply_add(self, frame, name, year, buttons):
            """ payout entry - apply changes """
            if name == "none":
                messagebox.showerror("Data Entry Error",
                                     "You must select a name.",
                                     parent=self.win.topframe)
                return
            for i in range(len(self.add_pay_periods)):
                pp = self.add_pay_periods[i].get().strip()
                hr = self.add_hours[i].get().strip()
                rt = self.add_rate[i].get().strip()
                amt = self.add_amount[i].get().strip()
                if pp and not isint(pp):
                    messagebox.showerror("Data Input Error",
                                         "Input error for {} in row {}. The pay period must be a number"
                                         .format(name, str(i + 1)),
                                         parent=self.win.topframe)
                    return
                if pp and int(pp) > 27:
                    messagebox.showerror("Data Input Error",
                                         "Input error for {} in row {}. The pay period can not be greater "
                                         "than 27".format(name, str(i + 1)),
                                         parent=self.win.topframe)
                    return
                if hr and amt:
                    messagebox.showerror("Data Input Error",
                                         "Input error for {} in row {}. You can not enter both hours and "
                                         "amount. You can only enter one or another, but not both. "
                                         "Awards can be in the form of "
                                         "hours at a given rate OR an amount.".format(name, str(i + 1)),
                                         parent=self.win.topframe)
                    return
                if rt and amt:
                    messagebox.showerror("Data Input Error",
                                         "Input error for {} in row {}. You can not enter both rate and "
                                         "amount. You can only enter one or another, but not both. "
                                         "Awards can be in the form of "
                                         "hours at a given rate OR an amount.".format(name, str(i + 1)),
                                         parent=self.win.topframe)
                    return
                if hr and not rt:
                    messagebox.showerror("Data Input Error",
                                         "Input error for {} in row {}. Hours must be a accompanied by a "
                                         "rate.".format(name, str(i + 1)),
                                         parent=self.win.topframe)
                    return
                if rt and not hr:
                    messagebox.showerror("Data Input Error",
                                         "Input error for {} in row {}. Rate must be a accompanied by a "
                                         "hours.".format(name, str(i + 1)),
                                         parent=self.win.topframe)
                    return
                if hr and not isfloat(hr):
                    messagebox.showerror("Data Input Error", "Input error for {} in row {}. Hours must be a number."
                                         .format(name, str(i + 1)),
                                         parent=self.win.topframe)
                    return
                if hr and '.' in hr:
                    s_hrs = hr.split(".")
                    if len(s_hrs[1]) > 2:
                        messagebox.showerror("Data Input Error",
                                             "Input error for {} in row {}. Hours must have no "
                                             "more than 2 decimal places.".format(name, str(i + 1)),
                                             parent=self.win.topframe)
                        return
                if rt and amt:
                    messagebox.showerror("Data Input Error",
                                         "Input error for {} in row {}. You can not enter both rate and "
                                         "amount. You can only enter one or the other, but not both. "
                                         "Awards can be in the form of "
                                         "hours at a given rate OR an amount.".format(name, str(i + 1)),
                                         parent=self.win.topframe)
                    return
                if rt and not isfloat(rt):
                    messagebox.showerror("Data Input Error",
                                         "Input error for {} in row {}. Rate must be a number."
                                         .format(name, str(i + 1)),
                                         parent=self.win.topframe)
                    return
                if rt and '.' in rt:
                    s_rate = rt.split(".")
                    if len(s_rate[1]) > 2:
                        messagebox.showerror("Data Input Error",
                                             "Input error for {} in row {}. Rates must have no "
                                             "more than 2 decimal places.".format(name, str(i + 1)),
                                             parent=self.win.topframe)
                        return
                if rt and float(rt) > 10:
                    messagebox.showerror("Data Input Error",
                                         "Input error for {} in row {}. Values greater than 10 are not "
                                         "accepted. \n"
                                         "Note the following rates would be expressed as: \n "
                                         "additional %50         .50 or just .5 \n"
                                         "straight time rate     1.00 or just 1 \n"
                                         "overtime rate          1.50 or 1.5 \n"
                                         "penalty rate           2.00 or just 2".format(name, str(i + 1)),
                                         parent=self.win.topframe)
                    return
                if amt and not isfloat(amt):
                    messagebox.showerror("Data Input Error",
                                         "Input error for {} in row {}. Amounts can only be expressed as "
                                         "numbers. No special characters, such as $ are allowed."
                                         .format(name, str(i + 1)),
                                         parent=self.win.topframe)
                    return
                if amt and '.' in amt:
                    s_amt = amt.split(".")
                    if len(s_amt[1]) > 2:
                        messagebox.showerror("Data Input Error",
                                             "Input error for {} in row {}. Amounts must have no "
                                             "more than 2 decimal places.".format(name, str(i + 1)),
                                             parent=self.win.topframe)
                        return
            pb_label = Label(buttons, text="Updating Changes: ")  # make label for progress bar
            pb_label.grid(row=1, column=2)
            pb = ttk.Progressbar(buttons, length=200, mode="determinate")  # create progress bar
            pb.grid(row=1, column=3)
            pb["maximum"] = len(self.add_pay_periods) * 2  # set length of progress bar
            pb.start()
            sql = "DELETE FROM informalc_payouts WHERE year='%s' and carrier_name='%s'" % (year, name)
            pb["value"] = len(self.add_pay_periods)  # increment progress bar
            buttons.update()
            commit(sql)
            ii = len(self.add_pay_periods)
            count = 0
            paydays = []
            for i in range(len(self.add_pay_periods)):
                if self.add_pay_periods[i].get().strip() != "":
                    if self.add_hours[i].get().strip() != "" and self.add_rate[i].get().strip() != "" \
                            or self.add_amount[i].get().strip() != "":
                        pp = self.add_pay_periods[i].get().zfill(2)
                        one = "1"
                        pp += one  # format pp so it can fit in find_pp()
                        dt = find_pp(int(year),
                                     pp)  # returns the starting date of the pp when given year and pay period
                        dt += timedelta(days=20)
                        paydays.append(dt)
                        sql = "INSERT INTO informalc_payouts (year,pp,payday,carrier_name,hours,rate,amount) " \
                              "VALUES('%s','%s','%s','%s','%s','%s','%s')" \
                              % (year, self.add_pay_periods[i].get().strip(), paydays[i], name,
                                 self.add_hours[i].get().strip(), self.add_rate[i].get().strip(),
                                 self.add_amount[i].get().strip())
                        commit(sql)
                        count += 1
                        ii += 1
                        pb["value"] = ii  # increment progress bar
                        buttons.update()
            pb.stop()  # stop and destroy the progress bar
            pb_label.destroy()  # destroy the label for the progress bar
            pb.destroy()
            array = []
            selection = "none"
            msg = "Update: {} records for {} have been recorded in the database.".format(count, name)
            self.add(frame, array, selection, year, msg)

        def add_plus(self, frame, payouts):
            """ pay out entry """
            if len(payouts) == 0:
                self.add_pay_periods.append(StringVar(frame))  # set up array of stringvars for hours,rate,amount
                self.add_hours.append(StringVar(frame))
                self.add_rate.append(StringVar(frame))
                self.add_amount.append(StringVar(frame))
                Entry(frame, textvariable=self.add_pay_periods[len(self.add_pay_periods) - 1], width=10) \
                    .grid(row=len(self.add_pay_periods) + 6, column=0, pady=5, padx=5, sticky="w")
                Entry(frame, textvariable=self.add_hours[len(self.add_hours) - 1], width=10) \
                    .grid(row=len(self.add_hours) + 6, column=1, pady=5, padx=5)
                Entry(frame, textvariable=self.add_rate[len(self.add_rate) - 1], width=10) \
                    .grid(row=len(self.add_rate) + 6, column=2, pady=5, padx=5)
                Entry(frame, textvariable=self.add_amount[len(self.add_amount) - 1], width=10) \
                    .grid(row=len(self.add_amount) + 6, column=3, pady=5, padx=5)
            else:
                for i in range(len(payouts)):
                    self.add_pay_periods.append(StringVar(frame))  # set up array of stringvars for hours,rate,amount
                    self.add_hours.append(StringVar(frame))
                    self.add_rate.append(StringVar(frame))
                    self.add_amount.append(StringVar(frame))
                    self.add_pay_periods[i].set(payouts[i][1])
                    self.add_hours[i].set(payouts[i][4])
                    self.add_rate[i].set(payouts[i][5])
                    self.add_amount[i].set(payouts[i][6])
                    Entry(frame, textvariable=self.add_pay_periods[i], width=10) \
                        .grid(row=len(self.add_pay_periods) + 6, column=0, sticky="w")
                    Entry(frame, textvariable=self.add_hours[i], width=10) \
                        .grid(row=len(self.add_hours) + 6, column=1, pady=5, padx=5)
                    Entry(frame, textvariable=self.add_rate[i], width=10) \
                        .grid(row=len(self.add_rate) + 6, column=2, pady=5, padx=5)
                    Entry(frame, textvariable=self.add_amount[i], width=10) \
                        .grid(row=len(self.add_amount) + 6, column=3, pady=5, padx=5)

        def add(self, frame, array, selection, year, msg):
            """ pay out entry - add payout. """
            empty_array = []
            self.add_pay_periods = []
            self.add_hours = []
            self.add_rate = []
            self.add_amount = []
            self.win = MakeWindow()
            self.win.create(frame)
            Label(self.win.body, text="Informal C: Payout Entry", font=macadj("bold", "Helvetica 18")) \
                .grid(row=0, column=0, sticky="w", columnspan=5)
            Label(self.win.body, text="").grid(row=1)
            if selection != "none":
                Label(self.win.body, text=array[int(selection[0])], font="bold") \
                    .grid(row=2, column=0, sticky="w", columnspan=5)
                name = array[int(selection[0])]
                Label(self.win.body, text="Year: {}".format(year)).grid(row=3, column=0, sticky="w")
                Label(self.win.body, text="").grid(row=4)
                Label(self.win.body, text="PP", width=10, fg="grey").grid(row=5, column=0, sticky="w")
                Label(self.win.body, text="Hours", width=10, fg="grey").grid(row=5, column=1, sticky="w")
                Label(self.win.body, text="Rate", width=10, fg="grey").grid(row=5, column=2, sticky="w")
                Label(self.win.body, text="Amount", width=10, fg="grey").grid(row=5, column=3, sticky="w")
                Button(self.win.body, text="Add Payouts", width=10,
                       command=lambda: self.add_plus(self.win.body, empty_array)) \
                    .grid(row=5, column=4, sticky="w")
                sql = "SELECT * FROM informalc_payouts WHERE year ='%s' and carrier_name='%s'ORDER BY pp" \
                      % (year, name)
                payouts = inquire(sql)
                self.add_plus(self.win.body, payouts)
            else:
                Label(self.win.body, text="Select a carrier from the carrier list.").grid(row=2, column=0, sticky="w",
                                                                                          columnspan=5)
                name = "none"
            if msg != "":  # display a message when there is a message
                Label(self.win.buttons, text=msg, fg="red", width=60, anchor="w") \
                    .grid(row=0, column=0, columnspan=4, sticky="w")
            Button(self.win.buttons, text="Go Back", width=20,
                   command=lambda: self.goback(self.win.topframe)) \
                .grid(row=1, column=0, sticky="w")
            Button(self.win.buttons, text="Apply", width=20,
                   command=lambda: self.apply_add(self.win.topframe, name, year, self.win.buttons)) \
                .grid(row=1, column=1, sticky="w")
            Label(self.win.buttons, text="", width=10).grid(row=1, column=2)
            Label(self.win.buttons, text="", width=10).grid(row=1, column=3)
            self.win.finish()

        def goback(self, frame):
            """ pay out entry - go back and destroy the companion window if it still exist. """
            try:
                self.poe_listbox_.destroy()
            except TclError:
                pass
            self.poe_search(frame)

        def poe_listbox(self, dt_year, station, dt_start, year):
            """ pay out entry - create a listbox which allows the user to add carriers. """
            poe_root = Tk()
            self.poe_listbox_ = poe_root  # set the value
            poe_root.title("KLUSTERBOX")
            titlebar_icon(poe_root)  # place icon in titlebar
            x_position = projvar.root.winfo_x() + 450
            y_position = projvar.root.winfo_y() - 25
            poe_root.geometry("%dx%d+%d+%d" % (240, 600, x_position, y_position))
            n_f = Frame(poe_root)
            n_f.pack()
            n_buttons = Canvas(n_f)  # button bar
            n_buttons.pack(fill=BOTH, side=BOTTOM)
            Label(n_f, text="Carrier List", font=macadj("bold", "Helvetica 18")).pack(anchor="w")
            Label(n_f, text="{} Station:".format(station.get())).pack(anchor="w")
            Label(n_f, text="{} though {}".format(dt_year.strftime("%Y"), dt_start.strftime("%Y"))).pack(anchor="w")
            Label(n_f, text="").pack()
            scrollbar = Scrollbar(n_f, orient=VERTICAL)
            listbox = Listbox(n_f, selectmode="single", yscrollcommand=scrollbar.set)
            listbox.config(height=100, width=50)
            c_list = informalc_gen_clist(dt_start, dt_year, station.get())
            for name in c_list:
                listbox.insert(END, name)
            scrollbar.config(command=listbox.yview)
            scrollbar.pack(side=RIGHT, fill=Y)
            listbox.pack(side=LEFT, expand=1)
            msg = ""
            Button(n_buttons, text="Add Carrier", width=10,
                   command=lambda: self.add(self.win.topframe, c_list, listbox.curselection(), year, msg)) \
                .pack(side=LEFT, anchor="w")

            Button(n_buttons, text="Close", width=10,
                   command=lambda: (poe_root.destroy())).pack(side=LEFT, anchor="w")

    class PayoutReport:
        """
        this generates reports of pay outs for carriers using information generated by the Payout Entry class.
        """

        def __init__(self, parent):
            self.parent = parent
            self.win = None

        def informalc_por(self, frame):
            """ pay out report - allows user to conduct a search. """
            self.win = MakeWindow()
            self.win.create(frame)
            afterdate = StringVar(self.win.topframe)
            beforedate = StringVar(self.win.topframe)
            station = StringVar(self.win.topframe)
            station_options = projvar.list_of_stations
            if "out of station" in station_options:
                station_options.remove("out of station")
            station.set("undefined")
            backdate = StringVar(self.win.topframe)
            backdate.set("1")
            Label(self.win.body, text="Informal C: Payout Report Search Criteria",
                  font=macadj("bold", "Helvetica 18")) \
                .grid(row=0, column=0, columnspan=4, sticky="w")
            Label(self.win.body, text="").grid(row=1)
            Label(self.win.body, text="Enter range of dates and select station") \
                .grid(row=2, column=0, columnspan=4, sticky="w")
            Label(self.win.body, text="\tProvide dates in mm/dd/yyyy format.", fg="grey") \
                .grid(row=3, column=0, columnspan=4, sticky="w")
            Label(self.win.body, text="", width=20).grid(row=4, column=0)
            Label(self.win.body, text="After Date: ").grid(row=4, column=1, sticky="w")
            Entry(self.win.body, textvariable=afterdate, width=16).grid(row=4, column=2, sticky="w")
            Label(self.win.body, text="Before Date: ").grid(row=5, column=1, sticky="w")
            Entry(self.win.body, textvariable=beforedate, width=16).grid(row=5, column=2, sticky="w")
            Label(self.win.body, text="Station: ").grid(row=6, column=1, sticky="w")
            om_station = OptionMenu(self.win.body, station, *station_options)
            om_station.config(width=28)
            om_station.grid(row=6, column=2, columnspan=2)
            Label(self.win.body, text="Build the carrier list by going back how many years?") \
                .grid(row=7, column=0, columnspan=4, sticky="w")
            Label(self.win.body, text="Back Date: ").grid(row=8, column=1, sticky="w")
            om_backdate = OptionMenu(self.win.body, backdate, "1", "2", "3", "4", "5", "6", "7", "8", "9", "10")
            om_backdate.config(width=5)
            om_backdate.grid(row=8, column=2, sticky="w")
            button_alignment = macadj("w", "center")
            Button(self.win.buttons, text="Go Back", width=16, anchor=button_alignment,
                   command=lambda: self.parent.informalc(self.win.topframe)).grid(row=0, column=0)
            Label(self.win.buttons, text="Report: ", width=16).grid(row=0, column=1)
            Button(self.win.buttons, text="All Carriers", width=16, anchor=button_alignment,
                   command=lambda: self.por_all(afterdate, beforedate, station, backdate)) \
                .grid(row=0, column=2)
            self.win.finish()

        def por_all(self, afterdate, beforedate, station, backdate):
            """ pay out report. generates text report for all. """
            if not informalc_date_checker(self.win.topframe, afterdate, "After Date"):
                return
            if not informalc_date_checker(self.win.topframe, beforedate, "Before Date"):
                return
            start = informalc_date_converter(afterdate)
            end = informalc_date_converter(beforedate)
            if start > end:
                messagebox.showerror("Data Entry Error",
                                     "The After Date can not be earlier than the Before Date",
                                     parent=self.win.topframe)
                return
            if station.get() == "undefined":
                messagebox.showerror("Data Entry Error",
                                     "You must select a station. ",
                                     parent=self.win.topframe)
                return
            weeks = int(backdate.get()) * 52
            clist_start = start - timedelta(weeks=weeks)
            carrier_list = informalc_gen_clist(clist_start, end, station.get())

            stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = "infc_grv_list" + "_" + stamp + ".txt"
            report = open(dir_path('infc_grv') + filename, "w")
            report.write("  Payouts Report\n\n")
            report.write(
                "  Range of Dates: " + start.strftime("%b %d, %Y") + " - " + end.strftime("%b %d, %Y") + "\n\n")

            for name in carrier_list:
                sql = "SELECT * FROM informalc_payouts WHERE carrier_name = '%s' AND payday BETWEEN '%s' AND '%s' " \
                      "ORDER BY payday DESC" % (name, start, end)
                results = inquire(sql)
                if results:
                    payxamt = 0
                    payxadj = 0
                    report.write("  " + name + "\n\n")
                    report.write("    PP          Payday          Hours   Rate  Adjusted      Amount\n")
                    report.write("    --------------------------------------------------------------\n")
                    for result in results:
                        hour = 0.0
                        rate = 0.0
                        amt = 0.0
                        if result[4]:
                            hour = float(result[4])
                        if result[5]:
                            rate = float(result[5])
                        if result[6]:
                            amt = float(result[6])
                        if hour and rate:
                            payxadj += hour * rate
                        if amt:
                            payxamt += amt
                        pp = result[0] + "-" + result[1].zfill(2)
                        payday = dt_converter(result[2]).strftime("%b %d, %Y")
                        if result[4]:
                            hours = "{0:.2f}".format(float(result[4]))
                        else:
                            hours = "---"
                        if result[5]:
                            rate = "{0:.2f}".format(float(result[5]))
                        else:
                            rate = "---"
                        if result[4] and result[5]:
                            adj = "{0:.2f}".format(float(result[4]) * float(result[5]))
                        else:
                            adj = "---"
                        if result[6]:
                            amt = "{0:.2f}".format(float(result[6]))
                        else:
                            amt = "---"
                        report.write(
                            '    {:<5}{:>17}{:>9}{:>7}{:>10}{:>12}\n'.format(pp, payday, hours, rate, adj, amt))
                    report.write("    --------------------------------------------------------------\n")
                    report.write("    {:<40}{:>10}\n".format("Payouts adjusted to straight time", "{0:.2f}"
                                                             .format(float(payxadj))))
                    report.write("    {:<38}{:>24}\n".format("Payouts as flat dollar amount", "{0:.2f}"
                                                             .format(float(payxamt))))
                    report.write("\n\n\n")

            report.close()
            if sys.platform == "win32":
                os.startfile(dir_path('infc_grv') + filename)
            if sys.platform == "linux":
                subprocess.call(["xdg-open", 'kb_sub/infc_grv/' + filename])
            if sys.platform == "darwin":
                subprocess.call(["open", dir_path('infc_grv') + filename])


class InformalCSettings:
    """
    this creates a screen which will allow the user to change configure informal c 'results per page',
    'decision options', 'issue options', etc
    """

    def __init__(self):
        self.win = None
        self.informalc_result_limit = 0
        self.custom_issue = []  # a list of custom issues from informalc issues categories.
        self.custom_decision = []  # a list of custom decisions from informalc decision categories
        self.result_limit = None  # stringvar
        self.addcustomissue = None  # stringvar
        self.addcustomindex = None  # stringvar
        self.addcustomarticle = None  # stringvar
        self.addcustomdecision = None  # stringvar
        self.adddecindex = None  # stringvar
        self.adddectype = None  # stringvar
        self.row = 0
        self.status_update = None  # a label widget for status report
        self.max_issue_index = "0"  # the biggest value for issue index in the db
        self.max_decision_index = "0"  # the biggest value for decision index in the db

    def create(self, frame):
        """ this is a master method for calling other methods in the class in sequence. """
        self.win = MakeWindow()
        self.win.create(frame)
        self.get_stringvars()
        self.get_recs()
        self.build()
        self.button_frame()
        self.win.finish()

    def get_stringvars(self):
        """ create the stringvars """
        self.result_limit = StringVar(self.win.body)
        self.addcustomissue = StringVar(self.win.body)
        self.addcustomindex = StringVar(self.win.body)
        self.addcustomarticle = StringVar(self.win.body)
        self.addcustomdecision = StringVar(self.win.body)
        self.adddecindex = StringVar(self.win.body)
        self.adddectype = StringVar(self.win.body)

    def get_recs(self):
        """ get records from the database and define variables. """

        def find_available_index(sqlresult):
            """ accepts a list of distinct numeric strings, converts them to integers, and returns
            the lowest available value not in the distinct list as a string """
            array = distinctresult_to_list(sqlresult)  # convert the results to a list
            int_array = [int(x) for x in array]  # convert all strings to int
            for i in range(1, 999):
                if i in int_array:  # ignore if the number is already in use
                    pass
                else:
                    return str(i)  # return the first number not in use as a string

        sql = "SELECT tolerance FROM tolerances WHERE category = '%s'" % "informalc_result_limit"
        results = inquire(sql)
        self.informalc_result_limit = results[0][0]
        self.result_limit.set(self.informalc_result_limit)
        sql = "SELECT * FROM informalc_decisioncategories WHERE standard = 'False'"
        self.custom_decision = inquire(sql)
        self.custom_decision = issuedecisionresult_sorter(self.custom_decision)  # sort results by first value
        sql = "SELECT * FROM informalc_issuescategories WHERE standard = 'False'"
        self.custom_issue = inquire(sql)
        self.custom_issue = issuedecisionresult_sorter(self.custom_issue)  # sort results by first value
        sql = "SELECT DISTINCT(ssindex) FROM informalc_issuescategories"
        result = inquire(sql)
        self.max_issue_index = find_available_index(result)  # find the largest value
        sql = "SELECT DISTINCT(ssindex) FROM informalc_decisioncategories"
        result = inquire(sql)
        self.max_decision_index = find_available_index(result)  # find the largest value

    def build(self):
        """ build the screens """
        self.row = 0
        Label(self.win.body, text="Informal C Settings", font=macadj("bold", "Helvetica 18"), anchor="w") \
            .grid(row=self.row, sticky="w", columnspan=macadj(14, 15))
        self.row += 1
        Label(self.win.body, text=" ").grid(row=self.row, column=0)
        self.row += 1
        # ----------------------------------------------------------------------------------- search results per page
        text = macadj("Search Results __________________________________",
                      "Search Results ______________________________________")
        Label(self.win.body, text=text, anchor="w",
              fg="blue").grid(row=self.row, column=0, columnspan=macadj(14, 15), sticky="w")
        self.row += 1
        Label(self.win.body, text="Results Per Page:  ", anchor="w").grid(row=self.row, column=0, sticky="w")
        e = Entry(self.win.body, width=macadj(6, 5), textvariable=self.result_limit)
        e.grid(row=self.row, column=macadj(13, 13), sticky=macadj("e", "e"), pady=3)
        self.row += 1
        b = Button(self.win.body, width=5, anchor=macadj("e", "e"), text="ENTER",
                   command=lambda: self.update_result_limit())
        b.grid(row=self.row, column=macadj(13, 13), sticky=macadj("e", "e"))
        self.row += 1
        Label(self.win.body, text=" ").grid(row=self.row, column=0)  # blank line for reabability
        self.row += 1
        # --------------------------------------------------------------------------------------------- issue options
        text = macadj("Issue Options ___________________________________",
                      "Issue Options _______________________________________")
        Label(self.win.body, text=text, anchor="w",
              fg="blue").grid(row=self.row, column=0, columnspan=macadj(14, 15), sticky="w")
        self.row += 1
        Label(self.win.body, text="Available Issue Options:  ", anchor="w").grid(row=self.row, column=0, sticky="w")
        Button(self.win.body, width=5, anchor=macadj("w", "center"), text="list",
               command=lambda: InformalCOptions().issue_options(self.win.topframe)). \
            grid(row=self.row, column=13, sticky="e")
        self.row += 1
        Label(self.win.body, text=" ").grid(row=self.row, column=0)  # blank line for reabability
        self.row += 1
        # -------------------------------------------------------------------------------------------- add custom issue
        addissueframe = Frame(self.win.body)
        addissueframe.grid(row=self.row, column=0, columnspan=macadj(14, 15), sticky="w")
        self.addcustomindex.set(str(self.max_issue_index))
        self.addcustomarticle.set("")
        self.addcustomissue.set("")  # assign stringvars an empty value
        Label(addissueframe, text="Add New Custom Issue ").grid(row=0, column=0, columnspan=3, sticky="w")
        Label(addissueframe, text="Index", fg="grey").grid(row=1, column=0, sticky="w")
        Label(addissueframe, text="Article", fg="grey").grid(row=1, column=1, sticky="w")
        Label(addissueframe, text="Issue", fg="grey").grid(row=1, column=2, sticky="w")
        e = Entry(addissueframe, width=macadj(5, 4), textvariable=self.addcustomindex)
        e.grid(row=2, column=0, sticky="w", pady=3)
        e = Entry(addissueframe, width=6, textvariable=self.addcustomarticle)
        e.grid(row=2, column=1, sticky="w", pady=3)
        e = Entry(addissueframe, width=macadj(30, 28), textvariable=self.addcustomissue)
        e.grid(row=2, column=2, sticky="w", pady=3)
        Button(addissueframe, width=5, anchor=macadj("w", "center"), text="add",
               command=lambda: self.add_customissue()). \
            grid(row=3, column=2, sticky="e")
        self.row += 1
        Label(self.win.body, text=" ").grid(row=self.row, column=0)  # blank line for reabability
        self.row += 1
        # -------------------------------------------------------------------------------- show / delete custom issues
        customissueframe = Frame(self.win.body)
        customissueframe.grid(row=self.row, column=0, columnspan=macadj(14, 15), sticky="w")
        if not self.custom_issue:
            text = "There are no custom issue options to show."
            Label(customissueframe, text=text).grid(row=0, column=0, sticky="w")
        else:
            text = "Available Custom Issue Options"
            Label(customissueframe, text=text).grid(row=0, column=0, sticky="w", columnspan=3)
            Label(customissueframe, text="Index", fg="grey").grid(row=1, column=0, sticky="w")
            Label(customissueframe, text="Article", fg="grey").grid(row=1, column=1, sticky="w")
            Label(customissueframe, text="Issue", fg="grey").grid(row=1, column=2, sticky="w")
        row = 2
        for ci in self.custom_issue:
            Label(customissueframe, text=ci[0], width=macadj(4, 5), anchor="w", borderwidth=1,
                  relief="groove", pady=3).grid(row=row, column=0, sticky="w")  # index
            Label(customissueframe, text=ci[1], width=macadj(5, 7), anchor="w", borderwidth=1,
                  relief="groove", pady=3).grid(row=row, column=1, sticky="w")  # article
            Label(customissueframe, text=ci[2], width=macadj(19, 23), anchor="w", borderwidth=1,
                  relief="groove", pady=3).grid(row=row, column=2, sticky="w")  # issue
            Button(customissueframe, text="delete",  # button x
                   command=lambda delete_issue=ci[2]: (self.delete_customissue(delete_issue)))\
                .grid(row=row, column=3, sticky="w")
            row += 1
        self.row += 1
        Label(self.win.body, text="").grid(row=self.row)  # blank line for reabability
        self.row += 1
        # ------------------------------------------------------------------------------------------- decision options
        text = macadj("Decision Options ________________________________",
                      "Decision Options ____________________________________")
        Label(self.win.body, text=text, anchor="w",
              fg="blue").grid(row=self.row, column=0, columnspan=macadj(14, 15), sticky="w")
        self.row += 1
        Label(self.win.body, text="Available Decision Options: ", anchor="w").grid(row=self.row, column=0, sticky="w")
        Button(self.win.body, width=5, anchor=macadj("w", "center"), text="list",
               command=lambda: InformalCOptions().decision_options(self.win.topframe)). \
            grid(row=self.row, column=13, sticky="e")
        self.row += 1
        Label(self.win.body, text=" ").grid(row=self.row, column=0)  # blank line for reabability
        self.row += 1
        # ----------------------------------------------------------------------------------------- add custom decision
        adddecisionframe = Frame(self.win.body)
        adddecisionframe.grid(row=self.row, column=0, columnspan=macadj(14, 15), sticky="w")
        self.addcustomdecision.set("")  # assign stringvars an empty value
        self.adddecindex.set(str(self.max_decision_index))
        self.adddectype.set("general")
        Label(adddecisionframe, text="Add New Custom Decision ").grid(row=0, column=0, columnspan=3, sticky="w")
        Label(adddecisionframe, text="Index", fg="grey").grid(row=1, column=0, sticky="w")
        Label(adddecisionframe, text="Type", fg="grey").grid(row=1, column=1, sticky="w")
        Label(adddecisionframe, text="Decision", fg="grey").grid(row=1, column=2, sticky="w")
        e = Entry(adddecisionframe, width=macadj(5, 4), textvariable=self.adddecindex)
        e.grid(row=2, column=0, sticky="w", pady=3)
        e = Entry(adddecisionframe, width=8, textvariable=self.adddectype)
        e.grid(row=2, column=1, sticky="w", pady=3)
        e = Entry(adddecisionframe, width=macadj(27, 26), textvariable=self.addcustomdecision)
        e.grid(row=2, column=2, sticky="w", pady=3)
        Button(adddecisionframe, width=5, anchor=macadj("w", "center"), text="add",
               command=lambda: self.add_customdecision()).grid(row=3, column=2, sticky="e")
        self.row += 1
        Label(self.win.body, text="").grid(row=self.row)  # blank line for reabability
        self.row += 1
        # ----------------------------------------------------------------------------- show / delete custom decisions
        customdecisionframe = Frame(self.win.body)
        customdecisionframe.grid(row=self.row, column=0, columnspan=macadj(14, 15), sticky="w")
        if not self.custom_decision:
            text = "There are no custom decision options to show."
            Label(customdecisionframe, text=text).grid(row=0, column=0, sticky="w")
        else:
            text = "Available Custom Decision Options"
            Label(customdecisionframe, text=text).grid(row=0, column=0, sticky="w", columnspan=3)
            Label(customdecisionframe, text="Index", fg="grey").grid(row=1, column=0, sticky="w")
            Label(customdecisionframe, text="Type", fg="grey").grid(row=1, column=1, sticky="w")
            Label(customdecisionframe, text="Decision", fg="grey").grid(row=1, column=2, sticky="w")
        row = 2
        for cd in self.custom_decision:
            Label(customdecisionframe, text=cd[0], width=macadj(4, 5), anchor="w", borderwidth=1,
                  relief="groove", pady=3).grid(row=row, column=0, sticky="w")
            Label(customdecisionframe, text=cd[1], width=macadj(6, 7), anchor="w", borderwidth=1,
                  relief="groove", pady=3).grid(row=row, column=1, sticky="w")
            Label(customdecisionframe, text=cd[2], width=macadj(18, 23), anchor="w", borderwidth=1,
                  relief="groove", pady=3).grid(row=row, column=2, sticky="w")
            Button(customdecisionframe, text="delete",
                   command=lambda delete_decision=cd[2]: (self.delete_customdecision(delete_decision)))\
                .grid(row=row, column=3, sticky="w")
            row += 1
        self.row += 1
        Label(self.win.body, text="").grid(row=self.row)  # blank line for reabability
        self.row += 1

    def button_frame(self):
        """ Display buttons and status update message """
        button = Button(self.win.buttons)
        button.config(text="Go Back", width=20,
                      command=lambda: MainFrame().start(frame=self.win.topframe))
        if sys.platform == "win32":
            button.config(anchor="w")
        button.pack(side=LEFT)
        self.status_update = Label(self.win.buttons, text="", fg="red")
        self.status_update.pack(side=LEFT)

    def update_result_limit(self):
        """ apply the informal c result per page limit into the db after a check """
        result_limit = self.result_limit.get()
        if not result_limit:
            msg = "The Results Per Page can not be blank or zero. "
            messagebox.showerror("Informal C Settings", msg, parent=self.win.topframe)
            self.status_update.config(text="")  # empty the status update message
            return
        if not isint(result_limit):
            msg = "The Results Per Page must be an integer"
            messagebox.showerror("Informal C Settings", msg, parent=self.win.topframe)
            self.status_update.config(text="")  # empty the status update message
            return
        result_limit = int(result_limit)
        if not 0 < result_limit < 101:
            msg = "The Results Per Page must be between 0 and 201"
            messagebox.showerror("Informal C Settings", msg, parent=self.win.topframe)
            self.status_update.config(text="")  # empty the status update message
            return
        sql = "UPDATE tolerances SET tolerance='%s'WHERE category='%s'" % \
              (result_limit, "informalc_result_limit")
        commit(sql)
        msg = "Results Per Page updated: {}".format(result_limit)
        self.status_update.config(text="{}".format(msg))

    def add_customissue(self):
        """ run when the button 'add' for new custom issue is pressed. """
        index = self.addcustomindex.get().strip()
        article = self.addcustomarticle.get().strip()
        issue = self.addcustomissue.get().lower().strip()
        # return if any of the checks fail
        if not IndexArticleChecker().check_all(self.win.topframe, index, "issue index"):
            return
        if not IndexArticleChecker().check_all(self.win.topframe, article, "article"):
            return
        if not IssueDecisionChecker().check_all(self.win.topframe, issue, "issue"):
            return
        index = str(int(index))  # convert string to int to string to eliminate leading zeros.
        article = str(int(article))  # convert string to int to string to eliminate leading zeros.
        sql = "INSERT INTO informalc_issuescategories(ssindex, article, issue, standard)" \
              "VALUES('%s', '%s', '%s', '%s')" % (index, article, issue, "False")
        commit(sql)
        InformalCSettings().create(self.win.topframe)

    def delete_customissue(self, delete_issue):
        """ run when the 'delete' button is pressed in the display of custom issue options.
        this will delete the selected custom issue option. """
        sql = "DELETE FROM informalc_issuescategories WHERE issue = '%s'" % delete_issue
        commit(sql)
        InformalCSettings().create(self.win.topframe)

    def add_customdecision(self):
        """ run when the button 'add' for new custom decision is pressed. """
        index = self.adddecindex.get().strip()
        dectype = self.adddectype.get().strip().lower()
        # if no value was entered for decision type, use "general" as a default
        dectype = Convert(dectype).empty_returns_str("general")
        decision = self.addcustomdecision.get().lower().strip()
        # return if any of the checks fail
        if not IndexArticleChecker().check_all(self.win.topframe, index, "decision index"):
            return
        if not DecisionTypeChecker().check_all(self.win.topframe, dectype):
            return
        if not IssueDecisionChecker().check_all(self.win.topframe, decision, "decision"):
            return
        index = str(int(index))  # convert string to int to string to eliminate leading zeros.
        sql = "INSERT INTO informalc_decisioncategories(ssindex, type, decision, standard)" \
              "VALUES('%s', '%s', '%s', '%s')" % (index, dectype, decision, "False")
        commit(sql)
        InformalCSettings().create(self.win.topframe)

    def delete_customdecision(self, delete_decision):
        """ run when the 'delete' button is pressed in the display of custom decision options.
        this will delete the selected custom decision option. """
        sql = "DELETE FROM informalc_decisioncategories WHERE decision = '%s'" % delete_decision
        commit(sql)
        InformalCSettings().create(self.win.topframe)


class OtDistribution:
    """
    creates a screen to allow the user to configure overtime distribution settings and generate a spreadsheet.
    """

    def __init__(self):
        self.frame = None
        self.win = None
        self.row = 0
        self.quartinvran_year = None  # StringVar for investigation range
        self.quartinvran_quarter = None
        self.quartinvran_station = None
        self.new_quartinvran_year = None
        self.new_quartinvran_quarter = None
        self.new_quartinvran_station = None
        self.stations_minus_outofstation = None
        self.carrierlist = []  # distinct list of carriers by station and quarter
        self.recset = []  # recset of otdl carriers
        self.eligible_carriers = []  # all carriers on otdl during quarter
        self.ineligible_carriers = []  # carriers with no otdl rec during quarter, but a rec in otdl prefs
        self.startdate = datetime(1, 1, 1)
        self.enddate = datetime(1, 1, 1)
        self.station = ""
        self.quarter = ""
        self.range = None
        self.list_option_otdl = None
        self.list_option_wal = None
        self.list_option_nl = None
        self.list_option_aux = None
        self.list_option_ptf = None
        self.list_option_array = []
        self.status_update = ""

    def create(self, frame):
        """ called from the main screen to build ot preferences screen """
        self.frame = frame
        self.win = MakeWindow()
        self.win.create(self.frame)
        self.startup_stringvars()
        self.setup_listoption_stringvars()
        self.create_lower()

    def re_create(self, frame):
        """ called from the ot preferences screen when invran is changed. """
        self.row = 0  # re initialize vars
        self.startdate = datetime(1, 1, 1)
        self.enddate = datetime(1, 1, 1)
        self.station = ""
        self.quarter = ""
        self.frame = frame  # define the frame
        self.win = MakeWindow()
        self.re_startup_stringvars()
        self.setup_listoption_stringvars()
        self.create_lower()

    def create_lower(self):
        """ a continuation of create or re-create methods. """
        self.get_quarter()
        self.get_stations_list()
        self.get_dates()  # get startdate, enddate and station
        self.win.create(self.frame)
        self.build_quarterinvran()
        self.investigation_status()
        self.build_range()
        self.build_list_options()
        self.buttons_frame()
        self.win.finish()

    def get_stations_list(self):
        """ get a list of stations for station optionmenu """
        self.stations_minus_outofstation = projvar.list_of_stations[:]
        self.stations_minus_outofstation.remove("out of station")
        if len(self.stations_minus_outofstation) == 0:
            self.stations_minus_outofstation.append("undefined")

    def get_dates(self):
        """ find startdate, enddate and station """
        year = int(self.quartinvran_year.get())
        startdate = (datetime(year, 1, 1), datetime(year, 4, 1), datetime(year, 7, 1), datetime(year, 10, 1))
        enddate = (datetime(year, 3, 31), datetime(year, 6, 30), datetime(year, 9, 30), datetime(year, 12, 31))
        self.startdate = startdate[int(self.quartinvran_quarter.get()) - 1]
        self.enddate = enddate[int(self.quartinvran_quarter.get()) - 1]
        if self.quartinvran_station.get() == "undefined":
            self.station = ""
        else:
            self.station = self.quartinvran_station.get()

    def build_quarterinvran(self):
        """ build widgets to change the investigation range. """
        Label(self.win.body, text="Overtime Distribution", font=macadj("bold", "Helvetica 18"), anchor="w") \
            .grid(row=self.row, column=0, sticky="w", columnspan=20)
        self.row += 1
        Label(self.win.body, text="").grid(row=self.row, column=0)
        self.row += 1
        Label(self.win.body, text="QUARTERLY INVESTIGATION RANGE") \
            .grid(row=self.row, column=0, columnspan=20, sticky="w")
        self.row += 1
        Label(self.win.body, text=macadj("Year: ", "Year:"), fg="Gray", anchor="w") \
            .grid(row=self.row, column=0, sticky="w")
        Entry(self.win.body, width=macadj(5, 4), textvariable=self.quartinvran_year) \
            .grid(row=self.row, column=1, sticky="w")
        Label(self.win.body, text=macadj("Quarter: ", "Quarter:"), fg="Gray") \
            .grid(row=self.row, column=2, sticky="w")
        Entry(self.win.body, width=macadj(2, 1), textvariable=self.quartinvran_quarter) \
            .grid(row=self.row, column=3, sticky="w")
        Label(self.win.body, text=macadj("Station: ", "Station:"), fg="Gray") \
            .grid(row=self.row, column=4, sticky="w")
        om_station = OptionMenu(self.win.body, self.quartinvran_station, *self.stations_minus_outofstation)
        om_station.config(width=macadj(31, 23))
        om_station.grid(row=self.row, column=5, columnspan=4, sticky=W, padx=2)
        # set and reset buttons for investigation range
        Button(self.win.body, text="Set", width=macadj(5, 6), bg=macadj("green", "SystemButtonFace"),
               fg=macadj("white", "green"), command=lambda: self.set_invran()).grid(row=self.row, column=9, padx=2)
        Button(self.win.body, text="Reset", width=macadj(5, 6), bg=macadj("red", "SystemButtonFace"),
               fg=macadj("white", "red")).grid(row=self.row, column=10, padx=2)
        self.row += 1

    def investigation_status(self):
        """ provide message on status of investigation range """
        Label(self.win.body, text="").grid(row=self.row, column=0)
        self.row += 1
        Label(self.win.body, text="WEEKLY INVESTIGATION RANGE") \
            .grid(row=self.row, column=0, columnspan=20, sticky="w")
        self.row += 1
        # Investigation date SET/NOT SET notification
        if projvar.invran_weekly_span is None:
            Label(self.win.body, text="Investigation date/range not set", fg="red") \
                .grid(row=self.row, column=0, columnspan=8, sticky="w")
        elif projvar.invran_weekly_span == 0:  # if the investigation range is one day
            f_date = projvar.invran_date.strftime("%a - %b %d, %Y")
            Label(self.win.body, text="Day Set: {} --> Pay Period: {}".format(f_date, projvar.pay_period), fg="red") \
                .grid(row=self.row, column=0, columnspan=8, sticky="w")
        else:
            # if the investigation range is weekly
            f_date = projvar.invran_date_week[0].strftime("%a - %b %d, %Y")
            end_f_date = projvar.invran_date_week[6].strftime("%a - %b %d, %Y")
            Label(self.win.body, text="{0} through {1} --> Pay Period: {2}"
                  .format(f_date, end_f_date, projvar.pay_period), fg="red") \
                .grid(row=self.row, column=0, columnspan=8, sticky="w")

    def build_range(self):
        """ build widgets for changing the range to weekly or quarterly """
        self.row += 1
        Label(self.win.body, text="").grid(row=self.row, column=0)
        self.row += 1
        Label(self.win.body, text="Spread Sheet Range: ").grid(row=self.row, column=0, columnspan=8, sticky="w")
        self.row += 1
        self.range = StringVar(self.win.body)
        self.range.set('weekly')
        Radiobutton(self.win.body, text="Quarterly", variable=self.range, value='quarterly', justify=LEFT) \
            .grid(row=self.row, column=1, sticky=W, columnspan=3)
        self.row += 1
        Radiobutton(self.win.body, text="Weekly", variable=self.range, value='weekly', justify=LEFT) \
            .grid(row=self.row, column=1, sticky=W, columnspan=3)

    def build_list_options(self):
        """ build widgets for selecting list statuses to include. """
        self.row += 1
        Label(self.win.body, text="").grid(row=self.row, column=0)
        self.row += 1
        Label(self.win.body, text="List Options: ").grid(row=self.row, column=0, columnspan=8, sticky="w")
        self.row += 1
        Checkbutton(self.win.body, text="OTDL", variable=self.list_option_otdl, justify=LEFT) \
            .grid(row=self.row, column=1, sticky=W, columnspan=3)
        self.row += 1
        Checkbutton(self.win.body, text="Work Assignment", variable=self.list_option_wal, justify=LEFT) \
            .grid(row=self.row, column=1, sticky=W, columnspan=3)
        self.row += 1
        Checkbutton(self.win.body, text="No List", variable=self.list_option_nl, justify=LEFT) \
            .grid(row=self.row, column=1, sticky=W, columnspan=3)
        self.row += 1
        Checkbutton(self.win.body, text="Auxiliary", variable=self.list_option_aux, justify=LEFT) \
            .grid(row=self.row, column=1, sticky=W, columnspan=3)
        self.row += 1
        Checkbutton(self.win.body, text="Part Time Flex", variable=self.list_option_ptf, justify=LEFT) \
            .grid(row=self.row, column=1, sticky=W, columnspan=3)
        self.row += 1

    def set_invran(self):
        """ sets the investigation range. """
        if not self.check_quarterinvran():
            return
        self.re_create(self.win.topframe)

    def error_msg(self, text):
        """ generates error messageboxes. """
        messagebox.showerror("OTDL Preferences", text, parent=self.win.topframe)

    def check_quarterinvran(self):
        """ checks values for the investigation range date. """
        if not isint(self.quartinvran_year.get()):
            self.error_msg("The year must be a numeric.")
            return False
        if not len(self.quartinvran_year.get()) == 4:
            self.error_msg("Year must have four digits.")
            return False
        if not isint(self.quartinvran_quarter.get()):
            self.error_msg("The quarter must be an integer.")
            return False
        if int(self.quartinvran_quarter.get()) not in (1, 2, 3, 4):
            self.error_msg("Acceptable values for Quarter are limited to 1, 2, 3 or 4.")
            return False
        if self.quartinvran_station.get() == "undefined":
            self.error_msg("You must select a station to set the investigation range.")
            return False
        self.new_quartinvran_year = self.quartinvran_year.get()
        self.new_quartinvran_quarter = self.quartinvran_quarter.get()
        self.new_quartinvran_station = self.quartinvran_station.get()
        return True

    def startup_stringvars(self):
        """ sets up the stringvars for the investigation range and station. """
        if projvar.invran_weekly_span is None:  # if no investigation range is set
            date = datetime.now()
            station = "undefined"
        elif projvar.invran_weekly_span:  # if the investigation range is weekly
            date = projvar.invran_date_week[6]
            station = projvar.invran_station
        else:
            date = projvar.invran_date  # if the investigation range is daily
            station = projvar.invran_station
        year = date.strftime("%Y")
        month = date.strftime("%m")
        quarter = Quarter(month).find()  # get the quarter from the month
        self.quartinvran_year = StringVar(self.win.body)
        self.quartinvran_quarter = StringVar(self.win.body)
        self.quartinvran_station = StringVar(self.win.body)
        self.quartinvran_year.set(year)
        self.quartinvran_quarter.set(quarter)
        self.quartinvran_station.set(station)

    def re_startup_stringvars(self):
        """ re initializes stringvars """
        self.quartinvran_year = StringVar(self.win.body)
        self.quartinvran_quarter = StringVar(self.win.body)
        self.quartinvran_station = StringVar(self.win.body)
        self.quartinvran_year.set(self.new_quartinvran_year)
        self.quartinvran_quarter.set(self.new_quartinvran_quarter)
        self.quartinvran_station.set(self.new_quartinvran_station)

    def setup_listoption_stringvars(self):
        """ sets up the intvars for the option menus. """
        self.list_option_otdl = IntVar(self.win.body)
        self.list_option_wal = IntVar(self.win.body)
        self.list_option_nl = IntVar(self.win.body)
        self.list_option_aux = IntVar(self.win.body)
        self.list_option_ptf = IntVar(self.win.body)
        self.list_option_otdl.set(0)
        self.list_option_wal.set(1)
        self.list_option_nl.set(1)
        self.list_option_aux.set(0)
        self.list_option_ptf.set(0)

    def get_quarter(self):
        """ creates quarter in format "2021-3" """
        self.quarter = self.quartinvran_year.get() + "-" + self.quartinvran_quarter.get()

    def buttons_frame(self):
        """ creates the buttons on the bottom of the screen. also creates a status update. """
        button = Button(self.win.buttons)
        button.config(text="Go Back", width=macadj(18, 12),
                      command=lambda: MainFrame().start(frame=self.win.topframe))
        if sys.platform == "win32":
            button.config(anchor="w")
        button.pack(side=LEFT)
        # generate spreadsheet
        button = Button(self.win.buttons)
        button.config(text="SpreadSheet", width=macadj(17, 12),
                      command=lambda: (self.set_listoption_array(), OTDistriSpreadsheet().create
                      (self.win.topframe, self.startdate, self.quartinvran_station.get(), self.range.get(),
                       self.list_option_array)))
        if sys.platform == "win32":
            button.config(anchor="w")
        button.pack(side=LEFT)
        self.status_update = Label(self.win.buttons, text="", fg="red")
        self.status_update.pack(side=LEFT)

    def set_listoption_array(self):
        """ get the values from the option menus """
        self.list_option_array = []
        options = ("otdl", "wal", "nl", "aux", "ptf")
        strvars = (self.list_option_otdl.get(), self.list_option_wal.get(), self.list_option_nl.get(),
                   self.list_option_aux.get(), self.list_option_ptf.get())
        for i in range(len(strvars)):
            if strvars[i]:
                self.list_option_array.append(options[i])


class OtEquitability:
    """
    This class creates a window where the user can configure the ot equitability spreadsheet.
    """

    def __init__(self):
        self.frame = None
        self.win = None
        self.row = 0
        self.quartinvran_year = None  # StringVar for investigation range
        self.quartinvran_quarter = None
        self.quartinvran_station = None
        self.new_quartinvran_year = None  # place values in these when setting a new investigation range
        self.new_quartinvran_quarter = None
        self.new_quartinvran_station = None
        self.stations_minus_outofstation = None
        self.carrierlist = []
        self.recset = []
        self.startdate = datetime(1, 1, 1)
        self.enddate = datetime(1, 1, 1)
        self.station = ""
        self.quarter = ""
        self.pref_var = []  # build an array of stringvars for ot preference
        self.makeup_var = []  # build an array of stringvars for ot makeups
        self.onrec_prefs_carriers = []
        self.onrec_prefs = []
        self.onrec_makeups = []
        self.status_update = None
        self.delete_report = []  # list of ineligible carriers to be deleted from otdl prefence table
        self.eligible_carriers = []  # carriers on the otdl during the quarter from carriers table
        self.ineligible_carriers = []  # carriers with no otdl rec during quarter, but a rec in otdl prefs

    def create(self, frame):
        """ called from the main screen to build ot preferences screen"""
        self.frame = frame
        self.win = MakeWindow()
        self.startup_stringvars()
        self.create_lower()
        self.win.finish()

    def create_from_refusals(self, frame, enddate, station):
        """ called from the refusals screen to recreate the window """
        self.frame = frame
        self.station = station
        self.win = MakeWindow()
        self.setup_stringvars_from_refusals(enddate, station)
        self.create_lower()
        self.win.finish()

    def re_create(self, frame):
        """ called from the ot preferences screen when invran is changed. """
        self.row = 0  # re initialize vars
        self.carrierlist = []  # distinct list of carriers by station and quarter
        self.recset = []  # recset of otdl carriers
        self.eligible_carriers = []  # all carriers on otdl during quarter
        self.ineligible_carriers = []  # carriers with no otdl rec during quarter, but a rec in otdl prefs
        self.startdate = datetime(1, 1, 1)
        self.enddate = datetime(1, 1, 1)
        self.station = ""
        self.quarter = ""
        self.pref_var = []
        self.makeup_var = []
        self.onrec_prefs_carriers = []
        self.onrec_prefs = []
        self.onrec_makeups = []
        # self.status_update = None
        self.delete_report = []  # list of ineligible carriers to be deleted from otdl prefence table
        self.frame = frame  # define the frame
        self.win = MakeWindow()
        self.re_startup_stringvars()
        self.create_lower()
        self.win.finish()

    def create_lower(self):
        """ the bottom segment of the create method used by multiple create methods. """
        self.get_quarter()
        self.get_stations_list()
        self.win.create(self.frame)
        self.build_invran()
        self.get_dates()  # get startdate, enddate and station
        self.get_carrierlist()
        self.get_recsets()
        self.get_eligible_carriers()
        self.get_onrecs_set_stringvars()
        self.get_onrec_pref_carriers()
        self.get_ineligible()
        self.delete_ineligible()
        self.build_header()
        self.build_main()
        self.deletion_report()
        self.buttons_frame()

    def startup_stringvars(self):
        """ defines stringvars. """
        if projvar.invran_weekly_span is None:  # if no investigation range is set
            date = datetime.now()
            station = "undefined"
        elif projvar.invran_weekly_span:  # if the investigation range is weekly
            date = projvar.invran_date_week[6]
            station = projvar.invran_station
        else:
            date = projvar.invran_date  # if the investigation range is daily
            station = projvar.invran_station
        year = date.strftime("%Y")
        month = date.strftime("%m")
        quarter = Quarter(month).find()  # get the quarter from the month
        self.quartinvran_year = StringVar(self.win.body)
        self.quartinvran_quarter = StringVar(self.win.body)
        self.quartinvran_station = StringVar(self.win.body)
        self.quartinvran_year.set(year)
        self.quartinvran_quarter.set(quarter)
        self.quartinvran_station.set(station)

    def setup_stringvars_from_refusals(self, enddate, station):
        """ re creates string vars when called from refusals screen. """
        year = enddate.strftime("%Y")
        month = enddate.strftime("%m")
        quarter = Quarter(month).find()  # get the quarter from the month
        self.quartinvran_year = StringVar(self.win.body)
        self.quartinvran_quarter = StringVar(self.win.body)
        self.quartinvran_station = StringVar(self.win.body)
        self.quartinvran_year.set(year)
        self.quartinvran_quarter.set(quarter)
        self.quartinvran_station.set(station)

    def re_startup_stringvars(self):
        """ defines string vars. called when screen is refreshed. """
        self.quartinvran_year = StringVar(self.win.body)
        self.quartinvran_quarter = StringVar(self.win.body)
        self.quartinvran_station = StringVar(self.win.body)
        self.quartinvran_year.set(self.new_quartinvran_year)
        self.quartinvran_quarter.set(self.new_quartinvran_quarter)
        self.quartinvran_station.set(self.new_quartinvran_station)

    def get_quarter(self):
        """ creates quarter in format "2021-3" """
        self.quarter = self.quartinvran_year.get() + "-" + self.quartinvran_quarter.get()

    def get_stations_list(self):
        """ get a list of stations for station optionmenu """
        self.stations_minus_outofstation = projvar.list_of_stations[:]
        self.stations_minus_outofstation.remove("out of station")
        if len(self.stations_minus_outofstation) == 0:
            self.stations_minus_outofstation.append("undefined")

    def get_dates(self):
        """ find startdate, enddate and station """
        year = int(self.quartinvran_year.get())
        startdate = (datetime(year, 1, 1), datetime(year, 4, 1), datetime(year, 7, 1), datetime(year, 10, 1))
        enddate = (datetime(year, 3, 31), datetime(year, 6, 30), datetime(year, 9, 30), datetime(year, 12, 31))
        self.startdate = startdate[int(self.quartinvran_quarter.get()) - 1]
        self.enddate = enddate[int(self.quartinvran_quarter.get()) - 1]
        if self.quartinvran_station.get() == "undefined":
            self.station = ""
        else:
            self.station = self.quartinvran_station.get()

    def get_carrierlist(self):
        """ defines the carrier list. """
        self.carrierlist = CarrierList(self.startdate, self.enddate, self.station).get_distinct()

    def get_recsets(self):
        """ gets clock rings for a carrier and defines recset (record set). """
        for carrier in self.carrierlist:
            otlist = ("otdl",)
            rec = QuarterRecs(carrier[0], self.startdate, self.enddate, self.station).get_filtered_recs(otlist)
            if rec:
                self.recset.append(rec)

    def build_invran(self):
        """ creates widgets which allow the user to adjust the investigation range. """
        Label(self.win.body, text="OTDL Preferences", font=macadj("bold", "Helvetica 18"), anchor="w") \
            .grid(row=self.row, column=0, sticky="w", columnspan=20)
        self.row += 1
        Label(self.win.body, text="").grid(row=self.row, column=0)
        self.row += 1
        Label(self.win.body, text="QUARTERLY INVESTIGATION RANGE").grid(row=self.row, column=0, columnspan=20,
                                                                        sticky="w")
        self.row += 1
        Label(self.win.body, text=macadj("Year: ", "Year:"), fg="Gray", anchor="w") \
            .grid(row=self.row, column=0, sticky="w")
        Entry(self.win.body, width=macadj(5, 4), textvariable=self.quartinvran_year) \
            .grid(row=self.row, column=1, sticky="w")
        Label(self.win.body, text=macadj("Quarter: ", "Quarter:"), fg="Gray") \
            .grid(row=self.row, column=2, sticky="w")
        Entry(self.win.body, width=macadj(2, 1), textvariable=self.quartinvran_quarter) \
            .grid(row=self.row, column=3, sticky="w")
        Label(self.win.body, text=macadj("Station: ", "Station:"), fg="Gray") \
            .grid(row=self.row, column=4, sticky="w")
        om_station = OptionMenu(self.win.body, self.quartinvran_station, *self.stations_minus_outofstation)
        om_station.config(width=macadj(31, 23))
        om_station.grid(row=self.row, column=5, columnspan=4, sticky=W, padx=2)
        # set and reset buttons for investigation range
        Button(self.win.body, text="Set", width=macadj(5, 6), bg=macadj("green", "SystemButtonFace"),
               fg=macadj("white", "green"), command=lambda: self.set_invran()).grid(row=self.row, column=9, padx=2)
        Button(self.win.body, text="Reset", width=macadj(5, 6), bg=macadj("red", "SystemButtonFace"),
               fg=macadj("white", "red")).grid(row=self.row, column=10, padx=2)
        self.row += 1

    def set_invran(self):
        """ sets the investigation range """
        if not self.check_quarterinvran():
            return
        self.re_create(self.win.topframe)

    def error_msg(self, text):
        """ generates an error message. """
        messagebox.showerror("OTDL Preferences", text, parent=self.win.topframe)

    def check_quarterinvran(self):
        """ checks the investigation range. """
        if not isint(self.quartinvran_year.get()):
            self.error_msg("The year must be a numeric.")
            return False
        if not len(self.quartinvran_year.get()) == 4:
            self.error_msg("Year must have four digits.")
            return False
        if not isint(self.quartinvran_quarter.get()):
            self.error_msg("The quarter must be an integer.")
            return False
        if int(self.quartinvran_quarter.get()) not in (1, 2, 3, 4):
            self.error_msg("Acceptable values for Quarter are limited to 1, 2, 3 or 4.")
            return False
        if self.quartinvran_station.get() == "undefined":
            self.error_msg("You must select a station to set the investigation range.")
            return False
        self.new_quartinvran_year = self.quartinvran_year.get()
        self.new_quartinvran_quarter = self.quartinvran_quarter.get()
        self.new_quartinvran_station = self.quartinvran_station.get()
        return True

    def get_status(self, recs):
        """ returns true if the carrier's last record is otdl and the station is correct. """
        if recs[0][2] == "otdl" and recs[0][5] == self.station:
            return "on"
        return "off"

    @staticmethod
    def check_consistancy(recs):
        """ check that carriers on list have not gotten off then on again. """
        off_list = False
        on_list = False
        for rec in reversed(recs):
            if off_list:
                if rec[2] == "otdl":
                    on_list = True
            if rec[2] != "otdl":
                off_list = True
        if off_list and on_list:
            return True
        return False

    def get_eligible_carriers(self):
        """ builds array of carriers on otdl at any point during quarter from carrier table """
        for carrier in self.recset:
            self.eligible_carriers.append(carrier[0][1])

    def get_pref(self, carrier):
        """ pull otdl preferences from dbase - insert if there is no preference. """
        sql = "SELECT preference FROM otdl_preference WHERE carrier_name = '%s' and quarter = '%s' and station = '%s'" \
              % (carrier, self.quarter, self.station)
        pref = inquire(sql)
        if not pref:
            sql = "INSERT INTO otdl_preference (quarter, carrier_name, preference, station, makeups) " \
                  "VALUES('%s', '%s', '%s', '%s', '%s')" \
                  % (self.quarter, carrier, "12", self.station, "")
            commit(sql)
            return ['12', ]
        else:
            return pref[0]

    def get_makeups(self, carrier):
        """ pull makeups from the dbase """
        sql = "SELECT makeups FROM otdl_preference WHERE carrier_name = '%s' and quarter = '%s' and station = '%s'" \
              % (carrier, self.quarter, self.station)
        makeups = inquire(sql)
        if not makeups:
            return 0
        return makeups[0]

    def get_onrecs_set_stringvars(self):
        """ sets stringvars for carriers. """
        i = 0
        for carrier in self.eligible_carriers:
            self.pref_var.append(StringVar(self.win.body))  # build array of string vars for otdl preferences
            self.makeup_var.append(StringVar(self.win.body))  # build array of string vars for make ups
            pref = self.get_pref(carrier)  # call method to inquire otdl preference table
            makeup = self.get_makeups(carrier)[0]  # call method to inquire otdl preference table
            makeup = Convert(makeup).empty_not_zero()  # use empty string instead of zero
            self.pref_var[i].set(pref[0])  # set the preference stringvar
            self.makeup_var[i].set(makeup)
            self.onrec_prefs.append(pref[0])  # build the array of otdl preferences from otdl preferences table.
            self.onrec_makeups.append(makeup)
            i += 1

    def get_onrec_pref_carriers(self):
        """ get the otdl preference for each carriers. """
        sql = "SELECT carrier_name FROM otdl_preference WHERE quarter = '%s'and station = '%s'" \
              % (self.quarter, self.station)
        pref = inquire(sql)
        for carrier in pref:
            self.onrec_prefs_carriers.append(carrier[0])

    def get_ineligible(self):
        """ fills the ineligible carriers array. """
        for pref_carrier in self.onrec_prefs_carriers:
            if pref_carrier not in self.eligible_carriers:
                self.ineligible_carriers.append(pref_carrier)

    def delete_ineligible(self):
        """ removes ineligible carriers from the otdl preference table. """
        for carrier in self.ineligible_carriers:
            sql = "DELETE FROM otdl_preference WHERE quarter = '%s' AND carrier_name = '%s' AND station = '%s'" \
                  % (self.quarter, carrier, self.station)
            commit(sql)
            self.delete_report.append(carrier)

    def deletion_report(self):
        """ creates a message box for deleted carriers. """
        if len(self.delete_report) > 0:
            deleted_list = ""
            for name in self.delete_report:
                deleted_list += "      " + name + "\n"
            msg = "The OTDL Preference records has been deleted for quarter {} for the following " \
                  "carriers:\n\n{}\nThis is a routine maintenance action.".format(self.quarter, deleted_list)
            messagebox.showinfo("OTDL Preferences", msg, parent=self.win.body)

    def carrier_report(self, recs, consistant):
        """ generates a text file which shows carrier list status during the investigation range. """
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = "report_carrier_history" + "_" + stamp + ".txt"
        report = open(dir_path('report') + filename, "w")
        report.write("\nCarrier List Status History\n\n")
        report.write('   Showing all list status changes for {} during quarter {}\n\n'.format(recs[0][1], self.quarter))
        report.write('{:<16}{:<8}{:<25}\n'.format("Date Effective", "List", "Station"))
        report.write('---------------------------------------------\n')
        i = 1
        for line in recs:
            report.write('{:<16}{:<8}{:<25}\n'
                         .format(dt_converter(line[0]).strftime("%m/%d/%Y"), line[2], line[5]))
            if i % 3 == 0:
                report.write('---------------------------------------------\n')
            i += 1
        if consistant == "error":
            report.write('\n')
            report.write('>>>Consistency Error: \n'
                         'OTDL Carriers can not get back on the Over Time Desired List once they \n'
                         'have gotten off during the quarter. This will raise an  \"error\" \n'
                         'message in the Check column. If this is a mistake, edit the carrier\'s \n'
                         'status history. \n')
        report.close()
        if sys.platform == "win32":  # open the text document
            os.startfile(dir_path('report') + filename)
        if sys.platform == "linux":
            subprocess.call(["xdg-open", 'kb_sub/report/' + filename])
        if sys.platform == "darwin":
            subprocess.call(["open", dir_path('report') + filename])

    def build_header(self):
        """ build the header for the screen. """
        Label(self.win.body, text="Name", fg="Gray").grid(row=self.row, column=1, sticky="w")
        Label(self.win.body, text="Preference", fg="Gray").grid(row=self.row, column=5, sticky="w")
        Label(self.win.body, text="Make up", fg="Gray").grid(row=self.row, column=6, sticky="w")
        Label(self.win.body, text="Status", fg="Gray").grid(row=self.row, column=7, sticky="w")
        Label(self.win.body, text="Check", fg="Gray").grid(row=self.row, column=8, sticky="w")
        Label(self.win.body, text="Report", fg="Gray").grid(row=self.row, column=9, sticky="w")
        Label(self.win.body, text="Refusal", fg="Gray").grid(row=self.row, column=10, sticky="w")
        self.row += 1

    def build_main(self):
        """ builds the main part of the screen. """
        i = 0
        for carrier in self.recset:
            Label(self.win.body, text=i + 1, anchor="w").grid(row=self.row, column=0, sticky="w")
            Label(self.win.body, text=carrier[0][1], anchor="w").grid(row=self.row, column=1, columnspan=4, sticky="w")
            om_pref = OptionMenu(self.win.body, self.pref_var[i], "12", "10", "track")
            om_pref.config(width=4)
            om_pref.grid(row=self.row, column=5, sticky="w")
            Entry(self.win.body, textvariable=self.makeup_var[i], width=macadj(8, 6), justify='right') \
                .grid(row=self.row, column=6, sticky="w")  # make ups entry field
            status = "on"
            fg = "black"
            if self.get_status(carrier) == "off":  # if there is an error, display in red
                status = "off"
                fg = "red"
            Label(self.win.body, text=status, anchor="w", fg=fg).grid(row=self.row, column=7, sticky="w")
            consistant = "ok"
            fg = "black"
            if self.check_consistancy(carrier):  # if there is an error, display in red
                consistant = "error"
                fg = "red"
            Label(self.win.body, text=consistant, fg=fg, anchor="w").grid(row=self.row, column=8, sticky="w")
            Button(self.win.body, text="report",
                   command=lambda car=carrier, con=consistant: self.carrier_report(car, con)) \
                .grid(row=self.row, column=9, sticky="w")
            Button(self.win.body, text="refusals",
                   command=lambda car=carrier[0][1]: RefusalWin().create(self.win.topframe, car,
                                                                         self.startdate, self.enddate, self.station)) \
                .grid(row=self.row, column=10, sticky="w")
            self.row += 1
            i += 1

    def check_all(self):
        """ router for checking times """
        for i in range(len(self.onrec_makeups)):
            if not self.check_each(i):
                return False
        return True

    def check_each(self, i):
        """ checks time values. """
        carrier = self.recset[i][0][1]
        makeup = self.makeup_var[i].get()  # call method to inquire otdl preference table
        if RingTimeChecker(makeup).check_for_zeros():
            return True
        if not RingTimeChecker(makeup).check_numeric():
            text = "The Make up value for {} must be a number.".format(carrier)
            self.error_msg(text)
            return False
        if not RingTimeChecker(makeup).over_5000():
            text = "The Make up value for {} must not exceed 5000.".format(carrier)
            self.error_msg(text)
            return False
        if not RingTimeChecker(makeup).less_than_zero():
            text = "The Make up value for {} must not be less than zero.".format(carrier)
            self.error_msg(text)
            return False
        if not RingTimeChecker(makeup).count_decimals_place():
            text = "The Make up value for {} can not have more than two decimal places.".format(carrier)
            self.error_msg(text)
            return False
        return True

    def apply(self, home):
        """ applies to update the otdl preferences. """
        if not self.check_all():
            return
        updates = 0
        for i in range(len(self.onrec_prefs)):
            update = False
            if self.onrec_prefs[i] != self.pref_var[i].get():
                carrier = self.recset[i][0][1]
                sql = "UPDATE otdl_preference SET preference = '%s' WHERE carrier_name = '%s' AND quarter = '%s' " \
                      "AND station = '%s'" % (self.pref_var[i].get(), carrier, self.quarter, self.station)
                commit(sql)
                update = True
            if self.onrec_makeups[i] != self.makeup_var[i].get():
                carrier = self.recset[i][0][1]
                makeup = Convert(self.makeup_var[i].get()).empty_not_zero()
                makeup = Convert(makeup).empty_or_hunredths()
                sql = "UPDATE otdl_preference SET makeups = '%s' WHERE carrier_name = '%s' AND quarter = '%s' " \
                      "AND station = '%s'" % (makeup, carrier, self.quarter, self.station)
                commit(sql)
                update = True
            if update:
                updates += 1
        if home:
            MainFrame().start(frame=self.win.topframe)
        else:
            self.status_report(updates)
            self.reset_onrecs_and_vars()

    def reset_onrecs_and_vars(self):
        """ sets the preferences and make ups. """
        for i in range(len(self.pref_var)):
            pref = self.pref_var[i].get()
            makeup = Convert(self.makeup_var[i].get()).empty_not_zero()
            makeup = Convert(makeup).empty_or_hunredths()
            self.onrec_prefs[i] = pref
            self.onrec_makeups[i] = makeup
            self.makeup_var[i].set(makeup)

    def status_report(self, updates):
        """ generates the status update """
        msg = "{} Record{} Updated.".format(updates, Handler(updates).plurals())
        self.status_update.config(text="{}".format(msg))

    def buttons_frame(self):
        """ generates the frame at the bottom of the screen. """
        button = Button(self.win.buttons)
        button.config(text="Submit", width=macadj(17, 12),
                      command=lambda: self.apply(True))  # apply and return to main screen
        if sys.platform == "win32":
            button.config(anchor="w")
        button.pack(side=LEFT)
        button = Button(self.win.buttons)
        button.config(text="Apply", width=macadj(18, 12),
                      command=lambda: self.apply(False))  # apply and no not return to main
        if sys.platform == "win32":
            button.config(anchor="w")
        button.pack(side=LEFT)
        button = Button(self.win.buttons)
        button.config(text="Go Back", width=macadj(18, 12),
                      command=lambda: MainFrame().start(frame=self.win.topframe))
        if sys.platform == "win32":
            button.config(anchor="w")
        button.pack(side=LEFT)
        # generate spreadsheet
        button = Button(self.win.buttons)
        button.config(text="SpreadSheet", width=macadj(17, 12),
                      command=lambda: OTEquitSpreadsheet()
                      .create(self.win.topframe, self.startdate, self.quartinvran_station.get()))
        if sys.platform == "win32":
            button.config(anchor="w")
        button.pack(side=LEFT)
        self.status_update = Label(self.win.buttons, text="", fg="red")
        self.status_update.pack(side=LEFT)


class RefusalWin:
    """ create a window for refusals for otdl equitability """

    def __init__(self):
        self.frame = None
        self.win = None
        self.row = 0
        self.carrier_name = ""
        self.startdate = datetime(1, 1, 1)
        self.enddate = datetime(1, 1, 1)
        self.station = ""
        self.time_vars = []  # a list of stringvars of refusal times
        self.type_vars = []  # a list of stringvars of refusal types/indicators.
        self.ref_dates = []  # a list of datetime objects corrosponding to refusal times and types
        self.displaydate = []  # a list of strings providing the date of the refusals
        self.refset = []  # a list of refusals for the quarter
        self.onrec_time = []  # a list of the refusal time in the database
        self.onrec_type = []  # a list of the refusal type/indicator in the database
        self.onrec_displaydate = []
        self.status_update = None

    def create(self, frame, carrier, startdate, enddate, station):
        """ a master method for running other methods in proper order. """
        self.frame = frame
        self.carrier_name = carrier
        self.startdate = startdate
        self.enddate = enddate
        self.station = station
        self.win = MakeWindow()
        self.win.create(self.frame)
        self.get_refset()
        self.setup_vars_and_stringvars()
        self.build_header()
        self.build()
        self.build_bottom()
        self.buttons_frame()
        self.win.finish()

    def get_refset(self):
        """ get refusals from database """
        sql = "SELECT * FROM refusals WHERE refusal_date between '%s' and '%s' and carrier_name = '%s' " \
              "ORDER BY refusal_date" % (self.startdate, self.enddate, self.carrier_name)
        self.refset = inquire(sql)

    def setup_vars_and_stringvars(self):
        """ set up the string vars """
        i = 0
        date = self.startdate  # this will be the first date
        while date != self.enddate + timedelta(days=1):  # for each date in the quarter
            self.time_vars.append(StringVar(self.win.body))  # create a stringvar for time
            self.type_vars.append(StringVar(self.win.body))  # create a stringvar for type
            self.ref_dates.append(date)  # create a list of datetime objs corrosponding to the time/type vars
            displaydate = date.strftime("%m") + "/" + date.strftime("%d")  # make a string of date eg 07/29
            self.displaydate.append(displaydate)  # create a list of dates as string corrosponding to time/type vars
            self.onrec_time.append("")  # create the onrec time array
            self.onrec_type.append("")  # create the onrec type array
            for line in self.refset:  # loop through refset for refusals on that date
                if dt_converter(line[0]) == date:  # if there is a match
                    self.type_vars[i].set(line[2])  # set the stringvar for type
                    self.time_vars[i].set(line[3])  # set the stringvar for time
                    self.onrec_type[i] = line[2]  # change the onrec type to type from refset
                    self.onrec_time[i] = line[3]  # change the onrec time to time from refset
                    # create list of dates with records in the database as a string date eg 07/29
                    self.onrec_displaydate.append(dt_converter(line[0]).strftime("%m") + "/" + date.strftime("%d"))
            date += timedelta(days=1)
            i += 1  # increment the counter

    def start_column(self):
        """ returns the column position of the startdate """
        days = ("Saturday", "Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday")
        i = 0
        for day in days:  # loop through tuple of days
            if self.startdate.strftime("%A") == day:  # if the startdate matches the day
                return i  # return the index of the tuple
            i += 3  # increment the counter

    def build_header(self):
        """ build the screen header """
        Label(self.win.body, text="Refusals: {}".format(self.carrier_name),
              font=macadj("bold", "Helvetica 18"), anchor="w") \
            .grid(row=self.row, column=0, sticky="w", columnspan=27)
        self.row += 1
        Label(self.win.body, text="").grid(row=self.row)
        self.row += 1
        Label(self.win.body, text="Investigation Range: {} though {}"
              .format(self.startdate.strftime("%m/%d/%Y"), self.enddate.strftime("%m/%d/%Y")), fg="red") \
            .grid(row=self.row, columnspan=macadj(20, 27), sticky="w")
        self.row += 1
        Label(self.win.body, text="Station: {}".format(self.station)) \
            .grid(row=self.row, columnspan=macadj(20, 27), sticky="w")
        self.row += 1
        text = "Fill in the Refusal Indicator (optional) in the small field and any Refusal " \
               "Times in the large field. "
        Label(self.win.body, text=text, anchor="e", justify=LEFT).grid(row=self.row, columnspan=macadj(20, 27))
        self.row += 1
        Label(self.win.body, text="").grid(row=self.row)
        self.row += 1
        column = 0
        days = ("Sat", "Sun", "Mon", "Tue", "Wed", "Thu", "Fri")
        for day in days:
            Label(self.win.body, width=macadj(7, 3), text=day, anchor="w", fg="Blue") \
                .grid(row=self.row, column=column + 1, columnspan=3, sticky="w")
            column += 3
        self.row += 1

    def build(self):
        """ build labels and entry fields for refusals and refusal indicators. """
        column = self.start_column()
        for i in range(len(self.time_vars)):
            Label(self.win.body, width=macadj(2, 0), text="").grid(row=self.row, column=column)  # blank column
            column += 1
            Label(self.win.body, width=macadj(7, 4), text=self.displaydate[i], fg="Gray", anchor="w") \
                .grid(row=self.row, column=column, columnspan=2, sticky="w")  # display date
            Entry(self.win.body, width=macadj(2, 1), textvariable=self.type_vars[i]) \
                .grid(row=self.row + 1, column=column, sticky="w")  # entry field for type
            column += 1
            Entry(self.win.body, width=macadj(5, 4), textvariable=self.time_vars[i]) \
                .grid(row=self.row + 1, column=column, sticky="w")  # entry field for time
            column += 1
            if column >= 21:  # if the row is full
                column = 0  # reset column position to begining
                self.row += 2  # and start a new row

    def build_bottom(self):
        """ builds label for status update at the bottom of the screen. """
        for _ in range(3):
            self.row += 1
            Label(self.win.body, text="").grid(row=self.row)

    def buttons_frame(self):
        """ builds buttons on the bottom of the screen. """
        button = Button(self.win.buttons)
        button.config(text="Submit", width=macadj(20, 21),
                      command=lambda: self.apply(True))  # apply and do no return to main screen
        if sys.platform == "win32":
            button.config(anchor="w")
        button.pack(side=LEFT)

        button = Button(self.win.buttons)
        button.config(text="Apply", width=macadj(20, 21),
                      command=lambda: self.apply(False))  # apply and do no return to main screen
        if sys.platform == "win32":
            button.config(anchor="w")
        button.pack(side=LEFT)

        button = Button(self.win.buttons)
        button.config(text="Go Back", width=macadj(20, 21),
                      command=lambda: OtEquitability()
                      .create_from_refusals(self.win.topframe, self.enddate, self.station))
        if sys.platform == "win32":
            button.config(anchor="w")
        button.pack(side=LEFT)

        self.status_update = Label(self.win.buttons, text="", fg="red")
        self.status_update.pack(side=LEFT)

    def apply(self, home):
        """ loop through all stringvars and check for errors """
        for i in range(len(self.type_vars)):
            if not self.checktypes(i):  # check the refusal indicator
                return  # return if there is an error
            if not self.checktimes(i):  # check the refusal time
                return  # return if there is an error
        # if all checks pass - input/update/dalete the refusals database table
        for i in range(len(self.type_vars)):
            time_var = self.time_vars[i].get().strip()
            if not self.match_type(i) or not self.match_time(i):
                if self.displaydate[i] not in self.onrec_displaydate:  # if there is no record with that date
                    self.insert(i)
                if self.displaydate[i] in self.onrec_displaydate:  # if there is a record with that date
                    if not time_var:  # if the time is blank
                        self.delete(i)  # delete the record
                    else:  # if there is a time
                        self.update(i)  # update the record
        if home:  # return to the OT Preference screen
            OtEquitability().create_from_refusals(self.win.topframe, self.enddate, self.station)
        else:  # create a new object and recreate the window
            RefusalWin().create(self.win.topframe, self.carrier_name, self.startdate, self.enddate, self.station)

    def checktypes(self, i):
        """ checks the refusal indicator to make sure it is propely formatted. """
        type_var = self.type_vars[i].get().strip()
        time_var = self.time_vars[i].get().strip()
        if RefusalTypeChecker(type_var).is_empty():
            return True
        if not RefusalTypeChecker(type_var).is_one():
            messagebox.showerror("Refusal Tracking",
                                 "The Refusal indicator for {} must be only one character".format(self.displaydate[i]),
                                 parent=self.win.body)
            return False
        if not RefusalTypeChecker(type_var).is_letter():
            messagebox.showerror("Refusal Tracking",
                                 "The Refusal indicator for {} must be a letter".format(self.displaydate[i]),
                                 parent=self.win.body)
            return False
        if type_var and not time_var:
            messagebox.showerror("Refusal Tracking",
                                 "The refusal indicator for {} is not accompanied with a refusal time."
                                 .format(self.displaydate[i]),
                                 parent=self.win.body)
        return True

    def checktimes(self, i):
        """ checks leave times for proper formatting. """
        time_var = self.time_vars[i].get().strip()
        if RingTimeChecker(time_var).check_for_zeros():  # if blank or zero, skip all other checks
            return True
        if not RingTimeChecker(time_var).check_numeric():
            text = "The Refusal time for {} must be a numeric value.".format(self.displaydate[i])
            messagebox.showerror("Refusal Tracking", text, parent=self.win.topframe)
            return False
        if not RingTimeChecker(time_var).over_24():
            text = "The Refusal time for {} must be less than 24.".format(self.displaydate[i])
            messagebox.showerror("Refusal Tracking", text, parent=self.win.topframe)
            return False
        if not RingTimeChecker(time_var).less_than_zero():
            text = "The Refusal time for {} must be greater than or equal to 0.".format(self.displaydate[i])
            messagebox.showerror("Refusal Tracking", text, parent=self.win.topframe)
            return False
        if not RingTimeChecker(time_var).count_decimals_place():
            text = "The Refusal time for {} must not have more than 2 decimal places.".format(self.displaydate[i])
            messagebox.showerror("Refusal Tracking", text, parent=self.win.topframe)
            return False
        return True

    def match_type(self, i):
        """ check if the newly inputed type matchs the type in the database """
        type_var = self.type_vars[i].get().strip()  # the newly inputed type
        onrec = self.onrec_type[i]  # the type on record in the database
        if type_var == onrec:
            return True
        return False

    def match_time(self, i):
        """ check if the newly inputed time matchs the time in the database """
        time_var = self.time_vars[i].get().strip()  # the newly inputed time
        onrec = self.onrec_time[i]  # the time on record in the database
        if time_var == onrec:
            return True
        return False

    def insert(self, i):
        """ # insert a new record into the dbase """
        type_var = self.type_vars[i].get().strip()
        time_var = Convert(self.time_vars[i].get().strip()).hundredths()
        sql = "INSERT INTO Refusals (refusal_date, carrier_name, refusal_type, refusal_time) " \
              "VALUES('%s', '%s', '%s', '%s')" % (self.ref_dates[i], self.carrier_name, type_var, time_var)
        commit(sql)

    def update(self, i):
        """ update an existing record in the dbase """
        type_var = self.type_vars[i].get().strip()
        time_var = Convert(self.time_vars[i].get().strip()).hundredths()
        sql = "UPDATE Refusals SET refusal_type = '%s', refusal_time = '%s' WHERE refusal_date = '%s' " \
              "and carrier_name = '%s'" % (type_var, time_var, self.ref_dates[i], self.carrier_name)
        commit(sql)

    def delete(self, i):
        """ delete the record from the dbase """
        sql = "DELETE FROM Refusals WHERE refusal_date = '%s' and carrier_name = '%s'" \
              % (self.ref_dates[i], self.carrier_name)
        commit(sql)


class SpeedWorkBookGet:
    """
    this class gets the speedsheet and opens it.
    """

    def __init__(self):
        pass

    @staticmethod
    def get_filepath():
        """ get the file path"""
        if projvar.platform == "macapp" or projvar.platform == "winapp":
            return os.path.join(os.path.sep, os.path.expanduser("~"), 'Documents', 'klusterbox', 'speedsheets')
        else:
            return 'kb_sub/speedsheets'

    def get_file(self):
        """ returns the file path if there is one. else no selection/invalid selection. """
        path_ = self.get_filepath()
        file_path = filedialog.askopenfilename(initialdir=path_, filetypes=[("Excel files", "*.xlsx")])
        if file_path[-5:].lower() == ".xlsx":
            return file_path
        elif file_path == "":
            return "no selection"
        else:
            return "invalid selection"

    def open_file(self, frame, interject):
        """ gets the file and calls the speedsheet check and progress bar. """
        global pb_flag
        pb_flag = True
        file_path = self.get_file()
        if file_path == "no selection":
            return
        elif file_path == "invalid selection":
            messagebox.showerror("Report Generator",
                                 "The file you have selected is not an .xlsx file. "
                                 "You must select a file with a .xlsx extension.",
                                 parent=frame)
            return
        else:
            pb = ProgressBarIn(title="Klusterbox", label="SpeedSheeets Loading",
                               text="Loading and reading workbook. This could take a minute")
            wb = SpeedLoadThread(file_path)  # open workbook in separate thread
            wb.start()  # start loading workbook
            pb.start_up()  # start progress bar
            wb.join()  # wait for loading workbook to finish
            pb.stop()  # stop the progress bar and destroy the object
            SpeedSheetCheck(frame, wb.workbook, file_path, interject).check()  # check the speedsheet


class SpeedLoadThread(Thread):
    """ use multithreading to load workbook while progress bar runs """

    def __init__(self, path_):
        Thread.__init__(self)
        self.path_ = path_
        self.workbook = ""

    def run(self):
        """ runs the speedsheet loading. """
        global pb_flag  # this will signal when the thread has ended to end the progress bar
        wb = load_workbook(self.path_)  # load xlsx doc with openpyxl
        self.workbook = wb
        pb_flag = False


class SpeedSheetCheck:
    """
    this class checks the speedsheet. sends rows to speedcarriercheck and speedringcheck.
    """
    def __init__(self, frame, wb, path_, interject):
        self.frame = frame
        self.wb = wb
        self.path_ = path_
        self.interject = interject  # True = add to database/ False = pre-check
        self.fullreport = False  # True = shows full report/ False = only show errors and attn warnings
        self.carrier_count = 0
        self.rings_count = 0
        self.fatal_rpt = 0
        self.fyi_rpt = 0
        self.add_rpt = 0
        self.rings_fatal_rpt = 0
        self.rings_fyi_rpt = 0
        self.rings_add_rpt = 0
        self.ns_xlate = {}
        self.ns_rotate_mode = True
        self.triad_routefirst = False
        self.ns_true_rev = {}
        self.ns_false_rev = {}
        self.ns_custom = {}
        self.filename = ReportName("speedsheet_precheck").create()  # generate a name for the report
        self.report = open(dir_path('report') + self.filename, "w")  # open the report
        self.station = ""
        self.i_range = True  # investigation range is one week unless changed
        self.start_date = datetime(1, 1, 1)
        self.end_date = datetime(1, 1, 1)
        self.name = ""
        self.allowaddrecs = True
        self.name_mentioned = False
        self.pb = ProgressBarDe(label="SpeedSheet Checking")
        self.sheets = []
        self.sheet_count = 0
        self.sheet_rowcount = []
        self.all_inclusive = True
        self.start_row = 7
        self.modulus = 8
        self.step = 2

    def check(self):
        """ master method for running other methods and returns to the mainframe. """
        try:
            date_array = [1, 1, 1]
            self.set_ns_preference()
            self.set_triad_routefirst()
            self.get_fullreport()
            if date_array:
                projvar.try_absorber = True  # use project variable to absorb error from try/except statement.
            if self.ns_rotate_mode is not None and self.set_all_inclusive(0):
                self.set_sheet_facts()
                self.set_dates()
                self.set_ns_dictionaries()
                self.set_station()
                self.start_reporter()
                self.checking()
                self.reporter()
                date_array = Convert(self.start_date).datetime_separation()  # get the date to reset globals
                Globals().set(date_array[0], date_array[1], date_array[2], self.i_range, self.station, "None")
                MainFrame().start(frame=self.frame)
            else:
                self.pb.delete()  # stop and destroy progress bar
                self.showerror()
        except KeyError:  # if wrong type of file is selected, there will be an error
            self.pb.delete()  # stop and destroy progress bar
            self.showerror()

    def set_ns_preference(self):
        """ are ns day preferences rotating or fixed? """
        rotation = self.wb["by employee id"].cell(row=3, column=12).value  # get the ns day mode preference.
        if rotation.lower() not in ("r", "f"):
            self.ns_rotate_mode = None
        elif rotation == "r":
            self.ns_rotate_mode = True
        else:
            self.ns_rotate_mode = False

    def set_triad_routefirst(self):
        """ what is the move notation preference? True shows moves notation as 'route+time+time'
         False will show the mover notation as 'time+time+route' """
        string = self.wb["by employee id"].cell(row=4, column=12).value  # get move notation preference.
        self.triad_routefirst = Convert(string).str_to_bool()

    def get_fullreport(self):
        """ get the full report setting from the tolerances table - if True, the full report including add/fyi
        will be shown when the precheck or input into database is run. - if False, the report will only show
        errors and attention warnings. """
        sql = "SELECT tolerance FROM tolerances WHERE category = 'speedsheets_fullreport'"
        result = inquire(sql)  # get spreadsheet settings from database
        self.fullreport = result[0][0]
        self.fullreport = Convert(self.fullreport).str_to_bool()

    def set_all_inclusive(self, sheet_count):
        """ is the speedsheet all inclusive/ carrier only. """
        all_in = self.wb["by employee id"].cell(row=1, column=1).value
        if all_in == "Speedsheet - All Inclusive Weekly":
            if sheet_count == 0:  # adjust setting for the first sheet
                self.start_row = 7
                self.step = 1
                self.modulus = 8
            else:  # adjust settings for sheets after the first
                self.start_row = 6
                self.step = 2
            return True  # default settings from __init__ do not need changing
        elif all_in == "Speedsheet - All Inclusive Daily":
            if sheet_count == 0:  # adjust setting for the first sheet
                self.start_row = 7
                self.step = 1
                self.modulus = 2
            else:  # adjust settings for sheets after the first
                self.start_row = 6
                self.step = 0
            return True
        elif all_in == "Speedsheet - Carriers":
            self.all_inclusive = False
            self.start_row = 5
            self.step = 0
            self.modulus = 1
            return True
        else:
            return False

    def set_sheet_facts(self):
        """ get the worksheet names and number worksheets. """
        self.sheets = self.wb.sheetnames  # get the names of the worksheets as a list
        self.sheet_count = len(self.sheets)  # get the number of worksheets

    def set_dates(self):
        """ set the dates and the investigation range based on speedsheet input """
        datecell = self.wb[self.sheets[0]].cell(row=2, column=2).value  # get the date or range of dates
        if len(datecell) < 12:  # if the investigation range is daily
            self.start_date = Convert(datecell).backslashdate_to_datetime()  # convert formatted date to datetime
            self.end_date = self.start_date  # since daily, dates are the same
            self.i_range = False  # change the range since it is daily
        else:  # if the investigation range is weekly
            d = datecell.split(" through ")  # split the date into two
            self.start_date = Convert(d[0]).backslashdate_to_datetime()  # convert formatted date to datetime
            self.end_date = Convert(d[1]).backslashdate_to_datetime()

    def set_ns_dictionaries(self):
        """ gets the nsday as a dictionary. """
        ns_obj = NsDayDict(self.start_date)  # get the ns day object
        self.ns_xlate = ns_obj.get()  # get ns day dictionary
        self.ns_true_rev = ns_obj.get_rev(True)  # get ns day dictionary for rotating days
        self.ns_false_rev = ns_obj.get_rev(False)  # get ns day dictionary for fixed days
        self.ns_custom = ns_obj.custom_config()  # shows custom ns day configurations for reports

    def set_station(self):
        """ gets the station from the speedsheet. """
        self.station = self.wb[self.sheets[0]].cell(row=2, column=11).value  # get the station.

    def start_reporter(self):
        """ starts the report. """
        self.report.write("\nSpeedSheet Pre-Check Report \n")
        self.report.write(">>> {}\n".format(self.path_))

    def row_count(self):
        """ get a count of all rows for all sheets - need for progress bar """
        total_rows = 0
        for i in range(self.sheet_count):
            ws = self.wb[self.sheets[i]]  # assign the worksheet object
            row_count = ws.max_row  # get the total amount of rows in the worksheet
            self.sheet_rowcount.append(row_count)
            total_rows += row_count
        return total_rows

    def showerror(self):
        """ message box for showing errors. """
        messagebox.showerror("Klusterbox SpeedSheets",
                             "SpeedSheets Precheck or Input has failed. \n"
                             "Either you have selected a spreadsheet that is not \n"
                             "a SpeedSheet or your Speedsheet is corrupted. \n"
                             "Suggestion: Verify that the file you are selecting \n "
                             "is a SpeedSheet. \n"
                             "Suggestion: Try re-generating the SpeedSheet.",
                             parent=self.frame)

    def checking(self):
        """ reads rows and send to SpeedCarrierCheck or SpeedRingCheck. """
        is_name = False  # initialize bool for speedcell name
        count_diff = self.sheet_count * self.start_row  # subtract top five/six rows from the row count
        self.pb.max_count(self.row_count() - count_diff)  # get total count of rows for the progress bar
        self.pb.start_up()  # start up the progress bar
        pb_counter = 0  # initialize the progress bar counter
        for i in range(self.sheet_count):  # loop once for each worksheet in the workbook
            self.set_all_inclusive(i)  # set the start_row, step and modulus
            ws = self.wb[self.sheets[i]]  # assign the worksheet object
            row_count = ws.max_row  # get the total amount of rows in the worksheet
            for ii in range(self.start_row,
                            row_count + 1):  # loop through all rows, start with row 5 or 6 until the end
                self.pb.move_count(pb_counter)
                if (ii + self.step) % self.modulus == 0:  # if the row is a carrier record
                    if ws.cell(row=ii, column=2).value is not None:  # if the carrier record has a carrier name
                        self.name_mentioned = False  # keeps names from being repeated in reports
                        self.carrier_count += 1  # get a count of the carriers for reports
                        is_name = True  # bool: the speedcell has a name
                        day = Handler(ws.cell(row=ii, column=1).value).nonetype()
                        name = Handler(ws.cell(row=ii, column=2).value).nonetype()
                        list_stat = Handler(ws.cell(row=ii, column=6).value).nonetype()
                        nsday = Handler(ws.cell(row=ii, column=7).value).ns_nonetype()
                        route = Handler(ws.cell(row=ii, column=8).value).nonetype()
                        empid = Handler(ws.cell(row=ii, column=12).value).nonetype()
                        self.name = name
                        self.pb.change_text("Reading Speedcell: {}".format(name))  # update text for progress bar
                        SpeedCarrierCheck(self, self.sheets[i], ii, name, day, list_stat, nsday, route,
                                          empid).check_all()
                    else:
                        is_name = False  # the speedcell does not have a name
                        self.pb.change_text("Detected empty Speedcell.")  # update text for progress bar
                else:
                    # if the speedcell has a name and passed carrier test, get the rings
                    if is_name and self.allowaddrecs:
                        self.rings_count += 1
                        # Handler().nonetype will convert any nonetypes to empty stings
                        day = Handler(ws.cell(row=ii, column=1).value).nonetype()
                        hours = Handler(ws.cell(row=ii, column=2).value).nonetype()
                        bt = Handler(ws.cell(row=ii, column=3).value).nonetype()
                        moves = Handler(ws.cell(row=ii, column=4).value).nonetype()
                        rs = Handler(ws.cell(row=ii, column=8).value).nonetype()
                        et = Handler(ws.cell(row=ii, column=9).value).nonetype()
                        codes = Handler(ws.cell(row=ii, column=10).value).nonetype()
                        lv_type = Handler(ws.cell(row=ii, column=11).value).nonetype()
                        lv_time = Handler(ws.cell(row=ii, column=12).value).nonetype()
                        SpeedRingCheck(self, self.sheets[i], ii, day, hours, bt, moves, rs, et, codes,
                                       lv_type, lv_time).check()
                pb_counter += 1
        self.pb.stop()

    def reporter(self):
        """ writes the report """
        self.report.write("\n\n----------------------------------")
        # build report summary for carrier checks
        self.report.write("\n\nSpeedSheet Carrier Check Complete.\n\n")
        msg = "carrier{} checked".format(Handler(self.carrier_count).plurals())
        self.report.write('{:>6}  {:<40}\n'.format(self.carrier_count, msg))
        msg = "fatal error{} found".format(Handler(self.fatal_rpt).plurals())
        self.report.write('{:>6}  {:<40}\n'.format(self.fatal_rpt, msg))
        if self.interject:
            msg = "addition{} made".format(Handler(self.add_rpt).plurals())
            self.report.write('{:>6}  {:<40}\n'.format(self.add_rpt, msg))
        else:
            msg = "fyi notification{}".format(Handler(self.fyi_rpt).plurals())
            self.report.write('{:>6}  {:<40}\n'.format(self.fyi_rpt, msg))
        # build report summary for rings checks
        self.report.write("\n\nSpeedSheet Rings Check Complete.\n\n")
        msg = "ring{} checked".format(Handler(self.rings_count).plurals())
        self.report.write('{:>6}  {:<40}\n'.format(self.rings_count, msg))
        msg = "fatal error{} found".format(Handler(self.rings_fatal_rpt).plurals())
        self.report.write('{:>6}  {:<40}\n'.format(self.rings_fatal_rpt, msg))
        if self.interject:
            msg = "addition{} made".format(Handler(self.rings_add_rpt).plurals())
            self.report.write('{:>6}  {:<40}\n'.format(self.rings_add_rpt, msg))
        else:
            msg = "fyi notification{}".format(Handler(self.rings_fyi_rpt).plurals())
            self.report.write('{:>6}  {:<40}\n'.format(self.rings_fyi_rpt, msg))
        # close out the report and open in notepad
        self.report.close()
        if sys.platform == "win32":  # open the text document
            os.startfile(dir_path('report') + self.filename)
        if sys.platform == "linux":
            subprocess.call(["xdg-open", 'kb_sub/report/' + self.filename])
        if sys.platform == "darwin":
            subprocess.call(["open", dir_path('report') + self.filename])


class DatabaseAdmin:
    """
    a class for the management of the database.
    """

    def __init__(self):
        self.win = None
        self.dbase_location = None

    def run(self, frame):
        """ a screen for controlling database maintenance. """
        self.win = MakeWindow()
        self.win.create(frame)
        r = 0
        Label(self.win.body, text="Database Maintenance", font=macadj("bold", "Helvetica 18"), anchor="w") \
            .grid(row=r, sticky="w", columnspan=4)
        r += 1
        Label(self.win.body, text="").grid(row=r)
        r += 1
        Label(self.win.body, text="Database Records").grid(row=r, sticky="w", columnspan=4)
        r += 1
        Label(self.win.body, text="                    ").grid(row=r, column=0, sticky="w")
        r += 1
        # get and display number of records for rings3
        sql = "SELECT COUNT (*) FROM rings3"
        results = inquire(sql)
        Label(self.win.body, text=results, anchor="e", fg="red").grid(row=r, column=0, sticky="e")
        Label(self.win.body, text=" total records in rings table").grid(row=r, column=1, sticky="w")
        r += 1
        # get and display number of records for unique carriers in rings3
        sql = "SELECT COUNT (DISTINCT carrier_name) FROM rings3"
        results = inquire(sql)
        Label(self.win.body, text=results, anchor="e", fg="red").grid(row=r, column=0, sticky="e")
        Label(self.win.body, text=" distinct carrier names in rings table").grid(row=r, column=1, sticky="w")
        r += 1
        # get and display number of records for unique days in rings3
        sql = "SELECT COUNT (DISTINCT rings_date) FROM rings3"
        results = inquire(sql)
        Label(self.win.body, text=results, anchor="e", fg="red").grid(row=r, column=0, sticky="e")
        Label(self.win.body, text=" distinct days in rings table").grid(row=r, column=1, sticky="w")
        r += 1
        # get and display number of records for carriers
        sql = "SELECT COUNT (*) FROM carriers"
        results = inquire(sql)
        Label(self.win.body, text=results, anchor="e", fg="red").grid(row=r, column=0, sticky="e")
        Label(self.win.body, text=" total records in carriers table").grid(row=r, column=1, sticky=W)
        r += 1
        # get and display number of records for distinct carrier names from carriers
        sql = "SELECT COUNT (DISTINCT carrier_name) FROM carriers"
        results = inquire(sql)
        Label(self.win.body, text=results, anchor="e", fg="red").grid(row=r, column=0, sticky="e")
        Label(self.win.body, text=" distinct carrier names in carriers table").grid(row=r, column=1, sticky=W)
        r += 1
        # get and display number of records for stations
        sql = "SELECT COUNT (*) FROM stations"
        results = inquire(sql)
        Label(self.win.body, text=results, anchor="e", fg="red").grid(row=r, column=0, sticky="e")
        Label(self.win.body, text=" total records in station table (this includes \'out of station\')") \
            .grid(row=r, column=1, sticky="w")
        r += 1
        # find orphaned rings from deceased carriers
        sql = "SELECT DISTINCT carrier_name FROM carriers"
        carriers_results = inquire(sql)
        sql = "SELECT DISTINCT carrier_name FROM rings3"
        rings_results = inquire(sql)
        deceased = [x for x in rings_results if x not in carriers_results]
        Label(self.win.body, text=len(deceased), anchor="e", fg="red").grid(row=r, column=0, sticky="e")
        Label(self.win.body, text=" \'deceased\' carriers in rings table").grid(row=r, column=1, sticky=W)
        r += 1
        if len(deceased) > 0:
            Label(self.win.body, text="").grid(row=r, column=0, sticky="w")
            r += 1
            Button(self.win.body, text="clean",
                   command=lambda: (self.database_clean_rings(), self.run(self.win.topframe))) \
                .grid(row=r, column=0, sticky="w")
            Label(self.win.body, text="Delete rings records where carriers no longer exist (recommended)") \
                .grid(row=r, column=1, sticky="w", columnspan=6)
            r += 1
            Label(self.win.body, text="").grid(row=r, column=0, sticky="w")
            r += 1
        sql = "SELECT DISTINCT station FROM carriers"
        all_stations = inquire(sql)
        sql = "SELECT station FROM stations"
        good_stations = inquire(sql)
        deceased_cars = [x for x in all_stations if x not in good_stations]
        Label(self.win.body, text=len(deceased_cars), anchor="e", fg="red").grid(row=r, column=0, sticky="e")
        Label(self.win.body, text=" \'deceased\' stations in carriers table").grid(row=r, column=1, sticky=W)
        r += 1
        if len(deceased_cars) > 0:
            Label(self.win.body, text="").grid(row=r, column=0, sticky="w")
            r += 1
            Button(self.win.body, text="clean",
                   command=lambda: (self.database_clean_carriers(), self.run(self.win.topframe))) \
                .grid(row=r, column=0, sticky="w")
            Label(self.win.body, text="Delete carrier records where station no longer exist (recommended)") \
                .grid(row=r, column=1, sticky="w", columnspan=6)
            r += 1
        if projvar.invran_station is None:
            Label(self.win.body, text="").grid(row=r, column=0, sticky="w")
            r += 1
            Label(self.win.body, text="Database Records, {} Specific".format(projvar.invran_station)) \
                .grid(row=r, sticky="w", columnspan=4)
            r += 1
            Label(self.win.body, text="To see results from other stations, change station "
                                      "in the investigation range", fg="grey") \
                .grid(row=r, column=0, sticky="w", columnspan=6)
            r += 1
            Label(self.win.body, text="                    ").grid(row=r, column=0, sticky="w")
            r += 1
            # get and display number of records for carriers
            sql = "SELECT COUNT (*) FROM carriers WHERE station = '%s'" % projvar.invran_station
            results = inquire(sql)
            Label(self.win.body, text=results, anchor="e", fg="red").grid(row=r, column=0, sticky="e")
            Label(self.win.body, text=" total records in carriers table").grid(row=r, column=1, sticky=W)
            r += 1
            # get and display number of records for distinct carrier names from carriers
            sql = "SELECT COUNT (DISTINCT carrier_name) FROM carriers WHERE station = '%s'" % projvar.invran_station
            results = inquire(sql)
            Label(self.win.body, text=results, anchor="e", fg="red").grid(row=r, column=0, sticky="e")
            Label(self.win.body, text=" distinct carrier names in carriers table").grid(row=r, column=1, sticky=W)
            r += 1
        if "out of station" in projvar.list_of_stations:
            Label(self.win.body, text="").grid(row=r, column=0, sticky="w")
            r += 1
            Label(self.win.body, text="Database Records, for \"{}\"".format("out of station")) \
                .grid(row=r, sticky="w", columnspan=4)
            r += 1
            Label(self.win.body, text="                    ").grid(row=r, column=0, sticky="w")
            r += 1
            # get and display number of records for carriers
            sql = "SELECT COUNT (*) FROM carriers WHERE station = '%s'" % "out of station"
            results = inquire(sql)
            Label(self.win.body, text=results, anchor="e", fg="red").grid(row=r, column=0, sticky="e")
            Label(self.win.body, text=" total records in carriers table").grid(row=r, column=1, sticky=W)
            r += 1
            # get and display number of records for distinct carrier names from carriers
            sql = "SELECT COUNT (DISTINCT carrier_name) FROM carriers WHERE station = '%s'" % "out of station"
            results = inquire(sql)
            Label(self.win.body, text=results, anchor="e", fg="red").grid(row=r, column=0, sticky="e")
            Label(self.win.body, text=" distinct carrier names in carriers table").grid(row=r, column=1, sticky=W)
        r += 1
        Label(self.win.body, text="").grid(row=r)
        r += 1
        #  Backup database
        backup_frame = Frame(self.win.body)
        backup_frame.grid(row=r, column=0, columnspan=6, sticky=W)
        rr = 0
        Label(backup_frame, text="Open Database Location:").grid(row=rr, column=0, columnspan=6, sticky=W)
        rr += 1
        Label(backup_frame, text="This will open the (hidden) .klusterbox folder containing the Klusterbox "
                                 "database. The Klusterbox database is an \'.sqlite\' file which is generated "
                                 "when Klusterbox starts. The file must be located in the .klusterbox folder "
                                 "and named \'mandates.sqlite\' in order for it to be recognized by Klusterbox. ",
              wraplength=500, justify=LEFT, anchor="w", fg="grey").grid(row=rr, sticky="w", columnspan=6, column=0)
        rr += 1
        Label(backup_frame, text="").grid(row=rr)
        rr += 1
        Label(backup_frame, text="Find Database: ").grid(row=rr, column=0, sticky=W)
        Button(backup_frame, text="Find", width=8,
               command=lambda: self.file_dialogue()).grid(row=rr, column=1, sticky=W, padx=20)
        rr += 1
        Label(backup_frame, text="").grid(row=rr)
        rr += 1
        Label(backup_frame, text="Back up your database:").grid(row=rr, column=0, columnspan=6, sticky=W)
        rr += 1
        Label(backup_frame, text="You can save a copy of your database to anywhere you want within your "
                                 "computer or an external drive. You can give the database any name you like, but "
                                 "you must rename it \'mandates.sqlite\'. and return it to the .klusterbox "
                                 "folder in order for Klusterbox to recognize it.",
              wraplength=500, justify=LEFT, anchor="w", fg="grey").grid(row=rr, sticky="w", columnspan=6, column=0)
        rr += 1
        Label(backup_frame, text="").grid(row=rr)
        rr += 1
        Label(backup_frame, text="Back Up: ").grid(row=rr, column=0, sticky=W)
        Button(backup_frame, text="Save As", width=8,
               command=lambda: self.backup(self.win.topframe)).grid(row=rr, column=1, sticky=W, padx=20)
        rr += 1
        Label(backup_frame, text="").grid(row=rr)
        r += 1
        Label(self.win.body, text="").grid(row=r)
        r += 1
        #  Clock Rings summary
        rings_frame = Frame(self.win.body)
        rings_frame.grid(row=r, column=0, columnspan=6, sticky=W)
        rings_station = StringVar(rings_frame)
        rr = 0
        Label(rings_frame, text="Clock Rings Summary Report by Station:").grid(row=rr, column=0, columnspan=6, sticky=W)
        rr += 1
        Label(rings_frame, text="").grid(row=rr)
        rr += 1
        Label(rings_frame, text="Station: ").grid(row=rr, column=0, sticky=W)
        om_rings = OptionMenu(rings_frame, rings_station, *projvar.list_of_stations)
        om_rings.config(width=20)
        if projvar.invran_station is None:
            present_station = projvar.invran_station
        else:
            present_station = "select a station"
        rings_station.set(present_station)
        om_rings.grid(row=rr, column=1, sticky=W)
        Button(rings_frame, text="Report", width=8,
               command=lambda: self.database_rings_report(self.win.topframe, rings_station.get())) \
            .grid(row=rr, column=2, sticky=W, padx=20)
        rr += 1
        Label(rings_frame, text="").grid(row=rr)
        r += 1
        # declare variables for Delete Database Records
        clean1_range = StringVar(self.win.body)
        clean1_date = StringVar(self.win.body)
        clean1_table = StringVar(self.win.body)
        clean1_station = StringVar(self.win.body)
        # create frame and widgets for Delete Database Records
        cleaner_frame1 = Frame(self.win.body)
        cleaner_frame1.grid(row=r, columnspan=6)
        rr = 0
        Label(cleaner_frame1, text="Delete Database Records (Remove records from database per given specifications)",
              anchor="w").grid(row=rr, sticky="w", columnspan=4, column=0)
        rr += 1
        Label(cleaner_frame1, text="* format all date fields as mm/dd/yyyy, failure to do so will return an error",
              anchor="w", fg="grey").grid(row=rr, sticky="w", columnspan=4, column=0)
        rr += 1
        Label(cleaner_frame1, text="                                               ").grid(row=rr, column=5)
        rr += 1
        Label(cleaner_frame1, text="Delete Records: ", anchor="w").grid(row=rr, sticky="w", column=0)
        Radiobutton(cleaner_frame1, text="before and on date", variable=clean1_range, value="before",
                    anchor="w").grid(row=rr, sticky="w", column=1)
        rr += 1
        Radiobutton(cleaner_frame1, text="entered date only", variable=clean1_range, value="this_date",
                    anchor="w").grid(row=rr, sticky="w", column=1)
        rr += 1
        Radiobutton(cleaner_frame1, text="after and on date", variable=clean1_range, value="after",
                    anchor="w").grid(row=rr, sticky="w", column=1)
        rr += 1
        Radiobutton(cleaner_frame1, text="all dates", variable=clean1_range, value="all", anchor="w") \
            .grid(row=rr, sticky="w", column=1)
        clean1_range.set("after")
        r += 1
        # create frame and widgets for Delete Database Records
        cleaner_frame2 = Frame(self.win.body)
        cleaner_frame2.grid(row=r, columnspan=6, sticky="w")
        rrr = 0
        Label(cleaner_frame2, text="date* ", anchor="e").grid(row=rrr, column=0, sticky="e")
        Entry(cleaner_frame2, textvariable=clean1_date, width=macadj(12, 8), justify='right') \
            .grid(row=rrr, column=1, sticky="w")
        Label(cleaner_frame2, text="         table", anchor="e").grid(row=rrr, column=2, sticky="e")
        table_options = ("carriers + index", "carriers", "name index", "seniority", "clock rings", "all")
        om1_table = OptionMenu(cleaner_frame2, clean1_table, *table_options)
        clean1_table.set(table_options[-1])
        if sys.platform != "darwin":  # if the platform is not darwin
            om1_table.config(width=20, anchor="w")
        else:
            om1_table.config(width=20)
        om1_table.grid(row=rrr, column=3, sticky="w")
        rrr += 1
        station_options = projvar.list_of_stations[:]  # use splice to make copy of list without creating alias
        station_options.append("all stations")
        Label(cleaner_frame2, text="stations", anchor="e").grid(row=rrr, column=2, sticky="e")
        om1_station = OptionMenu(cleaner_frame2, clean1_station, *station_options)
        clean1_station.set(station_options[-1])
        if sys.platform != "darwin":  # if the platform is not darwin
            om1_station.config(width=20, anchor="w")
        else:
            om1_station.config(width=20)
        om1_station.grid(row=rrr, column=3, sticky="w")
        Button(cleaner_frame2, text="delete", width=macadj(6, 5),
               command=lambda: self.database_delete_records
               (frame, self.win.topframe, clean1_range, clean1_date, "x", clean1_table, clean1_station)) \
            .grid(row=rrr, column=4, sticky="w")
        rrr += 1
        Label(cleaner_frame2, text="").grid(row=rrr)
        rrr += 1
        # declare variables for Delete Database Records
        clean2_range = StringVar(self.win.body)
        clean2_startdate = StringVar(self.win.body)
        clean2_enddate = StringVar(self.win.body)
        clean2_table = StringVar(self.win.body)
        clean2_station = StringVar(self.win.body)
        rr += 1
        Label(cleaner_frame2, text="Delete Records within a specified range: ", anchor="w") \
            .grid(row=rrr, sticky="w", column=0, columnspan=6)
        rrr += 1
        Label(cleaner_frame2, text="* format all date fields as mm/dd/yyyy, failure to do so will return an error",
              anchor="w", fg="grey").grid(row=rrr, sticky="w", columnspan=4)
        rrr += 1
        # declare range as "between by default
        clean2_range.set("between")
        Label(cleaner_frame2, text="     start date* ", anchor="e").grid(row=rrr, column=0, sticky="e")
        Entry(cleaner_frame2, textvariable=clean2_startdate, width=macadj(12, 8), justify='right') \
            .grid(row=rrr, column=1, sticky="w")
        Label(cleaner_frame2, text="         table", anchor="e").grid(row=rrr, column=2, sticky="e")
        om2_table = OptionMenu(cleaner_frame2, clean2_table, *table_options)
        clean2_table.set(table_options[-1])
        if sys.platform != "darwin":  # if the platform is not darwin
            om2_table.config(width=20, anchor="w")
        else:
            om2_table.config(width=20)
        om2_table.grid(row=rrr, column=3, sticky="w")
        rrr += 1
        Label(cleaner_frame2, text="end date* ", anchor="e").grid(row=rrr, column=0, sticky="e")
        Entry(cleaner_frame2, textvariable=clean2_enddate, width=macadj(12, 8), justify='right') \
            .grid(row=rrr, column=1, sticky="w")
        Label(cleaner_frame2, text="stations", anchor="e").grid(row=rrr, column=2, sticky="e")
        om2_station = OptionMenu(cleaner_frame2, clean2_station, *station_options)
        clean2_station.set(station_options[-1])
        if sys.platform != "darwin":  # if the platform is not darwin
            om2_station.config(width=20, anchor="w")
        else:
            om2_station.config(width=20)
        om2_station.grid(row=rrr, column=3, sticky="w")
        Button(cleaner_frame2, text="delete", width=macadj(6, 5),
               command=lambda: self.database_delete_records(frame, self.win.topframe, clean2_range, clean2_startdate,
                                                            clean2_enddate, clean2_table, clean2_station)) \
            .grid(row=rrr, column=4, sticky="w")
        rrr += 1
        Label(cleaner_frame2, text="").grid(row=rrr)
        rrr += 1
        Label(cleaner_frame2, text="Reset Database - Delete and Rebuild the Database (all information will be lost)") \
            .grid(row=rrr, sticky="w", column=0, columnspan=6)
        rrr += 1
        Label(cleaner_frame2, text="").grid(row=rrr)
        rrr += 1
        Button(cleaner_frame2, text="Reset", width=10, padx=5, fg=macadj("white", "red"), bg=macadj("red", "white"),
               command=lambda: self.database_reset(frame, self.win.topframe)) \
            .grid(row=rrr, column=0, sticky="w")
        rrr += 1
        Label(cleaner_frame2, text="").grid(row=rrr)
        rrr += 1
        Label(cleaner_frame2, text="").grid(row=rrr)
        r += 1
        button = Button(self.win.buttons)
        button.config(text="Go Back", width=20, command=lambda: MainFrame().start(frame=self.win.topframe))
        if sys.platform == "win32":  # center the widget text for mac
            button.config(anchor="w")
        button.pack(side=LEFT)
        self.win.finish()

    @staticmethod
    def database_clean_rings():
        """ cleans the database from carriers who are no longer in the carriers table, but remain in the
        rings table. """
        sql = "SELECT DISTINCT carrier_name FROM carriers"
        carriers_results = inquire(sql)
        sql = "SELECT DISTINCT carrier_name FROM rings3"
        rings_results = inquire(sql)
        deceased = [x for x in rings_results if x not in carriers_results]
        pb_root = Tk()  # create a window for the progress bar
        pb_root.title("Deleting Orphaned Clock Rings")
        titlebar_icon(pb_root)  # place icon in titlebar
        pb_label = Label(pb_root, text="Running Process: ", anchor="w")  # make label for progress bar
        pb_label.grid(row=0, column=0, sticky="w")
        pb = ttk.Progressbar(pb_root, length=400, mode="determinate")  # create progress bar
        pb.grid(row=0, column=1, sticky="w")
        pb_text = Label(pb_root, text="", anchor="w")
        pb_text.grid(row=1, column=0, columnspan=2, sticky="w")
        steps = len(deceased)
        pb_count = 0
        pb["maximum"] = steps  # set length of progress bar
        pb.start()
        for dead in deceased:
            pb_text.config(text="Deleting clock rings for: {}".format(dead[0]))
            pb_count += 1
            pb["value"] = pb_count  # increment progress bar
            # change text for progress bar
            pb_root.update()
            sql = "DELETE FROM rings3 WHERE carrier_name='%s'" % dead
            commit(sql)
        pb_text.config(text="Deleting NULL clock rings.")
        pb_root.update()
        sql = "DELETE FROM rings3 WHERE carrier_name IS NULL"
        commit(sql)
        sql = "DELETE FROM rings3 WHERE total='%s' and code='%s' and leave_type ='%s'" % ("", 'none', '0.0')
        commit(sql)
        pb.stop()  # stop and destroy the progress bar
        pb_label.destroy()  # destroy the label for the progress bar
        pb.destroy()
        pb_root.destroy()

    @staticmethod
    def database_clean_carriers():
        """ delete carrier records where station no longer exist """
        sql = "SELECT DISTINCT station FROM carriers"
        all_stations = inquire(sql)
        sql = "SELECT station FROM stations"
        good_stations = inquire(sql)
        deceased = [x for x in all_stations if x not in good_stations]
        pb_root = Tk()  # create a window for the progress bar
        pb_root.title("Deleting Orphaned Clock Rings")
        titlebar_icon(pb_root)  # place icon in titlebar
        pb_label = Label(pb_root, text="Running Process: ", anchor="w")  # make label for progress bar
        pb_label.grid(row=0, column=0, sticky="w")
        pb = ttk.Progressbar(pb_root, length=400, mode="determinate")  # create progress bar
        pb.grid(row=0, column=1, sticky="w")
        pb_text = Label(pb_root, text="", anchor="w")
        pb_text.grid(row=1, column=0, columnspan=2, sticky="w")
        steps = len(deceased)
        pb_count = 0
        pb["maximum"] = steps  # set length of progress bar
        pb.start()
        for dead in deceased:
            sql = "DELETE FROM carriers WHERE station ='%s'" % (dead[0])
            commit(sql)
            pb_count += 1
            pb_text.config(text="Deleting carrier records for: {}".format(dead[0]))
            pb["value"] = pb_count  # increment progress bar
            pb_root.update()
        sql = "DELETE FROM rings3 WHERE carrier_name IS Null"
        commit(sql)
        pb.stop()  # stop and destroy the progress bar
        pb_label.destroy()  # destroy the label for the progress bar
        pb.destroy()
        pb_root.destroy()

    @staticmethod
    def database_reset(masterframe, frame):
        """ deletes the database and rebuilds it. """
        if not messagebox.askokcancel("Delete Database",
                                      "This action will delete your database and all information inside it."
                                      "This includes carrier information, rings information, settings as "
                                      "well as any informal c data. The database will be rebuilt and will be "
                                      "like new. "
                                      "\n\n This action can not be reversed."
                                      "\n\n Are you sure you want to proceed?", parent=frame):
            return
        path_ = "kb_sub/mandates.sqlite"
        if projvar.platform == "macapp":
            path_ = os.path.expanduser("~") + '/Documents/.klusterbox/mandates.sqlite'
        if projvar.platform == "winapp":
            path_ = os.path.expanduser("~") + '\\Documents\\.klusterbox\\mandates.sqlite'
        if projvar.platform == "py":
            path_ = "kb_sub/mandates.sqlite"
        try:
            if os.path.exists(path_):
                os.remove(path_)
        except FileNotFoundError:
            pass
        except (sqlite3.OperationalError, PermissionError):
            messagebox.showerror("Access Error",
                                 "Klusterbox can not delete the database as it is being used by another "
                                 "application. Close the database in the other application and retry.",
                                 parent=frame)
            return
        frame.destroy()
        masterframe.destroy()
        Globals().reset()  # reset initial value of globals
        DataBase().setup()
        StartUp().start()

    def database_delete_records(self, masterframe, frame, time_range, date, end_date, table, stations):
        """ deletes records from the database. """
        db_date = datetime(1, 1, 1)
        db_end_date = datetime(1, 1, 1)
        table_array = []
        if time_range.get() != "all":
            if not informalc_date_checker(frame, date, "date"):
                return
        if time_range.get() == "between":
            if not informalc_date_checker(frame, end_date, "end date"):
                return
        if table.get() == "" or stations.get() == "":
            if messagebox.showerror("Database Maintenance",
                                    "You must select a table and a station. ",
                                    parent=frame):
                return
        if not messagebox.askokcancel("Database Maintenance",
                                      "This action will delete records from the database. \n\n"
                                      "This action is irreversible. \n\n"
                                      "Are you sure you want to proceed?",
                                      parent=frame):
            return
        #  convert date to format usable by sqlite
        if time_range.get() != "all":
            d = date.get().split("/")
            db_date = datetime(int(d[2]), int(d[0]), int(d[1]))
        if time_range.get() == "between":
            d = end_date.get().split("/")
            db_end_date = datetime(int(d[2]), int(d[0]), int(d[1]))
        # define the station array to loop
        if stations.get() == "all stations":
            station_array = projvar.list_of_stations[:]
        else:
            station_array = [stations.get()]
        # define the table array to loop
        if table.get() == "all":
            table_array = ["rings3", "name_index", "seniority", "carriers", "stations", "station_index"]
        elif table.get() == "carriers + index":
            table_array = ["carriers", "name_index", "seniority"]
        elif table.get() == "carriers":
            table_array = ["carriers"]
        elif table.get() == "name index":
            table_array = ["name_index"]
        elif table.get() == "seniority":
            table_array = ["seniority"]
        elif table.get() == "clock rings":
            table_array = ["rings3"]
        #  short cuts to delete all records in table
        if time_range.get() == "all" and stations.get() == "all stations":
            for tab in table_array:
                if tab == "stations":
                    sql = "DELETE FROM stations"
                    commit(sql)
                if tab == "station_index":
                    sql = "DELETE FROM station_index"
                    commit(sql)
                if tab == "name_index":
                    sql = "DELETE FROM name_index"
                    commit(sql)
                if tab == "seniority":
                    sql = "DELETE FROM seniority"
                    commit(sql)
                if tab == "carriers":
                    sql = "DELETE FROM carriers"
                    commit(sql)
                if tab == "rings3":
                    sql = "DELETE FROM rings3"
                    commit(sql)
                if tab == "stations":
                    sql = "DELETE FROM stations"
                    commit(sql)
                    sql = "INSERT INTO stations (station) VALUES('%s')" % "out of station"
                    commit(sql)
                    del projvar.list_of_stations[:]
                    projvar.list_of_stations.append("out of station")

                    Globals().reset()  # reset investigation range
            messagebox.showinfo("Database Maintenance",
                                "Success! The database has been cleaned of the specified records.",
                                parent=frame)
            frame.destroy()
            self.run(masterframe)
            return
        # loop for great justice
        operator = ""
        for stat in station_array:
            for tab in table_array:
                # delete all rings associated with station
                if tab == "stations":
                    if stat != "out of station":
                        sql = "DELETE FROM stations WHERE station = '%s'" % stat
                        commit(sql)
                    if stat != "out of station":
                        projvar.list_of_stations.remove(stat)
                    if projvar.invran_station == stat:
                        Globals().reset()  # reset initial value of globals
                if tab == "station_index":
                    if stat != "out of station":
                        sql = "DELETE FROM station_index WHERE kb_station = '%s'" % stat
                        commit(sql)
                if tab == "rings3":
                    # determine operator based on time_range
                    if time_range.get() == "before":
                        operator = " AND rings_date <= '%s'" % db_date
                    elif time_range.get() == "this_date":
                        operator = " AND rings_date = '%s'" % db_date
                    elif time_range.get() == "after":
                        operator = " AND rings_date >= '%s'" % db_date
                    elif time_range.get() == "all":
                        operator = ""
                    elif time_range.get() == "between":
                        operator = " AND rings_date BETWEEN '%s' AND '%s'" % (db_date, db_end_date)
                    sql = "SELECT DISTINCT carrier_name FROM carriers WHERE station = '%s' ORDER BY carrier_name" \
                          % stat
                    result = inquire(sql)
                    pb_root = Tk()  # create a window for the progress bar
                    pb_root.title("Deleting Clock Rings from {}".format(stat))
                    titlebar_icon(pb_root)
                    pb_label = Label(pb_root, text="Running Process: ", anchor="w")  # make label for progress bar
                    pb_label.grid(row=0, column=0, sticky="w")
                    pb = ttk.Progressbar(pb_root, length=400, mode="determinate")  # create progress bar
                    pb.grid(row=0, column=1, sticky="w")
                    pb_text = Label(pb_root, text="", anchor="w")
                    pb_text.grid(row=1, column=0, columnspan=2, sticky="w")
                    steps = len(result)
                    pb_count = 0
                    pb["maximum"] = steps  # set length of progress bar
                    pb.start()
                    for name in result:
                        pb_count += 1
                        active_station = []
                        # get all records for the carrier
                        sql = "SELECT * FROM carriers WHERE carrier_name= '%s' ORDER BY effective_date" % name[0]
                        result_1 = inquire(sql)
                        start_search = True
                        start = ''
                        end = ''
                        # build the active_station array - find dates where carrier entered/left station
                        for r in result_1:
                            if r[5] == stat and start_search is True:
                                start = r
                                start_search = False
                            if r[5] != stat and start_search is False:
                                end = r
                                active_station.append([start, end])
                                start = ''
                                end = ''
                                start_search = True
                        if not start_search:
                            active_station.append([start, end])
                        for active in active_station:
                            if active[1] != '':
                                sql = "DELETE FROM rings3 WHERE rings_date " \
                                      "BETWEEN '%s' AND '%s'{} AND carrier_name = '%s' " \
                                          .format(operator) % (active[0][0], active[1][0], name[0])
                                commit(sql)
                                # change text for progress bar
                                pb_text.config(text="Deleting in range rings for: {} - {} through {}"
                                               .format(name[0], active[0][0], active[1][0]))
                                pb["value"] = pb_count  # increment progress bar
                                pb_root.update()
                            else:
                                sql = "DELETE FROM rings3 WHERE rings_date >= '%s'{} AND carrier_name = '%s' " \
                                          .format(operator) % (active[0][0], name[0])
                                commit(sql)
                                # change text for progress bar
                                pb_text.config(
                                    text="Deleting in range rings for: {} - {} +".format(name[0], active[0][0]))
                                pb_root.update()
                                pb["value"] = pb_count  # increment progress bar
                    pb.stop()  # stop and destroy the progress bar
                    pb_label.destroy()  # destroy the label for the progress bar
                    pb.destroy()
                    pb_root.destroy()
                if tab == "name_index":
                    sql = "SELECT DISTINCT carrier_name FROM carriers WHERE station = '%s'" \
                          % stat
                    results = inquire(sql)
                    pb_root = Tk()  # create a window for the progress bar
                    pb_root.title("Deleting Clock Rings from {}".format(stat))
                    titlebar_icon(pb_root)
                    pb_label = Label(pb_root, text="Running Process: ", anchor="w")  # make label for progress bar
                    pb_label.grid(row=0, column=0, sticky="w")
                    pb = ttk.Progressbar(pb_root, length=400, mode="determinate")  # create progress bar
                    pb.grid(row=0, column=1, sticky="w")
                    pb_text = Label(pb_root, text="", anchor="w")
                    pb_text.grid(row=1, column=0, columnspan=2, sticky="w")
                    steps = len(results)
                    pb_count = 0
                    pb["maximum"] = steps  # set length of progress bar
                    pb.start()
                    for car in results:
                        sql = "DELETE FROM name_index WHERE kb_name='%s'" % car[0]
                        commit(sql)
                        pb_count += 1
                        pb_text.config(text="Deleting name index for: {}".format(car[0]))
                        pb["value"] = pb_count  # increment progress bar
                        pb_root.update()
                    pb.stop()  # stop and destroy the progress bar
                    pb_label.destroy()  # destroy the label for the progress bar
                    pb.destroy()
                    pb_root.destroy()
                if tab == "carriers":
                    # determine operator based on time_range
                    if time_range.get() == "before":
                        operator = "AND effective_date <= '%s'" % db_date
                    elif time_range.get() == "this_date":
                        operator = "AND effective_date = '%s'" % db_date
                    elif time_range.get() == "after":
                        operator = "AND effective_date >= '%s'" % db_date
                    elif time_range.get() == "all":
                        operator = ""
                    elif time_range.get() == "between":
                        operator = "AND '%s' <= effective_date <= '%s'" % (db_date, db_end_date)
                    sql = "SELECT DISTINCT carrier_name FROM carriers WHERE station = '%s' {}".format(operator) % stat
                    results = inquire(sql)
                    pb_root = Tk()  # create a window for the progress bar
                    pb_root.title("Deleting Carrier Records from {}".format(stat))
                    titlebar_icon(pb_root)
                    pb_label = Label(pb_root, text="Running Process: ", anchor="w")  # make label for progress bar
                    pb_label.grid(row=0, column=0, sticky="w")
                    pb = ttk.Progressbar(pb_root, length=400, mode="determinate")  # create progress bar
                    pb.grid(row=0, column=1, sticky="w")
                    pb_text = Label(pb_root, text="", anchor="w")
                    pb_text.grid(row=1, column=0, columnspan=2, sticky="w")
                    steps = len(results)
                    pb_count = 0
                    pb["maximum"] = steps  # set length of progress bar
                    pb.start()
                    for car in results:
                        pb_text.config(
                            text="Deleting clock rings for: {}".format(car[0]))  # change text for progress bar
                        pb_count += 1
                        pb["value"] = pb_count  # increment progress bar
                        pb_root.update()
                        sql = "SELECT * FROM carriers WHERE  carrier_name = '%s' {}".format(operator) % car[0]
                        car_ask = inquire(sql)
                        outside_station = False
                        for cc in car_ask:  # look for rings where the station doesn't match or out of station
                            if cc[5] != "out of station" or cc[5] != stat:
                                outside_station = True
                        if not outside_station:
                            for carr in results:
                                # update all records where station/carrier match to 'out of station'
                                sql = "UPDATE carriers SET station='%s' WHERE carrier_name ='%s' AND station='%s' {}" \
                                          .format(operator) % ("out of station", carr[0], stat)
                                commit(sql)
                                # find redundancies where two 'out of station' records are adjacent.
                                sql = "SELECT * FROM carriers WHERE carrier_name ='%s' " \
                                      "ORDER BY carrier_name, effective_date" % carr[0]
                                car_results = inquire(sql)
                                duplicates = []
                                for i in range(len(car_results)):
                                    if i != len(car_results) - 1:  # if the loop has not reached the end of the list
                                        # if the name current and next name are the same
                                        if car_results[i][5] == 'out of station' and \
                                                car_results[i + 1][5] == 'out of station':
                                            duplicates.append(i + 1)
                                for d in duplicates:
                                    sql = "DELETE FROM carriers WHERE effective_date='%s' and carrier_name='%s'" % (
                                        car_results[d][0], car_results[d][1])
                                    commit(sql)
                                # find and delete records where a carrier has only 'one out of station' record
                                sql = "SELECT station FROM carriers WHERE carrier_name = '%s'" \
                                      % carr[0]
                                if len(inquire(sql)) == 1:
                                    sql = "DELETE FROM carriers WHERE carrier_name = '%s' AND station = '%s'" \
                                          % (carr[0], "out of station")
                                    commit(sql)
                        else:
                            sql = "DELETE FROM carriers WHERE carrier_name = '%s' {}".format(operator) % car[0]
                            commit(sql)
                    pb.stop()  # stop and destroy the progress bar
                    pb_label.destroy()  # destroy the label for the progress bar
                    pb.destroy()
                    pb_root.destroy()
        messagebox.showinfo("Database Maintenance",
                            "Success! The database has been cleaned of the specified records.",
                            parent=frame)
        frame.destroy()
        self.run(masterframe)

    def database_delete_carriers(self, frame, station):
        """ build a screen to delete carriers. """
        self.win = MakeWindow()
        self.win.create(frame)
        Label(self.win.body, text="Delete Carriers", font=macadj("bold", "Helvetica 18")) \
            .grid(row=0, column=0, sticky="w")
        Label(self.win.body, text="").grid(row=1, column=0)
        Label(self.win.body, text="Select the station to see all carriers who have ever worked "
                                  "at the station - past and present. \nDeleting the carrier will"
                                  "result in all records for that carrier being deleted. This "
                                  "includes clock \nrings and name indexes. ", justify=LEFT) \
            .grid(row=2, column=0, sticky="w", columnspan=6)
        Label(self.win.body, text="").grid(row=3, column=0)
        Label(self.win.body, text="Select Station: ", anchor="w").grid(row=4, column=0, sticky="w")
        station_selection = StringVar(self.win.body)
        om_station = OptionMenu(self.win.body, station_selection, *projvar.list_of_stations)
        om_station.config(width=30, anchor="w")
        om_station.grid(row=5, column=0, columnspan=2, sticky="w")
        if station == "x":
            station_selection.set("Select a station")
        else:
            station_selection.set(station)
        Button(self.win.body, text="select", width=macadj(14, 12), anchor="w",
               command=lambda: self.database_chg_station(self.win.topframe, station_selection)) \
            .grid(row=5, column=2, sticky="w")
        Label(self.win.body, text="                ",
              anchor="w").grid(row=5, column=3, sticky="w")
        Label(self.win.body, text="").grid(row=6, column=0)
        sql = "SELECT DISTINCT carrier_name FROM carriers WHERE station = '%s' " \
              "ORDER BY carrier_name ASC" % station
        results = inquire(sql)
        if station != "x":
            Label(self.win.body, text="Carriers of {}".format(station), anchor="w").grid(row=7, column=0, sticky="w")
        results_frame = Frame(self.win.body)
        results_frame.grid(row=8, columnspan=4)
        i = 0
        car_vars = []
        if len(results) == 0 and station != "x":
            Label(results_frame, text="", anchor="w").grid(row=i, column=2, sticky="w")
            i += 1
            Label(results_frame, text="After a search, no carrier records were found in the Klustebox database",
                  anchor="w").grid(row=i, column=0, columnspan=3, sticky="w")
            Label(results_frame, text="                                    ",
                  anchor="w").grid(row=i, column=3, sticky="w")
        for name in results:
            sql = "SELECT MAX(effective_date), station FROM carriers WHERE carrier_name = '%s'" % name
            top_rec = inquire(sql)
            var = BooleanVar()
            chk = Checkbutton(results_frame, text=name[0], variable=var, anchor="w")
            chk.grid(row=i, column=0, sticky="w")
            car_vars.append((name[0], var))
            Label(results_frame, text=dt_converter(top_rec[0][0]).strftime("%m/%d/%Y"), anchor="w") \
                .grid(row=i, column=1, sticky="w")
            Label(results_frame, text="     ", anchor="w").grid(row=i, column=2, sticky="w")
            Label(results_frame, text=top_rec[0][1], anchor="w").grid(row=i, column=3, sticky="w")
            Label(results_frame, text="                 ", anchor="w").grid(row=i, column=4, sticky="w")
            i += 1
        # apply and close buttons
        button_apply = Button(self.win.buttons)
        button_back = Button(self.win.buttons)
        button_apply.config(text="Apply", width=15,
                            command=lambda: self.database_delete_carriers_apply(self.win.topframe,
                                                                                station_selection, car_vars))
        button_back.config(text="Go Back", width=15, command=lambda: MainFrame().start(frame=self.win.topframe))
        if sys.platform == "win32":
            button_apply.config(anchor="w")
            button_back.config(anchor="w")
        button_apply.pack(side=LEFT)
        button_back.pack(side=LEFT)
        self.win.finish()

    def database_chg_station(self, frame, station):
        """ delete the carrier in a station. """
        if station.get() == "Select a station":
            station_string = "x"
        else:
            station_string = station.get()
        self.database_delete_carriers(frame, station_string)

    def database_delete_carriers_apply(self, frame, station, car_vars):
        """ delete carriers from the database. """
        if station.get() == "Select a station":
            station_string = "x"
        else:
            station_string = station.get()

        del_holder = []
        for pair in car_vars:
            if pair[1].get():
                del_holder.append(pair[0])
        if len(del_holder) > 0:
            if messagebox.askokcancel("Delete Carrier Records",
                                      "Are you sure you want to delete {} carriers, \n"
                                      "along with all their clock rings and name indexes? \n\n"
                                      "This action is not reversible.".format(len(del_holder)),
                                      parent=frame):
                pb_root = Tk()  # create a window for the progress bar
                pb_root.title("Deleting Carrier Records")
                titlebar_icon(pb_root)
                pb_label = Label(pb_root, text="Running Process: ", anchor="w")  # make label for progress bar
                pb_label.grid(row=0, column=0, sticky="w")
                pb = ttk.Progressbar(pb_root, length=400, mode="determinate")  # create progress bar
                pb.grid(row=0, column=1, sticky="w")
                pb_text = Label(pb_root, text="", anchor="w")
                pb_text.grid(row=1, column=0, columnspan=2, sticky="w")
                steps = len(del_holder)
                pb_count = 0
                pb["maximum"] = steps  # set length of progress bar
                pb.start()
                for name in del_holder:
                    pb_count += 1
                    # change text for progress bar
                    pb_text.config(text="Deleting records for: {}".format(name))
                    pb_root.update()
                    pb["value"] = pb_count  # increment progress bar
                    sql = "DELETE FROM rings3 WHERE carrier_name = '%s'" % name
                    commit(sql)
                    sql = "DELETE FROM carriers WHERE carrier_name = '%s'" % name
                    commit(sql)
                    sql = "DELETE FROM name_index WHERE kb_name = '%s'" % name
                    commit(sql)
                    sql = "DELETE FROM seniority WHERE name = '%s'" % name
                    commit(sql)
                pb.stop()  # stop and destroy the progress bar
                pb_label.destroy()  # destroy the label for the progress bar
                pb.destroy()
                pb_root.destroy()
                self.database_delete_carriers(frame, station_string)
            else:
                return

    @staticmethod
    def database_rings_report(frame, station):
        """ generate a report summary of all clock rings for the station """
        gross_dates = []  # captures all dates of rings for given station
        # master_dates = []  # a distinct collection of dates for given station
        unique_dates = []
        # get a distinct list of all carriers who have ever been at the station
        sql = "SELECT DISTINCT carrier_name FROM carriers WHERE station = '%s' ORDER BY carrier_name" \
              % station
        results = inquire(sql)
        pb = ProgressBarDe(title="Clock Rings Summary", label="Gathering Data")
        pb.max_count(len(results))
        pb.start_up()
        count = 0
        for name in results:
            count += 1
            pb.move_count(count)
            pb.change_text(f"reading {name[0]}")
            active_station = []
            # get all records for the carrier from the carriers table
            sql = "SELECT * FROM carriers WHERE carrier_name= '%s' ORDER BY effective_date" % name[0]
            result_1 = inquire(sql)
            start_search = True
            start = ''
            end = ''
            # build the active_station array - find dates where carrier entered/left station
            for r in result_1:
                if r[5] == station and start_search:
                    start = r
                    start_search = False
                if r[5] != station and not start_search:
                    end = r
                    active_station.append([start, end])
                    start = ''
                    end = ''
                    start_search = True
            if not start_search:
                active_station.append([start, end])
            for active in active_station:
                if active[1] != '':
                    sql = "SELECT rings_date FROM rings3 WHERE rings_date " \
                          "BETWEEN '%s' AND '%s' AND carrier_name = '%s' " \
                          % (active[0][0], active[1][0], name[0])
                    the_dates = inquire(sql)
                    for td in the_dates:
                        gross_dates.append(td[0])
                else:
                    sql = "SELECT rings_date FROM rings3 WHERE rings_date >= '%s' AND carrier_name = '%s' " \
                          % (active[0][0], name[0])
                    the_dates = inquire(sql)
                    for td in the_dates:
                        gross_dates.append(td[0])
        pb.stop()
        for gd in gross_dates:  # get a list of unique dates
            if gd not in unique_dates:
                unique_dates.append(gd)
        unique_dates.sort(reverse=True)  # sort the unique dates in reverse order
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = "clock_rings_summary" + "_" + stamp + ".txt"
        pb = ProgressBarDe(title="Clock Rings Summary", label="Building Report")
        pb.max_count(len(unique_dates))
        pb.start_up()
        try:
            report = open(dir_path('report') + filename, "w")
            report.write("\nClock Rings Summary Report\n\n\n")
            report.write('   Showing results for:\n')
            report.write('   Station: {}\n'.format(station))
            report.write('\n')
            report.write('{:>4}  {:<26} {:<24}\n'.format("", "Date", "Records Available"))
            report.write('      --------------------------------------------\n')
            i = 1
            for line in unique_dates:
                report.write('{:>4}  {:<26} {:<24}\n'
                             .format("", dt_converter(line).strftime("%m/%d/%Y - %a"), gross_dates.count(line)))
                if i % 3 == 0:
                    report.write('      --------------------------------------------\n')
                pb.move_count(i)
                pb.change_text(f"building date: {dt_converter(line).strftime('%Y')}")
                i += 1
            report.write('\n')
            report.write('Total distinct dates for which clock ring records are available: {:<9}\n'.format(i - 1))
            report.close()
            if sys.platform == "win32":  # open the text document
                os.startfile(dir_path('report') + filename)
            if sys.platform == "linux":
                subprocess.call(["xdg-open", 'kb_sub/report/' + filename])
            if sys.platform == "darwin":
                subprocess.call(["open", dir_path('report') + filename])
        except PermissionError:
            pb.delete()
            messagebox.showerror("Report Generator", "The report failed to generate.", parent=frame)
        pb.stop()

    @staticmethod
    def carrier_list_cleaning(frame):
        """ cleans the database of duplicate database_delete_records """
        sql = "SELECT * FROM carriers ORDER BY carrier_name, effective_date"
        results = inquire(sql)
        duplicates = []
        for i in range(len(results)):
            if i != len(results) - 1:  # if the loop has not reached the end of the list
                if results[i][1] == results[i + 1][1] and \
                        results[i][2] == results[i + 1][2] and \
                        results[i][3] == results[i + 1][3] and \
                        results[i][4] == results[i + 1][4] and \
                        results[i][5] == results[i + 1][5]:  # if the name current and next name are the same
                    duplicates.append(i + 1)
        ok = False
        if len(duplicates) > 0:
            ok = messagebox.askokcancel("Database Maintenance",
                                        "Did you want to eliminate database redundancies? \n"
                                        "{} redundancies have been found in the database \n"
                                        "This is recommended maintenance.".format(len(duplicates)),
                                        parent=frame)
        if ok:
            pb_root = Tk()  # create a window for the progress bar
            pb_root.title("Database Maintenance")
            titlebar_icon(pb_root)  # place icon in titlebar
            pb_label = Label(pb_root, text="Updating Changes: ")  # make label for progress bar
            pb_label.pack(side=LEFT)
            pb = ttk.Progressbar(pb_root, length=400, mode="determinate")  # create progress bar
            pb.pack(side=LEFT)
            pb["maximum"] = len(duplicates)  # set length of progress bar
            pb.start()
            i = 0
            for d in duplicates:
                pb["value"] = i  # increment progress bar
                sql = "DELETE FROM carriers WHERE effective_date='%s' and carrier_name='%s'" % (
                    results[d][0], results[d][1])
                commit(sql)
                pb_root.update()
                i += 1
            pb.stop()  # stop and destroy the progress bar
            pb_label.destroy()  # destroy the label for the progress bar
            pb.destroy()
            pb_root.destroy()
            messagebox.showinfo("Database Maintenance",
                                "All redundancies have been eliminated from the carrier list.",
                                parent=frame)
            MainFrame().start(frame=frame)
        if not ok:
            messagebox.showinfo("Database Maintenance",
                                "No redundancies have been found in the carrier list.",
                                parent=frame)
        del duplicates[:]

    @staticmethod
    def backup(frame):
        """ create a copy of the database and save it in a folder designated by the user """
        filepath = filedialog.asksaveasfilename(filetypes=[("sqlite files", '*.sqlite')])
        if not filepath:
            return
        filepath = BuildPath().add_extension(filepath, "sqlite")
        dbasepath = BuildPath().location_dbase()
        try:
            copyfile(dbasepath, filepath)
            messagebox.showinfo("Klusterbox Database Back Up",
                                "Database back up successful. The location of the database back up file "
                                "is at {}".format(filepath),
                                parent=frame)
        except PermissionError:
            pass

    def dbase_loc(self):
        """ provides the location of the program """
        if sys.platform == "darwin":
            if projvar.platform == "macapp":
                self.dbase_location = os.path.expanduser("~") + '/Documents/.klusterbox'
            if projvar.platform == "py":
                self.dbase_location = os.getcwd() + '/kb_sub'
        else:
            if projvar.platform == "winapp":
                self.dbase_location = os.path.expanduser("~") + '\\Documents\\.klusterbox'
            else:
                self.dbase_location = os.getcwd() + '\\kb_sub'

    def file_dialogue(self):
        """ opens file folders to access generated kbreports """
        self.dbase_loc()  # get the location of the dbase and put it in the self.dbase_location var
        if not os.path.isdir(self.dbase_location):
            os.makedirs(self.dbase_location)
        filedialog.askopenfilename(initialdir=self.dbase_location)


class CarrierHistory:
    """ report window. generates the carrier status history screen. """

    def __init__(self):
        self.win = None
        self.carrier = None  # the carrier name
        self.effective_date = None  # the effective date - from the investigation range
        self.station = None  # the station as passed to create

    def create(self, frame, station):
        """ fills the screen with widgets. """
        self.station = station  # this updates the self.station variable
        self.get_effective_date()  # this will get the effective date from the investigation range
        self.win = MakeWindow()
        self.win.create(frame)
        Label(self.win.body, text="Carriers History", font=macadj("bold", "Helvetica 18")) \
            .grid(row=0, column=0, sticky="w")
        Label(self.win.body, text="").grid(row=1, column=0)
        Label(self.win.body, wraplength=macadj(500, 550),
              text="Select the station to see all carriers who have ever worked "
              "at the station - past and present.\n\n "
              "To move carriers back to the station, select Restore or Insert. "
              "New records for Restored or Inserted carriers will show the carriers with "
              "no list status, no ns day and no station. ", justify=LEFT)\
            .grid(row=2, column=0, sticky="w", columnspan=6)
        Label(self.win.body, text="").grid(row=3, column=0)
        Label(self.win.body, text="Select Station: ", anchor="w").grid(row=4, column=0, sticky="w")
        station_selection = StringVar(self.win.body)
        om_station = OptionMenu(self.win.body, station_selection, *projvar.list_of_stations)
        if sys.platform != "darwin":  # if the platform is not darwin
            om_station.config(width=30, anchor="w")
        else:
            om_station.config(width=30)
        om_station.grid(row=5, column=0, columnspan=2, sticky="w")
        if station == "x":
            station_selection.set("Select a station")
        else:
            station_selection.set(station)
        Button(self.win.body, text="select", width=macadj(14, 12), anchor="w",
               command=lambda: self.change_station(station_selection)) \
            .grid(row=5, column=2, sticky="w")
        Label(self.win.body, text="").grid(row=6, column=0)
        sql = "SELECT DISTINCT carrier_name FROM carriers WHERE station = '%s' " \
              "ORDER BY carrier_name ASC" % station
        results = inquire(sql)
        if station != "x":
            Label(self.win.body, text="Carriers of {}".format(station), anchor="w").grid(row=7, column=0, sticky="w")
        results_frame = Frame(self.win.body)
        results_frame.grid(row=8, columnspan=4)
        i = 0
        if station != "x":
            if len(results) > 0:
                Label(results_frame, text="Name", anchor="w", fg="grey").grid(row=i, column=0, sticky="w")
                Label(results_frame, text="Last Date", anchor="w", fg="grey") \
                    .grid(row=i, column=1, columnspan=2, sticky="w")
                Label(results_frame, text="Station", anchor="w", fg="grey").grid(row=i, column=3, sticky="w")
            elif len(results) == 0:
                Label(results_frame, text="", anchor="w").grid(row=i, column=0, sticky="w")
                i += 1
                Label(results_frame, text="After a search, no results were found in the klusterbox database.",
                      anchor="w") \
                    .grid(row=i, column=0, sticky="w")
        i += 1
        for name in results:
            sql = "SELECT MAX(effective_date), station FROM carriers WHERE carrier_name = '%s' and " \
                  "effective_date <= '%s'" % (name[0], self.effective_date)
            top_rec = inquire(sql)
            # name label
            Label(results_frame, text=name[0], anchor="w").grid(row=i, column=0, sticky="w")
            if top_rec[0][0]:  # if there is a record <= the investigation range.
                # date label
                Label(results_frame, text=dt_converter(top_rec[0][0]).strftime("%m/%d/%Y"), anchor="w") \
                    .grid(row=i, column=1, sticky="w")
                Label(results_frame, text="     ", anchor="w").grid(row=i, column=2, sticky="w")
                # station label
                Label(results_frame, text=top_rec[0][1], anchor="w").grid(row=i, column=3, sticky="w")
            else:  # if there is no current record, display "no record yet"
                Label(results_frame, text=">>> No record yet", anchor="w",
                      foreground="blue").grid(row=i, column=1, columnspan=2)
            Label(results_frame, text="     ", anchor="w").grid(row=i, column=4, sticky="w")
            # report button
            Button(results_frame, text="Report", anchor="w",
                   command=lambda in_line=name: Reports(self.win.topframe).rpt_carrier_history(in_line[0])) \
                .grid(row=i, column=5, sticky="w")
            Label(results_frame, text=" ", anchor="w").grid(row=i, column=6, sticky="w")
            # rings button
            Button(results_frame, text="Rings", anchor="w",
                   command=lambda in_line=name: Reports(self.win.topframe).rpt_all_rings(in_line[0])) \
                .grid(row=i, column=7, sticky="w")
            Label(results_frame, text=" ", anchor="w").grid(row=i, column=8, sticky="w")
            # restore button
            if station != "out of station":  # do not display if station is 'out of station'
                if top_rec[0][1] != station:  # do not display if carrier is currently at station.
                    text = "Restore"  # the default text for the restore button is "restore"
                    if not top_rec[0][1]:  # if there is no record for the carrier before this date...
                        text = "Insert"  # change the text in "insert"
                    Button(results_frame, text=text, anchor="w", width=macadj(6, 6),
                           command=lambda in_line=name: self.restore(in_line[0])) \
                        .grid(row=i, column=9, sticky="w")
            Label(results_frame, text="         ", anchor="w").grid(row=i, column=10, sticky="w")
            i += 1
        # close button
        button = Button(self.win.buttons)
        button.config(text="Go Back", width=macadj(20, 20),
                      command=lambda: MainFrame().start(frame=self.win.topframe))
        if sys.platform == "win32":
            button.config(anchor="w")
        button.pack(side=LEFT)
        self.win.finish()

    def change_station(self, station):
        """ gets the station """
        if station.get() == "Select a station":
            new_station = "x"
        else:
            new_station = station.get()
        self.create(self.win.topframe, new_station)

    def restore(self, carrier):
        """ this method moves the carrier from 'out of station' back to the current station.
        when the 'restore' button is pressed, the method is called. """
        self.carrier = carrier  # set the self.carrier var
        # sql select from carriers table when name and date match
        sql = "SELECT * FROM carriers WHERE carrier_name = '%s' and effective_date = '%s'" \
              % (self.carrier, self.effective_date)
        carrier_rec = inquire(sql)
        if carrier_rec:  # if there is a record - update the record in the carrier table.
            sql = "UPDATE carriers SET station = '%s' WHERE carrier_name = '%s' and effective_date = '%s'" \
                  % (self.station, self.carrier, self.effective_date)
            commit(sql)
        else:  # else - if there is no result - insert a record into carrier table.
            sql = "INSERT INTO carriers (effective_date,carrier_name,list_status,ns_day,route_s,station)" \
                  " VALUES('%s','%s','%s','%s','%s','%s')" \
                  % (self.effective_date, self.carrier, "nl", "none", "", self.station)
            commit(sql)
        self.create(self.win.topframe, self.station)

    def get_effective_date(self):
        """ this will get the effective date of the change in carrier status """
        if projvar.invran_weekly_span:  # if the invran is weekly - go with first day of week
            self.effective_date = projvar.invran_date_week[0]
        else:  # if the invran is daily - go with the day.
            self.effective_date = projvar.invran_date


class PdfSplitter:
    """
    The PDF Splitter. Builds a screen that allows the user to split a PDF.
    """

    def __init__(self):
        self.subject_path = None
        self.frame = None
        self.win = None
        self.new_path = None
        self.firstpage = None
        self.lastpage = None

    def get_file_path(self):
        """ Created for pdf splitter - gets a pdf file """
        path_ = dir_filedialog()  # get the pdf file
        file_path = filedialog.askopenfilename(initialdir=path_,
                                               filetypes=[("PDF files", "*.pdf")], title="Select PDF")
        self.subject_path.set(file_path)

    def get_new_path(self):
        """ Created for pdf splitter - creates/overwrites a pdf file """
        path_ = dir_filedialog()
        save_filename = filedialog.asksaveasfilename(initialdir=path_,
                                                     filetypes=[("PDF files", "*.pdf")], title="Overwrite/Create PDF")
        self.new_path.set(save_filename)

    def pdf_splitter_apply(self):
        """ check for empty fields / return if there are any errors """
        subject_path = self.subject_path.get().strip()
        firstpage = self.firstpage.get()
        lastpage = self.lastpage.get()
        new_path = self.new_path.get().strip()

        if subject_path == "":
            messagebox.showerror("Klusterbox PDF Splitter",
                                 "You must select a pdf file to split.",
                                 parent=self.win.topframe)
            return
        if new_path == "":
            messagebox.showerror("Klusterbox PDF Splitter",
                                 "You must designate a destination"
                                 " and a name for the df file you are creating.",
                                 parent=self.win.topframe)
            return
        # if the last characters are not .pdf then add the extension
        if new_path[-4:] != ".pdf":
            new_path += ".pdf"
        if firstpage > lastpage:
            messagebox.showerror("Klusterbox PDF Splitter",
                                 "The First Page of the document can not be "
                                 "higher than the Last Page.",
                                 parent=self.win.topframe)
            return
        try:
            pdf = PdfFileReader(subject_path)
            pdf_writer = PdfFileWriter()
            for page in range(firstpage - 1, lastpage):
                pdf_writer.addPage(pdf.getPage(page))
            with open(new_path, 'wb') as out:
                pdf_writer.write(out)
            if messagebox.askokcancel("Klusterbox PDF Splitter",
                                      "PDF file has been split sucessfully."
                                      "Do you want to open the pdf file?",
                                      parent=self.win.topframe):
                if sys.platform == "win32":
                    os.startfile(new_path)
                if sys.platform == "linux":
                    subprocess.call(["xdg-open", new_path])
                if sys.platform == "darwin":
                    subprocess.call(["open", new_path])
        except PermissionError:
            messagebox.showerror("Klusterbox PDF Splitter",
                                 "The PDF splitting has failed. \n"
                                 "It could be that that the pages set to be split don't exist \n"
                                 "or \n"
                                 "the pdf can't be split by this program due to formatting issues. \n"
                                 "For better results try www.sodapdf.com, google chrome or Adobe Acrobat "
                                 "Pro DC",
                                 parent=self.win.topframe)

    def run(self, frame):
        """ PDF Splitter - builds a screen to so the user can split pdfs. """
        self.win = MakeWindow()
        self.win.create(frame)
        Label(self.win.body, text="PDF Splitter", font=macadj("bold", "Helvetica 18"), anchor="w") \
            .grid(row=1, column=1, columnspan=4, sticky="w")
        Label(self.win.body, text="").grid(row=2)
        Label(self.win.body, text="Select pdf file you want to split:") \
            .grid(row=3, column=1, columnspan=4, sticky="w")
        self.subject_path = StringVar(self.win.body)
        Entry(self.win.body, textvariable=self.subject_path, width=macadj(95, 50)).grid(row=4, column=1, columnspan=4)
        Button(self.win.body, text="Select", width="10", command=lambda: self.get_file_path()) \
            .grid(row=5, column=1, sticky="w")
        Label(self.win.body, text="").grid(row=6)
        Label(self.win.body, text="Select range of pages you want to use to create the new file:") \
            .grid(row=7, column=1, columnspan=4, sticky="w")
        Label(self.win.body, text="First Page:  ").grid(row=8, column=1, sticky="e")
        self.firstpage = IntVar(self.win.body)
        Entry(self.win.body, textvariable=self.firstpage, width=8).grid(row=8, column=2, sticky="w")
        self.firstpage.set(1)
        Label(self.win.body, text="Last Page:  ").grid(row=9, column=1, sticky="e")
        self.lastpage = IntVar(self.win.body)
        Entry(self.win.body, textvariable=self.lastpage, width=8).grid(row=9, column=2, sticky="w")
        self.lastpage.set(1)
        Label(self.win.body, text="").grid(row=10)
        Label(self.win.body, text="Select pdf file you want to over write or a create a new file:") \
            .grid(row=11, column=1, columnspan=4, sticky="w")
        self.new_path = StringVar(self.win.body)
        Entry(self.win.body, textvariable=self.new_path, width=macadj(95, 50)) \
            .grid(row=12, column=1, columnspan=4, sticky="w")
        Button(self.win.body, text="Select", width="10", command=lambda: self.get_new_path()) \
            .grid(row=13, column=1, sticky="w")
        Label(self.win.body, text="").grid(row=14)
        Label(self.win.body, text="If all fields are filled out, split the file.") \
            .grid(row=15, column=1, columnspan=3, sticky="w")
        Button(self.win.body, text="Split PDF", width="10",
               command=lambda: self.pdf_splitter_apply()).grid(row=15, column=4, sticky="e")
        button_back = Button(self.win.buttons)
        button_back.config(text="Go Back", width=20, command=lambda: MainFrame().start(frame=self.win.topframe))
        if sys.platform == "win32":
            button_back.config(anchor="w")
        button_back.pack(side=LEFT)
        self.win.finish()


class AutoDataEntry:
    """
    this class allows the user to input Employee Everything Reports in the csv format. The reports must be for an
    entire service week. No longer or shorter. The Auto Data Entry reads the reports, ask for input from the user,
    and inputs the data into the database.
    """

    def __init__(self):
        self.frame = None
        self.file_path = None
        self.a_file = None  # the openned csv file
        self.tacs_station = None  # the station from the ee report
        self.t_range = None  # false - one day/ true - weekly ee report
        self.t_date = None  # the starting date of the pp
        self.station_index = []  # create a list of klusterbox stations
        self.possible_stations = []  # array of all stations in stations table minus station index
        self.tacs_list = []  # Get the names from tacs report
        self.check_these = []
        self.new_carrier = []  # new carriers who have duplicate names send these to auto indexer 6
        self.name_sorter = []  # stores stringvar objects for name pairing in an array
        self.to_addname = []  # initialize array of names to be added.
        self.tried_names = []
        self.is_mac = macadj(False, True)  # returns True for mac, False if not mac
        self.csv_fix = None
        self.target_file = None
        self.future_carriers = []  # carriers with recs in the future, but not the past.

    def run(self, frame):
        """ calls auto set up to get needed csv file. """
        self.frame = frame
        self.AutoSetUp(self).run(self.frame)

    def get_file(self):
        """ read the csv file and assign to self.a_file attribute """
        self.target_file = open(self.file_path, newline="")
        self.a_file = reader(self.target_file)

    def go_back(self, frame):
        """
        This first closes the opened csv file is being read
        Then destroys the temporary csv file created by CsvRepair() and referenced by self.file_path.
        Then the MainFrame() is called to return the user to the main screen.
        This is called with self.parent.go_back(frame)
        """
        self.target_file.close()
        self.csv_fix.destroy()
        MainFrame().start(frame=frame)

    class AutoSetUp:
        """ gets the needed csv file and reads the pay period"""

        def __init__(self, parent):
            self.parent = parent
            self.frame = None
            self.tacs_pp = None  # pay period read from csv file
            self.tacs_index = []  # create a list of tacs station names
            self.kb_stations = []  # array of all stations in stations table

        def run(self, frame):
            """ a master method for running methods in proper order. """
            self.frame = frame
            if not self.get_path():  # get the path to the employee everything report
                return  # return if invalid response
            self.parent.csv_fix = CsvRepair()  # create a CsvRepair object
            # returns a file path for a checked and, if needed, fixed csv file.
            self.parent.file_path = self.parent.csv_fix.run(self.parent.file_path)
            self.auto_precheck()  # delete recs from name index which don't have corresponding recs in carriers table
            self.parent.get_file()  # read the csv file and assign to self.a_file attribute
            if not self.check_file():  # check for invalid file, find station and pay period
                return  # return if invalid response
            if not self.check_range():  # check that file covers full service week
                return  # return if invalid response
            if not self.check_tacs_station():  # check that file has a station
                return  # return if invalid response
            self.get_tacs_date()  # get the date from tacs
            self.get_stations()  # build arrays of stations
            if self.parent.tacs_station not in self.tacs_index:
                self.parent.AutoIndexer1(self.parent).run(self.frame)
            else:
                self.parent.AutoIndexer2(self.parent).run(self.frame)

        def get_path(self):
            """ get the path to the employee everything report or return False """
            path_ = dir_filedialog()
            self.parent.file_path = filedialog.askopenfilename(initialdir=path_,
                                                               filetypes=[("Excel files", "*.csv *.xls")])
            if self.parent.file_path == "":  # if there is no selections - end
                return False
            elif self.parent.file_path[-4:].lower() == ".csv" or self.parent.file_path[-4:].lower() == ".xls":
                return True
            else:  # if an csv nor xls is selected - end
                messagebox.showerror("Report Generator",
                                     "The file you have selected is not a .csv or .xls file. "
                                     "You must select a file with a .csv or .xls extension.",
                                     parent=self.frame)
                return False

        @staticmethod
        def auto_precheck():
            """ delete any records from name index which don't have corresponding records in carriers table """
            sql = "SELECT kb_name FROM name_index"
            kb_name = inquire(sql)
            sql = "SELECT carrier_name FROM carriers"
            results = inquire(sql)
            carriers = []
            for item in results:
                if item not in carriers:
                    carriers.append(item)
            # create progressbar
            pb = ProgressBarDe(title="Database Maintenance", label="Updating Changes: ")
            pb.max_count(len(kb_name))
            pb.start_up()
            i = 0
            for name in kb_name:
                pb.move_count(i)  # increment progress bar
                if name not in carriers:
                    sql = "DELETE FROM name_index WHERE kb_name = '%s'" % name
                    commit(sql)
                i += 1
            pb.stop()  # stop and destroy the progress bar

        def check_file(self):
            """ check for invalid file, find station and pay period """
            self.parent.get_file()  # read the csv file
            cc = 0
            for line in self.parent.a_file:
                if cc == 0 and line[0][:8] != "TAC500R3":
                    messagebox.showwarning("File Selection Error",
                                           "The selected file does not appear to be an "
                                           "Employee Everything report.",
                                           parent=self.frame)
                    return False
                if cc == 3:
                    self.tacs_pp = line[0]  # find the pay period
                    self.parent.tacs_station = line[2]  # find the station
                    break
                cc += 1
            return True

        def check_range(self):  # check that file covers full service week
            """
            self.parent.a_file is not refreshed. So the loop will will start on line 4 of the csv file. The
            loop will read 150 lines of the code and pick up all the range_days to ensure that a full week is
            covered.
            """
            cc = 0
            range_days = ["Saturday", "Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
            for line in self.parent.a_file:  # find the range
                if line[18] in range_days:
                    range_days.remove(line[18])
                if cc == 150:
                    break  # survey 150 lines before breaking to anaylize results.
                cc += 1
            if len(range_days) > 5:
                self.parent.t_range = False  # set the range
                messagebox.showwarning("File Selection Error",
                                       "Employee Everything Reports that cover only one day /n"
                                       "are not supported in this version of Klusterbox.",
                                       parent=self.parent.frame)
                return False
            else:
                self.parent.t_range = True
            return True

        def check_tacs_station(self):
            """ make sure the csv has a stations """
            if len(self.parent.tacs_station) == 0:
                messagebox.showwarning("Auto Data Entry Error",
                                       "The Employee Everything Report is corrupt. Data Entry will stop.  \n"
                                       "The Employee Everything Report does not include "
                                       "information about the station. This could be caused by an error of the pdf "
                                       "converter. If you can obtain an Employee Everything Report from management in "
                                       "csv format, you should have better results.",
                                       parent=self.frame)
                return False
            return True

        def get_tacs_date(self):
            """ get the tacs date expressed as pay period """
            year = int(self.tacs_pp[:-3])
            pp = self.tacs_pp[-3:]
            self.parent.t_date = find_pp(year, pp)  # returns the starting date of the pp when given year and pay period

        def get_stations(self):
            """ inquires the database to get a list of stations. """
            sql = "SELECT tacs_station, kb_station, finance_num FROM station_index"
            results = inquire(sql)
            for line in results:
                self.parent.station_index.append(line[1])  # build station index
                self.tacs_index.append(line[0])  # build tacs_index
            sql = "SELECT station FROM stations"
            results = inquire(sql)
            for record in results:
                self.kb_stations.append(record[0])  # build kb_stations
            for item in self.kb_stations:
                self.parent.possible_stations.append(item)  # build possible stations
            self.parent.station_index.append("out of station")
            self.parent.possible_stations = \
                [x for x in self.parent.possible_stations if x not in self.parent.station_index]

    class AutoIndexer1:  # The station pairing screen
        """ This screen will only appear if station does not have a record in the station index table.
         If there is not a record in the staton index table, this screen will allow to pair the station with
         a station in the stations table which is not already paired in the station index. Or the screen will
         allow the user to enter a completely new station which will be added to the station table and the
         station index table. """

        def __init__(self, parent):
            self.parent = parent
            self.frame = None
            self.win = None  # creates a window object
            self.station_sorter = None
            self.station_new = None

        def run(self, frame):
            """ master method for running other methods. """
            self.frame = frame
            self.get_window_object()
            self.station_screen()

        def get_window_object(self):
            """ creates the window object. """
            self.win = MakeWindow()
            self.win.create(self.frame)

        def station_screen(self):
            """ pair station from tacs to correct station in klusterbox/ part 1 """
            Label(self.win.body, text="Station Pairing", font=macadj("bold", "Helvetica 18"), pady=10) \
                .grid(row=0, column=0, columnspan=4, sticky=W)  # page contents
            Label(self.win.body, text="Match the station detected from TACS with a pre-existing station\n "
                                      "or use ADD STATION to add the station if there isn't a match.", justify=LEFT) \
                .grid(row=1, column=0, columnspan=4, sticky=W)
            Label(self.win.body, text="Detected Station: ", anchor="w").grid(row=2, column=0, sticky="w")
            Label(self.win.body, text=self.parent.tacs_station, fg="blue").grid(row=3, column=0, columnspan=4)
            Label(self.win.body, text="Select Station: ", anchor="w").grid(row=4, column=0, sticky=W)
            self.station_sorter = StringVar(self.win.body)
            station_options = ["select matching station"] + self.parent.possible_stations + ["ADD STATION"]
            self.station_sorter.set(station_options[0])
            option_menu = OptionMenu(self.win.body, self.station_sorter, *station_options)
            option_menu.config(width=30)
            option_menu.grid(row=5, column=0, columnspan=2, sticky=W)
            Label(self.win.body, text=" ", justify=LEFT).grid(row=6, column=0, sticky=W)
            Label(self.win.body, text="If the station is not present in the drop down menu, select  \n "
                                      "ADD STATION from the menu and enter the new station name \n"
                                      "below to pair it with the station originating the report", justify=LEFT) \
                .grid(row=7, column=0, columnspan=4, sticky=W)
            Label(self.win.body, text=" ", justify=LEFT).grid(row=8, column=0, sticky=W)
            Label(self.win.body, text="Enter New Station Name: ", anchor="w") \
                .grid(row=9, column=0, columnspan=4, sticky=W)
            # insert entry for station name
            self.station_new = StringVar(self.win.body)
            Entry(self.win.body, width=35, textvariable=self.station_new).grid(row=10, column=0, columnspan=4, sticky=W)
            button_cancel = Button(self.win.buttons)  # cancel button
            button_cancel.config(text="Go Back", width=20, command=lambda: self.parent.go_back(self.win.topframe))
            if sys.platform == "win32":
                button_cancel.config(anchor="w")
            button_cancel.pack(side=LEFT)
            button_apply = Button(self.win.buttons)  # apply button
            button_apply.config(text="Submit", width=20, command=lambda: self.apply())
            if sys.platform == "win32":
                button_apply.config(anchor="w")
            button_apply.pack(side=LEFT)
            self.win.finish()  # close out the window function

        def apply(self):
            """ the method runs when the submit button is pressed. """
            if self.check():  # if the user entered data passes all checks
                self.insert()  # insert the user entered data into the database
                self.parent.AutoIndexer2(self.parent).run(self.win.topframe)
            else:  # if the user entered data fails the checks
                frame = self.win.topframe  # store the frame object so __init__ does not destroy it
                self.__init__(self.parent)  # re initialize the class
                self.run(frame)  # re run the methods of the class

        def check(self):
            """ this method ensures that the station input is valid. """
            self.station_new = self.station_new.get()
            self.station_new = self.station_new.strip()
            """ user didn't select station from the option menu or didn't select ADD STATION and entered a station 
            in the entry widget """
            if self.station_sorter.get() == "select matching station":  # user selected the label and not a station
                messagebox.showerror("Data Entry Error",
                                     "You must select a station or ADD STATION",
                                     parent=self.win.topframe)
                return False
            """ user selected add station but gave no station """
            if self.station_sorter.get() == "ADD STATION" and self.station_new == "":
                messagebox.showerror("Data Entry Error",
                                     "You must provide a name for the new station.",
                                     parent=self.win.topframe)
                return False
            """ user selected station and added station - error """
            if self.station_sorter.get() != "ADD STATION" and self.station_new != "":
                messagebox.showerror("Data Entry Error",
                                     "You can not select a station from the drop down menu AND enter "
                                     "a station in the text field.",
                                     parent=self.win.topframe)
                return False
            return True

        def insert(self):
            """ this inputs the station into the database and updates project variables. """
            if self.station_sorter.get() == "ADD STATION":
                """ if the user is using ADD STATION  to enter a new station not in the option menu """
                # add the new station to the stations table if it is not already there.
                if self.station_new not in projvar.list_of_stations:
                    sql = "INSERT INTO stations (station) VALUES('%s')" % self.station_new
                    commit(sql)
                    projvar.list_of_stations.append(self.station_new)  # add station to list of stations
                    DovBase().minimum_recs(self.station_new)  # put minimum recs into dov table for new station.
                # add the station to the station index
                sql = "INSERT INTO station_index (tacs_station, kb_station, finance_num) VALUES('%s','%s','%s')" \
                      % (self.parent.tacs_station, self.station_new, "")
                commit(sql)
                messagebox.showinfo("Database Updated",
                                    "The {} station has been added to the list of stations automatically "
                                    "recognized.".format(self.station_new),
                                    parent=self.win.topframe)
            else:
                # if the carrier is selecting a station from the drop down menu. add the station to the
                # station index
                sql = "INSERT INTO station_index (tacs_station, kb_station, finance_num) VALUES('%s','%s','%s')" \
                      % (self.parent.tacs_station, self.station_sorter.get(), "")
                commit(sql)
                messagebox.showinfo("Database Updated",
                                    "The {} station has been paired to the {} station. In the future, this association "
                                    "will be automatically recognized."
                                    .format(self.parent.tacs_station, self.station_sorter.get()),
                                    parent=self.win.topframe)

    class AutoIndexer2:  # Search for name matchs #1
        """ This screen will give the user the opportunity to pair carriers with records in the carrier table
        to new carriers from the employee everything report. Carriers with names that match exactly will be
        matched/paired automatically. Only carriers with no record in the name index will appear in this screen.
        If a new carrier has a name exactly matching an existing carrier, that carrier's employee id number
        will be added to the end of their name because the name is a unique identifier. """

        def __init__(self, parent):
            self.parent = parent
            self.frame = None
            self.name_index = []  # create a list of klusterbox names
            self.id_index = []  # create a list of emp ids
            self.c_list = []  # create a list of unique names from carrier list (a set)
            self.to_remove = []  # intialized array of names to be removed from tacs names
            self.win = None
            self.possible_names = []  # an array of possible matches of kb names and tacs names
            self.possible_match = False  # False if there are no possible matches ever

        def run(self, frame):
            """ namepairing_create """
            self.frame = frame
            self.set_globals()
            self.get_carrier_indexes()
            self.get_carrier_list()
            self.parent.get_file()  # read the csv file and assign to self.a_file attribute
            self.get_tacslist()
            self.remove_tacs_duplicates()
            self.insert_into_nameindex()
            self.get_new_carrier()
            self.get_future_carriers()
            self.limit_tacslist()
            self.get_name_index()
            self.namepairing_router()

        def set_globals(self):
            """ sets globals for the investigation range. This will change the main frame when it is called. """
            s_year = self.parent.t_date.strftime("%Y")
            s_mo = self.parent.t_date.strftime("%m")
            s_day = self.parent.t_date.strftime("%d")
            sql = "SELECT kb_station FROM station_index WHERE tacs_station = '%s'" % self.parent.tacs_station
            station = inquire(sql)
            Globals().set(s_year, s_mo, s_day, self.parent.t_range, station[0][0], "None")

        def get_carrier_indexes(self):
            """ gets carrier names and employee ids from the database."""
            sql = "SELECT tacs_name, kb_name, emp_id FROM name_index ORDER BY kb_name"
            results = inquire(sql)
            for line in results:
                self.name_index.append(line[1])  # create a list of klusterbox names
                self.id_index.append(line[2].zfill(8))  # create a list of emp ids

        def get_carrier_list(self):
            """ this method creates a list of carriers who are currently at the station during the dates of the
            investigation range. This is stored in the self.c_list array"""
            carrier_list = gen_carrier_list()  # generate an in range carrier list
            for each in carrier_list:  # create a list of unique names from carrier list (a set)
                if each[1] not in self.c_list:
                    self.c_list.append(each[1])

        def get_tacslist(self):
            """ Get the names from tacs report and create tacs_list """
            good_jobs = ("134", "844", "434")
            cc = 0
            for line in self.parent.a_file:
                if cc > 1 and line[19] in good_jobs:
                    # create a note for carrier's assignment - reg w/route, reg floater or aux
                    route = line[25].zfill(6)
                    lvl = line[23].zfill(2)
                    if line[19] == "134" and lvl == "01":
                        tac_route = route[1] + route[2] + route[3] + route[4] + route[5]
                        assignment = "reg " + Handler(tac_route).routes_adj()
                    elif line[19] == "134" and lvl == "02":
                        assignment = "reg " + "floater"
                    elif line[19] == "434":
                        assignment = "part time flex"
                    elif line[19] == "844":
                        assignment = "auxiliary"
                    else:
                        assignment = "undetected"
                    lastname = line[5].lower().replace("\'", " ")
                    add_to_list = [line[4].zfill(8), lastname, line[6].lower(),
                                   assignment]  # create list to insert in list
                    self.parent.tacs_list.append(add_to_list)
                cc += 1

        def remove_tacs_duplicates(self):
            """ I think this removes any multiple BASE/TEMP lines. """
            holder = ["", "", "", ""]  # find the duplicates and remove them where there is both BASE and TEMP
            put_back = []
            for item in self.parent.tacs_list:  # crawler goes down the list to identify Temp entries
                if item[0] == holder[0]:
                    if item == holder:
                        self.to_remove.append(holder)  # remove both records
                    if item != holder:
                        self.to_remove.append(holder)
                        self.to_remove.append(item)
                    put_back.append(item)  # put the later record back in the list
                holder = item  # hold the record to compare in the next loop
            # remove the duplicates
            self.parent.tacs_list = [x for x in self.parent.tacs_list if x not in self.to_remove]
            for record in put_back:  # put the Temp record back into the tacs_list
                self.parent.tacs_list.append(record)
            self.parent.tacs_list.sort(key=itemgetter(1))  # re-alphabetize the list of carriers

        def insert_into_nameindex(self):
            """ inserts new carriers into the name index. """
            sql = ""
            add = 0  # create tallies for reports
            rec = 0
            out = 0
            # carriers who are already or newly placed in name index - remove them from further processing
            self.to_remove = []
            pb = ProgressBarDe(title="Database Maintenance", label="Updating Changes: ")  # create progressbar
            pb.max_count(len(self.parent.tacs_list))  # set length of progress bar
            pb.start_up()
            i = 0
            for each in self.parent.tacs_list:
                pb.move_count(i)  # increment progress bar
                tac_str = "{}, {}".format(each[1], each[2])  # tac str is last name and first initial from tacs report
                # if there is an identical match between kb and tacs names:
                if tac_str in self.c_list and each[0] not in self.id_index:
                    # if there is a dup name / need a complete list of carrier names from index
                    if tac_str in self.name_index:
                        # maybe just pass information via new_carrier and add later
                        self.parent.new_carrier.append(each)
                    else:  # go ahead and pair the emp id with the name in carriers
                        sql = "INSERT INTO name_index (tacs_name, kb_name, emp_id ) VALUES('%s','%s','%s')" \
                              % (tac_str, tac_str, each[0])
                        self.name_index.append(tac_str)
                        self.id_index.append(each[0])
                    add += 1
                    commit(sql)  # commit the sql
                    self.to_remove.append(each[0])
                    self.name_index.append(tac_str)
                elif each[0] in self.id_index:  # RECOGNIZED -  the emp id is already in the name index
                    self.to_remove.append(each[0])
                    self.parent.check_these.append(each)
                    rec += 1
                    self.checkfortacsname(each)  # check to see if the name index record has a tacs name in it.
                else:
                    out += 1
                i += 1
            pb.stop()  # stop and destroy the progress bar

        @staticmethod
        def checkfortacsname(carrier_rec):
            """ check to see if the name index record has a tacs name in it.
                insert the tac name if the name index record does not have it. """
            sql = "SELECT tacs_name FROM name_index WHERE emp_id = '%s'" % carrier_rec[0]
            results = inquire(sql)  # execute the query
            if not results[0][0]:  # if there is no name in the tacs_name column
                tacs_str = carrier_rec[1] + ", " + carrier_rec[2]
                sql = "UPDATE name_index SET tacs_name = '%s' WHERE emp_id = '%s'" % (tacs_str, carrier_rec[0])
                commit(sql)  # commit the sql to the database

        def get_new_carrier(self):
            """ find the carriers in name_index who have records w/ eff dates in the future """
            dont_check = []  # remove items from check these if future carriers are found
            for name in self.parent.check_these:
                sql = "SELECT kb_name FROM name_index WHERE emp_id = '%s'" % name[0]
                result = inquire(sql)
                kb_name = result[0][0]  # capture the klusterbox name from name index
                sql = "SELECT effective_date,carrier_name FROM carriers " \
                      "WHERE carrier_name = '%s' AND effective_date <= '%s' " \
                      "ORDER BY effective_date DESC" % (kb_name, projvar.invran_date_week[0])
                result = inquire(sql)
                if not result:
                    self.parent.new_carrier.append(name)  # will add as new carrier in AI 3
                    dont_check.append(name[0])  # removes from check these array
                    self.to_remove.append(name[0])  # removes from tacs list
            # removes don't check from check these
            self.parent.check_these = [x for x in self.parent.check_these if x[0] not in dont_check]

        def get_new_carriers_not_indexed(self):
            """ need to add carriers from to_addname with no eariler records. """
            pass

        def get_future_carriers(self):
            """ get a list of carriers from tacs list who have records in the future, but not past. """
            for t in self.parent.tacs_list:
                for n in self.parent.new_carrier:
                    if t[0] == n[0]:  # if there is a match for emp ids
                        self.parent.future_carriers.append(t)  # build the future carriers array

        def limit_tacslist(self):
            """ deletes from the tacs list. """
            self.parent.tacs_list = [x for x in self.parent.tacs_list if x[0] not in self.parent.new_carrier]
            self.parent.tacs_list = [x for x in self.parent.tacs_list if x[0] not in self.to_remove]

        def get_name_index(self):
            """ gets names and employee ids and puts them into arrays. """
            sql = "SELECT tacs_name, kb_name, emp_id FROM name_index ORDER BY kb_name"
            results = inquire(sql)
            for item in self.name_index:
                self.parent.tried_names.append(item)
            self.name_index = []  # create a list of klusterbox names
            for line in results:
                self.name_index.append(line[1])

        def namepairing_router(self):
            """ route to appropriate function based on array contents """
            # all tacs list resolved/ nothing to check
            if len(self.parent.tacs_list) < 1 and len(self.parent.new_carrier) < 1 and len(self.parent.check_these) < 1:
                self.parent.AutoIndexer6(self.parent).run(self.frame)  # to straight to entering rings
            # all tacs list resolved/ new names unresolved
            elif len(self.parent.tacs_list) < 1 and len(self.parent.new_carrier) > 0:
                self.parent.AutoIndexer4(self.parent).run(self.frame)  # add new carriers in AI4
            # tacs and new carriers resolved/ carriers to check
            elif len(self.parent.tacs_list) < 1 and len(self.parent.new_carrier) < 1 and \
                    len(self.parent.check_these) > 0:
                # step to AI  to check discrepancies
                self.parent.AutoIndexer5(self.parent).run(self.frame)
            else:  # If there are candidates sort, generate PAIRING SCREEN 1
                self.namepairing_screen()

        def namepairing_screen(self):
            """ Pairing screen #1 """
            self.c_list = [x for x in self.c_list if x not in self.name_index]
            self.win = MakeWindow()
            self.win.create(self.frame)
            Label(self.win.body, text="Search for Name Matches #1", font=macadj("bold", "Helvetica 18"), pady=10) \
                .grid(row=0, column=0, sticky="w", columnspan=10)  # page contents
            wintext = "Look for possible matches for each unrecognized name. If the name has already been entered " \
                      "manually, you \n should be able to find it on this screen or the next. It is possible " \
                      "that the " \
                      "name has no match, if that is \n the case then select \"ADD NAME\" in the next screen. " \
                      "You can " \
                      "change the default between \"NOT FOUND\" and \n \"DISCARD\" using the buttons below. " \
                      "Information from TACS is shown in blue\n\n"
            mactext = \
                "Look for possible matches for each unrecognized name. If the name has already been \n" \
                "entered manually, you should be able to find it on this screen or the next. It is \n" \
                "possible that the name has no match, if that is the case then select \"ADD NAME\" in \n" \
                "the next screen. You can change the default between \"NOT FOUND\" and \"DISCARD\" \n" \
                "using the buttons below. Information from TACS is shown in blue\n\n"
            text = "Investigation Range: {0} through {1}\n\n".format(
                projvar.invran_date_week[0].strftime("%a - %b %d, %Y"),
                projvar.invran_date_week[6].strftime("%a - %b %d, %Y"))
            Label(self.win.body, text=macadj(wintext, mactext) + text, justify=LEFT) \
                .grid(row=1, column=0, columnspan=10, sticky="w")
            Button(self.win.body, text="DISCARD", width=10,
                   command=lambda: self.indexer_default(self.parent.name_sorter, i + 1, name_options, 1)) \
                .grid(row=2, column=3, sticky="w", columnspan=2)
            Label(self.win.body, text="switch default to DISCARD").grid(row=2, column=1, sticky="w", columnspan=2)
            Button(self.win.body, text="NOT FOUND", width=10,
                   command=lambda: self.indexer_default(self.parent.name_sorter, i + 1, name_options, 0)) \
                .grid(row=3, column=3, sticky="w", columnspan=2)
            Label(self.win.body, text="switch default to NOT FOUND").grid(row=3, column=1, sticky="w", columnspan=2)
            Label(self.win.body, text="").grid(row=4, column=0)
            Label(self.win.body, text="Name", fg="grey").grid(row=5, column=1, sticky="w")
            Label(self.win.body, text="Assignment", fg="grey").grid(row=5, column=2, sticky="w")
            Label(self.win.body, text="Candidates", fg="grey").grid(row=5, column=3, sticky="w")
            cc = 6
            i = 0
            color = "blue"
            for t_name in self.parent.tacs_list:  # for each name in the tacs report
                self.get_possible_names(t_name)  # fill the self.possible names array
                if self.possible_names:
                    Label(self.win.body, text=str(i + 1), anchor="w").grid(row=cc, column=0, sticky="w")
                    Label(self.win.body, text=t_name[1] + ", " + t_name[2], anchor="w", width=15, fg=color) \
                        .grid(row=cc, column=1, sticky="w")  # name
                    Label(self.win.body, text=t_name[3], anchor="w", width=10, fg=color) \
                        .grid(row=cc, column=2, sticky="w")  # assignment
                name_options = ["NOT FOUND", "DISCARD"] + self.possible_names
                self.parent.name_sorter.append(StringVar(self.win.body))
                option_menu = OptionMenu(self.win.body, self.parent.name_sorter[i], *name_options)
                self.parent.name_sorter[i].set(name_options[0])
                option_menu.config(width=15)
                if self.possible_names:
                    option_menu.grid(row=cc, column=3, sticky="w")  # possible matches
                    if len(self.possible_names) == 1:  # display indicator for possible matches
                        Label(self.win.body, text=str(len(self.possible_names)) + " name") \
                            .grid(row=cc, column=4, sticky="w")
                    elif len(self.possible_names) > 1:
                        Label(self.win.body, text=str(len(self.possible_names)) + " names") \
                            .grid(row=cc, column=4, sticky="w")
                    else:  # display indicator for possible matches
                        Label(self.win.body, text="no match", fg="grey").grid(row=cc, column=4, sticky="w")
                cc += 1
                i += 1
            Button(self.win.buttons, text="Continue", width=macadj(15, 16),
                   command=lambda: self.parent.AutoIndexer3(self.parent).run(self.win.topframe)).grid(row=0, column=0)
            Button(self.win.buttons, text="Cancel", width=macadj(15, 16),
                   command=lambda: self.parent.go_back(self.win.topframe)).grid(row=0, column=1)
            if not self.possible_match:  # if there are no possible matches to any carrier names
                self.parent.AutoIndexer3(self.parent).run(self.win.topframe)  # go to next screen
            else:
                self.win.finish()  # otherwise stay on this screen

        def get_possible_names(self, t_name):
            """ gets possible name matches for names not in the name index. """
            self.possible_names = []
            for c_name in self.c_list:
                """ if the first letter of a carrier name has no record in the name index and matches the 
                 carrier name from the tacs report - append the name to the possible names array so it cann 
                 be used in the option menu. """
                if c_name[0:3] == t_name[1][0:3]:
                    self.possible_names.append(c_name)
                    self.parent.tried_names.append(c_name)
                    self.possible_match = True

        @staticmethod
        def indexer_default(widget, count, options, choice):
            """ changes the default for the optionmenu widget """
            for i in range(count - 1):
                widget[i].set(options[choice])

    class AutoIndexer3:
        """ Carrier pairing screen - allows users to match new carrier entries to carriers already in klusterbox."""

        def __init__(self, parent):
            self.parent = parent
            self.frame = None
            self.to_remove = []  # intialized array of names to be removed from tacs names
            self.to_nameindex = []  # initialize array of names to be be paired in name index
            self.c_list = []
            self.win = None
            self.n_index = []  # an array of all klusterbox names with records in the name index table
            self.to_chg = []  # changes for apply ai3
            self.new_name = []  # array of new names which have been modified with emp id

        def run(self, frame):
            """ master method for running other methods. """
            self.frame = frame
            self.apply_namepairing_1()  # apply pairing screen
            # if empty tacs list and something in check these
            if len(self.parent.tacs_list) < 1 and len(self.parent.check_these) > 0:
                self.parent.AutoIndexer5(self.parent).run(self.frame)
            elif len(self.parent.tacs_list) < 1 and len(self.parent.check_these) < 1:
                self.parent.AutoIndexer6(self.parent).run(self.frame)
            else:
                self.build_namepairing_options()
                self.namepairing_screen_2()  # create pairing screen #2

        def apply_namepairing_1(self):
            """ apply pairing screen #1 / AutoIndexer 2 """
            i = 0  # count iterations of loops
            dis = 0  # count of discarded items
            out = 0  # count of unresolved items
            pair = 0  # count of added items
            self.to_remove = []  # intialized array of names to be removed from tacs names
            not_found = []  # initialize array of names to be futher analyzed.
            for item in self.parent.name_sorter:
                if item.get() == "DISCARD":
                    self.to_remove.append(self.parent.tacs_list[i][0])
                    dis += 1
                elif item.get() == "NOT FOUND":
                    not_found.append(self.parent.tacs_list[i])
                    out += 1
                else:
                    to_add = [self.parent.tacs_list[i], item.get()]
                    self.to_nameindex.append(to_add)
                    self.to_remove.append(self.parent.tacs_list[i][0])
                    self.parent.check_these.append(self.parent.tacs_list[i])
                    pair += 1
                i += 1
            self.parent.tacs_list = [x for x in self.parent.tacs_list if x[0] not in self.to_remove]
            for item in self.to_nameindex:
                # tac str is last name and first initial from tacs report
                tac_str = "{}, {}".format(item[0][1], item[0][2])
                sql = "INSERT INTO name_index (tacs_name, kb_name, emp_id) VALUES('%s','%s','%s')" \
                      % (tac_str, item[1], item[0][0])
                commit(sql)
            """
        # message screens to summerize output
            messagebox.showinfo("Processing Carriers", 
            "{} Carrier names were paired to names in klusterbox\n"
            "{} Carrier names were discarded.\n"
            "{} Carrier names have not been handled."
            .format(pair, dis, out), parent=frame)
            """

        def build_namepairing_options(self):
            """ build possible names for option menus """
            sql = "SELECT kb_name FROM name_index"
            results = inquire(sql)
            name_result = []  # create a list of klusterbox names
            for line in results:
                name_result.append(line[0])
            sql = "SELECT carrier_name FROM carriers ORDER BY carrier_name"  # get all names from the carrier list
            results = inquire(sql)  # call function to access database
            for item in results:
                if item[0] not in self.c_list and item[0] not in self.parent.tried_names and item[0] not in name_result:
                    self.c_list.append(item[0])

        def namepairing_screen_2(self):
            """ create pairing screen #2 """
            self.win = MakeWindow()
            self.win.create(self.frame)
            self.parent.name_sorter = []  # page contents
            Label(self.win.body, text="Search for Name Matches #2", font=macadj("bold", "Helvetica 18"), pady=10) \
                .grid(row=0, column=0, sticky="w", columnspan=10)  # page contents
            wintext = \
                "Look for possible matches for each unrecognized name. If the name has already been entered " \
                "manually, \n" \
                " you should be able to find it on this screen. It is possible that the name has no match, if " \
                "that is \n" \
                "the case then select \"ADD NAME\" in this screen. You can change the default between \"ADD NAME\" " \
                "and \n" \
                "\"DISCARD\" using the buttons below. Information from TACS is shown in blue\n\n"
            mactext = \
                "Look for possible matches for each unrecognized name. If the name has already been \n" \
                "entered manually, you should be able to find it on this screen. It is possible that \n" \
                "the name has no match, if that is the case then select \"ADD NAME\" in this screen. \n" \
                "You can change the default between \"ADD NAME\" and \"DISCARD\" using the buttons \n" \
                "below. Information from TACS is shown in blue\n\n"
            text = "Investigation Range: {0} through {1}\n\n".format(
                projvar.invran_date_week[0].strftime("%a - %b %d, %Y"),
                projvar.invran_date_week[6].strftime("%a - %b %d, %Y"))
            Label(self.win.body, text=macadj(wintext, mactext) + text, justify=LEFT) \
                .grid(row=1, column=0, columnspan=10, sticky="w")
            Button(self.win.body, text="DISCARD", width=10,
                   command=lambda: self.indexer_default(self.parent.name_sorter, i + 1, name_options, 1)) \
                .grid(row=2, column=3, sticky="w", columnspan=2)
            Label(self.win.body, text="switch default to DISCARD").grid(row=2, column=1, sticky="w", columnspan=2)
            Button(self.win.body, text="ADD NAME", width=10,
                   command=lambda: self.indexer_default(self.parent.name_sorter, i + 1, name_options, 0)) \
                .grid(row=3, column=3, sticky="w", columnspan=2)
            Label(self.win.body, text="switch default to ADD NAME").grid(row=3, column=1, sticky="w", columnspan=2)
            Label(self.win.body, text="").grid(row=4, column=0)
            Label(self.win.body, text="Name", fg="grey").grid(row=5, column=1, sticky="w")
            Label(self.win.body, text="Assignment", fg="grey").grid(row=5, column=2, sticky="w")
            Label(self.win.body, text="Candidates", fg="grey").grid(row=5, column=3, sticky="w")
            cc = 6  # item and grid row counter
            i = 0  # count iterations of the loop
            color = "blue"
            for t_name in self.parent.tacs_list:
                possible_names = []
                Label(self.win.body, text=str(i + 1), anchor="w").grid(row=cc, column=0)
                fullname = t_name[1] + ", " + t_name[2]
                Label(self.win.body, text=fullname, anchor="w", width=15, fg=color) \
                    .grid(row=cc, column=1)  # name
                Label(self.win.body, text=t_name[3], anchor="w", width=10, fg=color) \
                    .grid(row=cc, column=2)  # assignment
                # build option menu for unmatched tacs names
                for c_name in self.c_list:
                    if c_name[0] == t_name[1][0]:
                        possible_names.append(c_name)
                name_options = ["ADD NAME", "DISCARD"] + possible_names
                self.parent.name_sorter.append(StringVar(self.win.body))
                option_menu = OptionMenu(self.win.body, self.parent.name_sorter[i], *name_options)
                if fullname in possible_names:
                    self.parent.name_sorter[i].set(fullname)
                else:
                    self.parent.name_sorter[i].set(name_options[0])
                option_menu.config(width=15)
                option_menu.grid(row=cc, column=3)  # possible matches
                if len(possible_names) == 1:  # display indicator for possible matches
                    Label(self.win.body, text=str(len(possible_names)) + " name").grid(row=cc, column=4)
                if len(possible_names) > 1:
                    Label(self.win.body, text=str(len(possible_names)) + " names").grid(row=cc, column=4)
                cc += 1
                i += 1
            Button(self.win.buttons, text="Continue", width=macadj(15, 16),
                   command=lambda: self.ai3_apply()) \
                .grid(row=0, column=0)
            Button(self.win.buttons, text="Cancel", width=macadj(15, 16),
                   command=lambda: self.parent.go_back(self.win.topframe)).grid(row=0, column=1)
            self.win.finish()

        @staticmethod
        def indexer_default(widget, count, options, choice):
            """ changes the default for the optionmenu widget """
            for i in range(count - 1):
                widget[i].set(options[choice])

        def ai3_apply(self):
            """ apply pairing screen 2 """
            self.build_n_index()
            self.ai3_apply_sort()  # discard, add or pair name
            self.insert_to_nameindex()  # add names to name index
            self.insert_to_addname()  # add names to name index
            self.get_future_carriers()  # get carriers with no prior emp id/ no past rec but a future rec.
            # self.apply_ai3_report()  # message screens to summerize output
            self.build_addname()  # build to_addname array
            if len(self.parent.to_addname) > 0:  # route conditional on arrays
                self.parent.AutoIndexer4(self.parent).run(self.win.topframe)
            elif len(self.parent.check_these) > 0:
                self.parent.AutoIndexer5(self.parent).run(self.win.topframe)
            else:
                self.parent.AutoIndexer6(self.parent).run(self.win.topframe)

        def build_n_index(self):
            """ creates an array of all klusterbox names with records in the name index table. """
            sql = "SELECT tacs_name, kb_name, emp_id FROM name_index ORDER BY kb_name"
            results = inquire(sql)
            self.n_index = []  # create a list of klusterbox names
            for line in results:  # loop to fill arrays
                self.n_index.append(line[1])

        def ai3_apply_sort(self):
            """ discard, add or pair name """
            i = 0  # count iterations of the loops..
            for item in self.parent.name_sorter:  # sort passed data from auto index 4
                if item.get() == "DISCARD":
                    self.to_remove.append(self.parent.tacs_list[i][0])
                    # dis += 1  # count of discarded items
                elif item.get() == "ADD NAME":
                    self.parent.to_addname.append(self.parent.tacs_list[i])
                    # add += 1
                else:
                    to_add = [self.parent.tacs_list[i], item.get()]
                    self.to_nameindex.append(to_add)
                    self.to_remove.append(self.parent.tacs_list[i][0])
                    self.parent.check_these.append(self.parent.tacs_list[i])
                    # pair += 1  # count of paired items
                i += 1

        def insert_to_nameindex(self):
            """ add names to name index """
            pb_label = Label(self.win.buttons, text="Updating Changes: ")  # make label for progress bar
            pb_label.grid(row=0, column=2)
            pb = ttk.Progressbar(self.win.buttons, length=250, mode="determinate")  # create progress bar
            pb.grid(row=0, column=3)
            pb["maximum"] = len(self.to_nameindex)  # set length of progress bar
            pb.start()
            i = 0
            for item in self.to_nameindex:  # when a name from the optionmenu was selected
                if self.no_record(item[0][0]):  # check for a record in name index by employee id #
                    # tac str is last name and first initial from tacs report
                    tac_str = "{}, {}".format(item[0][1], item[0][2])
                    sql = "INSERT INTO name_index (tacs_name, kb_name, emp_id) VALUES('%s','%s','%s')" \
                          % (tac_str, item[1], item[0][0])
                    commit(sql)
                pb["value"] = i  # increment progress bar
                self.win.buttons.update()  # update the progress bar
                i += 1
            pb.stop()  # stop and destroy the progress bar
            pb_label.destroy()  # destroy the label for the progress bar
            pb.destroy()

        @staticmethod
        def no_record(empid):
            """ check for a record in name index by employee id """
            sql = "SELECT emp_id FROM name_index WHERE emp_id = '%s'" % empid
            result = inquire(sql)
            if result:
                return False  # if there is a record
            return True  # if there is no record

        def insert_to_addname(self):
            """ add names to name index """
            self.to_chg = []  # array of items from to_addname where the name needs to be modified with emp id
            self.new_name = []  # array of new names which have been modified with emp id
            for name in self.parent.new_carrier:
                self.parent.to_addname.append(name)  # add new carriers in list to be added to carrier table
            # delete any item from new carrier that has been added to addname
            self.parent.new_carrier = [x for x in self.parent.new_carrier if x not in self.parent.to_addname]
            pb_label = Label(self.win.buttons, text="Updating Changes: ")  # make label for progress bar
            pb_label.grid(row=0, column=2)
            pb = ttk.Progressbar(self.win.buttons, length=250, mode="determinate")  # create progress bar
            pb.grid(row=0, column=3)
            pb["maximum"] = len(self.parent.to_addname)  # set length of progress bar
            pb.start()
            i = 0
            for item in self.parent.to_addname:  # when add name was selected from option menu
                pb["value"] = i  # increment progress bar
                tacs_str = "{}, {}".format(item[1], item[2])  # tacs str is last name and first initial from tacs report
                kb_str = "{}, {}".format(item[1], item[2])  # kb str is last name and first initial from tacs report
                if kb_str in self.n_index or kb_str in self.c_list:  # detect matches with name index
                    sql = "SELECT emp_id, kb_name FROM name_index WHERE emp_id = '%s'" % item[0]
                    result = inquire(sql)
                    if not result:
                        kb_str = "{} {}".format(kb_str, item[0])
                        self.to_chg.append(item)
                        mod_name = "{} {}".format(item[2], item[0])
                        self.new_name.append(mod_name)
                    if result:  # if the carrier is in the name index
                        # if the kb name is not the same in the name index record - change name
                        if result[0][1] != kb_str:
                            self.to_chg.append(item)
                            mod_name = result[0][1].split(",")
                            mod_name = mod_name[1].strip()
                            self.new_name.append(mod_name)
                self.n_index.append(kb_str)  # add to n_index array so dups can be detected
                sql = "SELECT emp_id FROM name_index WHERE emp_id = '%s'" % item[0]
                result = inquire(sql)
                if not result:
                    sql = "INSERT INTO name_index (tacs_name, kb_name, emp_id) VALUES('%s','%s','%s')" \
                          % (tacs_str, str(kb_str), item[0])
                    commit(sql)
                self.win.buttons.update()  # update the progress bar
                i += 1
            pb.stop()  # stop and destroy the progress bar
            pb_label.destroy()  # destroy the label for the progress bar
            pb.destroy()

        def get_future_carriers(self):
            """ make list of carriers which rec in the future but none present or past. """
            # make a list of candidates out of all names selected and add names in ai3.
            possible_future = self.parent.check_these + self.parent.to_addname
            for name in possible_future:
                # get the klusterbox name with the employee id.
                sql = "SELECT kb_name FROM name_index WHERE emp_id = '%s'" % name[0]
                result = inquire(sql)
                kb_name = result[0][0]  # capture the klusterbox name from name index
                # check if there is any record in the past.
                sql = "SELECT effective_date,carrier_name FROM carriers " \
                      "WHERE carrier_name = '%s' AND effective_date <= '%s' " \
                      "ORDER BY effective_date DESC" % (kb_name, projvar.invran_date_week[0])
                result = inquire(sql)
                if not result:  # if there is not a record in the past...
                    if name not in self.parent.future_carriers:  # if they are not already in the array...
                        self.parent.future_carriers.append(name)  # add them so they can have rec added in ai4

        def apply_ai3_report(self):
            """ message screens to summerize output """
            messagebox.showinfo("Processing Carriers", "{} Carrier names were added to the database\n"
                                                       "{} Carrier names were paired to names in klusterbox\n"
                                                       "{} Carrier names were discarded.\n"
                                .format(len(self.parent.to_addname), len(self.to_nameindex),
                                        len(self.to_remove)), parent=self.win.topframe)

        def build_addname(self):
            """ add carriers with emp id#s to  to_addname array """
            count = 0  # swap out the names which have been modified in self.parent.to_addname
            for item in self.to_chg:  # for each item to be swapped
                self.parent.to_addname.remove(item)  # clear out the old one
                # create a modified array with modified name
                mod_str = [item[0], item[1], self.new_name[count], item[3]]
                self.parent.to_addname.append(mod_str)  # put in the new one
                count += 1

    class AutoIndexer4:
        """
        input new carrier information after a check
        """

        def __init__(self, parent):
            self.parent = parent
            self.frame = None
            self.opt_nsday = []  # make an array of "day / color" options for option menu
            self.full_ns_dict = {}
            self.ns_dict = {}  # create dictionary for ns day data
            self.eff_date = None  # effective date for apply
            self.station = None  # station as stringvar for apply
            self.changecount = None
            self.win = None
            self.ai4_carrier_name = []  # create array for carrier names
            self.ai4_l_s = []  # create array for list status
            self.ai4_l_ns = []  # create array for ns days
            self.ai4_route = []  # create array for route/s
            self.clean_ns = None  # temp string var for ns day

        def run(self, frame):
            """ add new carriers to carrier table / pairing screen #3 """
            self.frame = frame
            self.add_future_carriers()
            self.ai4_opt_nsday()
            self.ai4_full_ns_dict()
            self.ai4_ns_dict()
            self.ai4_screen()

        def add_future_carriers(self):
            """ add carriers with records in future but need new one for present."""
            change = False  # identifies if a change occurs - change will trigger a sort.
            for fc in self.parent.future_carriers:  # for each carrier in future_carriers.
                if fc not in self.parent.to_addname:  # if they are not in to_addname
                    self.parent.to_addname.append(fc)  # add them to to_addname
                    change = True
            if change:  # if there is a change to the to_addname array.
                self.parent.to_addname.sort(key=itemgetter(1))  # sort the array by name.

        def ai4_opt_nsday(self):
            """ get ns structure preference from database """
            sql = "SELECT tolerance FROM tolerances WHERE category='%s'" % "ns_auto_pref"
            result = inquire(sql)
            ns_toggle = result[0][0]  # modify available ns days per ns_toggle
            if ns_toggle == "rotation":
                remove_array = ("sat", "mon", "tue", "wed", "thu", "fri")
            else:
                remove_array = ("green", "brown", "red", "black", "yellow", "blue")
            ns_code_mod = dict()  # copy the projvar.ns_code dict to ns_code_mod using dict()
            for key in projvar.ns_code:
                ns_code_mod[key] = projvar.ns_code[key]
            for key in remove_array:
                if key in ns_code_mod:
                    del ns_code_mod[key]  # modify available ns days per ns_toggle
            for each in ns_code_mod:  #
                ns_option = ns_code_mod[each] + " - " + each  # make a string for each day/color
                if each == "none":
                    ns_option = "       " + " - " + each  # if the ns day is "none" - make a special string
                self.opt_nsday.append(ns_option)

        def ai4_full_ns_dict(self):
            """ get a dictionary of ns days and colors. """
            days = ("Saturday", "Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday")
            for each in self.opt_nsday:  # Make a dictionary to match full days and option menu options
                for day in days:
                    if day[:3] == each[:3]:
                        self.full_ns_dict[day] = each  # creates full_ns_dict
                if each[-4:] == "none":
                    ns_option = "       " + " - " + "none"  # if the ns day is "none" - make a special string
                    self.full_ns_dict["None"] = ns_option  # creates full_ns_dict None option

        def ai4_ns_dict(self):
            """ creates the ns dict variable with carrier emp id and name. """
            results = gen_ns_dict(self.parent.file_path, self.parent.to_addname)  # returns id and name
            for ids in results:  # loop to fill dictionary with ns day info
                self.ns_dict[ids[0]] = ids[1]
            return self.ns_dict

        def ai4_screen(self):
            """ builds the screen and fills it with widget. """
            self.win = MakeWindow()
            self.win.create(self.frame)
            Label(self.win.body, text="Input New Carriers", font=macadj("bold", "Helvetica 18"), pady=10) \
                .grid(row=0, column=0, sticky="w", columnspan=6)  # Pairing Screen #3
            wintext = \
                "Enter in information for carriers not already recorded in the Klusterbox database. You can use " \
                "the TACS \n" \
                "information (shown in blue),as a guide if it is accurate. As OTDL/WAL information is not in TACS, " \
                "it is \n" \
                "not shown and this information will have to requested from management. Routes must be only 4 " \
                "digits \n" \
                "long. In cases were there are multiple routes, the routes must be separated by a \"/\" backslash.\n\n"
            mactext = \
                "Enter in information for carriers not already recorded in the Klusterbox database. You can \n" \
                "use the TACS information (shown in blue),as a guide if it is accurate. As OTDL/WAL \n" \
                "information is not in TACS, it is not shown and this information will have to requested \n" \
                "from management. Routes must be only 4 digits long. In cases were there are multiple \n" \
                "routes, the routes must be separated by a \"/\" backslash.\n\n"
            text = "Investigation Range: {0} through {1}\n\n" \
                .format(projvar.invran_date_week[0].strftime("%a - %b %d, %Y"),
                        projvar.invran_date_week[6].strftime("%a - %b %d, %Y"))
            # is_mac = macadj(False, True)
            Label(self.win.body, text=macadj(wintext, mactext) + text, justify=LEFT) \
                .grid(row=1, column=0, sticky="w", columnspan=6)
            y = 2  # count for the row
            Label(self.win.body, text="Name", fg="Grey").grid(row=y, column=0, sticky="w")
            Label(self.win.body, text=macadj("List Status", "List"), fg="Grey").grid(row=y, column=1, sticky="w")
            Label(self.win.body, text="NS Day", fg="Grey").grid(row=y, column=2, sticky="w")
            Label(self.win.body, text="Route_s", fg="Grey").grid(row=y, column=3, sticky="w")
            if not self.parent.is_mac:
                Label(self.win.body, text="Station", fg="Grey").grid(row=y, column=4, sticky="w")
                Label(self.win.body, text="              ", fg="Grey").grid(row=y, column=5, sticky="w")
            y += 1
            i = 0  # count the instances of the array
            color = "blue"
            for name in self.parent.to_addname:
                Label(self.win.body, text=name[1] + ", " + name[2], fg=color).grid(row=y, column=0, sticky="w")
                self.ai4_carrier_name.append(str(name[1] + ", " + name[2]))
                Label(self.win.body, text=macadj("not in record", "unknown"), fg=color) \
                    .grid(row=y, column=1, sticky="w")
                Label(self.win.body, text=str(self.ns_dict[name[0]]), fg=color).grid(row=y, column=2, sticky="w")
                Label(self.win.body, text=name[3], fg=color).grid(row=y, column=3, sticky="w")
                if not self.parent.is_mac:
                    Label(self.win.body, text=projvar.invran_station, fg=color).grid(row=y, column=4, sticky="w")
                y += 1
                list_options = ("otdl", "wal", "nl", "ptf", "aux")  # create optionmenu for list status
                if name[3] == "auxiliary":
                    lx = 4  # configure defaults for list status
                elif name[3] == "part time flex":
                    lx = 3  # set as ptf
                else:
                    lx = 2  # set as 'nl' if not 'aux'
                self.ai4_l_s.append(StringVar(self.win.body))
                self.ai4_l_s[i].set(list_options[lx])  # set the list status
                list_status = OptionMenu(self.win.body, self.ai4_l_s[i], *list_options)
                list_status.config(width=macadj(5, 4))
                list_status.grid(row=y, column=1, sticky="w")
                self.ai4_l_ns.append(StringVar(self.win.body))  # create optionmenu for ns days
                self.ai4_l_ns[i].set(self.full_ns_dict[str(self.ns_dict[name[0]])])  # set ns day default
                ns_day = OptionMenu(self.win.body, self.ai4_l_ns[i], *self.opt_nsday)
                ns_day.config(width=macadj(12, 10))
                ns_day.grid(row=y, column=2, sticky="w")
                self.ai4_route.append(StringVar(self.win.body))  # create entry field for route
                # create entry for routes
                Entry(self.win.body, width=24, textvariable=self.ai4_route[i]).grid(row=y, column=3, sticky="w")
                if "reg " in name[3] and name[3] != "reg floater":
                    rte = name[3].replace("reg ", "")
                else:
                    rte = ""
                self.ai4_route[i].set(rte)
                y += 1
                i += 1
                Label(self.win.body, text="").grid(row=y, column=0, sticky="w")
                y += 1
            Button(self.win.buttons, text="Continue", width=macadj(15, 16),
                   command=lambda: self.ai4_apply()).pack(side=LEFT)
            Button(self.win.buttons, text="Cancel", width=macadj(15, 16),
                   command=lambda: self.parent.go_back(self.win.topframe)).pack(side=LEFT)
            self.win.finish()

        def ai4_apply(self):
            """ adds new carriers to the carriers table """
            self.ai4_date()  # get the effective date
            self.ai4_station()  # get the station as a stringvar (apply2 reads station as stringvar)
            if self.ai4_check():
                self.ai4_count_change()
                # route conditional to arrays
                if len(self.changecount) >= len(self.ai4_carrier_name) and len(self.parent.check_these) > 0:
                    self.parent.AutoIndexer5(self.parent).run(self.win.topframe)
                elif len(self.changecount) >= len(self.ai4_carrier_name):
                    self.parent.AutoIndexer6(self.parent).run(self.win.topframe)
                else:
                    return
            else:
                frame = self.win.topframe  # prevent the object from being obliterated by rerunning __init__
                self.__init__(self.parent)  # re initialize the child class
                self.run(frame)

        def ai4_date(self):
            """ get the effective date """
            self.eff_date = projvar.invran_date_week[0]  # if investigation range is weekly
            if not projvar.invran_weekly_span:  # if investigation range is daily
                self.eff_date = projvar.invran_date

        def ai4_station(self):
            """ get the station as a stringvar (apply2 reads station as stringvar) """
            self.station = StringVar(self.win.body)  # put station var in a StringVar object
            self.station.set(projvar.invran_station)

        def ai4_check(self):
            """ check and enter carrier info """
            pb_label = Label(self.win.buttons, text="Updating Changes: ")  # make label for progress bar
            pb_label.pack(side=LEFT)
            pb = ttk.Progressbar(self.win.buttons, length=250, mode="determinate")  # create progress bar
            pb.pack(side=LEFT)
            pb["maximum"] = len(self.ai4_carrier_name)  # set length of progress bar
            pb.start()
            for i in range(len(self.ai4_carrier_name)):
                pb["value"] = i  # increment progress bar
                passed_ns = self.ai4_l_ns[i].get().split(" - ")  # clean the passed ns day data
                self.clean_ns = StringVar(self.win.body)  # put ns day var in StringVar object
                self.clean_ns.set(passed_ns[1])
                # check moves/route and enter data into rings table
                if not self.check_and_apply(i):
                    return False
                self.win.buttons.update()
            pb.stop()  # stop and destroy the progress bar
            pb_label.destroy()  # destroy the label for the progress bar
            pb.destroy()
            return True

        def check_and_apply(self, i):
            """ adds new carriers to the carriers table """
            # simplify the variables.
            date = self.eff_date  # the effective date - first day of the investigation range.
            name = self.ai4_carrier_name[i]
            list_ = self.ai4_l_s[i].get()
            nsday = self.clean_ns.get()
            route = self.ai4_route[i].get()
            station = self.station.get()
            # check the route
            if not self.check_route(route):  # return False if the checks fail
                return False
            route = Handler(route).routes_adj()  # simplify any five digit route numbers when possible
            route = Handler(route).route_zeros_to_empty()  # convert routes "0000" to empty strings
            sql = "SELECT effective_date, carrier_name,list_status, ns_day,route_s, station, rowid FROM carriers " \
                  "WHERE carrier_name = '%s' and effective_date = '%s' ORDER BY effective_date" \
                  % (name, date)
            results = inquire(sql)
            if len(results) == 0:
                sql = "INSERT INTO carriers (effective_date,carrier_name,list_status,ns_day,route_s,station)" \
                      " VALUES('%s','%s','%s','%s','%s','%s')" % \
                      (date, name, list_, nsday, route, station)
                commit(sql)
            elif len(results) == 1:
                sql = "UPDATE carriers SET list_status='%s',ns_day='%s',route_s='%s',station='%s' " \
                      "WHERE effective_date = '%s' and carrier_name = '%s'" % \
                      (list_, nsday, route, station, date, name)
                commit(sql)
            elif len(results) > 1:
                sql = "DELETE FROM carriers WHERE effective_date ='%s' and carrier_name = '%s'" % \
                      (date, name)
                commit(sql)
                sql = "INSERT INTO carriers (effective_date,carrier_name,list_status,ns_day,route_s,station)" \
                      " VALUES('%s','%s','%s','%s','%s','%s')" \
                      % (date, name, list_, nsday, route, station)
                commit(sql)
            return True

        def check_route(self, route):
            """ checks the route for auto data entry manual entries. """
            routecheck = RouteChecker(route, frame=self.win.topframe)  # create RouteChecker object
            if routecheck.is_empty():  # if there is an empty string, skip other checks...
                return True
            if not routecheck.check_all():  # check numeric, lenght and route numbers...
                return False  # return False if the checks fail.
            return True

        def ai4_count_change(self):
            """ get count of carrier changes for current day """
            self.changecount = []
            for name in self.ai4_carrier_name:
                sql = "SELECT * FROM carriers WHERE carrier_name == '%s' and effective_date == '%s'" \
                      % (name, self.eff_date)
                result = inquire(sql)
                if result:
                    self.changecount.append(result)

    class AutoIndexer5:
        """
        discrepancy resolution screen
        """

        def __init__(self, parent):
            self.parent = parent
            self.frame = None
            self.opt_nsday = []  # make an array of "day / color" options for option menu
            self.ns_opt_dict = {}  # creates a dictionary of ns colors/ options for menu
            self.full_ns_dict = {}
            self.ns_dict = {}  # create dictionary for ns day data
            self.name_dict = {}  # generate dictionary for emp id to kb_name
            self.ai5_carrier_list = None
            self.code_ns = None
            self.win = None
            self.y = 1  # count for the row
            self.i = 0  # count the instances of the array for the screen
            self.carrier_name = []  # create array for carrier names  # attributes for screen
            self.l_s = []  # create array for list status
            self.l_ns = []  # create array for ns days
            self.e_route = []  # create array for routes
            self.l_station = []  # create array for stations
            self.aux_list_tuple = ("aux", "ptf")
            self.reg_list_tuple = ("nl", "wal", "otdl")
            self.skip_this_screen = True
            self.color = "blue"  # the display color of information from tacs
            self.eff_date = None  # effective date for check and apply
            self.clean_ns = None  # temp string var for ns day

        def run(self, frame):
            """ master method for running other methods. """
            self.frame = frame
            if len(self.parent.check_these) == 0:
                self.parent.AutoIndexer6(self.parent).run(self.frame)
            else:
                self.parent.check_these.sort(key=itemgetter(1))  # sort the incoming tacs information
                self.ai5_opt_nsday()  # creates the option menu options for ns day menu
                self.ai5_ns_dict()  # create dictionary for ns day data
                self.ai5_nameindex_dict()  # generate dictionary for emp id to kb_name
                self.ai5_carrierlist()  # generate list of only names from 'in range carrier list'
                self.ai5_nscode()  # generate reverse ns code dict
                self.ai5_screen()

        def ai5_opt_nsday(self):
            """ creates the option menu options for ns day menu """
            for each in projvar.ns_code:
                ns_option = projvar.ns_code[each] + " - " + each  # make a string for each day/color
                self.ns_opt_dict[each] = ns_option
                if each == "none":
                    ns_option = "       " + " - " + each  # if the ns day is "none" - make a special string
                    self.ns_opt_dict[each] = ns_option
                self.opt_nsday.append(ns_option)

        def ai5_ns_dict(self):
            """ create dictionary for ns day data """
            results = gen_ns_dict(self.parent.file_path, self.parent.check_these)  # returns id and name
            for ids in results:  # loop to fill dictionary with ns day info
                self.ns_dict[ids[0]] = ids[1]

        def ai5_nameindex_dict(self):
            """ generate dictionary for emp id to kb_name """
            sql = "SELECT tacs_name, kb_name, emp_id FROM name_index ORDER BY kb_name"
            results = inquire(sql)
            for line in results:  # loop to fill arrays
                self.name_dict[line[2]] = line[1]

        def ai5_carrierlist(self):
            """ generate list of only names from 'in range carrier list' """
            names_list = []
            self.ai5_carrier_list = gen_carrier_list()  # generate an in range carrier list
            for name in self.ai5_carrier_list:
                names_list.append(name[1])
            remainders = []  # find carriers in 'check these' but not in 'in range carrier list' aka 'remainders'
            for name in self.parent.check_these:
                if self.name_dict[name[0]] not in names_list:
                    remainders.append(name)
            for name in remainders:  # get carriers data from carriers for remainders
                sql = "SELECT * FROM carriers WHERE carrier_name = '%s' and effective_date <= '%s'" \
                      "ORDER BY effective_date desc" % (self.name_dict[name[0]], projvar.invran_date_week[0])
                result = inquire(sql)
                # self.ai5_carrier_list.append(list(result[0]))
                if result:
                    self.ai5_carrier_list.append(result[0])
            self.ai5_carrier_list.sort(key=itemgetter(1))  # resort carrier list after additions

        def ai5_nscode(self):
            """ generate reverse ns code dict """
            self.code_ns = NsDayDict(projvar.invran_date_week[0]).gen_rev_ns_dict()

        def ai5_screen(self):
            """ master method for creating the screen. """
            self.win = MakeWindow()
            self.win.create(self.frame)
            self.ai5_screen_header()
            self.ai5_screen_labels()
            self.ai5_find_discrepancies()
            self.ai5_screen_buttons()

        def ai5_screen_header(self):
            """ creates the widgets which fill the top part of the screen. """
            header = Frame(self.win.body)
            header.grid(row=0, columnspan=6, sticky="w")
            Label(header, text="Discrepancy Resolution Screen", font=macadj("bold", "Helvetica 18"), pady=10) \
                .grid(row=0, sticky="w")
            Label(header, text="Correct "
                               "any discrepancies and inconsistencies that exist "
                               "between the incoming TACS data (in blue) \n"
                               "and the information currently recorded in the Klusterbox "
                               "database (below in the entry fields and \n"
                               "option menus)to reflect the carrier's status accurately. "
                               "This will update the Klusterbox database. \n"
                               "Routes must 4  or 5 digits long. In cases where there "
                               "are multiple routes, the routes must be \n"
                               "separated by a \"/\" backslash.\n\n"
                               "Investigation Range: {0} through {1}\n\n"
                  .format(projvar.invran_date_week[0].strftime("%a - %b %d, %Y"),
                          projvar.invran_date_week[6].strftime("%a - %b %d, %Y")), justify=LEFT) \
                .grid(row=1, sticky="w")

        def ai5_screen_labels(self):
            """ creates column header labels. """
            if not self.parent.is_mac:  # skip labels if the os is mac
                Label(self.win.body, text="    ", fg="Grey").grid(row=self.y, column=0, sticky="w")
                Label(self.win.body, text=macadj("List Status", "List"), fg="Grey") \
                    .grid(row=self.y, column=1, sticky="w")
                Label(self.win.body, text="NS Day", fg="Grey").grid(row=self.y, column=2, sticky="w")
                Label(self.win.body, text="Route_s", fg="Grey").grid(row=self.y, column=3, sticky="w")
                Label(self.win.body, text="Station", fg="Grey").grid(row=self.y, column=4, sticky="w")
                Label(self.win.body, text=macadj("             ", ""), fg="Grey") \
                    .grid(row=self.y, column=5, sticky="w")
                self.y += 1

        def ai5_find_discrepancies(self):
            """ look for any discrepancies in carrier list """
            tlist = ()
            tnsday = "none"
            troute = ""
            for name in self.parent.check_these:
                for k_name in self.ai5_carrier_list:
                    if self.name_dict[name[0]] == k_name[1]:  # if the names match
                        if name[3] == "auxiliary":  # parse assignments from tacs list
                            tlist = self.aux_list_tuple
                            tnsday = "none"
                            troute = ""
                        if name[3] == "part time flex":  # parse assignments from tacs list
                            tlist = self.aux_list_tuple
                            tnsday = "none"
                            troute = ""
                        if name[3][-4:].isnumeric():
                            tlist = self.reg_list_tuple
                            tnsday = self.code_ns[str(self.ns_dict[name[0]])]
                            troute = name[3][-4:]
                        if name[3][-7:] == "floater":
                            tlist = self.reg_list_tuple
                            tnsday = self.code_ns[str(self.ns_dict[name[0]])]
                            troute = "floater"
                        if name[3] == "undetected":
                            tlist = "undetected"
                            tnsday = self.code_ns[str(self.ns_dict[name[0]])]
                            troute = "undetected"
                        discrepancy = False
                        # check tacs data against data in carriers table/ klusterbox
                        if k_name[2] not in tlist:  # check list status
                            discrepancy = True
                        if k_name[3] != tnsday:  # check nsday
                            discrepancy = True
                        k_rte_len = len(k_name[4].split('/'))  # check route
                        if k_rte_len == 0:  # check if route is aux
                            if troute != "":
                                discrepancy = True
                        if k_rte_len == 1:  # check if route is regular
                            if troute != k_name[4]:
                                discrepancy = True
                        if k_rte_len == 5:  # check if route is floater
                            if troute != "floater":
                                discrepancy = True
                        if projvar.invran_station != k_name[5]:  # check if station is correct
                            discrepancy = True
                        if discrepancy:  # if there are no discrepancies, then skip the screen
                            self.skip_this_screen = False
                            self.ai5_display_discrepancies(name, k_name)

        def ai5_display_discrepancies(self, name, k_name):
            """ displays the main body of the screen. """
            name_f = Frame(self.win.body)  # create separate frame for names
            name_f.grid(row=self.y, columnspan=6, sticky="w")
            Label(name_f, text="Name: ", fg="Grey").grid(row=0, column=0, sticky="w")
            Label(name_f, text=name[1] + ", " + name[2], fg=self.color).grid(row=0, column=1, sticky="w")
            Label(name_f, text=" / " + k_name[1]).grid(row=0, column=2, sticky="w")
            self.y += 1
            if not self.parent.is_mac:
                Label(self.win.body, text="    ", fg=self.color).grid(row=self.y, column=0, sticky="w")
            Label(self.win.body, text=macadj("not in record", "unknown"), fg=self.color) \
                .grid(row=self.y, column=1, sticky="w")
            Label(self.win.body, text=str(self.ns_dict[name[0]]), fg=self.color).grid(row=self.y, column=2, sticky="w")
            Label(self.win.body, text=name[3], fg=self.color).grid(row=self.y, column=3, sticky="w")
            Label(self.win.body, text=projvar.invran_station, fg=self.color).grid(row=self.y, column=4, sticky="w")
            self.y += 1
            self.carrier_name.append(k_name[1])  # add kb name to the array
            list_options = ("otdl", "wal", "nl", "ptf", "aux")  # create optionmenu for list status
            self.l_s.append(StringVar(self.win.body))
            self.l_s[self.i].set(k_name[2])  # set the list status
            list_status = OptionMenu(self.win.body, self.l_s[self.i], *list_options)
            list_status.config(width=macadj(6, 4))
            list_status.grid(row=self.y, column=1, sticky="w")
            self.l_ns.append(StringVar(self.win.body))  # create optionmenu for ns days
            self.l_ns[self.i].set(self.ns_opt_dict[k_name[3]])  # set ns day default
            ns_day = OptionMenu(self.win.body, self.l_ns[self.i], *self.opt_nsday)
            ns_day.config(width=macadj(12, 8))
            ns_day.grid(row=self.y, column=2, sticky="w")
            self.e_route.append(StringVar(self.win.body))  # create entry field for route
            Entry(self.win.body, width=25, textvariable=self.e_route[self.i]) \
                .grid(row=self.y, column=3, sticky="w")  # create entry for routes
            self.e_route[self.i].set(k_name[4])
            self.l_station.append(StringVar(self.win.body))
            self.l_station[self.i].set(k_name[5])
            list_station = OptionMenu(self.win.body, self.l_station[self.i], *projvar.list_of_stations)
            list_station.config(width=macadj(25, 18))
            list_station.grid(row=self.y, column=4, sticky="w")
            self.y += 1
            Label(self.win.body, text="").grid(row=self.y, column=1)
            self.y += 1
            self.i += 1

        def ai5_screen_buttons(self):
            """ displays the buttons on the bottom of the page. """
            Button(self.win.buttons, text="Continue", width=macadj(15, 16),
                   command=lambda: self.ai5_apply()).pack(side=LEFT)
            Button(self.win.buttons, text="Cancel", width=macadj(15, 16),
                   command=lambda: self.parent.go_back(self.win.topframe)).pack(side=LEFT)
            if self.skip_this_screen:
                self.parent.AutoIndexer6(self.parent).run(self.win.topframe)
            else:
                self.win.finish()  # get rear window objects

        def ai5_apply(self):
            """ generate progressbar - sends data to be checked """
            self.eff_date = projvar.invran_date_week[0]  # if investigation range is weekly
            if not projvar.invran_weekly_span:  # if investigation range is daily
                self.eff_date = projvar.invran_date
            pb_label = Label(self.win.buttons, text="Updating Changes: ")  # make label for progress bar
            pb_label.pack(side=LEFT)
            pb = ttk.Progressbar(self.win.buttons, length=250, mode="determinate")  # create progress bar
            pb.pack(side=LEFT)
            pb["maximum"] = len(self.carrier_name)  # set length of progress bar
            pb.start()
            for i in range(len(self.carrier_name)):
                pb["value"] = i  # increment progress bar
                passed_ns = self.l_ns[i].get().split(" - ")  # clean the passed ns day data
                self.clean_ns = StringVar(self.win.topframe)  # put ns day var in StringVar object
                self.clean_ns.set(passed_ns[1])
                if not self.check_and_apply(i):
                    frame = self.win.topframe  # prevent the object from being obliterated by rerunning __init__
                    self.__init__(self.parent)  # re initialize the child class
                    self.run(frame)
                    return
                self.win.buttons.update()
            pb.stop()  # stop and destroy the progress bar
            pb_label.destroy()  # destroy the label for the progress bar
            pb.destroy()
            self.parent.AutoIndexer6(self.parent).run(self.win.topframe)

        def check_and_apply(self, i):
            """ adds new carriers to the carriers table """
            # simplify the variables.
            date = self.eff_date  # the effective date - first day of the investigation range.
            name = self.carrier_name[i]
            list_ = self.l_s[i].get()
            nsday = self.clean_ns.get()
            route = self.e_route[i].get()
            station = self.l_station[i].get()
            # check the route
            if not self.check_route(route):
                return False
            route = Handler(route).routes_adj()
            route = Handler(route).route_zeros_to_empty()
            sql = "SELECT effective_date, carrier_name,list_status, ns_day,route_s, station, rowid FROM carriers " \
                  "WHERE carrier_name = '%s' and effective_date = '%s' ORDER BY effective_date" \
                  % (name, date)
            results = inquire(sql)
            if len(results) == 0:
                sql = "INSERT INTO carriers (effective_date,carrier_name,list_status,ns_day,route_s,station)" \
                      " VALUES('%s','%s','%s','%s','%s','%s')" % \
                      (date, name, list_, nsday, route, station)
                commit(sql)
            elif len(results) == 1:
                sql = "UPDATE carriers SET list_status='%s',ns_day='%s',route_s='%s',station='%s' " \
                      "WHERE effective_date = '%s' and carrier_name = '%s'" % \
                      (list_, nsday, route, station, date, name)
                commit(sql)
            elif len(results) > 1:
                sql = "DELETE FROM carriers WHERE effective_date ='%s' and carrier_name = '%s'" % \
                      (date, name)
                commit(sql)
                sql = "INSERT INTO carriers (effective_date,carrier_name,list_status,ns_day,route_s,station)" \
                      " VALUES('%s','%s','%s','%s','%s','%s')" \
                      % (date, name, list_, nsday, route, station)
                commit(sql)
            return True

        def check_route(self, route):
            """ checks the route for auto data entry manual entries. """
            routecheck = RouteChecker(route, frame=self.win.topframe)  # create RouteChecker object
            if routecheck.is_empty():  # if there is an empty string, skip other checks...
                return True
            if not routecheck.check_all():  # check numeric, lenght and route numbers...
                return False  # return False if the checks fail.
            return True

    class AutoIndexer6:
        """
        detect carriers who are no longer in station
        """

        def __init__(self, parent):
            self.parent = parent
            self.frame = None
            self.names_list = []  # list of carriers in investigation range.
            self.filtered_ids = []  # filter the tacs ids to only good jobs
            self.t_names = []  # matches emp id to the kb name
            self.ex_carrier = []  # carriers in carrier list but not tacs data
            self.win = None
            self.y = 1  # count for the row
            self.carrier_name = []
            self.list_status = []
            self.ns_day = []
            self.route = []
            self.station = []
            self.new_station = []
            self.cc = 0

        def run(self, frame):
            """ master method for running the class. """
            self.frame = frame
            self.ai6_nameslist()  # create the names list array
            self.ai6_filtered_ids()
            self.ai6_t_names()
            self.ai6_ex_carriers()
            if len(self.ex_carrier) == 0:
                self.parent.AutoSkimmer(self.parent).run(self.frame)
            else:
                self.ai6_screen()  # create the 'carriers no longer in station' screen
                self.ai6_screen_header()
                self.ai6_screen_labels()
                self.ai6_screen_loop()
                self.ai6_screen_buttons()
                self.win.finish()

        def ai6_nameslist(self):
            """ list who are not in the TACS list"""
            carrier_list = gen_carrier_list()  # create names_list array
            for name in carrier_list:  # eliminate duplicate names
                if name[1] not in self.names_list:
                    self.names_list.append(name[1])

        def ai6_filtered_ids(self):
            """ filter the tacs ids to get the good good_jobs """
            self.parent.get_file()  # read the csv file
            tacs_ids = []  # generate tacs list
            good_jobs = ("844", "134", "434")
            to_add = ("x", "x")  # create placeholder for
            for line in self.parent.a_file:
                if len(line) > 19:  # if there are enough items in the line
                    if line[18] == "Temp":
                        to_add = (line[4].zfill(8), line[19])
                    elif line[19] != "Temp" or line[19] != "Base":
                        if to_add != ("x", "x"):  # if not placeholder
                            tacs_ids.append(to_add)  # add tacs data to the array
                            to_add = ("x", "x")  # reset placeholder
                    if line[18] == "Base":
                        to_add = (line[4].zfill(8), line[19])
            self.filtered_ids = []  # filter the tacs ids to only good jobs
            for item in tacs_ids:
                if item[1] in good_jobs:
                    self.filtered_ids.append(item)
            del tacs_ids

        def ai6_t_names(self):
            """ get carrier names from the employee ids. """
            for name in self.filtered_ids:  #
                sql = "SELECT kb_name FROM name_index WHERE emp_id = '%s'" % (name[0])
                result = inquire(sql)  # check dbase for a match
                if result:  # if there is a match in the dbase, then add data to array
                    self.t_names.append(result[0][0])

        def ai6_ex_carriers(self):
            """ get a list of carriers no longer in the station """
            for name in self.names_list:  # for each name in carrier list
                if name not in self.t_names:  # if they are not also in the tacs data
                    self.ex_carrier.append(name)  # then add them to the array

        def ai6_screen(self):
            """ create the window object. """
            self.win = MakeWindow()
            self.win.create(self.frame)

        def ai6_screen_header(self):
            """ creates the header labels for the top of the screen. """
            header = Frame(self.win.body)
            header.grid(row=0, columnspan=5, sticky="w")
            Label(header, text="Carriers No Longer At Station", font=macadj("bold", "Helvetica 18"), pady=10) \
                .grid(row=0, sticky="w")
            wintext = "Klusterbox has detected that the following carriers may no longer be at the station. " \
                      "If they are no longer at the\n station, then please use the option menu below to move " \
                      "them to the correct station (if listed). If the correct \nis not listed or the carrier " \
                      "is no longer working for the post office, then select \"out of station\".\n\n"
            mactext = \
                "Klusterbox has detected that the following carriers may no longer be at the station. If they \n" \
                "are no longer at the station, then please use the option menu below to move them to the \n" \
                "correct station (if listed). If the correct is not listed or the carrier is no longer working \n" \
                "for the post office, then select \"out of station\".\n\n"
            text = "Investigation Range: {0} through {1}\n\n".format(
                projvar.invran_date_week[0].strftime("%a - %b %d, %Y"),
                projvar.invran_date_week[6].strftime("%a - %b %d, %Y"))
            Label(header, text=macadj(wintext, mactext) + text, justify=LEFT).grid(row=1, sticky="w")

        def ai6_screen_labels(self):
            """ creates the labels to be used as column headers. """
            Label(self.win.body, text="Name", fg="Grey").grid(row=self.y, column=0, sticky="w")
            Label(self.win.body, text=macadj("List Status", "List"), fg="Grey").grid(row=self.y, column=1, sticky="w")
            if sys.platform != "darwin":  # if the platform is not darwin
                Label(self.win.body, text="Route_s", fg="Grey").grid(row=self.y, column=2, sticky="w")
            Label(self.win.body, text="Station", fg="Grey").grid(row=self.y, column=3, sticky="w")
            Label(self.win.body, text="             ", fg="Grey").grid(row=self.y, column=4, sticky="w")
            self.y += 1

        def ai6_screen_loop(self):
            """ generate widges for all carriers no longer at station. """
            for name in self.ex_carrier:
                sql = "SELECT * FROM carriers WHERE carrier_name = '%s' and effective_date <= '%s' " \
                      "ORDER BY effective_date DESC" \
                      % (name, projvar.invran_date_week[0])
                result = inquire(sql)
                self.carrier_name.append(StringVar(self.win.body))  # store name
                self.carrier_name[self.cc].set(result[0][1])
                Button(self.win.body, text=result[0][1], relief=RIDGE, width=25, anchor="w") \
                    .grid(row=self.y, column=0, sticky="w")  # name
                self.list_status.append(StringVar(self.win.body))  # store list status
                self.list_status[self.cc].set(result[0][2])
                Button(self.win.body, text=result[0][2], relief=RIDGE, width=7, anchor="w") \
                    .grid(row=self.y, column=1, sticky="w")  # list
                self.ns_day.append(StringVar(self.win.body))  # store ns day
                self.ns_day[self.cc].set(result[0][3])
                self.route.append(StringVar(self.win.body))  # store route
                self.route[self.cc].set(result[0][4])
                if sys.platform != "darwin":  # if the platform is not darwin
                    Button(self.win.body, text=result[0][4], relief=RIDGE, width=20, anchor="w") \
                        .grid(row=self.y, column=2, sticky="w")  # route
                self.station.append(StringVar(self.win.body))  # store station
                self.station[self.cc].set(result[0][5])
                self.new_station.append(StringVar(self.win.body))
                self.new_station[self.cc].set("out of station")
                stat_om = OptionMenu(self.win.body, self.new_station[self.cc], *projvar.list_of_stations)  # station
                if sys.platform != "darwin":  # if the platform is not darwin
                    stat_om.config(width=25, anchor="w")
                else:
                    stat_om.config(width=25)
                stat_om.grid(row=self.y, column=3, sticky="w")
                Label(self.win.body, text="                     ").grid(row=self.y, column=4)
                self.cc += 1
                self.y += 1

        def ai6_screen_buttons(self):
            """ creates buttons on the bottom of the page. """
            Button(self.win.buttons, text="Continue", width=macadj(15, 16),
                   command=lambda: self.ai6_apply()).pack(side=LEFT)
            Button(self.win.buttons, text="Cancel", width=macadj(15, 16),
                   command=lambda: self.parent.go_back(self.win.topframe)).pack(side=LEFT)

        def ai6_apply(self):
            """ executes when apply is pressed. """
            pb_label = Label(self.win.buttons, text="Updating Changes: ")  # make label for progress bar
            pb_label.pack(side=LEFT)
            pb = ttk.Progressbar(self.win.buttons, length=250, mode="determinate")  # create progress bar
            pb.pack(side=LEFT)
            pb["maximum"] = len(self.carrier_name)  # set length of progress bar
            pb.start()
            for i in range(len(self.carrier_name)):
                pb["value"] = i  # increment progress bar
                if self.station[i].get() != self.new_station[i].get():  # if there is a change of station
                    self.check_and_apply(i)
                self.win.buttons.update()
            pb.stop()  # stop and destroy the progress bar
            pb_label.destroy()  # destroy the label for the progress bar
            pb.destroy()
            self.parent.AutoSkimmer(self.parent).run(self.win.topframe)

        def check_and_apply(self, i):
            """ adds new carriers to the carriers table """
            # simplify the variables.
            date = projvar.invran_date_week[0]  # the effective date - first day of the investigation range.
            name = self.carrier_name[i].get()
            list_ = self.list_status[i].get()
            nsday = self.ns_day[i].get()
            route = self.route[i].get()
            station = self.new_station[i].get()
            sql = "SELECT effective_date, carrier_name,list_status, ns_day,route_s, station, rowid FROM carriers " \
                  "WHERE carrier_name = '%s' and effective_date = '%s' ORDER BY effective_date" \
                  % (name, date)
            results = inquire(sql)
            if len(results) == 0:
                sql = "INSERT INTO carriers (effective_date,carrier_name,list_status,ns_day,route_s,station)" \
                      " VALUES('%s','%s','%s','%s','%s','%s')" % \
                      (date, name, list_, nsday, route, station)
                commit(sql)
            elif len(results) == 1:
                sql = "UPDATE carriers SET list_status='%s',ns_day='%s',route_s='%s',station='%s' " \
                      "WHERE effective_date = '%s' and carrier_name = '%s'" % \
                      (list_, nsday, route, station, date, name)
                commit(sql)
            elif len(results) > 1:
                sql = "DELETE FROM carriers WHERE effective_date ='%s' and carrier_name = '%s'" % \
                      (date, name)
                commit(sql)
                sql = "INSERT INTO carriers (effective_date,carrier_name,list_status,ns_day,route_s,station)" \
                      " VALUES('%s','%s','%s','%s','%s','%s')" \
                      % (date, name, list_, nsday, route, station)
                commit(sql)
            return True

    class AutoSkimmer:
        """
        This class enters in the clock rings by reading the employee everything report csv. While the above
        classes focused on the Base and Temp lines, this class focus on the lines dealing with hours worked,
        paid leave, unpaid leave, begin tour, moves an end tour.
        """

        def __init__(self, parent):
            self.parent = parent
            self.frame = None
            self.skippers = None
            self.days = ("Saturday", "Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday")
            self.mv_codes = ("BT", "MV", "ET")
            self.day_dict = {}  # make a dictionary for each day in the week
            self.carrier_lines = []
            self.weekly_protoarray = []  # an array of daily_protoarrays
            self.daily_protoarray = []  # returned from daily analysis
            self.newest_carrier = []
            self.kb_name = ""  # carrier name from nameindex table
            self.routes = []  # get route/s
            self.c_code = "none"  # notes ns day hit
            self.mv_triad = []  # triad is route#, start time off route, end time off route
            self.mv_str = ""  # moves
            self.hr_52 = ""  # paid leave
            self.rs = ""  # return to station time
            self.lv_type = ""  # 5200 leave type
            self.lv_time = ""  # 5200 leave time
            self.bt = ""  # begin tour
            self.et = ""  # end tour
            self.current_array = []  # product of skim weekly - formatted data to dbase input
            # variables for build_protoarray()
            self.daily_rings = []
            self.daily_line = []
            self.day_name = ""
            self.day_hr_52 = 0.0  # work hours
            self.day_hr_55 = 0.0  # annual leave
            self.day_hr_56 = 0.0  # sick leave
            self.day_hr_58 = 0.0  # holiday leave
            self.day_hr_62 = 0.0  # guaranteed time
            self.day_hr_86 = 0.0  # other paid leave
            self.day_rs = 0  # the return to station time for the proto array
            self.day_code = ""
            self.day_moves = []
            self.day_leave_type = []
            self.day_leave_time = []
            self.day_final_leave_type = ""
            self.day_final_leave_time = 0.0
            self.day_bt = 0  # the begin tour time for the proto array
            self.day_et = 0  # the end tour time for the proto array
            self.day_dayofweek = None
            # variables for fix carrier lines
            self.new_order = []

        def run(self, frame):
            """ master method for running other methods. """
            self.frame = frame
            self.skim_configs()  # get configuration settings
            self.carrier_list_cleaning()  # cleans the database of duplicate records
            self.skim_day_dict()  # make a dictionary for each day in the week
            if not self.skim_check_csv():  # checks for employee everything report
                self.parent.go_back(self.frame)  # quit and return to main screen
            else:
                if not messagebox.askokcancel("Automatic Rings Entry",
                                              "Do you want to automatically enter the rings?",
                                              parent=self.frame):
                    self.parent.go_back(self.frame)  # quit and return to main screen
                else:
                    self.skim_enter_rings()
                    messagebox.showinfo("Automatic Rings Entry",
                                        "The Employee Everything Report has been sucessfully inputed into the database",
                                        parent=self.frame)
                    self.parent.go_back(self.frame)  # quit and return to main screen

        def skim_configs(self):
            """ get configuration settings """
            sql = "SELECT code FROM skippers"  # get skippers data from dbase
            results = inquire(sql)
            self.skippers = []  # fill the array for skippers
            for item in results:
                self.skippers.append(item[0])

        def skim_day_dict(self):
            """ make a dictionary for each day in the week """
            x = 0
            for item in self.days:
                self.day_dict[item] = projvar.invran_date_week[x]
                x += 1

        def skim_check_csv(self):
            """ checks for employee everything report """
            self.parent.get_file()  # read the csv file
            for line in self.parent.a_file:
                if line[0][:8] == "TAC500R3":
                    return True
                else:
                    messagebox.showwarning("File Selection Error",
                                           "The selected file does not appear to be an "
                                           "Employee Everything report.", parent=self.frame)
                    return False

        def skim_enter_rings(self):
            """
            Takes the entire csv file, goes line by line and breaks it up into one chunk per carrier
            and sends it to the skim weekly for further breakdown by day
            """
            self.parent.get_file()  # read the csv file
            # row_count = sum(1 for row in self.parent.a_file)  # get number of rows in csv file
            row_count = sum(1 for _ in self.parent.a_file)  # get number of rows in csv file
            self.parent.get_file()  # read the csv file
            pb = ProgressBarDe(title="Entering Carrier Rings", label="Updating Rings: ")
            pb.max_count(int(row_count))
            pb.start_up()
            i = 0
            cc = 0
            good_id = "no"
            for line in self.parent.a_file:
                pb.move_count(i)
                if cc != 0:
                    if good_id != line[4] and good_id != "no":  # if new carrier_lines or employee
                        self.skim_weekly()  # trigger analysis
                        del self.carrier_lines[:]  # empty array
                        good_id = "no"  # reset trigger
                    # find first line of specific carrier_lines
                    if line[18] == "Base" and line[19] in ("844", "134", "434"):
                        good_id = line[4]  # set trigger to id of carriers who are FT or aux carriers
                        self.carrier_lines.append(line)  # gather times and moves for anaylsis
                        pb.change_text("Entering rings for {}".format(line[5]))
                    if good_id == line[4] and line[18] != "Base":
                        if line[18] in self.days:  # get the hours for each day
                            self.carrier_lines.append(line)  # gather times and moves for anaylsis
                        if line[19] in self.mv_codes and line[32] != "(W)Ring Deleted From PC":
                            self.carrier_lines.append(line)  # gather times and moves for anaylsis
                        pb.change_text("Entering rings for {}".format(line[5]))
                cc += 1
                i += 1
            self.skim_weekly()  # when loop ends, run final analysis
            del self.carrier_lines[:]  # empty array
            pb.stop()  # stop and destroy the progress bar

        def skim_weekly(self):
            """
            Takes the carrier lines sent by enter rings method and sends it to input rings to convert
            it into an array of proto arrays - one for each day collected in input rings
            """
            self.fix_carrierlines()
            del self.weekly_protoarray[:]  # delete prior input rings
            self.skim_input_rings()  # build an array of protoarrays
            if self.weekly_protoarray[0] is not None:
                result = self.skim_check_nameindex()  # get the carriers employee id number
                if result:  # if there is an employee id number in the name index, then continue
                    if self.skim_check_carriers(result):  # get the kb name which correlates to the emp id
                        self.skim_get_routes()  # create an array of the carrier's routes for self.routes
                        for i in range(len(self.weekly_protoarray)):  # loop for each day of carrier information
                            self.daily_protoarray = self.weekly_protoarray[i]
                            """ should be dealing with input rings and not protoarray as input rings is a storage 
                            array for the daily protoarrays"""
                            self.skim_detect_nsday()  # find if the day is an ns day
                            self.skim_detect_moves()  # find the moves if any
                            self.skim_get_movestring()
                            if self.skim_get_hour52():
                                self.skim_returntostation()
                                self.skim_get_leavetime()
                                self.skim_begintour()
                                self.skim_endtour()
                                self.skim_current_array()
                                self.skim_input_update()

        def fix_carrierlines(self):
            """
            This method solves a problem of lines in the employee everything report being slightly out of
            sequence such as a move coming before a begin tour or a move coming after an end tour. This method
            rearranges the begin tour, moves and end tour lines to make sure begin tour rings come first, moves
            are in the middle and end tours are last. Lines which are not BT, Moves or ET keep their original
            positions.
            """
            self.new_order = []  # carrier lines restructured
            moves_holder = []
            for line in self.carrier_lines:
                if line[19] in self.mv_codes:  # mv_codes is ("BT", "MV", "ET")
                    moves_holder.append(line)  # captures the BT, MV or ET lines
                else:
                    if moves_holder:  # if there are BT, MV or ET lines in the move holder
                        self.fix_carrierline_moves(moves_holder)  # call a method to put them in proper order
                        del moves_holder[:]  # delete the contents of the array
                    self.new_order.append(line)  # non-BT, MV or ET lines go straight to the new order array.
            if moves_holder:  # at the end of the loop check if there are BT, MV or ET lines in the move holder
                self.fix_carrierline_moves(moves_holder)  # call a method to put them in proper order
            self.carrier_lines = self.new_order[:]  # carrier lines is over written with correctly order array

        def fix_carrierline_moves(self, moves_holder):
            """ puts the BT, MV and ET lines in proper order """
            bt_array = []  # holds begin tour lines
            mv_array = []  # hold moves lines
            et_array = []  # holds end tour lines
            for move in moves_holder:  # loop through the BT, MV or ET lines
                if move[19] == "BT":
                    if not bt_array:  # only proceed if the bt array is empty - capture earliest valve
                        bt_array.append(move)  # capture begin tours in an array
                if move[19] == "MV":
                    mv_array.append(move)  # capture moves in an array
                if move[19] == "ET":
                    et_array.append(move)  # captures end tours in an array
            for move in (bt_array + mv_array + et_array):  # with the lines in the proper order...
                self.new_order.append(move)  # put the BT, MV or ET lines into the new order array

        def skim_input_rings(self):
            """
            Takes The carrier lines from enter rings method and creates a daily protoarrays for the
            investigation range then compiles them in a weekly protoarray.
            """
            rings = []
            good_day = "no"
            for line in self.carrier_lines:
                if line[18] in self.days and line[18] != good_day and good_day != "no":
                    to_input = self.build_protoarray(rings)  # returns the protoarray for one day
                    self.weekly_protoarray.append(to_input)
                    del rings[:]
                    good_day = line[18]
                if line[18] == "Base" and line[19] in ("844", "134", "434"):  # find first line of specific carrier
                    continue  # gather base line data
                elif line[18] == "Temp" and line[19] in ("844", "134", "434"):  # find first line of specific carrier
                    continue  # gather base line data
                else:
                    if line[18] in self.days and line[18] == good_day:
                        rings.append(line)
                    if line[18] in self.days and good_day == "no":  # day change triggers
                        good_day = line[18]
                        rings.append(line)
                    if line[18] not in self.days:
                        rings.append(line)
            to_input = self.build_protoarray(rings)  # call function for last line  # returns the protoarray for one day
            self.weekly_protoarray.append(to_input)  # add the daily proto array to the weekly proto array

        def skim_check_nameindex(self):
            """ check if the carrier is in the name index """
            sql = "SELECT kb_name FROM name_index WHERE emp_id = '%s'" % self.weekly_protoarray[0][1]
            result = inquire(sql)  # check to verify that they are in the name index
            return result  # if there is a match in the name index, then continue

        def skim_check_carriers(self, result):
            """ get the most recent record for the carrier. """
            self.kb_name = result[0][0]  # get the kb name which correlates to the emp id
            for line in self.weekly_protoarray:
                self.daily_protoarray = line
                sql = "SELECT effective_date, carrier_name, list_status, ns_day, route_s FROM" \
                      " carriers WHERE carrier_name = '%s' and effective_date <= '%s' " \
                      "ORDER BY effective_date DESC" % (self.kb_name, self.day_dict[self.daily_protoarray[0]])
                result = inquire(sql)
                for array in result:  # find the most recent carrier record
                    eff_date = datetime.strptime(array[0], '%Y-%m-%d %H:%M:%S')
                    if eff_date <= self.day_dict[self.daily_protoarray[0]]:
                        self.newest_carrier = array
                        break  # stop. we only need the most recent record
                if result:
                    return True
                return False

        def skim_detect_nsday(self):
            """ find the code, if any  / as of version 4.003 otdl carriers are allowed ns day code """
            if self.newest_carrier[2] in ("nl", "wal"):  # if the current day matches the ns day
                if self.day_dict[self.daily_protoarray[0]].strftime("%a") == \
                        projvar.ns_code[self.newest_carrier[3]] and \
                        float(self.daily_protoarray[2]) > 0:
                    self.c_code = "ns day"  # enter the code
                else:
                    self.c_code = "none"  # enter the code
            elif self.newest_carrier[2] == "otdl":  # if the current day matches the ns day
                if self.day_dict[self.daily_protoarray[0]].strftime("%a") == \
                        projvar.ns_code[self.newest_carrier[3]] and \
                        float(self.daily_protoarray[2]) > 0:
                    self.c_code = "ns day"  # enter the code
                else:
                    if self.daily_protoarray[4] == "":
                        self.c_code = "none"  # self.daily_protoarray[4] is the code from proto-array
                    else:
                        self.c_code = self.daily_protoarray[4]  # can be sick or annual
                pass
            elif self.newest_carrier[2] in ("ptf", "aux"):
                if self.daily_protoarray[4] == "":
                    self.c_code = "none"  # self.daily_protoarray[4] is the code from proto-array
                else:
                    self.c_code = self.daily_protoarray[4]  # can be sick or annual
            else:
                self.c_code = "none"

        def skim_get_routes(self):
            """ create an array for self.routes """
            self.routes = []
            if self.newest_carrier[4] != "":
                self.routes = self.newest_carrier[4].split("/")

        def skim_detect_moves(self):
            """ find the moves if any """
            self.mv_triad = []  # triad is route number, start time off route, end time off route
            route_holder = ""
            if len(self.routes) > 0:  # if the route is in kb
                pair = "closed"  # trigger opens when a move set needs to be closed
                for m in self.daily_protoarray[5]:  # loop through all the rings
                    mv_time = Convert(m[1]).zero_or_hundredths()  # assign move time variable and format
                    if m[3] not in self.routes and pair == "closed":
                        if m[3] == "0000" and m[2] in self.skippers:  # sometimes off route is not off route
                            continue
                        else:
                            route_holder = m[3]  # hold route to put at end of triad
                            self.mv_triad.append(mv_time)  # add start time to second place of triad
                            pair = "open"
                    if m[3] in self.routes and pair == "open":
                        self.mv_triad.append(mv_time)  # add end time to third place of triad
                        self.mv_triad.append(route_holder)
                        pair = "closed"
                if pair == "open":  # if open at end, then close it with the last ring
                    # assign move time variable and format for the last move if pair == 'open'
                    mv_time = \
                        Convert(self.daily_protoarray[5][len(self.daily_protoarray[5]) - 1][1]).zero_or_hundredths()
                    self.mv_triad.append(mv_time)
                    self.mv_triad.append(route_holder)

        def skim_get_movestring(self):
            """ format array as string to fit in dbase """
            self.mv_str = ','.join(self.mv_triad)

        def skim_get_hour52(self):
            """ get paid leave """
            # if hours worked > 0 or there is a code or a leave type
            if float(self.daily_protoarray[2]) > 0 or self.c_code != "none" or self.daily_protoarray[6] != "":
                hr_52 = self.daily_protoarray[2]  # assign 5200 hours variable
                if RingTimeChecker(hr_52).check_for_zeros():  # adjust hr_52to version 4 record standards
                    self.hr_52 = ""
                else:
                    self.hr_52 = Convert(hr_52).hundredths()
                return True
            return False

        def skim_returntostation(self):
            """ assign return to station variable """
            rs = self.daily_protoarray[3]
            if RingTimeChecker(rs).check_for_zeros():  # adjust rs to version 4 record standards
                self.rs = ""
            else:
                self.rs = Convert(rs).hundredths()

        def skim_begintour(self):
            """ assign the begin tour variable. """
            bt = self.daily_protoarray[8]
            if RingTimeChecker(bt).check_for_zeros():  # adjust bt to version 4 record standards
                self.bt = ""
            else:
                self.bt = Convert(bt).hundredths()

        def skim_endtour(self):
            """ assign the end tour variable. """
            et = self.daily_protoarray[9]
            if RingTimeChecker(et).check_for_zeros():  # adjust et to version 4 record standards
                self.et = ""
            else:
                self.et = Convert(et).hundredths()

        def skim_get_leavetime(self):
            """ check and handle leave time. """
            lv_time = float(self.daily_protoarray[7])  # assign leave time variable
            self.lv_type = Convert(self.daily_protoarray[6]).none_not_empty()  # adjust lv type to version 4 standards
            if RingTimeChecker(lv_time).check_for_zeros():  # adjust lv time to version 4 record standards
                self.lv_time = ""
            else:
                self.lv_time = Convert(lv_time).hundredths()

        def skim_current_array(self):
            """ build the current array """
            self.current_array = [str(self.day_dict[self.daily_protoarray[0]]), self.kb_name, self.hr_52, self.rs,
                                  self.c_code, self.mv_str, self.lv_type, self.lv_time, self.bt, self.et]

        def skim_input_update(self):
            """ check rings table to see if record already exist."""
            sql = "SELECT * FROM rings3 WHERE carrier_name = '%s' and rings_date = '%s'" % (
                self.kb_name, self.day_dict[self.daily_protoarray[0]])
            result = inquire(sql)
            if len(result) == 0:
                sql = "INSERT INTO rings3 (rings_date, carrier_name, total, " \
                      "rs, code, moves, leave_type, leave_time, bt, et) " \
                      "VALUES('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s')" % \
                      (self.current_array[0], self.current_array[1], self.current_array[2], self.current_array[3],
                       self.current_array[4], self.current_array[5], self.current_array[6], self.current_array[7],
                       self.current_array[8], self.current_array[9])
                commit(sql)
            else:
                sql = "UPDATE rings3 SET total='%s', rs='%s', code='%s', moves='%s'," \
                      "leave_type ='%s', leave_time='%s', bt='%s', et='%s'" \
                      "WHERE rings_date = '%s' and carrier_name = '%s'" % (
                          self.current_array[2], self.current_array[3], self.current_array[4], self.current_array[5],
                          self.current_array[6], self.current_array[7], self.current_array[8], self.current_array[9],
                          self.current_array[0], self.current_array[1])
                commit(sql)

        def build_protoarray(self, rings):
            """ build the protoarray. One days worth of rings are sent in rings arg an put into daily rings.
            """
            self.daily_rings = rings
            self.skim_daily_initialize()  # zero out all daily values for each iteration
            if len(self.daily_rings) > 0:
                self.skim_name()  # get the carrier id from the tacs data
                for line in self.daily_rings:
                    self.daily_line = line
                    # get 5200 or non 5200 times for TOTAL, code, leave_type and leave_time
                    if self.daily_line[18] in self.days:
                        self.skim_dayofweek()  # get the day of the week from the tacs data line as self.day_dayofweek
                        self.skim_get_hours()  # get hours as for the day as self.day_hr_52, etc
                        if float(self.day_hr_55) > 0 or float(self.day_hr_56) > 0 or float(self.day_hr_58) > 0 or \
                                float(self.day_hr_62) > 0 or float(self.day_hr_86) > 0:
                            self.skim_daily_leavetime()  # fill day leave type and time variables
                        self.skim_get_code()  # detects annual or sick leave for day_code variable
                    # get the RETURN TO OFFICE time
                    if self.daily_line[19] == "MV" and self.daily_line[23][:3] == "722":
                        self.skim_get_returntostation()  # get return to station time and fill day_rs variable
                    if self.daily_line[19] in self.mv_codes:  # get the MOVES
                        self.skim_get_moves()  # build an array of moves for the day
                    if self.daily_line[19] == "BT":  # get the begin tour time
                        self.skim_get_begintour()  # define self.day_bt
                    if self.daily_line[19] == "ET":  # get the end tour time
                        self.skim_get_endtour()  # define self.day_et
                proto_array = [self.day_dayofweek, self.day_name, self.day_hr_52, self.day_rs, self.day_code,
                               self.day_moves, self.day_final_leave_type, self.day_final_leave_time,
                               self.day_bt, self.day_et]
                return proto_array  # send it back to auto weekly analysis()

        def skim_daily_initialize(self):
            """ initialize variables for build_protoarray()self """
            self.day_hr_52 = 0.0  # work hours
            self.day_hr_55 = 0.0  # annual leave
            self.day_hr_56 = 0.0  # sick leave
            self.day_hr_58 = 0.0  # holiday leave
            self.day_hr_62 = 0.0  # guaranteed time
            self.day_hr_86 = 0.0  # other paid leave
            self.day_rs = 0
            self.day_code = ""
            self.day_moves = []
            self.day_leave_type = []
            self.day_leave_time = []
            self.day_final_leave_type = ""
            self.day_final_leave_time = 0.0
            self.day_bt = 0
            self.day_et = 0
            self.day_dayofweek = None

        def skim_name(self):
            """ get the carrier id from the tacs data """
            self.day_name = self.daily_rings[0][4].zfill(8)  # Get NAME

        def skim_dayofweek(self):
            """ get the day of the week from the tacs data line """
            self.day_dayofweek = self.daily_line[18]

        def skim_get_hours(self):
            """ identify different types of hours. """
            spt_20 = self.daily_line[20].split(':')  # split to get code and hours
            # get second and third digits of the of the split line 20 or spt_20
            spt_20_mod = "".join([spt_20[0][1], spt_20[0][2]])
            if spt_20_mod == "52":
                self.day_hr_52 = spt_20[1]  # get the total hours worked
            if spt_20_mod == "55":
                self.day_hr_55 = spt_20[1]  # get the annual leave hours
            if spt_20_mod == "56":
                self.day_hr_56 = spt_20[1]  # get the sick leave hours
            if spt_20_mod == "58":
                self.day_hr_58 = spt_20[1]  # get the holiday leave hours
            if spt_20_mod == "62":
                self.day_hr_62 = spt_20[1]  # get the guaranteed time hours
            if spt_20_mod == "86":
                self.day_hr_86 = spt_20[1]  # get other leave hours

        def skim_daily_leavetime(self):
            """ fill day leave type and time variables """
            if float(self.day_hr_55) > 0:
                self.day_leave_type.append("annual")
                self.day_leave_time.append(self.day_hr_55)
            if float(self.day_hr_56) > 0:
                self.day_leave_type.append("sick")
                self.day_leave_time.append(self.day_hr_56)
            if float(self.day_hr_58) > 0:
                self.day_leave_type.append("holiday")
                self.day_leave_time.append(self.day_hr_58)
            if float(self.day_hr_62) > 0:
                self.day_leave_type.append("guaranteed")
                self.day_leave_time.append(self.day_hr_62)
            if float(self.day_hr_86) > 0:
                self.day_leave_type.append("other")
                self.day_leave_time.append(self.day_hr_86)
            if len(self.day_leave_type) > 1:
                self.day_final_leave_type = "combo"
                self.day_final_leave_time = \
                    float(self.day_hr_55) + float(self.day_hr_56) + float(self.day_hr_58) + \
                    float(self.day_hr_62) + float(self.day_hr_86)
            elif len(self.day_leave_type) == 1:
                self.day_final_leave_type = self.day_leave_type[0]
                self.day_final_leave_time = self.day_leave_time[0]
            else:
                self.day_final_leave_type = ""
                self.day_final_leave_time = 0.0

        def skim_get_code(self):
            """ detects annual or sick leave for day_code variable """
            if float(self.day_hr_55) > 1:
                self.day_code = "annual"  # alter CODE if annual leave was used
            if float(self.day_hr_56) > 1:
                self.day_code = "sick"  # alter code if sick leave was used

        def skim_get_returntostation(self):
            """ get return to station time and fill day_rs variable """
            self.day_rs = self.daily_line[21]  # save the last occurrence.

        def skim_get_moves(self):
            """ build an array of moves for the day """
            route_z = self.daily_line[24].zfill(6)  # because some reports omit leading zeros
            # reformat route to 5 digit format
            route = route_z[1] + route_z[2] + route_z[3] + route_z[4] + route_z[5]  # build 5 digit route number
            route = Handler(route).routes_adj()  # convert to 4 digits if route < 100
            # MV code, time off, time on, route
            mv_data = [self.daily_line[19], self.daily_line[21], self.daily_line[23][:3], route]
            self.day_moves.append(mv_data)

        def skim_get_begintour(self):
            """ the the begin tour time for the proto array"""
            self.day_bt = self.daily_line[21]

        def skim_get_endtour(self):
            """ get the end tour time for the proto array """
            self.day_et = self.daily_line[21]

        @staticmethod
        def carrier_list_cleaning():
            """ cleans the database of duplicate records """
            sql = "SELECT * FROM carriers ORDER BY carrier_name, effective_date"
            results = inquire(sql)
            duplicates = []
            for i in range(len(results)):
                if i != len(results) - 1:  # if the loop has not reached the end of the list
                    if results[i][1] == results[i + 1][1] and \
                            results[i][2] == results[i + 1][2] and \
                            results[i][3] == results[i + 1][3] and \
                            results[i][4] == results[i + 1][4] and \
                            results[i][5] == results[i + 1][5]:  # if the name current and next name are the same
                        duplicates.append(i + 1)
            if len(duplicates) > 0:
                pb_root = Tk()  # create a window for the progress bar
                pb_root.title("Database Maintenance")
                titlebar_icon(pb_root)  # place icon in titlebar
                pb_label = Label(pb_root, text="Updating Changes: ")  # make label for progress bar
                pb_label.pack(side=LEFT)
                pb = ttk.Progressbar(pb_root, length=400, mode="determinate")  # create progress bar
                pb.pack(side=LEFT)
                pb["maximum"] = len(duplicates)  # set length of progress bar
                pb.start()
                i = 0
                for d in duplicates:
                    pb["value"] = i  # increment progress bar
                    sql = "DELETE FROM carriers WHERE effective_date='%s' and carrier_name='%s'" % (
                        results[d][0], results[d][1])
                    commit(sql)
                    pb_root.update()
                    i += 1
                pb.stop()  # stop and destroy the progress bar
                pb_label.destroy()  # destroy the label for the progress bar
                pb.destroy()
                pb_root.destroy()
            del duplicates[:]


class StartUp:
    """
    This class creates a screen which is displayed if there are no stations in the station index. It will show up
    when the program is first started.
    """

    def __init__(self):
        self.win = None
        self.new_station = None

    def start(self):
        """ a master method for running the other methods in proper order. """
        self.win = MakeWindow()
        self.win.create(None)
        self.new_station = StringVar(self.win.body)
        self.build()
        self.buttons_frame()
        self.win.finish()

    def build(self):
        """ fills the screen with widgets. """
        Label(self.win.body, text="Welcome to Klusterbox", font=macadj("bold", "Helvetica 18")) \
            .grid(row=0, columnspan=2, sticky="w")
        Label(self.win.body, text="version: {}".format(version)).grid(row=1, columnspan=2, sticky="w")
        Label(self.win.body, text="", pady=20).grid(row=2, column=0)
        # enter new stations
        Label(self.win.body, text="To get started, please enter your station name:", pady=5) \
            .grid(row=3, columnspan=2, sticky="w")
        e = Entry(self.win.body, width=35, textvariable=self.new_station)
        e.grid(row=4, column=0, sticky="w")
        self.new_station.set("")
        Button(self.win.body, width=5, anchor="w", text="ENTER",
               command=lambda: self.apply_startup()).grid(row=4, column=1, sticky="w")
        Label(self.win.body, text="", pady=20).grid(row=5, columnspan=2, sticky="w")
        Label(self.win.body, text="Or you can exit to the main screen and enter your\n"
                                  "station by going to Management > list of stations.").grid(row=6, columnspan=2,
                                                                                             sticky="w")
        Button(self.win.body, width=5, text="EXIT",
               command=lambda: MainFrame().start(frame=self.win.topframe)).grid(row=7, columnspan=2, sticky="e")

    def buttons_frame(self):
        """ creates a label at the bottom of the screen. """
        Label(self.win.buttons, text="").pack()

    def apply_startup(self):
        """ run checks after the user presses apply. """
        station = self.new_station.get().strip()  # simplify the var name
        if not station:
            messagebox.showerror("Prohibited Action",
                                 "You can not enter a blank entry for a station.",
                                 parent=self.win.body)
            return
        sql = "INSERT INTO stations (station) VALUES('%s')" % station
        commit(sql)
        projvar.list_of_stations.append(station)
        # access list of stations from database
        sql = "SELECT * FROM stations ORDER BY station"
        results = inquire(sql)
        # define and populate list of stations variable
        del projvar.list_of_stations[:]
        for stat in results:
            projvar.list_of_stations.append(stat[0])
        # create minimum records in dov table for the station
        DovBase().minimum_recs(station)
        MainFrame().start(frame=self.win.topframe)  # load new frame


class GenConfig:
    """
    This class sets up the GUI Configuration screen used for setting mouse wheel orientation, investigation range
    display mode, overtime rings limiter, and tour rings display.
    """

    def __init__(self, frame):
        self.frame = frame
        self.win = None
        self.wheel_selection = None  # stringvar
        self.nav_selection = None  # stringvar
        self.invran_mode = None  # stringvar
        self.ot_rings_limiter = None  # stringvar
        self.tourrings_var = None  # stringvar
        self.spreadsheet_pref = None  # stringvar
        self.nav_result = None  # True (1) for button - mac navigation, False (0) for pulldown menu
        self.tourrings = None  # True to show BT/ET rings, False to hide
        self.rings_limiter = None  # ot rings limiter status from tolerance table
        self.invran_result = None  # investigation range mode from tolerance table
        self.spreadsheet_result = None  # the spreadsheet_pref from the tolerance table.
        self.row = 0
        self.status_update = None  # a label widget for status report

    def create(self):
        """ this is a master method for calling other methods in the class in sequence. """
        self.get_settings()
        self.get_window_object()
        self.get_stringvars()
        self.build()
        self.button_frame()
        self.win.finish()

    def get_settings(self):
        """ get records from the database and define variables. """
        sql = "SELECT tolerance FROM tolerances WHERE category = '%s'" % "mousewheel"
        results = inquire(sql)
        projvar.mousewheel = int(results[0][0])
        sql = "SELECT tolerance FROM tolerances WHERE category = '%s'" % "mac_navigation"
        results = inquire(sql)
        self.nav_result = int(results[0][0])
        sql = "SELECT tolerance FROM tolerances WHERE category = '%s'" % "invran_mode"
        results = inquire(sql)
        self.invran_result = results[0][0]
        sql = "SELECT tolerance FROM tolerances WHERE category = '%s'" % "spreadsheet_pref"
        results = inquire(sql)
        self.spreadsheet_result = results[0][0]
        sql = "SELECT tolerance FROM tolerances WHERE category = '%s'" % "ot_rings_limiter"
        results = inquire(sql)
        rings_limiter = results[0][0]
        self.rings_limiter = Convert(rings_limiter).bool_to_onoff()  # convert the bool to on or off
        sql = "SELECT tolerance FROM tolerances WHERE category = '%s'" % "tourrings"
        results = inquire(sql)
        tourrings = results[0][0]
        self.tourrings = Convert(tourrings).bool_to_onoff()  # convert the bool to on or off

    def get_window_object(self):
        """ create the window object and define self.win """
        self.win = MakeWindow()
        self.win.create(self.frame)

    def get_stringvars(self):
        """ create the stringvars """
        self.wheel_selection = StringVar(self.win.body)
        self.nav_selection = StringVar(self.win.body)
        self.invran_mode = StringVar(self.win.body)
        self.spreadsheet_pref = StringVar(self.win.body)
        self.ot_rings_limiter = StringVar(self.win.body)
        self.tourrings_var = StringVar(self.win.body)

    def build(self):
        """ build the screens """
        Label(self.win.body, text="General Configurations", font=macadj("bold", "Helvetica 18"), anchor="w") \
            .grid(row=self.row, sticky="w", columnspan=14)
        self.row += 1
        Label(self.win.body, text=" ").grid(row=self.row, column=0)
        self.row += 1
        text = macadj("Interface Configurations __________________________________",
                      "Interface Configurations _________________________")
        Label(self.win.body, text=text, anchor="w",
              fg="blue").grid(row=self.row, column=0, columnspan=14, sticky="w")
        self.row += 1
        # mousewheel scrolling direction
        Label(self.win.body, text="Mouse Wheel Scrolling:  ", anchor="w").grid(row=self.row, column=0, sticky="w")
        om_wheel = OptionMenu(self.win.body, self.wheel_selection, "natural", "reverse")  # option menu configuration
        om_wheel.config(width=7)
        om_wheel.grid(row=self.row, column=1)
        if projvar.mousewheel == 1:
            self.wheel_selection.set("natural")
        else:
            self.wheel_selection.set("reverse")
        Button(self.win.body, text="set", width=7, command=lambda: self.apply_mousewheel()).grid(row=self.row, column=2)
        self.row += 1
        Label(self.win.body, text=" ").grid(row=self.row, column=0)
        self.row += 1
        # pulldown or navigation button
        Label(self.win.body, text="Navigation Preference:  ", anchor="w").grid(row=self.row, column=0, sticky="w")
        om_wheel = OptionMenu(self.win.body, self.nav_selection, "pulldown", "button")  # option menu config
        om_wheel.config(width=7)
        om_wheel.grid(row=self.row, column=1)
        if self.nav_result:
            self.nav_selection.set("button")
        else:
            self.nav_selection.set("pulldown")
        Button(self.win.body, text="set", width=7, command=lambda: self.__apply_navigation())\
            .grid(row=self.row, column=2)
        self.row += 1
        Label(self.win.body, text=" ").grid(row=self.row, column=0)
        self.row += 1
        # investigation range mode
        Label(self.win.body, text="Investigation Range Mode:  ", anchor="w").grid(row=self.row, column=0, sticky="w")
        om_invran_mode = OptionMenu(self.win.body, self.invran_mode, "original", "simple", "no labels")
        om_invran_mode.config(width=7)
        om_invran_mode.grid(row=self.row, column=1)
        self.invran_mode.set(self.invran_result)
        Button(self.win.body, text="set", width=7,
               command=lambda: self.apply_invran_mode()).grid(row=self.row, column=2)
        self.row += 1
        Label(self.win.body, text=" ").grid(row=self.row, column=0)
        self.row += 1
        # spreadsheet preference - any new options must be added to MainFrame().define_spreadsheet_button().
        pref_options = (
            "Mandates",
            "Over Max",
            "OT Equitability",
            "OT Distribution",
            "Mandates_4",
            "Off Bid"
        )
        Label(self.win.body, text="Spreadsheet Preference:  ", anchor="w").grid(row=self.row, column=0, sticky="w")
        om_sheet = OptionMenu(self.win.body, self.spreadsheet_pref, *pref_options)  # option menu configuration below
        om_sheet_width = macadj(18, 14)
        om_sheet.config(width=om_sheet_width)
        om_sheet.grid(row=self.row, column=1, columnspan=2)
        self.spreadsheet_pref.set(self.spreadsheet_result)
        self.row += 1
        Button(self.win.body, text="set", width=7,
               command=lambda: self.apply_spreadsheet_pref()).grid(row=self.row, column=2)
        self.row += 1
        Label(self.win.body, text=" ").grid(row=self.row, column=0)
        self.row += 1
        # overtime rings limiter - rename to "Hide OTDL Move Rings"
        Label(self.win.body, text="Hide OTDL Move Rings:  ", anchor="w").grid(row=self.row, column=0, sticky="w")
        om_rings = OptionMenu(self.win.body, self.ot_rings_limiter, "on", "off")  # option menu configuration below
        om_rings.config(width=7)
        om_rings.grid(row=self.row, column=1)
        self.ot_rings_limiter.set(self.rings_limiter)
        Button(self.win.body, text="set", width=7,
               command=lambda: self.apply_rings_limiter()).grid(row=self.row, column=2)
        self.row += 1
        Label(self.win.body, text=" ").grid(row=self.row, column=0)
        self.row += 1
        # tourrings - show bt et rings
        text = macadj("Expanded Clock Rings ____________________________________",
                      "Expanded Clock Rings ___________________________")
        Label(self.win.body, text=text, anchor="w",
              fg="blue").grid(row=self.row, column=0, columnspan=14, sticky="w")
        self.row += 1
        Label(self.win.body, text="Show BT/ET Rings:  ", anchor="w").grid(row=self.row, column=0, sticky="w")
        om_tourrings = OptionMenu(self.win.body, self.tourrings_var, "on", "off")  # option menu configuration below
        om_tourrings.config(width=7)
        om_tourrings.grid(row=self.row, column=1)
        self.tourrings_var.set(self.tourrings)
        Button(self.win.body, text="set", width=7,
               command=lambda: self.apply_tourrings()).grid(row=self.row, column=2)

    def button_frame(self):
        """ Display buttons and status update message """
        button = Button(self.win.buttons)
        button.config(text="Go Back", width=20, command=lambda: MainFrame().start(frame=self.win.topframe))
        if sys.platform == "win32":
            button.config(anchor="w")
        button.pack(side=LEFT)
        self.status_update = Label(self.win.buttons, text="", fg="red")
        self.status_update.pack(side=LEFT)

    def apply_rings_limiter(self):
        """ apply the ot rings limiter """
        if self.ot_rings_limiter.get() == "on":
            rings_limiter = int(1)
        else:
            rings_limiter = int(0)
        sql = "UPDATE tolerances SET tolerance='%s'WHERE category='%s'" % (rings_limiter, "ot_rings_limiter")
        commit(sql)
        msg = "Hide OTDL Move Rings updated: {}".format(self.ot_rings_limiter.get())
        self.status_update.config(text="{}".format(msg))

    def apply_spreadsheet_pref(self):
        """ apply the spreadsheet preference. """
        sql = "UPDATE tolerances SET tolerance='%s'WHERE category='%s'" % \
              (self.spreadsheet_pref.get(), "spreadsheet_pref")
        commit(sql)
        msg = "Spreadsheet Preference updated: {}".format(self.spreadsheet_pref.get())
        self.status_update.config(text="{}".format(msg))

    def apply_invran_mode(self):
        """ apply investigation range mode. """
        sql = "UPDATE tolerances SET tolerance='%s'WHERE category='%s'" % (self.invran_mode.get(), "invran_mode")
        commit(sql)
        msg = "Investigation Range mode updated: {}".format(self.invran_mode.get())
        self.status_update.config(text="{}".format(msg))

    def apply_mousewheel(self):
        """ apply mouse wheel configuration """
        if self.wheel_selection.get() == "natural":
            wheel_multiple = int(1)
            projvar.mousewheel = int(1)  # sets the project variable
        else:  # if the self.wheel_selection.get() == "reverse"
            wheel_multiple = int(-1)
            projvar.mousewheel = int(-1)  # sets the project variable
        sql = "UPDATE tolerances SET tolerance='%s' WHERE category='%s'" % (wheel_multiple, "mousewheel")
        commit(sql)
        msg = "Mousescroll direction updated: {}".format(self.wheel_selection.get())
        self.status_update.config(text="{}".format(msg))

    def __apply_navigation(self):
        """ apply the navigation method - either a pulldown menu or a button.
        the button method is prefered for the mac, as pulldown menus cause the program to crash.
        True (1) will create a navigation button and page, False (0) will use the pulldown menu. """
        if self.nav_selection.get() == "pulldown":
            nav_value = 0
            projvar.mac_navigation = 0
        else:  # if the self.nav_selection.get() == "pulldown"
            nav_value = 1
            projvar.mac_navigation = 1
        sql = "UPDATE tolerances SET tolerance='%s' WHERE category='%s'" % (nav_value, "mac_navigation")
        commit(sql)
        msg = "Navigation preference updated: {}".format(self.nav_selection.get())
        self.status_update.config(text="{}".format(msg))

    def apply_tourrings(self):
        """ apply tour rings """
        if self.tourrings_var.get() == "on":  # convert tourrings to boolean values
            tourrings = int(1)
        else:
            tourrings = int(0)
        sql = "UPDATE tolerances SET tolerance='%s' WHERE category='%s'" % (tourrings, "tourrings")
        commit(sql)
        msg = "Show BT/ET rings updated: {}".format(self.tourrings_var.get())
        self.status_update.config(text="{}".format(msg))


class StationList:
    """
    creates a window for users to view, change, rename and delete stations from the station list.
    """

    def __init__(self):
        self.win = None

    def station_list(self, frame):
        """
        creates a screen that allows the user to add stations.
        """
        self.win = MakeWindow()
        self.win.create(frame)
        # page title
        row = 0
        Label(self.win.body, text="Manage Station List", font=macadj("Arial 12", "Helvetica 18")) \
            .grid(row=row, columnspan=13, sticky="w")
        row += 1
        Label(self.win.body, text="").grid(row=row)
        row += 1
        # enter new stations
        new_name = StringVar(self.win.body)
        text = macadj("Enter New Station __________________________________",
                      "Enter New Station ______________________")
        Label(self.win.body, text=text, pady=5, fg="blue").grid(row=row, columnspan=3, sticky="w")
        row += 1
        e = Entry(self.win.body, width=macadj(35, 24), textvariable=new_name)
        e.grid(row=row, column=0, sticky="w")
        new_name.set("")
        Button(self.win.body, width=5, anchor="w", text="ENTER",
               command=lambda: self.apply_station("enter", new_name)). \
            grid(row=row, column=1, sticky="w")
        row += 1
        Label(self.win.body, text="").grid(row=row)
        row += 1
        # list current list of stations and delete buttons.
        sql = "SELECT * FROM stations ORDER BY station"
        results = inquire(sql)
        text = macadj("List Of Stations _____________________________________",
                      "List Of Stations ________________________")
        Label(self.win.body, text=text, fg="blue", pady=5).grid(row=row, columnspan=13, sticky="w")
        row += 1
        for record in results:
            button_width = macadj(30, 25)
            Button(self.win.body, text=record[0], width=button_width, anchor="w").grid(row=row, column=0, sticky="w")
            Button(self.win.body, text="delete",
                   command=lambda x=record[0]: self.apply_station("delete", x)) \
                .grid(row=row, column=1, sticky="w")
            row += 1

        if len(results) > 1:
            Label(self.win.body, text="").grid(row=row)
            row += 1
            # change names of stations
            text = macadj("Change Station Name ______________________________",
                          "Change Station Name ___________________")
            Label(self.win.body, text=text, fg="blue").grid(row=row, column=0, columnspan=13, sticky="w")
            row += 1
            all_stations = []
            for rec in results:
                all_stations.append(rec[0])
            if "out of station" in all_stations:
                all_stations.remove("out of station")
            old_station = StringVar(self.win.body)
            om = OptionMenu(self.win.body, old_station, *all_stations)
            om_width = macadj("35", "27")
            om.config(width=om_width)
            om.grid(row=row, column=0, sticky="w", columnspan=2)
            row += 1
            old_station.set("select a station")
            Label(self.win.body, text="enter a new name:").grid(row=row, column=0, sticky="w")
            row += 1
            new_station = StringVar(self.win.body)
            Entry(self.win.body, textvariable=new_station, width=macadj("35", "24")).grid(row=row, column=0, sticky="w")
            new_station.set("enter a new station name")
            Button(self.win.body, text="update", command=lambda: self.station_update_apply(old_station, new_station)) \
                .grid(row=row, column=1, sticky="w")
            row += 1
        # find and display list of unique stations
        Label(self.win.body, text="").grid(row=row)
        row += 1
        text = macadj("List Of Stations _____________________________________",
                      "List Of Stations ________________________")
        Label(self.win.body, text=text, pady=5, fg="blue").grid(row=row, columnspan=13, sticky="w")
        row += 1
        Label(self.win.body, text="(referenced in carrier database)", pady=5) \
            .grid(row=row, columnspan=3, sticky="w")
        row += 1
        unique_station = []
        sql = "SELECT * FROM carriers"
        results = inquire(sql)
        for name in results:
            if name[5] not in unique_station:
                unique_station.append(name[5])
        unique_station = sorted(unique_station, key=str.lower)
        count = 1
        for ss in unique_station:
            Label(self.win.body, text="{}.  {}".format(count, ss)).grid(row=row, columnspan=2, sticky="w")
            count += 1
            row += 1
        # the screen has just one button - go back which returns to the main screeen.
        gobackbutton = Button(self.win.buttons)
        gobackbutton.config(text="Go Back", width=20, command=lambda: MainFrame().start(frame=self.win.topframe))
        if sys.platform == "win32":
            gobackbutton.config(anchor="w")
        gobackbutton.pack(side=LEFT)
        self.win.finish()  # main loop, etc

    def apply_station(self, switch, station):
        """ checks and enters stations into the station table. """
        if switch == "enter":
            if station.get().strip() == "" or station.get().strip() == "x":
                messagebox.showerror("Prohibited Action",
                                     "You can not enter a blank entry for a station.",
                                     parent=self.win.body)
                return
            if station.get() in projvar.list_of_stations:
                messagebox.showerror("Prohibited Action",
                                     "That station is already in the list of stations.",
                                     parent=self.win.body)
                return
            sql = "INSERT INTO stations (station) VALUES('%s')" % (station.get().strip())
            commit(sql)
            projvar.list_of_stations.append(station.get())
            # add minimum recs to DOV table
            DovBase().minimum_recs(station.get())
        if switch == "delete":
            if station == "out of station":
                text = "You can not delete the \"out of station\" listing."
                messagebox.showerror("Action not allowed", text, parent=self.win.body)
                return
            if messagebox.askokcancel("Delete Station",
                                      "Are you sure you want to delete {}? \n"
                                      "The station will be deleted and maintenance actions will\n"
                                      "clean any orphan carriers, clock rings and indexes from\n"
                                      "database. This can not be reversed.".format(station),
                                      parent=self.win.body):
                sql = "DELETE FROM stations WHERE station='%s'" % station
                commit(sql)
                DatabaseAdmin().database_clean_carriers()
                DatabaseAdmin().database_clean_rings()
                if projvar.invran_station == station:
                    Globals().reset()  # reset initial value of globals
            # delete from DOV table
            sql = "DELETE FROM dov WHERE station = '%s'" % station
            commit(sql)
        # access list of stations from database
        sql = "SELECT * FROM stations ORDER BY station"
        results = inquire(sql)
        # define and populate list of stations variable
        del projvar.list_of_stations[:]
        for stat in results:
            projvar.list_of_stations.append(stat[0])
        self.station_list(self.win.topframe)

    def station_update_apply(self, old_station, new_station):
        """ change the name of a station. """
        if old_station.get() == "select a station":
            messagebox.showerror("Prohibited Action",
                                 "Please select a station.",
                                 parent=self.win.body)
            return
        if new_station.get().strip() == "" or \
                new_station.get() == "enter a new station name" or \
                new_station.get().strip() == "x":
            messagebox.showerror("Prohibited Action",
                                 "You can not enter a blank entry for a station.",
                                 parent=self.win.body)
            return
        if projvar.invran_station == old_station.get():
            Globals().reset()  # reset initial value of globals
        go_ahead = True
        duplicate = False
        if new_station.get() in projvar.list_of_stations:
            go_ahead = messagebox.askokcancel("Duplicate Detected",
                                              "This station already exist in the list of stations. "
                                              "If you proceed, all records for {} will be merged with "
                                              "records from {}. Do you want to proceed?"
                                              .format(old_station.get(), new_station.get()),
                                              parent=self.win.body)
            duplicate = True
        if duplicate and go_ahead:
            sql = "DELETE FROM stations WHERE station='%s'" % old_station.get()
            commit(sql)
            projvar.list_of_stations.remove(new_station.get())
        if go_ahead:
            # update in stations table
            sql = "UPDATE stations SET station='%s' WHERE station='%s'" % (new_station.get(), old_station.get())
            commit(sql)
            # update in carriers table
            sql = "UPDATE carriers SET station='%s' WHERE station='%s'" % (new_station.get(), old_station.get())
            commit(sql)
            # update in station index table
            sql = "UPDATE station_index SET kb_station='%s' WHERE kb_station='%s'" % \
                  (new_station.get(), old_station.get())
            commit(sql)
            # update in DOV table
            sql = "UPDATE dov SET station='%s' WHERE station='%s'" % (new_station.get(), old_station.get())
            commit(sql)
            """ update the the project variable for list of stations: """
            projvar.list_of_stations.append(new_station.get())  # add the new station name
            projvar.list_of_stations.remove(old_station.get())  # remove the old station name
            self.station_list(self.win.topframe)
        if not go_ahead:
            return


class SetDov:
    """
    creates a scren where the user can change, delete and reset the dispatch of value for each day - Saturday through
    Friday
    """

    def __init__(self):
        self.frame = None
        self.win = None
        self.autofill = None  # stringvar for autofill
        self.dovsat = StringVar()  # stringvar for saturday dov
        self.dovsun = StringVar()  # stringvar for sunday dov
        self.dovmon = StringVar()  # stringvar for monday dov
        self.dovtue = StringVar()  # stringvar for tuesday dov
        self.dovwed = StringVar()  # stringvar for wednesday dov
        self.dovthu = StringVar()  # stringvar for thursday dov
        self.dovfri = StringVar()  # stringvar for friday dov
        self.dovarray = [self.dovsat, self.dovsun, self.dovmon, self.dovtue, self.dovwed, self.dovthu, self.dovfri]
        self.days = ("Saturday", "Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday")
        self.day = ("sat", "sun", "mon", "tue", "wed", "thu", "fri")
        self.checksat = BooleanVar()  # bool for saturday "temporary"
        self.checksun = BooleanVar()  # bool for sunday "temporary"
        self.checkmon = BooleanVar()  # bool for monday "temporary"
        self.checktue = BooleanVar()  # bool for tuesday "temporary"
        self.checkwed = BooleanVar()  # bool for wednesday "temporary"
        self.checkthu = BooleanVar()  # bool for thursday "temporary"
        self.checkfri = BooleanVar()  # bool for friday "temporary"
        self.checkarray = [self.checksat, self.checksun, self.checkmon, self.checktue, self.checkwed,
                           self.checkthu, self.checkfri]
        # all records in the dov table
        self.onrecsat = []
        self.onrecsun = []
        self.onrecmon = []
        self.onrectue = []
        self.onrecwed = []
        self.onrecthu = []
        self.onrecfri = []
        self.onrecarray = [self.onrecsat, self.onrecsun, self.onrecmon, self.onrectue, self.onrecwed, self.onrecthu,
                           self.onrecfri]
        # shows record if there is one for the current day, if blank if there is no record for the current day
        self.now_onrecsat = False
        self.now_onrecsun = False
        self.now_onrecmon = False
        self.now_onrectue = False
        self.now_onrecwed = False
        self.now_onrecthu = False
        self.now_onrecfri = False
        self.now_onrecarray = [self.now_onrecsat, self.now_onrecsun, self.now_onrecmon, self.now_onrectue,
                               self.now_onrecwed, self.now_onrecthu, self.now_onrecfri]
        # get the most recent record which is not temporary
        self.perm_onrecsat = []
        self.perm_onrecsun = []
        self.perm_onrecmon = []
        self.perm_onrectue = []
        self.perm_onrecwed = []
        self.perm_onrecthu = []
        self.perm_onrecfri = []
        self.perm_onrecarray = [self.perm_onrecsat, self.perm_onrecsun, self.perm_onrecmon, self.perm_onrectue,
                                self.perm_onrecwed, self.perm_onrecthu, self.perm_onrecfri]
        self.addsat = ""
        self.addsun = ""
        self.addmon = ""
        self.addtue = ""
        self.addwed = ""
        self.addthu = ""
        self.addfri = ""
        self.addarray = [self.addsat, self.addsun, self.addmon, self.addtue, self.addwed, self.addthu, self.addfri]
        self.insert_counter = 0
        self.update_counter = 0
        self.status_update = None

    def run(self, frame):
        """ a master method for running the other methods in the proper order. """
        self.frame = frame
        self.win = MakeWindow()
        self.win.create(self.frame)
        self.get_onrecs()
        self.get_now_onrecs()
        self.get_premrecs()
        self.make_stringvars()
        self.set_stringvars()
        self.build_screen()
        self.buttons_frame()
        self.win.finish()

    def get_onrecs(self):
        """ get the records currently in the database """
        for i in range(len(self.days)):
            sql = "SELECT * FROM dov WHERE eff_date <= '%s' AND station = '%s' AND day = '%s' " \
                  "ORDER BY eff_date DESC" % \
                  (projvar.invran_date_week[0], projvar.invran_station, self.day[i])
            result = inquire(sql)
            self.onrecarray[i] = result

    def get_now_onrecs(self):
        """ sets now onrecs to True if there is a record for the current day. That value is put into an array of
        seven days - one boolean for each day. """
        for i in range(len(self.onrecarray)):
            if self.onrecarray[i][0]:
                onrec_date = self.onrecarray[i][0][0]
                onrec_day = self.day[i]
                invran_date = Convert(projvar.invran_date_week[0]).dt_to_str()
                invran_day = Convert(projvar.invran_date_week[i]).dt_to_day_str()
                if onrec_date == invran_date:  # compare dates - is always saturday in the service week
                    if onrec_day == invran_day:  # compare days - mon vs tue, etc
                        self.now_onrecarray[i] = True

    def get_premrecs(self):
        """ get the permenent records - all recs that have temp set to False """
        for i in range(len(self.onrecarray)):
            for rec in self.onrecarray[i]:
                if rec[4] == 'False':
                    self.perm_onrecarray[i] = rec
                    break

    def make_stringvars(self):
        """ define and create the stringvars using a loop"""
        self.autofill = StringVar(self.win.body)
        for i in range(7):
            self.dovarray[i] = StringVar(self.win.body)

    def set_stringvars(self):
        """ set the values for the stringvars using a loop """
        for i in range(7):
            recfortoday = self.now_onrecarray[i]
            if recfortoday:  # use the most recent record - which is for the current day
                daily_temp = Convert(self.onrecarray[i][0][4]).str_to_bool()  # make boolean a proper boolean
                self.dovarray[i].set(self.onrecarray[i][0][3])  # get time - time is fourth item in results
                self.checkarray[i].set(daily_temp)
            else:  # use the most recent record where "temp" is False
                daily_temp = Convert(self.perm_onrecarray[i][4]).str_to_bool()  # make boolean a proper boolean
                self.dovarray[i].set(self.perm_onrecarray[i][3])  # get time - time is fourth item in results
                self.checkarray[i].set(daily_temp)

    def build_screen(self):
        """ generates the widgets to build the screen. """
        row = 0
        Label(self.win.body, text="Dispatch of Value Settings", font=macadj("Arial 12", "Helvetica 18")) \
            .grid(row=row, columnspan=14, sticky="w")
        row += 1
        Label(self.win.body, text="").grid(row=row)
        row += 1
        text = macadj("Autofill Dispatch of Value (DOV) _____________________________",
                      "Autofill Dispatch of Value (DOV) ___________________")
        Label(self.win.body, text=text, pady=5, fg="blue").grid(row=row, columnspan=14, sticky="w")
        row += 1
        Entry(self.win.body, width=7, textvariable=self.autofill).grid(row=row, column=0, sticky="w")
        Button(self.win.body, width=5, text="autofill", command=lambda: self.applyautofill()) \
            .grid(row=row, column=4, sticky="w")
        row += 1
        text = "Use Autofill to fill a value as a time formatted in military time with clicks, not minutes, " \
               "for all days at once. Then press apply or submit."
        Label(self.win.body, text=text, wraplength=300, anchor="w", justify=LEFT) \
            .grid(row=row, columnspan=14, sticky="w")
        row += 1
        Label(self.win.body, text="").grid(row=row)
        row += 1
        text = macadj("Set Dispatch of Value (DOV) __________________________________",
                      "Set Dispatch of Value (DOV) ______________________")
        Label(self.win.body, text=text, pady=5, fg="blue").grid(row=row, columnspan=14, sticky="w")
        row += 1
        for i in range(len(self.dovarray)):
            Entry(self.win.body, width=7, textvariable=self.dovarray[i]).grid(row=row, column=0, sticky="w")
            Checkbutton(self.win.body, variable=self.checkarray[i], onvalue=True, offvalue=False). \
                grid(row=row, column=1, sticky="w")
            Label(self.win.body, width=10, text=self.days[i], anchor="w").grid(row=row, column=2, sticky="w")
            date = Convert(projvar.invran_date_week[i]).dt_to_backslash_str()
            Label(self.win.body, width=10, text=date, anchor="w").grid(row=row, column=3, sticky="w")
            row += 1
        text = "Fill in the daily dispatch of value times. Use a military time with clicks not minutes. Click the " \
               "checkbox next to the time if the daily time is temporary and one time only. This allow you to change " \
               "the DOV for just one day before going back to the previous setting. "
        Label(self.win.body, text=text, wraplength=300, anchor="w", justify=LEFT). \
            grid(row=row, columnspan=14, sticky="w")
        row += 1
        Label(self.win.body, text="").grid(row=row)  # whitespace
        row += 1
        text = macadj("Dispatch of Value History ____________________________________",
                      "Dispatch of Value History ________________________")
        Label(self.win.body, text=text, pady=5, fg="blue").grid(row=row, columnspan=14, sticky="w")
        row += 1
        Label(self.win.body, text="Generate Text Report: ", anchor="w") \
            .grid(row=row, column=0, sticky="w", columnspan=10)
        Button(self.win.body, text="Report", width=5, command=lambda: self.generate_report()).grid(row=row, column=4)
        row += 1
        Label(self.win.body, text="Delete History: ", anchor="w") \
            .grid(row=row, column=0, sticky="w", columnspan=10)
        Button(self.win.body, text="Delete", width=5, command=lambda: self.delete_history()).grid(row=row, column=4)
        row += 1
        text = "Delete History will delete all DOV records for the station except for the default records with " \
               "the default settings. After the records are deleted, you can use this screen to fill in the proper " \
               "values."
        Label(self.win.body, text=text, wraplength=300, anchor="w", justify=LEFT). \
            grid(row=row, columnspan=14, sticky="w")
        row += 1
        Label(self.win.body, text="").grid(row=row)  # whitespace

    def buttons_frame(self):
        """ configures the widgets on the bottom of the frame """
        button_submit = Button(self.win.buttons)
        button_apply = Button(self.win.buttons)
        button_back = Button(self.win.buttons)
        button_submit.config(text="Submit", command=lambda: self.apply(True))
        button_apply.config(text="Apply", command=lambda: self.apply(False))
        button_back.config(text="Go Back", command=lambda: MainFrame().start(frame=self.win.topframe))
        if sys.platform == "win32":
            button_submit.config(width=15, anchor="w")
            button_apply.config(width=15, anchor="w")
            button_back.config(width=15, anchor="w")
        else:
            button_submit.config(width=9)
            button_apply.config(width=9)
            button_back.config(width=9)
        button_submit.pack(side=LEFT)
        button_apply.pack(side=LEFT)
        button_back.pack(side=LEFT)
        self.status_update = Label(self.win.buttons, text="", fg="red")
        self.status_update.pack(side=LEFT)

    def applyautofill(self):
        """ automatically sets the values for the entry widgets for all days of DOV times"""
        for i in range(len(self.dovarray)):  # in self.autofill.get():
            self.dovarray[i].set(self.autofill.get())

    def apply(self, goback):
        """ check and enter new dov values into the database """
        if not self.check():
            return
        self.enter_database()
        self.route(goback)

    def check(self):
        """ check the values and return False if there is an error. """
        for i in range(7):
            if not RingTimeChecker(self.dovarray[i].get()).check_numeric():
                messagebox.showerror("Dispatch of Value Error",
                                     "The Dispatch of Value for {} must be a number.".format(self.days[i]),
                                     parent=self.win.body)
                return False
            if not RingTimeChecker(self.dovarray[i].get()).count_decimals_place():
                messagebox.showerror("Dispatch of Value Error",
                                     "The Dispatch of Value for {} must not have more than one "
                                     "decimal.".format(self.days[i]),
                                     parent=self.win.body)
                return False
            if RingTimeChecker(self.dovarray[i].get()).check_for_zeros():
                messagebox.showerror("Dispatch of Value Error",
                                     "The Dispatch of Value for {} must not be empty or zero.".format(self.days[i]),
                                     parent=self.win.body)
                return False
            if not RingTimeChecker(self.dovarray[i].get()).less_than_zero():
                messagebox.showerror("Dispatch of Value Error",
                                     "The Dispatch of Value for {} must not be less than zero.".format(self.days[i]),
                                     parent=self.win.body)
                return
            if not RingTimeChecker(self.dovarray[i].get()).over_24():
                messagebox.showerror("Dispatch of Value Error",
                                     "The Dispatch of Value for {} must not more than 24.".format(self.days[i]),
                                     parent=self.win.body)
                return False
            # return if the number can not be made into a float.
            to_add = RingTimeChecker(self.dovarray[i].get()).make_float()
            if not to_add:
                return False
            self.addarray[i] = "{:.2f}".format(to_add)
        return True

    def enter_database(self):
        """ input/ update records to the dov table of the database """
        for i in range(7):
            time_ = self.addarray[i]  # simplify the time to be updated
            temp = self.checkarray[i].get()
            onrec_time = self.onrecarray[i][0][3]
            onrec_temp = Convert(self.onrecarray[i][0][4]).str_to_bool()
            if self.now_onrecarray[i]:  # if there is already a record for the same day...
                # or the time/temp is different
                if onrec_time != time_ or onrec_temp != temp:
                    sql = "UPDATE dov SET dov_time='%s', temp='%s' " \
                          "WHERE eff_date='%s' AND station='%s' AND day='%s'" % \
                          (time_, temp, projvar.invran_date_week[0],
                           projvar.invran_station, self.day[i])
                    commit(sql)
                    self.update_counter += 1
            else:  # if there is no record for the same day
                if temp:  # if the temp box is checked
                    self.insert_database(i)  # make a new record in the dov table
                else:  # if the temp box is not checked
                    if onrec_temp:  # if the current or earliest record in the database is temporary
                        if self.perm_onrecarray[i][3] != time_:  # if there is a diffence in the time.
                            self.insert_database(i)  # make a new record in the dov table
                    else:  # if the current or earliest record in the database is not temporary.
                        if onrec_time != time_ or onrec_temp != temp:  # if the time/temp is different
                            self.insert_database(i)  # make a new record in the dov table

    def insert_database(self, i):
        """ make a new record in the dov table. """
        time_ = self.addarray[i]  # simplify the time to be updated
        temp = self.checkarray[i].get()
        sql = "INSERT INTO dov (eff_date, station, day, dov_time, temp) " \
              "VALUES('%s', '%s', '%s', '%s', '%s')" % \
              (projvar.invran_date_week[0], projvar.invran_station, self.day[i], time_, temp)
        commit(sql)
        self.insert_counter += 1

    def generate_report(self):
        """ generate a report showing all dispatch of value times for each week with records. """
        history_array = []  # holds an array of seven days for each week where recs exist.
        date_array = []  # holds an array of dates where records exist
        sql = "SELECT DISTINCT eff_date FROM dov WHERE station = '%s' ORDER BY eff_date DESC" % projvar.invran_station
        unique_dates = inquire(sql)
        for date in unique_dates:
            date_array.append(date[0])
            sql = "SELECT * FROM dov WHERE station = '%s' and eff_date = '%s'" % (projvar.invran_station, date[0])
            recs = inquire(sql)
            wk_array = ["", "", "", "", "", "", ""]
            for i in range(len(self.day)):  # once per day of the week - 7 times
                to_add = ""  # initialize with empty string - also serves as a default value
                for rec in recs:  # check each record where effective date and station match
                    if rec[2] == self.day[i]:
                        to_add = rec[3]  # put the time in a holder variable.
                        if rec[4] == "True":  # if temp is true
                            to_add += "*"  # astrick will denote a temporary/one time value.
                wk_array[i] = to_add
            history_array.append(wk_array)
        Reports(self.win.topframe).rpt_dov_history(date_array, history_array)

    def delete_history(self):
        """ this will delete all records for the station in the dov table then will recreate the default records. """
        if not messagebox.askokcancel("Dispatch of Value Settings",
                                      "Are you sure you want to delete the DOV history? All DOV records will "
                                      "be deleted and new default records will be generated. ",
                                      parent=self.win.body):
            return
        sql = "DELETE FROM dov WHERE station='%s'" % projvar.invran_station
        commit(sql)
        DovBase().minimum_recs(projvar.invran_station)
        self.get_onrecs()
        self.get_now_onrecs()
        self.get_premrecs()
        self.set_stringvars()
        msg = "All DOV records deleted - default values reset."
        self.status_update.config(text="{}".format(msg))

    def route(self, goback):
        """ re run screen or return to main screen. """
        if goback:
            MainFrame().start(frame=self.win.topframe)
        else:
            self.get_onrecs()
            self.get_now_onrecs()
            self.get_premrecs()
            self.set_stringvars()
            msg = self.create_msg()
            self.status_update.config(text="{}".format(msg))

    def create_msg(self):
        """ builds the msg for the status update """
        msg = ""
        if self.insert_counter:
            s = ""
            if self.insert_counter > 1:
                s = "s"
            msg += str(self.insert_counter) + " record{} inserted.  ".format(s)
        if self.update_counter:
            s = ""
            if self.update_counter > 1:
                s = "s"
            msg += str(self.update_counter) + " record{} updated.  ".format(s)
        if not self.insert_counter and not self.update_counter:
            msg = "No changes made"
        self.insert_counter = 0  # reset counters
        self.update_counter = 0
        return msg


class Tolerances:
    """
    creates a screen where the user can view, change and reset defaults on tolerances.
    """

    def __init__(self):
        self.win = None
        self.msg = ""

    def tolerances(self, frame):
        """ creatses a screen where the user can change tolerances. """
        self.win = MakeWindow()
        self.win.create(frame)
        # page contents
        sql = "SELECT * FROM tolerances"
        results = inquire(sql)
        ot_own_rt = StringVar(self.win.body)
        ot_tol = StringVar(self.win.body)
        av_tol = StringVar(self.win.body)
        Label(self.win.body, text="Tolerances", font=macadj("bold", "Helvetica 18"), anchor="w") \
            .grid(row=0, column=0, columnspan=4, sticky="w")
        Label(self.win.body, text=" ").grid(row=1, column=0, columnspan=4, sticky="w")
        Label(self.win.body, text="Overtime on own route", width=20, anchor="w") \
            .grid(row=2, column=0, ipady=5, sticky="w")
        Entry(self.win.body, width=5, textvariable=ot_own_rt).grid(row=2, column=1, padx=4)
        Button(self.win.body, width=5, text="change",
               command=lambda: self.apply_tolerance(ot_own_rt.get(), "ot_own_rt")) \
            .grid(row=2, column=2, padx=4)
        Button(self.win.body, width=5, text="info",
               command=lambda: Messenger(self.win.topframe).tolerance_info("OT_own_route")) \
            .grid(row=2, column=3, padx=4)
        Label(self.win.body, text="Overtime off own route").grid(row=3, column=0, ipady=5, sticky="w")
        Entry(self.win.body, width=5, textvariable=ot_tol).grid(row=3, column=1)
        Button(self.win.body, width=5, text="change",
               command=lambda: self.apply_tolerance(ot_tol.get(), "ot_tol")) \
            .grid(row=3, column=2)
        Button(self.win.body, width=5, text="info",
               command=lambda: Messenger(self.win.topframe).tolerance_info("OT_off_route")) \
            .grid(row=3, column=3)
        Label(self.win.body, text="Availability tolerance").grid(row=4, column=0, ipady=5, sticky="w")
        Entry(self.win.body, width=5, textvariable=av_tol).grid(row=4, column=1)
        Button(self.win.body, width=5, text="change",
               command=lambda: self.apply_tolerance(av_tol.get(), "av_tol")) \
            .grid(row=4, column=2)
        Button(self.win.body, width=5, text="info",
               command=lambda: Messenger(self.win.topframe).tolerance_info("availability")) \
            .grid(row=4, column=3)
        dashes = ""
        dashcount = 59
        if sys.platform == "darwin":
            dashcount = 47
        for _ in range(dashcount):
            dashes += "_"
        Label(self.win.body, text=dashes, pady=5, fg="blue").grid(row=5, columnspan=4, sticky="w")
        Label(self.win.body, text="Recommended settings").grid(row=6, column=0, ipady=5, sticky="w")
        Button(self.win.body, width=5, text="set",
               command=lambda: self.tolerance_presets("default")).grid(row=6, column=2)
        Label(self.win.body, text="Set tolerances to zero").grid(row=7, column=0, ipady=5, sticky="w")
        Button(self.win.body, width=5, text="set",
               command=lambda: self.tolerance_presets("zero")).grid(row=7, column=2)
        ot_own_rt.set(results[0][2])
        ot_tol.set(results[1][2])
        av_tol.set(results[2][2])
        # the bottom button
        gobackbutton = Button(self.win.buttons, text="Go Back", width=20,
                              command=lambda: MainFrame().start(frame=self.win.topframe))
        if sys.platform == "win32":
            gobackbutton.config(anchor="w")
        gobackbutton.pack(side=LEFT)
        Label(self.win.buttons, text=self.msg, fg="red").pack(side=LEFT)
        self.win.finish()

    def apply_tolerance(self, tolerance, tolerance_type):
        """ checks tolerances. """
        "ot_own_rt"
        "ot_tol"
        "av_tol"
        tol_dict = {"ot_own_rt": "overtime on own route", "ot_tol": "non-otdl overtime",
                    "av_tol": "otdl/aux availability"}
        if not isfloat(tolerance):
            text = "You must enter a number."
            messagebox.showerror("Tolerance value entry error", text, parent=self.win.body)
            return
        if tolerance.strip() == "":
            text = "You must enter a numeric value for tolerances"
            messagebox.showerror("Tolerance value entry error", text, parent=self.win.body)
            return
        if float(tolerance) < 0:
            text = "Values must be equal to or greater than zero."
            messagebox.showerror("Tolerance value entry error", text, parent=self.win.body)
            return
        if float(tolerance) > 1:
            text = "You must enter a value less than one."
            messagebox.showerror("Tolerance value entry error", text, parent=self.win.body)
            return
        if float(tolerance) < 1:
            number = tolerance.split('.')
            if len(number) == 2:
                if len(number[1]) > 2:
                    text = "Value cannot exceed two decimal places."
                    messagebox.showerror("Tolerance value entry error", text, parent=self.win.body)
            else:
                if len(number[0]) > 2:
                    text = "Value cannot exceed two decimal places."
                    messagebox.showerror("Tolerance value entry error", text, parent=self.win.body)
        sql = "UPDATE tolerances SET tolerance ='%s' WHERE category = '%s'" % (tolerance, tolerance_type)
        commit(sql)
        self.msg = "Tolerance for {} has been updated to {}.".format(tol_dict[tolerance_type], tolerance)
        self.tolerances(self.win.topframe)

    def tolerance_presets(self, order):
        """ defines defaults for tolerances. """
        num = None
        if order == "default":
            num = ".25"
            self.msg = "Default tolerance settings have been restored."
        if order == "zero":
            num = "0"
            self.msg = "No tolerances. All values have been set to zero."
        types = ("ot_own_rt", "ot_tol", "av_tol")
        for t in types:
            sql = "UPDATE tolerances SET tolerance ='%s' WHERE category = '%s'" % (num, t)
            commit(sql)

        self.tolerances(self.win.topframe)


class SpreadsheetConfig:
    """
    creates a window that allows the user to adjust the settings for spreadsheets.
    """

    def __init__(self):
        self.frame = None
        self.win = None
        self.minrows_limit = 100  # hardcoded limit of min rows
        self.min_nl = 0.0
        self.min_wal = 0.0
        self.min_otdl = 0.0
        self.min_aux = 0.0
        self.pb_nl_wal = True  # page break between no list and work assignment
        self.pb_wal_otdl = True  # page break between work assignment and otdl
        self.pb_otdl_aux = True  # page break between otdl and auxiliary
        self.min4_nl = 0.0  # minimum rows for mandates no.4
        self.min4_wal = 0.0
        self.min4_otdl = 0.0
        self.min4_aux = 0.0
        self.pb4_nl_wal = True  # page break between no list and work assignment for mandates no.4
        self.pb4_wal_aux = True  # page break between work assignment and otdl for mandates no.4
        self.pb4_aux_otdl = True  # page break between otdl and auxiliary for mandates no.4
        self.man4_dis_limit = None
        self.min_overmax = 0.0
        self.overmax_12hour = None
        self.overmax_wal_dec = None
        self.offbid_distinctpages = None  # off bid distinct page
        self.offbid_maxpivot = None  # off bid maximum pivot
        self.min_ot_equit = None  # minimum rows for ot equitability spreadsheet
        self.ot_calc_pref = None  # overtime calcuations preference for otdl equitability
        self.min_ot_dist = None  # minimum rows for ot distribution spreadsheet
        self.ot_calc_pref_dist = None  # overtime calcuations preference for otdl distribution
        self.min_nl_var = None
        self.min_wal_var = None
        self.min_otdl_var = None
        self.min_aux_var = None
        self.pb_nl_wal_var = None  # page break between no list and work assignment
        self.pb_wal_otdl_var = None  # page break between work assignment and otdl
        self.pb_otdl_aux_var = None  # page break between otdl and auxiliary
        self.min4_nl_var = None  # stringvar for mandates no.4
        self.min4_wal_var = None
        self.min4_otdl_var = None
        self.min4_aux_var = None
        self.pb4_nl_wal_var = None  # page break between no list and work assignment for mandates no.4
        self.pb4_wal_aux_var = None  # page break between work assignment and aux for mandates no.4
        self.pb4_aux_otdl_var = None  # page break between auxiliary and otdl for mandates no.4
        self.man4_dis_limit_var = None  # mandates no.4 display limiter
        self.min_overmax_var = None  # minimum rows for overmax
        self.overmax_12hour_var = None  # 12 and 60 hour option for wal 12 hour daily limit
        self.overmax_wal_dec_var = None  # 12 and 60 hour option for wal dec exemption
        self.offbid_distinctpages_var = None  # off bid spreadsheet: creates distinct pages for each carrier
        self.offbid_maxpivot_var = None  # off bid spreadsheet: maximum pivot
        self.min_otdl_var = None  # minimum rows for ot equitability
        self.min_ot_equit_var = None  # minimum rows for ot equitability spreadsheet
        self.ot_calc_pref_var = None  # overtime calcuations preference for otdl equitability
        self.min_ot_dist_var = None  # minimum rows for ot distribution spreadsheet
        self.ot_calc_pref_dist_var = None  # overtime calcuations preference for otdl distribution
        self.status_update = None  # Label(self.win.buttons, text="", fg="red")
        self.report_counter = 0
        self.check_i = 0  # the iteration of the apply/check method
        self.add_min_nl = 0.0  # prep values to be entered into database
        self.add_min_wal = 0.0
        self.add_min_otdl = 0.0
        self.add_min_aux = 0.0
        self.add_pb_nl_wal = True  # page break between no list and work assignment
        self.add_pb_wal_otdl = True  # page break between work assignment and otdl
        self.add_pb_otdl_aux = True  # page break between otdl and auxiliary
        self.add_min4_nl = 0.0  # prep values to be entered into database for mandates no.4
        self.add_min4_wal = 0.0
        self.add_min4_otdl = 0.0
        self.add_min4_aux = 0.0
        self.add_pb4_nl_wal = True  # page break between no list and work assignment for mandates no.4
        self.add_pb4_wal_aux = True  # page break between work assignment and otdl for mandates no.4
        self.add_pb4_aux_otdl = True  # page break between otdl and auxiliary for mandates no.4
        self.add_man4_dis_limit = None  # mandates no.4 display limiter
        self.add_min_overmax = 0.0
        self.add_offbid_maxpivot = 0.0
        self.add_overmax_12hour = None
        self.add_overmax_wal_dec = None
        self.add_offbid_distinctpages = None
        self.add_min_ot_equit = None
        self.add_ot_calc_pref = None
        self.add_min_ot_dist = None  # minimum rows for ot distribution spreadsheet
        self.add_ot_calc_pref_dist = None  # overtime calcuations preference for otdl distribution

    def start(self, frame):
        """ a master method for controlling other methods. """
        self.frame = frame
        self.win = MakeWindow()
        self.win.create(self.frame)
        self.get_settings()
        self.build_stringvars()
        self.set_stringvars()
        self.build()
        self.buttons_frame()
        self.win.finish()

    def get_settings(self):
        """ gets the current settings from the database. """
        sql = "SELECT tolerance FROM tolerances"
        results = inquire(sql)  # get spreadsheet settings from database
        self.min_nl = results[3][0]
        self.min_wal = results[4][0]
        self.min_otdl = results[5][0]
        self.min_aux = results[6][0]
        self.min_overmax = results[14][0]
        self.overmax_12hour = results[44][0]
        self.overmax_wal_dec = results[45][0]  # work assignment list total violation december exemption
        # get values for off bid assignment spreadsheets
        self.offbid_distinctpages = results[41][0]  # off bid distinct pages
        self.offbid_maxpivot = results[42][0]  # off bid maximum pivot
        self.pb_nl_wal = results[21][0]  # page break between no list and work assignment
        self.pb_wal_otdl = results[22][0]  # page break between work assignment and otdl
        self.pb_otdl_aux = results[23][0]  # page break between otdl and auxiliary
        # get values for mandates no.4 spreadsheets
        self.min4_nl = results[32][0]
        self.min4_wal = results[33][0]
        self.min4_otdl = results[34][0]
        self.min4_aux = results[35][0]
        self.pb4_nl_wal = results[36][0]  # page break between no list and work assignment
        self.pb4_wal_aux = results[37][0]  # page break between work assignment and aux
        self.pb4_aux_otdl = results[38][0]  # page break between auxiliary and otdl
        self.man4_dis_limit = results[39][0]  # mandates no.4 display limiter
        # convert bool to "on" or "off"
        self.pb_nl_wal = Convert(self.pb_nl_wal).strbool_to_onoff()
        self.pb_wal_otdl = Convert(self.pb_wal_otdl).strbool_to_onoff()
        self.pb_otdl_aux = Convert(self.pb_otdl_aux).strbool_to_onoff()
        self.pb4_nl_wal = Convert(self.pb4_nl_wal).strbool_to_onoff()
        self.pb4_wal_aux = Convert(self.pb4_wal_aux).strbool_to_onoff()
        self.pb4_aux_otdl = Convert(self.pb4_aux_otdl).strbool_to_onoff()
        self.overmax_12hour = Convert(self.overmax_12hour).strbool_to_onoff()
        self.overmax_wal_dec = Convert(self.overmax_wal_dec).strbool_to_onoff()
        self.offbid_distinctpages = Convert(self.offbid_distinctpages).strbool_to_onoff()
        # otdl equitability vars
        self.min_ot_equit = results[25][0]  # minimum rows
        self.ot_calc_pref = results[26][0]  # ot calculation preference
        # overtime distribution vars
        self.min_ot_dist = results[27][0]  # minimum rows
        self.ot_calc_pref_dist = results[28][0]  # ot calculations preference

    def build_stringvars(self):
        """ create stringvars """
        self.min_nl_var = StringVar(self.win.body)
        self.min_wal_var = StringVar(self.win.body)
        self.min_otdl_var = StringVar(self.win.body)
        self.min_aux_var = StringVar(self.win.body)
        self.min_overmax_var = StringVar(self.win.body)
        self.overmax_12hour_var = StringVar(self.win.body)
        self.overmax_wal_dec_var = StringVar(self.win.body)
        self.offbid_distinctpages_var = StringVar(self.win.body)
        self.offbid_maxpivot_var = StringVar(self.win.body)
        self.pb_nl_wal_var = StringVar(self.win.body)
        self.pb_wal_otdl_var = StringVar(self.win.body)
        self.pb_otdl_aux_var = StringVar(self.win.body)
        self.min4_nl_var = StringVar(self.win.body)
        self.min4_wal_var = StringVar(self.win.body)
        self.min4_otdl_var = StringVar(self.win.body)
        self.min4_aux_var = StringVar(self.win.body)
        self.pb4_nl_wal_var = StringVar(self.win.body)
        self.pb4_wal_aux_var = StringVar(self.win.body)
        self.pb4_aux_otdl_var = StringVar(self.win.body)
        self.man4_dis_limit_var = StringVar(self.win.body)
        self.min_ot_equit_var = StringVar(self.win.body)
        self.ot_calc_pref_var = StringVar(self.win.body)
        self.min_ot_dist_var = StringVar(self.win.body)
        self.ot_calc_pref_dist_var = StringVar(self.win.body)

    def set_stringvars(self):
        """ set stringvar values """
        self.min_nl_var.set(self.min_nl)
        self.min_wal_var.set(self.min_wal)
        self.min_otdl_var.set(self.min_otdl)
        self.min_aux_var.set(self.min_aux)
        self.min_overmax_var.set(self.min_overmax)
        self.overmax_12hour_var.set(self.overmax_12hour)
        self.overmax_wal_dec_var.set(self.overmax_wal_dec)
        self.offbid_distinctpages_var.set(self.offbid_distinctpages)
        self.offbid_maxpivot_var.set(self.offbid_maxpivot)
        self.pb_nl_wal_var.set(self.pb_nl_wal)
        self.pb_wal_otdl_var.set(self.pb_wal_otdl)
        self.pb_otdl_aux_var.set(self.pb_otdl_aux)
        self.min4_nl_var.set(self.min4_nl)
        self.min4_wal_var.set(self.min4_wal)
        self.min4_otdl_var.set(self.min4_otdl)
        self.min4_aux_var.set(self.min4_aux)
        self.pb4_nl_wal_var.set(self.pb4_nl_wal)
        self.pb4_wal_aux_var.set(self.pb4_wal_aux)
        self.pb4_aux_otdl_var.set(self.pb4_aux_otdl)
        self.man4_dis_limit_var.set(self.man4_dis_limit)
        self.min_ot_equit_var.set(self.min_ot_equit)
        self.ot_calc_pref_var.set(self.ot_calc_pref)
        self.min_ot_dist_var.set(self.min_ot_dist)
        self.ot_calc_pref_dist_var.set(self.ot_calc_pref_dist)

    def build(self):
        """ fills the window with widgets. """
        row = 0
        Label(self.win.body, text="Spreadsheet Settings",
              font=macadj("bold", "Helvetica 18"), anchor="w").grid(row=row, sticky="w", columnspan=14)
        row += 1
        Label(self.win.body, text="").grid(row=row, column=0)
        row += 1
        text = macadj("Improper Mandate Spreadsheets __________________________________________",
                      "Improper Mandate Spreadsheets __________________________________")
        Label(self.win.body, text=text, anchor="w",
              fg="blue").grid(row=row, column=0, columnspan=114, sticky="w")
        row += 1
        Label(self.win.body, text="Minimum rows for No List Carriers", width=macadj(30, 26), anchor="w") \
            .grid(row=row, column=0, ipady=5, sticky="w")
        Entry(self.win.body, width=5, textvariable=self.min_nl_var).grid(row=row, column=1, padx=4, sticky="e")
        Button(self.win.body, width=5, text="info",
               command=lambda: Messenger(self.win.topframe).tolerance_info("min_nl")).grid(row=row, column=2, padx=4)
        row += 1
        Label(self.win.body, text="Minimum rows for Work Assignment", width=macadj(30, 26), anchor="w") \
            .grid(row=row, column=0, ipady=5, sticky="w")
        Entry(self.win.body, width=5, textvariable=self.min_wal_var).grid(row=row, column=1, padx=4, sticky="e")
        Button(self.win.body, width=5, text="info",
               command=lambda: Messenger(self.win.topframe).tolerance_info("min_wal")).grid(row=row, column=2, padx=4)
        row += 1
        Label(self.win.body, text="Minimum rows for OT Desired", width=macadj(30, 26), anchor="w") \
            .grid(row=row, column=0, ipady=5, sticky="w")
        Entry(self.win.body, width=5, textvariable=self.min_otdl_var).grid(row=row, column=1, padx=4, sticky="e")
        Button(self.win.body, width=5, text="info",
               command=lambda: Messenger(self.win.topframe).tolerance_info("min_otdl")).grid(row=row, column=2, padx=4)
        row += 1
        Label(self.win.body, text="Minimum rows for Auxiliary", width=macadj(30, 26), anchor="w") \
            .grid(row=row, column=0, ipady=5, sticky="w")
        Entry(self.win.body, width=5, textvariable=self.min_aux_var).grid(row=row, column=1, padx=4, sticky="e")
        Button(self.win.body, width=5, text="info",
               command=lambda: Messenger(self.win.topframe).tolerance_info("min_aux")).grid(row=row, column=2, padx=4)
        row += 1
        Label(self.win.body, text="").grid(row=row, column=0)
        row += 1
        Label(self.win.body, text="Page Breaks Between List:", anchor="w").grid(row=row, column=0, sticky="w")
        row += 1
        # Page break between no list and work assignment
        Label(self.win.body, text="  No List and Work Assignment", width=macadj(30, 26), anchor="w") \
            .grid(row=row, column=0, ipady=5, sticky="w")
        om_pb_1 = OptionMenu(self.win.body, self.pb_nl_wal_var, "on", "off")
        om_pb_1.config(width=3)
        om_pb_1.grid(row=row, column=1, padx=4, sticky="e")
        Button(self.win.body, width=5, text="info",
               command=lambda: Messenger(self.win.topframe).tolerance_info("pb_nl_wal")) \
            .grid(row=row, column=2, padx=4)
        row += 1
        # Page break between no list and work assignment
        Label(self.win.body, text="  Work Assignment and OT Desired", width=macadj(30, 26), anchor="w") \
            .grid(row=row, column=0, ipady=5, sticky="w")
        om_pb_2 = OptionMenu(self.win.body, self.pb_wal_otdl_var, "on", "off")
        om_pb_2.config(width=3)
        om_pb_2.grid(row=row, column=1, padx=4, sticky="e")
        Button(self.win.body, width=5, text="info",
               command=lambda: Messenger(self.win.topframe).tolerance_info("pb_wal_otdl")) \
            .grid(row=row, column=2, padx=4)
        row += 1
        # Page break between no list and work assignment
        Label(self.win.body, text="  OT Desired and Auxiliary", width=macadj(30, 26), anchor="w") \
            .grid(row=row, column=0, ipady=5, sticky="w")
        om_pb_3 = OptionMenu(self.win.body, self.pb_otdl_aux_var, "on", "off")
        om_pb_3.config(width=3)
        om_pb_3.grid(row=row, column=1, padx=4, sticky="e")
        Button(self.win.body, width=5, text="info",
               command=lambda: Messenger(self.win.topframe).tolerance_info("pb_otdl_aux")) \
            .grid(row=row, column=2, padx=4)
        row += 1
        # Display header for 12 and 60 Hour Violations Spread Sheet
        Label(self.win.body, text="").grid(row=row, column=0)
        row += 1

        text = macadj("Improper Mandate No. 4 Spreadsheets ____________________________________",
                      "Improper Mandate No. 4 Spreadsheets ____________________________")
        Label(self.win.body, text=text, anchor="w",
              fg="blue").grid(row=row, column=0, columnspan=14, sticky="w")
        row += 1
        Label(self.win.body, text="Minimum rows for No List Carriers", width=macadj(30, 26), anchor="w") \
            .grid(row=row, column=0, ipady=5, sticky="w")
        Entry(self.win.body, width=5, textvariable=self.min4_nl_var).grid(row=row, column=1, padx=4, sticky="e")
        Button(self.win.body, width=5, text="info",
               command=lambda: Messenger(self.win.topframe).tolerance_info("min_nl")).grid(row=row, column=2, padx=4)
        row += 1
        Label(self.win.body, text="Minimum rows for Work Assignment", width=macadj(30, 26), anchor="w") \
            .grid(row=row, column=0, ipady=5, sticky="w")
        Entry(self.win.body, width=5, textvariable=self.min4_wal_var).grid(row=row, column=1, padx=4, sticky="e")
        Button(self.win.body, width=5, text="info",
               command=lambda: Messenger(self.win.topframe).tolerance_info("min_wal")).grid(row=row, column=2, padx=4)
        row += 1
        Label(self.win.body, text="Minimum rows for OT Desired", width=macadj(30, 26), anchor="w") \
            .grid(row=row, column=0, ipady=5, sticky="w")
        Entry(self.win.body, width=5, textvariable=self.min4_otdl_var).grid(row=row, column=1, padx=4, sticky="e")
        Button(self.win.body, width=5, text="info",
               command=lambda: Messenger(self.win.topframe).tolerance_info("min_otdl")).grid(row=row, column=2, padx=4)
        row += 1
        Label(self.win.body, text="Minimum rows for Auxiliary", width=macadj(30, 26), anchor="w") \
            .grid(row=row, column=0, ipady=5, sticky="w")
        Entry(self.win.body, width=5, textvariable=self.min4_aux_var).grid(row=row, column=1, padx=4, sticky="e")
        Button(self.win.body, width=5, text="info",
               command=lambda: Messenger(self.win.topframe).tolerance_info("min_aux")).grid(row=row, column=2, padx=4)
        row += 1
        Label(self.win.body, text="").grid(row=row, column=0)
        row += 1
        Label(self.win.body, text="Page Breaks Between List:", anchor="w").grid(row=row, column=0, sticky="w")
        row += 1
        # Page break between no list and work assignment
        Label(self.win.body, text="  No List and Work Assignment", width=macadj(30, 26), anchor="w") \
            .grid(row=row, column=0, ipady=5, sticky="w")
        om_pb_1 = OptionMenu(self.win.body, self.pb4_nl_wal_var, "on", "off")
        om_pb_1.config(width=3)
        om_pb_1.grid(row=row, column=1, padx=4, sticky="e")
        Button(self.win.body, width=5, text="info",
               command=lambda: Messenger(self.win.topframe).tolerance_info("pb_nl_wal")) \
            .grid(row=row, column=2, padx=4)
        row += 1
        # Page break between work assignment and auxiliary
        Label(self.win.body, text="  Work Assignment and Auxiliary", width=macadj(30, 26), anchor="w") \
            .grid(row=row, column=0, ipady=5, sticky="w")
        om_pb_2 = OptionMenu(self.win.body, self.pb4_wal_aux_var, "on", "off")
        om_pb_2.config(width=3)
        om_pb_2.grid(row=row, column=1, padx=4, sticky="e")
        Button(self.win.body, width=5, text="info",
               command=lambda: Messenger(self.win.topframe).tolerance_info("pb_wal_aux")) \
            .grid(row=row, column=2, padx=4)
        row += 1
        # Page break between auxiliary and ot desired
        Label(self.win.body, text=" Auxiliary and OT Desired", width=macadj(30, 26), anchor="w") \
            .grid(row=row, column=0, ipady=5, sticky="w")
        om_pb_3 = OptionMenu(self.win.body, self.pb4_aux_otdl_var, "on", "off")
        om_pb_3.config(width=3)
        om_pb_3.grid(row=row, column=1, padx=4, sticky="e")
        Button(self.win.body, width=5, text="info",
               command=lambda: Messenger(self.win.topframe).tolerance_info("pb_aux_otdl")) \
            .grid(row=row, column=2, padx=4)
        row += 1
        Label(self.win.body, text="Mandates No. 4 Display Limiter", width=macadj(30, 26), anchor="w") \
            .grid(row=row, column=0, ipady=5, sticky="w")
        # display_limiter options: show all, only workdays, only mandates
        om_dis_lim = OptionMenu(self.win.body, self.man4_dis_limit_var, "show all", "only workdays", "only mandates")
        om_dis_lim.config(width=12)
        om_dis_lim.grid(row=row, column=1, padx=4, sticky="e")
        Button(self.win.body, width=5, text="info",
               command=lambda: Messenger(self.win.topframe).tolerance_info("man4_dis_limit")) \
            .grid(row=row, column=2, padx=4)
        row += 1
        Label(self.win.body, text="").grid(row=row, column=0)
        row += 1
        text = macadj("12 and 60 Hour Violations Spreadsheets ___________________________________",
                      "12 and 60 Hour Violations Spreadsheets __________________________")
        Label(self.win.body, text=text, anchor="w",
              fg="blue").grid(row=row, column=0, columnspan=14, sticky="w")
        row += 1
        # Display widgets for 12 and 60 Hour Violations Spread Sheet
        Label(self.win.body, text="Minimum rows for Over Max", width=macadj(30, 26), anchor="w") \
            .grid(row=row, column=0, ipady=5, sticky="w")
        Entry(self.win.body, width=5, textvariable=self.min_overmax_var).grid(row=row, column=1, padx=4, sticky="e")
        Button(self.win.body, width=5, text="info",
               command=lambda: Messenger(self.win.topframe).tolerance_info("min_overmax")) \
            .grid(row=row, column=2, padx=4)
        row += 1
        # Display widget for 12 hour simplification option
        Label(self.win.body, text="WAL 12 Hour Violation", width=macadj(30, 26), anchor="w") \
            .grid(row=row, column=0, ipady=5, sticky="w")
        om_12simple = OptionMenu(self.win.body, self.overmax_12hour_var, "on", "off")
        om_12simple.config(width=7)
        om_12simple.grid(row=row, column=1, padx=4, sticky="e")
        Button(self.win.body, width=5, text="info",
               command=lambda: Messenger(self.win.topframe).tolerance_info("wal_12_hour")) \
            .grid(row=row, column=2, padx=4)
        row += 1

        # Display widget for wal december exemption
        Label(self.win.body, text="WAL December Exemption", width=macadj(30, 26), anchor="w") \
            .grid(row=row, column=0, ipady=5, sticky="w")
        om_wal_dec = OptionMenu(self.win.body, self.overmax_wal_dec_var, "on", "off")
        om_wal_dec.config(width=7)
        om_wal_dec.grid(row=row, column=1, padx=4, sticky="e")
        Button(self.win.body, width=5, text="info",
               command=lambda: Messenger(self.win.topframe).tolerance_info("wal_dec_exempt")) \
            .grid(row=row, column=2, padx=4)
        row += 1
        Label(self.win.body, text="").grid(row=row, column=0)
        row += 1

        text = macadj("Off Bid Assignment Violations Spreadsheets _______________________________",
                      "Off Bid Assignment Violations Spreadsheets ______________________")
        Label(self.win.body, text=text, anchor="w",
              fg="blue").grid(row=row, column=0, columnspan=14, sticky="w")
        row += 1
        Label(self.win.body, text="Distinct Pages for Violations", width=macadj(30, 26), anchor="w") \
            .grid(row=row, column=0, ipady=5, sticky="w")
        om_offbid = OptionMenu(self.win.body, self.offbid_distinctpages_var, "on", "off")
        om_offbid.config(width=7)
        om_offbid.grid(row=row, column=1, padx=4, sticky="e")
        Button(self.win.body, width=5, text="info",
               command=lambda: Messenger(self.win.topframe).tolerance_info("offbid_distinctpage")) \
            .grid(row=row, column=2, padx=4)
        row += 1
        Label(self.win.body, text="Maximum Pivot", width=macadj(30, 26), anchor="w") \
            .grid(row=row, column=0, ipady=5, sticky="w")
        Entry(self.win.body, width=5, textvariable=self.offbid_maxpivot_var).grid(row=row, column=1, padx=4, sticky="e")
        Button(self.win.body, width=5, text="info",
               command=lambda: Messenger(self.win.topframe).tolerance_info("offbid_maxpivot")) \
            .grid(row=row, column=2, padx=4)
        row += 1

        # Display header for OTDL Equitability Spread Sheet
        Label(self.win.body, text="").grid(row=row, column=0)
        row += 1
        text = macadj("OTDL Equitability Spreadsheets ____________________________________________",
                      "OTDL Equitability Spreadsheets __________________________________")
        Label(self.win.body, text=text, anchor="w",
              fg="blue").grid(row=row, column=0, columnspan=14, sticky="w")
        row += 1
        # Display widgets for OTDL Equitability Spread Sheet
        Label(self.win.body, text="Minimum rows for OTDL Equitability", width=macadj(30, 26), anchor="w") \
            .grid(row=row, column=0, ipady=5, sticky="w")
        Entry(self.win.body, width=5, textvariable=self.min_ot_equit_var).grid(row=row, column=1, padx=4, sticky="e")
        Button(self.win.body, width=5, text="info",
               command=lambda: Messenger(self.win.topframe).tolerance_info("min_ot_equit")) \
            .grid(row=row, column=2, padx=4)
        row += 1
        Label(self.win.body, text="Overtime Calculation Preference", width=macadj(30, 26), anchor="w") \
            .grid(row=row, column=0, ipady=5, sticky="w")
        om_ot_equit = OptionMenu(self.win.body, self.ot_calc_pref_var, "all", "off_route")
        om_ot_equit.config(width=7)
        om_ot_equit.grid(row=row, column=1, padx=4, sticky="e")
        Button(self.win.body, width=5, text="info",
               command=lambda: Messenger(self.win.topframe).tolerance_info("ot_calc_pref")) \
            .grid(row=row, column=2, padx=4)
        row += 1
        Label(self.win.body, text="").grid(row=row, column=0)
        row += 1
        # Display header for Overtime Distribution Spread Sheet
        text = macadj("Overtime Distribution Spreadsheets _______________________________________",
                      "Overtime Distribution Spreadsheets _______________________________")
        Label(self.win.body, text=text, anchor="w",
              fg="blue").grid(row=row, column=0, columnspan=14, sticky="w")
        row += 1
        # Display widgets for Overtime Distribution Spread Sheet
        Label(self.win.body, text="Minimum rows for Overtime Distribution", width=macadj(30, 27), anchor="w") \
            .grid(row=row, column=0, ipady=5, sticky="w")
        Entry(self.win.body, width=5, textvariable=self.min_ot_dist_var).grid(row=row, column=1, padx=4, sticky="e")
        Button(self.win.body, width=5, text="info",
               command=lambda: Messenger(self.win.topframe).tolerance_info("min_ot_dist")) \
            .grid(row=row, column=2, padx=4)
        row += 1
        Label(self.win.body, text="Overtime Calculation Preference", width=macadj(30, 26), anchor="w") \
            .grid(row=row, column=0, ipady=5, sticky="w")
        om_ot_equit = OptionMenu(self.win.body, self.ot_calc_pref_dist_var, "all", "off_route")
        om_ot_equit.config(width=7)
        om_ot_equit.grid(row=row, column=1, padx=4, sticky="e")
        Button(self.win.body, width=5, text="info",
               command=lambda: Messenger(self.win.topframe).tolerance_info("ot_calc_pref_dist")) \
            .grid(row=row, column=2, padx=4)
        row += 1
        Label(self.win.body, text="").grid(row=row, column=0)

        row += 1
        dashes = ""
        dashcount = 77
        if sys.platform == "darwin":
            dashcount = 60
        for _ in range(dashcount):
            dashes += "_"
        Label(self.win.body, text=dashes, pady=5, fg="blue").grid(row=row, columnspan=14, sticky="w")
        row += 1
        Label(self.win.body, text="Restore Defaults").grid(row=row, column=0, ipady=5, sticky="w")
        Button(self.win.body, width=5, text="set", command=lambda: self.min_ss_presets("default")) \
            .grid(row=row, column=2)
        row += 1
        Label(self.win.body, text="Set rows to one").grid(row=row, column=0, ipady=5, sticky="w")
        Button(self.win.body, width=5, text="set", command=lambda: self.min_ss_presets("one")) \
            .grid(row=row, column=2)

    def buttons_frame(self):
        """ configures the widgets on the bottom of the frame """
        button_submit = Button(self.win.buttons)
        button_apply = Button(self.win.buttons)
        button_back = Button(self.win.buttons)
        button_submit.config(text="Submit", command=lambda: self.apply(True))
        button_apply.config(text="Apply", command=lambda: self.apply(False))
        button_back.config(text="Go Back", command=lambda: MainFrame().start(frame=self.win.topframe))
        if sys.platform == "win32":
            button_submit.config(width=15, anchor="w")
            button_apply.config(width=15, anchor="w")
            button_back.config(width=15, anchor="w")
        else:
            button_submit.config(width=9)
            button_apply.config(width=9)
            button_back.config(width=9)
        button_submit.pack(side=LEFT)
        button_apply.pack(side=LEFT)
        button_back.pack(side=LEFT)
        self.status_update = Label(self.win.buttons, text="", fg="red")
        self.status_update.pack(side=LEFT)

    def min_ss_presets(self, order):
        """ defines the presets. """
        num = "25"  # default for improper mandates
        num4 = "19"  # default for improper mandates no. 4
        over_num = "30"  # default for over max
        ot_num = "19"  # default for otdl equitability minimum rows
        ot_dist_num = "25"  # default for ot distribution minimum rows
        msg = "Minimum rows reset to default. "
        if order == "one":
            num = "1"
            num4 = "1"
            over_num = "1"
            ot_num = "1"
            ot_dist_num = "1"
            msg = "Minimum rows set to one. "
        self.status_update.config(text="{}".format(msg))
        # set minimum rows for improper mandates
        types = ("min_ss_nl", "min_ss_wal", "min_ss_otdl", "min_ss_aux")
        for t in types:  # set minimum row values for improper mandate spreadsheet
            sql = "UPDATE tolerances SET tolerance ='%s' WHERE category = '%s'" % (num, t)
            commit(sql)
        # set minimum rows for improper mandates no. 4
        types = ("min4_ss_nl", "min4_ss_wal", "min4_ss_otdl", "min4_ss_aux")
        for t in types:  # set minimum row values for improper mandate spreadsheet no.4
            sql = "UPDATE tolerances SET tolerance ='%s' WHERE category = '%s'" % (num4, t)
            commit(sql)
        # set minimum row value for overmax spreadsheet
        sql = "UPDATE tolerances SET tolerance ='%s' WHERE category = '%s'" % (over_num, "min_ss_overmax")
        commit(sql)
        # set minimum row value for otdl equitability
        sql = "UPDATE tolerances SET tolerance ='%s' WHERE category = '%s'" % (ot_num, "min_ot_equit")
        commit(sql)
        # set minimum row value for ot distribution
        sql = "UPDATE tolerances SET tolerance ='%s' WHERE category = '%s'" % (ot_dist_num, "min_ot_dist")
        commit(sql)
        pagebreaks = ("pb_nl_wal", "pb_wal_otdl", "pb_otdl_aux", "pb4_nl_wal", "pb4_wal_aux", "pb4_aux_otdl")
        if order == "default":
            for pb in pagebreaks:
                sql = "UPDATE tolerances SET tolerance ='%s' WHERE category = '%s'" % ("True", pb)
                commit(sql)
            sql = "UPDATE tolerances SET tolerance ='%s' WHERE category = '%s'" % ("off_route", "ot_calc_pref")
            commit(sql)
            sql = "UPDATE tolerances SET tolerance ='%s' WHERE category = '%s'" % ("off_route", "ot_calc_pref_dist")
            commit(sql)
            sql = "UPDATE tolerances SET tolerance ='%s' WHERE category = '%s'" % ("show all", "man4_dis_limit")
            commit(sql)
        self.get_settings()
        self.set_stringvars()

    def check(self, var):
        """ checks entries for minimum rows. """
        current_var = ("No List minimum rows", "Work Assignment minimum rows", "OT Desired minimum rows",
                       "Auxiliary minimum rows", "No List minimum rows no.4", "Work Assignment minimum rows no.4",
                       "OT Desired minimum rows no.4",
                       "Auxiliary minimum rows no.4", "Over Max minimum rows", "OTDL Equitability minimum rows")
        if MinrowsChecker(var).is_empty():
            return True
        if not MinrowsChecker(var).is_numeric():
            text = "The value must be a number for {}".format(current_var[self.check_i])
            messagebox.showerror("Minimum Row Value Entry Error", text, parent=self.win.body)
            return False
        if not MinrowsChecker(var).no_decimals():
            text = "Numbers with decimals are not allowed for {}".format(current_var[self.check_i])
            messagebox.showerror("Minimum Row Value Entry Error", text, parent=self.win.body)
            return False
        if not MinrowsChecker(var).not_negative():
            text = "Numbers less than zero are not allowed for {}".format(current_var[self.check_i])
            messagebox.showerror("Minimum Row Value Entry Error", text, parent=self.win.body)
            return False
        if not MinrowsChecker(var).not_zero():
            text = "Numbers less than one are not allowed for {}".format(current_var[self.check_i])
            messagebox.showerror("Minimum Row Value Entry Error", text, parent=self.win.body)
            return False
        if not MinrowsChecker(var).within_limit(self.minrows_limit):
            text = "Numbers greater than {} are not allowed for {}" \
                .format(self.minrows_limit, current_var[self.check_i])
            messagebox.showerror("Minimum Row Value Entry Error", text, parent=self.win.body)
            return False
        return True

    def check_float(self, var):
        """ checks entries for floating values. """
        current_var = ("Off bid maximum pivot",)
        if RingTimeChecker(var).check_for_zeros():  # skip all checks if the value is zero or empty
            return True
        if not RingTimeChecker(var).check_numeric():
            text = "{} must be a number".format(current_var[self.check_i])
            messagebox.showerror("Data Entry Error", text, parent=self.win.body)
            return False
        if not RingTimeChecker(var).over_8():
            text = "{} must be an 8 or zero or a number in between.".format(current_var[self.check_i])
            messagebox.showerror("Data Entry Error", text, parent=self.win.body)
            return False
        if not RingTimeChecker(var).less_than_zero():
            text = "{} must be an 8 or zero or a number in between.".format(current_var[self.check_i])
            messagebox.showerror("Data Entry Error", text, parent=self.win.body)
            return False
        if not RingTimeChecker(var).count_decimals_place():
            text = "The Make up value for {} can not have more than two decimal places." \
                .format(current_var[self.check_i])
            messagebox.showerror("Data Entry Error", text, parent=self.win.body)
            return False
        return True

    def apply(self, go_home):
        """ checks and updates spreadsheet settings """
        onrecs_min = (self.min_nl, self.min_wal, self.min_otdl, self.min_aux,
                      self.min4_nl, self.min4_wal, self.min4_otdl, self.min4_aux,
                      self.min_overmax, self.min_ot_equit,
                      self.min_ot_dist)
        onrecs_float = (self.offbid_maxpivot,)
        # current records for page breaks and distinct pages options
        onrecs_breaks = (self.pb_nl_wal, self.pb_wal_otdl, self.pb_otdl_aux,
                         self.pb4_nl_wal, self.pb4_wal_aux, self.pb4_aux_otdl,
                         self.overmax_12hour, self.overmax_wal_dec,
                         self.offbid_distinctpages)
        onrecs_misc = (self.man4_dis_limit, self.ot_calc_pref, self.ot_calc_pref_dist)
        check_these = (self.min_nl_var.get(), self.min_wal_var.get(), self.min_otdl_var.get(), self.min_aux_var.get(),
                       self.min4_nl_var.get(), self.min4_wal_var.get(), self.min4_otdl_var.get(),
                       self.min4_aux_var.get(), self.min_overmax_var.get(), self.min_ot_equit_var.get(),
                       self.min_ot_dist_var.get())
        check_float = (self.offbid_maxpivot_var.get(),)
        add_these = [self.add_min_nl, self.add_min_wal, self.add_min_otdl, self.add_min_aux,
                     self.add_min4_nl, self.add_min4_wal, self.add_min4_otdl, self.add_min4_aux,
                     self.add_min_overmax, self.add_min_ot_equit, self.add_min_ot_dist]
        add_float = [self.add_offbid_maxpivot, ]
        categories = ("min_ss_nl", "min_ss_wal", "min_ss_otdl", "min_ss_aux",
                      "min4_ss_nl", "min4_ss_wal", "min4_ss_otdl", "min4_ss_aux",
                      "min_ss_overmax", "min_ot_equit", "min_ot_dist")
        float_categories = ("offbid_maxpivot",)
        # page breaks and distinct pages option menu items
        pbs = (self.pb_nl_wal_var.get(), self.pb_wal_otdl_var.get(), self.pb_otdl_aux_var.get(),
               self.pb4_nl_wal_var.get(), self.pb4_wal_aux_var.get(), self.pb4_aux_otdl_var.get(),
               self.overmax_12hour_var.get(), self.overmax_wal_dec_var.get(), self.offbid_distinctpages_var.get())
        add_pbs = [self.add_pb_nl_wal, self.add_pb_wal_otdl, self.add_pb_otdl_aux,
                   self.add_pb4_nl_wal, self.add_pb4_wal_aux, self.add_pb4_aux_otdl,
                   self.add_overmax_12hour, self.add_overmax_wal_dec, self.add_offbid_distinctpages]
        # the settings as they are named in the tolerances table of the database
        pb_categories = ("pb_nl_wal", "pb_wal_otdl", "pb_otdl_aux",
                         "pb4_nl_wal", "pb4_wal_aux", "pb4_aux_otdl",
                         "wal_12_hour", "wal_dec_exempt", "offbid_distinctpage")
        # misc stringvars
        misc = (self.man4_dis_limit_var.get(), self.ot_calc_pref_var.get(), self.ot_calc_pref_dist_var.get())
        # misc values to update to database
        add_misc = [self.add_man4_dis_limit, self.add_ot_calc_pref, self.add_ot_calc_pref_dist]
        # list of records in the tolerance table.
        misc_categories = ("man4_dis_limit", "ot_calc_pref", "ot_calc_pref_dist")
        self.check_i = 0
        for var in check_these:  # check each of the minimum rows stringvars
            if not self.check(var):  # if any fail
                return  # stop the method
            self.check_i += 1
        self.check_i = 0
        for var in check_float:  # check each of the float value stringvars - off route max pivot
            if not self.check_float(var):  # if any fail
                return  # stop the method
            self.check_i += 1
        for i in range(len(check_these)):
            add_this = Convert(check_these[i]).zero_not_empty()  # replace empty strings with a zero
            add_these[i] = Handler(add_this).format_str_as_int()  # format the string as an int
            if onrecs_min[i] != add_these[i]:
                sql = "UPDATE tolerances SET tolerance ='%s' WHERE category = '%s'" % (add_these[i], categories[i])
                commit(sql)
                self.report_counter += 1
        for i in range(len(check_float)):
            add_this = Convert(check_float[i]).zero_not_empty()  # replace empty string with a zero
            add_float[i] = Handler(add_this).format_str_as_float()  # format the string as a float
            if onrecs_float[i] != add_float[i]:
                sql = "UPDATE tolerances SET tolerance ='%s' WHERE category = '%s'" \
                      % (add_float[i], float_categories[i])
                commit(sql)
                self.report_counter += 1
        for i in range(len(pbs)):  # loop through pagebreak stringvars
            add_pbs[i] = Convert(pbs[i]).onoff_to_bool()
            if onrecs_breaks[i] != str(pbs[i]):
                sql = "UPDATE tolerances SET tolerance ='%s' WHERE category = '%s'" % (add_pbs[i], pb_categories[i])
                commit(sql)
                self.report_counter += 1
        for i in range(len(misc)):  # loop through misc/otdl calculation preferences stringvar
            add_misc[i] = str(misc[i])
            if onrecs_misc[i] != str(misc[i]):
                sql = "UPDATE tolerances SET tolerance ='%s' WHERE category = '%s'" % (add_misc[i], misc_categories[i])
                commit(sql)
                self.report_counter += 1
        if go_home:
            MainFrame().start(frame=self.win.topframe)
        else:
            self.write_report()
            self.get_settings()
            self.set_stringvars()

    def write_report(self):
        """ changes the status at the bottom of the screen. """
        text = "No Records Updated"
        if self.report_counter:
            text = "{} Record{} Updated" \
                .format(self.report_counter, Handler(self.report_counter).plurals())
        self.status_update.config(text=text)
        self.report_counter = 0


class NsConfig:
    """
    creates a screen that allows the user to view and customize ns day settings in a manner that is appropiate
    to a station or branch.
    """

    def __init__(self):
        self.win = None

    def ns_config(self, frame):
        """ generate Non-Scheduled Day Configurations page to configure ns day settings """
        if projvar.invran_day is None:
            messagebox.showerror("Non-Scheduled Day Configurations",
                                 "You must set the Investigation Range before changing the NS Day Configurations.",
                                 parent=frame)
            return
        sql = "SELECT * FROM ns_configuration"
        result = inquire(sql)
        self.win = MakeWindow()
        self.win.create(frame)
        Label(self.win.body, text="Non-Scheduled Day Configurations", font=macadj("bold", "Helvetica 18"), anchor="w") \
            .grid(row=0, sticky="w", columnspan=4)
        Label(self.win.body, text=" ").grid(row=1, column=0)
        text = macadj("Change Configuration ________________________________________________",
                      "Change Configuration _________________________________________________")
        Label(self.win.body, text=text, anchor="w",
              fg="blue").grid(row=2, column=0, columnspan=4, sticky="w")
        # Label(self.win.body, text="Change Configuration").grid(row=2, sticky="w", columnspan=4)
        f_date = projvar.invran_date_week[0].strftime("%a - %b %d, %Y")
        end_f_date = projvar.invran_date_week[6].strftime("%a - %b %d, %Y")
        Label(self.win.body, text="Investigation Range: {0} through {1}".format(f_date, end_f_date),
              foreground="red").grid(row=3, column=0, sticky="w", columnspan=4)
        Label(self.win.body, text="Pay Period: {0}".format(projvar.pay_period),
              foreground="red").grid(row=4, column=0, sticky="w", columnspan=4)
        Label(self.win.body, text=" ").grid(row=5, column=0, sticky="w", columnspan=4)
        Label(self.win.body, text="Day", foreground="grey").grid(row=6, column=0, sticky="w")  # column headers
        Label(self.win.body, text="Name", foreground="grey").grid(row=6, column=1, sticky="w")
        Label(self.win.body, text="Color", foreground="grey").grid(row=6, column=2, sticky="w")
        Label(self.win.body, text="Default", foreground="grey").grid(row=6, column=3, sticky="w")
        yellow_text = StringVar(self.win.body)  # declare variables
        blue_text = StringVar(self.win.body)
        green_text = StringVar(self.win.body)
        brown_text = StringVar(self.win.body)
        red_text = StringVar(self.win.body)
        black_text = StringVar(self.win.body)
        text_array = [yellow_text, blue_text, green_text, brown_text, red_text, black_text]
        color_array = (
            "black", "blue", "brown", "brown4", "dark green", "deep pink", "forest green", "gold", "gray10", "green",
            "navy", "orange", "purple", "red", "red3", "saddle brown", "yellow", "yellow2")
        yellow_color = StringVar(self.win.body)
        blue_color = StringVar(self.win.body)
        green_color = StringVar(self.win.body)
        brown_color = StringVar(self.win.body)
        red_color = StringVar(self.win.body)
        black_color = StringVar(self.win.body)
        fill_array = [yellow_color, blue_color, green_color, brown_color, red_color, black_color]
        Label(self.win.body, text="{}".format(projvar.ns_code['yellow'])) \
            .grid(row=7, column=0, sticky="w")  # yellow row
        Entry(self.win.body, textvariable=yellow_text, width=10).grid(row=7, column=1, sticky="w")
        yellow_text.set(result[0][2])
        om_yellow = OptionMenu(self.win.body, yellow_color, *color_array)
        yellow_color.set(result[0][1])
        om_yellow.config(width=13, anchor=macadj("w", "center"))
        om_yellow.grid(row=7, column=2, sticky="w")
        Label(self.win.body, text="yellow").grid(row=7, column=3, sticky="w")
        Label(self.win.body, text="{}".format(projvar.ns_code['blue'])).grid(row=8, column=0, sticky="w")  # blue row
        Entry(self.win.body, textvariable=blue_text, width=10).grid(row=8, column=1, sticky="w")
        blue_text.set(result[1][2])
        om_blue = OptionMenu(self.win.body, blue_color, *color_array)
        blue_color.set(result[1][1])
        om_blue.config(width=13, anchor=macadj("w", "center"))
        om_blue.grid(row=8, column=2, sticky="w")
        Label(self.win.body, text="blue").grid(row=8, column=3, sticky="w")
        Label(self.win.body, text="{}".format(projvar.ns_code['green'])).grid(row=9, column=0, sticky="w")  # green row
        Entry(self.win.body, textvariable=green_text, width=10).grid(row=9, column=1, sticky="w")
        green_text.set(result[2][2])
        om_green = OptionMenu(self.win.body, green_color, *color_array)
        green_color.set(result[2][1])
        om_green.config(width=13, anchor=macadj("w", "center"))
        om_green.grid(row=9, column=2, sticky="w")
        Label(self.win.body, text="green").grid(row=9, column=3, sticky="w")
        Label(self.win.body, text="{}".format(projvar.ns_code['brown'])).grid(row=10, column=0, sticky="w")  # brown row
        Entry(self.win.body, textvariable=brown_text, width=10).grid(row=10, column=1, sticky="w")
        brown_text.set(result[3][2])
        om_brown = OptionMenu(self.win.body, brown_color, *color_array)
        brown_color.set(result[3][1])
        om_brown.config(width=13, anchor=macadj("w", "center"))
        om_brown.grid(row=10, column=2, sticky="w")
        Label(self.win.body, text="brown").grid(row=10, column=3, sticky="w")
        Label(self.win.body, text="{}".format(projvar.ns_code['red'])).grid(row=11, column=0, sticky="w")  # red row
        Entry(self.win.body, textvariable=red_text, width=10).grid(row=11, column=1, sticky="w")
        red_text.set(result[4][2])
        om_red = OptionMenu(self.win.body, red_color, *color_array)
        red_color.set(result[4][1])
        om_red.config(width=13, anchor=macadj("w", "center"))
        om_red.grid(row=11, column=2, sticky="w")
        Label(self.win.body, text="red").grid(row=11, column=3, sticky="w")
        Label(self.win.body, text="{}".format(projvar.ns_code['black'])).grid(row=12, column=0, sticky="w")  # black row
        Entry(self.win.body, textvariable=black_text, width=10).grid(row=12, column=1, sticky="w")
        black_text.set(result[5][2])
        om_black = OptionMenu(self.win.body, black_color, *color_array)
        black_color.set(result[5][1])
        om_black.config(width=13, anchor=macadj("w", "center"))
        om_black.grid(row=12, column=2, sticky="w")
        Label(self.win.body, text="black").grid(row=12, column=3, sticky="w")
        Label(self.win.body, text=" ").grid(row=13)
        Button(self.win.body, text="set", width=10, command=lambda: self.ns_config_apply(text_array, fill_array)) \
            .grid(row=14, column=3)
        Label(self.win.body, text=" ").grid(row=15)
        text = macadj("Restore Defaults ______________________________________________________",
                      "Restore Defaults ____________________________________________________")
        Label(self.win.body, text=text, anchor="w",
              fg="blue").grid(row=16, column=0, columnspan=4, sticky="w")
        # Label(self.win.body, text="Restore Defaults").grid(row=16)
        Button(self.win.body, text="reset", width=10, command=lambda: self.ns_config_reset()).grid(row=17, column=3)
        button_back = Button(self.win.buttons)
        button_back.config(text="Go Back", width=20, command=lambda: MainFrame().start(frame=self.win.topframe))
        if sys.platform == "win32":
            button_back.config(anchor="w")
        button_back.pack(side=LEFT)
        self.win.finish()

    def ns_config_apply(self, text_array, color_array):
        """ set ns configurations from Non-Scheduled Day Configurations page """
        for t in text_array:
            if len(t.get()) > 6:
                messagebox.showerror("Non_Scheduled Day Configuration",
                                     "Names must not be longer than 6 characters.",
                                     parent=self.win.body)
                return
            if len(t.get()) < 1:
                messagebox.showerror("Non_Scheduled Day Configuration",
                                     "Names must not be shorter than 1 character.",
                                     parent=self.win.body)
                return
        color = ("yellow", "blue", "green", "brown", "red", "black")
        for i in range(6):
            sql = "UPDATE ns_configuration SET custom_name ='%s' WHERE ns_name = '%s'" % (text_array[i].get(), color[i])
            commit(sql)
            sql = "UPDATE ns_configuration SET fill_color ='%s' WHERE ns_name = '%s'" % (color_array[i].get(), color[i])
            commit(sql)
        self.ns_config(self.win.topframe)

    def ns_config_reset(self):
        """ reset ns day configurations from Non-Scheduled Day Configurations page """
        fill = ("gold", "navy", "forest green", "saddle brown", "red3", "gray10")
        color = ("yellow", "blue", "green", "brown", "red", "black")
        for i in range(6):
            sql = "UPDATE ns_configuration SET custom_name ='%s' WHERE ns_name = '%s'" % (color[i], color[i])
            commit(sql)
            sql = "UPDATE ns_configuration SET fill_color ='%s' WHERE ns_name = '%s'" % (fill[i], color[i])
            commit(sql)
        self.ns_config(self.win.topframe)


class SpeedConfig:
    """
    builds a screen that allows the user to configure Speedsheets.
    """

    def __init__(self, frame):
        self.frame = frame
        self.win = MakeWindow()
        self.ns_mode = StringVar(self.win.body)  # create stringvars
        self.fullreport = StringVar(self.win.body)
        self.movesnotation = StringVar(self.win.body)
        self.abc_breakdown = StringVar(self.win.body)
        self.min_empid = StringVar(self.win.body)
        self.min_alpha = StringVar(self.win.body)
        self.min_abc = StringVar(self.win.body)
        self.status_update = Label(self.win.buttons, text="", fg="red")

    def create(self):
        """ builds the widgets that fill the page. """
        self.win.create(self.frame)
        Label(self.win.body, text="SpeedSheet Configurations", font=macadj("bold", "Helvetica 18"), anchor="w") \
            .grid(row=0, sticky="w", columnspan=4)

        Label(self.win.body, text=" ").grid(row=1, column=0)
        self.set_stringvars()

        Label(self.win.body, text="NS Day Preferred Mode: ", width=macadj(40, 30), anchor="w") \
            .grid(row=2, column=0, ipady=5, sticky="w")
        ns_pref = OptionMenu(self.win.body, self.ns_mode, "rotating", "fixed")
        ns_pref.config(width=macadj(9, 9))
        if sys.platform == "win32":
            ns_pref.config(anchor="w")
        ns_pref.grid(row=2, column=1, columnspan=2, sticky="w", padx=4)
        Button(self.win.body, width=5, text="change",
               command=lambda: self.apply_ns_mode()).grid(row=2, column=3, padx=4)

        Label(self.win.body, text="Show Full Report: ", width=macadj(40, 30), anchor="w") \
            .grid(row=3, column=0, ipady=5, sticky="w")
        fullrpt_pref = OptionMenu(self.win.body, self.fullreport, "True", "False")
        fullrpt_pref.config(width=macadj(9, 9))
        if sys.platform == "win32":
            fullrpt_pref.config(anchor="w")
        fullrpt_pref.grid(row=3, column=1, columnspan=2, sticky="w", padx=4)
        Button(self.win.body, width=5, text="change",
               command=lambda: self.apply_fullreport()).grid(row=3, column=3, padx=4)

        Label(self.win.body, text="Moves Notation - route first: ", width=macadj(40, 30), anchor="w") \
            .grid(row=4, column=0, ipady=5, sticky="w")
        moves_pref = OptionMenu(self.win.body, self.movesnotation, "True", "False")
        moves_pref.config(width=macadj(9, 9))
        if sys.platform == "win32":
            moves_pref.config(anchor="w")
        moves_pref.grid(row=4, column=1, columnspan=2, sticky="w", padx=4)
        Button(self.win.body, width=5, text="change",
               command=lambda: self.apply_movenotation()).grid(row=4, column=3, padx=4)

        Label(self.win.body, text="Minimum rows for SpeedSheets", width=macadj(30, 30), anchor="w") \
            .grid(row=5, column=0, ipady=5, sticky="w")
        Label(self.win.body, text="Alphabetical Breakdown (multiple tabs)", width=macadj(40, 30), anchor="w") \
            .grid(row=6, column=0, ipady=5, sticky="w")
        opt_breakdown = OptionMenu(self.win.body, self.abc_breakdown, "True", "False")
        opt_breakdown.config(width=macadj(9, 9))
        if sys.platform == "win32":
            opt_breakdown.config(anchor="w")
        opt_breakdown.grid(row=6, column=1, columnspan=2, sticky="w", padx=4)
        Button(self.win.body, width=5, text="change",
               command=lambda: self.apply_abc_breakdown()).grid(row=6, column=3, padx=4)
        Label(self.win.body, text="Minimum rows for Employee ID tab", width=macadj(40, 30), anchor="w") \
            .grid(row=7, column=0, ipady=5, sticky="w")
        Entry(self.win.body, width=5, textvariable=self.min_empid).grid(row=7, column=1, padx=4)
        Button(self.win.body, width=5, text="change",
               command=lambda: self.apply_min_empid()).grid(row=7, column=2, padx=4)
        Button(self.win.body, width=5, text="info",
               command=lambda: self.info("min_spd_empid")) \
            .grid(row=7, column=3, padx=4)
        Label(self.win.body, text="Minimum rows for Alphabetically tab", width=macadj(40, 30), anchor="w") \
            .grid(row=8, column=0, ipady=5, sticky="w")
        Entry(self.win.body, width=5, textvariable=self.min_alpha).grid(row=8, column=1, padx=4)
        Button(self.win.body, width=5, text="change",
               command=lambda: self.apply_min_alpha()).grid(row=8, column=2, padx=4)
        Button(self.win.body, width=5, text="info",
               command=lambda: self.info("min_spd_alpha")) \
            .grid(row=8, column=3, padx=4)
        Label(self.win.body, text="Minimum rows for Alphabetical breakdown tabs", width=macadj(40, 35), anchor="w") \
            .grid(row=9, column=0, ipady=5, sticky="w")
        Entry(self.win.body, width=5, textvariable=self.min_abc).grid(row=9, column=1, padx=4)
        Button(self.win.body, width=5, text="change",
               command=lambda: self.apply_min_abc()) \
            .grid(row=9, column=2, padx=4)
        Button(self.win.body, width=5, text="info", command=lambda: self.info("min_spd_abc")) \
            .grid(row=9, column=3, padx=4)

        text = macadj("________________________________________________________________________________________",
                      "__________________________________________________________________")
        Label(self.win.body,
              text=text, pady=5, fg="blue").grid(row=10, columnspan=5, sticky="w")
        Label(self.win.body, text="Restore Defaults").grid(row=11, column=0, ipady=5, sticky="w")
        Button(self.win.body, width=5, text="set",
               command=lambda: self.preset_default()).grid(row=11, column=3)
        Label(self.win.body, text="High Settings").grid(row=12, column=0, ipady=5, sticky="w")
        Button(self.win.body, width=5, text="set",
               command=lambda: self.preset_high()).grid(row=12, column=3)
        Label(self.win.body, text="Low Settings").grid(row=13, column=0, ipady=5, sticky="w")
        Button(self.win.body, width=5, text="set",
               command=lambda: self.preset_low()).grid(row=13, column=3)
        self.buttons_frame()

    def buttons_frame(self):
        """ builds the buttons and status update at the bottom of the page. """
        button = Button(self.win.buttons)
        button.config(text="Go Back", width=20, command=lambda: MainFrame().start(frame=self.win.topframe))
        if sys.platform == "win32":
            button.config(anchor="w")
        button.pack(side=LEFT)
        self.status_update.pack(side=LEFT)
        self.win.finish()

    def apply_ns_mode(self):
        """ applies change to ns preference mode. """
        if self.ns_mode.get() == "rotating":
            value = True
        else:
            value = False
        msg = "NS Day Preferred Mode updated: {}".format(self.ns_mode.get())
        self.commit_to_base(value, "speedcell_ns_rotate_mode", msg)

    def apply_fullreport(self):
        """ applies change to full report. """
        if self.fullreport.get() == "True":
            value = True
        else:
            value = False
        msg = "Full report setting updated: {}".format(self.fullreport.get())
        self.commit_to_base(value, "speedsheets_fullreport", msg)

    def apply_movenotation(self):
        """ applies change to full report. """
        if self.movesnotation.get() == "True":
            value = True
        else:
            value = False
        msg = "Move notation setting updated: {}".format(self.movesnotation.get())
        self.commit_to_base(value, "triad_routefirst", msg)

    def apply_abc_breakdown(self):
        """ appplies change to abc breakdown preference - True/False. """
        msg = "Alphabetical Breakdown (multiple tabs) updated: {}".format(self.abc_breakdown.get())
        self.commit_to_base(self.abc_breakdown.get(), "abc_breakdown", msg)

    def apply_min_empid(self):
        """ applies changes to minimum rows for the employee id speedsheet. """
        if self.check(self.min_empid.get()) is None:
            msg = "Minimum rows for Employee ID tab updated: {}".format(self.min_empid.get())
            self.commit_to_base(self.min_empid.get(), "min_spd_empid", msg)

    def apply_min_alpha(self):
        """ applies changes to minimum rows for alphabetical speedsheets. """
        if self.check(self.min_alpha.get()) is None:
            msg = "Minimum rows for Alphabetically tab updated: {}".format(self.min_alpha.get())
            self.commit_to_base(self.min_alpha.get(), "min_spd_alpha", msg)

    def apply_min_abc(self):
        """ applies changes to minimum rows for alphabetical breakdown speedsheets. """
        if self.check(self.min_abc.get()) is None:
            if self.check_abc(self.min_abc.get()) is None:
                msg = "Minimum rows for Alphabetical breakdown tabs updated: {}".format(self.min_abc.get())
                self.commit_to_base(self.min_abc.get(), "min_spd_abc", msg)

    def commit_to_base(self, value, setting, msg):
        """ commits to tolerances table. """
        sql = "UPDATE tolerances SET tolerance ='%s' WHERE category = '%s'" % \
              (value, setting)
        commit(sql)
        self.set_stringvars()
        self.status_update.config(text="{}".format(msg))

    def check(self, value):
        """ check values for minimum rows """
        if not isint(value):
            text = "You must enter a number with no decimals. "
            messagebox.showerror("Tolerance value entry error",
                                 text,
                                 parent=self.win.topframe)
            return False
        if value.strip() == "":
            text = "You must enter a numeric value for tolerances"
            messagebox.showerror("Tolerance value entry error",
                                 text,
                                 parent=self.win.topframe)
            return False
        if float(value) < 5:
            text = "Values must be equal to or greater than five."
            messagebox.showerror("Tolerance value entry error",
                                 text,
                                 parent=self.win.topframe)

            return False
        if float(value) > 500:
            text = "You must enter a value less five hundred."
            messagebox.showerror("Tolerance value entry error",
                                 text,
                                 parent=self.win.topframe)
            return False

    def check_abc(self, value):
        """ checks the arg to make sure it is less than 50. """
        if float(value) > 50:
            text = "You must enter a value less than fifty."
            messagebox.showerror("Tolerance value entry error",
                                 text,
                                 parent=self.win.topframe)
            return False

    def preset_default(self):
        """ sets the normal defaults. """
        empid = "50"
        alpha = "50"
        abc = "10"
        self.preset_to_base(self, empid, alpha, abc)
        self.status_update.config(text="Default Minimum Row Settings Restored")

    def preset_high(self):
        """ a high setting for defaults. """
        empid = "150"
        alpha = "150"
        abc = "40"
        self.preset_to_base(self, empid, alpha, abc)
        self.status_update.config(text="High Minimum Row Settings Enabled")

    def preset_low(self):
        """ a low setting for defaults. """
        empid = "10"
        alpha = "10"
        abc = "5"
        self.preset_to_base(self, empid, alpha, abc)
        self.status_update.config(text="Low Minimum Row Settings Enabled")

    @staticmethod
    def preset_to_base(self, empid, alpha, abc):
        """ abc breakdown is false in all cases """
        sql = "UPDATE tolerances SET tolerance ='%s' WHERE category = '%s'" % ("False", "abc_breakdown")
        commit(sql)
        sql = "UPDATE tolerances SET tolerance ='%s' WHERE category = '%s'" % (empid, "min_spd_empid")
        commit(sql)
        sql = "UPDATE tolerances SET tolerance ='%s' WHERE category = '%s'" % (alpha, "min_spd_alpha")
        commit(sql)
        sql = "UPDATE tolerances SET tolerance ='%s' WHERE category = '%s'" % (abc, "min_spd_abc")
        commit(sql)
        self.set_stringvars()

    def set_stringvars(self):
        """ gets settings and sets stringvars. """
        setting = SpeedSettings()  # retrieve settings from tolerance table in dbase
        if setting.speedcell_ns_rotate_mode:
            self.ns_mode.set("rotating")
        else:
            self.ns_mode.set("fixed")
        self.fullreport.set(str(setting.speedsheet_fullreport))  # convert to str, else you get a 0 or 1
        self.movesnotation.set(str(setting.triad_routefirst))  # convert to str, else you get a  or 1
        self.abc_breakdown.set(str(setting.abc_breakdown))  # convert to str, else you get a 0 or 1
        self.min_empid.set(setting.min_empid)
        self.min_alpha.set(setting.min_alpha)
        self.min_abc.set(setting.min_abc)

    def info(self, switch):
        """ controls messages to messagebox. """
        text = ""
        if switch == "min_spd_empid":
            text = "Sets the minimum number of rows for the " \
                   "Employee Id tab of the All Inclusive Speedsheet. \n\n" \
                   "Enter a value between 5 and 500"
        if switch == "min_spd_alpha":
            text = "Sets the minimum number of rows for the " \
                   "Alphabetical tab of the All Inclusive Speedsheet. \n\n" \
                   "Enter a value between 5 and 500"
        if switch == "min_spd_abc":
            text = "Sets the minimum number of rows for the " \
                   "Alphabetical breakdown tabs of the All Inclusive Speedsheet. \n\n" \
                   "Enter a value between 5 and 50"
        messagebox.showinfo("SpeedSheet Minimum Rows", text, parent=self.win.topframe)


class AdeSettings:
    """
    allows the user to view, change and customized the Automatic Data Entry (ADE) settings.
    """

    def __init__(self):
        self.win = None

    def start(self, frame):
        """ creates window that allows the user to adjust the settings for the ADE. """
        i = None
        self.win = MakeWindow()
        self.win.create(frame)
        r = 0
        Label(self.win.body, text="Auto Data Entry Settings", font=macadj("bold", "Helvetica 18")) \
            .grid(row=r, column=0, sticky="w", columnspan=14)
        r += 1
        Label(self.win.body, text="").grid(row=r)
        r += 1
        text = macadj("NS Day Structure Preference ________________________________",
                      "NS Day Structure Preference _________________________")
        Label(self.win.body, text=text, anchor="w",
              fg="blue").grid(row=r, column=0, columnspan=14, sticky="w")
        r += 1
        ns_structure = StringVar(self.win.body)
        sql = "SELECT tolerance FROM tolerances WHERE category='%s'" % "ns_auto_pref"
        result = inquire(sql)
        Radiobutton(self.win.body, text="rotation", variable=ns_structure, value="rotation") \
            .grid(row=r, column=1, sticky="e")
        Radiobutton(self.win.body, text="fixed", variable=ns_structure, value="fixed") \
            .grid(row=r, column=2, sticky="w")
        ns_structure.set(result[0][0])
        r += 1
        Button(self.win.body, text="Set", width=5, command=lambda: self.ns_structure(ns_structure)) \
            .grid(row=r, column=3)
        r += 1
        Label(self.win.body, text="").grid(row=r, column=1)
        r += 1
        text = macadj("List of TACS MODS Codes __________________________________",
                      "List of TACS MODS Codes ___________________________")
        Label(self.win.body, text=text, anchor="w",
              fg="blue").grid(row=r, column=0, columnspan=14, sticky="w")
        r += 1
        Label(self.win.body, text="(to exclude from Auto Data Entry moves).") \
            .grid(row=r, column=0, columnspan=14, sticky="w")
        r += 1
        Label(self.win.body, text="code", fg="grey", anchor="w") \
            .grid(row=r, column=0)
        Label(self.win.body, text="description", fg="grey", anchor="w") \
            .grid(row=r, column=1, columnspan=2)
        sql = "SELECT * FROM skippers"
        results = inquire(sql)
        r += 1
        if len(results) > 0:
            for i in range(len(results)):
                Button(self.win.body, text=results[i][0], anchor="w", width=5) \
                    .grid(row=i + r, column=0)  # display code
                Button(self.win.body, text=results[i][1], anchor="w", width=30) \
                    .grid(row=i + r, column=1, columnspan=2)  # display description
                Button(self.win.body, text="delete",
                       command=lambda x=i: self.codes_delete(results[x])) \
                    .grid(row=i + r, column=3)
        else:
            Label(self.win.body, text="No Exceptions Listed.", anchor="w") \
                .grid(row=r, column=0, sticky="w", columnspan=3)
            i = 1
        r += i
        r += 1
        Label(self.win.body, text="").grid(row=r, column=2)
        r += 1
        text = macadj("Add New Code _____________________________________________",
                      "Add New Code ____________________________________")
        Label(self.win.body, text=text, anchor="w",
              fg="blue").grid(row=r, column=0, columnspan=14, sticky="w")
        r += 1
        new_code = StringVar(self.win.body)
        new_descp = StringVar(self.win.body)
        Label(self.win.body, text="code", fg="grey", anchor="w").grid(row=r, column=0)
        Label(self.win.body, text="description", fg="grey", anchor="w").grid(row=r, column=1, columnspan=2)
        r += 1

        Entry(self.win.body, textvariable=new_code, width=macadj(6, 4)).grid(row=r, column=0)  # add new code
        Entry(self.win.body, textvariable=new_descp, width=macadj(35, 27)).grid(row=r, column=1, columnspan=2)
        Button(self.win.body, text="Add", width=5,
               command=lambda: self.codes_add(new_code, new_descp)) \
            .grid(row=r, column=3)
        r += 1
        Label(self.win.body, text="").grid(row=r, column=0)
        r += 1
        Label(self.win.body, text="Restore Defaults").grid(row=r, column=1, columnspan=2, sticky="e")
        Button(self.win.body, text="Set", width=5,
               command=lambda: self.codes_default()).grid(row=r, column=3)
        r += 1
        Label(self.win.body, text="").grid(row=r, column=0)
        r += 1
        button = Button(self.win.buttons)
        button.config(text="Go Back", width=20, command=lambda: MainFrame().start(frame=self.win.topframe))
        if sys.platform == "win32":
            button.config(anchor="w")
        button.pack(side=LEFT)

        self.win.finish()

    def ns_structure(self, ns_structure):
        """ method of updating the ns day preference for the ADE. """
        sql = "UPDATE tolerances SET tolerance='%s'WHERE category='%s'" % (ns_structure.get(), "ns_auto_pref")
        commit(sql)
        messagebox.showinfo("Settings Updated",
                            "Auto Data Entry settings have been updated.",
                            parent=self.win.body)

    def codes_delete(self, to_delete):
        """ method of deleting operation numbers which are ignored by the automatic data entry """
        sql = "DELETE FROM skippers WHERE code='%s'" % to_delete[0]
        commit(sql)
        self.start(self.win.topframe)

    def codes_add(self, code, description):
        """ checks and enters operation codes skipped by ADE. """
        sql = "SELECT code FROM skippers"
        results = inquire(sql)
        existing_codes = []
        for item in results:
            existing_codes.append(item[0])
        prohibited_codes = ('721', '722')
        if code.get() in prohibited_codes:
            messagebox.showerror("Data Entry Error",
                                 "It is prohibited to exclude code {}"
                                 .format(code.get(),
                                         parent=self.win.body))
            return
        if code.get() in existing_codes:
            messagebox.showerror("Data Entry Error",
                                 "This code had already been entered.",
                                 parent=self.win.body)
            return
        if code.get().isdigit() == FALSE:
            messagebox.showerror("Data Entry Error",
                                 "TACS code must contain only numbers.",
                                 parent=self.win.body)
            return
        if len(code.get()) > 3 or len(code.get()) < 3:
            messagebox.showerror("Data Entry Error",
                                 "TACS code must be 3 digits long.",
                                 parent=self.win.body)
            return
        if len(description.get()) > 39:
            messagebox.showerror("Data Enty Error",
                                 "Please limit description to less than 40 characters.",
                                 parent=self.win.body)
            return
        sql = "INSERT INTO skippers(code,description) VALUES('%s','%s')" % (code.get(), description.get())
        commit(sql)
        self.start(self.win.topframe)

    def codes_default(self):
        """ resets the defaults operation codes skipped by ADE. """
        sql = "DELETE FROM skippers"
        commit(sql)
        # put records in the skippers table
        skip_these = [["354", "stand by"], ["613", "stewards time"], ["743", "route maintenance"]]
        for rec in skip_these:
            sql = "INSERT OR IGNORE INTO skippers(code, description) VALUES ('%s','%s')" % (rec[0], rec[1])
            commit(sql)
        self.start(self.win.topframe)


class PdfConvertConfig:
    """
    creates a screen where the user can view and configure the pdf converter
    """

    def __init__(self):
        self.win = None
        self.errorrpt = None
        self.rawrpt = None
        self.txtreader = None
        self.msg = ""

    def start(self, frame):
        """ a screen for updating the pdf converter settings. """
        sql = "SELECT tolerance FROM tolerances WHERE category ='%s'" % "pdf_error_rpt"
        result = inquire(sql)
        self.errorrpt = result[0][0]
        self.win = MakeWindow()
        self.win.create(frame)
        Label(self.win.body, text="PDF Converter Settings", font=macadj("bold", "Helvetica 18"), anchor="w") \
            .grid(row=0, sticky="w", columnspan=4)
        Label(self.win.body, text=" ").grid(row=1, column=0)
        # Label(self.win.body, text="Generate Reports for PDF Converter").grid(row=2, sticky="w", columnspan=4)
        text = macadj("Generate Reports for PDF Converter __________________________",
                      "Generate Reports for PDF Converter _____________________")
        Label(self.win.body, text=text, anchor="w",
              fg="blue").grid(row=2, column=0, columnspan=4, sticky="w")
        # Label(self.win.body, text=" ").grid(row=3, column=0)
        Label(self.win.body, text="Error Report", width=33, anchor="w").grid(row=4, column=0, sticky="w")
        error_selection = StringVar(self.win.body)
        om_error = OptionMenu(self.win.body, error_selection, "on", "off")  # option menu configuration below
        om_error.grid(row=4, column=1)
        error_selection.set(result[0][0])
        sql = "SELECT tolerance FROM tolerances WHERE category ='%s'" % "pdf_raw_rpt"
        result = inquire(sql)
        self.rawrpt = result[0][0]
        Label(self.win.body, text="Raw Output Report", width=15, anchor="w") \
            .grid(row=5, column=0, sticky="w")
        raw_selection = StringVar(self.win.body)
        om_raw = OptionMenu(self.win.body, raw_selection, "on", "off")  # option menu configuration below
        om_raw.grid(row=5, column=1)
        raw_selection.set(result[0][0])
        Label(self.win.body, text=" ").grid(row=6, column=0)
        # allow user to read from a text file to bypass the pdfminer
        text = macadj("Generate Reports from Text File ______________________________",
                      "Generate Reports from Text File ________________________")
        Label(self.win.body, text=text, anchor="w",
              fg="blue").grid(row=7, column=0, columnspan=4, sticky="w")
        Label(self.win.body, text="     (where a text file of pdfminer output has been generated)") \
            .grid(row=8, sticky="w", columnspan=4)
        # Label(self.win.body, text=" ").grid(row=9, column=0)
        sql = "SELECT tolerance FROM tolerances WHERE category ='%s'" % "pdf_text_reader"
        result = inquire(sql)
        self.txtreader = result[0][0]
        Label(self.win.body, text="Read from txt file", width=15, anchor="w").grid(row=10, column=0, sticky="w")
        txt_selection = StringVar(self.win.body)
        om_txt = OptionMenu(self.win.body, txt_selection, "on", "off")
        om_txt.grid(row=10, column=1)  # option menu configuration below
        txt_selection.set(result[0][0])
        Label(self.win.body, text=" ").grid(row=11, column=0)
        if sys.platform == "darwin":  # option menu configuration
            om_error.config(width=5)
            om_raw.config(width=5)
            om_txt.config(width=5)
        else:
            om_error.config(width=5, anchor="w")
            om_raw.config(width=5, anchor="w")
            om_txt.config(width=5, anchor="w")
        button = Button(self.win.buttons)
        button.config(text="Go Back", width=15, command=lambda: MainFrame().start(frame=self.win.topframe))
        if sys.platform == "win32":
            button.config(anchor="w")
        button.pack(side=LEFT)
        button = Button(self.win.buttons)
        button.config(text="Apply", width=15, command=lambda: self.apply(error_selection, raw_selection, txt_selection))
        if sys.platform == "win32":
            button.config(anchor="w")
        button.pack(side=LEFT)
        Label(self.win.buttons, text=self.msg, fg="red").pack(side=LEFT)
        self.win.finish()

    def apply(self, error, raw, txt):
        """ updates the settings for the pdf converter. """
        update_counter = 0
        sql = "UPDATE tolerances SET tolerance='%s'WHERE category='%s'" % (error.get(), "pdf_error_rpt")
        if self.errorrpt != error.get():
            commit(sql)
            update_counter += 1
        sql = "UPDATE tolerances SET tolerance='%s'WHERE category='%s'" % (raw.get(), "pdf_raw_rpt")
        if self.rawrpt != raw.get():
            commit(sql)
            update_counter += 1
        sql = "UPDATE tolerances SET tolerance='%s'WHERE category='%s'" % (txt.get(), "pdf_text_reader")
        if self.txtreader != txt.get():
            commit(sql)
            update_counter += 1
        self.get_msg(update_counter)
        self.start(self.win.topframe)

    def get_msg(self, counter):
        """ update the message on the button bar at the bottom of the screen when records are added."""
        if not counter:
            self.msg = "No Records Updated."
        elif counter == 1:
            self.msg = "One Record Updated."
        elif counter == 2:
            self.msg = "Two Records Updated."
        else:
            self.msg = "All Records Updated. "


class NameIndex:
    """
    This creates a screen the user can use to view carrier names as they appear in tacs and klusterbox as well as the
    employee id numbers which are used by Auto Data Entry and Speedsheets.
    """

    def __init__(self):
        self.win = None

    def name_index_screen(self, frame):
        """ creates a screen which shows all records in the name index. """
        sql = "SELECT * FROM name_index ORDER BY kb_name"
        results = inquire(sql)
        self.win = MakeWindow()
        self.win.create(frame)
        x = 0
        if len(results) == 0:
            Label(self.win.body, text="The Name Index is empty").grid(row=0, column=x)
        else:
            Label(self.win.body, text="Name Index Management", font=macadj("bold", "Helvetica 18")) \
                .grid(row=x, column=0, sticky="w", columnspan=2)  # page header
            x += 1
            Label(self.win.body, text="").grid(row=x, column=0, sticky="w")
            x += 1
            Label(self.win.body, text="TACS Name").grid(row=x, column=1, sticky="w")  # column headers
            Label(self.win.body, text="Klusterbox Name").grid(row=x, column=2, sticky="w")
            Label(self.win.body, text="Emp ID").grid(row=x, column=3, sticky="w")
            x += 1
            for item in results:  # loop for names in the index
                Label(self.win.body, text=str(x - 2), anchor="w").grid(row=x, column=0)
                Button(self.win.body, text=" " + item[0], anchor="w", width=20, relief=RIDGE).grid(row=x, column=1)
                Button(self.win.body, text=" " + item[1], anchor="w", width=20, relief=RIDGE).grid(row=x, column=2)
                Button(self.win.body, text=" " + item[2], anchor="w", width=8, relief=RIDGE).grid(row=x, column=3)
                Button(self.win.body, text="delete", anchor="w", width=5, relief=RIDGE,
                       command=lambda xx=item[2]: self.apply_nameindexer_list(xx)).grid(row=x, column=4)
                x += 1
            Button(self.win.body, text="Delete All", width="15",
                   command=lambda: self.del_all_nameindexer()) \
                .grid(row=x, column=0, columnspan=5, sticky="e")
        Button(self.win.buttons, text="Go Back", width=20,
               command=lambda: MainFrame().start(frame=self.win.topframe)).pack(side=LEFT)
        self.win.finish()

    def apply_nameindexer_list(self, x):
        """ deletes a carrier/record from the name index. """
        sql = "DELETE FROM name_index WHERE emp_id = '%s'" % x
        commit(sql)
        self.name_index_screen(self.win.topframe)

    def del_all_nameindexer(self):
        """ deletes everything from the name index. """
        sql = "DELETE FROM name_index"
        commit(sql)
        self.name_index_screen(self.win.topframe)


class StationIndex:
    """
    creates a screen which the user can use to display, change and delete station and station indexes.
    """

    def __init__(self):
        self.win = None  # the window object
        self.results = None  # search results from all records in station index.
        self.frame = []  # rename function: holds topframe
        self.passframe = []  # rename function: holds the frame name of the station to be renamed
        self.tacs = []  # rename function: holds the tacs name
        self.kb = []  # rename function: holds the klusterbox name of the station to be changed.
        self.newname = []  # rename function: holds the new name of the station.
        self.rename_button = []  # rename function: holds a button widget
        self.all_stations = []  # rename function: holds all the stations in the station list except out of station.

    def reinitialize(self):
        """ re initialize the arrays to empty out any entries previously accumulated. """
        self.results = None  # search results from all records in station index.
        self.frame = []  # rename function: holds topframe
        self.passframe = []  # rename function: holds the frame name of the station to be renamed
        self.tacs = []  # rename function: holds the tacs name
        self.kb = []  # rename function: holds the klusterbox name of the station to be changed.
        self.newname = []  # rename function: holds the new name of the station.
        self.rename_button = []  # rename function: holds a button widget
        self.all_stations = []  # rename function: holds all the stations in the station list except out of station.

    def get_all_stations(self):
        """ this provides a list of stations in the station list, but not in the station index. """
        sql = "SELECT * FROM stations"
        results = inquire(sql)
        for rec in results:
            self.all_stations.append(rec[0])  # get all stations in database.
        sql = "SELECT * FROM station_index"
        self.results = inquire(sql)
        for rec in self.results:
            if rec[1] in self.all_stations:
                self.all_stations.remove(rec[1])  # remove any station in station index
        self.all_stations.remove("out of station")  # remove out of station.

    def station_index_mgmt(self, frame):
        """ creates a screen that allows the user to adjust the station index. """
        self.reinitialize()
        self.get_all_stations()  # provides a list of stations in the station list, but not in the station index.
        self.win = MakeWindow()
        self.win.create(frame)
        self.frame = self.win.topframe  # get the topframe for page reloading.
        g = 0  # a counter for the row
        Label(self.win.body, text="Station Index Management", font=macadj("bold", "Helvetica 18")) \
            .grid(row=g, column=0, sticky="w")
        Label(self.win.body, text="").grid(row=g + 1, column=0)
        g += 2
        if len(self.results) == 0:
            Label(self.win.body, text="There are no stations in the station index") \
                .grid(row=g, column=0, sticky="w")
            g += 1
        else:
            header_frame = Frame(self.win.body, width=500)
            header_frame.grid(row=g, column=0, sticky="w")
            Label(header_frame, text="TACS Station Name", width=macadj(30, 25), anchor="w") \
                .grid(row=0, column=0, sticky="w")
            Label(header_frame, text="Klusterbox Station Name", width=macadj(30, 25), anchor="w") \
                .grid(row=0, column=1, sticky="w")
            g += 1
            f = 0  # initialize number for frame
            frame = []  # initialize array for frame
            for record in self.results:
                self.tacs.append(record[0])
                self.kb.append(record[1])
                to_add = "station_frame" + str(f)  # give the new frame a name
                frame.append(to_add)  # add the frame to the array
                frame[f] = Frame(self.win.body, width=500)  # create the frame widget
                frame[f].grid(row=g, padx=5, sticky="w")  # grid the widget
                self.passframe.append(frame[f])  # use attribute to hold the frame name.
                self.newname.append(StringVar(self.win.topframe))
                Button(frame[f], text=record[0], width=macadj(30, 25), anchor="w").grid(row=0, column=0)
                Button(frame[f], text=record[1], width=macadj(30, 25), anchor="w").grid(row=0, column=1)
                to_add = Button(frame[f], text="rename", width=6)
                self.rename_button.append(to_add)
                self.rename_button[f]['command'] = lambda x=f: self.station_index_rename(x)
                self.rename_button[f].grid(row=0, column=2)
                delete_button = Button(frame[f], text="delete", width=6,
                                       command=lambda x=f: self.station_rec_del(x))
                delete_button.grid(row=0, column=3)
                f += 1
                g += 1
            Label(self.win.body, text="", height=1).grid(row=g)
            Button(self.win.body, text="Delete All", width="15",
                   command=lambda: (self.stationindexer_del_all(self.win.topframe))) \
                .grid(row=g + 1, column=0, columnspan=3, sticky="e")
        button = Button(self.win.buttons)
        button.config(text="Go Back", width=20, command=lambda: MainFrame().start(frame=self.win.topframe))
        if sys.platform == "win32":
            button.config(anchor="w")
        button.pack(side=LEFT)
        self.win.finish()

    def station_index_rename_apply(self, f):
        """ rename a station in the station index. """
        sql = "UPDATE station_index SET kb_station='%s' WHERE tacs_station='%s'" % \
              (self.newname[f].get(), self.tacs[f])
        commit(sql)
        self.station_index_mgmt(self.frame)

    def station_index_rename(self, f):
        """ widgets allow the user to select a new name for the kb station from a the stations list. """
        self.rename_button[f].destroy()
        Button(self.passframe[f], text=" ", width=6).grid(row=0, column=2)
        if len(self.all_stations) > 0:
            Label(self.passframe[f], text="update station name:  ", anchor="e").grid(row=1, column=0, sticky="e")
            # set up station option menu and variable
            om_station = OptionMenu(self.passframe[f], self.newname[f], *self.all_stations)
            om_station.config(width=28, anchor="w")
            om_station.grid(row=1, column=1)
            self.newname[f].set(self.kb[f])
            Button(self.passframe[f], text="rename",
                   command=lambda: self.station_index_rename_apply(f)) \
                .grid(row=1, column=2)
        else:
            Label(self.passframe[f], text="No Unassigned Stations Available") \
                .grid(row=1, column=0, columnspan=2, sticky="e")

    def station_rec_del(self, f):
        """ delete a record from the station index. """
        sql = "DELETE FROM station_index WHERE tacs_station = '%s' and kb_station='%s'" % \
              (self.tacs[f], self.kb[f])
        commit(sql)
        self.station_index_mgmt(self.frame)

    def stationindexer_del_all(self, frame):
        """ deletes everything from the station index. """
        sql = "DELETE FROM station_index"
        commit(sql)
        self.station_index_mgmt(frame)


class AboutKlusterbox:
    """
    a class for displaying the About Klusterbox screen. Will display the version number, release date, contact
    information and source code.
    """

    def __init__(self):
        self.win = None
        self.frame = None
        self.photo = None

    def start(self, frame):
        """ a master method to run other methods in proper order. """
        self.frame = frame
        self.win = MakeWindow()
        self.win.create(self.frame)
        self.build()
        self.button_frame()
        self.win.finish()

    def build(self):
        """ fills the screen with widgets. """
        r = 0  # set row counter
        if projvar.platform == "macapp":
            path_ = os.path.join(os.path.sep, 'Applications', 'klusterbox.app', 'Contents', 'Resources',
                                 'kb_about.jpg')
        elif projvar.platform == "winapp":
            path_ = os.path.join(os.path.sep, os.getcwd(), 'kb_about.jpg')
        else:
            path_ = os.path.join(os.path.sep, os.getcwd(), 'kb_sub', 'kb_images', 'kb_about.jpg')
        try:
            self.photo = ImageTk.PhotoImage(Image.open(path_))
            Label(self.win.body, image=self.photo).grid(row=r, column=0, columnspan=10, sticky="w")
        except (TclError, FileNotFoundError):
            pass
        r += 1
        Label(self.win.body, text="").grid(row=r)
        r += 1
        Label(self.win.body, text="Klusterbox", font=macadj("bold", "Helvetica 18"), fg="red", anchor=W) \
            .grid(row=r, column=0, sticky="w", columnspan=6)
        r += 1
        Label(self.win.body, text="").grid(row=r)
        r += 1
        Label(self.win.body, text="version: {}".format(version), anchor=W) \
            .grid(row=r, column=0, sticky="w", columnspan=6)
        r += 1
        Label(self.win.body, text="release date: {}".format(release_date), anchor=W) \
            .grid(row=r, column=0, sticky="w", columnspan=6)
        r += 1
        Label(self.win.body, text="created by Thomas Weeks", anchor=W).grid(row=r, column=0, sticky="w",
                                                                            columnspan=6)
        r += 1
        Label(self.win.body, text="Original release: October 2018", anchor=W) \
            .grid(row=r, column=0, sticky="w", columnspan=6)
        r += 1
        Label(self.win.body, text=" ", anchor=W).grid(row=r, column=0, sticky="w", columnspan=6)
        r += 1
        Label(self.win.body, text="comments and criticisms are welcome", anchor=W, fg="red") \
            .grid(row=r, column=0, sticky="w", columnspan=6)
        r += 1
        Label(self.win.body, text=" ", anchor=W).grid(row=r, column=0, sticky="w", columnspan=6)
        r += 1
        Label(self.win.body, text="contact information: ", anchor=W).grid(row=r, column=0, sticky="w", columnspan=6)
        r += 1
        Label(self.win.body, text="Thomas Weeks", anchor=W).grid(row=r, column=0, sticky="w", columnspan=6)
        r += 1
        Label(self.win.body, text="    tomweeks@klusterbox.com", anchor=W) \
            .grid(row=r, column=0, sticky="w", columnspan=6)
        r += 1
        Label(self.win.body, text="I've found that some emails get filtered out by the junk folder so", anchor=W) \
            .grid(row=r, column=0, sticky="w", columnspan=6)
        r += 1
        Label(self.win.body, text="Message me on Facebook Messenger:", anchor=W) \
            .grid(row=r, column=0, sticky="w", columnspan=6)
        r += 1
        kb_link = Label(self.win.body, text="    facebook.com/thomas.weeks.artist", fg="blue", cursor="hand2")
        kb_link.grid(row=r, columnspan=6, sticky="w")
        kb_link.bind("<Button-1>", lambda e: self.callback("http://www.facebook.com/thomas.weeks.artist"))
        r += 1
        Label(self.win.body, text="    720.280.0415", anchor=W).grid(row=r, column=0, sticky="w", columnspan=6)
        r += 1
        Label(self.win.body, text="").grid(row=r)
        r += 1
        Label(self.win.body, text="For the lastest updates on Klusterbox check out the official Klusterbox") \
            .grid(row=r, columnspan=6, sticky="w")
        r += 1
        Label(self.win.body, text="website at:").grid(row=r, columnspan=6, sticky="w")
        r += 1
        kb_link = Label(self.win.body, text="    www.klusterbox.com", fg="blue", cursor="hand2")
        kb_link.grid(row=r, columnspan=6, sticky="w")
        kb_link.bind("<Button-1>", lambda e: self.callback("http://klusterbox.com"))
        r += 1
        Label(self.win.body, text="Also look on Facebook for Klusterbox - Software for NALC Stewards at:") \
            .grid(row=r, columnspan=6, sticky="w")
        r += 1
        fb_link = Label(self.win.body, text="    www.facebook.com/klusterbox", fg="blue", cursor="hand2")
        fb_link.grid(row=r, columnspan=6, sticky="w")
        fb_link.bind("<Button-1>", lambda e: self.callback("http://www.facebook.com/klusterbox"))
        r += 1
        Label(self.win.body, text="Like, Follow and Share!").grid(row=r, columnspan=6, sticky="w")
        r += 1
        Label(self.win.body, text="").grid(row=r)
        r += 1
        Label(self.win.body, text="Project Documentation", font=macadj("bold", "Helvetica 16"), anchor=W) \
            .grid(row=r, column=0, sticky="w", columnspan=3)
        Label(self.win.body, text="                                             ").grid(row=r, column=3)
        Label(self.win.body, text="                                             ").grid(row=r, column=4)
        r += 1
        Label(self.win.body, text="").grid(row=r)
        r += 1
        Button(self.win.body, text="read", width=macadj(7, 7), command=lambda: self.open_docs("readme.txt")) \
            .grid(row=r, column=0, sticky="w")
        Label(self.win.body, text="Read Me", anchor=E).grid(row=r, column=1, sticky="w")
        r += 1
        Label(self.win.body, text="").grid(row=r)
        r += 1
        Button(self.win.body, text="read", width=macadj(7, 7), command=lambda: self.open_docs("history.txt")) \
            .grid(row=r, column=0, sticky="w")
        Label(self.win.body, text="History", anchor=E).grid(row=r, column=1, sticky="w")
        r += 1
        Label(self.win.body, text="").grid(row=r)
        r += 1
        Button(self.win.body, text="read", width=macadj(7, 7), command=lambda: self.open_docs("LICENSE.txt")) \
            .grid(row=r, column=0, sticky="w")
        Label(self.win.body, text="License", anchor=E).grid(row=r, column=1, sticky="w")
        r += 1
        Label(self.win.body, text="").grid(row=r)
        r += 1
        """
        Enter all modules imported by klusterbox below as part of the sourcecode tuple. All modules must be in the 
        klusterbox project folder.
        """
        sourcecode = ("klusterbox.py",
                      "projvar.py",
                      "kbtoolbox.py",
                      "kbdatabase.py",
                      "kbreports.py",
                      "kbspreadsheets.py",
                      "kbspeedsheets.py",
                      "kbequitability.py",
                      "kbcsv_repair.py",
                      "kbpdfhandling.py",
                      "kbcsv_reader.py",
                      "kbenterrings.py",
                      "kbfixes.py",
                      "kbinformalc.py"
                      )
        for i in range(len(sourcecode)):
            Button(self.win.body, text="read", width=macadj(7, 7),
                   command=lambda source=sourcecode[i]: self.open_docs(source)).grid(row=r, column=0, sticky="w")
            Label(self.win.body, text="Source Code - {}".format(sourcecode[i]), anchor=E) \
                .grid(row=r, column=1, sticky="w")
            r += 1
            Label(self.win.body, text="").grid(row=r)
            r += 1
        Button(self.win.body, text="read", width=macadj(7, 7), command=lambda: self.open_docs("requirements.txt")) \
            .grid(row=r, column=0, sticky="w")
        Label(self.win.body, text="python requirements", anchor=E).grid(row=r, column=1, sticky="w")

    def button_frame(self):
        """ builds the buttons on the bottom of the screen. """
        button = Button(self.win.buttons)
        button.config(text="Go Back", width=20, command=lambda: MainFrame().start(frame=self.win.topframe))
        if sys.platform == "win32":
            button.config(anchor="w")
        button.pack(side=LEFT)

    def open_docs(self, doc):
        """ opens docs in the about_klusterbox() function """
        try:
            if sys.platform == "win32":
                if projvar.platform == "py":
                    try:
                        path_ = doc
                        os.startfile(path_)  # in IDE the files are in the project folder
                    except FileNotFoundError:
                        path_ = os.path.join(os.path.sep, os.getcwd(), 'kb_sub', doc)
                        os.startfile(path_)  # in KB legacy the files are in the kb_sub folder
                if projvar.platform == "winapp":
                    path_ = os.path.join(os.path.sep, os.getcwd(), doc)
                    os.startfile(path_)
            if sys.platform == "linux":
                subprocess.call(doc)
            if sys.platform == "darwin":
                if projvar.platform == "macapp":
                    path_ = os.path.join(os.path.sep, 'Applications', 'klusterbox.app', 'Contents', 'Resources', doc)
                    subprocess.call(["open", path_])
                if projvar.platform == "py":
                    subprocess.call(["open", doc])
        except FileNotFoundError:
            messagebox.showerror("Project Documents",
                                 "The document was not opened or found.",
                                 parent=self.win.body)

    @staticmethod
    def callback(url):
        """ open hyperlinks at about_klusterbox() """
        open_new(url)


class MassInput:
    """
    creates screen where users can change multiple characteristics for multiple carriers at one time.
    """

    def __init__(self):
        self.win = None
        self.array_var = None
        self.mi_list = None
        self.mi_nsday = None
        self.mi_station = None
        self.mi_route = None
        self.pass_date = None
        self.mi_date = None
        self.mi_sort = None
        self.apply_date = None

    def initialize(self):
        """ initialize all the arrays to empty """
        self.array_var = []
        self.mi_list = []
        self.mi_nsday = []
        self.mi_station = []
        self.mi_route = []
        self.pass_date = IntVar(self.win.body)
        self.mi_date = StringVar(self.win.body)
        self.mi_sort = StringVar(self.win.body)

    def mass_input(self, frame, day, sort):
        """ creates the mass input screen that allows the user to update the list status, ns day or station for
        multiple carries. """
        self.win = MakeWindow()
        self.win.create(frame)
        self.initialize()
        optionframe = Frame(self.win.body)
        optionframe.grid(row=0, columnspan=10, sticky=W)
        # set up the option menus to order results by day and sort criteria.
        opt_day = ["Sat", "Sun", "Mon", "Tue", "Wed", "Thu", "Fri"]
        opt_sort = ["name", "list", "ns day"]
        self.mi_date.set(day)
        if projvar.invran_weekly_span:  # if investigation range is daily
            self.mi_date.set(day)
            om1 = OptionMenu(optionframe, self.mi_date, *opt_day)  # option menu of days
            om1.config(width="5")
            om1.grid(row=0, column=0)
        self.mi_sort.set(sort)
        om2 = OptionMenu(optionframe, self.mi_sort, *opt_sort)  # option menu of list statuses
        om2.grid(row=0, column=1)
        om2.config(width="8")
        Button(optionframe, text="set", width=10,  # button to set
               command=lambda: self.mass_input(self.win.topframe, self.mi_date.get(), self.mi_sort.get())) \
            .grid(row=0, column=2)
        # figure out the day and display
        if projvar.invran_weekly_span:  # if investigation range is weekly
            for i in range(len(projvar.invran_date_week)):
                if opt_day[i] == day:
                    self.apply_date = projvar.invran_date_week[i]  # save the date for the apply method.
                    f_date = projvar.invran_date_week[i].strftime("%a - %b %d, %Y")
                    self.pass_date.set(i)
                    Label(self.win.body, text="Showing results for {}"
                          .format(f_date), font=macadj("bold", "Helvetica 18"), justify=LEFT) \
                        .grid(row=1, column=0, columnspan=10, sticky=W)
        if not projvar.invran_weekly_span:  # if investigation range is daily
            for i in range(len(opt_day)):
                if projvar.invran_date.strftime("%a") == opt_day[i]:
                    self.apply_date = projvar.invran_date  # save the date for the apply method.
                    f_date = projvar.invran_date.strftime("%a - %b %d, %Y")
                    self.pass_date.set(i)
                    Label(self.win.body, text="Showing results for {}"
                          .format(f_date), font=macadj("bold", "Helvetica 18"), justify=LEFT) \
                        .grid(row=1, column=0, columnspan=6, sticky=W)
        # access database
        sql = ""
        for i in range(len(projvar.invran_date_week)):
            if opt_day[i] == day:
                if projvar.invran_weekly_span:  # if investigation range is weekly
                    sql = "SELECT effective_date, carrier_name,list_status, ns_day,route_s, station, rowid" \
                          " FROM carriers WHERE effective_date <= '%s'" \
                          "ORDER BY carrier_name, effective_date" % (projvar.invran_date_week[i])
                else:
                    sql = "SELECT effective_date, carrier_name,list_status, ns_day,route_s, station, rowid" \
                          " FROM carriers WHERE effective_date <= '%s'" \
                          "ORDER BY carrier_name, effective_date" % projvar.invran_date
        results = inquire(sql)
        # initialize arrays for data sorting
        carrier_list = []
        candidates = []
        otdl_array = []
        wal_array = []
        nl_array = []
        ptf_array = []
        aux_array = []
        yellow_array = []
        blue_array = []
        green_array = []
        brown_array = []
        red_array = []
        black_array = []
        none_array = []
        # take raw data and sort into appropiate arrays
        for i in range(len(results)):
            candidates.append(results[i])  # put name into candidates array
            jump = "no"  # triggers an analysis of the candidates array
            if i != len(results) - 1:  # if the loop has not reached the end of the list
                if results[i][1] == results[i + 1][1]:  # if the name current and next name are the same
                    jump = "yes"  # bypasses an analysis of the candidates array
            if jump == "no":
                winner = max(candidates, key=itemgetter(0))  # select the most recent record
                if winner[5] == projvar.invran_station:  # if that record matches the current station...
                    carrier_list.append(winner)  # then insert that record in the carrier list
                    if sort == "list":  # sort carrier list by ot list if selected
                        if winner[2] == "otdl":
                            otdl_array.append(winner)
                        if winner[2] == "wal":
                            wal_array.append(winner)
                        if winner[2] == "nl":
                            nl_array.append(winner)
                        if winner[2] == "ptf":
                            ptf_array.append(winner)
                        if winner[2] == "aux":
                            aux_array.append(winner)
                    if sort == "ns day":  # sort carrier list by ns day if selected
                        if winner[3] == "yellow":
                            yellow_array.append(winner)
                        if winner[3] == "blue":
                            blue_array.append(winner)
                        if winner[3] == "green":
                            green_array.append(winner)
                        if winner[3] == "brown":
                            brown_array.append(winner)
                        if winner[3] == "red":
                            red_array.append(winner)
                        if winner[3] == "black":
                            black_array.append(winner)
                        if winner[3] == "none":
                            none_array.append(winner)
            del candidates[:]
        # Display results XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX
        i = 4
        list_header = ""
        # set up first header
        if sort == "name":
            for car in carrier_list:
                self.array_var.append(car)
            list_header = "carrier list"
        if sort == "list":
            self.array_var = nl_array + wal_array + otdl_array + ptf_array + aux_array
            if len(nl_array) > 0:
                list_header = "nl"
            else:
                list_header = " "
        if sort == "ns day":
            self.array_var = yellow_array + blue_array + green_array + brown_array + red_array + black_array + \
                             none_array
            if len(yellow_array) > 0:
                list_header = "yellow"
            else:
                list_header = " "
        Label(self.win.body, text=list_header).grid(row=i, column=0)
        i += 1
        sql = "SELECT * FROM ns_configuration"
        ns_results = inquire(sql)
        ns_dict = {}  # build dictionary for ns days
        days = ("sat", "mon", "tue", "wed", "thu", "fri")
        for r in ns_results:  # build dictionary for rotating ns days
            ns_dict[r[0]] = r[2]
        for d in days:  # expand dictionary for fixed days
            ns_dict[d] = "fixed: " + d
        ns_dict["none"] = "none"  # add "none" to dictionary
        # intialize arrays for option menus
        opt_list = "nl", "wal", "otdl", "aux", "ptf"
        nsk = []
        days = ("sat", "mon", "tue", "wed", "thu", "fri")
        for each in projvar.ns_code.keys():
            nsk.append(each)  # make an array of projvar.ns_code keys
        opt_nsday = []  # make an array of "day / color" options for option menu
        for each in projvar.ns_code:
            ns_option = projvar.ns_code[each] + "  " + ns_dict[each]  # make a string for each day/color
            if each in days:
                ns_option = "fixed:" + "  " + each  # if the ns day is fixed - make a special string
            if each == "none":
                ns_option = "---" + "  " + each  # if the ns day is "none" - make a special string
            opt_nsday.append(ns_option)
        count = 0
        for record in self.array_var:  # loop to put information on to window
            # set up color
            if i & 1:
                color = "light yellow"
            else:
                color = "white"
            if sort == "list":
                if list_header != record[2]:
                    list_header = record[2]
                    Label(self.win.body, text=list_header).grid(row=i, column=0)
                    i += 1
            if sort == "ns day":
                if list_header != record[3]:
                    list_header = record[3]
                    Label(self.win.body, text=list_header).grid(row=i, column=0)
                    i += 1
            # set up carrier name button and variable
            Button(self.win.body, text=record[1], width=macadj(24, 20), anchor="w", bg=color, bd=0) \
                .grid(row=i, column=0)
            # set up list status option menu and variable
            self.mi_list.append(StringVar(self.win.body))
            om_list = OptionMenu(self.win.body, self.mi_list[count], *opt_list)  # configuration below
            om_list.grid(row=i, column=1, ipadx=0)
            self.mi_list[count].set(record[2])
            # set up ns day option menu and variable
            self.mi_nsday.append(StringVar(self.win.body))
            om_nsday = OptionMenu(self.win.body, self.mi_nsday[count], *opt_nsday)  # configuration below
            om_nsday.grid(row=i, column=2)
            ns_index = nsk.index(record[3])
            self.mi_nsday[count].set(opt_nsday[ns_index])
            # set up station option menu and variable
            self.mi_station.append(StringVar(self.win.body))
            # configuration below
            om_station = OptionMenu(self.win.body, self.mi_station[count], *projvar.list_of_stations)
            om_station.grid(row=i, column=3)
            self.mi_station[count].set(record[5])
            # adjust optionmenu configuration by platform
            if sys.platform == "darwin":
                om_list.config(width=4, bg=color)
                om_nsday.config(width=9, bg=color)
                om_station.config(width=18, bg=color)
            else:
                om_list.config(width=5, anchor="w", bg=color, relief='ridge', bd=0)
                om_nsday.config(width=10, anchor="w", bg=color, relief='ridge', bd=0)
                om_station.config(width=28, anchor="w", bg=color, relief='ridge', bd=0)
            # set up route variable - not visible but passed along with other variables
            self.mi_route.append(StringVar(self.win.body))
            self.mi_route[count].set(record[4])
            count += 1
            i += 1
        del carrier_list[:]
        self.build_buttons()
        self.win.finish()

    def build_buttons(self):
        """ build the buttons on the bottom of the page """
        button_submit = Button(self.win.buttons, text="Submit", width=15, command=lambda: self.apply_mi(goback=True))
        button_apply = Button(self.win.buttons, text="Apply", width=15, command=lambda: self.apply_mi())
        button_back = Button(self.win.buttons, text="Go Back", width=15,
                             command=lambda: MainFrame().start(frame=self.win.topframe))

        if sys.platform == "win32":
            button_submit.config(anchor="w")
            button_apply.config(anchor="w")
            button_back.config(anchor="w")
        button_submit.pack(side=LEFT)
        button_apply.pack(side=LEFT)
        button_back.pack(side=LEFT)

    def apply_mi(self, goback=False):
        """ enter changes from multiple input into database """
        x = self.pass_date.get()
        year = IntVar()
        month = IntVar()
        day = IntVar()
        y = projvar.invran_date_week[x].strftime("%Y").lstrip("0")
        m = projvar.invran_date_week[x].strftime("%m").lstrip("0")
        d = projvar.invran_date_week[x].strftime("%d").lstrip("0")
        year.set(y)
        month.set(m)
        day.set(d)
        sql = "SELECT * FROM ns_configuration"
        ns_results = inquire(sql)
        ns_dict = {}  # build dictionary for ns days
        for r in ns_results:  # build dictionary for rotating ns days
            ns_dict[r[2]] = r[0]
        ns_dict["none"] = "none"  # add "none" to dictionary
        for i in range(len(self.array_var)):  # loop through all received data
            if "fixed: " not in self.mi_nsday[i].get():
                passed_ns = self.mi_nsday[i].get().split("  ")  # break apart the day/color_code
                self.mi_nsday[i].set(ns_dict[passed_ns[1]])  # match color_code to proper color_code in dict and set
            else:
                passed_ns = self.mi_nsday[i].get().split("  ")  # do not subject the fixed to the dictionary
                self.mi_nsday[i].set(passed_ns[1])
            # if there is a differance, then put the new record in the database
            if self.array_var[i][2] != self.mi_list[i].get() or self.array_var[i][3] != self.mi_nsday[i].get() \
                    or self.array_var[i][5] != self.mi_station[i].get():
                self.apply(i)
        if goback:
            MainFrame().start(frame=self.win.topframe)
        else:
            self.mass_input(self.win.topframe, self.mi_date.get(), self.mi_sort.get())

    def apply(self, i):
        """ executes to insert or update changes from mass input. """
        date = self.apply_date  # simplify the variable names
        name = self.array_var[i][1]
        list_ = self.mi_list[i].get()
        nsday = self.mi_nsday[i].get()
        route = self.mi_route[i].get()
        station = self.mi_station[i].get()
        sql = "SELECT effective_date, carrier_name,list_status, ns_day,route_s, station, rowid FROM carriers " \
              "WHERE carrier_name = '%s' and effective_date = '%s' ORDER BY effective_date" % \
              (self.array_var[i][1], self.apply_date)
        results = inquire(sql)
        if len(results) == 0:
            sql = "INSERT INTO carriers (effective_date,carrier_name,list_status,ns_day,route_s,station)" \
                  " VALUES('%s','%s','%s','%s','%s','%s')" \
                  % (date, name, list_, nsday, route, station)
            commit(sql)
        elif len(results) == 1:
            sql = "UPDATE carriers SET list_status='%s',ns_day='%s',route_s='%s',station='%s' " \
                  "WHERE effective_date = '%s' and carrier_name = '%s'" % \
                  (list_, nsday, route, station, date, name)
            commit(sql)
        elif len(results) > 1:
            sql = "DELETE FROM carriers WHERE effective_date ='%s' and carrier_name = '%s'" % (date, name)
            commit(sql)
            sql = "INSERT INTO carriers (effective_date,carrier_name,list_status,ns_day,route_s,station)" \
                  " VALUES('%s','%s','%s','%s','%s','%s')" \
                  % (date, name, list_, nsday, route, station)
            commit(sql)


class CarrierInput:
    """
    provides screens for users to view carrier characteristics, add, edit and delete.
    """

    def __init__(self):
        self.input_type = None  # 3 types: new, edit, update
        self.carrier = ""  # a string for the carrier's name used only in edit and update
        self.win = None
        self.status_label = None  # the label widget in the buttons method.
        self.ns_dict = None
        self.ns_color_dict = None
        # set up vars
        self.month = None
        self.day = None
        self.year = None
        self.name = None  # last name only or full name with first initial
        self.fname = None  # first initial (only used with new input type)
        self.id = None
        self.seniority = None
        self.ls = None
        self.ns = None
        self.route = None
        self.station = None
        # onrecs - Carrier information on record and already in the database, used only for edit and update
        self.onrecs = None
        self.onrec_ls = None
        self.onrec_ns = None
        self.onrec_route = None
        self.onrec_station = None
        self.onrecs_id = None
        self.onrecs_seniority = None
        self.name_set = []  # get a list of carrier names for new carriers and name changes (edit).
        # new carrier specific
        self.carrier_set = []  # get a list of carriers and effective dates for new carriers.
        # edit carrier specific
        self.chg_name = None
        self.status = ""  # status message.
        # update carrier specific
        self.rowid = None

    def initialize_vars(self):
        """ initialize the variables """
        self.year = StringVar(self.win.body)  # define variables for date
        self.month = IntVar(self.win.body)
        self.day = IntVar(self.win.body)
        self.name = StringVar(self.win.body)  # can be last name or full name
        self.fname = StringVar(self.win.body)  # used only for new carriers
        self.chg_name = StringVar(self.win.body)  # used only for edit carriers
        self.id = StringVar(self.win.body)
        self.seniority = StringVar(self.win.body)
        self.ls = StringVar(self.win.body)
        self.route = StringVar(self.win.body)
        self.ns = StringVar(self.win.body)
        self.station = StringVar(self.win.body)

    def set_new_vars(self):
        """ set the vars for new carrier entries. """
        self.year.set(projvar.invran_year)  # dates are set to the investigation range.
        self.month.set(projvar.invran_month)
        self.day.set(projvar.invran_day)
        self.name.set("")  # all other information is blank
        self.fname.set("")
        self.id.set("")
        self.seniority.set("")
        self.ls.set(value="nl")  # default is 'no list'
        self.route.set("")
        self.ns.set("none")  # default non schedule day is none
        self.station.set(projvar.invran_station)  # default value

    def get_onrecs(self):
        """ get the record for the carrier during the investigation range. """
        sql = "SELECT effective_date, carrier_name, list_status, ns_day, route_s, station, rowid" \
              " FROM carriers WHERE carrier_name = '%s' ORDER BY effective_date DESC" % self.carrier
        self.onrecs = inquire(sql)  # used for status change history
        self.onrec_ls = self.onrecs[0][2]  # used to set stringvars
        self.onrec_ns = self.onrecs[0][3]
        self.onrec_route = self.onrecs[0][4]
        self.onrec_station = self.onrecs[0][5]
        sql = "SELECT emp_id FROM name_index WHERE kb_name = '%s'" % self.carrier
        name_result = inquire(sql)  # the employee id on record
        self.onrecs_id = ""  # employee id default value is an empty string
        if name_result:  # if there is a record
            self.onrecs_id = name_result[0][0]
        sql = "SELECT senior_date FROM seniority WHERE name = '%s'" % self.carrier
        senior_result = inquire(sql)  # the seniority date on record
        self.onrecs_seniority = ""  # seniority date default value is an empty string
        if senior_result:  # if there is a record
            if senior_result[0][0]:  # if the seniority date is not an empty string
                # convert the datetime obj string into a backslash date
                self.onrecs_seniority = Convert(senior_result[0][0]).dtstring_to_backslashdate()

    def set_edit_vars(self):
        """ set the vars for an existing carrier. """
        self.year.set(projvar.invran_year)  # dates are set to the investigation range.
        self.month.set(projvar.invran_month)
        self.day.set(projvar.invran_day)
        self.name.set(self.carrier)  # set the to carrier name
        self.chg_name.set(self.carrier)
        self.id.set(self.onrecs_id)
        self.seniority.set(self.onrecs_seniority)
        self.ls.set(self.onrec_ls)  # default is 'no list'
        self.route.set(self.onrec_route)
        self.ns.set(self.onrec_ns)  # default non schedule day is none
        self.station.set(self.onrec_station)  # default value

    def set_update_vars(self, onrecs):
        """ set the vars for updating a carrier record. """
        self.year.set(int(onrecs[0][:4]))  # dates are set value passed from on recs.
        self.month.set(int(onrecs[0][5:7]))
        self.day.set(int(onrecs[0][8:10]))
        self.name.set(onrecs[1])  # set the to carrier name
        self.chg_name.set(self.carrier)
        self.ls.set(onrecs[2])  # value=onrec[2]
        self.route.set(onrecs[4])
        self.ns.set(onrecs[3])
        self.station.set(onrecs[5])

    def get_nsdicts(self):
        """ get ns day color configurations """
        sql = "SELECT * FROM ns_configuration"
        ns_results = inquire(sql)
        self.ns_dict = {}  # build dictionary for ns days
        self.ns_color_dict = {}
        days = ("sat", "mon", "tue", "wed", "thu", "fri")
        for r in ns_results:  # build dictionary for rotating ns days
            self.ns_dict[r[0]] = r[2]
            self.ns_color_dict[r[0]] = r[1]  # build dictionary for ns fill colors
        for d in days:  # expand dictionary for fixed days
            self.ns_dict[d] = "fixed: " + d
            self.ns_color_dict[d] = "teal"
        self.ns_dict["none"] = "none"  # add "none" to dictionary
        self.ns_color_dict["none"] = "teal"

    def restart_new_carriers(self, frame):
        """ reinitialize and restart New Carrier when the user hits Apply. """
        self.ns_dict = None
        self.ns_color_dict = None
        # set up vars
        self.month = None
        self.day = None
        self.year = None
        self.carrier = ""
        self.name = None
        self.fname = None
        self.id = None
        self.seniority = None
        self.ls = None
        self.ns = None
        self.route = None
        self.station = None
        self.name_set = []  # get a list of carrier names for new carriers
        self.carrier_set = []  # get a list of carriers and effective dates for new carriers.
        self.new_carriers(frame)  # restart the new carriers method.

    def new_carriers(self, frame):
        """ window for inputting new carriers """
        self.input_type = "new"
        self.get_nsdicts()
        self.win = MakeWindow()
        self.win.create(frame)
        self.initialize_vars()
        self.set_new_vars()
        self.title()
        self.date()
        self.get_name()
        self.id_seniority()
        self.list_status()
        self.nsday()
        self.get_route()
        self.get_station()
        self.buttons()
        self.win.finish()

    def restart_edit_carriers(self, frame, carrier):
        """ reinitialize and restart edit carriers when user changes name/ hits apply"""
        self.ns_dict = None
        self.ns_color_dict = None
        # set up vars
        self.month = None
        self.day = None
        self.year = None
        self.name = None  # last name only or full name with first initial
        self.id = None
        self.seniority = None
        self.ls = None
        self.ns = None
        self.route = None
        self.station = None
        # onrecs - Carrier information on record and already in the database, used only for edit and update
        self.onrec_ls = None
        self.onrec_ns = None
        self.onrec_route = None
        self.onrec_station = None
        self.name_set = []  # get a list of carrier names for new carriers and name changes (edit).
        # edit carrier specific
        self.chg_name = None
        self.edit_carriers(frame, carrier)  # restart the edit carriers method.

    def edit_carriers(self, frame, carrier):
        """ window for editing existing carriers - creating new records for a carrier. """
        self.input_type = "edit"
        self.carrier = carrier
        self.get_nsdicts()
        self.win = MakeWindow()
        self.win.create(frame)
        self.initialize_vars()
        self.get_onrecs()  # get the information on record for the carrier
        self.set_edit_vars()  # set the string/int vars for the carrier
        self.title()
        self.date()
        self.get_name()
        self.id_seniority()
        self.list_status()
        self.nsday()
        self.get_route()
        self.get_station()
        self.delete_button()
        self.reports()
        self.status_history()
        self.buttons()
        self.win.finish()

    def update_carrier(self, frame, onrec):
        """ window of updating existing carrier records. """
        self.input_type = "update"
        self.rowid = onrec[6]  # the row id of the record to be updated.
        self.get_nsdicts()
        self.win = MakeWindow()
        self.win.create(frame)
        self.set_update_vars(onrec)  # set the string/int vars to values passed in onrec
        self.title()
        self.date()
        self.get_name()
        self.list_status()
        self.nsday()
        self.get_route()
        self.get_station()
        self.buttons()
        self.win.finish()

    def title(self):
        """ set the title for new carrier input"""
        title_f = Frame(self.win.body)
        title_f.grid(row=0, sticky=W, pady=5)  # put frame on grid
        text = "Enter New Carrier"  # default for new input type.
        if self.input_type == "edit":
            text = "Edit Carrier Information"
        if self.input_type == "update":
            text = "Update Carrier Record"
        Label(title_f, text=text, font=macadj("bold", "Helvetica 18")).grid(row=0, column=0, columnspan=4)

    def date(self):
        """ set up the date widgets. """
        date_frame = Frame(self.win.body)  # define frame
        date_frame.grid(row=1, sticky=W, pady=5)  # put frame on grid
        text = macadj("Effective Date _______________________________",
                      "Effective Date _________________________")
        Label(date_frame, text=text, anchor="w", fg="blue").grid(row=0, column=0, columnspan=20, sticky="w")
        Label(date_frame, text="Month", fg=macadj("grey", "grey"), anchor="w").grid(row=1, column=0)
        Label(date_frame, text="Day", fg=macadj("grey", "grey"), anchor="w").grid(row=1, column=1)
        Label(date_frame, text="Year", fg=macadj("grey", "grey"), anchor="w").grid(row=1, column=2)
        Label(date_frame, text="          ").grid(row=1, column=3)
        om_month = OptionMenu(date_frame, self.month, "1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "11", "12")
        om_month.config(width=2)
        om_month.grid(row=2, column=0, sticky=W)
        om_day = OptionMenu(date_frame, self.day, "1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "11", "12", "13",
                            "14", "15", "16", "17", "18", "19", "20", "21", "22", "23", "24", "25", "26", "27", "28",
                            "29", "30", "31")
        om_day.config(width=2)
        om_day.grid(row=2, column=1, sticky=W)
        Label(date_frame, text="          ").grid(row=2, column=3)
        Entry(date_frame, width=6, textvariable=self.year).grid(row=2, column=2, sticky=W)

    def get_name(self):
        """ enter the carrier's name"""
        name_frame = Frame(self.win.body, pady=2)
        name_frame.grid(row=2, sticky=W, pady=5)
        text = macadj("Carrier Name _______________________________",
                      "Carrier Name _________________________")
        Label(name_frame, text=text, anchor="w", fg="blue").grid(row=0, column=0, columnspan=3, sticky="w")
        if self.input_type == "new":
            Label(name_frame, text=" Last Name: ", width=22, anchor="w", background=macadj("gray95", "white"),
                  fg=macadj("black", "grey")).grid(row=1, column=0, sticky=W)
            Label(name_frame, text=" 1st Initial ", width=7, anchor="w", background=macadj("gray95", "white"),
                  fg=macadj("black", "grey")).grid(row=1, column=1, sticky=W)
            Entry(name_frame, width=macadj(27, 22), textvariable=self.name).grid(row=2, column=0, sticky=W)
            Entry(name_frame, width=macadj(8, 6), textvariable=self.fname).grid(row=2, column=1, sticky=W)
        if self.input_type == "edit":
            Label(name_frame, text=" Carrier Name: {}".format(self.carrier), anchor="w",
                  background=macadj("gray95", "white"), fg=macadj("black", "black"), width=30) \
                .grid(row=1, column=0, columnspan=4, sticky=W)
            Entry(name_frame, width=macadj(37, 29), textvariable=self.chg_name) \
                .grid(row=2, column=0, columnspan=4, sticky=W)
            Label(name_frame, text="Change Name: ").grid(row=3, column=0, sticky=W)
            Button(name_frame, width=7, text="Update",
                   command=lambda: self.name_change()).grid(row=3, column=1, sticky=W, pady=6)
        if self.input_type == "update":
            Label(name_frame, text=self.carrier, anchor="w",
                  background=macadj("gray95", "white"), fg=macadj("black", "black"), width=30) \
                .grid(row=1, column=0, columnspan=4, sticky=W)

    def id_seniority(self):
        """ display widgets for the employee id and seniority date """
        id_frame = Frame(self.win.body, pady=5)
        id_frame.grid(row=3, sticky=W, pady=5)
        Label(id_frame, text=" Employee ID:   ", anchor="w").grid(row=0, column=0, sticky="w")
        length = 10
        if self.input_type == "new":
            length = 21
        Entry(id_frame, width=macadj(length, length), textvariable=self.id).grid(row=0, column=1, sticky=W)
        Label(id_frame, text=" Seniority Date:   ", anchor="w").grid(row=1, column=0, sticky="w")
        Entry(id_frame, width=macadj(length, length), textvariable=self.seniority).grid(row=1, column=1, sticky=W)
        if self.input_type == "edit":
            Button(id_frame, width=7, text="Update",
                   command=lambda: self.id_change()).grid(row=0, column=2, sticky=W, padx=12)
            Button(id_frame, width=7, text="Update",
                   command=lambda: self.seniority_change()).grid(row=1, column=2, sticky=W, padx=12)

    def list_status(self):
        """ set up the list status """
        list_frame = Frame(self.win.body, pady=5)
        list_frame.grid(row=4, sticky=W, pady=5)
        text = macadj("List Status _________________________________",
                      "List Status ___________________________")
        Label(list_frame, text=text, anchor="w", fg="blue").grid(row=0, column=0, columnspan=3, sticky="w")

        Radiobutton(list_frame, text="OTDL", variable=self.ls, value='otdl', justify=LEFT) \
            .grid(row=1, column=0, sticky=W)
        Radiobutton(list_frame, text="Work Assignment", variable=self.ls, value='wal', justify=LEFT) \
            .grid(row=1, column=1, sticky=W)
        Radiobutton(list_frame, text="No List", variable=self.ls, value='nl', justify=LEFT) \
            .grid(row=2, column=0, sticky=W)
        Radiobutton(list_frame, text="Auxiliary", variable=self.ls, value='aux', justify=LEFT) \
            .grid(row=2, column=1, sticky=W)
        Radiobutton(list_frame, text="Part Time Flex", variable=self.ls, value='ptf', justify=LEFT) \
            .grid(row=3, column=1, sticky=W)

    def nsday(self):
        """ set up the ns day"""
        ns_frame = Frame(self.win.body, pady=5)
        ns_frame.grid(row=5, sticky=W, pady=5)
        text = macadj("Non Scheduled Day ________________________",
                      "Non Scheduled Day ____________________")
        Label(ns_frame, text=text, anchor="w", fg="blue").grid(row=0, column=0, columnspan=3, sticky="w")
        Radiobutton(ns_frame, text="{}:   yellow".format(projvar.ns_code['yellow']), variable=self.ns, value="yellow",
                    indicatoron=macadj(0, 1), width=15, anchor="w", bg=macadj("grey", "white"),
                    fg=macadj("white", "black"), selectcolor=self.ns_color_dict["yellow"]).grid(row=1, column=0)
        Radiobutton(ns_frame, text="{}:   blue".format(projvar.ns_code['blue']), variable=self.ns, value="blue",
                    indicatoron=macadj(0, 1), width=15, anchor="w", bg=macadj("grey", "white"),
                    fg=macadj("white", "black"), selectcolor=self.ns_color_dict["blue"]).grid(row=2, column=0)
        Radiobutton(ns_frame, text="{}:   green".format(projvar.ns_code['green']), variable=self.ns, value="green",
                    indicatoron=macadj(0, 1), width=15, anchor="w", bg=macadj("grey", "white"),
                    fg=macadj("white", "black"), selectcolor=self.ns_color_dict["green"]).grid(row=3, column=0)
        Radiobutton(ns_frame, text="{}:   brown".format(projvar.ns_code['brown']), variable=self.ns, value="brown",
                    indicatoron=macadj(0, 1), width=15, anchor="w", bg=macadj("grey", "white"),
                    fg=macadj("white", "black"), selectcolor=self.ns_color_dict["brown"]).grid(row=1, column=1)
        Radiobutton(ns_frame, text="{}:   red".format(projvar.ns_code['red']), variable=self.ns, value="red",
                    indicatoron=macadj(0, 1), width=15, anchor="w", bg=macadj("grey", "white"),
                    fg=macadj("white", "black"), selectcolor=self.ns_color_dict["red"]).grid(row=2, column=1)
        Radiobutton(ns_frame, text="{}:   black".format(projvar.ns_code['black']), variable=self.ns, value="black",
                    indicatoron=macadj(0, 1), width=15, anchor="w", bg=macadj("grey", "white"),
                    fg=macadj("white", "black"), selectcolor=self.ns_color_dict["black"]).grid(row=3, column=1)
        Label(ns_frame, text=" Fixed:", anchor="w").grid(row=4, column=0, sticky="w")
        Radiobutton(ns_frame, text="none", variable=self.ns, value="none", indicatoron=macadj(0, 1),
                    width=15, anchor="w") \
            .grid(row=4, column=1)
        Radiobutton(ns_frame, text="none", variable=self.ns, value="none", indicatoron=macadj(0, 1),
                    width=15, bg=macadj("grey", "white"), fg=macadj("white", "black"),
                    selectcolor=self.ns_color_dict["none"], anchor="w").grid(row=4, column=1)
        Radiobutton(ns_frame, text="Sat:   fixed", variable=self.ns, value="sat", bg=macadj("grey", "white"),
                    fg=macadj("white", "black"),
                    selectcolor=self.ns_color_dict["sat"], indicatoron=macadj(0, 1), width=15,
                    anchor="w").grid(row=5, column=0)
        Radiobutton(ns_frame, text="Mon:   fixed", variable=self.ns, value="mon", bg=macadj("grey", "white"),
                    fg=macadj("white", "black"),
                    selectcolor=self.ns_color_dict["mon"], indicatoron=macadj(0, 1),
                    width=15, anchor="w").grid(row=5, column=1)
        Radiobutton(ns_frame, text="Tue:   fixed", variable=self.ns, value="tue", bg=macadj("grey", "white"),
                    fg=macadj("white", "black"),
                    selectcolor=self.ns_color_dict["tue"], indicatoron=macadj(0, 1),
                    width=15, anchor="w").grid(row=6, column=0)
        Radiobutton(ns_frame, text="Wed:   fixed", variable=self.ns, value="wed", bg=macadj("grey", "white"),
                    fg=macadj("white", "black"),
                    selectcolor=self.ns_color_dict["wed"], indicatoron=macadj(0, 1),
                    width=15, anchor="w").grid(row=6, column=1)
        Radiobutton(ns_frame, text="Thu:   fixed", variable=self.ns, value="thu", bg=macadj("grey", "white"),
                    fg=macadj("white", "black"),
                    selectcolor=self.ns_color_dict["thu"], indicatoron=macadj(0, 1),
                    width=15, anchor="w").grid(row=7, column=0)
        Radiobutton(ns_frame, text="Fri:   fixed", variable=self.ns, value="fri", bg=macadj("grey", "white"),
                    fg=macadj("white", "black"),
                    selectcolor=self.ns_color_dict["fri"], indicatoron=macadj(0, 1),
                    width=15, anchor="w").grid(row=7, column=1)

    def get_route(self):
        """ set route entry field """
        route_frame = Frame(self.win.body, pady=2)
        route_frame.grid(row=6, sticky=W)
        text = macadj("Route _______________________________________",
                      "Route _______________________________")
        Label(route_frame, text=text, anchor="w", fg="blue").grid(row=0, column=0, columnspan=3, sticky="w")
        # Label(route_frame, text=" Route/s", width=30, anchor="w", background=macadj("gray95", "grey"),
        #       fg=macadj("black", "white")).grid(row=1, column=0, sticky=W)
        Entry(route_frame, width=macadj(37, 29), textvariable=self.route).grid(row=1, column=0, sticky=W)

    def get_station(self):
        """ set station option menu"""
        station_frame = Frame(self.win.body, pady=5)
        station_frame.grid(row=7, sticky=W, pady=5)
        text = macadj("Station _____________________________________",
                      "Station ______________________________")
        Label(station_frame, text=text, anchor="w", fg="blue").grid(row=0, column=0, columnspan=3, sticky="w")
        om_stat = OptionMenu(station_frame, self.station, *projvar.list_of_stations)
        om_stat.config(width=macadj(30, 27))
        om_stat.grid(row=1, column=0, sticky=W)

    def delete_button(self):
        """ delete button - allows user to delete all records of the carrier. """
        delete_frame = Frame(self.win.body, pady=5)
        delete_frame.grid(row=8, sticky=W, pady=5)
        text = macadj("Delete Carrier ______________________________",
                      "Delete Carrier ________________________")
        Label(delete_frame, text=text, anchor="w", fg="blue").grid(row=0, column=0, columnspan=3, sticky="w")
        Label(delete_frame, text="Delete carrier and all associated records. ", anchor="w") \
            .grid(row=1, column=0, columnspan=3, sticky=W)
        Button(delete_frame, text="Delete", width=15,
               bg=macadj("red3", "white"), fg=macadj("white", "red"),
               command=lambda: self.purge_carrier()).grid(row=3, column=0, sticky=W, padx=8)

    def reports(self):
        """ create a button that allows the user to view all the carrier's records. """
        report_frame = Frame(self.win.body, pady=5)
        report_frame.grid(row=9, sticky=W, pady=5)
        text = macadj("Status Change Report _______________________",
                      "Status Change Report __________________")
        Label(report_frame, text=text, anchor="w", fg="blue").grid(row=0, column=0, columnspan=3, sticky="w")
        Label(report_frame, text="Generate Report: ", anchor="w").grid(row=1, column=0, sticky=W)
        Button(report_frame, text="Report", width=10,
               command=lambda: Reports(self.win.topframe).rpt_carrier_history(self.carrier)) \
            .grid(row=1, column=1, sticky=W, padx=10)

    def status_history(self):
        """ History of status changes """
        history_frame = Frame(self.win.body, pady=5)
        history_frame.grid(row=10, sticky=W, pady=5)
        text = macadj("Status Change History ______________________",
                      "Status Change History _________________")
        Label(history_frame, text=text, anchor="w", fg="blue").grid(row=0, column=0, columnspan=3, sticky="w")
        row_line = 1
        for line in self.onrecs:
            con_date = datetime.strptime(line[0], "%Y-%m-%d %H:%M:%S")  # convert str to datetime obj.
            Label(history_frame, width=25, text="date: {}".format(str(con_date.strftime("%b %d, %Y"))), anchor="w") \
                .grid(row=row_line, column=0, sticky=W, columnspan=4)
            row_line += 1
            Label(history_frame, width=25, text="list status: {}".format(line[2]), anchor="w") \
                .grid(row=row_line, column=0, sticky=W, columnspan=4)
            row_line += 1
            Label(history_frame, width=25, text="ns day: {}".format(self.ns_dict[line[3]]), anchor="w") \
                .grid(row=row_line, column=0, sticky=W, columnspan=4)
            row_line += 1
            Label(history_frame, width=35, text="route: {}".format(line[4]), anchor="w") \
                .grid(row=row_line, column=0, sticky=W, columnspan=4)
            row_line += 1
            Label(history_frame, width=25, text="station: {}".format(line[5]), anchor="w") \
                .grid(row=row_line, column=0, sticky=W, columnspan=4)
            row_line += 1
            button_alignment = macadj("w", "center")
            Button(history_frame, width=14, text="edit", anchor=button_alignment,
                   command=lambda rec=line: self.update_carrier(self.win.topframe, rec)) \
                .grid(row=row_line, column=0, sticky=W, )
            Button(history_frame, width=14, text="delete", anchor=button_alignment,
                   command=lambda rec=line: self.delete_rec(rec)) \
                .grid(row=row_line, column=1, sticky=W)
            Label(history_frame, text="                             ").grid(row=row_line, column=2, sticky=W)
            row_line += 1

    def buttons(self):
        """ define and display the buttons on the bottom of the screen. """
        button_submit = Button(self.win.buttons)  # buttons at bottom of screen
        button_apply = Button(self.win.buttons)
        button_back = Button(self.win.buttons)
        if self.input_type == "new":
            button_submit.config(text="Submit", command=lambda: self.nc_apply(goback=True))
            button_apply.config(text="Apply", command=lambda: self.nc_apply())
        if self.input_type == "edit":
            button_submit.config(text="Submit", command=lambda: self.ec_apply(goback=True))
            button_apply.config(text="Apply", command=lambda: self.ec_apply())
        button_back.config(text="Go Back", command=lambda: MainFrame().start(frame=self.win.topframe))
        if self.input_type == "update":
            button_submit.config(text="Submit", command=lambda: self.update_apply(goback=True))
            button_apply.config(text="Apply", command=lambda: self.update_apply())
            button_back.config(text="Go Back",
                               command=lambda: self.restart_edit_carriers(self.win.topframe, self.carrier))
        if sys.platform == "win32":
            button_submit.config(anchor="w", width=10)
            button_apply.config(anchor="w", width=10)
            button_back.config(anchor="w", width=10)
        else:
            button_submit.config(width=11)
            button_apply.config(width=11)
            button_back.config(width=11)
        button_submit.pack(side=LEFT)
        button_apply.pack(side=LEFT)
        button_back.pack(side=LEFT)
        self.status_label = Label(self.win.buttons, text=self.status, fg="red")
        self.status_label.pack(side=LEFT)

    def nc_apply(self, goback=False):
        """ executes to check then enter in new carrier information into the database. """
        if not self.check_date():  # check the date
            return
        adddate = str(datetime(int(self.year.get()), self.month.get(), self.day.get(), 00, 00, 00))
        # check the carrier name
        addname = self.join_names()  # join the last name and first initial
        if not addname:  # if method returns False then return.
            return  # join name returns a name or False.
        self.get_name_set()  # get a list of carriers in the carriers table.
        if addname in self.name_set:  # if the name is already in the carriers table...
            if not messagebox.askokcancel("New Carrier Input Warning",
                                          "This carrier name is already in the database.\n"
                                          "Did you want to proceed? \n"
                                          "Pressing Ok will create a new record for an existing carrier. ",
                                          parent=self.win.topframe):
                return
        for pair in self.carrier_set:
            if pair[0] == addname and pair[1] == adddate:
                messagebox.showwarning("New Carrier - Prohibited Action",
                                       "There is a pre existing record for this carrier on this day.\n"
                                       "You can not update that record using this window.\n"
                                       "To edit/ delete this record, return to the main page and press\n"
                                       "\"edit\" to the right of the carrier's name. ",
                                       parent=self.win.topframe)
                return
        new_id = self.id.get().strip()  # get the employee id
        if not EmpIdChecker().run_newcarrier(new_id, self.win.body):  # run all checks for employee id
            return  # return if error
        new_senior_date = self.seniority.get().strip()  # get the seniority date
        if not SeniorityChecker().run_manual(new_senior_date, self.win.body):  # run all checks for seniority date
            return  # will give error messages and return if False
        if not self.check_route():  # check the carrier route
            return  # return if error
        # once all checks have be completed and passed, input data into database
        addroute = Handler(self.route.get()).routes_adj()  # convert 5 digit route numbers to 4 digits.
        if addroute == "0000":  # convert route 0000 to empty string.
            addroute = ""
        # add record for carrier table
        sql = "INSERT INTO carriers (effective_date,carrier_name,list_status,ns_day,route_s,station)" \
              " VALUES('%s','%s','%s','%s','%s','%s')" \
              % (adddate, addname, self.ls.get(), self.ns.get(), addroute, self.station.get())
        commit(sql)
        # add record for name_index table
        if new_id:
            sql = "INSERT INTO name_index (tacs_name, kb_name, emp_id) " \
                  "VALUES('%s', '%s', '%s')" \
                  % ("", addname, new_id)
            commit(sql)
        self.carrier = addname  # input seniority relies on self.carrier to name carrier
        self.input_seniority()  # add record for seniority
        self.status = "{} was added.".format(addname)
        if goback:
            MainFrame().start(frame=self.win.topframe)
        else:
            self.restart_new_carriers(self.win.topframe)

    def ec_apply(self, goback=False):
        """ executes to check then enter a new carrier record into the database. """
        if not self.check_date():  # check the date
            return
        adddate = str(datetime(int(self.year.get()), self.month.get(), self.day.get(), 00, 00, 00))
        if not self.check_route():  # check the carrier route
            return
        addroute = Handler(self.route.get()).routes_adj()  # convert 5 digit route numbers to 4 digits.
        if addroute == "0000":  # convert route 0000 to empty string.
            addroute = ""
        sql = "SELECT effective_date, carrier_name,list_status, ns_day,route_s, station, rowid FROM carriers " \
              "WHERE carrier_name = '%s' and effective_date = '%s' ORDER BY effective_date" % \
              (self.carrier, adddate)
        results = inquire(sql)
        if len(results) == 0:
            sql = "INSERT INTO carriers (effective_date,carrier_name,list_status,ns_day,route_s,station)" \
                  " VALUES('%s','%s','%s','%s','%s','%s')" % \
                  (adddate, self.carrier, self.ls.get(), self.ns.get(), addroute, self.station.get())
            commit(sql)
            self.status = "Carrier record added."
        elif len(results) == 1:
            sql = "UPDATE carriers SET list_status='%s',ns_day='%s',route_s='%s',station='%s' " \
                  "WHERE effective_date = '%s' and carrier_name = '%s'" % \
                  (self.ls.get(), self.ns.get(), addroute, self.station.get(), adddate, self.carrier)
            commit(sql)
            self.status = "Carrier record updated."
        elif len(results) > 1:
            sql = "DELETE FROM carriers WHERE effective_date ='%s' and carrier_name = '%s'" % \
                  (adddate, self.carrier)
            commit(sql)
            sql = "INSERT INTO carriers (effective_date,carrier_name,list_status,ns_day,route_s,station)" \
                  " VALUES('%s','%s','%s','%s','%s','%s')" \
                  % (adddate, self.carrier, self.ls.get(), self.ns.get(), addroute, self.station.get())
            commit(sql)
            self.status = "Carrier record updated. "
        if goback:
            MainFrame().start(frame=self.win.topframe)
        else:
            self.restart_edit_carriers(self.win.topframe, self.carrier)

    def update_apply(self, goback=False):
        """ checks and then updates a record for a carrier. """
        if not self.check_date():  # check the date
            return
        adddate = str(datetime(int(self.year.get()), self.month.get(), self.day.get(), 00, 00, 00))
        if not self.check_route():  # check the carrier route
            return
        addroute = Handler(self.route.get()).routes_adj()  # convert 5 digit route numbers to 4 digits.
        if addroute == "0000":  # convert route 0000 to empty string.
            addroute = ""
        sql = "UPDATE carriers SET effective_date='%s',list_status='%s',ns_day='%s',route_s='%s',station='%s' " \
              "WHERE rowid = '%s'" % \
              (adddate, self.ls.get(), self.ns.get(), addroute, self.station.get(), self.rowid)
        commit(sql)
        self.status = "Carrier record updated. "
        if goback:
            MainFrame().start(frame=self.win.topframe)
        else:
            self.restart_edit_carriers(self.win.topframe, self.carrier)

    def check_date(self):
        """ checks dates for months and days from option menus and years with entry widgets. """
        checkdate = DateChecker(self.win.body, self.month.get(), self.day.get(), self.year.get())
        if not checkdate.check_int():  # check that the year is an integer.
            return False
        if not checkdate.check_year():  # check that the year is within an acceptable range.
            return False
        if not checkdate.try_date():  # if the checks returned False, then return.
            return False
        return True

    def join_names(self):
        """ check and join last name and first initial."""
        if len(self.name.get()) < 1:
            messagebox.showerror("Name input error",
                                 "You must enter a name.",
                                 parent=self.win.topframe)
            return False
        if len(self.fname.get()) < 1:
            messagebox.showerror("Name input error",
                                 "You must enter a first initial or name.",
                                 parent=self.win.topframe)
            return False
        if len(self.fname.get()) > 1:
            if not messagebox.askyesno("Caution",
                                       "It is recommended that you use only the first initial of the first"
                                       "name unless it is necessary to create a unique identifier, such as"
                                       "when you have two identical names that must be distinguished."
                                       "Do you want to proceed?",
                                       parent=self.win.topframe):
                return False

        name = self.name.get().strip() + ", " + self.fname.get().strip()
        checkname = NameChecker(name, frame=self.win.body)  # run checks in the toolbox
        if not checkname.check_characters():  # make sure only letters and special characters are in name.
            return False
        if not checkname.check_length():  # if the length is no more than 28 characters.
            return False
        return self.name.get().strip().lower() + ", " + self.fname.get().strip().lower()

    def get_name_set(self):
        """ get a distinct list of carrier names from the carriers table. """
        sql = "SELECT carrier_name, effective_date FROM carriers"
        result = inquire(sql)
        for x in result:
            self.carrier_set.append([x[0], x[1]])  # add the carrier name and effective date
            self.name_set.append(x[0])  # add the carrier name to the name set.

    def check_route(self):
        """ check the route. return True if the route is blank. """
        checkroute = RouteChecker(self.route.get(), frame=self.win.topframe)
        if checkroute.is_empty():  # if the route is an empty string, return True.
            return True
        if not checkroute.check_numeric():
            return False
        if not checkroute.check_array():
            return False
        if not checkroute.check_length():
            return False
        return True

    def id_change(self):
        """ enter or change the employee id number """
        if not EmpIdChecker().run_manual(self.id.get(), self.onrecs_id, self.win.body):
            self.id.set(self.onrecs_id)  # revert field to emp id in dbase or empty string.
            return
        self.input_id()  # if all checks pass, input into database
        if self.input_type == "edit":
            self.status = "Employee ID updated"
            self.status_label.config(text=self.status)
            projvar.root.update()

    def input_id(self):
        """ input the employee id into the name_index table. """
        new_id = self.id.get().strip()
        sql = "SELECT emp_id FROM name_index WHERE kb_name = '%s'" % self.carrier
        result = inquire(sql)  # the employee id on record
        if result and not new_id:  # if the field is blank - delete the record for the carrier
            sql = "DELETE FROM name_index WHERE kb_name = '%s'" % self.carrier
            commit(sql)
            self.onrecs_id = ""
            self.id.set("")
            return
        elif not result:  # if there is no record in the name_index table
            sql = "INSERT INTO name_index (tacs_name, kb_name, emp_id) " \
                  "VALUES('%s', '%s', '%s')" \
                  % ("", self.carrier, new_id)
            commit(sql)
        else:
            sql = "UPDATE name_index SET emp_id = '%s' WHERE kb_name = '%s'" % (new_id, self.carrier)
            commit(sql)
        sql = "SELECT emp_id FROM name_index WHERE kb_name = '%s'" % self.carrier
        result = inquire(sql)
        self.onrecs_id = result[0][0]
        self.id.set(result[0][0])

    def seniority_change(self):
        """ enter or change the employee seniority date """
        new_senior_date = self.seniority.get().strip()
        if not SeniorityChecker().run_manual(new_senior_date, self.win.body):  # run all checks
            self.seniority.set(self.onrecs_seniority)  # revert field to semiority in dbase or empty string.
            return  # will give error messages and return if False
        self.input_seniority()  # if all checks pass, input into database
        if self.input_type == "edit":
            sql = "SELECT senior_date FROM seniority WHERE name = '%s'" % self.carrier
            result = inquire(sql)  # retrieve the seniority date you just created/ modified
            senior_date = result[0][0]
            date_ = ""  # the default is an empty string
            if senior_date:  # if the result inquiry is not an empty string
                date_ = Convert(senior_date).dtstring_to_backslashdate()  # convert the datetime to a backslash date
            self.onrecs_seniority = date_  # update the seniority date on record
            self.seniority.set(date_)  # update the seniority date field
            self.status = "Seniority Date updated"  # update the status label on bottom of the screen
            self.status_label.config(text=self.status)
            projvar.root.update()  # root update

    def input_seniority(self):
        """ input the seniority into the seniority table """
        new_senior_date = self.seniority.get().strip()
        date_ = ""
        if not new_senior_date:  # do not reformat if the string is empty
            pass
        else:
            date_ = Convert(new_senior_date).backslashdate_to_datetime()  # create a datetime object string
        sql = "SELECT senior_date FROM seniority WHERE name = '%s'" % self.carrier
        result = inquire(sql)  # the seniordate on record
        if not result:  # if there is no record in the seniority table
            sql = "INSERT INTO seniority (name, senior_date) " \
                  "VALUES('%s', '%s')" % (self.carrier, date_)
            commit(sql)  # since no record exist - create it
        else:
            sql = "UPDATE seniority SET senior_date = '%s' WHERE name = '%s'" % (date_, self.carrier)
            commit(sql)  # since a record does exist - modify it

    def name_change(self):
        """ change the name of the carrier in the Edit input type """
        c_name = self.chg_name.get()  # get name from the stringvar
        if len(c_name) < 1:
            messagebox.showerror("Change Name Error",
                                 "You must enter a name.",
                                 parent=self.win.topframe)
            return
        checkname = NameChecker(c_name, frame=self.win.body)  # run checks in the toolbox
        if not checkname.check_characters():  # make sure only letters and special characters are in name.
            return
        if not checkname.check_length():  # if the length is no more than 28 characters.
            return
        if not checkname.check_comma():  # if there is no comma in the name - there must be one
            return
        if not checkname.check_initial():  # checks for more than one character in first initial place
            if not messagebox.askokcancel("Change Name Warning",
                                          "It is recommended that first initials only consist of one "
                                          "letter unless adding more is necessary to creating a unique name. \n"
                                          "Do you want to proceed?",
                                          parent=self.win.topframe):
                return
        c_name = c_name.strip().lower()  # strip out any whitespace and convert to all lowercase.
        self.get_name_set()  # get a list of carriers in the carriers table.
        if c_name in self.name_set:  # if the name is already in the carriers table...
            messagebox.showerror("Change Name Error",
                                 "This carrier name is already in the database.\n"
                                 "You can not change the carrier's name to a name that is already being used.",
                                 parent=self.win.topframe)
            return
        if not messagebox.askokcancel("Name Change",
                                      "This will change the name {} to {} in all records. "
                                      "Are you sure?".format(self.carrier, self.chg_name.get()),
                                      parent=self.win.topframe):
            return
        tables = ("carriers", "informalc_awards2", "informalc_payouts", "otdl_preference", "refusals", "rings3",
                  "seniority", "name_index")
        columns = ("carrier_name", "carrier_name", "carrier_name", "carrier_name", "carrier_name", "carrier_name",
                   "name", "kb_name")
        for i in range(len(tables)):
            sql = "SELECT {} FROM {} WHERE {} = '%s'".format(columns[i], tables[i], columns[i]) % self.carrier
            result = inquire(sql)  # look for record
            if result:
                sql = "UPDATE {} SET {} = '%s' WHERE {} = '%s'".format(tables[i], columns[i], columns[i]) \
                      % (c_name, self.carrier)
                commit(sql)
        self.status = "Carrier name change applied."
        self.restart_edit_carriers(self.win.topframe, c_name)

    def purge_carrier(self):
        """ executes to delete all carrier records along with rings and name index from the database. """
        if not messagebox.askokcancel("Delete Carrier",
                                      "This will delete the carrier and all records associated with "
                                      "this carrier, including rings and name index.\n\n"
                                      "If this carrier has left the station, quit, been fired or retired "
                                      "you should change station to \"out of station\" and not delete. \n\n"
                                      "This can not be reversed.",
                                      parent=self.win.topframe):
            return
        sql = "DELETE FROM carriers WHERE carrier_name = '%s'" % self.carrier
        commit(sql)
        sql = "DELETE FROM rings3 WHERE carrier_name= '%s'" % self.carrier
        commit(sql)
        sql = "DELETE FROM name_index WHERE kb_name = '%s'" % self.carrier
        commit(sql)
        sql = "DELETE FROM seniority WHERE name = '%s'" % self.carrier
        commit(sql)
        MainFrame().start(frame=self.win.topframe)

    def delete_rec(self, onrec):
        """ executes when a carrier is deleted. """
        sql = "DELETE FROM carriers WHERE rowid = '%s'" % onrec[6]
        commit(sql)
        sql = "SELECT carrier_name FROM carriers WHERE carrier_name = '%s'" % onrec[1]
        results = inquire(sql)
        if len(results) > 0:
            self.status = "Carrier record deleted. "
            self.restart_edit_carriers(self.win.topframe, onrec[1])
        else:
            MainFrame().start(frame=self.win.topframe)


class MainFrame:
    """
    This is the main screen where the carrier list and all pull down menus are displayed.
    """

    def __init__(self):
        self.win = None
        self.nav = None  # an window for the optional mac navigation method
        self.invest_frame = None
        self.main_frame = None
        self.start_year = None
        self.start_month = None  # stringvars
        self.start_day = None
        self.i_range = None  # investigation range boolean
        self.invran = None  # investigation range stringvar
        self.start_date = None
        self.end_date = None
        self.station = None
        self.carrier_list = []
        self.invran_date = None  # investigation range date
        self.stations_minus_outofstation = []  # list of stations
        self.invran_result = None
        self.spreadsheet_pref = "improper_mandate"
        self.ot_date = None  # an argument passed to the OT Equitability Spreadsheet - a date in the
        # ot distribution spreadsheet will show only work assignment and no list carriers
        self.listoptions = ("wal", "nl")

    def start(self, frame=None):
        """ master method for controlling methods in class """
        self.win = MakeWindow()
        self.win.create(frame)  # create the window
        self.invest_frame = Frame(self.win.body)
        self.main_frame = Frame(self.win.body)
        self.invest_frame.pack()  # put the investigation frame in the window
        self.main_frame.pack()  # puts the mainframe in the window
        self.set_dates()
        self.make_stringvars()
        self.get_carrierlist()  # call CarrierList to get Carrier Rec Set
        if not projvar.mac_navigation:
            self.pulldown_menu()  # create a pulldown menu, and add it to the menu bar
        self.set_investigation_vars()  # set the stringvars for the investigation range
        self.get_stations_list()  # get a list of stations for station optionmenu
        self.get_invran_mode()  # get the investigation range mode. alternate widget layouts for investigation range
        if self.invran_result in ("simple", "no labels"):
            self.investigation_range_simple()  # configure widgets for setting investigation range
        else:
            self.investigation_range()  # configure widgets for setting investigation range
        self.get_spreadsheet_preference()  # configure what spreadsheet will generate if spreadsheet button pushed.
        self.investigation_status()  # provide message on status of investigation range
        if projvar.invran_station is None:  # if the investigation range is not set
            self.invran_not_set()  # investigation range not set screen
        else:
            if self.carrier_list:  # is the carrier is has contents
                self.show_carrierlist()  # show the carrier list
            else:  # if the carrier list is empty
                self.empty_carrierlist()  # the carrier list is empty screen
        self.bottom_of_frame()  # place necessary code to mainloop the window
        DataBaseFix().empty_in_rings3(frame)  # check for a bug where empties appear in the date of rings3
        self.win.finish()  # close the window

    def set_dates(self):
        """ gets the start and end dates """
        self.start_date = projvar.invran_date
        self.end_date = projvar.invran_date
        if projvar.invran_weekly_span:
            self.start_date = projvar.invran_date_week[0]
            self.end_date = projvar.invran_date_week[6]
        self.ot_date = projvar.invran_date  # build argument for ot equitability spreadsheet
        if projvar.invran_weekly_span:  # if the investigation range is weekly
            self.ot_date = projvar.invran_date_week[6]  # pass the last day of the investigation range as datetime

    def make_stringvars(self):
        """ create stringvars """
        self.start_year = StringVar(self.win.body)
        self.start_month = StringVar(self.win.body)
        self.start_day = StringVar(self.win.body)
        self.invran_date = StringVar(self.win.body)
        self.i_range = BooleanVar(self.win.body)
        self.invran = StringVar(self.win.body)
        self.station = StringVar(self.invest_frame)

    def get_carrierlist(self):
        """ call CarrierList to get Carrier Rec Set """
        self.carrier_list = CarrierList(self.start_date, self.end_date, projvar.invran_station).get()

    def set_investigation_vars(self):
        """ set the stringvars for the investigation range """
        now = datetime.now()
        self.start_month.set(now.month)  # default setting is now
        self.start_day.set(now.day)
        self.start_year.set(now.year)
        self.invran_date.set(now.strftime("%m/%d/%Y"))
        self.station.set("undefined")  # default value
        if projvar.invran_month:  # set month if a month is set
            self.start_month.set(projvar.invran_month)
        if projvar.invran_day:  # set day if a day is set
            self.start_day.set(projvar.invran_day)
        if projvar.invran_year:  # set year if a year is set
            self.start_year.set(projvar.invran_year)
        if projvar.invran_weekly_span:
            self.invran_date.set(projvar.invran_date_week[0].strftime("%m/%d/%Y"))
        elif projvar.invran_weekly_span is False:
            self.invran_date.set(projvar.invran_date.strftime("%m/%d/%Y"))
        if projvar.invran_station:
            self.station.set(projvar.invran_station)
        if projvar.invran_weekly_span is None:  # investigation range weekly/true or daily/false or none
            self.i_range.set(True)
            self.invran.set("week")
        elif not projvar.invran_weekly_span:  # if investigation range is daily
            self.i_range.set(False)
            self.invran.set("day")
        else:  # if investigation range is weekly
            self.i_range.set(True)
            self.invran.set("week")

    def get_stations_list(self):
        """ get a list of stations for station optionmenu """
        self.stations_minus_outofstation = projvar.list_of_stations[:]
        if "out of station" in self.stations_minus_outofstation:
            self.stations_minus_outofstation.remove("out of station")
        if len(self.stations_minus_outofstation) == 0:
            self.stations_minus_outofstation.append("undefined")

    def get_invran_mode(self):
        """ get the investigation range mode """
        sql = "SELECT tolerance FROM tolerances WHERE category = '%s'" % "invran_mode"
        results = inquire(sql)
        self.invran_result = results[0][0]

    def get_spreadsheet_preference(self):
        """ get the spreadsheet preference from the tolerances table for use in the define_spreadsheet() method. """
        sql = "SELECT tolerance FROM tolerances WHERE category = '%s'" % "spreadsheet_pref"
        results = inquire(sql)
        self.spreadsheet_pref = results[0][0]

    def investigation_range_simple(self):
        """ executes if the investigation range is configured to simple or no labels in gui configution """
        Label(self.invest_frame, text="INVESTIGATION RANGE").grid(row=0, column=0, columnspan=2, sticky=W)
        nav_button = Button(self.invest_frame, text="Navigation", width=macadj(12, 14),
                            command=lambda: self.mac_navigation())
        if projvar.mac_navigation:  # conditional on user preference for navigation
            nav_button.grid(row=0, column=3, columnspan=2, padx=2)  # only show nav button for mac platform
        if self.invran_result != "no labels":  # create a label row
            Label(self.invest_frame, text="Date: ", fg="grey").grid(row=1, column=0, sticky=W)
            Label(self.invest_frame, text="Range: ", fg="grey").grid(row=1, column=1, sticky=W)
            Label(self.invest_frame, text="Station: ", fg="grey").grid(row=1, column=2, sticky=W)
            Label(self.invest_frame, text="Set/Reset: ", fg="grey").grid(row=1, column=3, columnspan=2, sticky=W)
        # create widget row
        Entry(self.invest_frame, textvariable=self.invran_date, width=macadj(14, 9), justify='center') \
            .grid(row=2, column=0, padx=2)
        om_range = OptionMenu(self.invest_frame, self.invran, "week", "day")
        om_range.config(width=4)
        om_range.grid(row=2, column=1, sticky=W, padx=2)
        om_station = OptionMenu(self.invest_frame, self.station, *self.stations_minus_outofstation)
        om_station.config(width=macadj(31, 29))
        om_station.grid(row=2, column=2, sticky=W, padx=2)
        # set and reset buttons for investigation range
        Button(self.invest_frame, text="Set", width=macadj(5, 6), bg=macadj("green", "SystemButtonFace"),
               fg=macadj("white", "green"), command=lambda: self.call_globals()).grid(row=2, column=3, padx=2)
        Button(self.invest_frame, text="Reset", width=macadj(5, 6), bg=macadj("red", "SystemButtonFace"),
               fg=macadj("white", "red"), command=lambda: self.reset_globals()).grid(row=2, column=4, padx=2)

    def investigation_range(self):
        """ configure widgets for setting investigation range """
        nav_button = Button(self.invest_frame, text="Navigation", width=macadj(18, 20),
                            command=lambda: self.mac_navigation())
        if projvar.mac_navigation:
            nav_button.grid(row=0, column=6, columnspan=2, padx=2)  # only show nav button for mac platform
        Label(self.invest_frame, text="INVESTIGATION RANGE").grid(row=1, column=0, columnspan=2)
        om_month = OptionMenu(self.invest_frame, self.start_month, "1", "2", "3", "4", "5", "6", "7", "8", "9",
                              "10", "11", "12")
        om_month.config(width=2)
        om_month.grid(row=1, column=2)
        om_day = OptionMenu(self.invest_frame, self.start_day, "1", "2", "3", "4", "5", "6", "7", "8",
                            "9", "10", "11", "12", "13", "14", "15", "16", "17", "18", "19", "20",
                            "21", "22", "23", "24", "25", "26", "27", "28", "29", "30", "31")
        om_day.config(width=2)
        om_day.grid(row=1, column=3)
        date_year = Entry(self.invest_frame, width=6, textvariable=self.start_year)
        date_year.grid(row=1, column=4)
        Label(self.invest_frame, text="RANGE", width=macadj(6, 8)).grid(row=1, column=5)
        if projvar.invran_weekly_span is None:
            self.i_range.set(True)
        elif not projvar.invran_weekly_span:  # if investigation range is daily
            self.i_range.set(False)
        else:  # if investigation range is weekly
            self.i_range.set(True)
        Radiobutton(self.invest_frame, text="weekly", variable=self.i_range, value=True,
                    width=macadj(6, 7), anchor="w").grid(row=1, column=6)
        Radiobutton(self.invest_frame, text="daily", variable=self.i_range, value=False,
                    width=macadj(6, 7), anchor="w").grid(row=1, column=7)
        # set station option menu
        Label(self.invest_frame, text="STATION", anchor="w").grid(row=2, column=0, sticky=W)
        om = OptionMenu(self.invest_frame, self.station, *self.stations_minus_outofstation)
        om.config(width=macadj(40, 34))
        om.grid(row=2, column=1, columnspan=5, sticky=W)
        # set and reset buttons for investigation range
        Button(self.invest_frame, text="Set", width=macadj(8, 9),
               bg=macadj("green", "SystemButtonFace"), fg=macadj("white", "green"),
               command=lambda: self.make_globals(self.start_year.get(), self.start_month.get(),
                                                 self.start_day.get(), self.i_range.get(), self.station.get(),
                                                 self.win.topframe)) \
            .grid(row=2, column=6)
        Button(self.invest_frame, text="Reset", width=macadj(8, 9), bg=macadj("red", "SystemButtonFace"),
               fg=macadj("white", "red"), command=lambda: self.reset_globals()).grid(row=2, column=7)

    def make_globals(self, year, month, day, i_range, station, frame):
        """ sets the globals and then restarts the class. """
        if not Globals().set(year, month, day, i_range, station, frame):
            return
        self.__init__()  # re initialize the class
        self.start(frame)  # start again

    def reset_globals(self):
        """ resets the globals and then restarts the class. """
        frame = self.win.topframe  # capture the frame object so it isn't destroyed by next line.
        Globals().reset()
        self.__init__()  # re initialize the class
        self.start(frame)  # start again

    def call_globals(self):
        """ breaks down the date and checks each segment before setting the globals to reflect investigation range. """
        msg_rear = "\n Dates must be formatted as \"mm/dd/yyyy\".\n" \
                   "Month must be expressed as number between 1 and 12.\n" \
                   "Day must be expressed as a number between 1 and 31.\n" \
                   "Year must be have four digits and be above 0010. "
        breakdown = BackSlashDateChecker(self.invran_date.get())
        if not breakdown.count_backslashes():
            msg = "The date must have 2 backslashes. " + msg_rear
            messagebox.showerror("Set Investigation Range", msg, parent=self.invest_frame)
            return
        breakdown.breaker()  # fully form the backslashdatechecker object
        if not breakdown.check_numeric():
            msg = "All month, day and year must be numbers. " + msg_rear
            messagebox.showerror("Set Investigation Range", msg, parent=self.invest_frame)
            return
        if not breakdown.check_minimums():
            msg = "All month, day and year must be greater than zero. " + msg_rear
            messagebox.showerror("Set Investigation Range", msg, parent=self.invest_frame)
            return
        if not breakdown.check_month():
            msg = "The value provided for the month is not acceptable. " + msg_rear
            messagebox.showerror("Set Investigation Range", msg, parent=self.invest_frame)
            return
        if not breakdown.check_day():
            msg = "The value provided for the day is not acceptable. " + msg_rear
            messagebox.showerror("Set Investigation Range", msg, parent=self.invest_frame)
            return
        if not breakdown.check_year():
            msg = "The value provided for the year is not acceptable. " + msg_rear
            messagebox.showerror("Set Investigation Range", msg, parent=self.invest_frame)
            return
        if not breakdown.valid_date():
            msg = "The investigation date is not valid. " + msg_rear
            messagebox.showerror("Set Investigation Range", msg, parent=self.invest_frame)
            return
        invest_range = True
        if self.invran.get() == "day":
            invest_range = False
        if not self.make_globals(breakdown.year, breakdown.month, breakdown.day, invest_range, self.station.get(),
                                 self.win.topframe):
            return

    def investigation_status(self):
        """ provide message on status of investigation range """
        # Investigation date SET/NOT SET notification
        if projvar.invran_weekly_span is None:
            Label(self.invest_frame, text="----> Investigation date/range not set", foreground="red") \
                .grid(row=3, column=0, columnspan=8, sticky="w")
        elif projvar.invran_weekly_span == 0:  # if the investigation range is one day
            f_date = projvar.invran_date.strftime("%a - %b %d, %Y")
            Label(self.invest_frame, text="---> Day Set: {} --> Pay Period: {}".format(f_date, projvar.pay_period),
                  foreground="red").grid(row=3, column=0, columnspan=8, sticky="w")
        else:
            # if the investigation range is weekly
            f_date = projvar.invran_date_week[0].strftime("%a - %b %d, %Y")
            end_f_date = projvar.invran_date_week[6].strftime("%a - %b %d, %Y")
            Label(self.invest_frame, text="---> Range Set: {0} through {1} --> Pay Period: {2}"
                  .format(f_date, end_f_date, projvar.pay_period),
                  foreground="red").grid(row=3, column=0, columnspan=8, sticky="w")

    def invran_not_set(self):
        """ executes if investigation range is not set"""
        Button(self.main_frame, text="Automatic Data Entry", width=30,
               command=lambda: AutoDataEntry().run(self.win.topframe)).grid(row=0, column=1, pady=5)
        Button(self.main_frame, text="Informal C", width=30,
               command=lambda: InformalC().informalc(self.win.topframe)).grid(row=1, column=1, pady=5)
        Button(self.main_frame, text="Quit", width=30, command=lambda: projvar.root.destroy()) \
            .grid(row=2, column=1, pady=5)

    def empty_carrierlist(self):
        """ the carrier list is empty """
        Label(self.main_frame, text="").grid(row=0, column=0)
        Label(self.main_frame, text="The carrier list is empty. ", font=macadj("bold", "Helvetica 18")) \
            .grid(row=1, column=0, sticky="w")
        Label(self.main_frame, text="").grid(row=2, column=0)
        Label(self.main_frame, text="Build the carrier list either with the New Carrier feature, Speedsheets or "
                                    "by running the Automatic Data Entry Feature.", wraplength=500,
              justify=LEFT, anchor="w").grid(row=3, column=0)

    def show_carrierlist(self):
        """ investigation range is set and carrier list is not empty """
        Label(self.main_frame, text="Name (click for Rings)", fg="grey").grid(row=0, column=1, sticky="w")
        Label(self.main_frame, text="List", fg="grey").grid(row=0, column=2, sticky="w")
        Label(self.main_frame, text="N/S", fg="grey").grid(row=0, column=3, sticky="w")
        Label(self.main_frame, text="Route", fg="grey").grid(row=0, column=4, sticky="w")
        Label(self.main_frame, text="Edit", fg="grey").grid(row=0, column=5, sticky="w")
        r = 1
        i = 0
        ii = 1
        for line in self.carrier_list:
            rec_count = 0
            # detect any out of station records and modify recset - function returns arrays with (startdate, carrier)
            line = CarrierRecFilter(line, self.start_date).detect_outofstation(projvar.invran_station)
            # if the row is even, then choose a color for it
            if i & 1:
                color = "light yellow"
            else:
                color = "white"
            for rec in line:
                if rec_count == 0:  # display the first row of carrier recs
                    Label(self.main_frame, text=ii).grid(row=r, column=0)  # display count
                    Button(self.main_frame, text=rec[1], width=macadj(25, 23), bg=color, anchor="w",
                           command=lambda x=rec: EnterRings(x[1]).start()).grid(row=r, column=1)
                    Button(self.main_frame, text="edit", width=4, bg=color, anchor="w",
                           command=lambda x=rec[1]: CarrierInput().edit_carriers(self.win.topframe, x)) \
                        .grid(row=r, column=5)
                    ii += 1
                else:  # display non first rows of carrier recs
                    dt = datetime.strptime(rec[0], "%Y-%m-%d %H:%M:%S")
                    Button(self.main_frame, text=dt.strftime("%a"), width=macadj(25, 23), bg=color, anchor="e") \
                        .grid(row=r, column=1)
                    Button(self.main_frame, text="", width=4, bg=color) \
                        .grid(row=r, column=5)
                if len(rec) > 2:  # because "out of station" recs only have two items
                    # list
                    Button(self.main_frame, text=rec[2], width=macadj(3, 4), bg=color, anchor="w").grid(row=r, column=2)
                    day_off = projvar.ns_code[rec[3]].lower()
                    Button(self.main_frame, text=day_off, width=4, bg=color, anchor="w").grid(row=r, column=3)  # nsday
                    Button(self.main_frame, text=rec[4], width=25, bg=color, anchor="w") \
                        .grid(row=r, column=4)  # route
                    rec_count += 1
                else:
                    Button(self.main_frame, text="out of station", width=35, bg=color) \
                        .grid(row=r, column=2, columnspan=3)
                r += 1
                rec_count += 1
            i += 1
            r += 1

    def mac_navigation(self):
        """ create a screen for navigation to be used instead of the pulldown menu for macOS """
        self.nav = MakeWindow()
        self.nav.create(self.win.topframe)
        # options with a first value of 0 are enabled buttons.
        # options with a first value of 1 will be disabled if the investigation range is not set.
        # options with a first value of 2 are labels."""

        options = (
            # ---------------------------------------------------------------------------------------------- basic
            (2, "Basic Operations __________________________", ""),
            (0, "Go Back", lambda: MainFrame().start(frame=self.nav.topframe)),
            (0, "Save All", lambda: save_all(self.nav.topframe)),
            (1, "New Carrier", lambda: CarrierInput().new_carriers(self.nav.topframe)),  # if invran set
            (1, "Multiple Input", lambda dd="Sat", ss="name": MassInput().mass_input(self.nav.topframe, dd, ss)),
            (1, "Mandates Spreadsheet", lambda: ImpManSpreadsheet().create(self.nav.topframe)),
            (1, "Mandates No.4 Spreadsheet", lambda: ImpManSpreadsheet4().create(self.nav.topframe)),
            (1, "Over Max Spreadsheet", lambda: OvermaxSpreadsheet().create(self.nav.topframe)),
            (1, "Off Bid Spreadsheet", lambda: OffbidSpreadsheet().create(self.nav.topframe)),
            (1, "OT Equitability Spreadsheet", lambda: OTEquitSpreadsheet()
             .create(self.nav.topframe, self.ot_date, projvar.invran_station)),
            (1, "OT Distribution Spreadsheet", lambda: OTDistriSpreadsheet()
             .create(self.nav.topframe, projvar.invran_date_week[0],
                     projvar.invran_station, "weekly", self.listoptions)),
            (1, "Availability Spreadsheet", lambda: OtAvailSpreadsheet().create(self.nav.topframe)),
            (0, "OT Preferences", lambda: OtEquitability().create(self.nav.topframe)),
            (1, "OT Distribution", lambda: OtDistribution().create(self.nav.topframe)),
            (0, "Informal C", lambda: InformalC().informalc(self.nav.topframe)),
            (0, "Location", lambda: Messenger(self.nav.topframe).location_klusterbox()),
            (0, "About Klusterbox", lambda: AboutKlusterbox().start(self.nav.topframe)),
            (0, "View Out of Station", lambda: self.make_globals(self.start_year.get(),
                                                                 self.start_month.get(),
                                                                 self.start_day.get(),
                                                                 self.i_range.get(),
                                                                 "out of station",
                                                                 self.nav.topframe)),
            (0, "Quit", lambda: projvar.root.destroy()),
            # ------------------------------------------------------------------------------------------ reader
            (2, "Reader Operations __________________________", ""),
            (0, "Automatic Data Entry", lambda: AutoDataEntry().run(self.nav.topframe)),
            (0, "Auto Over Max Finder", lambda: MaxHr().run(self.nav.topframe)),
            (0, "Everything Report Reader", lambda: ee_skimmer(self.nav.topframe)),
            (0, "PDF Converter", lambda: PdfConverter().run(self.nav.topframe)),
            (0, "PDF Splitter", lambda: PdfSplitter().run(self.nav.topframe)),
            # ------------------------------------------------------------------------------------------- report
            (2, "Report Operations __________________________", ""),
            (0, "TACS Cheat Sheet", lambda: CheatSheet().tacs_cheatsheet()),
            (1, "Carrier Route and NS Day", lambda: Reports(self.nav.topframe).rpt_carrier()),
            (1, "Carrier Route", lambda: Reports(self.nav.topframe).rpt_carrier_route()),
            (1, "Carrier NS Day", lambda: Reports(self.nav.topframe).rpt_carrier_nsday()),
            (1, "Carrier by List", lambda: Reports(self.nav.topframe).rpt_carrier_by_list()),
            (1, "Carrier History", lambda: CarrierHistory().create(self.nav.topframe, projvar.invran_station)),
            (1, "Carrier Seniority", lambda: Reports(self.nav.topframe).rpt_carrier_seniority()),
            (1, "Carrier Seniority and ID", lambda: Reports(self.nav.topframe).rpt_carrier_seniority_id()),
            (1, "Clock Rings Summary", lambda: DatabaseAdmin().database_rings_report(self.nav.topframe,
                                                                                     projvar.invran_station)),
            (0, "Pay Period Guide Generator", lambda: Reports(self.nav.topframe).pay_period_guide()),
            # --------------------------------------------------------------------------------------- speedsheets
            (2, "Speedsheet Operations __________________________", ""),
            (1, "Generate All Inclusive", lambda: SpeedSheetGen(self.nav.topframe, True).gen()),
            (1, "Generate Carrier", lambda: SpeedSheetGen(self.nav.topframe, False).gen()),
            (0, "Pre-check", lambda: SpeedWorkBookGet().open_file(self.nav.topframe, False)),
            (0, "Input to Database", lambda: SpeedWorkBookGet().open_file(self.nav.topframe, True)),
            (0, "Cheatsheet", lambda: CheatSheet().spdsht_cheatsheet()),
            (0, "Instructions", lambda: OpenText().open_docs(self.nav.body, 'speedsheet_instructions.txt')),
            (0, "Speedsheet Archive", lambda: Archive().file_dialogue(dir_path('speedsheets'))),
            (0, "Clear Archive", lambda: Archive().remove_file_var(self.nav.topframe, 'speedsheets')),
            # ----------------------------------------------------------------------------------------- archive
            (2, "Archive Operations __________________________", ""),
            (0, "Mandates Spreadsheet", lambda: Archive().file_dialogue(dir_path('spreadsheets'))),
            (3, "delete", lambda: Archive().remove_file_var(self.nav.topframe, 'spreadsheets')),
            (0, "Mandates No. 4", lambda: Archive().file_dialogue(dir_path('mandates_4'))),
            (3, "delete", lambda: Archive().remove_file_var(self.nav.topframe, 'mandates_4')),
            (0, "Over Max Spreadsheet", lambda: Archive().file_dialogue(dir_path('over_max_spreadsheet'))),
            (3, "delete", lambda: Archive().remove_file_var(self.nav.topframe, 'over_max_spreadsheet')),
            (0, "Speedsheets", lambda: Archive().file_dialogue(dir_path('speedsheets'))),
            (3, "delete", lambda: Archive().remove_file_var(self.nav.topframe, 'speedsheets')),
            (0, "Over Max Finder", lambda: Archive().file_dialogue(dir_path('over_max'))),
            (3, "delete", lambda: Archive().remove_file_var(self.nav.topframe, 'over_max')),
            (0, "Off Bid Assignment", lambda: Archive().file_dialogue(dir_path('off_bid'))),
            (3, "delete", lambda: Archive().remove_file_var(self.nav.topframe, 'off_bid')),
            (0, "OT Equitability", lambda: Archive().file_dialogue(dir_path('ot_equitability'))),
            (3, "delete", lambda: Archive().remove_file_var(self.nav.topframe, 'ot_equitability')),
            (0, "OT Distribution", lambda: Archive().file_dialogue(dir_path('ot_distribution'))),
            (3, "delete", lambda: Archive().remove_file_var(self.nav.topframe, 'ot_distribution')),
            (0, "Everything Report", lambda: Archive().file_dialogue(dir_path('ee_reader'))),
            (3, "delete", lambda: Archive().remove_file_var(self.nav.topframe, 'ee_reader')),
            (0, "Weekly Availability", lambda: Archive().file_dialogue(dir_path('weekly_availability'))),
            (3, "delete", lambda: Archive().remove_file_var(self.nav.topframe, 'weekly_availability')),
            (0, "Pay Period Guide", lambda: Archive().file_dialogue(dir_path('pp_guide'))),
            (3, "delete", lambda: Archive().remove_file_var(self.nav.topframe, 'pp_guide')),
            (0, "Clear All Archives", lambda: Archive().clear_all(self.nav.topframe)),
            # ---------------------------------------------------------------------------------------- management
            (2, "Management Operations __________________________", ""),
            (0, "General Configurations", lambda: GenConfig(self.nav.topframe).create()),
            (0, "List of Stations", lambda: StationList().station_list(self.nav.topframe)),
            (1, "Set Dispatch of Value", lambda: SetDov().run(self.nav.topframe)),
            (0, "Tolerances", lambda: Tolerances().tolerances(self.nav.topframe)),
            (0, "Spreadsheet Settings", lambda: SpreadsheetConfig().start(self.nav.topframe)),
            (1, "NS Day Configurations", lambda: NsConfig().ns_config(self.nav.topframe)),
            (0, "Speedsheet Settings", lambda: SpeedConfig(self.nav.topframe).create()),
            (0, "Informal C Settings", lambda: InformalCSettings().create(self.nav.topframe)),
            (0, "Auto Data Entry Settings", lambda: AdeSettings().start(self.nav.topframe)),
            (0, "PDF Converter Settings", lambda: PdfConvertConfig().start(self.nav.topframe)),
            (0, "Database", lambda: (self.nav.topframe.destroy(), DatabaseAdmin().run(self.nav.topframe))),
            (0, "Delete Carriers",
             lambda: DatabaseAdmin().database_delete_carriers(self.nav.topframe, projvar.invran_station)),
            (0, "Clean Carrier List", lambda: DatabaseAdmin().carrier_list_cleaning(self.nav.topframe)),
            (0, "Name Index", lambda: NameIndex().name_index_screen(self.nav.topframe)),
            (0, "Station Index", lambda: StationIndex().station_index_mgmt(self.nav.topframe)),
        )
        i = 0
        row = 0
        for _ in range(len(options)):
            if options[i][0] == 3:  # if the option is a delete button
                button = Button(self.nav.body, text=options[i][1], width=5, anchor="w", padx=5,
                                activebackground="grey", highlightcolor="red", command=options[i][2])
                button.grid(row=row - 1, column=2, sticky="w")
            elif options[i][0] == 2:  # if the option is a header
                label = Label(self.nav.body, text=options[i][1], fg="blue", width=26, anchor="w")
                label.grid(row=row, column=1, pady=5, sticky="w", columnspan=3)
                row += 1
            else:  # the option is a button
                button = Button(self.nav.body, text=options[i][1], width=21, anchor="w", padx=5,
                                activebackground="grey", highlightcolor="red", command=options[i][2])
                button.grid(row=row, column=1, sticky="w")
                if not projvar.invran_day and options[i][0]:  # disable the button until invran is set.
                    button.config(state=DISABLED)
                row += 1
            i += 1
        Button(self.nav.buttons, text="Quit", width=macadj(13, 13), command=projvar.root.destroy).pack(side=LEFT)
        Button(self.nav.buttons, text="Go Back", width=15,
               command=lambda: MainFrame().start(frame=self.nav.topframe)).pack(side=LEFT)
        self.nav.finish()

    def pulldown_menu(self):
        """ create a pulldown menu, and add it to the menu bar """
        menubar = Menu(self.win.topframe)
        # file menu
        basic_menu = Menu(menubar, tearoff=0)
        basic_menu.add_command(label="Save All", command=lambda: save_all(self.win.topframe))
        basic_menu.add_separator()
        basic_menu.add_command(label="New Carrier", command=lambda: CarrierInput().new_carriers(self.win.topframe))
        basic_menu.add_command(label="Multiple Input",
                               command=lambda dd="Sat", ss="name":
                               MassInput().mass_input(self.win.topframe, dd, ss))
        basic_menu.add_command(label="Mandates Spreadsheet",
                               command=lambda: ImpManSpreadsheet().create(self.win.topframe))
        basic_menu.add_command(label="Mandates No.4 Spreadsheet",
                               command=lambda: ImpManSpreadsheet4().create(self.win.topframe))
        basic_menu.add_command(label="Over Max Spreadsheet",
                               command=lambda: OvermaxSpreadsheet().create(self.win.topframe))
        basic_menu.add_command(label="Off Bid Spreadsheet",
                               command=lambda: OffbidSpreadsheet().create(self.win.topframe))
        basic_menu.add_command(label="OT Equitability Spreadsheet",
                               command=lambda: OTEquitSpreadsheet().create(self.win.topframe,
                                                                           self.ot_date, self.station.get()))
        basic_menu.add_command(label="OT Distribution Spreadsheet", command=lambda: OTDistriSpreadsheet()
                               .create(self.win.topframe, projvar.invran_date_week[0], self.station.get(),
                                       "weekly", self.listoptions))
        basic_menu.add_command(label="Availability Spreadsheet",
                               command=lambda: OtAvailSpreadsheet().create(self.win.topframe))
        basic_menu.add_separator()
        basic_menu.add_command(label="OT Preferences", command=lambda: OtEquitability().create(self.win.topframe))
        basic_menu.add_command(label="OT Distribution", command=lambda: OtDistribution().create(self.win.topframe))
        basic_menu.add_separator()
        basic_menu.add_command(label="Informal C", command=lambda: InformalC().informalc(self.win.topframe))
        basic_menu.add_separator()
        basic_menu.add_command(label="Location", command=lambda: Messenger(self.win.topframe).location_klusterbox())
        basic_menu.add_command(label="About Klusterbox",
                               command=lambda: AboutKlusterbox().start(self.win.topframe))
        basic_menu.add_separator()
        basic_menu.add_command(label="View Out of Station",
                               command=lambda: self.make_globals(self.start_year.get(), self.start_month.get(),
                                                                 self.start_day.get(), self.i_range.get(),
                                                                 "out of station", self.win.topframe))
        basic_menu.add_separator()
        basic_menu.add_command(label="Quit", command=lambda: projvar.root.destroy())
        # gray out options if no investigation range is set
        if projvar.invran_day is None:
            basic_menu.entryconfig(2, state=DISABLED)
            basic_menu.entryconfig(3, state=DISABLED)
            basic_menu.entryconfig(4, state=DISABLED)
            basic_menu.entryconfig(5, state=DISABLED)
            basic_menu.entryconfig(6, state=DISABLED)
            basic_menu.entryconfig(7, state=DISABLED)
            basic_menu.entryconfig(8, state=DISABLED)
        menubar.add_cascade(label="Basic", menu=basic_menu)
        # automated menu
        automated_menu = Menu(menubar, tearoff=0)
        automated_menu.add_command(label="Automatic Data Entry",
                                   command=lambda: AutoDataEntry().run(self.win.topframe))
        automated_menu.add_separator()
        automated_menu.add_command(label=" Auto Over Max Finder", command=lambda: MaxHr().run(self.win.topframe))
        automated_menu.add_command(label="Everything Report Reader", command=lambda: ee_skimmer(self.win.topframe))
        automated_menu.add_separator()
        automated_menu.add_command(label="PDF Converter", command=lambda: PdfConverter().run(self.win.topframe))
        automated_menu.add_command(label="PDF Splitter", command=lambda: PdfSplitter().run(self.win.topframe))
        menubar.add_cascade(label="Readers", menu=automated_menu)
        # reports menu
        reports_menu = Menu(menubar, tearoff=0)
        reports_menu.add_command(label="TACS Cheat Sheet",
                                 command=lambda: CheatSheet().tacs_cheatsheet())
        reports_menu.add_separator()
        reports_menu.add_command(label="Carrier Route and NS Day",
                                 command=lambda: Reports(self.win.topframe).rpt_carrier())
        reports_menu.add_command(label="Carrier Route",
                                 command=lambda: Reports(self.win.topframe).rpt_carrier_route())
        reports_menu.add_command(label="Carrier NS Day",
                                 command=lambda: Reports(self.win.topframe).rpt_carrier_nsday())
        reports_menu.add_command(label="Carrier by List",
                                 command=lambda: Reports(self.win.topframe).rpt_carrier_by_list())
        reports_menu.add_command(label="Carrier History",
                                 command=lambda: CarrierHistory().create(self.win.topframe, projvar.invran_station))
        reports_menu.add_command(label="Carrier Seniority",
                                 command=lambda: Reports(self.win.topframe).rpt_carrier_seniority())
        reports_menu.add_command(label="Carrier Seniority and ID",
                                 command=lambda: Reports(self.win.topframe).rpt_carrier_seniority_id())
        reports_menu.add_separator()
        reports_menu.add_command(label="Clock Rings Summary",
                                 command=lambda: DatabaseAdmin().database_rings_report(self.win.topframe,
                                                                                       projvar.invran_station))
        reports_menu.add_separator()
        reports_menu.add_command(label="Pay Period Guide Generator",
                                 command=lambda: Reports(self.win.topframe).pay_period_guide())
        if projvar.invran_day is None:
            reports_menu.entryconfig(2, state=DISABLED)
            reports_menu.entryconfig(3, state=DISABLED)
            reports_menu.entryconfig(4, state=DISABLED)
            reports_menu.entryconfig(5, state=DISABLED)
            reports_menu.entryconfig(6, state=DISABLED)
            reports_menu.entryconfig(7, state=DISABLED)
        menubar.add_cascade(label="Reports", menu=reports_menu)
        # speedsheeet menu
        speed_menu = Menu(menubar, tearoff=0)
        speed_menu.add_command(label="Generate All Inclusive",
                               command=lambda: SpeedSheetGen(self.win.topframe, True).gen())
        speed_menu.add_command(label="Generate Carrier",
                               command=lambda: SpeedSheetGen(self.win.topframe, False).gen())
        speed_menu.add_command(label="Pre-check",
                               command=lambda: SpeedWorkBookGet().open_file(self.win.topframe, False))
        speed_menu.add_command(label="Input to Database",
                               command=lambda: SpeedWorkBookGet().open_file(self.win.topframe, True))
        if projvar.invran_day is None:
            speed_menu.entryconfig(0, state=DISABLED)
            speed_menu.entryconfig(1, state=DISABLED)
        speed_menu.add_separator()
        speed_menu.add_command(label="Cheatsheet",
                               command=lambda: CheatSheet().spdsht_cheatsheet())
        speed_menu.add_command(label="Instructions",
                               command=lambda: OpenText().open_docs(self.win.body, 'speedsheet_instructions.txt'))
        speed_menu.add_command(label="Speedsheet Archive",
                               command=lambda: Archive().file_dialogue(dir_path('speedsheets')))
        speed_menu.add_command(label="Clear Archive",
                               command=lambda: Archive().remove_file_var(self.win.topframe, 'speedsheets'))
        menubar.add_cascade(label="Speedsheet", menu=speed_menu)
        # archive menu
        reportsarchive_menu = Menu(menubar, tearoff=0)
        reportsarchive_menu.add_command(label="Mandates Spreadsheet",
                                        command=lambda: Archive().file_dialogue(dir_path('spreadsheets')))
        reportsarchive_menu.add_command(label="Mandates No. 4",
                                        command=lambda: Archive().file_dialogue(dir_path('mandates_4')))
        reportsarchive_menu.add_command(label="Over Max Spreadsheet",
                                        command=lambda: Archive().file_dialogue(dir_path('over_max_spreadsheet')))
        reportsarchive_menu.add_command(label="Speedsheets",
                                        command=lambda: Archive().file_dialogue(dir_path('speedsheets')))
        reportsarchive_menu.add_command(label="Over Max Finder",
                                        command=lambda: Archive().file_dialogue(dir_path('over_max')))
        reportsarchive_menu.add_command(label="Off Bid Assignment",
                                        command=lambda: Archive().file_dialogue(dir_path('off_bid')))
        reportsarchive_menu.add_command(label="OT Equitability",
                                        command=lambda: Archive().file_dialogue(dir_path('ot_equitability')))
        reportsarchive_menu.add_command(label="OT Distribution",
                                        command=lambda: Archive().file_dialogue(dir_path('ot_distribution')))
        reportsarchive_menu.add_command(label="Everything Report",
                                        command=lambda: Archive().file_dialogue(dir_path('ee_reader')))
        reportsarchive_menu.add_command(label="Weekly Availability",
                                        command=lambda: Archive().file_dialogue(dir_path('weekly_availability')))
        reportsarchive_menu.add_command(label="Pay Period Guide",
                                        command=lambda: Archive().file_dialogue(dir_path('pp_guide')))
        reportsarchive_menu.add_separator()
        cleararchive = Menu(reportsarchive_menu, tearoff=0)
        cleararchive.add_command(label="Mandates Spreadsheet",
                                 command=lambda: Archive().remove_file_var(self.win.topframe, 'spreadsheets'))
        cleararchive.add_command(label="Mandates No. 4",
                                 command=lambda: Archive().remove_file_var(self.win.topframe, 'mandates_4'))
        cleararchive.add_command(label="Over Max Spreadsheet",
                                 command=lambda: Archive().remove_file_var(self.win.topframe, 'over_max_spreadsheet'))
        cleararchive.add_command(label="Speedsheets",
                                 command=lambda: Archive().remove_file_var(self.win.topframe, 'speedsheets'))
        cleararchive.add_command(label="Over Max Finder",
                                 command=lambda: Archive().remove_file_var(self.win.topframe, 'over_max'))
        cleararchive.add_command(label="Off Bid Assignment",
                                 command=lambda: Archive().remove_file_var(self.win.topframe, 'off_bid'))
        cleararchive.add_command(label="OT Equitability",
                                 command=lambda: Archive().remove_file_var(self.win.topframe, 'ot_equitability'))
        cleararchive.add_command(label="OT Distribution",
                                 command=lambda: Archive().remove_file_var(self.win.topframe, 'ot_distribution'))
        cleararchive.add_command(label="Everything Report",
                                 command=lambda: Archive().remove_file_var(self.win.topframe, 'ee_reader'))
        cleararchive.add_command(label="Weekly Availability",
                                 command=lambda: Archive().remove_file_var(self.win.topframe, 'weekly_availability'))
        cleararchive.add_command(label="Pay Period Guide",
                                 command=lambda: Archive().remove_file_var(self.win.topframe, 'pp_guide'))
        reportsarchive_menu.add_cascade(label="Clear Archive", menu=cleararchive)
        menubar.add_cascade(label="Archive", menu=reportsarchive_menu)
        reportsarchive_menu.add_command(label="Clear All Archives",
                                        command=lambda: Archive().clear_all(self.win.topframe))
        # management menu
        management_menu = Menu(menubar, tearoff=0)
        management_menu.add_command(label="General Configurations",
                                    command=lambda: GenConfig(self.win.topframe).create())
        management_menu.add_separator()
        management_menu.add_command(label="List of Stations",
                                    command=lambda: StationList().station_list(self.win.topframe))
        management_menu.add_command(label="Set Dispatch of Value",
                                    command=lambda: SetDov().run(self.win.topframe))
        management_menu.add_command(label="Tolerances", command=lambda: Tolerances().tolerances(self.win.topframe))
        management_menu.add_command(label="Spreadsheet Settings",
                                    command=lambda: SpreadsheetConfig().start(self.win.topframe))
        management_menu.add_command(label="NS Day Configurations",
                                    command=lambda: NsConfig().ns_config(self.win.topframe))
        if projvar.invran_day is None:
            management_menu.entryconfig(3, state=DISABLED)  # disable the Set DOV if invran is not set.
            management_menu.entryconfig(6, state=DISABLED)  # disable ns day configurations if invran is not set.
        management_menu.add_command(label="Speedsheet Settings",
                                    command=lambda: SpeedConfig(self.win.topframe).create())
        management_menu.add_command(label="Informal C Settings",
                                    command=lambda: InformalCSettings().create(self.win.topframe))
        management_menu.add_separator()
        management_menu.add_command(label="Auto Data Entry Settings",
                                    command=lambda: AdeSettings().start(self.win.topframe))
        management_menu.add_command(label="PDF Converter Settings",
                                    command=lambda: PdfConvertConfig().start(self.win.topframe))
        management_menu.add_separator()
        management_menu.add_command(label="Database",
                                    command=lambda: (self.win.topframe.destroy(),
                                                     DatabaseAdmin().run(self.win.topframe)))
        management_menu.add_command(label="Delete Carriers",
                                    command=lambda: DatabaseAdmin().database_delete_carriers(self.win.topframe,
                                                                                             projvar.invran_station))
        management_menu.add_command(label="Clean Carrier List",
                                    command=lambda: DatabaseAdmin().carrier_list_cleaning(self.win.topframe))
        management_menu.add_separator()
        management_menu.add_command(label="Name Index",
                                    command=lambda: NameIndex().name_index_screen(self.win.topframe))
        management_menu.add_command(label="Station Index",
                                    command=lambda: StationIndex().station_index_mgmt(self.win.topframe))
        menubar.add_cascade(label="Management", menu=management_menu)
        projvar.root.config(menu=menubar)
        projvar.root.update()  # root update

    def define_spreadsheet_button(self):
        """ determine what happens when the spreadsheet button on the bottom of the page is pressed. """
        if self.spreadsheet_pref == "Mandates":
            ImpManSpreadsheet().create(self.win.topframe)
        if self.spreadsheet_pref == "Over Max":
            OvermaxSpreadsheet().create(self.win.topframe)
        if self.spreadsheet_pref == "OT Equitability":
            OTEquitSpreadsheet().create(self.win.topframe, self.ot_date, self.station.get())
        if self.spreadsheet_pref == "OT Distribution":
            OTDistriSpreadsheet().create(self.win.topframe, projvar.invran_date_week[0],
                                         self.station.get(), "weekly", self.listoptions)
        if self.spreadsheet_pref == "Mandates_4":
            ImpManSpreadsheet4().create(self.win.topframe)
        if self.spreadsheet_pref == "Off Bid":
            OffbidSpreadsheet().create(self.win.topframe)

    def bottom_of_frame(self):
        """ configure buttons on the bottom of the frame """
        nc_button = Button(self.win.buttons, text="New Carrier", width=macadj(13, 13),
                           command=lambda: CarrierInput().new_carriers(self.win.topframe))
        nc_button.pack(side=LEFT)
        if not projvar.invran_day:
            nc_button.config(state=DISABLED)
        mi_button = Button(self.win.buttons, text="Multi Input", width=macadj(13, 13),
                           command=lambda dd="Sat", ss="name": MassInput().mass_input(self.win.topframe, dd, ss))
        mi_button.pack(side=LEFT)
        if not projvar.invran_day:
            mi_button.config(state=DISABLED)
        ade_button = Button(self.win.buttons, text="Auto Data Entry", width=macadj(12, 12),
                            command=lambda: AutoDataEntry().run(self.win.topframe))
        ade_button.pack(side=LEFT)
        ss_button = Button(self.win.buttons, text="Spreadsheet", width=macadj(13, 13),
                           command=lambda: self.define_spreadsheet_button())
        ss_button.pack(side=LEFT)
        if not projvar.invran_day:
            ss_button.config(state=DISABLED)
        quit_button = Button(self.win.buttons, text="Quit", width=macadj(13, 13),
                             command=projvar.root.destroy)
        quit_button.pack(side=LEFT)


if __name__ == "__main__":
    """ this is where the program starts """
    global pb_flag  # global for multithreading
    setup_plaformvar()  # set up platform variable
    setup_dirs_by_platformvar()  # create klusterbox/.klusterbox or kb_sub directories if they don't exist
    DataBase().setup()  # set up the database
    Fixes().check(version)
    projvar.root = Tk()  # initialize root window
    position_x = 100  # initialize position and size for root window
    position_y = 50
    size_x = 625
    size_y = 600
    projvar.root.title("KLUSTERBOX version {}".format(version))
    titlebar_icon(projvar.root)  # place icon in titlebar
    projvar.root.geometry("%dx%d+%d+%d" % (size_x, size_y, position_x, position_y))
    if len(projvar.list_of_stations) < 2:  # if there are no stations in the stations list
        StartUp().start()  # a start up screen for first time use
    else:
        Archive().remove_file(dir_path_check('report'))  # empty out folders
        Archive().remove_file(dir_path_check('infc_grv'))
        MainFrame().start()  # get the show on the road
